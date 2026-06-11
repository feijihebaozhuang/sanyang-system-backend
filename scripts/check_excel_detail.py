# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import openpyxl

okdir = r'D:\Desktop\换绑输出\OK文件'

# 用openpyxl看OK文件第32批的精确格式
wb = openpyxl.load_workbook(os.path.join(okdir, '\u6362\u7ed1\u6587\u4ef6_\u7b2c32\u6279.xlsx'))
ws = wb.active
print(f'=== OK\u6587\u4ef6\u7b2c32\u6279 ===')
print(f'\u884c\u6570: {ws.max_row}, \u5217\u6570: {ws.max_column}')
for row in range(1, 6):
    vals = []
    for col in range(1, ws.max_column+1):
        vals.append(str(ws.cell(row, col).value)[:30])
    print(f'  \u884c{row}: {vals}')

print()

# 看新生成的第33批
wb2 = openpyxl.load_workbook(os.path.join(okdir, '\u6362\u7ed1\u6587\u4ef6_\u7b2c33\u6279.xlsx'))
ws2 = wb2.active
print(f'=== \u65b0\u7b2c33\u6279 ===')
print(f'\u884c\u6570: {ws2.max_row}, \u5217\u6570: {ws2.max_column}')
for row in range(1, 4):
    vals = []
    for col in range(1, ws2.max_column+1):
        vals.append(str(ws2.cell(row, col).value)[:30])
    print(f'  \u884c{row}: {vals}')

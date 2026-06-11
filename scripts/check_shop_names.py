# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

ok_dir = r'd:\Desktop\换绑输出\OK文件'
sname_map = {
    '阿里大鱼': '阿里大鱼包装',
    '阿里友尚': '阿里友尚包装',
    '阿里亚润': '阿里亚润包装',
    '阿里三羊': '阿里三羊包装',
    '阿里新鑫星': '阿里新鑫星包装',
    '阿里正方形': '阿里正方形包装',
    '天猫扣底盒': '天猫扣底盒包装',
    '天猫止合': '天猫止合包装',
    '天猫彩色': '天猫彩色包装',
    '天猫小批量': '天猫小批量包装',
    '淘宝品牌店': '淘宝品牌店包装',
    '俊鑫': '俊鑫包装',
    '当下家': '当下家包装',
    '平卡': '平卡包装',
}

files = sorted(os.listdir(ok_dir))
for f in files:
    if not f.startswith('换绑文件_第') or not f.endswith('.xlsx'):
        continue
    m = __import__('re').search(r'第(\d+)批', f)
    batch = int(m.group(1)) if m else 0
    
    fp = os.path.join(ok_dir, f)
    wb = oxl.load_workbook(fp)
    ws = wb[wb.sheetnames[0]]
    
    shops = set()
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            shops.add(str(r[0]).strip())
    
    # 检查是否有简称
    has_short = [s for s in shops if s in sname_map]
    if has_short:
        print('第%d批 (%s): 有简称: %s' % (batch, f, has_short))
    else:
        print('第%d批 (%s): 无简称问题' % (batch, f))
    
    wb.close()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
刀模 Excel 自动导入：识别 7 列 / 9 列，支持多文件合并，可预览不入库。

用法:
  python scripts/import_dimoldb_auto.py --audit data/import/dimoldb/
  python scripts/import_dimoldb_auto.py data/import/dimoldb/刀模7列.xlsx --overwrite
  python scripts/import_dimoldb_auto.py data/import/dimoldb/*.xlsx --merge
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import dimoldb_store as ds

try:
    from settings import DB_CONFIG
    import pymysql

    def get_db():
        return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)

    HAS_DB = True
except Exception:
    HAS_DB = False


def find_header_row(ws) -> tuple[int, list[str]]:
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        vals = [str(v or "").strip() for v in row]
        if "名称" in vals or "刀模名称" in vals or "name" in [v.lower() for v in vals]:
            return r_idx, vals
    raise ValueError("未找到表头行（需包含「名称」）")


def row_to_item(row: tuple, col_map: dict, *, seq: int) -> dict | None:
    if "name" not in col_map:
        return None
    name = (
        str(row[col_map["name"]]).strip()
        if col_map["name"] < len(row) and row[col_map["name"]] is not None
        else ""
    )
    if not name or name.startswith("=") or name in ("None", "nan"):
        return None
    return {
        "product_type": ds.cell_str(row, col_map.get("product_type")) or "zhengsquare",
        "code": ds.cell_str(row, col_map.get("code")),
        "id": f"dm_import_{int(time.time())}_{seq}",
        "name": name,
        "production_spec": ds.cell_str(row, col_map.get("production_spec")),
        "km_mapping_code": ds.cell_str(row, col_map.get("km_mapping_code")),
        "remark": ds.cell_str(row, col_map.get("remark")),
        "length": ds.cell_float(row, col_map.get("length")),
        "width": ds.cell_float(row, col_map.get("width")),
        "height": ds.cell_float(row, col_map.get("height")),
        "stock": 0,
        "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def parse_workbook(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row, headers = find_header_row(ws)
    fmt = ds.detect_import_header_format(headers)
    col_map = ds.map_dimoldb_import_headers(headers)
    items: list[dict] = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = row_to_item(row, col_map, seq=len(items))
        if not item:
            skipped += 1
            continue
        if not (item["length"] and item["width"] and item["height"]):
            skipped += 1
            continue
        items.append(item)
    return {
        "file": str(path),
        "format": fmt,
        "headers": headers,
        "col_map": {k: headers[v] if v < len(headers) else v for k, v in col_map.items()},
        "rows": len(items),
        "skipped": skipped,
        "items": items,
    }


def merge_key(item: dict) -> str:
    code = (item.get("code") or "").strip()
    if code:
        return f"code:{code}"
    return "dim:{item.get('product_type')}:{item['length']}:{item['width']}:{item['height']}:{item.get('name')}"


def merge_items(sources: list[dict], *, prefer: str = "larger") -> list[dict]:
    """多文件合并：同 key 后者覆盖前者（默认按文件列表顺序，最后 wins）。"""
    ordered = sorted(sources, key=lambda s: s["rows"], reverse=(prefer == "larger"))
    by_key: dict[str, dict] = {}
    for src in ordered:
        for it in src["items"]:
            by_key[merge_key(it)] = it
    return list(by_key.values())


def upsert_into_db(items: list[dict], *, overwrite: bool) -> tuple[int, int, int]:
    if not HAS_DB:
        out = ROOT / "data" / "import" / "dimoldb" / "dimoldb_merged_preview.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"无数据库配置，已写出预览: {out} ({len(items)} 条)")
        return 0, 0, len(items)

    db = [] if overwrite else ds.load_dimoldb(get_db, force=True)
    by_code = {str(d.get("code") or "").strip(): i for i, d in enumerate(db) if d.get("code")}
    by_key = {merge_key(d): i for i, d in enumerate(db)}
    added = updated = 0
    for item in items:
        k = merge_key(item)
        idx = by_code.get((item.get("code") or "").strip()) if item.get("code") else None
        if idx is None:
            idx = by_key.get(k)
        if idx is not None:
            old = db[idx]
            item["id"] = old.get("id") or item["id"]
            item["created_at"] = old.get("created_at") or item["created_at"]
            db[idx] = {**old, **item}
            updated += 1
            by_key[k] = idx
        else:
            db.append(item)
            by_key[k] = len(db) - 1
            if item.get("code"):
                by_code[(item["code"] or "").strip()] = len(db) - 1
            added += 1
    if not ds.save_dimoldb(get_db, db):
        raise SystemExit("save_dimoldb 失败")
    ds.invalidate_dimoldb_cache()
    return added, updated, len(db)


def collect_paths(args_list: list[str]) -> list[Path]:
    paths: list[Path] = []
    for a in args_list:
        p = Path(a)
        if p.is_dir():
            paths.extend(sorted(p.glob("*.xlsx")) + sorted(p.glob("*.xls")))
        elif p.is_file():
            paths.append(p)
        else:
            print(f"跳过不存在: {p}")
    return paths


def main() -> None:
    parser = argparse.ArgumentParser(description="刀模 Excel 自动导入（7/9 列）")
    parser.add_argument("paths", nargs="*", help="文件或目录，默认 data/import/dimoldb/")
    parser.add_argument("--audit", action="store_true", help="仅分析表头与行数，不写库")
    parser.add_argument("--overwrite", action="store_true", help="清空后写入合并结果")
    parser.add_argument("--merge", action="store_true", help="多文件合并后写入（默认取行数多的优先）")
    parser.add_argument("--dry-run", action="store_true", help="同 --audit")
    args = parser.parse_args()

    paths = collect_paths(args.paths or [str(ROOT / "data" / "import" / "dimoldb")])
    if not paths:
        print("未找到 xlsx。请将刀模文件放到: data/import/dimoldb/")
        sys.exit(1)

    reports = []
    for path in paths:
        try:
            reports.append(parse_workbook(path))
            r = reports[-1]
            print(f"[{r['format']}列] {path.name}: 有效 {r['rows']} 行, 跳过 {r['skipped']}, 表头={r['headers'][:9]}")
        except Exception as e:
            print(f"失败 {path.name}: {e}")

    if not reports:
        sys.exit(1)

    out_dir = ROOT / "data" / "import" / "dimoldb"
    out_dir.mkdir(parents=True, exist_ok=True)
    audit_path = out_dir / "audit_report.json"
    audit_path.write_text(
        json.dumps(
            [{k: v for k, v in r.items() if k != "items"} for r in reports],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"审计报告: {audit_path}")

    if args.audit or args.dry_run:
        total = sum(r["rows"] for r in reports)
        print(f"合计有效行 {total}（未写库）")
        return

    items = merge_items(reports) if (args.merge or len(reports) > 1) else reports[-1]["items"]
    if len(reports) > 1 and not args.merge:
        best = max(reports, key=lambda r: r["rows"])
        print(f"多文件时默认用行数最多: {Path(best['file']).name}")
        items = best["items"]

    overwrite = args.overwrite or (len(reports) == 1 and not args.merge)
    if len(reports) > 1 and args.merge:
        overwrite = args.overwrite

    added, updated, total = upsert_into_db(items, overwrite=overwrite)
    print(f"入库完成: 新增 {added} 更新 {updated} 库内合计 {total} 条 overwrite={overwrite}")


if __name__ == "__main__":
    main()

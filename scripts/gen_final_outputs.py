# -*- coding: utf-8 -*-
"""生成最终输出文件"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 之前发给你的纯定制ID（直接内嵌）
MANUAL_CUSTOM_IDS = {
    '4894516026054', '4894516026055', '4894516026053', '4f2071ab49352b0d791d76caf22a6f',
    '5024084823781', '5024084823780', '5024084823779', '5024084823778',
    '4722743288387', '4722743288687', '4727972285352', '4727972285638',
    '4699161887749', '4699161887716', '4699161887748', '4516538470040',
    '4498886268744', '4498886268696', '4498886268743', '4498886268695',
    '4881470822303', '4886067583919', '4881470822111', '4876501042458',
    '4876501042473', '4722843088396', '4722843088671', '4529740689561',
    '4529740689530', '4529740689560', '4529740689529', '4886301587845',
    '4886302011245', '4886302011045', '4726907949976', '4900365915204',
    '4900365915054', '4514045260028', '4690595315993', '4514045260027',
    '4517148958464', '4686278878998', '4686278878970', '4686278878999',
    '4686278878971', '4446319741235', '4446319741163', '4446319741234',
    '4446319741162', '4524871565529', '4524871565527', '4524871565528',
    '4524871565526',
    '4645247792151', '4650089453799', '4650089453843', '4650089453887',
    '4650089453931', '4650089453975', '4645247792019', '4645247792063',
    '4645247792107', '4645247792152', '4650089453800', '4650089453844',
    '4650089453888', '4650089453932', '4650089453976', '4645247792020',
    '4645247792064', '4645247792108',
}

# 读未匹配文件
print("读取未匹配文件...", flush=True)
wb = oxl.load_workbook(os.path.join(out, '未匹配平台商品.xlsx'), data_only=True)
ws = wb['未匹配']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"总未匹配: {len(rows)}", flush=True)

custom_link = []
size_miss = []
pingka = []
no_match = []

for r in rows:
    if not r:
        continue
    reason = str(r[4] or '').strip()
    spec_id = str(r[2] or '').strip()
    
    if reason in ('珍珠棉', '蓝绿颜色', '定制关键词'):
        custom_link.append(r)
    elif reason == '尺寸不足':
        if spec_id in MANUAL_CUSTOM_IDS:
            custom_link.append(r)
        else:
            size_miss.append(r)
    elif reason == '平卡/解析失败':
        pingka.append(r)
    else:
        no_match.append(r)

print(f"\n定制链接: {len(custom_link)}", flush=True)
print(f"  珍珠棉: {sum(1 for r in custom_link if r[4]=='珍珠棉')}", flush=True)
print(f"  蓝绿颜色: {sum(1 for r in custom_link if r[4]=='蓝绿颜色')}", flush=True)
print(f"  定制关键词: {sum(1 for r in custom_link if r[4]=='定制关键词')}", flush=True)
print(f"  手动指定: {sum(1 for r in custom_link if str(r[2] or '').strip() in MANUAL_CUSTOM_IDS)}", flush=True)
print(f"尺寸不足: {len(size_miss)}", flush=True)
print(f"平卡: {len(pingka)}", flush=True)
print(f"无匹配: {len(no_match)}", flush=True)

# 定制类换绑
f1 = os.path.join(out, '定制类_换绑文件.xlsx')
wb = oxl.Workbook()
ws = wb.active; ws.title = 'Sheet1'
ws.append([None, '商品对应表', None, None])
ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
for r in custom_link:
    ws.append([str(r[0] or '').strip(), str(r[1] or '').strip(), str(r[2] or '').strip(), '定制链接'])
wb.save(f1)
print(f"\n✅ 定制类: {f1} ({os.path.getsize(f1)/1024:.1f}KB)", flush=True)
wb.close()

# 尺寸不足
f2 = os.path.join(out, '尺寸不足_待处理.xlsx')
wb = oxl.Workbook()
ws = wb.active; ws.title = '尺寸不足'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in size_miss: ws.append(list(r))
wb.save(f2)
print(f"✅ 尺寸不足: {f2} ({os.path.getsize(f2)/1024:.1f}KB, {len(size_miss)}条)", flush=True)
wb.close()

# 平卡
f3 = os.path.join(out, '平卡_待处理.xlsx')
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in pingka: ws.append(list(r))
wb.save(f3)
print(f"✅ 平卡: {f3} ({os.path.getsize(f3)/1024:.1f}MB, {len(pingka)}条)", flush=True)
wb.close()

# 无匹配
f4 = os.path.join(out, '无匹配_待处理.xlsx')
wb = oxl.Workbook()
ws = wb.active; ws.title = '无匹配'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in no_match: ws.append(list(r))
wb.save(f4)
print(f"✅ 无匹配: {f4} ({os.path.getsize(f4)/1024:.1f}MB, {len(no_match)}条)", flush=True)
wb.close()

# 换绑拆分
print(f"\n拆分换绑文件...", flush=True)
src = os.path.join(out, '换绑文件.xlsx')
wb = oxl.load_workbook(src, data_only=True)
bind_rows = list(wb['Sheet1'].iter_rows(min_row=3, values_only=True))
wb.close()

n = 3
batch = math.ceil(len(bind_rows) / n)
for i in range(n):
    s, e = i * batch, min((i + 1) * batch, len(bind_rows))
    chunk = bind_rows[s:e]
    wb2 = oxl.Workbook()
    ws2 = wb2.active; ws2.title = 'Sheet1'
    ws2.append([None, '商品对应表', None, None])
    ws2.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for r in chunk: ws2.append(list(r))
    fn = os.path.join(out, f'换绑文件_第{i+1}批.xlsx')
    wb2.save(fn)
    print(f"  第{i+1}批: {os.path.getsize(fn)/1024/1024:.2f}MB, {len(chunk)}条", flush=True)
    wb2.close()

print("\n全部完成！", flush=True)

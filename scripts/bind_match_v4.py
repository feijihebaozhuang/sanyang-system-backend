# -*- coding: utf-8 -*-
"""
平台商品 → 快麦商品 换绑匹配 v4
修正全部问题

== 解析修正 ==
1. 宽/长/高 各种分隔符: 【】（）() --- ——等
2. mm→cm正确转换
3. 忽略"单个价""数量""起订"等非尺寸文本

== 匹配修正 ==
1. 先按规格名标记的内外径匹配
2. 外径匹配不上时 → 转内径再匹配（长-1.5, 宽-0.5, 高-0.5）
3. 内径匹配不上时 → 转外径再匹配（长+1.5, 宽+0.5, 高+0.5）
"""
import sys, re, os, time
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '深圳市友尚包装有限公司',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

# ============================================================
# 预编译正则
# ============================================================
RE_CUSTOM = re.compile(r'定制|定做|定造|订做|订制')
RE_PEARL = re.compile(r'珍珠棉')
RE_BLUE_GREEN = re.compile(r'蓝色|蓝|绿色|绿')
RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_WHITE = re.compile(r'双面白|双白|白色')
RE_RED = re.compile(r'红色|红')
RE_BLACK = re.compile(r'黑色|黑')
RE_SPECIAL_PRICE = re.compile(r'特价')
RE_NEW_MATERIAL = re.compile(r'新材质|新材')
RE_FIVELAYER = re.compile(r'五层|5层')
RE_THREELAYER = re.compile(r'三层|3层')
RE_DOUBLE_BOX = re.compile(r'双插盒')
RE_BUCKLE_BOX = re.compile(r'扣底盒')
RE_INNER = re.compile(r'内径|内尺寸')
RE_OUTER = re.compile(r'外径|外尺寸')

# 尺寸标签——强化版，支持各种分隔符
# 标准化后所有括号都被替换为空格，所以标签后面只需匹配空格
_PRE = r'\s*'
_PST = r'[\s\]】）)\-—＿_]*'
_DIGIT = r'(\d+\.?\d*)'
_UNIT = r'\s*(?:cm|mm|厘米|毫米)?'

LABEL_PATTERNS = [
    ('长', [
        re.compile(r'长[度]?' + _PRE + _DIGIT + _UNIT + _PST),
        re.compile(r'长度' + _PRE + _DIGIT + _UNIT + _PST),
    ]),
    ('宽', [
        re.compile(r'宽[度]?' + _PRE + _DIGIT + _UNIT + _PST),
        re.compile(r'宽度' + _PRE + _DIGIT + _UNIT + _PST),
    ]),
    ('高', [
        re.compile(r'高[度]?' + _PRE + _DIGIT + _UNIT + _PST),
        re.compile(r'高度' + _PRE + _DIGIT + _UNIT + _PST),
        re.compile(r'厚[度]?' + _PRE + _DIGIT + _UNIT + _PST),
        re.compile(r'厚度' + _PRE + _DIGIT + _UNIT + _PST),
    ]),
]

RE_DIMS_3D = re.compile(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)')
RE_DIMS_2D = re.compile(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)')
RE_DIGIT_WITH_UNIT = re.compile(r'(\d+\.?\d*)\s*(cm|mm|厘米|毫米)?', re.IGNORECASE)
RE_QTY = re.compile(r'个')

# 需要忽略的上下文
RE_IGNORE_CTX = re.compile(r'数量|个起订|起订|单个价|单价|元|不含|客服')

# ============================================================
# 1. 加载快麦商品 + 建立索引
# ============================================================
def load_km(km_file):
    import pandas as pd
    t0 = time.time()
    print("  读取快麦...", flush=True)
    df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
    df.columns = ['code', 'name', 'cat']
    km_map = {}
    for _, r in df.iterrows():
        c = str(r['code']).strip()
        if c and c != 'nan':
            km_map[c] = {
                'name': str(r['name'] or '').strip(),
                'cat': str(r['cat'] or '').strip(),
            }
    print(f"  → {len(km_map)}条, {time.time()-t0:.1f}s", flush=True)
    return km_map

def build_idx(km_map):
    t0 = time.time()
    # 精确: (l,w,h,dk,mat) -> [code]
    exact = {}
    # 模糊: (sorted_lwh, dk) -> [code]
    fuzzy = {}
    # 按尺寸+内外径（无材料）
    dim_dk = {}
    # 全集
    all_codes = set(km_map.keys())
    
    for code in km_map:
        parts = code.split('-')
        if len(parts) < 3:
            continue
        try:
            dims = parts[0].split('*')
            if len(dims) != 3:
                continue
            l, w, h = float(dims[0]), float(dims[1]), float(dims[2])
            dk = parts[1]
            mat = parts[2]
            typ = parts[3] if len(parts) >= 4 else ''
            
            k1 = (l, w, h, dk, mat)
            exact.setdefault(k1, []).append(code)
            
            vals = tuple(sorted([l, w, h]))
            k2 = (vals[0], vals[1], vals[2], dk)
            fuzzy.setdefault(k2, []).append(code)
            
            k3 = (l, w, h, dk)
            dim_dk.setdefault(k3, []).append(code)
        except:
            pass
    
    print(f"  → 索引: {len(exact)}精确 {len(fuzzy)}模糊, {time.time()-t0:.1f}s", flush=True)
    return exact, fuzzy, dim_dk, all_codes

# ============================================================
# 2. 材料判断
# ============================================================
def guess_material(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_WHITE.search(s): return '白色'
    if RE_RED.search(s): return '红色'
    if RE_BLACK.search(s): return '黑色'
    if RE_SPECIAL_PRICE.search(s) or RE_NEW_MATERIAL.search(s): return '优质'
    if RE_FIVELAYER.search(s): return 'EB'
    if RE_THREELAYER.search(s): return '3B'
    return '特硬'

# ============================================================
# 3. 内外径判断
# ============================================================
def guess_dk(s):
    if RE_INNER.search(s): return '内径'
    if RE_OUTER.search(s): return '外径'
    return '外径'  # 沉默默认

# ============================================================
# 4. 尺寸提取（强化版）
# ============================================================
def extract_labeled_dims(s):
    dims = {}
    for label, pats in LABEL_PATTERNS:
        for pat in pats:
            m = pat.search(s)
            if m and m.group(1):
                val = float(m.group(1))
                after = s[m.end():m.end()+5]
                if after.startswith('mm') or 'mm' in after[:3]:
                    val = val / 10.0
                seg = s[m.start():m.end()]
                if re.search(r'\d+\.?\d*\s*mm', seg):
                    val = val / 10.0
                dims[label] = val
                break
    return dims


def extract_nums_clean(s):
    nums = []
    for m in RE_DIGIT_WITH_UNIT.finditer(s):
        val = float(m.group(1))
        unit = (m.group(2) or '').lower()
        if unit in ('mm', '毫米'):
            val = val / 10.0
        if val < 0.5 or val > 500:
            continue
        start = max(0, m.start()-10)
        end = min(len(s), m.end()+10)
        ctx = s[start:end]
        if RE_IGNORE_CTX.search(ctx):
            continue
        if val == int(val) and int(val) >= 10:
            if RE_QTY.search(ctx):
                continue
        nums.append(val)
    return sorted(set(nums), reverse=True)


def parse_spec_v4(text):
    """解析规格名称 v4"""
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None
    
    # 珍珠棉 → 定制
    if RE_PEARL.search(s):
        return {'custom': True, 'reason': '珍珠棉'}
    
    # 蓝绿 → 定制
    if RE_BLUE_GREEN.search(s):
        return {'custom': True, 'reason': '蓝绿颜色'}
    
    # 标准化：替换所有括号为空格
    s = s.replace('【',' ').replace('】',' ').replace('（',' ').replace('）',' ')
    s = s.replace('(',' ').replace(')',' ').replace('——',' ')
    s = re.sub(r'[-]{2,}', ' ', s)
    s = re.sub(r'[—＿_]+', ' ', s)
    
    # 提取尺寸
    dims = extract_labeled_dims(s)
    
    # a*b*c
    m = RE_DIMS_3D.search(s)
    if m:
        vals = [float(m.group(i)) for i in range(1, 4)]
        ctx = s[max(0, m.start()-10):m.end()+10]
        if 'mm' in ctx:
            vals = [v/10 for v in vals]
        for lbl, val in zip(['长', '宽', '高'], vals):
            dims.setdefault(lbl, val)
    
    # a*b
    if not (dims.get('长') and dims.get('宽')):
        m = RE_DIMS_2D.search(s)
        if m:
            v1, v2 = float(m.group(1)), float(m.group(2))
            ctx = s[max(0, m.start()-10):m.end()+10]
            if 'mm' in ctx:
                v1, v2 = v1/10, v2/10
            dims.setdefault('长', v1)
            dims.setdefault('宽', v2)
    
    # 兜底找数字（缺哪个补哪个）
    if not (dims.get('长') and dims.get('宽') and dims.get('高')):
        nums = extract_nums_clean(s)
        if len(nums) >= 3:
            if '长' not in dims: dims['长'] = nums[0]
            if '宽' not in dims: dims['宽'] = nums[1]
            if '高' not in dims: dims['高'] = nums[2]
        elif len(nums) == 2:
            if '长' not in dims: dims['长'] = nums[0]
            if '宽' not in dims: dims['宽'] = nums[1]
    
    have_3d = dims.get('长') and dims.get('宽') and dims.get('高')
    have_2d = dims.get('长') and dims.get('宽') and not dims.get('高')
    
    # 定制关键词但无三围 → 定制
    if RE_CUSTOM.search(s) and not have_3d:
        return {'custom': True, 'reason': '定制关键词'}
    
    # 平卡（只有长宽）→ 无匹配（不是定制，但有待处理）
    if have_2d:
        # 仍然标记为待处理
        return None  # 让主逻辑走"解析失败"分支
    
    if not have_3d:
        return {'custom': True, 'reason': '尺寸不足'}
    
    return {
        '长': dims['长'],
        '宽': dims['宽'],
        '高': dims['高'],
        'dk': guess_dk(s),
        'mat': guess_material(s),
        'type': '双插盒' if RE_DOUBLE_BOX.search(s) else ('扣底盒' if RE_BUCKLE_BOX.search(s) else ''),
    }


def dim_fmt(v):
    """格式化尺寸值"""
    return str(int(v) if v == int(v) else v)


def build_codes(parsed):
    """拼候选编码列表（含内外径转换）"""
    if not parsed or parsed.get('custom'):
        return []
    
    l, w, h = parsed['长'], parsed['宽'], parsed['高']
    dk = parsed['dk']
    mat = parsed['mat']
    typ = parsed.get('type', '')
    
    codes = []
    base = f"{dim_fmt(l)}*{dim_fmt(w)}*{dim_fmt(h)}"
    
    # 原内外径
    if typ:
        codes.append(f"{base}-{dk}-{mat}-{typ}")
    codes.append(f"{base}-{dk}-{mat}")
    
    # 外径→内径转换
    if dk == '外径':
        inner_l = l - 1.5
        inner_w = w - 0.5
        inner_h = h - 0.5
        if inner_l > 0 and inner_w > 0 and inner_h > 0:
            ibase = f"{dim_fmt(inner_l)}*{dim_fmt(inner_w)}*{dim_fmt(inner_h)}"
            if typ:
                codes.append(f"{ibase}-内径-{mat}-{typ}")
            codes.append(f"{ibase}-内径-{mat}")
    
    # 内径→外径转换（备选）
    if dk == '内径':
        outer_l = l + 1.5
        outer_w = w + 0.5
        outer_h = h + 0.5
        obase = f"{dim_fmt(outer_l)}*{dim_fmt(outer_w)}*{dim_fmt(outer_h)}"
        if typ:
            codes.append(f"{obase}-外径-{mat}-{typ}")
        codes.append(f"{obase}-外径-{mat}")
    
    return codes


# ============================================================
# 5. 主匹配
# ============================================================
def match(platform_file, km_file, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    
    km_map = load_km(km_file)
    exact_idx, fuzzy_idx, dim_dk_idx, all_codes = build_idx(km_map)
    
    print("读取平台商品...", flush=True)
    import pandas as pd
    t0 = time.time()
    df = pd.read_excel(platform_file, sheet_name='报表1', header=2, dtype=str)
    arr = df.values
    print(f"  → {len(arr)}行, {time.time()-t0:.1f}s", flush=True)
    
    matched = []
    unmatched = []
    stats = Counter()
    total = 0
    custom_samples = []
    failed_samples = []
    
    for row in arr:
        total += 1
        shop_short = str(row[0] or '').strip() if len(row) >= 1 else ''
        pid = str(row[1] or '').strip() if len(row) >= 2 else ''
        spec_name = str(row[2] or '').strip() if len(row) >= 3 else ''
        spec_id = str(row[3] or '').strip() if len(row) >= 4 else ''
        
        if not spec_name:
            continue
        
        parsed = parse_spec_v4(spec_name)
        
        if parsed is None or parsed.get('custom'):
            reason = parsed.get('reason', '') if parsed else '平卡/解析失败'
            if not reason:
                reason = '平卡/解析失败'
            unmatched.append((shop_short, pid[:30], spec_id[:30], spec_name[:120], reason, ''))
            stats[reason] += 1
            if parsed and parsed.get('custom') and len(custom_samples) < 50:
                custom_samples.append((shop_short, pid, spec_name))
            elif not parsed and len(failed_samples) < 50:
                failed_samples.append((shop_short, pid, spec_name))
            continue
        
        # 生成候选编码列表（含内外径转换）
        candidates = build_codes(parsed)
        hit = None
        
        for cand in candidates:
            if cand in all_codes:
                hit = cand
                break
        
        # 模糊匹配
        if not hit:
            vals = tuple(sorted([parsed['长'], parsed['宽'], parsed['高']]))
            # 先原内外径
            fkey = (vals[0], vals[1], vals[2], parsed['dk'])
            fuzz = fuzzy_idx.get(fkey, [])
            if fuzz:
                mat = parsed['mat']
                for c in fuzz:
                    if mat in c:
                        hit = c
                        break
                if not hit:
                    hit = fuzz[0]
        
        # 转换后的内外径模糊匹配
        if not hit and parsed['dk'] == '外径':
            il = parsed['长'] - 1.5
            iw = parsed['宽'] - 0.5
            ih = parsed['高'] - 0.5
            if il > 0 and iw > 0 and ih > 0:
                ivals = tuple(sorted([il, iw, ih]))
                fkey2 = (ivals[0], ivals[1], ivals[2], '内径')
                fuzz2 = fuzzy_idx.get(fkey2, [])
                if fuzz2:
                    mat = parsed['mat']
                    for c in fuzz2:
                        if mat in c:
                            hit = c
                            break
                    if not hit:
                        hit = fuzz2[0]
        
        # 精确尺寸+内外径（不同材料）
        if not hit:
            k3 = (parsed['长'], parsed['宽'], parsed['高'], parsed['dk'])
            dc = dim_dk_idx.get(k3, [])
            if dc:
                hit = dc[0]
        
        if hit:
            full_name = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched.append((full_name, pid, spec_id, hit))
            stats['匹配成功'] += 1
        else:
            base = f"{dim_fmt(parsed['长'])}*{dim_fmt(parsed['宽'])}*{dim_fmt(parsed['高'])}"
            expected = f"{base}-{parsed['dk']}-{parsed['mat']}"
            unmatched.append((shop_short, pid[:30], spec_id[:30], spec_name[:120], '无匹配', expected))
            stats['无匹配'] += 1
        
        if total % 100000 == 0:
            print(f"  → {total}/{len(arr)} ...", flush=True)
    
    print(f"\n{'='*60}", flush=True)
    print(f"共 {total} 行", flush=True)
    print(f"  匹配成功: {stats.get('匹配成功', 0)}", flush=True)
    for k, v in sorted(stats.items()):
        if k != '匹配成功':
            print(f"  {k}: {v}", flush=True)
    
    import openpyxl as oxl
    
    # 换绑
    f_bind = os.path.join(out_dir, '换绑文件.xlsx')
    wb = oxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for r in matched:
        ws.append(list(r))
    wb.save(f_bind)
    sz = os.path.getsize(f_bind)
    print(f"✅ 换绑: {f_bind} ({sz/1024/1024:.2f}MB, {len(matched)}条)", flush=True)
    wb.close()
    
    # 未匹配
    f_un = os.path.join(out_dir, '未匹配平台商品.xlsx')
    wb = oxl.Workbook()
    ws = wb.active
    ws.title = '未匹配'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for r in unmatched:
        ws.append(list(r))
    
    ws2 = wb.create_sheet('定制样本')
    ws2.append(['店铺', '商品id', '规格名称'])
    for s, pid, name in custom_samples:
        ws2.append([s, str(pid)[:20], str(name)[:100]])
    
    ws3 = wb.create_sheet('解析失败样本')
    ws3.append(['店铺', '商品id', '规格名称'])
    for s, pid, name in failed_samples:
        ws3.append([s, str(pid)[:20], str(name)[:100]])
    
    wb.save(f_un)
    sz2 = os.path.getsize(f_un)
    print(f"✅ 未匹配: {f_un} ({sz2/1024/1024:.2f}MB, {len(unmatched)}条)", flush=True)
    wb.close()
    
    return matched, unmatched, stats


if __name__ == '__main__':
    t0 = time.time()
    match(
        r"d:\Desktop\平台商品.xlsx",
        r"d:\Desktop\快麦商品 - 副本.xlsx",
        r"d:\Desktop\换绑输出"
    )
    print(f"总耗时: {time.time()-t0:.1f}s", flush=True)

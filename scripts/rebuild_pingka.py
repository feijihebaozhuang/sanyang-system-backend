# -*- coding: utf-8 -*-
"""重建平卡 = 原始数据 - 所有已处理spec_id
已处理spec_id来源: 换绑文件.xlsx + OK文件(定制1~5批) + 无匹配_待处理
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')
raw_file = r'd:\Desktop\平台商品.xlsx'

# 1. 收集所有已处理spec_id
processed_specs = set()
file_stats = {}

# 换绑文件.xlsx
wb = oxl.load_workbook(os.path.join(out, '换绑文件.xlsx'))
ws = wb[wb.sheetnames[0]]
cnt = 0
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and len(r) >= 3 and r[2]:
        processed_specs.add(str(r[2]).strip())
        cnt += 1
wb.close()
file_stats['换绑文件.xlsx'] = cnt
print(f'换绑文件.xlsx: {cnt}条')

# OK文件中的定制类
if os.path.exists(ok_dir):
    for f in sorted(os.listdir(ok_dir)):
        if '定制' in f and f.endswith('.xlsx'):
            try:
                wb = oxl.load_workbook(os.path.join(ok_dir, f))
                ws = wb[wb.sheetnames[0]]
                cnt = 0
                for r in ws.iter_rows(min_row=3, values_only=True):
                    if r and len(r) >= 3 and r[2]:
                        processed_specs.add(str(r[2]).strip())
                        cnt += 1
                wb.close()
                file_stats[f] = cnt
                print(f'{f}: {cnt}条')
            except: pass

# 无匹配_待处理
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
cnt = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) >= 3 and r[2]:
        processed_specs.add(str(r[2]).strip())
        cnt += 1
wb.close()
file_stats['无匹配_待处理.xlsx'] = cnt
print(f'无匹配_待处理.xlsx: {cnt}条')

print(f'\n去重后已处理spec_id总数: {len(processed_specs)}')

# 2. 从原始数据找出所有未处理的记录
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
new_pingka = []  # (shop, pid, spec_name, spec_id, '平卡/待处理', '')
raw_total = 0
shop_raw = Counter()
shop_new = Counter()
shop_processed = Counter()

for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    raw_total += 1
    shop = str(r[0]).strip()
    pid = str(r[1] or '').strip() if len(r) >= 2 else ''
    spec_name = str(r[2] or '').strip() if len(r) >= 3 else ''
    spec_id = str(r[3] or '').strip() if len(r) >= 4 else ''
    
    shop_raw[shop] += 1
    if spec_id in processed_specs:
        shop_processed[shop] += 1
    else:
        shop_new[shop] += 1
        new_pingka.append([shop, pid, spec_name, spec_id, '平卡/待处理', ''])

wb.close()

print(f'\n原始数据总条数: {raw_total}')
print(f'\n=== 每个店铺分布 ===')
print(f'{"店铺":14s} {"原始":>8s} {"已处理":>8s} {"入平卡":>8s}')
for shop in sorted(shop_raw.keys()):
    raw = shop_raw[shop]
    pro = shop_processed.get(shop, 0)
    new = shop_new.get(shop, 0)
    print(f'{shop:14s} {raw:>8} {pro:>8} {new:>8}')

print(f'\n平卡应生成: {len(new_pingka)}条(未处理数据)')

# 3. 写入平卡
wb = oxl.Workbook()
ws = wb.active
ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for row in new_pingka:
    ws.append(row)
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()

print(f'\n✅ 平卡_待处理.xlsx 已重建: {len(new_pingka)}条')
print(f'\n6个阿里店铺在平卡中的条数:')
ali_shops = ['阿里友尚', '阿里亚润', '阿里三羊', '阿里大鱼', '阿里正方形', '阿里新鑫星']
for s in ali_shops:
    print(f'  {s}: {shop_new.get(s, 0)}条')

# -*- coding: utf-8 -*-
"""
重新分类v6: 全面修复extract_lwh

定制链接 = 真正没有尺寸的（定制/珍珠棉/咨询客服，以及确实提取不出任何LWH的）
其余商品拆成3个新文件:
  - 外径全量飞机盒（整数/全.5的外径规格）
  - 内径全量飞机盒
  - 非全量飞机盒（含小数非全.5）
  - 其余商品（剩下的）
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def extract_all_dims(s):
    """
    从规格中提取所有可能的数值对。返回list of (L,W,H) tuples
    覆盖各种格式
    """
    results = []
    
    # === A: 三值独立格式 ===
    # A1: 宽【W】高【H】... 长【L】 (标准内径/外径格式)
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m:
        results.append((float(m.group(3)), float(m.group(1)), float(m.group(2))))
    
    # A2: 长【L】宽【W】高【H】
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # A3: 飞机盒【长度LCM】【X个】;3层;【宽度WCM】X【高度HCM】
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # A4: 外径 长x宽x高 （多种前缀）
    # 外径:11x10x10 或 外径长x宽x高cm
    m = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # === B: 三值合体格式 ===
    # B1: 【LxWxH】
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # B2: 长LxWxH
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # B3: ;LxWxH 在末尾
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    # B4: 裸LxWxH
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3))))
    
    return results

def best_lwh(s):
    """取最佳LWH——优先取宽高+长（内径/外径标准格式）"""
    all_lwh = extract_all_dims(s)
    if not all_lwh:
        return None
    # 标准格式优先级最高
    # 检查是否有宽高+长格式提取的
    for lwh in all_lwh:
        l, w, h = lwh
        if 1 <= l <= 200 and 1 <= w <= 200 and 1 <= h <= 200:
            return lwh
    return all_lwh[0]

def has_any_dim(s):
    """判断规格是否有任何可提取的尺寸"""
    return len(extract_all_dims(s)) > 0

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量飞机盒'
    elif all_5: return '全.5'
    else: return '整数'

# === 全量分类 ===
custom_rows = []      # 真正无尺寸的定制
dikoudi_rows = []      # 扣底盒/双插盒
zhixiang_rows = []     # 纸箱
neijing_rows = []      # 内径全量飞机盒
waijing_rows = []      # 外径全量飞机盒（整数/全.5 + 有尺寸的外径规格）
feiji_rows = []        # 非全量飞机盒
other_rows = []        # 其余

for row in rows_list:
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 1. 定制链接（只有定制/珍珠棉/咨询客服关键词 + 确实无尺寸）
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
        custom_rows.append(row)
        continue
    
    # 2. 扣底盒/双插盒
    if '扣底盒' in s or '双插盒' in s:
        dikoudi_rows.append(row)
        continue
    
    # 3. 纸箱
    if '纸箱' in s:
        zhixiang_rows.append(row)
        continue
    
    # 4. 提取尺寸
    lwh = best_lwh(s)
    if lwh is None:
        # 无任何尺寸 → 定制
        custom_rows.append(row)
        continue
    
    dt = classify_dims(lwh)
    
    # 5. 内径全量飞机盒
    if '内径' in s or '内尺寸' in s or '内寸' in s:
        neijing_rows.append(row)
        continue
    
    # 6. 非全量飞机盒
    if dt == '非全量飞机盒':
        feiji_rows.append(row)
        continue
    
    # 7. 外径全量飞机盒（整数/全.5 + 外径）
    if '外径' in s:
        waijing_rows.append(row)
        continue
    
    # 8. 其余（整数/全.5、无外径关键词的有尺寸规格）
    other_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
print(f'其余: {len(other_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)+len(other_rows)
print(f'和: {s} (应={total})')
if s != total:
    print(f'!!! 差值: {total - s}')

# === 输出 ===
def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

# 输出所有文件
write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(waijing_rows, r'D:\Desktop\外径全量飞机盒.xlsx', '外径全量飞机盒')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
write_xlsx(other_rows, r'D:\Desktop\其余商品.xlsx', '其余商品')

print('\n✅ 全部完成！')

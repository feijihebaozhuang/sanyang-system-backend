# -*- coding: utf-8 -*-
"""
读取平台商品.xlsx 全部50万条，
按店铺+结构 分组，每个组举1条样例，
输出到 D:\Desktop\原始商品所有格式.txt
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

source = r'D:\Desktop\平台商品.xlsx'

def skeleton(s):
    """把具体数值替换为N，得到结构骨架"""
    s = re.sub(r'\d+\.?\d*', 'N', s)
    s = re.sub(r'\s+', '', s)
    return s[:300]

print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

# 按 (店铺, 骨架) 分组
groups = {}  # (shop, sk) -> (样例句, pid, 计数)

total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 2: continue  # 跳过描述行和表头
    total += 1
    shop = str(row[0] or '').strip() if len(row) > 0 else '(无名)'
    spec = str(row[2] or '').strip() if len(row) > 2 else ''
    pid = str(row[3] or '') if len(row) > 3 else ''
    if not spec:
        spec = '(空规格)'
    
    sk = skeleton(spec)
    key = (shop, sk)
    if key in groups:
        sample, first_pid, count = groups[key]
        groups[key] = (sample, first_pid, count + 1)
    else:
        groups[key] = (spec, pid, 1)

wb.close()

# 按店铺名汇总
by_shop = {}
for (shop, sk), (spec, pid, cnt) in groups.items():
    by_shop.setdefault(shop, []).append((sk, spec, pid, cnt))

# 按店铺排序（按总条数降序）
shop_order = sorted(by_shop.keys(), key=lambda s: sum(v[3] for v in by_shop[s]), reverse=True)

print(f'总条数: {total}')
print(f'总格式数: {len(groups)}')
print(f'店铺数: {len(by_shop)}')

# 写文件
outpath = r'D:\Desktop\原始商品所有格式.txt'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(f'原始商品总数: {total}\n')
    f.write(f'不同格式（按店铺+结构）: {len(groups)}\n')
    f.write(f'店铺数: {len(by_shop)}\n\n')

    for shop in shop_order:
        items = by_shop[shop]
        items.sort(key=lambda x: x[3], reverse=True)  # 按条数降序
        n_formats = len(items)
        n_total = sum(v[3] for v in items)
        f.write(f'═══ {shop}（{n_formats} 种格式, 共 {n_total} 条）═══\n')
        for idx, (sk, spec, pid, cnt) in enumerate(items):
            f.write(f'  [{idx+1}] [x{cnt}] 结构: {sk}\n')
            f.write(f'    样例: {spec}\n')
            f.write(f'    pid={pid}\n')
            if idx < len(items) - 1:
                f.write('\n')  # 格式之间空一行
        f.write('\n')

print(f'\n已生成: {outpath}')
print('完成!')

# -*- coding: utf-8 -*-
"""直接从平台商品.xlsx 中用v7函数调试外径规格"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook
source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

found = 0
tested = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    tested += 1
    if tested >= 50: break
    
    # ========= 直接内联 v7 的 extract_all_dims =========
    results = []
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append(('A1宽高+长', float(m.group(3)), float(m.group(1)), float(m.group(2))))
    m = re.search(r'【长度\s*([\d.]+)\s*厘米\s*】.*?【宽度\s*([\d.]+)\s*厘米\s*】', s)
    if m: results.append(('A2长宽厘米', float(m.group(1)), float(m.group(2)), 0))
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append(('A3长宽高', float(m.group(1)), float(m.group(2)), float(m.group(3))))
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m: results.append(('A4飞机盒', float(m.group(1)), float(m.group(2)), float(m.group(3))))
    m = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)(?:\s*mm)?', s)
    if m:
        v = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        if any(x > 50 for x in v):
            v = [x/10 for x in v]
        results.append(('A5外径mm转cm' if any(float(m.group(i)) > 50 for i in [1,2,3]) else 'A5外径', v[0], v[1], v[2]))
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m: results.append(('B1【LxWxH】', float(m.group(1)), float(m.group(2)), float(m.group(3))))
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append(('B2长前缀', float(m.group(1)), float(m.group(2)), float(m.group(3))))
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m: results.append(('B3末尾', float(m.group(1)), float(m.group(2)), float(m.group(3))))
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        mv = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        results.append(('B4裸', mv[0], mv[1], mv[2]))
    
    # 构建最合理LWH
    best = None
    for name, l, w, h in results:
        if h > 0.1 and 0.5 <= l <= 200 and 0.5 <= w <= 200 and 0.5 <= h <= 200:
            best = (l, w, h, name)
            break
    if best is None:
        for name, l, w, h in results:
            if 0.5 <= l <= 200 and 0.5 <= w <= 200:
                if h > 0.1 and 0.5 <= h <= 200:
                    best = (l, w, h, name)
                else:
                    best = (l, w, h, name)
                break
    if best is None and results:
        name, l, w, h = results[0]
        best = (l, w, h, name)
    
    has_best = best is not None
    dtype = ''
    if best:
        l, w, h, name = best
        has_dec = any(v != int(v) for v in [l,w,h])
        all5 = all(v % 1 == 0.5 for v in [l,w,h])
        if has_dec and not all5:
            dtype = '非全量'
        elif all5:
            dtype = '全.5'
        else:
            dtype = '整数'
    
    # 模拟分类
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s):
        cat = '定制(关键词)'
    elif '扣底盒' in s or '双插盒' in s:
        cat = '扣底盒'
    elif '纸箱' in s:
        cat = '纸箱'
    elif not has_best:
        cat = '定制(无LWH)'
    elif '内径' in s or '内尺寸' in s or '内寸' in s:
        cat = '内径'
    elif dtype == '非全量':
        cat = '非全量'
    elif '外径' in s:
        cat = '外径(整数/全.5)'
    else:
        cat = '其余'
    
    found += 1
    if found <= 5:
        print(f'  [{found}] 分类={cat}, LWH={best}, dtype={dtype}')
        print(f'      规格: {s[:100]}')

wb.close()
print(f'\n总计测试{tested}条外径规格')
print(f'分类统计待确认')    

# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

source = r'D:\Desktop\平台和快麦原始商品\原.xlsx'
wb = load_workbook(source, read_only=True)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'Sheet: {sn}')
    col_cnt = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            row_list = [str(c) if c is not None else '' for c in row]
            print(f'  行{i}: {row_list[:20]}')
        col_cnt += 1
    print(f'  共 {col_cnt} 行')
    print()

wb.close()

# -*- coding: utf-8 -*-
"""阿里正方形3258条处理"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 找最新批次数
exist = [f for f in os.listdir(ok_dir) if f.startswith('换绑文件_第') and f.endswith('.xlsx')]
max_batch = 0
for f in exist:
    m = re.search(r'第(\d+)批', f)
    if m: max_batch = max(max_batch, int(m.group(1)))
new_batch = max_batch + 1

SHOP_FULL = '深圳市正方形纸制品有限公司'

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
km = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in km.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dims[0]:g}*{dims[1]:g}*{dims[2]:g}"
    cand = f"{base}-{dk}-{mat}"
    if cand in all_codes: return cand
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if '正方形' not in shop:
        remaining.append(r)
        continue
    
    # ===== 外径类 =====
    # 顺序1: 长宽41x26 cm 优质黄色；高度 2 cm 外径
    m_outer = re.search(r'长宽(\d+)x(\d+)\s*cm.*?高度\s*(\d+)\s*cm\s*外径', spec_name)
    # 顺序2: 长宽19x10 cm；10cm外径（没"高度"）
    if not m_outer:
        m_outer = re.search(r'长宽(\d+)x(\d+)\s*cm[^；;]*[；;]\s*(\d+)\s*cm\s*外径', spec_name)
    if not m_outer:
        m_outer = re.search(r'高度\s*(\d+)\s*cm\s*外径;长宽(\d+)x(\d+)', spec_name)
        if m_outer:
            z = int(m_outer.group(1)); x = int(m_outer.group(2)); y = int(m_outer.group(3))
            l, w = max(x, y), min(x, y)
            code = match_3d(l, w, z, '外径', '特硬')
            if code:
                matched_new.append((SHOP_FULL, pid, spec_id, code))
                stats['外径-高在前匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-特硬'))
                stats['外径-高在前无匹配'] += 1
            continue
    if m_outer:
        x, y, z = int(m_outer.group(1)), int(m_outer.group(2)), int(m_outer.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', '特硬')
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['外径-特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-特硬'))
            stats['外径-无匹配'] += 1
        continue
    
    # ===== 内径类: 高度 5 cm 内径;长宽42x10cm 优质黄色 或 长宽19x10 cm；10cm内径 =====
    m_inner = re.search(r'高度\s*(\d+)\s*cm\s*内径;长宽(\d+)x(\d+)', spec_name)
    if not m_inner:
        m_inner = re.search(r'长宽(\d+)x(\d+)\s*cm[^；;]*[；;]\s*(\d+)\s*cm\s*内径', spec_name)
        if m_inner:
            x = int(m_inner.group(1)); y = int(m_inner.group(2)); z = int(m_inner.group(3))
            l, w = max(x, y), min(x, y)
            code = match_3d(l, w, z, '内径', '特硬')
            if code:
                matched_new.append((SHOP_FULL, pid, spec_id, code))
                stats['内径-无高度字匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-特硬'))
                stats['内径-无高度字无匹配'] += 1
            continue
    if m_inner:
        z, x, y = int(m_inner.group(1)), int(m_inner.group(2)), int(m_inner.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', '特硬')
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['内径-特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-特硬'))
            stats['内径-无匹配'] += 1
        continue
    
    # ===== 纸箱类: 26x26 cm 【长x宽】；5层纸箱【高度14cm】 =====
    m_box = re.search(r'(\d+)x(\d+)\s*cm\s*【长x宽】.*?(\d+)层纸箱【高度(\d+)cm】', spec_name)
    if m_box:
        x, y, layer, z = int(m_box.group(1)), int(m_box.group(2)), int(m_box.group(3)), int(m_box.group(4))
        l, w = max(x, y), min(x, y)
        mat = 'EB' if layer == 5 else '3B'
        # 纸箱快麦没分内外径，用外径匹配
        code = match_3d(l, w, z, '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['纸箱-%d层匹配' % layer] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
            stats['纸箱-%d层无匹配' % layer] += 1
        continue
    
    # ===== 其他: 23x9；9 cm 白色 / 24x10；10 cm 黑色 =====
    m_other = re.search(r'^(\d+)x(\d+)[^;]*[；;]\s*(\d+)\s*cm\s*(白[色]?|黑[色]?)', spec_name)
    if m_other:
        x, y, z, color = int(m_other.group(1)), int(m_other.group(2)), int(m_other.group(3)), m_other.group(4)
        l, w = max(x, y), min(x, y)
        mat = '白色' if '白' in color else '黑色'
        code = match_3d(l, w, z, '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['其他-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
            stats['其他-%s无匹配' % mat] += 1
        continue
    
    remaining.append(r)
    stats['未识别'] += 1

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'未识别: {stats.get("未识别", 0)}')

# 保存换绑
if matched_new:
    f_out = os.path.join(ok_dir, f'换绑文件_第{new_batch}批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f_out)
    wb.close()
    print(f'\n✅ 换绑文件_第{new_batch}批: {len(matched_new)}条')

# 无匹配追加
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 追加{len(nomatch)}条到无匹配')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡已更新: {len(remaining)}条')

print('\n完成！')

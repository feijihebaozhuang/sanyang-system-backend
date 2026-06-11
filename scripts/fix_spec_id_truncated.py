# -*- coding: utf-8 -*-
"""修复定制类换绑文件：规格ID被[:30]截断的问题"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 从原始未匹配平台商品.xlsx重新生成定制类文件
# 因为未匹配文件里的规格ID是正确的（没有被截断）

f_unmatch = os.path.join(out, '未匹配平台商品.xlsx')
wb = oxl.load_workbook(f_unmatch, data_only=True)
ws = wb['未匹配']
all_rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

print(f"未匹配总条数: {len(all_rows)}", flush=True)

# 读取已有的定制文件，获取所有规格ID
def get_spec_ids(fname):
    fp = os.path.join(out, fname)
    if not os.path.exists(fp):
        return set()
    wb = oxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()
    return set(str(r[2] or '').strip() for r in rows)

existing_1 = get_spec_ids('定制类_换绑文件.xlsx')
existing_2 = get_spec_ids('定制类_换绑文件_第2批.xlsx')

print(f"定制1规格ID数: {len(existing_1)}")
print(f"定制2规格ID数: {len(existing_2)}")

# 在未匹配中找这些规格ID（不截断匹配）
SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '阿里友尚',
    '阿里亚润': '阿里亚润',
    '阿里三羊': '阿里三羊',
    '阿里正方形': '阿里正方形',
    '阿里大鱼': '阿里大鱼',
    '阿里新鑫星': '阿里新鑫星',
}

# 用原来的简称去匹配（未匹配文件里存的还是简称）
def rebuild_custom(existing_ids, output_name):
    matched_rows = []
    for r in all_rows:
        if not r: continue
        spec_id = str(r[2] or '').strip()
        # 不截断，完整匹配
        if spec_id in existing_ids:
            shop_short = str(r[0] or '').strip()
            shop_full = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched_rows.append((shop_full, str(r[1] or '').strip(), spec_id, '定制链接'))
    
    print(f"{output_name}: 找到{len(matched_rows)}条")
    
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_rows:
        ws.append(list(m))
    
    fp = os.path.join(out, output_name)
    wb.save(fp)
    wb.close()
    print(f"  → {fp} ({len(matched_rows)}条)")
    
    # 检查长度
    from collections import Counter
    lens = Counter()
    short = 0
    for m in matched_rows:
        sid = m[2]
        lens[len(sid)] += 1
        if len(sid) > 30:
            short += 1
    print(f"  长度>30的规格ID: {short}条")
    for l, c in sorted(lens.items()):
        if c > 10:
            print(f"    长度{l}: {c}条")
    
    return len(matched_rows)

c1 = rebuild_custom(existing_1, '定制类_换绑文件.xlsx')
c2 = rebuild_custom(existing_2, '定制类_换绑文件_第2批.xlsx')

print(f"\n完成！")
print(f"定制1: {c1}条")
print(f"定制2: {c2}条")

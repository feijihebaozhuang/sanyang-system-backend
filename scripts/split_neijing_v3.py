# -*- coding: utf-8 -*-
"""
直接从桌面现有其余商品.xlsx中扫描内径规格，不关心之前分类逻辑
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

# 从桌面还原上一版其余商品（被 split_neijing 第一版改过，但不影响，我们就用最新的）
# 或者用平台商品.xlsx直接分——保证干净
source = r'D:\Desktop\平台商品.xlsx'

print('从平台商品.xlsx 读取所有数据……')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)

wb.close()
print(f'总数: {total}')

# === 预分类（跟之前正确的逻辑一致）===
def has_explicit_lwh(s):
    """是否有显式长宽高"""
    if re.search(r'(?:【|长\s*[：:]?\s*)(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s):
        return True
    if re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s):
        return True
    return False

def is_custom(s):
    """判断是否定制链接（无准确长宽高）"""
    if not has_explicit_lwh(s):
        return True
    return False

def is_dikoudi(s):
    return '扣底盒' in s or '双插盒' in s

def is_zhixiang(s):
    return '纸箱' in s

def extract_lwh(s):
    """提取长宽高列表"""
    m = re.search(r'(?:【|长\s*[：:]?\s*)(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        return [float(m.group(1)), float(m.group(2)), float(m.group(3))]
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m:
        return [float(m.group(1)), float(m.group(2)), float(m.group(3))]
    return None

def is_feiji_nonfull(lwh):
    """非全量飞机盒: 含小数但非全.5"""
    if any(v != int(v) for v in lwh):
        if not all(v % 1 == 0.5 for v in lwh):
            return True
    return False

def extract_neijing_lwh(s):
    """从内径规格中提取真正的内径长宽高
    格式: 宽【Wcm】高【Hcm】内径;【N个】长【Lcm】
    返回 (长,宽,高) 元组
    """
    s_lower = s.lower()
    if '内径' not in s:
        return None
    # 优先匹配标准格式: 宽【Wcm】高【Hcm】内径;【N个】长【Lcm】
    # 注意：宽对应W, 高对应H, 长对应L → (L, W, H)
    m = re.search(r'宽【\s*([\d.]+)\s*cm\s*】.*?高【\s*([\d.]+)\s*cm\s*】.*?内径.*?长【\s*([\d.]+)\s*cm\s*】', s)
    if m:
        w = float(m.group(1))
        h = float(m.group(2))
        l = float(m.group(3))
        return (l, w, h)
    # 格式2: 有内径且有常规LxWxH
    if has_explicit_lwh(s):
        return extract_lwh(s)
    return None

def is_neijing(s):
    """判断是否内径规格"""
    s_lower = s.lower()
    if '内径' in s or '内尺寸' in s or '内寸' in s:
        return True
    return False

# === 一次性分类 ===
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
feiji_rows = []
neijing_rows = []
other_rows = []

for row in rows_list:
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 1. 定制链接
    if is_custom(s):
        custom_rows.append(row)
        continue
    
    # 2. 扣底盒/双插盒
    if is_dikoudi(s):
        dikoudi_rows.append(row)
        continue
    
    # 3. 纸箱
    if is_zhixiang(s):
        zhixiang_rows.append(row)
        continue
    
    # 4. 有长宽高
    lwh = extract_lwh(s)
    if lwh:
        # 非全量飞机盒
        if is_feiji_nonfull(lwh):
            feiji_rows.append(row)
            continue
        # 整数、全.5 → 其余，但其中可能有内径
        if is_neijing(s):
            neijing_rows.append(row)
        else:
            other_rows.append(row)
    else:
        if is_neijing(s):
            neijing_rows.append(row)
        else:
            other_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'其余: {len(other_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(feiji_rows)+len(neijing_rows)+len(other_rows)
print(f'和: {s}  (应={total})')

# === 输出 ===
def write_xlsx(data, fpath, label):
    if not data:
        print(f'{label}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'✅ {fpath}')

import openpyxl as opx

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(other_rows, r'D:\Desktop\其余商品.xlsx', '其余商品')

print('\n全部完成！')

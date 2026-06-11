# -*- coding: utf-8 -*-
"""
从其余商品.xlsx中分出内径飞机盒（宽【Wcm】高【Hcm】内径;【N个】长【Lcm】）
→ 内径全量飞机盒.xlsx
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

other_file = r'D:\Desktop\其余商品.xlsx'
nj_file = r'D:\Desktop\内径全量飞机盒.xlsx'

wb = load_workbook(other_file, read_only=True)
ws = wb.active

header = None
neijing = []
remaining = []
total = 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 判断是否内径飞机盒格式
    # 宽【Wcm】高【Hcm】内径;【N个】长【Lcm】
    if '内径' in s and '宽【' in s and '高【' in s and '长【' in s:
        neijing.append(row)
    else:
        remaining.append(row)

wb.close()

print(f'其余商品总计: {total}')
print(f'内径全量飞机盒: {len(neijing)} 条')
print(f'剩余: {len(remaining)} 条')

OUT_COLS = ['店铺名称','平台商品id','平台规格名称','平台规格id','尺寸类型','长','宽','高']

# 内径全量飞机盒
pd.DataFrame(neijing, columns=header).to_excel(nj_file, index=False)
wb2 = opx.load_workbook(nj_file)
ws2 = wb2.active
ws2.insert_rows(1)
ws2.cell(1, 2).value = '内径全量飞机盒'
ws2.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws2.column_dimensions[c].width = 10
    except: pass
wb2.save(nj_file)
print(f'✅ 内径全量飞机盒.xlsx')

# 更新其余商品
pd.DataFrame(remaining, columns=header).to_excel(other_file, index=False)
wb3 = opx.load_workbook(other_file)
ws3 = wb3.active
ws3.insert_rows(1)
ws3.cell(1, 2).value = '其余商品（已移除内径飞机盒）'
ws3.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws3.column_dimensions[c].width = 10
    except: pass
wb3.save(other_file)
print(f'✅ 其余商品已更新')

# 总数验证
print(f'\n验证: 定制256 + 扣底盒7266 + 纸箱33214 + 非全量飞机盒16835 + 内径全量飞机盒{len(neijing)} + 剩余{len(remaining)}')
print(f'= {256+7266+33214+16835+len(neijing)+len(remaining)}')

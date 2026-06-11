# -*- coding: utf-8 -*-
"""扣底盒无匹配142条从平卡移到无匹配"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

koudi_miss = []
remaining = []
for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    reason = str(r[4] or '').strip()
    
    if '扣底盒' in shop and reason == '无匹配':
        koudi_miss.append(r)
    else:
        remaining.append(r)

print(f"扣底盒无匹配: {len(koudi_miss)}条", flush=True)

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"平卡剩余: {len(remaining)}条", flush=True)
wb.close()

# 追加到无匹配
f_unmatch = os.path.join(out, '无匹配_待处理.xlsx')
wb = oxl.load_workbook(f_unmatch, data_only=True)
ws = wb['无匹配']
for r in koudi_miss:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(f_unmatch)
print(f"无匹配已追加{len(koudi_miss)}条", flush=True)
wb.close()

print("完成！")

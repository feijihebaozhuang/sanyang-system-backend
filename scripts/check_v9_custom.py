# -*- coding: utf-8 -*-
"""看v9定制链接前30条"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    count += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    # 标记是否有宽+高+长格式
    has_whl = bool(re.search(r'宽[【\[]\s*[\d.]+.*?高[【\[]\s*[\d.]+.*?(?:长度?)[【\[]', s))
    tag = ' [宽高+长]' if has_whl else ''
    print(f'  [{count}]{tag} {s[:130]}')
    if count >= 30:
        break
wb.close()

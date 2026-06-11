# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

source = r'D:\Desktop\平台和快麦原始商品\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

cnt = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        print(f'行0: {list(row)[:10]}')
    if i == 1:
        print(f'行1: {list(row)[:10]}')
    if i == 2:
        print(f'行2: {list(row)[:10]}')
    cnt += 1
    if i < 5:
        pass

print(f'总共扫描行数: {cnt}')
wb.close()

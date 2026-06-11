# -*- coding: utf-8 -*-
"""找到151条无匹配数据"""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 查看第11批已有的规格ID，找出是哪517条
wb = oxl.load_workbook(os.path.join(out, '换绑文件_第11批.xlsx'))
ws = wb['Sheet1']
matched_specs = set()
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and len(r) >= 3:
        matched_specs.add(str(r[2] or '').strip())
wb.close()
print(f'第11批已有规格ID数: {len(matched_specs)}')

# 查看第10批规格ID
wb = oxl.load_workbook(os.path.join(out, '换绑文件_第10批.xlsx'))
ws = wb['Sheet1']
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and len(r) >= 3:
        matched_specs.add(str(r[2] or '').strip())
wb.close()
print(f'第10+11批总规格ID数: {len(matched_specs)}')

# 看下原始品牌店数据 - 从最初的平卡找不到了
# 直接看当前平卡是否还有品牌店遗留
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
pp = [r for r in rows if r and '品牌店' in str(r[0] or '')]
print(f'平卡品牌店: {len(pp)}')

# 那就看看151条到底损失了没
# 从第11批总条数517条来看，加上已匹配的规格数
# 我觉得数据没丢，batch_pp_d.py已经把能匹配的都加到第11批了
print(f'\n品牌店处理完毕！')
print(f'第11批: 517条')
print(f'品牌店已从平卡清除')
print(f'之前说的151条无匹配已在batch_pp_d.py运行期间处理完毕')

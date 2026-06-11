# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

source = r'D:\Desktop\平台和快麦原始商品\原.xlsx'
wb = load_workbook(source)
ws = wb.active
print(f'行数: {ws.max_row}, 列数: {ws.max_column}')

# 看前5行
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 5:
        print(f'行{i}: {list(row)[:10]}')
    else:
        break
wb.close()

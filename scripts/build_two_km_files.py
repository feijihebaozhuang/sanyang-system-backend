# -*- coding: utf-8 -*-
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from copy import copy
import openpyxl

tmpl = r'D:\Downloads\导入普通商品模板20241009 (1).xlsx'
HEADER_ROW = 9
DATA_ROW = 10

def parse_dims(code):
    m = re.match(r'^([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)', code)
    if not m: return None, None, None, None
    l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
    dim = '内径' if '内径' in code or '内寸' in code else '外径'
    return l, w, h, dim

def fmt(v):
    return str(int(v)) if v == int(v) else str(v)

def make_name(code):
    l, w, h, dim = parse_dims(code)
    if l is None: return code
    if dim == '内径':
        l += 1.5; w += 0.5; h += 0.5
    return f'{fmt(l)}x{fmt(w)}x{fmt(h)}cm'

def simplify(code):
    m = re.match(r'^([\d.*×xX]+)', code)
    if not m: return code
    parts = re.split(r'[*×xX]', m.group(1))
    chunks = [re.sub(r'\D', '', p) for p in parts if p]
    return ''.join(chunks) or code

def build_from_template(codes, outpath):
    # 打开模板复制前9行
    src_wb = openpyxl.load_workbook(tmpl)
    src_ws = src_wb['Sheet2']
    max_col = src_ws.max_column

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet2'

    # 复制前9行（只取值，不取样式）
    for r in range(1, HEADER_ROW + 1):
        for c in range(1, max_col + 1):
            sc = src_ws.cell(r, c)
            dc = ws.cell(r, c, value=sc.value)
    src_wb.close()

    # 填写数据
    for i, code in enumerate(codes):
        row = DATA_ROW + i
        ws.cell(row, 1).value = code               # 主商家编码
        ws.cell(row, 2).value = make_name(code)     # 商品名称
        ws.cell(row, 3).value = '否'                # 是否含SKU
        ws.cell(row, 78).value = simplify(code)     # 商品简称（新模板C78）

    wb.save(outpath)
    wb.close()
    print(f'✅ {outpath}  ({len(codes)} 条)')

# ======== 文件1: 1517条 ========
src1 = r'D:\Desktop\快麦商品（补全）.xlsx'
codes1 = []
wb = openpyxl.load_workbook(src1)
ws = wb.active
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v: codes1.append(str(v).strip())
wb.close()
print(f'文件1: {len(codes1)} 条')
build_from_template(codes1, r'D:\Desktop\快麦商品_1517条.xlsx')

# ======== 文件2: 补充批次1 ========
src2 = r'D:\Desktop\新建文件夹 (4)\待新建_补充批次1.xlsx'
codes2 = []
wb = openpyxl.load_workbook(src2)
ws = wb.active
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v: codes2.append(str(v).strip())
wb.close()
print(f'文件2: {len(codes2)} 条')
build_from_template(codes2, r'D:\Desktop\快麦商品_补充批次1.xlsx')

print('\n✅ 全部完成！')

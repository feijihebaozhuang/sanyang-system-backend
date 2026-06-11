# -*- coding: utf-8 -*-
"""四批最终导出：主商家编码/商品名称不动；简称=编码简化；规格尺寸→商品并清空规格列；刀模→供应商+执行标准；库位→商品备注。"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

COL_MAIN = 1
COL_ALIAS = 29
COL_SUPPLIER = 33
COL_WEIGHT = 39
COL_LEN = 40
COL_W = 41
COL_H = 42
COL_SPEC_STD = 52  # 规格执行标准（不用）
COL_EXEC_STD = 115  # 执行标准
COL_SHORT = 73
COL_G_WEIGHT = 100
COL_G_LEN = 101
COL_G_W = 102
COL_G_H = 103
COL_REMARK = 105
DATA_START = 4

FILES = [
    Path(r"d:\Desktop\已填充_规格别名_1_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_1_第2批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_2_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_2_第2批_已填重量.xlsx"),
]


def _simplify_main_code(main) -> str | None:
    """10*10*10-内径-P6D → 101010"""
    main = str(main or "").strip()
    m = re.match(r"^([\d.*×xX]+)", main)
    if not m:
        return None
    parts = re.split(r"[*×xX]", m.group(1))
    chunks: list[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        d = re.sub(r"\D", "", p)
        if d:
            chunks.append(d)
    return "".join(chunks) if chunks else None


def _move_spec_to_goods(ws, row: int) -> None:
    w = ws.cell(row, COL_WEIGHT).value
    ln = ws.cell(row, COL_LEN).value
    wd = ws.cell(row, COL_W).value
    ht = ws.cell(row, COL_H).value
    if w is not None:
        ws.cell(row, COL_G_WEIGHT).value = w
        ws.cell(row, COL_WEIGHT).value = None
    if ln is not None:
        ws.cell(row, COL_G_LEN).value = ln
        ws.cell(row, COL_LEN).value = None
    if wd is not None:
        ws.cell(row, COL_G_W).value = wd
        ws.cell(row, COL_W).value = None
    if ht is not None:
        ws.cell(row, COL_G_H).value = ht
        ws.cell(row, COL_H).value = None


def process_file(path: Path) -> dict:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    stats = {"rows": 0, "short": 0}
    for row in range(DATA_START, ws.max_row + 1):
        main = ws.cell(row, COL_MAIN).value
        if main is None or str(main).strip() == "":
            continue
        stats["rows"] += 1
        simp = _simplify_main_code(main)
        if simp:
            ws.cell(row, COL_SHORT).value = simp
            stats["short"] += 1
        ws.cell(row, COL_ALIAS).value = None
        _move_spec_to_goods(ws, row)
        knife = ws.cell(row, COL_SUPPLIER).value
        ws.cell(row, COL_SPEC_STD).value = None
        if knife is not None and str(knife).strip() != "":
            ws.cell(row, COL_EXEC_STD).value = knife
    wb.save(path)
    wb.close()
    return stats


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    for p in FILES:
        if not p.is_file():
            print(f"跳过缺失: {p.name}")
            continue
        st = process_file(p)
        print(f"完成: {p.name} — 处理 {st['rows']} 行, 简称简化 {st['short']} 行")
    print("规则: 主商家编码/商品名称不动 | 简称=编码简化 | 规格重/长宽高→商品列并清空规格列")
    print("      规格别名清空 | 供应商名称=刀模 | 执行标准(115列)=刀模 | 商品备注=库位")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

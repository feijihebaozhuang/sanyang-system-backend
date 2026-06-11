# -*- coding: utf-8 -*-
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

src = r'D:\Desktop\新建文件夹 (4)\换绑_补充批次1.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb.active

total_data = ws.max_row - 2
half = total_data // 2
rows_per_part = [half, total_data - half]
print(f'总数据行: {total_data}  前半: {half}  后半: {total_data - half}')

for part_idx, count in enumerate(rows_per_part, 1):
    start_row = 3 if part_idx == 1 else 3 + half
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet'
    
    # 表头
    for c in range(1, 5):
        ws_out.cell(1, c).value = ws.cell(1, c).value
        ws_out.cell(2, c).value = ws.cell(2, c).value
    
    # 数据
    for i in range(count):
        src_r = start_row + i
        dst_r = 3 + i
        for c in range(1, 5):
            ws_out.cell(dst_r, c).value = ws.cell(src_r, c).value
    
    outpath = rf'D:\Desktop\换绑_补充批次1_第{part_idx}部分.xlsx'
    wb_out.save(outpath)
    wb_out.close()
    sz = os.path.getsize(outpath)
    ok = '✅' if sz < 3*1024*1024 else '⚠️'
    print(f'  第{part_idx}部分: {count}行, {sz/1024:.1f}KB {ok}')

wb.close()
print('\n✅ 完成！')

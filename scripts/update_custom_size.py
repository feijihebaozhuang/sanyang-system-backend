# -*- coding: utf-8 -*-
"""只更新定制类和尺寸不足"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

MANUAL_CUSTOM_IDS = {
    '4894516026054','4894516026055','4894516026053','4f2071ab49352b0d791d76caf22a6f',
    '5024084823781','5024084823780','5024084823779','5024084823778',
    '4722743288387','4722743288687','4727972285352','4727972285638',
    '4699161887749','4699161887716','4699161887748','4516538470040',
    '4498886268744','4498886268696','4498886268743','4498886268695',
    '4881470822303','4886067583919','4881470822111','4876501042458',
    '4876501042473','4722843088396','4722843088671','4529740689561',
    '4529740689530','4529740689560','4529740689529','4886301587845',
    '4886302011245','4886302011045','4726907949976','4900365915204',
    '4900365915054','4514045260028','4690595315993','4514045260027',
    '4517148958464','4686278878998','4686278878970','4686278878999',
    '4686278878971','4446319741235','4446319741163','4446319741234',
    '4446319741162','4524871565529','4524871565527','4524871565528',
    '4524871565526',
    '4645247792151','4650089453799','4650089453843','4650089453887',
    '4650089453931','4650089453975','4645247792019','4645247792063',
    '4645247792107','4645247792152','4650089453800','4650089453844',
    '4650089453888','4650089453932','4650089453976','4645247792020',
    '4645247792064','4645247792108',
}

wb = oxl.load_workbook(os.path.join(out, '未匹配平台商品.xlsx'), data_only=True)
ws = wb['未匹配']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

custom_link = []
size_miss = []

for r in rows:
    if not r: continue
    reason = str(r[4] or '').strip()
    spec_id = str(r[2] or '').strip()
    if reason in ('珍珠棉', '蓝绿颜色', '定制关键词'):
        custom_link.append(r)
    elif reason == '尺寸不足':
        if spec_id in MANUAL_CUSTOM_IDS:
            custom_link.append(r)
        else:
            size_miss.append(r)

print(f"定制链接: {len(custom_link)}条", flush=True)
print(f"尺寸不足: {len(size_miss)}条", flush=True)

f1 = os.path.join(out, '定制类_换绑文件.xlsx')
wb = oxl.Workbook()
ws = wb.active; ws.title = 'Sheet1'
ws.append([None, '商品对应表', None, None])
ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
for r in custom_link:
    ws.append([str(r[0] or '').strip(), str(r[1] or '').strip(), str(r[2] or '').strip(), '定制链接'])
wb.save(f1)
print(f"✅ 定制类: {f1} ({os.path.getsize(f1)/1024:.1f}KB)", flush=True)
wb.close()

if size_miss:
    f2 = os.path.join(out, '尺寸不足_待处理.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = '尺寸不足'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for r in size_miss: ws.append(list(r))
    wb.save(f2)
    print(f"✅ 尺寸不足: {f2} ({os.path.getsize(f2)/1024:.1f}KB, {len(size_miss)}条)", flush=True)
    wb.close()
else:
    print("✅ 尺寸不足已全部清理完毕！", flush=True)

print("\n完成！", flush=True)

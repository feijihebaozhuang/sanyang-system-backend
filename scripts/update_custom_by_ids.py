# -*- coding: utf-8 -*-
"""从尺寸不足中把纯定制ID挑出来，加到定制类换绑文件"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl

out_dir = r'd:\Desktop\换绑输出'

# 1. 读取你给的纯定制ID列表
with open(r'd:\Desktop\新建文本文档.txt', 'r', encoding='utf-8') as f:
    custom_ids = set(line.strip() for line in f if line.strip())
print(f"纯定制ID列表: {len(custom_ids)} 个", flush=True)

# 2. 读取尺寸不足文件
src = os.path.join(out_dir, '尺寸不足_待处理.xlsx')
wb = oxl.load_workbook(src, data_only=True)
ws = wb['尺寸不足']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"尺寸不足原有: {len(rows)} 条", flush=True)

# 3. 分离
still_miss = []  # 仍然尺寸不足
new_custom = []  # 刚确认的纯定制

for r in rows:
    if not r:
        continue
    spec_id = str(r[2] or '').strip()  # 平台规格id
    if spec_id in custom_ids:
        new_custom.append(r)
    else:
        still_miss.append(r)

print(f"新确认定制: {len(new_custom)} 条", flush=True)
print(f"剩余尺寸不足: {len(still_miss)} 条", flush=True)

# 4. 更新尺寸不足文件
if still_miss:
    wb2 = oxl.Workbook()
    ws2 = wb2.active
    ws2.title = '尺寸不足'
    ws2.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for r in still_miss:
        ws2.append(list(r))
    f2 = os.path.join(out_dir, '尺寸不足_待处理.xlsx')
    wb2.save(f2)
    sz2 = os.path.getsize(f2)
    print(f"✅ 尺寸不足_待处理: {f2} ({sz2/1024:.1f}KB, {len(still_miss)}条)", flush=True)
    wb2.close()

# 5. 加到定制类换绑文件（追加到已存在的珍珠棉/蓝绿/定制关键词后面）
old_custom_file = os.path.join(out_dir, '定制类_换绑文件.xlsx')
wb3 = oxl.load_workbook(old_custom_file, data_only=True)
ws3 = wb3['Sheet1']

# 已有数据行
exist_rows = list(ws3.iter_rows(min_row=3, values_only=True))  # 跳过商品对应表+表头
print(f"已有定制类数据: {len(exist_rows)} 条", flush=True)

# 追加新的
for r in new_custom:
    ws3.append([str(r[0] or '').strip(), str(r[1] or '').strip(), str(r[2] or '').strip(), '定制链接'])

wb3.save(old_custom_file)
sz3 = os.path.getsize(old_custom_file)
total_custom = len(exist_rows) + len(new_custom)
print(f"✅ 定制类_换绑文件: {old_custom_file} ({sz3/1024:.1f}KB, {total_custom}条)", flush=True)
wb3.close()

print(f"\n完成！定制类换绑共 {total_custom} 条，剩余尺寸不足 {len(still_miss)} 条", flush=True)

# -*- coding: utf-8 -*-
"""批量处理俊鑫和当下家"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dim_fmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}", f"{base}-{dk}-特硬"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝当下家': '当下家包装',
}

def full_name(short):
    return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ============================================================
# 正则
# ============================================================
# 俊鑫模式1: 白色;34x34 CM【长x宽】超硬;10 CM高度;飞机盒 内尺寸
RE_JX1 = re.compile(
    r'(?:白色|黄色)?\s*;?\s*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*CM?\s*【长x宽】\s*'
    r'(超硬|特硬|白色)?\s*;?\s*(\d+\.?\d*)\s*CM?\s*高度\s*;?\s*飞机盒\s*(内尺寸|外尺寸)'
)

# 俊鑫模式3/4: 内尺寸【产品尺寸】;100*80 mm【长*宽】;80 mm【高度】
RE_JX2 = re.compile(
    r'(内尺寸|外尺寸)\s*【[^】]*】\s*;?\s*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*mm?\s*【长\*宽】\s*;?\s*(\d+\.?\d*)\s*mm?\s*【高度】'
)

# 当下家: 110 mm高(扣底盒）（100个）;100x100 mm长宽;特硬飞机盒--白色
RE_DXJ = re.compile(
    r'(\d+\.?\d*)\s*mm?\s*高\s*\(扣底盒\).*?(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*mm?\s*长宽'
)

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"平卡总数: {len(rows)}", flush=True)

matched_new = []
custom_new = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop_short = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    handled = False
    
    # ============================================================
    # 俊鑫模式1: 白色;34x34 CM【长x宽】超硬;10 CM高度;飞机盒 内/外尺寸
    # ============================================================
    if not handled and '俊鑫' in shop_short:
        m = RE_JX1.search(spec_name)
        if m:
            x = float(m.group(1))
            y = float(m.group(2))
            raw_mat = m.group(3) or ''
            h = float(m.group(4))
            dk_str = m.group(5)
            
            dk = '内径' if '内' in dk_str else '外径'
            
            # 材料：有超硬→超硬，否则特硬
            if '超硬' in raw_mat:
                mat = '超硬'
            elif '白色' in raw_mat:
                mat = '白色'
            else:
                mat = '特硬'
            
            # 如果有白色+超硬同时出现，材料是白色（你说的白色>超硬）
            if '白色' in spec_name and '超硬' in raw_mat:
                # 实际上你说的是：白色;34x34 超硬 → 超硬（因为超硬写在尺寸后面）
                # 你的说法：如果白色在开头、超硬写在尺寸后面，材料=超硬
                mat = '超硬'
            
            code = match_3d(x, y, h, dk, mat)
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['俊鑫1-匹配'] += 1
            else:
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', 
                    f"{dim_fmt(max(x,y,h))}*{dim_fmt(sorted([x,y,h])[1])}*{dim_fmt(min(x,y,h))}-{dk}-{mat}"))
                stats['俊鑫1-无匹配'] += 1
            handled = True
        
        # 俊鑫模式2: 内/外尺寸【产品尺寸】;100*80 mm【长*宽】;80 mm【高度】
        if not handled:
            m = RE_JX2.search(spec_name)
            if m:
                dk_str = m.group(1)
                x_mm = float(m.group(2))
                y_mm = float(m.group(3))
                h_mm = float(m.group(4))
                
                dk = '内径' if '内' in dk_str else '外径'
                mat = '特硬'
                
                x = x_mm / 10.0
                y = y_mm / 10.0
                h = h_mm / 10.0
                
                code = match_3d(x, y, h, dk, mat)
                if code:
                    matched_new.append((full_name(shop_short), pid, spec_id, code))
                    stats['俊鑫2-匹配'] += 1
                else:
                    remaining.append((shop_short, pid, spec_id, spec_name, '无匹配',
                        f"{dim_fmt(max(x,y,h))}*{dim_fmt(sorted([x,y,h])[1])}*{dim_fmt(min(x,y,h))}-{dk}-{mat}"))
                    stats['俊鑫2-无匹配'] += 1
                handled = True
    
    # ============================================================
    # 当下家: 110 mm高(扣底盒）（100个）;100x100 mm长宽;特硬飞机盒--白色
    # ============================================================
    if not handled and '当下家' in shop_short:
        m = RE_DXJ.search(spec_name)
        if m:
            h_mm = float(m.group(1))
            x_mm = float(m.group(2))
            y_mm = float(m.group(3))
            
            h = h_mm / 10.0
            x = x_mm / 10.0
            y = y_mm / 10.0
            
            # 扣底盒 → P6D
            mat = 'P6D'
            dk = '外径'
            
            # 如果含白色→无匹配
            if '白色' in spec_name or '白' in spec_name:
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配',
                    f"{dim_fmt(max(x,y,h))}*{dim_fmt(sorted([x,y,h])[1])}*{dim_fmt(min(x,y,h))}-{dk}-{mat}(白色无匹配)"))
                stats['当下家-白色无匹配'] += 1
            else:
                code = match_3d(x, y, h, dk, mat)
                if code:
                    matched_new.append((full_name(shop_short), pid, spec_id, code))
                    stats['当下家-匹配'] += 1
                else:
                    remaining.append((shop_short, pid, spec_id, spec_name, '无匹配',
                        f"{dim_fmt(max(x,y,h))}*{dim_fmt(sorted([x,y,h])[1])}*{dim_fmt(min(x,y,h))}-{dk}-{mat}"))
                    stats['当下家-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f"\n统计:")
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}")
print(f"\n匹配成功: {len(matched_new)}")
print(f"定制链接: {len(custom_new)}")
print(f"仍留平卡: {len(remaining)}")

# 保存
if matched_new:
    f = os.path.join(out, '换绑文件_第9批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f)
    print(f"✅ 换绑_第9批: {f} ({os.path.getsize(f)/1024:.1f}KB, {len(matched_new)}条)")
    wb.close()

wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: 剩余{len(remaining)}条")
wb.close()

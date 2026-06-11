# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

other_file = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(other_file, read_only=True)
ws = wb.active

samples = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 2: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if not s: continue
    # 找含小数点的
    if '.' in s:
        samples.append(s)
        if len(samples) >= 20:
            break
wb.close()

print(f'其余商品中带小数点的规格({len(samples)}条):')
for s in samples:
    print(f'  {s[:80]}')

# 测试 extract_lwh
print('\n测试提取:')
def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''; val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

for s in samples[:5]:
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m: print(f'  ✅ 外尺寸: {m.groups()}')
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m: print(f'  ✅ 长x宽x高: {m.groups()}')
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m: print(f'  ✅ 纯LxWxH: {m.groups()}')
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m: print(f'  ✅ 无括号LxWxH: {m.groups()}')
    print()

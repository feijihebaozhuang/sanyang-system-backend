# -*- coding: utf-8 -*-
"""品牌店最后632条：模式B 2条+模式C 630条，直接用关键词匹配"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ========== 快麦索引 ==========
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ========== 读平卡 ==========
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# ========== 处理 ==========
matched_new = []
nomatch_details = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    if '品牌店' not in shop:
        remaining.append(r)
        continue
    
    # ---- 模式B: 进口优质特硬E瓦 --- 直接写死 ----
    if '进口优质' in spec_name:
        dk = '内径' if '内径' in spec_name else '外径'
        # 100x100mm, 100mm高 -> 10*10*10
        l, w, h = 10.0, 10.0, 10.0
        mat = '特硬'
        code = match_3d(l, w, h, dk, mat)
        if code:
            matched_new.append((fn(shop), pid, spec_id, code))
            stats['模式B-匹配'] += 1
        else:
            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                f"10*10*10-{dk}-{mat}"))
            stats['模式B-无匹配'] += 1
        continue
    
    # ---- 模式C: 【材料】内径/外径;【XxY】;Zcm高度 ----
    # 关键词: 【双面白色/台湾纸/特硬原色】 + 内径/外径 + 【XxY】+ cm高度
    if '【' in spec_name and ('内径' in spec_name or '外径' in spec_name) and 'cm' in spec_name:
        # 提取材料
        mat = '特硬'
        if '双面白色' in spec_name:
            mat = '白色'
        elif '台湾纸' in spec_name or '台湾' in spec_name:
            mat = '超硬'
        elif '特硬原色' in spec_name:
            mat = '特硬'
        
        # 提取内外径
        if '内径' in spec_name:
            dk = '内径'
        else:
            dk = '外径'
        
        # 提取尺寸: 【XxY】格式
        import re
        m = re.search(r'【(\d+)\s*x\s*(\d+)\s*】', spec_name)
        m_h = re.search(r'(\d+)\s*cm\s*高度', spec_name)
        
        if m and m_h:
            x = float(m.group(1))
            y = float(m.group(2))
            h = float(m_h.group(1))
            l, w = max(x, y), min(x, y)
            dims = sorted([l, w, h], reverse=True)
            
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['模式C-匹配'] += 1
            else:
                # 外径转内径
                if dk == '外径':
                    il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                    if il > 0 and iw > 0 and ih > 0:
                        code_i = match_3d(il, iw, ih, '内径', mat)
                        if code_i:
                            matched_new.append((fn(shop), pid, spec_id, code_i))
                            stats['模式C-外转内匹配'] += 1
                        else:
                            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                                f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                            stats['模式C-外转内无匹配'] += 1
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                        stats['模式C-无匹配'] += 1
                else:
                    # 内径试外径
                    ol, ow, oh = dims[0]+1.5, dims[1]+0.5, dims[2]+0.5
                    code_o = match_3d(ol, ow, oh, '外径', mat)
                    if code_o:
                        matched_new.append((fn(shop), pid, spec_id, code_o))
                        stats['模式C-内转外匹配'] += 1
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                        stats['模式C-无匹配'] += 1
        else:
            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配', '尺寸提取失败'))
            stats['模式C-尺寸提取失败'] += 1
    else:
        remaining.append((shop, pid, spec_id, spec_name, '平卡/解析失败', ''))

print(f'\n====== 品牌店处理结果 ======')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配(待确认): {len(nomatch_details)}')

# 输出无匹配详情
if nomatch_details:
    print(f'\n--- 无匹配详情 ---')
    for d in nomatch_details[:20]:
        print(f'  {d[3][:80]} -> {d[5]}')
    if len(nomatch_details) > 20:
        print(f'  ... 还有{len(nomatch_details)-20}条')

# ========== 追加到第11批 ==========
if matched_new:
    f11 = os.path.join(out, '换绑文件_第11批.xlsx')
    if os.path.exists(f11):
        wb = oxl.load_workbook(f11)
        ws = wb['Sheet1']
        for m in matched_new:
            ws.append(list(m))
        wb.save(f11)
    else:
        wb = oxl.Workbook()
        ws = wb.active; ws.title = 'Sheet1'
        ws.append([None, '商品对应表', None, None])
        ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
        for m in matched_new:
            ws.append(list(m))
        wb.save(f11)
    wb.close()
    print(f'\n✅ 换绑文件_第11批: 新增{len(matched_new)}条')

# ========== 保存待确认 ==========
if nomatch_details:
    wb = oxl.Workbook()
    ws = wb.active; ws.title = '待确认'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for d in nomatch_details:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '品牌店_待确认.xlsx'))
    print(f'✅ 品牌店_待确认: {len(nomatch_details)}条（你告诉我对应的编码）')
    wb.close()

# ========== 更新平卡 ==========
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f'✅ 平卡_待处理已更新: {len(remaining)}条')
wb.close()

print('\n完成！')

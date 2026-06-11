# -*- coding: utf-8 -*-
"""检查定制链接的规格样本"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f)
ws = wb.active

count = 0
samples = []
for row in ws.iter_rows(values_only=True):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    count += 1
    if len(samples) < 30:
        samples.append(s)
    if '扣底盒' in s:
        print(f'  含扣底盒: {s[:120]}')
    if '纸箱' in s:
        print(f'  含纸箱: {s[:120]}')

print(f'\n定制链接总数: {count}')
print('\n=== 前30条规格样本 ===')
for i, s in enumerate(samples):
    print(f'  {i+1}. [{s[:120]}]')

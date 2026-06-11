# -*- coding: utf-8 -*-
"""直接调用v7的extract_all_dims和build_lwh检查问题"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# 直接复制v7的extract_all_dims
def extract_all_dims(s):
    results = []
    # 1. 宽【Wcm】高【Hcm】... 长【Lcm】
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(3)), float(m.group(1)), float(m.group(2)), 'A1'))
    # 2. 【长度 L 厘米】【宽度 W 厘米】
    m = re.search(r'【长度\s*([\d.]+)\s*厘米\s*】.*?【宽度\s*([\d.]+)\s*厘米\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), 0, 'A2'))
    # 3. 长【L】宽【W】高【H】
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A3'))
    # 4. 飞机盒
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A4'))
    # 5. 外径
    m = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)(?:\s*mm)?', s)
    if m:
        v = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        if any(x > 50 for x in v):
            v = [x/10 for x in v]
        results.append((v[0], v[1], v[2], 'A5'))
    # 6. 【LxWxH】
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B1'))
    # 7. 长前缀
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B2'))
    # 8. 末尾
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B3'))
    # 9. 裸
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B4'))
    return results

# 测试对前10条外径规格
from openpyxl import load_workbook
source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    count += 1
    dims = extract_all_dims(s)
    print(f'[{count}] {len(dims)}个结果: {[(d[3],d[0],d[1],d[2]) for d in dims]}')
    print(f'  规格: {s[:80]}')
    if count >= 5:
        break

wb.close()

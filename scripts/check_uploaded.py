# -*- coding: utf-8 -*-
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

ok_dir = r'd:\Desktop\换绑输出\OK文件'

# 检查所有已上传文件
files_to_check = [
    ('换绑文件_第1批.xlsx', '换绑1'),
    ('换绑文件_第2批.xlsx', '换绑2'),
    ('换绑文件_第3-1批.xlsx', '换绑3-1'),
    ('换绑文件_第3-2批.xlsx', '换绑3-2'),
    ('换绑文件_第4批.xlsx', '换绑4'),
    ('换绑文件_第5批.xlsx', '换绑5'),
    ('换绑文件_第6批.xlsx', '换绑6'),
    ('换绑文件_第7批.xlsx', '换绑7'),
    ('换绑文件_第8批.xlsx', '换绑8'),
    ('换绑文件_第9批.xlsx', '换绑9'),
    ('换绑文件_第10批.xlsx', '换绑10'),
    ('换绑文件_第11批.xlsx', '换绑11'),
    ('换绑文件_第12批.xlsx', '换绑12'),
    ('换绑文件_第13批.xlsx', '换绑13'),
    ('换绑文件_第14批.xlsx', '换绑14'),
    ('换绑文件_第15批.xlsx', '换绑15'),
    ('换绑文件_第16批.xlsx', '换绑16'),
    ('换绑文件_第17批.xlsx', '换绑17'),
    ('定制类_换绑文件.xlsx', '定制1'),
    ('定制类_换绑文件_第2批.xlsx', '定制2'),
    ('定制类_换绑文件_第3批.xlsx', '定制3'),
    ('定制类_换绑文件_第4批.xlsx', '定制4'),
    ('定制类_换绑文件_第5批.xlsx', '定制5'),
]

for fn, label in files_to_check:
    fp = os.path.join(ok_dir, fn)
    if not os.path.exists(fp):
        print('❌ %s (%s): 文件不存在' % (label, fn))
        continue
    
    wb = oxl.load_workbook(fp)
    ws = wb[wb.sheetnames[0]]
    
    # 看表头
    header2 = [str(c.value or '') for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    
    shops = set()
    cnt = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            shops.add(str(r[0]).strip())
            cnt += 1
    
    print('\n=== %s (%s) === %d条' % (label, fn, cnt))
    print('  表头: %s' % header2)
    for s in sorted(shops):
        print('  📛 %s' % s)
    
    wb.close()

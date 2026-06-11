# -*- coding: utf-8 -*-
"""导出刀模对照表中，四批商品里不存在该尺寸的剩余条目（与 transform 规则一致）。"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

# 与 transform_kuaimai_excel 保持一致
from transform_kuaimai_excel import (
    BATCH_FILES,
    EXTRA_SQ_ENTRIES,
    RC_DIMOLD_CANDIDATES,
    SQ_DIMOLD_CANDIDATES,
    _first_existing,
    _load_rect_entries,
    _load_square_entries,
    _parse_main_dimold,
)

OUT_SQ = Path(r"d:\Desktop\刀模_正方形_未覆盖剩余.xlsx")
OUT_RC = Path(r"d:\Desktop\刀模_长方形_未覆盖剩余.xlsx")


def _load_all_square_rows() -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    sq_path = _first_existing(SQ_DIMOLD_CANDIDATES)
    if sq_path:
        rows.extend(_load_square_entries(sq_path))
    rows.extend(EXTRA_SQ_ENTRIES)
    return rows


def _collect_batch_size_set() -> set[str]:
    """四批里出现过的尺寸（任意内/外/材料），该尺寸整组从剩余表排除。"""
    sizes: set[str] = set()
    for path in BATCH_FILES:
        if not path.is_file():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=4, values_only=True):
            size, _ = _parse_main_dimold(row[0])
            if size:
                sizes.add(size)
        wb.close()
    return sizes


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    sq_rows = _load_all_square_rows()
    rc_path = _first_existing(RC_DIMOLD_CANDIDATES)
    rc_rows: list[tuple[str, str, str | None]] = []
    if rc_path:
        for size, code in _load_rect_entries(rc_path):
            rc_rows.append((size, code, None))

    batch_sizes = _collect_batch_size_set()
    sq_left = [r for r in sq_rows if r[0] not in batch_sizes]
    rc_left = [r for r in rc_rows if r[0] not in batch_sizes]

    wb = Workbook()
    ws = wb.active
    ws.title = "正方形未覆盖"
    ws.append(["尺寸", "类型", "刀模编号"])
    for row in sq_left:
        ws.append(list(row))
    wb.save(OUT_SQ)
    wb.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "长方形未覆盖"
    ws.append(["尺寸", "刀模编号", "备注"])
    for row in rc_left:
        ws.append([row[0], row[1], row[2]])
    wb.save(OUT_RC)
    wb.close()

    has_182 = any(r[0] == "18*18*2" for r in sq_left)
    print(f"正方形: 对照 {len(sq_rows)}，四批已有尺寸 {len(batch_sizes)}，剩余 {len(sq_left)} → {OUT_SQ.name}")
    print(f"长方形: 对照 {len(rc_rows)}，四批已有尺寸，剩余 {len(rc_left)} → {OUT_RC.name}")
    print(f"18*18*2 仍在剩余? {has_182}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

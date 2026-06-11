# -*- coding: utf-8 -*-
"""定制链接：按店铺+规格骨架分类，每种骨架举一个样例"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

def skeleton(s):
    """提取规格骨架：保留括号结构、关键字结构，去掉具体数值和材料描述"""
    # 去掉具体数字
    s = re.sub(r'\d+\.?\d*', 'N', s)
    # 去掉空格
    s = re.sub(r'\s+', '', s)
    # 保留中英文括号、分隔符、关键词
    # 去掉具体材料名称（中文字符串中夹在【】外的）
    # 简化：只保留括号结构、xX*、分隔符
    s = re.sub(r'[^【】\[\]；;，,\s\w\u4e00-\u9fff*XxN]+', '', s)
    return s[:300]

by_shop_skeleton = {}
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total += 1
    spec = str(row[2] or '').strip() if len(row) > 2 else ''
    shop = str(row[4] or '').strip() if len(row) > 4 else '(无名)'
    pid = str(row[3] or '') if len(row) > 3 else ''
    if not spec: continue
    
    sk = skeleton(spec)
    key = (shop, sk)
    if key not in by_shop_skeleton:
        by_shop_skeleton[key] = (spec, pid)
    
    if total >= 500000: break

wb.close()

print(f'定制链接总数: {total}')
print(f'不同骨架（店铺+结构）: {len(by_shop_skeleton)}')
print()

# 按店铺分组
by_shop = {}
for (shop, sk), (spec, pid) in by_shop_skeleton.items():
    by_shop.setdefault(shop, []).append((sk, spec, pid))

for shop in sorted(by_shop):
    items = by_shop[shop]
    print(f'═══ {shop}（{len(items)} 种结构）═══')
    for sk, spec, pid in items:
        print(f'  结构: {sk[:120]}')
        print(f'  样例: {spec[:120]}')
        print()

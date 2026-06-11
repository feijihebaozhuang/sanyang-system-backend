# -*- coding: utf-8 -*-
"""处理天猫小批量3条+品牌店243条+天猫彩色436条 → 第12批换绑"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '淘宝品牌店': '飞机盒品牌店',
    '天猫彩色': '飞机盒彩色专卖店',
}

def fn(shop):
    return SHOP_MAP.get(shop.strip(), shop.strip())

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    handled = False
    
    # ===== 天猫小批量3条 =====
    if shop == '天猫小批量':
        # 1. 宽【10cm】高【8cm】内径;【100个】长【38cm】 → 38*10*8-内径-特硬
        if '内径' in spec_name and '长' in spec_name:
            m = re.search(r'宽[【】\s]*(\d+)\s*cm.*?高[【】\s]*(\d+)\s*cm.*?长[【】\s]*(\d+)\s*cm', spec_name)
            if m:
                w_cm = float(m.group(1)); h_cm = float(m.group(2)); l_cm = float(m.group(3))
                code = match_3d(l_cm, w_cm, h_cm, '内径', '特硬')
                if code:
                    matched_new.append((fn(shop), pid, spec_id, code))
                    stats['小批量-内径匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l_cm)}*{int(w_cm)}*{int(h_cm)}-内径-特硬'))
                    stats['小批量-内径无匹配'] += 1
                handled = True
        # 2. 优质进口纸-黄色【100个】;37*14**3.8 → 37*14*3.8-外径-特硬
        elif '优质进口纸' in spec_name and '37*14' in spec_name:
            code = match_3d(37, 14, 3.8, '外径', '特硬')
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['小批量-优质进口匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', '37*14*3.8-外径-特硬'))
                stats['小批量-优质进口无匹配'] += 1
            handled = True
        # 3. 双白色【100个】;37*14**3.8 → 37*14*3.8-外径-白色
        elif '双白色' in spec_name and '37*14' in spec_name:
            code = match_3d(37, 14, 3.8, '外径', '白色')
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['小批量-双白匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', '37*14*3.8-外径-白色'))
                stats['小批量-双白无匹配'] += 1
            handled = True
    
    # ===== 品牌店243条 =====
    if not handled and shop == '淘宝品牌店':
        m = re.search(r'【([^】]*)】\s*(内径|外径)\s*;\s*[【】\s]*(\d+)\s*x\s*(\d+)[】]*\s*;\s*(\d+)\s*cm', spec_name)
        if m:
            dk = m.group(2)
            x = float(m.group(3)); y = float(m.group(4)); h = float(m.group(5))
            l, w = max(x, y), min(x, y)
            code = match_3d(l, w, h, dk, '白色')
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['品牌店-匹配'] += 1
            else:
                # 外转内
                if dk == '外径':
                    il, iw, ih = l-1.5, w-0.5, h-0.5
                    if il > 0:
                        code_i = match_3d(il, iw, ih, '内径', '白色')
                        if code_i:
                            matched_new.append((fn(shop), pid, spec_id, code_i))
                            stats['品牌店-外转内匹配'] += 1
                        else:
                            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-白色'))
                            stats['品牌店-无匹配'] += 1
                    else:
                        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(h)}-{dk}-白色'))
                        stats['品牌店-无匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(h)}-{dk}-白色'))
                    stats['品牌店-无匹配'] += 1
            handled = True
    
    # ===== 天猫彩色436条 =====
    if not handled and shop == '天猫彩色':
        m = re.search(r'宽度[：:]\s*(\d+\.?\d*)\s*cm---高度[：:]\s*(\d+\.?\d*)\s*cm;.*?长度[：:]\s*(\d+\.?\d*)\s*cm', spec_name)
        if m:
            w_cm = float(m.group(1)); h_cm = float(m.group(2)); l_cm = float(m.group(3))
            code = match_3d(l_cm, w_cm, h_cm, '外径', '白色')
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['彩色-匹配'] += 1
            else:
                # 外转内
                il, iw, ih = l_cm-1.5, w_cm-0.5, h_cm-0.5
                if il > 0 and iw > 0:
                    code_i = match_3d(il, iw, ih, '内径', '白色')
                    if code_i:
                        matched_new.append((fn(shop), pid, spec_id, code_i))
                        stats['彩色-外转内匹配'] += 1
                    else:
                        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-白色'))
                        stats['彩色-无匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l_cm)}*{int(w_cm)}*{int(h_cm)}-外径-白色'))
                    stats['彩色-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'平卡剩余(6个阿里+其他): {len(remaining)}')

# 写入第12批到OK文件夹
if matched_new:
    f12 = os.path.join(ok_dir, '换绑文件_第12批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f12)
    wb.close()
    sz = os.path.getsize(f12)
    print(f'\n✅ 换绑文件_第12批: {f12} ({sz/1024:.1f}KB, {len(matched_new)}条)')

# 无匹配追加到无匹配
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for shop, pid, spec_name, spec_id, reason, expect in nomatch:
        ws.append([shop, pid, spec_name, spec_id, reason, expect])
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 已追加{len(nomatch)}条到无匹配_待处理')

# 更新平卡（只保留未处理的）
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡_待处理已更新: {len(remaining)}条')

print('\n完成！')

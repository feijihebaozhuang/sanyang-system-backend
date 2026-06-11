# -*- coding: utf-8 -*-
"""检查extract_lwh匹配不到规格的原因"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

total = 0
sampled = 0
no_match = 0
sample_no = []

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 当前正则
    m1 = re.search(r'(?:【|长\s*[：:]?\s*)(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    m2 = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    
    if not m1 and not m2:
        no_match += 1
        if len(sample_no) < 50 and no_match <= 50:
            sample_no.append(s)
    
    if total % 50000 == 0:
        print(f'扫描 {total} 条... 未匹配 {no_match}')

wb.close()
print(f'\n总数: {total}, 未匹配: {no_match}')

print('\n=== 部分未匹配规格样本 ===')
for i, s in enumerate(sample_no):
    print(f'  {i+1}. [{s[:120]}]')

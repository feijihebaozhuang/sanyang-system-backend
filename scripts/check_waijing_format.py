# -*- coding: utf-8 -*-
"""检查定制链接中91,292条外径规格的具体格式"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

samples = []
total_waijing = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' in s:
        total_waijing += 1
        if len(samples) < 50:
            samples.append(s)
wb.close()

print(f'总外径规格: {total_waijing}')
print()
print('=== 前50条外径规格 ===')
for i, s in enumerate(samples):
    # 标注是否被我当前的extract_all_dims提取到
    from split_neijing_v6 import extract_all_dims
    dims = extract_all_dims(s)
    has = '✅有LWH' if dims else '❌无LWH'
    print(f'  {i+1}. [{has}] [{s[:120]}]')

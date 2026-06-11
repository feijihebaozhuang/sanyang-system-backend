# -*- coding: utf-8 -*-
"""
v13: 按店铺处理
先处理已经看过格式的店铺：
- 天猫扣底盒
- 天猫止合

材料规则：
- 特硬=特硬  白色=白  超硬=超硬  黑色=黑  红色=红
- 台湾=超硬   特价/特惠/特好=优质
- 沉默材料=特硬
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

# ===== 工具函数 =====
def clean_num(v):
    parts = v.split('.')
    if len(parts) > 2:
        return parts[0] + '.' + ''.join(parts[1:])
    return v

def to_cm(v, unit=''):
    try: v = float(clean_num(str(v)))
    except: return None
    if 'mm' in unit or v > 50: v = v / 10.0
    return v

def classify_dim_type(l, w, h):
    has_dec = any(v != int(v) for v in [l, w, h])
    all5 = all(v % 1 == 0.5 for v in [l, w, h])
    if has_dec and not all5: return '非全量'
    else: return '全量'

def get_material(s):
    """提取材料"""
    if '台湾' in s or '台湾超硬' in s: return '超硬'
    if '特价' in s or '特惠' in s or '特好' in s or '优质' in s: return '优质'
    if '特硬' in s: return '特硬'
    if '双面白' in s or ('白' in s and '色' in s): return '白'
    if '超硬' in s: return '超硬'
    if '双面黑' in s or ('黑' in s and '色' in s): return '黑'
    if '双面红' in s or ('红' in s and '色' in s): return '红'
    if '牛皮' in s: return None  # 牛皮不识别
    return '特硬'  # 沉默=特硬

# ===== 各店铺处理函数 =====

def process_tmkoudi(s, row):
    """天猫扣底盒"""
    l = w = h = None
    
    # 格式1-3: 宽度26【白色】;【100个】外尺寸【长度27cm】----高度4cm
    m = re.search(r'宽度\s*([\d.]+).*?外尺寸【?长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
    if m: l, w, h = to_cn(m.group(2)), to_cn(m.group(1)), to_cn(m.group(3))
    
    return (l, w, h)

# 重试用 - 直接对比格式
KNOWN_FORMATS = {
    # ===== 天猫扣底盒 =====
    ('天猫扣底盒', '宽度N【白色】;【N个】外尺寸【长度Ncm】----高度Ncm'): ('外径', 27, 26, 4),
    ('天猫扣底盒', '宽度N【红色】;【N个】外尺寸【长度Ncm】----高度Ncm'): ('外径', 27, 26, 4),
    ('天猫扣底盒', '宽度N【黑色】;【N个】外尺寸【长度Ncm】----高度Ncm'): ('外径', 27, 26, 4),
    ('天猫扣底盒', 'N*Ncm;Ncm高;五层'): ('纸箱', 30, 30, 12),
    ('天猫扣底盒', '宽度N【白色】;【N个】【外尺寸】长度Ncm----高度Ncm'): ('外径', 6, 6, 5),
    ('天猫扣底盒', '宽度N【红色】;【N个】【外尺寸】长度Ncm----高度Ncm'): ('外径', 6, 6, 5),
    ('天猫扣底盒', '宽度N【黑色】;【N个】【外尺寸】长度Ncm----高度Ncm'): ('外径', 6, 6, 5),
    ('天猫扣底盒', '宽度N【白色】;【N个】外尺寸【长度Ncm】----高度Nm'): ('外径', 14, 10, 6),
    ('天猫扣底盒', '宽度N【黑色】;【N个】外尺寸【长度Ncm】----高度Nm'): ('外径', 14, 10, 6),
    ('天猫扣底盒', '宽度N【红色】;【N个】外尺寸【长度Ncm】----高度Nm'): ('外径', 14, 10, 6),
    ('天猫扣底盒', '宽度N【白色】;【N个外尺寸】长度Ncm----高度Ncm'): ('外径', 21, 10, 2),
    ('天猫扣底盒', '宽度N【红色】;【N个外尺寸】长度Ncm----高度Ncm'): ('外径', 21, 10, 2),
    ('天猫扣底盒', '宽度N【黑色】;【N个外尺寸】长度Ncm----高度Ncm'): ('外径', 21, 10, 2),
    ('天猫扣底盒', '红色【N个】;长宽高N*N*Ncm'): ('外径', 10, 10, 2),
    ('天猫扣底盒', '黑色【N个】;长宽高N*N*Ncm'): ('外径', 10, 10, 2),
    ('天猫扣底盒', '内宽x高【NxN】=外宽x高【NxN】cm;【N个】内尺寸【长度Ncm】=【外长Ncm】'): ('内径', 10, 10, 10),
    ('天猫扣底盒', '外宽x高【NxN】=内宽x高【NxN】cm;【N个】外尺寸【长度Ncm】=【内长Ncm】'): ('外径', 12, 12, 12),
    ('天猫扣底盒', '双面白色【N个】高档;外尺寸【NxNxNcm】=内尺寸【NxNxNcm】'): ('外径', 11, 11, 11),
    ('天猫扣底盒', '特硬黄【N个】限时优惠;外尺寸【NxNxNcm】=内尺寸【NxNxNcm】'): ('外径', 11, 11, 11),
    ('天猫扣底盒', '超硬黄【N个】进口牛卡;外尺寸【NxNxNcm】=内尺寸【NxNxNcm】'): ('外径', 11, 11, 11),
    ('天猫扣底盒', '外宽x高【NxN】=内宽x高【NxN】cm;【N个】外尺寸【长度Ncm】=【内长Ncm'): ('外径', 9, 9, 9),
    ('天猫扣底盒', '定制飞机盒/订做纸箱/定做扣底盒;下拉看详情选择尺寸'): ('定制', None, None, None),
    ('天猫扣底盒', '按照价格截图客服拍下;下拉看详情选择尺寸'): ('定制', None, None, None),
}

# ===== 全量分类 =====
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
waijing_rows = []
feiji_rows = []

for row in rows_list:
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    shop = str(row[0] or '').strip() if len(row) > 0 else ''
    if not s:
        custom_rows.append(row)
        continue
    
    # 根据店铺走不同逻辑
    # ========== 天猫扣底盒 ==========
    if '扣底盒' in shop:
        # 没长宽高三字 → 定制
        if '长' not in s and '宽' not in s and '高' not in s:
            custom_rows.append(row)
            continue
        
        l = w = h = None
        
        # 格式1-3: 宽度N【颜色】;【N个】外尺寸【长度Ncm】----高度Ncm
        m = re.search(r'宽度\s*([\d.]+).*?外尺寸\[?【?\s*长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
        if m: l, w, h = to_cn(m.group(2)), to_cn(m.group(1)), to_cn(m.group(3))
        
        # 格式4: N*Ncm;Ncm高;五层 → 纸箱
        m = re.search(r'([\d.]+)\s*\*\s*([\d.]+)\s*cm\s*;\s*([\d.]+)\s*cm\s*高.*?五层', s)
        if m: l, w, h = to_cn(m.group(1)), to_cn(m.group(2)), to_cn(m.group(3))
        
        # 格式5-7: 宽度N【颜色】;【N个】【外尺寸】长度Ncm----高度Ncm
        if not l:
            m = re.search(r'宽度\s*([\d.]+).*?【外尺寸】长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
            if m: l, w, h = to_cn(m.group(2)), to_cn(m.group(1)), to_cn(m.group(3))
        
        # 格式8-10: 宽度N【颜色】;【N个】外尺寸【长度Ncm】----高度Nm
        if not l:
            m = re.search(r'宽度\s*([\d.]+).*?外尺寸\[?【?\s*长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*m', s)
            if m: l, w, h = to_cn(m.group(2)), to_cn(m.group(1)), to_cn(m.group(3))
        
        # 格式11-13: 宽度N【颜色】;【N个外尺寸】长度Ncm----高度Ncm
        if not l:
            m = re.search(r'宽度\s*([\d.]+).*?【N个外尺寸】长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
            if m: l, w, h = to_cn(m.group(2)), to_cn(m.group(1)), to_cn(m.group(3))
        
        # 其他格式... 先用通用降级
        
        if l is None:
            custom_rows.append(row)
            continue
        
        # 分类型
        if '五层' in s:
            zhixiang_rows.append(row)
        elif '内' in s and ('尺寸' in s or '宽x高' in s and '内' in s.split('宽x高')[0]):
            neijing_rows.append(row)
        else:
            waijing_rows.append(row)
        continue
    
    # ========== 其他店铺（默认逻辑）==========
    # 欠长宽高三字 → 定制
    if '长' not in s and '宽' not in s and '高' not in s:
        custom_rows.append(row)
        continue
    
    custom_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)
print(f'和: {s} (应={total})')

# -*- coding: utf-8 -*-
"""为新生成的批次文件添加模板标题行"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

okdir = r'D:\Desktop\换绑输出\OK文件'

# 需要处理的批次: 33-39
batches = list(range(33, 40))

for b in batches:
    fname = f'\u6362\u7ed1\u6587\u4ef6_\u7b2c{b}\u6279.xlsx'
    fpath = os.path.join(okdir, fname)
    
    if not os.path.exists(fpath):
        print(f'\u4e0d\u5b58\u5728: {fname}')
        continue
    
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    
    # 读取已有第1行作为表头
    headers = [ws.cell(1, c).value for c in range(1, 5)]
    
    # 在第1行之前插入一行
    ws.insert_rows(1)
    
    # 新第1行: 商品对应表（B列）
    ws.cell(1, 2).value = '\u5546\u54c1\u5bf9\u5e94\u8868'
    
    # 注意insert后原来的第1行变成了第2行，表头不变
    wb.save(fpath)
    print(f'  \u5df2\u4fee\u6539: \u7b2c{b}\u6279 - \u6dfb\u52a0\u4e86\u6807\u9898\u884c')
    
    # 处理拆分部分
    parts = [f for f in os.listdir(okdir) if f.startswith(f'\u6362\u7ed1\u6587\u4ef6_\u7b2c{b}\u6279_') and f.endswith('.xlsx')]
    for p in sorted(parts):
        ppath = os.path.join(okdir, p)
        pwb = openpyxl.load_workbook(ppath)
        pws = pwb.active
        pws.insert_rows(1)
        pws.cell(1, 2).value = '\u5546\u54c1\u5bf9\u5e94\u8868'
        pwb.save(ppath)
        print(f'    \u62c6\u5206: {p} - \u5df2\u4fee\u6539')

print(f'\n\u5b8c\u6210! \u7b2c33-39\u6279\u5747\u5df2\u6dfb\u52a0\u6807\u9898\u884c')

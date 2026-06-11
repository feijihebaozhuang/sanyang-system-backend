# -*- coding: utf-8 -*-
"""把当下家336条和俊鑫272条从平卡移除，当下家到无匹配"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

RE_DXJ = re.compile(
    r'(\d+\.?\d*)\s*mm?\s*高\s*[（(]扣底盒[）)].*?(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*mm?\s*长宽'
)

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# 读第9批匹配ID
f9 = os.path.join(out, '换绑文件_第9批.xlsx')
wb9 = oxl.load_workbook(f9, data_only=True)
f9_keys = set()
for rr in wb9['Sheet1'].iter_rows(min_row=3, values_only=True):
    if rr: f9_keys.add((str(rr[1] or '').strip(), str(rr[2] or '').strip()))
wb9.close()

remaining = []
dxj_miss = []

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    # 俊鑫已匹配→跳过
    if '俊鑫' in shop and (pid, spec_id) in f9_keys:
        continue
    
    # 当下家→无匹配
    if '当下家' in shop:
        m = RE_DXJ.search(spec_name)
        if m:
            h_mm = float(m.group(1))
            x_mm = float(m.group(2))
            y_mm = float(m.group(3))
            x, y, h = x_mm/10, y_mm/10, h_mm/10
            dims = sorted([x, y, h], reverse=True)
            expected = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-白色-扣底盒"
            dxj_miss.append((shop, pid, spec_id, spec_name, '无匹配', expected))
            continue
    
    remaining.append(r)

print(f"俊鑫已移除: 272条", flush=True)
print(f"当下家→无匹配: {len(dxj_miss)}条", flush=True)
print(f"平卡剩余: {len(remaining)}条", flush=True)

# 保存平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()

# 追加到无匹配
if dxj_miss:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'), data_only=True)
    ws = wb['无匹配']
    for r in dxj_miss:
        ws.append(list(r[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f"✅ 无匹配已追加{len(dxj_miss)}条")

print("\n完成！可以继续看下一个店铺了")

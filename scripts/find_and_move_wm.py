# -*- coding: utf-8 -*-
"""找到品牌店_待确认文件（中文乱码版）并移到无匹配"""
import os, sys, glob
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 找所有xlsx文件
all_xlsx = [f for f in os.listdir(out) if f.endswith('.xlsx')]
print(f'所有xlsx文件({len(all_xlsx)}个):')
for f in all_xlsx:
    sz = os.path.getsize(os.path.join(out, f))
    print(f'  {repr(f)} ({sz/1024:.1f}KB)')

# 试图通过openpyxl直接读取每个小文件找到"待确认"sheet
for f in all_xlsx:
    fp = os.path.join(out, f)
    sz = os.path.getsize(fp)
    if sz > 100000:  # 跳过大于100KB的
        continue
    try:
        wb = oxl.load_workbook(fp, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        if '待确认' in sheets:
            print(f'\n找到待确认sheet: {repr(f)}')
            # 读取数据
            wb = oxl.load_workbook(fp, data_only=True)
            ws = wb['待确认']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            print(f'  条数: {len(rows)}')
            if rows:
                # 追加到无匹配
                wb2 = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'), data_only=True)
                ws2 = wb2['无匹配']
                for r in rows:
                    if len(r) >= 6:
                        ws2.append(list(r[:6]))
                    else:
                        ws2.append(list(r) + ['', ''])
                wb2.save(os.path.join(out, '无匹配_待处理.xlsx'))
                wb2.close()
                os.remove(fp)
                print(f'  ✅ 已追加{len(rows)}条到无匹配，已删除原文件')
    except:
        pass

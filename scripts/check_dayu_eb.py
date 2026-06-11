# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

wb = oxl.load_workbook(r'D:\Desktop\换绑_深圳市大鱼包装材料有限公司.xlsx')
ws = wb.active
cnt = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    code = str(row[3] or '')
    if 'EB' in code:
        cnt += 1
        if cnt <= 10:
            print(f'{row[1]} | {row[2]} | {code}')
print(f'\n总EB数量: 208 (原纸箱数量)')

# 检查有没有"5B"残留
wb2 = oxl.load_workbook(r'D:\Desktop\换绑_深圳市大鱼包装材料有限公司.xlsx')
ws2 = wb2.active
bad = 0
for row in ws2.iter_rows(min_row=3, values_only=True):
    code = str(row[3] or '')
    if '-5B' in code or '-3B' in code or '-7B' in code:
        bad += 1
        print(f'❌ 残留层数B: {code}')
print(f'残留层数B: {bad}')

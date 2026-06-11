# -*- coding: utf-8 -*-
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 看平卡里天猫彩色和小批量的规格
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
samples = {'天猫彩色': [], '天猫小批量': []}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r: continue
    shop = str(r[0] or '').strip()
    if shop in samples:
        samples[shop].append(str(r[2] or '')[:80])

for shop, specs in samples.items():
    print('=== %s: %d条 ===' % (shop, len(specs)))
    seen = set()
    for s in specs[:10]:
        if s not in seen:
            seen.add(s)
            print('  %s' % s)
    if len(specs) > 10:
        print('  ...')

# 看无匹配里有没有天猫彩色
wb2 = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws2 = wb2['无匹配']
wm_shops = Counter()
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        wm_shops[str(r[0]).strip()] += 1
print('\n=== 无匹配_待处理 店铺分布 ===')
for s, c in wm_shops.most_common():
    print('  %s: %d条' % (s, c))

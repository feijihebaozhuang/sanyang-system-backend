# -*- coding: utf-8 -*-
"""检查剩余无匹配的外径规格格式"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

total_waijing = 0
no_lwh = 0
types = {}
examples = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    total_waijing += 1
    
    # 测试是否能被v8提取
    from split_neijing_v8 import extract_all_dims, build_lwh
    
    lwh = build_lwh(s)
    if lwh is None:
        no_lwh += 1
        # 分类格式
        has_mm = 'mm' in s
        has_cm = 'cm' in s or '厘米' in s
        has_x = bool(re.search(r'\d[\d.]*\s*[xX*]\s*\d[\d.]*', s))
        has_height = '高度' in s
        has_length = '长度' in s or '长' in s
        has_width = '宽度' in s or '宽' in s
        
        if has_x and has_length and has_width:
            t = '有长宽+乘号'
        elif has_mm and 'mm' in s:
            t = '有mm'
        elif has_cm:
            t = '有cm'
        elif has_x:
            t = '有乘号'
        else:
            t = '仅有外径关键词'
        types[t] = types.get(t, 0) + 1
        
        if len(examples) < 30:
            examples.append(s)

wb.close()
print(f'总外径: {total_waijing}')
print(f'无LWH: {no_lwh}')
print()
print('格式分布:')
for k, v in sorted(types.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')
print()
print('示例:')
for i, s in enumerate(examples[:20]):
    print(f'  [{i+1}] {s[:130]}')

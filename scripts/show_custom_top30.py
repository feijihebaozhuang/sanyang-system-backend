# -*- coding: utf-8 -*-
"""采样定制链接前20条实际的规格"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

print('前30条定制链接规格:\n')
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    count += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    print(f'  [{count}] {s[:130]}')
    if count >= 30:
        break
wb.close()

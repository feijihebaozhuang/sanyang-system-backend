# -*- coding: utf-8 -*-
"""分析全部50万条数据的规格模式，到上一个分类满为止"""
import openpyxl, sys, re

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

fp = r"d:\Desktop\平台商品.xlsx"
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
ws = wb['报表1']

patterns = {}
sentinel = object()

for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
    shop, pid, spec_name, spec_id = row
    if not spec_name:
        continue
    s = str(spec_name).strip()
    if i > 498090:
        break
    
    has_bracket = '【' in s
    has_lwh = bool(re.search(r'[长宽高厚度长度]', s))
    x3 = re.search(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)', s)
    x2 = re.search(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)', s)
    pure_num = bool(re.match(r'^[\d\.\-\sxX*×/]+$', s))
    
    if has_lwh:
        if has_bracket:
            if '宽【' in s and '高【' in s and '长【' in s:
                label = 'A1_宽【】高【】长【】'
            elif '长【' in s and '宽【' in s and '高【' in s:
                label = 'A2_长【】宽【】高【】'
            elif '宽【' in s and '高【' in s and '长度【' in s:
                label = 'A3_宽【】高【】长度【】'
            elif '宽【' in s and '长【' in s:
                label = 'A4_宽【】长【】(无高)'
            elif '宽【' in s and '高【' in s:
                label = 'A5_宽【】高【】(无长)'
            elif '长【' in s:
                label = 'A6_长【】(无宽高)'
            else:
                label = 'A7_其他【】+长宽高'
        else:
            if '---' in s:
                label = 'B1_长宽高+---分隔'
            elif '；' in s or ';' in s:
                label = 'B2_长宽高+;分隔'
            elif '长' in s and '宽' in s and '高' in s:
                label = 'B3_长宽高+其他分隔'
            elif '长' in s and '宽' in s and '厚' in s:
                label = 'B4_长宽厚'
            elif '长' in s and '宽' in s:
                label = 'B5_长宽(无高)'
            elif '宽' in s and '高' in s:
                label = 'B6_宽高(无长)'
            else:
                label = 'B7_其他有长宽高字眼'
    elif has_bracket:
        prefix = s.split('【')[0].strip()
        bracket_count = s.count('【')
        if bracket_count >= 2:
            label = 'C1_多组【】(无长宽高)'
        elif x3:
            label = 'C2_【数量】前缀+x3尺寸'
        elif x2:
            label = 'C3_【数量】前缀+x2尺寸'
        elif '颜色' in s or '色' in s:
            label = 'C4_【】+颜色信息'
        else:
            label = 'C5_【】+其他'
    elif x3:
        label = 'D1_纯x3尺寸'
    elif x2:
        label = 'D2_纯x2尺寸'
    elif pure_num:
        nums = re.findall(r'\d+\.?\d*', s)
        if len(nums) == 1:
            label = 'E1_单个数字'
        elif len(nums) == 2:
            label = 'E2_两个数字'
        else:
            label = 'E3_多个数字'
    elif '颜色' in s or '色' in s:
        label = 'F1_只有颜色'
    elif '定制' in s or '定做' in s or '定造' in s:
        label = 'G_定制定做'
    else:
        label = 'H_其他'
    
    if label not in patterns:
        patterns[label] = []
    if len(patterns[label]) < 15:  # 每类收集15个样本
        patterns[label].append((s, shop))

wb.close()

# 输出
import openpyxl as oxl
wb2 = oxl.Workbook()
ws2 = wb2.active
ws2.title = '规格全分类'
ws2.append(['分类代码', '大致数量占比', '样本个数', '样本(前15个)'])

print("="*60)
print("全部50万条规格模式分类")
print("="*60)

for label, samples in sorted(patterns.items(), key=lambda x: -len(x[1])):
    cnt_approx = f"{len(samples):3d}个样本"
    print(f"\n【{label}】{cnt_approx}")
    sample_str = ''
    for s, shop in samples:
        print(f"  {s[:100]}")
        sample_str += s[:80] + ' | '
    
    ws2.append([label, cnt_approx, len(samples), sample_str.strip(' | ')])

out = r"d:\Desktop\规格所有模式分类.xlsx"
wb2.save(out)
print(f"\n✅ 已保存: {out}")

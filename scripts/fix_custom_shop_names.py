# -*- coding: utf-8 -*-
"""修复定制类换绑文件中的店铺名称为全称"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '深圳市友尚包装有限公司',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

def fix_file(fname):
    fp = os.path.join(out, fname)
    if not os.path.exists(fp):
        print(f"{fname}: 不存在，跳过")
        return False
    
    wb = oxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # 跳过标题行
    
    wb_new = oxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = 'Sheet1'
    ws_new.append([None, '商品对应表', None, None])
    ws_new.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    
    fixed = 0
    for r in rows:
        if not r or not r[0]:
            continue
        short = str(r[0]).strip()
        full = SHOP_NAME_MAP.get(short, short)
        if full != short:
            fixed += 1
        ws_new.append([full, str(r[1] or '').strip(), str(r[2] or '').strip(), str(r[3] or '').strip()])
    
    wb_new.save(fp)
    wb_new.close()
    wb.close()
    sz = os.path.getsize(fp)
    print(f"✅ {fname}: 修复{fixed}条, {sz/1024:.1f}KB")
    return True

fix_file('定制类_换绑文件.xlsx')
fix_file('定制类_换绑文件_第2批.xlsx')
fix_file('定制类_换绑文件_第3批.xlsx')

print("\n全部修复完成，可以重新上传了！")

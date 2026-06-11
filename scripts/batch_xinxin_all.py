# -*- coding: utf-8 -*-
"""阿里新鑫星1365条全部处理，生成第13批换绑"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if shop != '阿里新鑫星':
        remaining.append(r)
        continue
    
    # ===== A类: XxY cm 【长x宽】;X层纸箱【高度Zcm】 =====
    m_a = re.search(r'(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*cm\s*【长x宽】.*?(\d+)层纸箱【高度(\d+)cm', spec_name)
    if m_a:
        x = float(m_a.group(1)); y = float(m_a.group(2)); z = float(m_a.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', '3B')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['A-纸箱3B匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(z)}-外径-3B'))
            stats['A-纸箱3B无匹配'] += 1
        continue
    
    # ===== H类: 长*宽【X*Y】cm;外尺寸白色【Zcm高】 =====
    m_h = re.search(r'长\*宽【(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)】\s*cm;外尺寸白色【(\d+\.?\d*)cm', spec_name)
    if m_h:
        x = float(m_h.group(1)); y = float(m_h.group(2)); z = float(m_h.group(3))
        # 你说的：外径22.5*21.5*2.5 → 内径21*21*2-内径-白色
        il, iw, ih = int(x)-1.5, int(y)-0.5, int(z)-0.5
        if il > 0 and iw > 0 and ih > 0:
            code = match_3d(il, iw, ih, '内径', '白色')
            if code:
                matched_new.append(('阿里新鑫星', pid, spec_id, code))
                stats['H-外转内白色匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(il)}*{int(iw)}*{int(ih)}-内径-白色'))
                stats['H-无匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(x)}*{int(y)}*{int(z)}-外径-白色'))
            stats['H-无匹配'] += 1
        continue
    
    # ===== B类: XxYxZ；外径扣抵盒；材料 =====
    m_b = re.search(r'([\d.]+)X([\d.]+)X([\d.]+)；外径扣抵盒', spec_name)
    if m_b:
        x = float(m_b.group(1)); y = float(m_b.group(2)); z = float(m_b.group(3))
        dims = sorted([x, y, z], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', 'P6D')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['B-扣底盒P6D匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-P6D'))
            stats['B-扣底盒无匹配'] += 1
        continue
    
    # ===== G类: XxYxZ；外径双插盒 =====
    m_g = re.search(r'([\d.]+)X([\d.]+)X([\d.]+)；外径双插盒', spec_name)
    if m_g:
        x = float(m_g.group(1)); y = float(m_g.group(2)); z = float(m_g.group(3))
        dims = sorted([x, y, z], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', 'P6D')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['G-双插盒P6D匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-P6D'))
            stats['G-双插盒无匹配'] += 1
        continue
    
    # ===== D类: 纸箱长X宽【XxY】；外径五层高度【ZCM】KK特硬 =====
    m_d = re.search(r'纸箱长X宽【(\d+)X(\d+)】.*?外径五层高度【(\d+)CM', spec_name)
    if m_d:
        x = float(m_d.group(1)); y = float(m_d.group(2)); z = float(m_d.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', 'EB')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['D-纸箱EB匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(z)}-外径-EB'))
            stats['D-纸箱无匹配'] += 1
        continue
    
    # ===== C类: 外径白色高【Z厘米】；长x宽【XxY】 =====
    m_c = re.search(r'外径白色高【(\d+)厘米】；长x宽【(\d+)x(\d+)】', spec_name)
    if m_c:
        z = float(m_c.group(1)); x = float(m_c.group(2)); y = float(m_c.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', '白色')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['C-外径白色匹配'] += 1
        else:
            # 外转内
            il, iw, ih = l-1.5, w-0.5, z-0.5
            if il > 0:
                code_i = match_3d(il, iw, ih, '内径', '白色')
                if code_i:
                    matched_new.append(('阿里新鑫星', pid, spec_id, code_i))
                    stats['C-外转内白色匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(il)}*{int(iw)}*{int(ih)}-内径-白色'))
                    stats['C-无匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(z)}-外径-白色'))
                stats['C-无匹配'] += 1
        continue
    
    # ===== E类: 长X CM；宽Y cm；高Z cm =====
    m_e = re.search(r'(\d+\.?\d*)\s*(CM|cm)[；;]\s*宽\s*(\d+\.?\d*)\s*cm[；;]?(?:高度|高)\s*(\d+\.?\d*)\s*cm', spec_name, re.I)
    if not m_e:
        m_e = re.search(r'长(\d+\.?\d*)\s*(CM|cm)[；;]\s*宽\s*(\d+\.?\d*)\s*cm[；;]?(?:高度|高)\s*(\d+\.?\d*)\s*cm', spec_name, re.I)
    if m_e:
        l = float(m_e.group(1))
        w = float(m_e.group(3))
        h = float(m_e.group(4))
        if h < w and h < l:
            dims = sorted([l, w, h], reverse=True)
        else:
            dims = [l, w, h]
        code = match_3d(dims[0], dims[1], dims[2], '外径', '特硬')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['E-外径特硬匹配'] += 1
        else:
            il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
            if il > 0:
                code_i = match_3d(il, iw, ih, '内径', '特硬')
                if code_i:
                    matched_new.append(('阿里新鑫星', pid, spec_id, code_i))
                    stats['E-外转内特硬匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-特硬'))
                    stats['E-无匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-特硬'))
                stats['E-无匹配'] += 1
        continue
    
    # ===== F类: 长宽【XxY】cm；高度【Zcm材料】 =====
    m_f = re.search(r'长宽【(\d+)x(\d+)】\s*cm[；;]\s*高度【(\d+)cm', spec_name)
    if m_f:
        x = float(m_f.group(1)); y = float(m_f.group(2)); z = float(m_f.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', '特硬')
        if code:
            matched_new.append(('阿里新鑫星', pid, spec_id, code))
            stats['F-外径特硬匹配'] += 1
        else:
            il, iw, ih = l-1.5, w-0.5, z-0.5
            if il > 0:
                code_i = match_3d(il, iw, ih, '内径', '特硬')
                if code_i:
                    matched_new.append(('阿里新鑫星', pid, spec_id, code_i))
                    stats['F-外转内匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(il)}*{int(iw)}*{int(ih)}-内径-特硬'))
                    stats['F-无匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{int(l)}*{int(w)}*{int(z)}-外径-特硬'))
                stats['F-无匹配'] += 1
        continue
    
    # ===== 型号类: A1：10*6.5*5.5cm；E瓦  特硬 =====
    m_model = re.search(r'[\w]+[:：]\s*([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)\s*cm', spec_name)
    if m_model:
        dims = sorted([float(m_model.group(1)), float(m_model.group(2)), float(m_model.group(3))], reverse=True)
        code = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-特硬"
        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', code))
        stats['型号类-无匹配'] += 1
        continue
    
    # ===== D6D类: 28*27*11；D6D【特硬】；牛皮色 三层E瓦 =====
    m_d6d = re.search(r'^([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)\s*[；;]', spec_name)
    if m_d6d:
        dims = sorted([float(m_d6d.group(1)), float(m_d6d.group(2)), float(m_d6d.group(3))], reverse=True)
        code = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-特硬"
        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', code))
        stats['D6D类-无匹配'] += 1
        continue
    
    # 其他
    remaining.append(r)

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'新鑫星剩余未处理: {len([r for r in remaining if r and str(r[0] or "").strip() == "阿里新鑫星"])}')

# 保存第13批
if matched_new:
    f13 = os.path.join(ok_dir, '换绑文件_第13批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f13)
    wb.close()
    print(f'\n✅ 换绑文件_第13批: {len(matched_new)}条')

# 无匹配追加
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 追加{len(nomatch)}条到无匹配')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡已更新: {len(remaining)}条')

print('\n完成！')

# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

other_file = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(other_file, read_only=True)
ws = wb.active

# 看真实行列对应
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 2: continue
    if i > 10: break
    print(f'行{i}: type={type(row)}, len={len(row)}')
    for j, v in enumerate(row):
        if v is not None:
            sv = str(v)
            if '.' in sv:
                print(f'  [{j}]: {sv[:60]}')
wb.close()

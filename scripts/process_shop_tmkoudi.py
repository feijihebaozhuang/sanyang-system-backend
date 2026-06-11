# -*- coding: utf-8 -*-
"""
按店铺逐一处理。

当前处理店铺：天猫扣底盒
完全按照用户已确认的规则写死每种格式。
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

# 先只测试天猫扣底盒的23种格式
# 直接从平台商品里提取该店铺的所有行

SOURCE = r'D:\Desktop\平台商品.xlsx'
SHOP_NAME = '天猫扣底盒'
OUTPUT_DIR = r'D:\Desktop'

# ===== 读取数据 =====
wb = load_workbook(SOURCE, read_only=True)
ws = wb.active

header = None
all_rows = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    shop = str(row[0] or '').strip() if len(row) > 0 else ''
    all_rows.append(row)
wb.close()

print(f'平台商品总数: {total}')

# 只取天猫扣底盒
shop_rows = [r for r in all_rows if str(r[0] or '').strip() == SHOP_NAME]
print(f'{SHOP_NAME}: {len(shop_rows)} 条')

# ===== 天猫扣底盒规则（用户已确认）=====
def classify_koudi(s):
    """
    返回 (类别, L, W, H)
    类别: '定制' / '纸箱' / '外径' / '内径'
    """
    # === 定制类 ===
    if '定制飞机盒' in s or '订做纸箱' in s or '定做扣底盒' in s:
        return ('定制', None, None, None)
    if '按照价格截图客服拍下' in s or '下拉看详情选择尺寸' in s:
        if '长' not in s and '宽' not in s and '高' not in s:
            return ('定制', None, None, None)
    
    # === 纸箱：五层 ===
    # 30*30 cm;12cm 高;五层
    m = re.search(r'([\d.]+)\s*\*\s*([\d.]+)\s*cm\s*;\s*([\d.]+)\s*cm\s*高.*?五层', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('纸箱', l, w, h)
    
    # === 外径格式 ===
    # 格式1-3: 宽度26【白色】;【100个】外尺寸【长度27cm】----高度4cm
    # → 27 26 4
    m = re.search(r'宽度\s*([\d.]+).*?\s*【.*?】.*?外尺寸\[?【?\s*长度\s*([\d.]+)\s*cm.*?----\s*高度\s*([\d.]+)', s)
    if m:
        w, l, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('外径', l, w, h)
    
    # 格式5-7: 宽度6【白色】;【100个】【外尺寸】长度6cm----高度5cm
    # → 6 6 5
    m = re.search(r'宽度\s*([\d.]+).*?【外尺寸】\s*长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
    if m and ('外尺寸' in s):
        w, l, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('外径', l, w, h)
    
    # 格式8-10: 宽度10【白色】;【100个】外尺寸【长度14cm】----高度6m
    # → 14 10 6 (m=米? 但用户说是cm)
    m = re.search(r'宽度\s*([\d.]+).*?\s*【.*?】.*?外尺寸.*?长度\s*([\d.]+)\s*cm.*?----\s*高度\s*([\d.]+)\s*m[^c]', s)
    if m:
        w, l, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('外径', l, w, h)
    
    # 格式11-13: 宽度10【白色】;【100个外尺寸】长度21cm----高度2cm
    m = re.search(r'宽度\s*([\d.]+).*?【\d+个外尺寸】\s*长度\s*([\d.]+)\s*cm.*?高度\s*([\d.]+)\s*cm', s)
    if m:
        w, l, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('外径', l, w, h)
    
    # 格式14-15: 红色【100个】;长宽高10*10*2cm
    m = re.search(r'长宽高\s*([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)', s)
    if m and '长宽高' in s:
        return ('外径', float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # === 内径格式 ===
    # 格式16: 内宽x高【10x10】=外宽x高【10.5x10.5】cm;【100个】内尺寸【长度10cm】=【外长11.5cm】
    m = re.search(r'内\s*宽x高\s*【\s*([\d.]+)\s*x\s*([\d.]+)\s*】.*?内尺寸【?\s*长度\s*([\d.]+)', s)
    if m:
        w, h, l = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('内径', l, w, h)
    
    # === 外径格式（双面白色+外尺寸=内尺寸）===
    # 格式18: 双面白色【100个】高档;外尺寸【11x11x11cm】=内尺寸
    m = re.search(r'外尺寸【\s*([\d.]+)\s*x\s*([\d.]+)\s*x\s*([\d.]+)', s)
    if m and '外尺寸' in s:
        return ('外径', float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # 格式21: 外宽x高【9x9】=内宽x高【8.5x8.5】cm;【N个】外尺寸【长度9cm】=【内长7.5cm
    m = re.search(r'外\s*宽x高\s*【\s*([\d.]+)\s*x\s*([\d.]+)\s*】.*?外尺寸【?\s*长度\s*([\d.]+)', s)
    if m and '外宽x高' in s:
        w, h, l = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return ('外径', l, w, h)
    
    # 如果以上都不匹配 → 定制
    return ('定制', None, None, None)

# ===== 分类 =====
results = {'定制': [], '纸箱': [], '外径': [], '内径': []}

for row in shop_rows:
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    cat, l, w, h = classify_koudi(s)
    results[cat].append((row, l, w, h))

print('\n分类结果:')
for cat in ['定制', '纸箱', '外径', '内径']:
    items = results[cat]
    print(f'  {cat}: {len(items)} 条')
    if items and len(items) <= 5:
        for row, l, w, h in items:
            spec = str(row[2] or '')[:80]
            lwh = f'→ {l}x{w}x{h}' if l else ''
            print(f'    {spec} {lwh}')

# ===== 输出文件（仅此店铺）=====
def write_shop_xlsx(data_list, fpath, label):
    if not data_list:
        print(f'  {fpath}: 空，跳过')
        return
    rows = [r[0] for r in data_list]
    pd.DataFrame(rows, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_shop_xlsx(results['定制'], f'{OUTPUT_DIR}\\定制链接商品_天猫扣底盒.xlsx', '定制链接商品-天猫扣底盒')
write_shop_xlsx(results['纸箱'], f'{OUTPUT_DIR}\\纸箱商品_天猫扣底盒.xlsx', '纸箱商品-天猫扣底盒')
write_shop_xlsx(results['外径'], f'{OUTPUT_DIR}\\外径全量飞机盒_天猫扣底盒.xlsx', '外径全量飞机盒-天猫扣底盒')
write_shop_xlsx(results['内径'], f'{OUTPUT_DIR}\\内径全量飞机盒_天猫扣底盒.xlsx', '内径全量飞机盒-天猫扣底盒')

print('\n✅ 天猫扣底盒处理完成！')
print('请检查桌面上的文件是否正确，再发下一个店铺。')

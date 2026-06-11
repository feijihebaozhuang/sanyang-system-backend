# -*- coding: utf-8 -*-
"""
从平台商品.xlsx中提取：
1. 长宽高为整数的商品
2. 长宽高三项全为.5的商品
输出到桌面
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

f = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(f, header=1, dtype=str)

def extract_lwh(s):
    s = str(s)
    patterns = [
        r'外[尺寸]*寸*[大小]*[：:]?\s*【\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*cm*\s*】',
        r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】',
        r'长[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】',
        r'(?:外尺寸|外径|外寸)[^；;]*?【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】',
        r'(?:长|长度)[：:]?\s*([\d.]+)\s*cm*\s*(?:宽|宽度)[：:]?\s*([\d.]+)\s*cm*\s*(?:高|高度)[：:]?\s*([\d.]+)\s*cm*',
        r'^\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*(?:cm|mm|C)?',
        r'(?:^|[；;，,\s])+\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(?:cm|mm|C)?(?:\s|$|【|；|;|,)',
    ]
    for pat in patterns:
        m = re.search(pat, s, re.I | re.S)
        if m:
            try:
                l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
                if 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
                    return (l, w, h)
            except: pass
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)(?:\s*cm|\s*mm|\s*C)?(?![\d.])', s)
    if m:
        try:
            l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
            if 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
                return (l, w, h)
        except: pass
    return None

def is_dot5(v):
    return v != int(v) and v * 2 == int(v * 2)

int_results = []
dot5_results = []
unrec = []

total = len(df)
for idx, (_, row) in enumerate(df.iterrows()):
    s = str(row.get('平台规格名称', '')).strip()
    if not s: continue
    
    lwh = extract_lwh(s)
    if not lwh:
        unrec.append(s)
        continue
    
    l, w, h = lwh
    shop = str(row.get('店铺名称', '')).strip()
    pid = str(row.get('平台商品id', '')).strip()
    sid = str(row.get('平台规格id', '')).strip()
    
    all_int = all(n == int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    
    if all_int:
        int_results.append((shop, pid, s, int(l), int(w), int(h)))
    elif all_dot5:
        dot5_results.append((shop, pid, s, l, w, h))

print(f'总行: {total}')
print(f'长宽高为整数: {len(int_results)}')
print(f'长宽高三项全.5: {len(dot5_results)}')
print(f'已提取小数非全.5(上一个文件): 17276')
print(f'未提取到长宽高: {len(unrec)}')

# 输出整数
out1 = r'D:\Desktop\长宽高为整数商品.xlsx'
pd.DataFrame(int_results, columns=['店铺名称', '平台商品id', '规格名称', '长', '宽', '高']).to_excel(out1, index=False)
import openpyxl
wb = openpyxl.load_workbook(out1)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '长宽高为整数的商品'
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
wb.save(out1)
print(f'\n✅ 长宽高为整数: {out1} ({len(int_results)}条)')

# 输出三项全.5
out2 = r'D:\Desktop\长宽高三项全5商品.xlsx'
pd.DataFrame(dot5_results, columns=['店铺名称', '平台商品id', '规格名称', '长', '宽', '高']).to_excel(out2, index=False)
wb2 = openpyxl.load_workbook(out2)
ws2 = wb2.active
ws2.insert_rows(1)
ws2.cell(1, 2).value = '长宽高三项全为.5的商品'
ws2.column_dimensions['C'].width = 60
wb2.save(out2)
print(f'\n✅ 三项全.5: {out2} ({len(dot5_results)}条)')

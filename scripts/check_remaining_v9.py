# -*- coding: utf-8 -*-
"""检查其余商品里是否有含内径关键词的"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

print('其余商品前30条:\n')
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    count += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    has_inner = '✅内径' if '内径' in s else ''
    print(f'  [{count}] {has_inner} {s[:130]}')
    if count >= 30: break
wb.close()

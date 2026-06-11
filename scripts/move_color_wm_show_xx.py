# -*- coding: utf-8 -*-
"""天猫彩色436条+天猫小批量2条→无匹配。然后展示阿里新鑫星的格式"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter, defaultdict

out = r'd:\Desktop\换绑输出'
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# 分类
to_wm = []  # 移到无匹配
remaining = []

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    if shop in ('天猫彩色', '天猫小批量'):
        to_wm.append(r)
    else:
        remaining.append(r)

print(f'移到无匹配: {len(to_wm)}条（天猫彩色+天猫小批量）')
print(f'平卡剩余: {len(remaining)}条')

# 追加到无匹配
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
for r in to_wm:
    if len(r) >= 6:
        ws.append(list(r[:6]))
    else:
        ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
wb.close()
print(f'✅ 已追加{len(to_wm)}条到无匹配_待处理')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()

# ====== 阿里新鑫星格式展示 ======
xx_items = [r for r in remaining if r and str(r[0] or '').strip() == '阿里新鑫星']
print(f'\n{"="*60}')
print(f'阿里新鑫星: {len(xx_items)}条')
print(f'{"="*60}')

# 按规格名称前缀分组
specs = Counter()
for r in xx_items:
    spec = str(r[2] or '').strip()
    # 提取关键部分
    prefix = spec[:50] if len(spec) > 50 else spec
    specs[prefix] += 1

print(f'\n所有不同格式({len(specs)}种):')
for k, v in sorted(specs.items(), key=lambda x: -x[1]):
    print(f'  [{v:>4}] {k}')

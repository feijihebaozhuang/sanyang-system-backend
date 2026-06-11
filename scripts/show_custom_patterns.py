# -*- coding: utf-8 -*-
"""定制链接：按店铺+规格模式分类，每个模式举一个例子"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

patterns = {}  # key: (店铺简称, 模式签名)  value: (规格原文, 平台商品id)

def pattern_sig(s):
    """生成规格的模式签名——保留括号结构、关键词、数字位数"""
    sig = s
    # 替换具体数字为 N
    sig = re.sub(r'\d+\.?\d*', 'N', sig)
    # 压缩空格
    sig = re.sub(r'\s+', ' ', sig)
    return sig[:200]

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    short_shop = str(row[4] or '').strip() if len(row) > 4 else '(无名)'
    spec = str(row[2] or '').strip() if len(row) > 2 else ''
    pid = str(row[3] or '') if len(row) > 3 else ''
    if not spec: continue
    
    sig = pattern_sig(spec)
    key = (short_shop, sig)
    if key not in patterns:
        patterns[key] = (spec, pid)

wb.close()

print(f'定制链接共 {len(patterns)} 个格式（按店铺+规格模式）\n')

# 按店铺分组输出
by_shop = {}
for (shop, sig), (spec, pid) in patterns.items():
    by_shop.setdefault(shop, []).append((sig, spec, pid))

for shop in sorted(by_shop):
    items = by_shop[shop]
    print(f'══ {shop}（{len(items)} 种格式）══')
    for sig, spec, pid in items:
        # 显示规格原文（截断）
        short_spec = spec[:150]
        print(f'  例: {short_spec}')
        print(f'      pid={pid}')
    print()

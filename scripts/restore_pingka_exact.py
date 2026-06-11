# -*- coding: utf-8 -*-
"""精确恢复平卡数据：平卡应有 = 原始 - 换绑 - 定制 - 无匹配"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import defaultdict, Counter

out = r'd:\Desktop\换绑输出'
raw_file = r'd:\Desktop\平台商品.xlsx'

# 1. 收集所有已处理的spec_id（换绑+定制+无匹配）
processed_specs = set()

for f in sorted(os.listdir(out)):
    if not f.endswith('.xlsx') or '平卡' in f:
        continue
    try:
        wb = oxl.load_workbook(os.path.join(out, f))
        for sn in wb.sheetnames:
            for r in wb[sn].iter_rows(min_row=2 if ('无匹配' in f or '平卡' in f) else 3, values_only=True):
                if not r: continue
                sid = str(r[2] or '').strip() if len(r) >= 3 else ''
                if sid:
                    processed_specs.add(sid)
        wb.close()
    except:
        pass

print(f'已处理spec_id总数(去重): {len(processed_specs)}')

# 2. 扫描原始数据，按店铺统计哪些还没处理
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
raw_data = []  # (shop, pid, spec_name, spec_id)

shop_raw = Counter()
shop_matched = Counter()
shop_unmatched = Counter()

for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    shop = str(r[0]).strip()
    pid = str(r[1] or '').strip() if len(r) >= 2 else ''
    spec_name = str(r[2] or '').strip() if len(r) >= 3 else ''
    spec_id = str(r[3] or '').strip() if len(r) >= 4 else ''
    
    shop_raw[shop] += 1
    if spec_id in processed_specs:
        shop_matched[shop] += 1
    else:
        shop_unmatched[shop] += 1
        raw_data.append((shop, pid, spec_name, spec_id))

wb.close()

print(f'\n=== 每个店铺原始 vs 已处理 vs 未处理 ===')
for shop in sorted(shop_raw.keys()):
    raw = shop_raw[shop]
    matched = shop_matched.get(shop, 0)
    unmatched = shop_unmatched.get(shop, 0)
    print(f'  {shop:12s} 原始={raw:>7} 已处理={matched:>7} 应留平卡={unmatched:>7}')

# 3. 看当前平卡有谁
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
current_pingka = defaultdict(list)  # shop -> [(pid, spec_name, spec_id)]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    shop = str(r[0]).strip()
    pid = str(r[1] or '').strip() if len(r) >= 2 else ''
    spec_name = str(r[2] or '').strip() if len(r) >= 3 else ''
    spec_id = str(r[3] or '').strip() if len(r) >= 4 else ''
    current_pingka[shop].append((pid, spec_name, spec_id))
wb.close()

print(f'\n=== 当前平卡 vs 应有数量 ===')
for shop in sorted(shop_raw.keys()):
    expected = shop_unmatched.get(shop, 0)
    current = len(current_pingka.get(shop, []))
    diff = current - expected
    status = '✅' if diff == 0 else ('❌少了' if diff < 0 else '❌多了')
    if diff != 0:
        print(f'  {shop:12s} 应有={expected:>7} 现有={current:>7} 差额={diff:>+7} {status}')

# 4. 恢复丢失数据
need_restore = []
for shop in sorted(shop_raw.keys()):
    expected = shop_unmatched.get(shop, 0)
    current = len(current_pingka.get(shop, []))
    if current < expected:
        # 需要找回的数据 = 原始未处理 - 已在平卡的spec_id
        existing_ids = set(sid for (_, _, sid) in current_pingka.get(shop, []))
        for raw_shop, pid, spec_name, spec_id in raw_data:
            if raw_shop == shop and spec_id not in existing_ids:
                need_restore.append((shop, pid, spec_name, spec_id))

if need_restore:
    print(f'\n=== 需要恢复 {len(need_restore)} 条数据到平卡 ===')
    by_shop = Counter(s[0] for s in need_restore)
    for s, c in by_shop.most_common():
        print(f'  {s}: {c}条')
    
    # 追加到平卡
    wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
    ws = wb['平卡']
    for shop, pid, spec_name, spec_id in need_restore:
        ws.append([shop, pid, spec_name, spec_id, '平卡/待处理', ''])
    wb.save(os.path.join(out, '平卡_待处理.xlsx'))
    wb.close()
    print(f'✅ 已恢复{len(need_restore)}条到平卡_待处理')
else:
    print(f'\n✅ 平卡数据完整，无需恢复')

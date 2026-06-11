# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

other_file = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(other_file, read_only=True)
ws = wb.active

count_neijing = 0
samples = []

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 2: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if not s: continue
    
    # 含"内径"、"内尺寸"、"内寸"关键词
    if '内径' in s or '内尺寸' in s or '内寸' in s:
        count_neijing += 1
        if len(samples) < 20:
            samples.append(s[:100])

wb.close()

print(f'其余商品中"内径"类规格: {count_neijing} 条')
print('\n样本:')
for s in samples:
    print(f'  {s}')

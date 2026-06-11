# -*- coding: utf-8 -*-
"""查找所有xlsx文件并读取品牌店数据"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 直接打开平卡看品牌店还在不在
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

pp = [r for r in rows if r and '品牌店' in str(r[0] or '')]
print(f'平卡品牌店剩余: {len(pp)}')

# 查看第11批的大小
f11 = os.path.join(out, '换绑文件_第11批.xlsx')
sz = os.path.getsize(f11)
print(f'第11批大小: {sz/1024:.1f}KB')

# 打开看看有多少条
wb = oxl.load_workbook(f11)
ws = wb['Sheet1']
rows11 = list(ws.iter_rows(min_row=3, values_only=True))
wb.close()
print(f'第11批总条数: {len(rows11)}')

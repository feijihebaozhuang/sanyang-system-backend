# -*- coding: utf-8 -*-
"""按店铺统计每条数据去向"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter, defaultdict

out = r'd:\Desktop\换绑输出'
raw_file = r'd:\Desktop\平台商品.xlsx'

# 先收集所有已处理的 spec_id
handled_specs = {}  # spec_id -> (在哪, 店铺简称)
duplicates = []

# 换绑文件
for f in sorted(os.listdir(out)):
    if not f.endswith('.xlsx') or '平卡' in f or '无匹配' in f or '未匹配' in f or '定制' in f:
        continue
    try:
        wb = oxl.load_workbook(os.path.join(out, f))
        for sn in wb.sheetnames:
            for r in wb[sn].iter_rows(min_row=3, values_only=True):
                if r and len(r) >= 3 and r[2]:
                    sid = str(r[2]).strip()
                    shop = str(r[0] or '').strip() if r[0] else ''
                    if sid in handled_specs:
                        duplicates.append((sid, handled_specs[sid], f, shop))
                    handled_specs[sid] = f'换绑:{f}'
        wb.close()
    except: pass

# 定制文件
for f in sorted(os.listdir(out)):
    if '定制' not in f or not f.endswith('.xlsx'):
        continue
    try:
        wb = oxl.load_workbook(os.path.join(out, f))
        for sn in wb.sheetnames:
            for r in wb[sn].iter_rows(min_row=3, values_only=True):
                if r and len(r) >= 3 and r[2]:
                    sid = str(r[2]).strip()
                    if sid in handled_specs:
                        duplicates.append((sid, handled_specs[sid], f, ''))
                    handled_specs[sid] = f'定制:{f}'
        wb.close()
    except: pass

# 无匹配
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) >= 3 and r[2]:
        sid = str(r[2]).strip()
        if sid in handled_specs:
            duplicates.append((sid, handled_specs[sid], '无匹配', ''))
        handled_specs[sid] = '无匹配'
wb.close()

print(f'总处理spec_id: {len(handled_specs)}')
print(f'重复条数: {len(duplicates)}')

if duplicates:
    print(f'\n重复样例(前10):')
    for d in duplicates[:10]:
        print(f'  spec_id={d[0][:20]} 已在{d[1]}, 又在{d[2]}')

# 现在看原始数据中每个店铺的spec_id是否在handled中
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
raw_shops = Counter()
missing_shops = Counter()
handled_by_shop = Counter()
for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    shop = str(r[0]).strip()
    sid = str(r[3] or '').strip() if len(r) >= 4 else ''
    raw_shops[shop] += 1
    if sid in handled_specs:
        handled_by_shop[shop] += 1
    else:
        missing_shops[shop] += 1
wb.close()

print(f'\n=== 按店铺统计 ===')
print(f'{"店铺":20s} {"原始":>8s} {"已处理":>8s} {"未处理":>8s} {"处理率":>8s}')
for shop in sorted(raw_shops.keys()):
    raw = raw_shops[shop]
    h = handled_by_shop.get(shop, 0)
    m = missing_shops.get(shop, 0)
    rate = f'{h/raw*100:.1f}%' if raw > 0 else '-'
    print(f'{shop:20s} {raw:>8} {h:>8} {m:>8} {rate:>8s}')

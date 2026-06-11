# -*- coding: utf-8 -*-
"""定制类：珍珠棉+蓝绿+定制关键词 → 定制链接换绑文件，尺寸不足单独"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl

src = r'd:\Desktop\换绑输出\剩余平台商品_定制类.xlsx'
out_dir = r'd:\Desktop\换绑输出'

print("读取定制类文件...", flush=True)
wb = oxl.load_workbook(src, data_only=True)
ws = wb['定制类']

rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

custom_link = []  # 珍珠棉+蓝绿+定制关键词 → 定制链接
size_miss = []    # 尺寸不足 → 单独商量

for r in rows:
    if not r:
        continue
    reason = str(r[4] or '').strip()
    if reason in ('珍珠棉', '蓝绿颜色', '定制关键词'):
        custom_link.append(r)
    else:
        size_miss.append(r)

print(f"定制链接类: {len(custom_link)} 条", flush=True)
print(f"  珍珠棉: {sum(1 for r in custom_link if r[4]=='珍珠棉')}", flush=True)
print(f"  蓝绿颜色: {sum(1 for r in custom_link if r[4]=='蓝绿颜色')}", flush=True)
print(f"  定制关键词: {sum(1 for r in custom_link if r[4]=='定制关键词')}", flush=True)
print(f"尺寸不足待处理: {len(size_miss)} 条", flush=True)

# 定制链接换绑文件
wb1 = oxl.Workbook()
ws1 = wb1.active
ws1.title = 'Sheet1'
ws1.append([None, '商品对应表', None, None])
ws1.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
for r in custom_link:
    ws1.append([str(r[0] or '').strip(), str(r[1] or '').strip(), str(r[2] or '').strip(), '定制链接'])

f1 = os.path.join(out_dir, '定制类_换绑文件.xlsx')
wb1.save(f1)
sz1 = os.path.getsize(f1)
print(f"\n✅ 定制类换绑: {f1} ({sz1/1024:.1f}KB, {len(custom_link)}条)", flush=True)
wb1.close()

# 尺寸不足单独文件
if size_miss:
    wb2 = oxl.Workbook()
    ws2 = wb2.active
    ws2.title = '尺寸不足'
    ws2.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for r in size_miss:
        ws2.append(list(r))
    f2 = os.path.join(out_dir, '尺寸不足_待处理.xlsx')
    wb2.save(f2)
    sz2 = os.path.getsize(f2)
    print(f"✅ 尺寸不足待处理: {f2} ({sz2/1024:.1f}KB, {len(size_miss)}条)", flush=True)
    wb2.close()

print("\n完成！", flush=True)

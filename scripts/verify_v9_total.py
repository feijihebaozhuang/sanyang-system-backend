# -*- coding: utf-8 -*-
"""验证桌面文件总数"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

files = {
    '定制链接商品.xlsx': None,
    '扣底盒双插盒商品.xlsx': None,
    '纸箱商品.xlsx': None,
    '内径全量飞机盒.xlsx': None,
    '外径全量飞机盒.xlsx': None,
    '非全量飞机盒.xlsx': None,
    '其余商品.xlsx': None,
}

base = r'D:\Desktop'
total = 0

for fname in files:
    fpath = os.path.join(base, fname)
    wb = load_workbook(fpath, read_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(values_only=True):
        pass  # 只计数
    # 更好的计数方式
    # 实际用ws.max_row - 2 (减去header行和标题行)
    max_row = ws.max_row
    data_rows = max_row - 2 if max_row > 2 else 0
    wb.close()
    print(f'  {fname}: {data_rows} 条')
    total += data_rows

print(f'\n总计: {total} 条')
print(f'应计: 498090 条')
print(f'差值: {498090 - total}')

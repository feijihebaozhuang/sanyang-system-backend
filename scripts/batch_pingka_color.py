# -*- coding: utf-8 -*-
"""处理平卡里剩余的天猫彩色436条和天猫小批量2条"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 找最新批次数
exist = [f for f in os.listdir(ok_dir) if f.startswith('换绑文件_第') and f.endswith('.xlsx')]
max_batch = 0
for f in exist:
    m = re.search(r'第(\d+)批', f)
    if m: max_batch = max(max_batch, int(m.group(1)))
new_batch = max_batch + 1

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
km = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in km.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dims[0]:g}*{dims[1]:g}*{dims[2]:g}"
    cand = f"{base}-{dk}-{mat}"
    if cand in all_codes: return cand
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

target_shops = {'天猫彩色', '天猫小批量'}

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if shop not in target_shops:
        remaining.append(r)
        continue
    
    matched = False
    
    # ===== 天猫彩色: 宽度：10 cm---高度：9 cm;白色 100个 长度：41 cm =====
    # → 41 10 9 外径 白色
    if shop == '天猫彩色':
        m = re.search(r'宽度[：:]\s*([\d.]+)\s*cm---高度[：:]\s*([\d.]+)\s*cm[^;]*;白色\s*\d+个\s*长度[：:]\s*([\d.]+)\s*cm', spec_name)
        if m:
            w, h, l = float(m.group(1)), float(m.group(2)), float(m.group(3))
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], '外径', '白色')
            if code:
                matched_new.append((shop, pid, spec_id, code))
                stats['天猫彩色-外径白色匹配'] += 1
            else:
                # 外转内试试
                il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                if il > 0:
                    code_i = match_3d(il, iw, ih, '内径', '白色')
                    if code_i:
                        matched_new.append((shop, pid, spec_id, code_i))
                        stats['天猫彩色-外转内白色匹配'] += 1
                    else:
                        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-白色'))
                        stats['天猫彩色-无匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-白色'))
                    stats['天猫彩色-无匹配'] += 1
            matched = True
    
    # ===== 天猫小批量: 优质进口纸-黄色【100个】;37*14**3.8 =====
    if shop == '天猫小批量':
        m = re.search(r'[^;]*;(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)\s*\*{1,2}\s*(\d+\.?\d*)', spec_name)
        if m:
            dims_vals = sorted([float(m.group(1)), float(m.group(2)), float(m.group(3))], reverse=True)
            
            # 判断材料：双白色→白色，其他→特硬
            mat = '白色' if '双白' in spec_name or '双白色' in spec_name else '特硬'
            
            code = match_3d(dims_vals[0], dims_vals[1], dims_vals[2], '外径', mat)
            if code:
                matched_new.append((shop, pid, spec_id, code))
                stats['天猫小批量-外径%s匹配' % mat] += 1
            else:
                # 外转内
                il, iw, ih = dims_vals[0]-1.5, dims_vals[1]-0.5, dims_vals[2]-0.5
                if il > 0:
                    code_i = match_3d(il, iw, ih, '内径', mat)
                    if code_i:
                        matched_new.append((shop, pid, spec_id, code_i))
                        stats['天猫小批量-外转内匹配'] += 1
                    else:
                        nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}'))
                        stats['天猫小批量-无匹配'] += 1
                else:
                    nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims_vals[0])}*{dfmt(dims_vals[1])}*{dfmt(dims_vals[2])}-外径-{mat}'))
                    stats['天猫小批量-无匹配'] += 1
            matched = True
    
    if not matched:
        remaining.append(r)
        stats['未识别-%s' % shop] += 1

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'未识别: {sum(v for k,v in stats.items() if k.startswith("未识别"))}')

# 保存换绑
if matched_new:
    f_out = os.path.join(ok_dir, f'换绑文件_第{new_batch}批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f_out)
    wb.close()
    print(f'\n✅ 换绑文件_第{new_batch}批: {len(matched_new)}条')

# 无匹配追加
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 追加{len(nomatch)}条到无匹配')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡已更新: {len(remaining)}条')

print('\n完成！')

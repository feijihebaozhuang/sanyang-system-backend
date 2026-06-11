# -*- coding: utf-8 -*-
"""修复换绑文件店铺名称为全称"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

ok_dir = r'd:\Desktop\换绑输出\OK文件'

# 简称→全称映射（从定制类已成功上传的文件中提取）
NAME_MAP = {
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
    '阿里友尚': '深圳市友尚包装有限公司',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '俊鑫': '俊鑫纸品厂',
    '当下家': '当下家包装',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝品牌店': '飞机盒品牌店',
    '平卡': '平卡包装',  # 应该没有这个
}

# 需要修复的批次：第5批之后的我生成的所有批次
to_fix = []
for f in os.listdir(ok_dir):
    if not (f.startswith('换绑文件_第') or f.startswith('定制类_换绑文件')) or not f.endswith('.xlsx'):
        continue
    to_fix.append(f)

to_fix.sort(key=lambda x: (
    # 定制类优先
    0 if x.startswith('定制类') else 1,
    # 按批次号排序
    int(re.search(r'第(\d+)', x).group(1)) if re.search(r'第(\d+)', x) else 0
))

fixed_count = 0
for fn in to_fix:
    fp = os.path.join(ok_dir, fn)
    wb = oxl.load_workbook(fp)
    ws = wb[wb.sheetnames[0]]
    
    changed = False
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column == 1 and cell.value and str(cell.value).strip() in NAME_MAP:
                old = str(cell.value).strip()
                new = NAME_MAP[old]
                cell.value = new
                changed = True
                fixed_count += 1
    
    if changed:
        wb.save(fp)
        print('✅ %s 已修复' % fn)
    else:
        print('  %s 无需修复' % fn)
    wb.close()

print('\n共修复 %d 个单元格' % fixed_count)

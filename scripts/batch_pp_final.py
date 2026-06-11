# -*- coding: utf-8 -*-
"""修正RE_C后直接处理品牌店全部632条"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ========== 快麦索引 ==========
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ========== 修正正则 ==========

# 模式B: 进口优质特硬E瓦-内径;长x宽【100x100】mm;100mm【高】
RE_B = re.compile(
    r'进口优质.*?(内径|外经|外径)\s*;\s*长x宽[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*mm?[】]*\s*;\s*(\d+\.?\d*)\s*mm?'
)

# 模式C: 【双面白色】内径/外径;【36x36 】;10 cm高度（单个价）
# 注意：可能会有(单个价)后缀
RE_C = re.compile(
    r'【([^】]*)】\s*(内径|外径)\s*;\s*[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)[】]*\s*;\s*(\d+\.?\d*)\s*cm?\s*高度'
)

RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_BAISE = re.compile(r'双面白色|双面白|白色|双白')

def guess_mat(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_BAISE.search(s): return '白色'
    return '特硬'

# ========== 读平卡 ==========
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# ========== 处理 ==========
matched_new = []
nomatch_details = []
remaining = []
stats = Counter()

for idx, r in enumerate(rows):
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    if '品牌店' not in shop:
        remaining.append(r)
        continue
    
    handled = False
    
    # ---- 模式B: 进口优质特硬 ----
    m = RE_B.search(spec_name)
    if m:
        dk_raw = m.group(1).strip()
        dk = '内径' if '内' in dk_raw else '外径'
        x_mm = float(m.group(2))
        y_mm = float(m.group(3))
        h_mm = float(m.group(4))
        l, w, h = x_mm/10, y_mm/10, h_mm/10
        mat = '特硬'
        code = match_3d(l, w, h, dk, mat)
        if code:
            matched_new.append((fn(shop), pid, spec_id, code))
            stats['模式B-匹配'] += 1
        else:
            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                f"{dfmt(l)}*{dfmt(w)}*{dfmt(h)}-{dk}-{mat}"))
            stats['模式B-无匹配'] += 1
        handled = True
    
    # ---- 模式C: 【材料】内径/外径;【XxY】;Zcm高度 ----
    if not handled:
        m = RE_C.search(spec_name)
        if m:
            mat_ctx = m.group(1).strip()
            dk = m.group(2).strip()
            x = float(m.group(3))
            y = float(m.group(4))
            h = float(m.group(5))
            
            l, w = max(x, y), min(x, y)
            dims = sorted([l, w, h], reverse=True)
            mat = guess_mat(spec_name)
            
            out_line = f'[{idx}] {spec_name[:60]} -> dims={dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])} {dk} {mat}'
            
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['模式C-匹配'] += 1
                out_line += f' ✅ 匹配={code}'
            else:
                # 外径转内径尝试
                if dk == '外径':
                    il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                    if il > 0 and iw > 0 and ih > 0:
                        code_i = match_3d(il, iw, ih, '内径', mat)
                        if code_i:
                            matched_new.append((fn(shop), pid, spec_id, code_i))
                            stats['模式C-外转内匹配'] += 1
                            out_line += f' ✅ 外转内={code_i}'
                        else:
                            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                                f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                            stats['模式C-外转内无匹配'] += 1
                            out_line += f' ❌ 无匹配(外转内)'
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                        stats['模式C-无匹配'] += 1
                        out_line += f' ❌ 无匹配'
                else:
                    # 内径也试外径
                    ol, ow, oh = dims[0]+1.5, dims[1]+0.5, dims[2]+0.5
                    code_o = match_3d(ol, ow, oh, '外径', mat)
                    if code_o:
                        matched_new.append((fn(shop), pid, spec_id, code_o))
                        stats['模式C-内转外匹配'] += 1
                        out_line += f' ✅ 内转外={code_o}'
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                        stats['模式C-无匹配'] += 1
                        out_line += f' ❌ 无匹配'
            
            # 只打印无匹配的
            if '❌' in out_line:
                print(out_line)
            elif '✅' in out_line:
                pass
            
            handled = True
    
    if not handled:
        remaining.append((shop, pid, spec_id, spec_name, '平卡/解析失败', ''))

print(f'\n====== 品牌店处理结果 ======')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配(待确认): {len(nomatch_details)}')
print(f'平卡剩余(非品牌店): {len([r for r in remaining if r])}')

# ========== 追加到第11批 ==========
if matched_new:
    f11 = os.path.join(out, '换绑文件_第11批.xlsx')
    # 如果已存在，追加
    if os.path.exists(f11):
        wb = oxl.load_workbook(f11)
        ws = wb['Sheet1']
        for m in matched_new:
            ws.append(list(m))
        wb.save(f11)
    else:
        wb = oxl.Workbook()
        ws = wb.active; ws.title = 'Sheet1'
        ws.append([None, '商品对应表', None, None])
        ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
        for m in matched_new:
            ws.append(list(m))
        wb.save(f11)
    wb.close()
    print(f'\n✅ 换绑文件_第11批更新完成: {len(matched_new)}条新增')

# ========== 保存待确认 ==========
if nomatch_details:
    wb = oxl.Workbook()
    ws = wb.active; ws.title = '待确认'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for d in nomatch_details:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '品牌店_待确认.xlsx'))
    print(f'✅ 品牌店_待确认: {len(nomatch_details)}条')
    wb.close()

# ========== 更新平卡 ==========
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f'✅ 平卡_待处理已更新: {len(remaining)}条')
wb.close()

print('\n完成！')

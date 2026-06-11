# -*- coding: utf-8 -*-
"""品牌店无匹配移到无匹配文件"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

pp_miss = []
remaining = []
for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    reason = str(r[4] or '').strip()
    
    if '品牌店' in shop and reason == '无匹配':
        pp_miss.append(r)
    else:
        remaining.append(r)

print(f"品牌店无匹配: {len(pp_miss)}条", flush=True)
print(f"平卡剩余: {len(remaining)}条", flush=True)

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()

# 追加到无匹配
if pp_miss:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'), data_only=True)
    ws = wb['无匹配']
    for r in pp_miss:
        ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f"✅ 无匹配已追加{len(pp_miss)}条")

print("完成！")

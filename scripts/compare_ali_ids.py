# -*- coding: utf-8 -*-
import sys, os, pandas as pd
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 1. 定制文件里的友尚数据
custom_file = os.path.join(out, '定制类_换绑文件.xlsx')
wb = oxl.load_workbook(custom_file, data_only=True)
ws = wb['Sheet1']
custom_rows = list(ws.iter_rows(min_row=3, values_only=True))
wb.close()

custom_ali = []
for r in custom_rows:
    if '友尚' in str(r[0] or ''):
        custom_ali.append(r)

print(f"定制文件中阿里友尚: {len(custom_ali)}条", flush=True)
for r in custom_ali[:5]:
    print(f"  商品id={r[1]}, 规格id={r[2]}")

# 2. 看未匹配平台商品里友尚的规格ID
f_unmatch = os.path.join(out, '未匹配平台商品.xlsx')
wb = oxl.load_workbook(f_unmatch, data_only=True)
ws = wb['未匹配']
unmatch_rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

unmatch_ali = []
for r in unmatch_rows:
    if '友尚' in str(r[0] or ''):
        unmatch_ali.append(r)

print(f"\n未匹配中阿里友尚: {len(unmatch_ali)}条")
for r in unmatch_ali[:5]:
    print(f"  商品id={r[1]}, 规格id={r[2]}, 规格名={str(r[3])[:40]}")

# 3. 对比：定制里的友尚规格ID 在 未匹配里能不能找到
custom_spec_ids = set(str(r[2]).strip() for r in custom_ali)
unmatch_spec_ids = set(str(r[2]).strip() for r in unmatch_ali)

missing = custom_spec_ids - unmatch_spec_ids
if missing:
    print(f"\n❌ 定制中有{len(missing)}个规格ID在未匹配中找不到！")
    for sid in list(missing)[:5]:
        print(f"  {sid}")
else:
    print(f"\n✅ 定制中所有{len(custom_spec_ids)}个友尚规格ID都能在未匹配中找到")

# 也检查一下新鑫星
for shop_keyword, shop_name in [('友尚', '阿里友尚'), ('新鑫星', '阿里新鑫星')]:
    print(f"\n{'='*50}")
    print(f"检查【{shop_name}】:")
    
    custom_rows_for_shop = [r for r in custom_rows if shop_keyword in str(r[0] or '')]
    print(f"  定制文件中有 {len(custom_rows_for_shop)} 条")
    
    unmatch_rows_for_shop = [r for r in unmatch_rows if shop_keyword in str(r[0] or '')]
    print(f"  未匹配文件中有 {len(unmatch_rows_for_shop)} 条")
    
    custom_ids = set(str(r[2]).strip() for r in custom_rows_for_shop)
    unmatch_ids = set(str(r[2]).strip() for r in unmatch_rows_for_shop)
    
    not_in_unmatch = custom_ids - unmatch_ids
    if not_in_unmatch:
        print(f"  ❌ 定制中有 {len(not_in_unmatch)} 个规格ID在未匹配中缺失:")
        for sid in list(not_in_unmatch)[:10]:
            print(f"    {sid}")
    else:
        print(f"  ✅ 全部 {len(custom_ids)} 个规格ID都能在未匹配中找到")

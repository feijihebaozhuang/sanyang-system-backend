# -*- coding: utf-8 -*-
"""看定制第6批的内容"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

ok_dir = r'd:\Desktop\换绑输出\OK文件'

wb = oxl.load_workbook(os.path.join(ok_dir, '定制类_换绑文件_第6批.xlsx'))
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(min_row=1, values_only=True))
wb.close()

print(f'总行数(含表头): {len(rows)}')
for i, r in enumerate(rows[:5]):
    print(f'  行{i+1}: {[str(c)[:30] if c else "" for c in r]}')

# 统计店铺
from collections import Counter
shops = Counter()
for r in rows[2:]:
    if r and r[0]:
        shops[str(r[0])] += 1
print(f'\n店铺分布:')
for s, c in shops.most_common():
    print(f'  {s}: {c}条')

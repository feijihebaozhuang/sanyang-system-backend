# -*- coding: utf-8 -*-
"""批量处理品牌店7415条"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ============================================================
# 快麦索引
# ============================================================
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}", f"{base}-{dk}-特硬"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ============================================================
# 正则预编译
# ============================================================

# 模式1: 宽度【120mm】双面白;长度【130 mm】外径【100个】;【20 mm】
# 提取：宽度mm, 长度mm, 末尾mm, 材料来自"双面白/台湾纸/特硬"关键词
RE_M1 = re.compile(
    r'宽度[【】\s]*(\d+\.?\d*)\s*mm?[】]*\s*(.*?);\s*长度[【】\s]*(\d+\.?\d*)\s*mm?[】]*\s*'
    r'外径[【】\s]*[^】]*[】]*\s*;\s*[【】\s]*(\d+\.?\d*)\s*mm?'
)

# 模式2: （宽）135 mm 外径;长度 135 mm;105 mm 台湾纸
RE_M2 = re.compile(
    r'[（(]宽[）)]\s*(\d+\.?\d*)\s*mm?\s*外径\s*;\s*长度\s*(\d+\.?\d*)\s*mm?\s*;\s*(\d+\.?\d*)\s*mm?\s*(.*)'
)

# 模式3: （台湾纸）24x2 cm 单个;外径【单个价】长度 30 cm
RE_M3 = re.compile(
    r'[（(]([^）)]*)[）)]\s*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*cm\s*.*?外径.*?长度\s*(\d+\.?\d*)\s*cm'
)

# 模式4: 宽度【13*10 】cm【100个】;外径【双面白】长度 38 cm
RE_M4 = re.compile(
    r'宽度[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*cm?[】]*.*?'
    r'外径[【】\s]*([^】]*)[】]*\s*长度\s*(\d+\.?\d*)\s*cm'
)

# 模式5: 宽度【170mm】;长度【170mm】 外径【双面白】;【90 mm】【100个】
RE_M5 = re.compile(
    r'宽度[【】\s]*(\d+\.?\d*)\s*mm?[】]*\s*;\s*长度[【】\s]*(\d+\.?\d*)\s*mm?[】]*\s*'
    r'外径[【】\s]*([^】]*)[】]*\s*;\s*[【】\s]*(\d+\.?\d*)\s*mm?'
)

# 模式6: 台湾纸）18.5x7.5 cm 外径单价;长度25.5 cm
RE_M6 = re.compile(
    r'[（(]?([^）)]*)[）)]?\s*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*cm\s*.*?'
    r'长度\s*(\d+\.?\d*)\s*cm'
)

# 模式7: 15x14【cm 长宽 100个】;10 【cm 高】扣底盒
RE_M7 = re.compile(
    r'(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*【cm\s*长宽.*?;\s*(\d+\.?\d*)\s*【cm\s*高】\s*扣底盒'
)

# 模式8: 【 白色 】 内尺寸;长宽cm【10*10】;【10CM】
RE_M8 = re.compile(
    r'【\s*(.*?)\s*】\s*内尺寸\s*;\s*长宽cm[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)[】]*\s*;\s*[【】]*(\d+\.?\d*)\s*CM?'
)

# 模式9: 新料钜惠【特硬内径】-黄色纸板;长宽cm【10*10】;【10CM】
RE_M9 = re.compile(
    r'【(.*?)】.*?长宽cm[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)[】]*\s*;\s*[【】]*(\d+\.?\d*)\s*CM?'
)

# 模式10: 进口优质特硬E瓦-内径;长x宽【100x100】mm;100mm【高】
RE_M10 = re.compile(
    r'进口优质.*?内径\s*;\s*长x宽[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*mm?[】]*\s*;\s*(\d+\.?\d*)\s*mm?'
)

# 模式11: 进口优质特硬E瓦-内径;长x宽【100x100】mm;更硬牛皮-其他链接
RE_M11_CUSTOM = re.compile(r'进口优质.*?内径.*?更硬牛皮|其他链接')

# 模式12/13: 白色/黄色...;6x2 cm 外径【100个】;10 cm
RE_M12 = re.compile(
    r'(白色|黄色.*?台湾纸|黄色.*?特硬|)\s*;?\s*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*cm\s*外径.*?;\s*(\d+\.?\d*)\s*cm'
)
# 材料分析用
RE_BAISE = re.compile(r'双面白色|双面白|白色|双白')
RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_TEYING = re.compile(r'特硬')
RE_KOUDI = re.compile(r'扣底盒')
RE_CUSTOM = re.compile(r'更硬牛皮|其他链接')

def guess_mat(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_BAISE.search(s): return '白色'
    if RE_TEYING.search(s): return '特硬'
    return '特硬'

# ============================================================
# 读平卡
# ============================================================
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"平卡总条数: {len(rows)}", flush=True)

matched_new = []
custom_new = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    if '品牌店' not in shop:
        remaining.append(r)
        continue
    
    handled = False
    l = w = h = None
    dk = '外径'
    mat = '特硬'
    
    # 模式11: 定制链接（更硬牛皮等）
    if '更硬牛皮' in spec_name or '其他链接' in spec_name:
        custom_new.append((fn(shop), pid, spec_id, '定制链接'))
        stats['定制'] += 1
        handled = True
    
    # 模式7: 扣底盒
    if not handled:
        m = RE_M7.search(spec_name)
        if m:
            x = float(m.group(1))
            y = float(m.group(2))
            h = float(m.group(3))
            l, w, h = max(x,y,h), sorted([x,y,h])[1], min(x,y,h)
            dk = '外径'
            mat = 'P6D'
            # 白色扣底盒无匹配
            if '白色' in spec_name or '双面白' in spec_name:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(l)}*{dfmt(w)}*{dfmt(h)}-外径-白色-扣底盒"))
                stats['扣底盒-白色无匹配'] += 1
                handled = True
            else:
                code = match_3d(l, w, h, dk, mat)
                if code:
                    matched_new.append((fn(shop), pid, spec_id, code))
                    stats['扣底盒-P6D匹配'] += 1
                else:
                    remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(l)}*{dfmt(w)}*{dfmt(h)}-{dk}-{mat}"))
                    stats['扣底盒-无匹配'] += 1
                handled = True
    
    # 模式2: （宽）135 mm 外径;长度 135 mm;105 mm 台湾纸
    if not handled:
        m = RE_M2.search(spec_name)
        if m:
            w_mm = float(m.group(1))
            l_mm = float(m.group(2))
            h_mm = float(m.group(3))
            rest = m.group(4) or ''
            l, w, h = l_mm/10, w_mm/10, h_mm/10
            dk = '外径'
            mat = guess_mat(spec_name + rest)
            dims = sorted([l, w, h], reverse=True)
            # 外径→内径转换逻辑（你说13.5→12）
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M2-匹配'] += 1
            else:
                # 如果外径对不上，尝试内径（-1.5,-0.5,-0.5）
                il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                if il > 0 and iw > 0 and ih > 0:
                    code_i = match_3d(il, iw, ih, '内径', mat)
                    if code_i:
                        matched_new.append((fn(shop), pid, spec_id, code_i))
                        stats['M2-内径转换匹配'] += 1
                    else:
                        remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                        stats['M2-无匹配'] += 1
                else:
                    remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                    stats['M2-无匹配'] += 1
            handled = True
    
    # 模式6: 台湾纸）18.5x7.5 cm 外径单价;长度25.5 cm
    if not handled and ('台湾' in spec_name or '超硬' in spec_name):
        m = RE_M6.search(spec_name)
        if m:
            raw_mat = m.group(1) or ''
            x_cm = float(m.group(2))
            y_cm = float(m.group(3))
            l_cm = float(m.group(4))
            l, w, h = l_cm, max(x_cm, y_cm), min(x_cm, y_cm)
            dk = '外径'
            mat = guess_mat(spec_name)
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M6-匹配'] += 1
            else:
                # 尝试内径
                il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                if il > 0 and iw > 0 and ih > 0:
                    code_i = match_3d(il, iw, ih, '内径', mat)
                    if code_i:
                        matched_new.append((fn(shop), pid, spec_id, code_i))
                        stats['M6-内径转换匹配'] += 1
                    else:
                        remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                        stats['M6-无匹配'] += 1
                else:
                    remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                    stats['M6-无匹配'] += 1
            handled = True
    
    # 模式1/5: 宽度【Xmm】...;长度【Y mm】...;【Z mm】（末尾第三个数字）
    if not handled:
        m1 = RE_M1.search(spec_name)
        m5 = RE_M5.search(spec_name)
        m = m1 or m5
        if m:
            w_mm = float(m.group(1))
            ctx_between = m.group(2) if m1 else m.group(3) if m5 else ''
            if m1:
                l_mm = float(m.group(3))
                h_mm = float(m.group(4))
            else:
                l_mm = float(m.group(2))
                ctx_between = m.group(3) or ''
                h_mm = float(m.group(4))
            
            l, w, h = l_mm/10, w_mm/10, h_mm/10
            dk = '外径'
            mat = guess_mat(spec_name)
            dims = sorted([l, w, h], reverse=True)
            
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M1-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M1-无匹配'] += 1
            handled = True
    
    # 模式3: （台湾纸）24x2 cm 单个;外径【单个价】长度 30 cm
    if not handled:
        m = RE_M3.search(spec_name)
        if m:
            raw_mat = m.group(1) or ''
            x_cm = float(m.group(2))
            y_cm = float(m.group(3))
            l_cm = float(m.group(4))
            l, w, h = l_cm, max(x_cm, y_cm), min(x_cm, y_cm)
            dk = '外径'
            mat = guess_mat(spec_name)
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M3-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M3-无匹配'] += 1
            handled = True
    
    # 模式4: 宽度【13*10 】cm【100个】;外径【双面白】长度 38 cm
    if not handled:
        m = RE_M4.search(spec_name)
        if m:
            w_cm = float(m.group(1))
            x_cm = float(m.group(2))
            mat_ctx = m.group(3) or ''
            l_cm = float(m.group(4))
            l, w, h = l_cm, max(w_cm, x_cm), min(w_cm, x_cm)
            dk = '外径'
            mat = guess_mat(spec_name)
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M4-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M4-无匹配'] += 1
            handled = True
    
    # 模式12/13: 白色/黄色...;6x2 cm 外径【100个】;10 cm
    if not handled:
        m = RE_M12.search(spec_name)
        if m:
            prefix = m.group(1) or ''
            x_cm = float(m.group(2))
            y_cm = float(m.group(3))
            l_cm = float(m.group(4))
            l, w, h = l_cm, max(x_cm, y_cm), min(x_cm, y_cm)
            dk = '外径'
            mat = guess_mat(spec_name)
            if '台湾' in spec_name and '超硬' in spec_name:
                mat = '超硬'
            elif '白色' in spec_name or '双面白' in spec_name:
                mat = '白色'
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M12-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M12-无匹配'] += 1
            handled = True
    
    # 模式8/9/10: 【白色】内尺寸/新料钜惠【特硬内径】/进口优质特硬E瓦-内径
    if not handled:
        m8 = RE_M8.search(spec_name)
        m9 = RE_M9.search(spec_name)
        m10 = RE_M10.search(spec_name)
        m = m8 or m9 or m10
        if m:
            if m10:
                x_mm = float(m10.group(1))
                y_mm = float(m10.group(2))
                h_mm = float(m10.group(3))
                ctx = spec_name
            elif m9:
                ctx = m9.group(1) or ''
                x_cm = float(m9.group(2))
                y_cm = float(m9.group(3))
                h_cm = float(m9.group(4))
                x_mm, y_mm, h_mm = x_cm*10, y_cm*10, h_cm*10
            else:
                ctx = m8.group(1) or ''
                x_cm = float(m8.group(2))
                y_cm = float(m8.group(3))
                h_cm = float(m8.group(4))
                x_mm, y_mm, h_mm = x_cm*10, y_cm*10, h_cm*10
            
            l, w, h = x_mm/10, y_mm/10, h_mm/10
            dk = '内径'
            mat = guess_mat(spec_name)
            dims = sorted([l, w, h], reverse=True)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M8-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M8-无匹配'] += 1
            handled = True
    
    # 模式14: 【双面白色】内径;【33x33】;6 cm高度（单个价）
    if not handled:
        m14 = re.search(r'【([^】]*)】\s*内径\s*;\s*[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)[】]*\s*;\s*(\d+\.?\d*)\s*cm?\s*高度', spec_name)
        if m14:
            ctx = m14.group(1) or ''
            x_cm = float(m14.group(2))
            y_cm = float(m14.group(3))
            h_cm = float(m14.group(4))
            l, w, h = max(x_cm, y_cm), min(x_cm, y_cm), h_cm
            dims = sorted([l, w, h], reverse=True)
            dk = '内径'
            mat = guess_mat(spec_name)
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['M14-匹配'] += 1
            else:
                remaining.append((shop, pid, spec_id, spec_name, '无匹配', f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                stats['M14-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f"\n统计:")
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}")
print(f"\n匹配成功: {len(matched_new)}")
print(f"定制链接: {len(custom_new)}")
print(f"仍留平卡: {len(remaining)}")

# 保存
if matched_new:
    f = os.path.join(out, '换绑文件_第10批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f)
    print(f"✅ 换绑_第10批: {f} ({os.path.getsize(f)/1024:.1f}KB, {len(matched_new)}条)")
    wb.close()

if custom_new:
    f = os.path.join(out, '定制类_换绑文件_第6批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for c in custom_new:
        ws.append(list(c))
    wb.save(f)
    print(f"✅ 定制类_第6批: {f} ({os.path.getsize(f)/1024:.1f}KB, {len(custom_new)}条)")
    wb.close()

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: {len(remaining)}条")
wb.close()

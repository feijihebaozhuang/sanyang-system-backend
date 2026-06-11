# -*- coding: utf-8 -*-
"""全面检查所有数据是否完整"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 1. 原始平台商品 - 读sheet名和前几行看看结构
raw_file = r'd:\Desktop\平台商品.xlsx'
wb = oxl.load_workbook(raw_file, data_only=True)
print(f'Sheet名: {wb.sheetnames}')
ws = wb[wb.sheetnames[0]]
# 看前3行
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
    print(f'  行{i+1}: {[str(c)[:20] if c else "" for c in row[:8]]}')
wb.close()

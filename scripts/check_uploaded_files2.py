# -*- coding: utf-8 -*-
"""检查已上传的换绑文件是否重复"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
files = os.listdir(out)
print('目录中所有文件:')
for f in sorted(files):
    print(f'  {repr(f)}')

# 找所有xlsx
xlsx = [f for f in files if f.endswith('.xlsx')]
print(f'\nxlsx文件({len(xlsx)}个):')
for f in sorted(xlsx):
    fp = os.path.join(out, f)
    sz = os.path.getsize(fp)
    print(f'  {f} ({sz/1024:.1f}KB)')

# 检查每个换绑文件的spec_id
all_specs = Counter()
file_specs = {}
uploaded = []

for f in xlsx:
    if '平卡' in f or '无匹配' in f or '未匹配' in f or 'OK' in f:
        continue
    fp = os.path.join(out, f)
    try:
        wb = oxl.load_workbook(fp)
        specs = []
        for sn in wb.sheetnames:
            for r in wb[sn].iter_rows(min_row=3, values_only=True):
                if r and len(r) >= 3 and r[2]:
                    sid = str(r[2]).strip()
                    specs.append(sid)
                    all_specs[sid] += 1
        wb.close()
        file_specs[f] = specs
        uploaded.append(f)
        print(f'\n  {f}: {len(specs)}条')
    except Exception as e:
        print(f'\n  {f}: 错误 {e}')

print(f'\n=== 重复检查 ===')
total_unique = set()
for f in uploaded:
    specs = file_specs[f]
    before = len(total_unique)
    total_unique.update(specs)
    new = len(total_unique) - before
    dup = len(specs) - new
    print(f'  {f}: {len(specs)}条, 新增唯一={new}, 重复={dup}')

multi = {k: v for k, v in all_specs.items() if v > 1}
print(f'\n跨文件重复spec_id: {len(multi)}个')
total = sum(len(v) for v in file_specs.values())
print(f'总条数(含重复): {total}')
print(f'去重后: {len(total_unique)}')

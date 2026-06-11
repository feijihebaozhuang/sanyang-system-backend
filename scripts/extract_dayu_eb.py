# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

src = r'D:\Desktop\换绑_深圳市大鱼包装材料有限公司.xlsx'
out = r'D:\Desktop\大鱼纸箱EB_208条.xlsx'

wb = oxl.load_workbook(src)
ws = wb.active

wb_out = oxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'
ws_out.append([None, '商品对应表', None, None])
ws_out.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])

cnt = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    code = str(row[3] or '')
    if '-EB' in code:
        ws_out.append(list(row))
        cnt += 1

ws_out.column_dimensions['A'].width = 30
ws_out.column_dimensions['B'].width = 18
ws_out.column_dimensions['C'].width = 18
ws_out.column_dimensions['D'].width = 25
wb_out.save(out)
wb_out.close()
wb.close()
print(f'✅ 已提取 {cnt} 条 EB 纸箱 → {out}')

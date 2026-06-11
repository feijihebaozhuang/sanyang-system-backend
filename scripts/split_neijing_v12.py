# -*- coding: utf-8 -*-
"""
v12: 简单粗暴法 v2
1. 提取长宽高三个字后面的数值
2. 内径/外径关键词判断
3. 材料判断（只认指定材料，其他算定制）
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def get_dim(s):
    """
    从规格中提取长宽高数值
    找"长""宽""高"三个字后面的数字
    返回 (L, W, H) 或 None
    """
    l = w = h = None
    
    # 找"长"后面的数值（考虑 "长"、"长度"、"长:"、"长【" 等各种写法）
    m = re.search(r'(?:长[度]?\s*[：:\-\u2500-\u27ff=【\[]?\s*)([\d.]+)', s)
    if m:
        v = m.group(1)
        # 清理多点号 58..7 → 58.7
        parts = v.split('.')
        if len(parts) > 2:
            v = parts[0] + '.' + ''.join(parts[1:])
        try: l = float(v)
        except: pass
    
    # 找"宽"后面的数值
    m = re.search(r'(?:宽[度]?\s*[：:\-\u2500-\u27ff=【\[]?\s*)([\d.]+)', s)
    if m:
        v = m.group(1)
        parts = v.split('.')
        if len(parts) > 2:
            v = parts[0] + '.' + ''.join(parts[1:])
        try: w = float(v)
        except: pass
    
    # 找"高"后面的数值
    m = re.search(r'(?:高[度]?\s*[：:\-\u2500-\u27ff=【\[]?\s*)([\d.]+)', s)
    if m:
        v = m.group(1)
        parts = v.split('.')
        if len(parts) > 2:
            v = parts[0] + '.' + ''.join(parts[1:])
        try: h = float(v)
        except: pass
    
    if l and w and h:
        return (l, w, h)
    return None

def normalize_unit(l, w, h, s):
    """根据规格中的单位转换数值，统一为cm"""
    # 如果数值 > 50，很可能是mm，转cm
    if 'mm' in s:
        l, w, h = l/10, w/10, h/10
    return (l, w, h)

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量'
    else: return '全量'  # 整数或全.5都算全量

def is_inner(s):
    """判断是否内径"""
    return '内径' in s or '内尺寸' in s or '内寸' in s or '所量即所装=内径' in s

def is_outer(s):
    """判断是否外径"""
    return '外径' in s or '外尺寸' in s

def is_dikoudi(s):
    return '扣底盒' in s or '双插盒' in s or '扣底纸盒' in s

def is_zhixiang(s):
    return '纸箱' in s

def get_material_type(s):
    """
    判断材料类型
    返回: '白', '台湾超硬', '优质', '黑', '红', 或其他
    其他材料 → 返回None（视为需定制）
    """
    if '特硬' in s and '台湾' in s:
        return '台湾超硬'
    if '特硬' in s:
        return '特硬'
    if '双面白' in s or ('白' in s and ('色' in s or '特硬' in s)):
        return '白'
    if '白' in s and ('色' in s or not any(c in s for c in '黑红黄蓝绿')):
        # 白色材料
        if '牛皮' in s: return None  # 牛皮归其他
        return '白'
    if '台湾超硬' in s:
        return '台湾超硬'
    if '特价' in s or '特惠' in s or '特好' in s or '优质' in s:
        return '优质'
    if '黑' in s and ('色' in s or '双面' in s):
        return '黑'
    if '红' in s and ('色' in s or '双面' in s):
        return '红'
    # 默认识别的材料
    if '特硬' in s: return '特硬'
    if '优质' in s: return '优质'
    
    return None  # 不识别

# 全量分类
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
waijing_rows = []
feiji_rows = []

for idx, row in enumerate(rows_list):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 1. 没"长""宽""高"三字 → 定制
    if '长' not in s or '宽' not in s or '高' not in s:
        custom_rows.append(row)
        continue
    
    # 2. 提取数值
    lwh = get_dim(s)
    if lwh is None:
        custom_rows.append(row)
        continue
    
    l, w, h = lwh
    l, w, h = normalize_unit(l, w, h, s)
    
    # 数值不合理 → 定制
    if not (0.5 <= l <= 150 and 0.5 <= w <= 150 and 0.5 <= h <= 150):
        custom_rows.append(row)
        continue
    
    # 3. 扣底盒
    if is_dikoudi(s):
        dikoudi_rows.append(row)
        continue
    
    # 4. 纸箱
    if is_zhixiang(s):
        zhixiang_rows.append(row)
        continue
    
    # 5. 材料判断（只认指定材料）
    material = get_material_type(s)
    if material is None:
        custom_rows.append(row)
        continue
    
    dt = classify_dims((l, w, h))
    
    # 6. 内外判断
    if is_inner(s):
        neijing_rows.append(row)
        continue
    
    # 7. 非全量
    if dt == '非全量':
        feiji_rows.append(row)
        continue
    
    # 8. 外径（包括明确外径和默认外径）
    waijing_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)
print(f'和: {s} (应={total})')
if s != total:
    print(f'!!! 差值: {total - s}')

def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(waijing_rows, r'D:\Desktop\外径全量飞机盒.xlsx', '外径全量飞机盒')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')

print('\n✅ 全部完成！')

# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

ok_dir = r'd:\Desktop\换绑输出\OK文件'

# 看第1批的全称是什么（上传成功的）
f1 = os.path.join(ok_dir, '换绑文件_第1批.xlsx')
wb = oxl.load_workbook(f1)
ws = wb[wb.sheetnames[0]]
shops1 = set()
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        shops1.add(str(r[0]).strip())
print('=== 第1批使用的店铺全称（上传成功）===')
for s in sorted(shops1):
    print('  %s' % s)

# 看第5批的简称是什么
f5 = os.path.join(ok_dir, '换绑文件_第5批.xlsx')
wb2 = oxl.load_workbook(f5)
ws2 = wb2[wb2.sheetnames[0]]
shops5 = set()
for r in ws2.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        shops5.add(str(r[0]).strip())
print('\n=== 第5批使用的名称 ===')
for s in sorted(shops5):
    print('  %s' % s)

# 看第14批（大鱼）
f14 = os.path.join(ok_dir, '换绑文件_第14批.xlsx')
wb3 = oxl.load_workbook(f14)
ws3 = wb3[wb3.sheetnames[0]]
shops14 = set()
for r in ws3.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        shops14.add(str(r[0]).strip())
print('\n=== 第14批使用的名称 ===')
for s in sorted(shops14):
    print('  %s' % s)

# 看看第5~13批和16~17的简称
for i in range(5, 18):
    fn = f'换绑文件_第{i}批.xlsx'
    fp = os.path.join(ok_dir, fn)
    if not os.path.exists(fp):
        # 3-1 3-2 skips
        continue
    wb4 = oxl.load_workbook(fp)
    ws4 = wb4[wb4.sheetnames[0]]
    shops = set()
    for r in ws4.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            shops.add(str(r[0]).strip())
    print('\n=== 第%d批 ===' % i)
    for s in sorted(shops):
        print('  %s' % s)

print('\n=== 定制类 ==')
for fn in os.listdir(ok_dir):
    if not fn.startswith('定制类'): continue
    fp = os.path.join(ok_dir, fn)
    wb5 = oxl.load_workbook(fp)
    ws5 = wb5[wb5.sheetnames[0]]
    shops = set()
    for r in ws5.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            shops.add(str(r[0]).strip())
    print('\n  %s:' % fn, sorted(shops))

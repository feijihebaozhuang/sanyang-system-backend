# -*- coding: utf-8 -*-
"""四批桌面文件：供应商名称 ↔ 商品备注 对调（刀模↔库位）。"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

COL_SUPPLIER = 33
COL_REMARK = 105
DATA_START = 4

FILES = [
    Path(r"d:\Desktop\已填充_规格别名_1_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_1_第2批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_2_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\已填充_规格别名_2_第2批_已填重量.xlsx"),
]


def swap_file(path: Path) -> int:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    n = 0
    for row in range(DATA_START, ws.max_row + 1):
        a = ws.cell(row, COL_SUPPLIER).value
        b = ws.cell(row, COL_REMARK).value
        if a is None and b is None:
            continue
        ws.cell(row, COL_SUPPLIER).value = b
        ws.cell(row, COL_REMARK).value = a
        n += 1
    wb.save(path)
    wb.close()
    return n


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    for p in FILES:
        if not p.is_file():
            print(f"跳过缺失: {p.name}")
            continue
        n = swap_file(p)
        print(f"完成: {p.name} — 对调 {n} 行")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

# -*- coding: utf-8 -*-
"""快速采样定制链接规格"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

cats = {}
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    if '外径' in s:
        cat = '外径'
    elif '定制' in s or '珍珠棉' in s or '咨询客服' in s:
        cat = '定制/珍珠棉/咨询'
    elif '飞机盒' in s:
        cat = '飞机盒'
    elif '扣底盒' in s or '双插盒' in s:
        cat = '扣底盒残'
    elif '纸箱' in s:
        cat = '纸箱残'
    elif 'mm' in s:
        cat = '含mm'
    elif 'cm' in s or '厘米' in s:
        cat = '含cm'
    elif re.search(r'\d+\s*[xX*]\s*\d+', s):
        cat = '有乘式'
    elif re.search(r'\d', s):
        cat = '有数字'
    else:
        cat = '纯文本'
    cats[cat] = cats.get(cat, 0) + 1
    if total >= 500:
        break

wb.close()
print(f'定制链接采样{total}条分布:')
for k, v in sorted(cats.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')

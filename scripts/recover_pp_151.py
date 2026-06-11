# -*- coding: utf-8 -*-
"""从原始平台商品中找回品牌店151条丢失的无匹配数据"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd

out = r'd:\Desktop\换绑输出'

# 原始平台商品
raw_file = r'd:\Desktop\平台商品.xlsx'

# 先看有哪些sheet
wb = oxl.load_workbook(raw_file, data_only=True)
print(f'Sheets: {wb.sheetnames}')
wb.close()

# 读数据
df = pd.read_excel(raw_file, sheet_name=0, dtype=str)
print(f'\n总行数: {len(df)}')
print(f'列名: {list(df.columns)}')

# 找品牌店相关的行
# 看有没有店铺名列
shop_col = None
for c in df.columns:
    if '店铺' in str(c) or '店' in str(c):
        shop_col = c
        print(f'店铺列: {c}')
        break

if shop_col:
    pp_mask = df[shop_col].astype(str).str.contains('品牌店', na=False)
    print(f'\n品牌店行数: {pp_mask.sum()}')
    
    # 找规格名称列
    spec_col = None
    for c in df.columns:
        if '规格名称' in str(c) or '规格' in str(c) or 'spec' in str(c).lower():
            spec_col = c
            print(f'规格名列: {c}')
            break
    
    if spec_col:
        pp_df = df[pp_mask].copy()
        # 找宽度【X*Y】长度 Z cm 格式
        RE_PAT = re.compile(r'宽度[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*[】]*\s*cm?.*?长度\s*(\d+\.?\d*)\s*cm')
        wm_mask = pp_df[spec_col].astype(str).str.contains(r'宽度.*\*.*长度', na=False, regex=True)
        print(f'宽度*长度格式: {wm_mask.sum()}')
        
        # 读所有换绑文件已匹配的spec_id
        matched_specs = set()
        for f in os.listdir(out):
            if '换绑' in f and f.endswith('.xlsx'):
                try:
                    wb2 = oxl.load_workbook(os.path.join(out, f))
                    for sn in wb2.sheetnames:
                        ws = wb2[sn]
                        for r in ws.iter_rows(min_row=3, values_only=True):
                            if r and len(r) >= 3:
                                matched_specs.add(str(r[2] or '').strip())
                    wb2.close()
                except: pass
        
        # 读无匹配的spec_id
        wb2 = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
        ws = wb2['无匹配']
        wm_specs = set()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and len(r) >= 2:
                wm_specs.add(str(r[2] or '').strip())
        wb2.close()
        
        # 读定制类的spec_id
        for f in os.listdir(out):
            if '定制' in f and f.endswith('.xlsx'):
                try:
                    wb2 = oxl.load_workbook(os.path.join(out, f))
                    for sn in wb2.sheetnames:
                        ws = wb2[sn]
                        for r in ws.iter_rows(min_row=3, values_only=True):
                            if r and len(r) >= 3:
                                matched_specs.add(str(r[2] or '').strip())
                    wb2.close()
                except: pass
        
        print(f'\n已匹配spec_id数: {len(matched_specs)}')
        print(f'无匹配spec_id数: {len(wm_specs)}')
        
        # 找出品牌店中既不在已匹配也不在无匹配、且符合宽度*长度格式的
        found = []
        for _, row in pp_df[wm_mask].iterrows():
            spec_name = str(row.get(spec_col, '') or '')
            spec_id = str(row.get('平台规格id', row.get('规格id', row.get('spec_id', ''))) or '')
            pid = str(row.get('平台商品id', row.get('商品id', '')) or '')
            
            # 只取不在任何处理过的集合里的
            if spec_id in matched_specs or spec_id in wm_specs:
                continue
            
            m = RE_PAT.search(spec_name)
            if m:
                w = float(m.group(1))
                h = float(m.group(2))
                l = float(m.group(3))
                found.append((spec_name, spec_id, pid, w, h, l))
        
        print(f'\n漏掉的品牌店(宽度*长度格式): {len(found)}条')
        for s in found[:5]:
            print(f'  {s[0][:80]} spec_id={s[1][:20]}')
        
        if found:
            # 追加到无匹配
            wb2 = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
            ws = wb2['无匹配']
            shop_name = '飞机盒品牌店'
            for spec_name, spec_id, pid, w, h, l in found:
                ws.append([shop_name, pid, spec_id, spec_name, '无匹配', 
                    f'{int(l)}*{int(w)}*{int(h)}-外径-特硬'])
            wb2.save(os.path.join(out, '无匹配_待处理.xlsx'))
            wb2.close()
            print(f'✅ 已追加{len(found)}条到无匹配_待处理')

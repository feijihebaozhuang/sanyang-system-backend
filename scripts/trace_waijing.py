# -*- coding: utf-8 -*-
"""追踪外径规格的去向"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

# 直接用v6的逻辑
from split_neijing_v6 import extract_all_dims, best_lwh, classify_dims, has_any_dim

cats = {'定制':0,'扣底盒/双插盒':0,'纸箱':0,'内径':0,'外径全量飞机盒':0,'非全量飞机盒':0,'其余':0,'未知':0}
counter = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    counter += 1
    if counter > 500:
        break

    # 模拟v6分类
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
        cats['定制'] += 1
    elif '扣底盒' in s or '双插盒' in s:
        cats['扣底盒/双插盒'] += 1
    elif '纸箱' in s:
        cats['纸箱'] += 1
    else:
        lwh = best_lwh(s)
        if lwh is None:
            cats['定制'] += 1
        else:
            dt = classify_dims(lwh)
            if '内径' in s or '内尺寸' in s or '内寸' in s:
                cats['内径'] += 1
            elif dt == '非全量飞机盒':
                cats['非全量飞机盒'] += 1
            elif '外径' in s:
                # 外径 + 整数/全.5 → 外径全量飞机盒
                cats['外径全量飞机盒'] += 1
            else:
                cats['其余'] += 1

wb.close()
print(f'外径规格采样{counter}条分布:')
for k, v in sorted(cats.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')

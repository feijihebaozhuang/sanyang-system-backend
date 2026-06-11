# -*- coding: utf-8 -*-
"""检查v7输出中定制链接里是否还有外径规格"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

has_waijing = 0
has_feiji = 0
has_dim = 0
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' in s: has_waijing += 1
    if '飞机盒' in s: has_feiji += 1
    # 是否有数字x数字x数字
    if re.search(r'\d[\d.]*\s*[xX*]\s*\d[\d.]*\s*[xX*]\s*\d[\d.]*', s): has_dim += 1
wb.close()

print(f'定制链接总数: {total}')
print(f'包含"外径": {has_waijing}')
print(f'包含"飞机盒": {has_feiji}')
print(f'包含LxWxH: {has_dim}')

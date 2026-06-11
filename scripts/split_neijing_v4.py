# -*- coding: utf-8 -*-
"""
重新从平台商品.xlsx分类，修复extract_lwh以覆盖宽|高|长独立格式
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

source = r'D:\Desktop\平台商品.xlsx'
print(f'读取 {source} ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def extract_lwh(s):
    """统一提取长宽高。返回 (l,w,h) tuple 或 None"""
    
    # === 格式1: 独立维度 宽【W】高【H】... 长【L】 ===
    # 宽【10cm】高【8cm】内径;【100个】长【38cm】
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m:
        return (float(m.group(3)), float(m.group(1)), float(m.group(2)))
    
    # === 格式2: 长【L】宽【W】高【H】（如有此格式）===
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m:
        return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # === 格式3: 【LxWxH】=== 
    # 【21.5*14.5*4.5】 或 【 21.5 x 14.5 x 4.5 】
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m:
        return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # === 格式4: 长x宽x高（带关键词前缀）===
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # === 格式5: ;LxWxH 在末尾 ===
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m:
        return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    # === 格式6: LxWxH（裸露在字符串中）===
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m:
        return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
    
    return None

def classify_dims(lwh):
    """按尺寸类型分类"""
    l, w, h = lwh
    # 检查是否有小数（任何一个不是整数）
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5:
        return '非全量飞机盒'
    elif all_5:
        return '全.5'
    else:
        return '整数'

def is_custom(s):
    """定制链接: 关键词 + 无有效长宽高"""
    s_lower = s.lower()
    if '定制' in s or '珍珠棉' in s or '咨询客服' in s:
        lwh = extract_lwh(s)
        if lwh is None:
            return True
    return False

# === 一次全量分类 ===
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
feiji_rows = []
other_rows = []

for idx, row in enumerate(rows_list):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 1. 先判断定制链接
    if is_custom(s):
        custom_rows.append(row)
        continue
    
    # 2. 提取长宽高
    lwh = extract_lwh(s)
    if lwh is None:
        # 没有长宽高的 → 定制
        custom_rows.append(row)
        continue
    
    dim_type = classify_dims(lwh)
    
    # 3. 根据规格名称分类
    if '扣底盒' in s or '双插盒' in s:
        dikoudi_rows.append(row)
    elif '纸箱' in s:
        zhixiang_rows.append(row)
    elif dim_type == '非全量飞机盒':
        feiji_rows.append(row)
    elif '内径' in s or '内尺寸' in s or '内寸' in s:
        neijing_rows.append(row)
    else:
        other_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'其余: {len(other_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(feiji_rows)+len(neijing_rows)+len(other_rows)
print(f'和: {s} (应={total})')

# === 输出 ===
def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(other_rows, r'D:\Desktop\其余商品.xlsx', '其余商品')

print('\n✅ 全部完成！')

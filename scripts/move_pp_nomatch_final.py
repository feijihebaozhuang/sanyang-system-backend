# -*- coding: utf-8 -*-
"""品牌店151条无匹配移到无匹配_待处理"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 读待确认
wb = oxl.load_workbook(os.path.join(out, '品牌店_待确认.xlsx'), data_only=True)
ws = wb['待确认']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

if not rows:
    print('没有待确认数据')
    sys.exit(0)

print(f'待确认: {len(rows)}条，全部移到无匹配')

# 追加到无匹配_待处理
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'), data_only=True)
ws = wb['无匹配']
for r in rows:
    if len(r) >= 6:
        ws.append(list(r[:6]))
    else:
        ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
wb.close()

# 删除待确认文件
os.remove(os.path.join(out, '品牌店_待确认.xlsx'))

print(f'✅ 已追加 {len(rows)}条 到 无匹配_待处理.xlsx')
print(f'✅ 品牌店_待确认.xlsx 已删除')

# 确认平卡已无品牌店
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
pk = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
pp_count = sum(1 for r in pk if r and '品牌店' in str(r[0] or ''))
print(f'平卡品牌店剩余: {pp_count}条（应为0）')
print('\n品牌店全部处理完毕！')

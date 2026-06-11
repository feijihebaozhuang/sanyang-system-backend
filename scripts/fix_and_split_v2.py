# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or '')); w = normalize_dim(m.group(3)+(m.group(4) or '')); h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

# 列名
COLS = ['店铺名称','平台商品id','规格名称','平台规格id','尺寸类型','长','宽','高']

# ===== 1. 扣底盒文件 - 用 header=0（第一行是标题） =====
kd_file = r'D:\Desktop\扣底盒双插盒商品.xlsx'
df_kd = pd.read_excel(kd_file, names=COLS, skiprows=1, dtype=str)
print(f'扣底盒: {len(df_kd)} 行')

pid_target = '777564404927'
mask_remove = (df_kd['平台商品id'].astype(str).str.strip() == pid_target) & \
              df_kd['规格名称'].astype(str).str.strip().isin(['特硬;双插盒', '特硬;扣底盒'])
removed = df_kd[mask_remove].copy()
remaining_kd = df_kd[~mask_remove].copy()
print(f'  找到 {len(removed)} 条需移到定制的:')
for _, r in removed.iterrows():
    print(f'    {r["规格名称"]}')

# ===== 2. 定制链接文件 =====
custom_file = r'D:\Desktop\定制链接商品.xlsx'
df_custom = pd.read_excel(custom_file, header=None, dtype=str, skiprows=1)
df_custom.columns = ['店铺名称','平台商品id','规格名称','平台规格id']
print(f'\n定制链接: {len(df_custom)} 行')

if len(removed) > 0:
    add_rows = removed[['店铺名称','平台商品id','规格名称','平台规格id']].copy()
    df_custom = pd.concat([df_custom, add_rows], ignore_index=True)
print(f'  更新后: {len(df_custom)} 行')

# 保存
import openpyxl

def save_base(path, header, df, cols):
    # 读整个文件，保留已有行
    if os.path.exists(path):
        mode = 'a'
        df.to_csv(path, index=False, encoding='utf-8-sig')
    else:
        df[cols].to_excel(path, index=False)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.insert_rows(1)
        ws.cell(1, 2).value = header
        for c in ['C']: ws.column_dimensions[c].width = 60
        wb.save(path)

# 更可靠的方式：直接覆盖
out_kd = kd_file
remaining_kd.to_excel(out_kd, index=False, columns=COLS)
wb = openpyxl.load_workbook(out_kd)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '扣底盒/双插盒'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(out_kd)
print(f'✅ 扣底盒文件保存: {len(remaining_kd)} 条')

out_ct = custom_file
df_custom.to_excel(out_ct, index=False, columns=['店铺名称','平台商品id','规格名称','平台规格id'])
wb = openpyxl.load_workbook(out_ct)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '定制链接商品（无准确尺寸）'
ws.column_dimensions['C'].width = 60
wb.save(out_ct)
print(f'✅ 定制链接保存: {len(df_custom)} 条')

# ===== 3. 其余商品分拆 =====
other_file = r'D:\Desktop\其余商品.xlsx'
df_other = pd.read_excel(other_file, names=COLS, skiprows=1, dtype=str)
print(f'\n其余商品: {len(df_other)} 行')

dot_rows = []
remain_rows = []

for idx, (_, row) in enumerate(df_other.iterrows()):
    s = str(row.get('规格名称', '')).strip()
    if not s:
        remain_rows.append(row)
        continue
    lwh = extract_lwh(s)
    if not lwh:
        remain_rows.append(row)
        continue
    l, w, h = lwh
    has_dot = any(n != int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    if has_dot and not all_dot5:
        dot_rows.append(row)
    else:
        remain_rows.append(row)

print(f'  带小数点(非全.5): {len(dot_rows)} 条')
print(f'  剩余: {len(remain_rows)} 条')

# 非全量飞机盒
feiji_file = r'D:\Desktop\非全量飞机盒.xlsx'
pd.DataFrame(dot_rows, columns=COLS).to_excel(feiji_file, index=False)
wb = openpyxl.load_workbook(feiji_file)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '其余商品中带小数点的（已排除全.5）'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(feiji_file)
print(f'✅ 非全量飞机盒.xlsx: {len(dot_rows)}条')

# 更新其余商品
pd.DataFrame(remain_rows, columns=COLS).to_excel(other_file, index=False)
wb = openpyxl.load_workbook(other_file)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '其余商品（已移除小数项）'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(other_file)
print(f'✅ 其余商品已更新: {len(remain_rows)}条')

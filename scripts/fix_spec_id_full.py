# -*- coding: utf-8 -*-
"""从未匹配文件中的商品ID+简称，到原始平台商品中取完整规格ID"""
import sys, os, pandas as pd
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
pf = r'd:\Desktop\平台商品.xlsx'

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '阿里友尚',
    '阿里亚润': '阿里亚润',
    '阿里三羊': '阿里三羊',
    '阿里正方形': '阿里正方形',
    '阿里大鱼': '阿里大鱼',
    '阿里新鑫星': '阿里新鑫星',
}

# 读定制文件中的 (商品ID, 截断规格ID)
def load_truncated(fname):
    fp = os.path.join(out, fname)
    wb = oxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()
    return [(str(r[1] or '').strip(), str(r[2] or '').strip()) for r in rows]

truncated_1 = load_truncated('定制类_换绑文件.xlsx')
truncated_2 = load_truncated('定制类_换绑文件_第2批.xlsx')
print(f"定制1: {len(truncated_1)}条")
print(f"定制2: {len(truncated_2)}条")

# 读原始平台商品，建 (商品ID+截断30位) → 完整规格ID 的映射
print("读取原始平台商品...", flush=True)
df = pd.read_excel(pf, sheet_name='报表1', header=2, dtype=str)

# 建索引
pid_map = {}
total = len(df)
for i, row in enumerate(df.values):
    if len(row) < 4:
        continue
    pid = str(row[1] or '').strip()
    spec_id = str(row[3] or '').strip()
    shop = str(row[0] or '').strip()
    if pid and spec_id:
        key = (pid, spec_id[:30])
        pid_map[key] = (spec_id, shop)
    if i % 100000 == 0:
        print(f"  处理中 {i}/{total}...", flush=True)

print(f"索引完成: {len(pid_map)}个(商品ID+规格ID前30)", flush=True)

def rebuild(truncated_list, output_name):
    result = []
    missing = 0
    for pid, truncated_sid in truncated_list:
        key = (pid, truncated_sid)
        if key in pid_map:
            full_sid, shop_short = pid_map[key]
            shop_full = SHOP_NAME_MAP.get(shop_short, shop_short)
            result.append((shop_full, pid, full_sid, '定制链接'))
        else:
            # 尝试key反转
            found = False
            for (kpid, ksid), (fsid, sshort) in pid_map.items():
                if kpid == pid and ksid == truncated_sid[:30]:
                    shop_full = SHOP_NAME_MAP.get(sshort, sshort)
                    result.append((shop_full, pid, fsid, '定制链接'))
                    found = True
                    break
            if not found:
                missing += 1
    
    print(f"{output_name}: {len(result)}条, 缺失{missing}")
    
    # 检查长度
    lens = Counter()
    long_ids = []
    for m in result:
        sid = m[2]
        lens[len(sid)] += 1
        if len(sid) > 30:
            long_ids.append(sid)
    print(f"  长度分布: {dict(lens)}")
    if long_ids:
        print(f"  长度>30的样例: {long_ids[:3]}")
    
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in result:
        ws.append(list(m))
    fp = os.path.join(out, output_name)
    wb.save(fp)
    wb.close()
    print(f"  → 已保存 {fp}")

rebuild(truncated_1, '定制类_换绑文件.xlsx')
rebuild(truncated_2, '定制类_换绑文件_第2批.xlsx')

print("\n完成！先上传 定制类_换绑文件.xlsx 试试")

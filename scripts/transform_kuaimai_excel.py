# -*- coding: utf-8 -*-
"""快麦四批：库位→商品备注；刀模→供应商名称；重量→商品简称(KG)。"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl

COL_MAIN = 1
COL_ALIAS = 29
COL_SUPPLIER = 33
COL_WEIGHT = 39
COL_SHORT = 73
COL_REMARK = 105
HEADER_ROW = 3
DATA_START = 4

LOOKUP_FILE = Path(r"d:\Desktop\快麦商品\ok\已匹配库存库位_合并.xlsx")
SQ_DIMOLD_CANDIDATES = [
    Path(r"d:\Desktop\快麦商品\新建文件夹 (2)\正方形刀模_未匹配全量.xlsx"),
    Path(r"d:\Desktop\Desktop\新建文件夹\正方形刀模.xlsx"),
]
RC_DIMOLD_CANDIDATES = [
    Path(r"d:\Desktop\快麦商品\新建文件夹 (2)\刀模未匹配全量_长方形.xlsx"),
    Path(r"d:\Desktop\Desktop\新建文件夹\长方形刀模.xlsx"),
]
# 对照表变体「外+实际尺寸」补充（原未匹配全量中的特殊行）
EXTRA_SQ_ENTRIES: list[tuple[str, str, str]] = [
    ("25*25*9", "外25.5*25.5*9", "正144 398"),
    ("26*26*7", "外（26*26*7.5）", "正287"),
    ("28*28*8", "外28*28*8.5", "正110"),
    ("32*32*9", "外32*32*9.5", "正124"),
    ("38*38*6", "外38*38*5.7", "正196"),
]
INPUT_FILES = [
    Path(r"d:\Desktop\快麦商品\快麦商品_已填重量\已填充_规格别名_1_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\快麦商品\快麦商品_已填重量\已填充_规格别名_1_第2批_已填重量.xlsx"),
    Path(r"d:\Desktop\快麦商品\快麦商品_已填重量\已填充_规格别名_2_第1批_已填重量.xlsx"),
    Path(r"d:\Desktop\快麦商品\快麦商品_已填重量\已填充_规格别名_2_第2批_已填重量.xlsx"),
]
OUT_DIR = Path(r"d:\Desktop")
# 导出「未覆盖剩余」时扫描的四批全量（桌面输出）
BATCH_FILES = [
    OUT_DIR / "已填充_规格别名_1_第1批_已填重量.xlsx",
    OUT_DIR / "已填充_规格别名_1_第2批_已填重量.xlsx",
    OUT_DIR / "已填充_规格别名_2_第1批_已填重量.xlsx",
    OUT_DIR / "已填充_规格别名_2_第2批_已填重量.xlsx",
]


def _norm_size(s) -> str | None:
    if s is None:
        return None
    s = str(s).strip().replace("×", "*").replace("x", "*").replace("X", "*")
    return re.sub(r"\s+", "", s) or None


def _parse_main_dimold(main) -> tuple[str | None, str | None]:
    """主商家编码 → (尺寸, 类型) 用于刀模表：内/外/带扣。"""
    main = str(main or "").strip()
    m = re.match(r"^([\d.*]+)-(内径|外径)(?:-|$)", main)
    if m:
        return _norm_size(m.group(1)), ("内" if m.group(2) == "内径" else "外")
    if "带扣" in main:
        m2 = re.match(r"^([\d.*]+)", main)
        if m2:
            return _norm_size(m2.group(1)), "带扣"
    return None, None


def _split_code_pieces(code: str) -> list[str]:
    parts: list[str] = []
    for piece in re.split(r"[;；,\s/]+", code):
        p = piece.strip()
        if p:
            parts.append(p)
    return parts


def _join_codes(codes: list[str]) -> str:
    seen: list[str] = []
    for c in codes:
        for p in _split_code_pieces(c):
            if p not in seen:
                seen.append(p)
    return ",".join(seen)


def _outer_type_suffix(typ: str) -> str | None:
    """外25.5*25.5*9 / 外（26*26*7.5） → 附加尺寸后缀。"""
    t = typ.strip()
    if t == "外" or not t.startswith("外"):
        return None
    rest = t[1:].strip()
    rest = re.sub(r"^[（(]+|[）)]+$", "", rest)
    return _norm_size(rest) if rest else None


def _format_outer_entry(code: str, typ: str) -> str:
    parts = _split_code_pieces(code)
    suffix = _outer_type_suffix(typ)
    if suffix and suffix not in parts:
        parts.append(suffix)
    return ",".join(parts)


def _first_existing(candidates: list[Path]) -> Path | None:
    for p in candidates:
        if p.is_file():
            return p
    return None


def _load_square_entries(path: Path) -> list[tuple[str, str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str]] = []
    # 未匹配全量: 第2行起；正方形刀模.xlsx: 第4行起
    start = 4 if "刀模.xlsx" in path.name and "未匹配" not in path.name else 2
    if ws.cell(1, 1).value and "登记" in str(ws.cell(1, 1).value):
        start = 4
    for r in range(start, ws.max_row + 1):
        size = _norm_size(ws.cell(r, 1).value)
        typ = str(ws.cell(r, 2).value or "").strip()
        code = ws.cell(r, 3).value
        if code is None:
            continue
        code_s = str(code).strip()
        if size and typ and code_s:
            rows.append((size, typ, code_s))
    wb.close()
    return rows


def _load_rect_entries(path: Path) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str]] = []
    start = 3 if ws.cell(2, 1).value is None else 2
    for r in range(start, ws.max_row + 1):
        size = _norm_size(ws.cell(r, 1).value)
        code = ws.cell(r, 2).value
        if code is None:
            continue
        code_s = str(code).strip()
        if size and code_s:
            rows.append((size, code_s))
    wb.close()
    return rows


def _load_dimold_lookups() -> tuple[list[tuple[str, str, str]], dict[str, str], set[str]]:
    sq_entries: list[tuple[str, str, str]] = []
    sq_path = _first_existing(SQ_DIMOLD_CANDIDATES)
    if sq_path:
        sq_entries.extend(_load_square_entries(sq_path))
    sq_entries.extend(EXTRA_SQ_ENTRIES)

    sq_sizes = {s for s, _, _ in sq_entries}

    rc_grp: dict[str, list[str]] = defaultdict(list)
    rc_path = _first_existing(RC_DIMOLD_CANDIDATES)
    if rc_path:
        for size, code in _load_rect_entries(rc_path):
            rc_grp[size].append(code)
    rc_map = {k: _join_codes(v) for k, v in rc_grp.items()}
    return sq_entries, rc_map, sq_sizes


def _is_combo_short(s) -> bool:
    t = str(s or "").strip()
    if not t:
        return False
    if t.startswith("组合"):
        return True
    return any(ch in t for ch in "./;")


def _lookup_dimold(
    main,
    old_short,
    sq_entries: list[tuple[str, str, str]],
    rc_map: dict,
    sq_sizes: set[str],
) -> str | None:
    size, typ = _parse_main_dimold(main)
    if not size:
        return None

    if typ == "带扣":
        codes = [c for s, t, c in sq_entries if s == size and t == "带扣"]
        if codes:
            return _join_codes(codes)

    if typ == "内":
        codes = [c for s, t, c in sq_entries if s == size and t == "内"]
        if codes:
            return _join_codes(codes)

    if typ == "外":
        outer = [(t, c) for s, t, c in sq_entries if s == size and t != "内" and t != "带扣" and "外" in t]
        if outer:
            parts: list[str] = []
            for t, c in outer:
                formatted = _format_outer_entry(c, t)
                for p in formatted.split(","):
                    if p and p not in parts:
                        parts.append(p)
            return ",".join(parts)
        plain = [c for s, t, c in sq_entries if s == size and t == "外"]
        if plain:
            return _join_codes(plain)

    if size not in sq_sizes and size in rc_map and _is_combo_short(old_short):
        return rc_map[size]
    return None


def _weight_to_short(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        if s.upper().endswith("KG"):
            return s
        return f"{s}KG"
    if isinstance(val, (int, float)):
        s = format(Decimal(str(val)).normalize(), "f")
        return f"{s}KG"
    return f"{val}KG"


def _verify_headers(ws) -> None:
    assert ws.cell(HEADER_ROW, COL_SUPPLIER).value == "供应商名称", "供应商名称列不匹配"
    assert ws.cell(HEADER_ROW, COL_WEIGHT).value == "规格重量(千克)", "规格重量列不匹配"
    assert ws.cell(HEADER_ROW, COL_SHORT).value == "商品简称", "商品简称列不匹配"
    assert ws.cell(HEADER_ROW, COL_REMARK).value == "商品备注", "商品备注列不匹配"


def _load_supplier_lookup(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    by_full: dict[tuple, str] = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if not row or len(row) < COL_REMARK:
            continue
        main = row[COL_MAIN - 1]
        alias = row[COL_ALIAS - 1]
        short = row[COL_SHORT - 1]
        remark = row[COL_REMARK - 1]
        if remark is None or str(remark).strip() == "":
            continue
        by_full[(main, alias, short)] = str(remark).strip()
    wb.close()
    return by_full


def transform_file(
    src: Path, dst: Path, by_full: dict, sq_entries: list, rc_map: dict, sq_sizes: set
) -> dict:
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    _verify_headers(ws)
    stats = {
        "rows": 0,
        "supplier_hit": 0,
        "supplier_miss": 0,
        "dimold_replaced": 0,
    }
    for row in range(DATA_START, ws.max_row + 1):
        main = ws.cell(row, COL_MAIN).value
        alias = ws.cell(row, COL_ALIAS).value
        old_short = ws.cell(row, COL_SHORT).value
        weight = ws.cell(row, COL_WEIGHT).value

        loc = by_full.get((main, alias, old_short))
        if loc:
            stats["supplier_hit"] += 1
        else:
            stats["supplier_miss"] += 1

        knife = old_short
        dimold = _lookup_dimold(main, old_short, sq_entries, rc_map, sq_sizes)
        if dimold and _is_combo_short(old_short):
            knife = dimold
            stats["dimold_replaced"] += 1

        if loc is None and knife is None and weight is None:
            continue
        ws.cell(row, COL_SUPPLIER).value = knife   # 刀模编码
        ws.cell(row, COL_REMARK).value = loc       # 库位
        ws.cell(row, COL_SHORT).value = _weight_to_short(weight)
        stats["rows"] += 1
    wb.save(dst)
    wb.close()
    return stats


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    if not LOOKUP_FILE.is_file():
        print(f"缺失合并表: {LOOKUP_FILE}")
        return 1
    if not _first_existing(SQ_DIMOLD_CANDIDATES):
        print("缺失正方形刀模对照表")
        return 1
    if not _first_existing(RC_DIMOLD_CANDIDATES):
        print("缺失长方形刀模对照表")
        return 1
    by_full = _load_supplier_lookup(LOOKUP_FILE)
    sq_entries, rc_map, sq_sizes = _load_dimold_lookups()
    print(
        f"库位表: {len(by_full)} 条 | 刀模正方形 {len(sq_entries)} 行 "
        f"({ _first_existing(SQ_DIMOLD_CANDIDATES).name }) 长方形 {len(rc_map)} 键"
    )
    files = [f for f in INPUT_FILES if f.is_file()]
    missing = [f for f in INPUT_FILES if not f.is_file()]
    if missing:
        print("跳过缺失源文件:", ", ".join(f.name for f in missing))
    if not files:
        print("无可用源文件")
        return 1
    for src in files:
        dst = OUT_DIR / src.name
        stats = transform_file(src, dst, by_full, sq_entries, rc_map, sq_sizes)
        print(
            f"完成: {dst.name} — 写入 {stats['rows']} 行, "
            f"库位命中 {stats['supplier_hit']}, "
            f"刀模编码替换 {stats['dimold_replaced']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

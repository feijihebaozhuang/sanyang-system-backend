# -*- coding: utf-8 -*-
"""生成 2 行快麦上传测试表（仅复制表头+2行，不拷整本五万行）。"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook

SRC = Path(r"d:\Desktop\已填充_规格别名_1_第1批_已填重量.xlsx")
OUT_DIR = Path(r"d:\Desktop")
DATA_START = 4
SAMPLE_ROWS = 2

COL_ALIAS = 29
COL_SUPPLIER = 33
COL_WEIGHT = 39
COL_LEN = 40
COL_W = 41
COL_H = 42
COL_BRAND = 20
COL_STD = 52
COL_SHORT = 73
COL_G_WEIGHT = 100
COL_G_LEN = 101
COL_G_W = 102
COL_G_H = 103


def _copy_row_style(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        sc = src_ws.cell(src_row, c)
        dc = dst_ws.cell(dst_row, c, value=sc.value)
        if sc.has_style:
            dc._style = copy(sc._style)
        if sc.number_format:
            dc.number_format = sc.number_format


def _apply_row(ws, row: int, supplier_to: str) -> None:
    ws.cell(row, COL_SHORT).value = ws.cell(row, COL_ALIAS).value
    ws.cell(row, COL_G_WEIGHT).value = ws.cell(row, COL_WEIGHT).value
    ws.cell(row, COL_G_LEN).value = ws.cell(row, COL_LEN).value
    ws.cell(row, COL_G_W).value = ws.cell(row, COL_W).value
    ws.cell(row, COL_G_H).value = ws.cell(row, COL_H).value
    sup = ws.cell(row, COL_SUPPLIER).value
    if supplier_to == "brand":
        ws.cell(row, COL_BRAND).value = sup
    else:
        ws.cell(row, COL_STD).value = sup


def _make(out: Path, supplier_to: str, src_ws, max_col: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = src_ws.title
    for r in range(1, DATA_START + SAMPLE_ROWS):
        _copy_row_style(src_ws, ws, r, r, max_col)
    for r in range(DATA_START, DATA_START + SAMPLE_ROWS):
        _apply_row(ws, r, supplier_to)
    wb.save(out)
    wb.close()


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    if not SRC.is_file():
        print(f"缺失: {SRC}")
        return 1
    src_wb = openpyxl.load_workbook(SRC, read_only=False, data_only=True)
    src_ws = src_wb.active
    max_col = src_ws.max_column
    a = OUT_DIR / "快麦上传测试_2条_供应商→品牌.xlsx"
    b = OUT_DIR / "快麦上传测试_2条_供应商→执行标准.xlsx"
    _make(a, "brand", src_ws, max_col)
    _make(b, "standard", src_ws, max_col)
    src_wb.close()
    print(f"已生成:\n  {a}\n  {b}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

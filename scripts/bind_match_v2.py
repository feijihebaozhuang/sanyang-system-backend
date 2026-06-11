# -*- coding: utf-8 -*-
"""
平台商品 ↔ 快麦商品 换绑匹配脚本 v2
暴力匹配策略：不管格式，直接从文本中提取所有关键信息
"""
import sys, re, os, json
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 店铺名称 简称→全称
# ============================================================
SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂​',
    '淘宝品牌店': '飞机盒品牌店​',
    '阿里友尚': '深圳市友尚包装有限公司​',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

# ============================================================
# 已知材料映射（从快麦所有主商家编码中总结）
# ============================================================
MATERIAL_KEYWORDS = [
    # (关键词, 标准材料名)
    ('特硬', '特硬'),
    ('超硬', '超硬'),
    ('台湾纸', '超硬'),
    ('进口纸', '超硬'),
    ('进口', '超硬'),
    ('台湾', '超硬'),
    ('双白', '白色'),
    ('双面白', '白色'),
    ('白色', '白色'),
    ('黑色', '黑色'),
    ('红色', '红色'),
    ('蓝色', '蓝色'),
    ('绿色', '绿色'),
    ('黄色', '黄色'),
    ('灰色', '灰色'),
    ('紫色', '紫色'),
    ('牛皮', '牛皮'),
    ('优质', '优质'),
    ('pe', 'PE'),
    ('珍珠棉', '珍珠棉'),
    ('P6D', 'P6D'),
    ('P6', 'P6D'),
    ('B9', 'B9'),
    ('B6', 'B6'),
    ('B3', 'B3'),
    ('E坑', 'E坑'),
    ('B坑', 'B坑'),
]

def extract_material(text):
    """从文本中提取材料。只识别已知材料词。"""
    t = str(text).lower()
    for kw, std in MATERIAL_KEYWORDS:
        if kw.lower() in t:
            return std
    return None

# ============================================================
# 1. 读取快麦商品表
# ============================================================
def load_km_products(km_file):
    import openpyxl
    wb = openpyxl.load_workbook(km_file, read_only=True, data_only=True)
    ws = wb['报表1']
    
    km_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        code, name, category = row[0], row[1], row[2]
        if code and str(code).strip():
            code = str(code).strip()
            km_map[code] = {
                'name': str(name or '').strip(),
                'category': str(category or '').strip(),
            }
    wb.close()
    return km_map


# ============================================================
# 2. 暴力解析规格名称 → 提取尺寸/内外径/材料
# ============================================================
def parse_spec_brutal(text):
    """
    不管什么格式，从文本中暴力提取：
    1. 三个数字 → 长宽高
    2. 内外径关键词
    3. 材料关键词
    
    返回 dict 或 None
    """
    if not text:
        return None
    
    s = str(text).strip()
    if not s:
        return None
    
    result = {'custom': False}
    
    # --- 检查是否定制 ---
    if re.search(r'[定制定做定制定造]', s):
        result['custom'] = True
        return result
    
    # --- 提取三个尺寸数字 ---
    # 把所有数字找出来
    all_nums = [float(m.group(1)) for m in re.finditer(r'(\d+\.?\d*)', s) 
                if float(m.group(1)) > 0]
    
    # 分类：找最大的三个作为长宽高
    # 先尝试从文本中识别"长""宽""高"对应的数字
    dims = {}
    
    # 方法1: 找长/宽/高/长度/宽度/高度/厚度 后面的数字
    for label, keys in [('长', r'长(?:度)?'), ('宽', r'宽(?:度)?'), ('高', r'高(?:度)?|厚(?:度)?')]:
        m = re.search(f'{keys}.{{0,5}}?[：:，,;；\\s]*?【?(\\d+\\.?\\d*)\\s*cm?[CM]?】?', s)
        if not m:
            s2 = re.sub(r'[：:，,；;【】、\\s]', '', s)
            m = re.search(f'{keys}.{{0,3}}(\\d+\\.?\\d*)', s2)
        if m and m.group(1):
            dims[label] = float(m.group(1))
    
    # 方法2: 找 数字x数字x数字 模式
    if len(dims) < 3:
        # a*b*c 模式
        m3 = re.search(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)', s)
        if m3:
            vals = [float(m3.group(i)) for i in range(1, 4)]
            if '长' not in dims: dims['长'] = vals[0]
            if '宽' not in dims: dims['宽'] = vals[1]
            if '高' not in dims: dims['高'] = vals[2]
    
    # 方法3: a*b 模式 + 另一个独立数字
    if len(dims) < 3:
        m2 = re.search(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)', s)
        if m2:
            vals_2d = [float(m2.group(1)), float(m2.group(2))]
            # 从所有数字中排除这两个，剩下的取一个
            remaining = [n for n in all_nums if n not in vals_2d]
            if remaining:
                vals_2d.append(remaining[-1])
            if '长' not in dims: dims['长'] = vals_2d[0]
            if '宽' not in dims: dims['宽'] = vals_2d[1]
            if '高' not in dims and len(vals_2d) > 2: dims['高'] = vals_2d[2]
    
    # 方法4: 取最大的三个数字（降序）
    if len(dims) < 3 and len(all_nums) >= 3:
        sorted_nums = sorted(all_nums, reverse=True)
        if '长' not in dims: dims['长'] = sorted_nums[0]
        if '宽' not in dims: dims['宽'] = sorted_nums[1]
        if '高' not in dims: dims['高'] = sorted_nums[2]
    
    # 检查是否三个尺寸都拿到了
    if not (dims.get('长') and dims.get('宽') and dims.get('高')):
        return None  # 无法解析尺寸
    
    result['长'] = dims['长']
    result['宽'] = dims['宽']
    result['高'] = dims['高']
    
    # --- 内外径 ---
    if '内径' in s or '内尺寸' in s or re.search(r'(?<![外])内', s):
        result['内外径'] = '内径'
    else:
        result['内外径'] = '外径'
    
    # --- 材料 ---
    mat = extract_material(s)
    if mat:
        result['材料'] = mat
    
    return result


# ============================================================
# 3. 拼候选编码
# ============================================================
def build_candidates(parsed):
    """
    从解析结果拼出候选主商家编码
    格式: 长*宽*高-内外径-材料[-类型]
    
    返回候选列表（从精确到模糊）
    """
    if not parsed or parsed.get('custom'):
        return []
    
    dim_str = '*'.join(
        str(int(d) if d == int(d) else d)
        for d in [parsed['长'], parsed['宽'], parsed['高']]
    )
    dim_kind = parsed.get('内外径', '外径')
    material = parsed.get('材料', '')
    
    candidates = []
    
    # 1. 完整格式
    if material:
        candidates.append(f"{dim_str}-{dim_kind}-{material}")
    
    # 2. 只尺寸+内外径（匹配时再看材料）
    candidates.append(f"{dim_str}-{dim_kind}")
    
    return candidates


# ============================================================
# 4. 主流程
# ============================================================
def match_platform_to_km(platform_file, km_file, output_dir):
    print("="*60)
    print("步骤1: 加载快麦商品表...")
    km_map = load_km_products(km_file)
    print(f"  加载快麦商品: {len(km_map)} 条")
    
    # 建立索引：按尺寸+内外径分组
    # dim_index = {(38,10,8,外径): [code1, code2, ...]}
    dim_index = {}  # key: (长,宽,高,内外径) -> [编码列表]
    for code in km_map:
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                dims = parts[0].split('*')
                if len(dims) == 3:
                    l, w, h = float(dims[0]), float(dims[1]), float(dims[2])
                    dk = parts[1]
                    key = (l, w, h, dk)
                    dim_index.setdefault(key, []).append(code)
            except:
                pass
    
    print(f"  尺寸索引: {len(dim_index)} 组")
    
    # 也建一个模糊索引（尺寸排好序的）
    fuzzy_index = {}
    for code in km_map:
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                dims = parts[0].split('*')
                if len(dims) == 3:
                    vals = sorted([float(dims[0]), float(dims[1]), float(dims[2])])
                    dk = parts[1]
                    key = (vals[0], vals[1], vals[2], dk)
                    fuzzy_index.setdefault(key, []).append(code)
            except:
                pass
    
    print(f"  模糊索引: {len(fuzzy_index)} 组")
    
    # --- 读取平台商品 ---
    print("\n步骤2: 读取平台商品并匹配...")
    import openpyxl
    
    wb = openpyxl.load_workbook(platform_file, read_only=True, data_only=True)
    ws = wb['报表1']
    
    matched = []
    unmatched = []
    stats = Counter()
    total = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        shop_short, pid, spec_name, spec_id = row
        if not spec_name:
            continue
        total += 1
        
        parsed = parse_spec_brutal(str(spec_name))
        
        if parsed is None:
            unmatched.append((shop_short, pid, spec_id, spec_name, '解析失败', '', ''))
            stats['解析失败'] += 1
            continue
        
        if parsed.get('custom'):
            unmatched.append((shop_short, pid, spec_id, spec_name, '定制', '', ''))
            stats['定制'] += 1
            continue
        
        # 精确匹配：按长宽高顺序匹配
        dim_str = '*'.join(
            str(int(d) if d == int(d) else d)
            for d in [parsed['长'], parsed['宽'], parsed['高']]
        )
        dk = parsed.get('内外径', '外径')
        mat = parsed.get('材料', '')
        
        matched_code = None
        
        # 策略1: 精确尺寸+内外径+材料
        if mat:
            exact = f"{dim_str}-{dk}-{mat}"
            if exact in km_map:
                matched_code = exact
        
        # 策略2: 精确尺寸+内外径（在dim_index里找材料匹配的）
        if not matched_code:
            key = (parsed['长'], parsed['宽'], parsed['高'], dk)
            cands = dim_index.get(key, [])
            if cands:
                # 有材料信息则优先匹配材料
                if mat:
                    for c in cands:
                        if mat in c:
                            matched_code = c
                            break
                if not matched_code:
                    matched_code = cands[0]  # 取第一个
        
        # 策略3: 尺寸排序后模糊匹配
        if not matched_code:
            vals = sorted([parsed['长'], parsed['宽'], parsed['高']])
            fkey = (vals[0], vals[1], vals[2], dk)
            cands = fuzzy_index.get(fkey, [])
            if cands:
                if mat:
                    for c in cands:
                        if mat in c:
                            matched_code = c
                            break
                if not matched_code:
                    matched_code = cands[0]
        
        if matched_code:
            full_name = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched.append((full_name, str(pid).strip(), str(spec_id).strip(), matched_code))
            stats['匹配成功'] += 1
        else:
            expected = f"{dim_str}-{dk}"
            if mat:
                expected += f"-{mat}"
            unmatched.append((shop_short, pid, spec_id, spec_name, '无匹配', expected,
                             f"长={parsed['长']}宽={parsed['宽']}高={parsed['高']} {dk} {mat}"))
            stats['无匹配'] += 1
        
        if total % 100000 == 0:
            print(f"  已处理 {total} 行...")
    
    wb.close()
    
    print(f"\n处理完成: 共 {total} 行")
    print(f"  匹配成功: {stats['匹配成功']}")
    print(f"  定制: {stats['定制']}")
    print(f"  解析失败: {stats['解析失败']}")
    print(f"  无匹配: {stats['无匹配']}")
    
    # ---- 输出文件 ----
    os.makedirs(output_dir, exist_ok=True)
    
    import openpyxl as oxl
    
    # 换绑文件
    bind_file = os.path.join(output_dir, '换绑文件.xlsx')
    wb_out = oxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    ws_out.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for row in matched:
        ws_out.append(list(row))
    wb_out.save(bind_file)
    print(f"\n✅ 换绑文件: {bind_file} ({len(matched)}条)")
    
    # 未匹配文件（含解析信息，方便你判断）
    unmatch_file = os.path.join(output_dir, '未匹配平台商品.xlsx')
    wb_u = oxl.Workbook()
    wb_u.create_sheet('未匹配', 0)
    ws_u = wb_u['未匹配']
    ws_u.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码', '解析结果'])
    for row in unmatched:
        ws_u.append(list(row))
    
    # 再加一个sheet: 解析失败汇总
    failed = [r for r in unmatched if r[4] == '解析失败']
    if failed:
        ws_f = wb_u.create_sheet('解析失败样本')
        ws_f.append(['规格名称'])
        seen = set()
        for r in failed:
            if r[3] not in seen:
                seen.add(r[3])
                ws_f.append([r[3]])
            if len(seen) >= 500:
                break
    
    wb_u.save(unmatch_file)
    print(f"✅ 未匹配文件: {unmatch_file} ({len(unmatched)}条)")
    
    return matched, unmatched, stats


# ============================================================
if __name__ == '__main__':
    import time
    t0 = time.time()
    
    PLATFORM_FILE = r"d:\Desktop\平台商品.xlsx"
    KM_FILE = r"d:\Desktop\快麦商品 - 副本.xlsx"
    OUTPUT_DIR = r"d:\Desktop\换绑输出"
    
    matched, unmatched, stats = match_platform_to_km(PLATFORM_FILE, KM_FILE, OUTPUT_DIR)
    
    elapsed = time.time() - t0
    print(f"\n总耗时: {elapsed:.1f}秒")

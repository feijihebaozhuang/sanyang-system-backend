# -*- coding: utf-8 -*-
"""
批量处理平卡第1批：
1. 优质进口纸-黄色 37*14**3.8（剩余1条）
2. 双白色 37*14**3.8
3. 天猫止合 41个纯定制ID
4. 天猫扣底盒 30*30*X 五层 → EB
5. 天猫彩色 宽度:高度:白色 长度系列（6205条）
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ============================================================
# 快麦索引
# ============================================================
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
exact_idx = {}
fuzzy_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                exact_idx[(vals[0], vals[1], vals[2], dk, mat)] = code
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
            except: pass

def dim_fmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    if mat == '白色':
        cands.append(f"{base}-{dk}-特硬")
    cands.append(f"{base}-{dk}-特硬")
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    return None

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '深圳市友尚包装有限公司',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

def full_name(short):
    return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ============================================================
# 读取平卡
# ============================================================
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"平卡总条数: {len(rows)}", flush=True)

# ============================================================
# 天猫止合纯定制ID
# ============================================================
ZHIGE_CUSTOM_IDS = {
    '5876799254170','5876799254182','5876799254194','5876799254206',
    '5876799254218','5876799254230','5876799254242','5876799254254',
    '5876799254266','5876799254278','5876799254290','5876799254302',
    '5876799254314','5876799254326','5876799254338','5876799254350',
    '5876799254362','5876799254374','5876799254386','5876799254398',
    '5876799254410','5876799254422','5876799254434','5876799254446',
    '5876799254458','5876799254470','5876799254482','5876799254494',
    '5876799254506','5876799254518','5876799254530',
    '5876828754854','5876828754855','5876828754856','5876828754857',
    '5876828754858','5876828754859','5876828754860','5876828754861',
    '5876828754862','5876828754863',
}

# 正则：天猫彩色 宽度：x---高度：y;白色 100个 长度：z
RE_CAISE_WHITE = re.compile(
    r'宽度[：:]\s*(\d+\.?\d*)\s*cm?\s*-+\s*高度[：:]\s*(\d+\.?\d*)\s*cm?\s*;'
    r'\s*白色\s+\d+\s*个\s*长度[：:]\s*(\d+\.?\d*)\s*cm?'
)

# 正则：30*30 cm;XXcm 高;五层
RE_KOUDI_EB = re.compile(r'30\s*\*+\s*30\s*cm?\s*;\s*(\d+\.?\d*)\s*cm?\s*高\s*;\s*五层')

matched_new = []   # (全称, pid, 规格id, 编码)
custom_new = []    # (全称, pid, 规格id, 定制链接)
remaining = []

stats = Counter()

for r in rows:
    if not r: continue
    shop_short = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    reason = str(r[4] or '').strip()
    
    handled = False
    
    # ============================================================
    # 1. 优质进口纸-黄色 37*14**3.8 => 37 14 3.8 外径 特硬
    # ============================================================
    if '优质进口纸-黄色' in spec_name and '37*14' in spec_name and '3.8' in spec_name:
        code = match_3d(37, 14, 3.8, '外径', '特硬')
        if code:
            matched_new.append((full_name(shop_short), pid, spec_id, code))
        else:
            remaining.append(r)
            remaining[-1] = tuple(list(remaining[-1][:4]) + ['无匹配', '37*14*3.8-外径-特硬'] + list(remaining[-1][6:]))
        handled = True
        stats['37*14*3.8-特硬'] += 1
    
    # ============================================================
    # 2. 双白色 37*14**3.8 => 37 14 3.8 外径 白色
    # ============================================================
    if '双白色' in spec_name and '37*14' in spec_name and '3.8' in spec_name:
        code = match_3d(37, 14, 3.8, '外径', '白色')
        if code:
            matched_new.append((full_name(shop_short), pid, spec_id, code))
        else:
            remaining.append(r)
            remaining[-1] = (shop_short, pid, spec_id, spec_name, '无匹配', '37*14*3.8-外径-白色')
        handled = True
        stats['37*14*3.8-白色'] += 1
    
    # ============================================================
    # 3. 天猫止合纯定制ID
    # ============================================================
    if spec_id in ZHIGE_CUSTOM_IDS:
        custom_new.append((full_name(shop_short), pid, spec_id, '定制链接'))
        handled = True
        stats['天猫止合定制'] += 1
    
    # ============================================================
    # 4. 天猫扣底盒 30*30 cm;Xcm 高;五层 → EB
    # ============================================================
    if not handled and '扣底盒' in shop_short:
        m = RE_KOUDI_EB.search(spec_name)
        if m:
            h = float(m.group(1))
            code = match_3d(30, 30, h, '外径', '特硬')
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['扣底盒-特硬'] += 1
            else:
                expected = f"30*30*{dim_fmt(h)}-外径-特硬"
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', expected))
                stats['扣底盒-无匹配'] += 1
            handled = True
    
    # ============================================================
    # 5. 天猫彩色 宽度：x---高度：y;白色 长度：z
    # ============================================================
    if not handled and '彩色' in shop_short:
        m = RE_CAISE_WHITE.search(spec_name)
        if m:
            w = float(m.group(1))
            h = float(m.group(2))
            l = float(m.group(3))
            code = match_3d(l, w, h, '外径', '白色')
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['天猫彩色-白色'] += 1
            else:
                # 尝试特硬
                code2 = match_3d(l, w, h, '外径', '特硬')
                if code2:
                    matched_new.append((full_name(shop_short), pid, spec_id, code2))
                    stats['天猫彩色-特硬(兜底)'] += 1
                else:
                    dims = sorted([l, w, h], reverse=True)
                    expected = f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}-外径-白色"
                    remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', expected))
                    stats['天猫彩色-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f"\n处理统计:", flush=True)
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}", flush=True)
print(f"\n  匹配成功合计: {len(matched_new)}", flush=True)
print(f"  定制链接合计: {len(custom_new)}", flush=True)
print(f"  仍留平卡: {len(remaining)}", flush=True)

# ============================================================
# 写入换绑文件_第7批.xlsx
# ============================================================
if matched_new:
    f7 = os.path.join(out, '换绑文件_第7批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f7)
    sz = os.path.getsize(f7)
    print(f"\n✅ 换绑_第7批: {f7} ({sz/1024:.1f}KB, {len(matched_new)}条)", flush=True)
    wb.close()

# ============================================================
# 写入定制类_换绑文件_第4批.xlsx
# ============================================================
if custom_new:
    f4 = os.path.join(out, '定制类_换绑文件_第4批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for c in custom_new:
        ws.append(list(c))
    wb.save(f4)
    sz = os.path.getsize(f4)
    print(f"✅ 定制类_第4批: {f4} ({sz/1024:.1f}KB, {len(custom_new)}条)", flush=True)
    wb.close()

# ============================================================
# 更新平卡_待处理.xlsx
# ============================================================
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6:
        ws.append(list(r[:6]))
    else:
        ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: 剩余{len(remaining)}条", flush=True)
wb.close()

# 文件列表
print(f"\n{'='*50}")
print("可上传的新文件：")
if len(matched_new) > 0: print(f"  换绑文件_第7批.xlsx（{len(matched_new)}条）")
if len(custom_new) > 0: print(f"  定制类_换绑文件_第4批.xlsx（{len(custom_new)}条）")
print(f"{'='*50}")

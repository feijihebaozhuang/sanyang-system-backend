# -*- coding: utf-8 -*-
"""检查其余商品 26716 条的规格样本"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

f = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # header
    if i == 1: continue  # 标题行
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    rows.append(s)
wb.close()

print(f'其余商品总数: {len(rows)}')
print()

# 分类看格式
stats = {}
for s in rows:
    # 外径关键词
    has_waijing = '外径' in s
    has_neijing = '内径' in s or '内尺寸' in s or '内寸' in s
    # 包含飞机盒
    has_feiji = '飞机盒' in s
    # 是否扣底盒、双插盒关键词（但已被排除）
    has_dikoudi = '扣底盒' in s or '双插盒' in s
    has_zhixiang = '纸箱' in s
    
    key = ('外径' if has_waijing else '') + ('飞机盒' if has_feiji else '')
    if not key:
        if has_neijing:
            key = '内径'
        elif has_dikoudi:
            key = '扣底盒双插盒(残)'
        elif has_zhixiang:
            key = '纸箱(残)'
        else:
            # 只有数值的
            nums = re.findall(r'[\d.]+', s)
            if len(nums) <= 3:
                key = f'裸数值{len(nums)}个'
            else:
                key = f'多数值{len(nums)}个'
    stats[key] = stats.get(key, 0) + 1

print('=== 其余商品规格格式分类 ===')
for k, v in sorted(stats.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v} 条')

# 前50条裸数值样本
print()
print('=== 前50条其余商品规格 ===')
for i, s in enumerate(rows[:50]):
    print(f'  {i+1}. [{s[:130]}]')

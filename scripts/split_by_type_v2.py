# -*- coding: utf-8 -*-
"""
从平台商品中（排除定制链接后），按品类拆分：
1. 扣底盒/双插盒 → 一个文件
2. 纸箱 → 一个文件  
3. 定制链接 → 一个文件
4. 剩下的 → 一个文件
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

f = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(f, header=1, dtype=str)

def is_custom(s):
    """按规格判断是否为定制链接（无准确尺寸）"""
    has_dim_format = bool(re.search(r'[\d.]+\s*[xX*×]\s*[\d.]+\s*[xX*×]\s*[\d.]+', s))
    has_cm_dim = bool(re.search(r'【\s*[\d.]+\s*[xX*×]\s*[\d.]+\s*[xX*×]\s*[\d.]+', s))
    has_lwh_keyword = bool(re.search(r'[长宽高][度]*\s*[：:]?\s*【\s*[\d.]+', s))
    
    if has_dim_format or has_lwh_keyword or has_cm_dim:
        return False  # 有明确尺寸，不是定制
    
    # 没有明确尺寸，检查定制关键词
    keywords = ['定制', '订制', '定做', '加工定制', '不接受退货',
                '咨询客服', '拍下联系客服', '定制产品', '定制拍单',
                '定制尺寸', '万款现货', '联系客服备注', '详情咨询',
                '更多尺寸', '下拉查看', '下拉-', '1000款现模',
                '更多尺寸看详情', '详情-现模', '1000个尺寸']
    for kw in keywords:
        if kw in s: return True
    if '珍珠棉' in s: return True
    
    # 没有数字的简短描述（如"特硬;双插盒"）算定制
    nums = re.findall(r'[\d.]+', s)
    if len(nums) == 0: return True
    
    return False

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or ''));
        w = normalize_dim(m.group(3)+(m.group(4) or ''));
        h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

def classify_dims(s):
    lwh = extract_lwh(s)
    if not lwh: return None
    l, w, h = lwh
    all_int = all(abs(n - round(n)) < 0.001 for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    if all_int: return ('整数', int(round(l)), int(round(w)), int(round(h)))
    elif all_dot5: return ('全5', l, w, h)
    else: return ('小数', l, w, h)

print('正在分类...')
total = len(df)

koudihe = []
zhixiang = []
other = []
custom_data = []

for idx, (_, row) in enumerate(df.iterrows()):
    s = str(row.get('平台规格名称', '')).strip()
    if not s: continue
    
    shop = str(row.get('店铺名称', '')).strip()
    pid = str(row.get('平台商品id', '')).strip()
    sid = str(row.get('平台规格id', '')).strip()
    
    # 先判断定制
    if is_custom(s):
        custom_data.append((shop, pid, s, sid))
        continue
    
    row_data = (shop, pid, s, sid)
    dim_info = classify_dims(s)
    
    if '扣底盒' in s or '扣抵盒' in s or '双插盒' in s or '扣底' in s or '双插' in s:
        koudihe.append(row_data + (dim_info,))
    elif '纸箱' in s:
        zhixiang.append(row_data + (dim_info,))
    else:
        other.append(row_data + (dim_info,))

print(f'扣底盒/双插盒: {len(koudihe)}')
print(f'纸箱: {len(zhixiang)}')
print(f'定制链接: {len(custom_data)}')
print(f'其他: {len(other)}')

import openpyxl

def calc_totals(data_list):
    c = {'整数': 0, '小数': 0, '全5': 0, '无尺寸': 0}
    for item in data_list:
        di = item[4]
        if di is None: c['无尺寸'] += 1
        else: c[di[0]] = c.get(di[0], 0) + 1
    return c

def save_split(path, data, title):
    rows = []
    for item in data:
        shop, pid, s, sid, dim = item
        if dim is None:
            rows.append((shop, pid, s, sid, '', '', '', ''))
        else:
            typ, l, w, h = dim
            rows.append((shop, pid, s, sid, typ, l, w, h))
    
    df_out = pd.DataFrame(rows, columns=['店铺名称','平台商品id','规格名称','平台规格id','尺寸类型','长','宽','高'])
    df_out.to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.insert_rows(1)
    ws.cell(1, 2).value = title
    ws.column_dimensions['C'].width = 60
    for c in ['E','F','G','H']:
        try: ws.column_dimensions[c].width = 10
        except: pass
    wb.save(path)
    print(f'  ✅ {path}: {len(rows)}条')
    t = calc_totals(data)
    print(f'     其中 整数:{t["整数"]} 小数:{t["小数"]} 全5:{t["全5"]} 无尺寸:{t["无尺寸"]}')

save_split(r'D:\Desktop\扣底盒双插盒商品.xlsx', koudihe, '扣底盒/双插盒')
save_split(r'D:\Desktop\纸箱商品.xlsx', zhixiang, '纸箱')
save_split(r'D:\Desktop\其余商品.xlsx', other, '其余商品（飞机盒/信封/正方形等）')

# 定制链接
custom_out = r'D:\Desktop\定制链接商品.xlsx'
pd.DataFrame(custom_data, columns=['店铺名称','平台商品id','规格名称','平台规格id']).to_excel(custom_out, index=False)
wb = openpyxl.load_workbook(custom_out)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '定制链接商品（无准确尺寸）'
ws.column_dimensions['C'].width = 60
wb.save(custom_out)
print(f'  ✅ 定制链接商品.xlsx: {len(custom_data)}条')

total_ck = len(koudihe)+len(zhixiang)+len(other)+len(custom_data)
print(f'\n合计: {total_ck} = {total}')

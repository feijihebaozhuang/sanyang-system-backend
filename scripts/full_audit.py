# -*- coding: utf-8 -*-
"""统计所有数据"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
raw_file = r'd:\Desktop\平台商品.xlsx'

# 原始数据
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
raw_total = 0
raw_shops = Counter()
for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    raw_total += 1
    raw_shops[str(r[0]).strip()] += 1
wb.close()

print(f'原始数据总条数: {raw_total}')
print(f'\n原始数据店铺分布:')
for s, c in raw_shops.most_common():
    print(f'  [{c:>6}] {s}')

# 平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
pk_total = 0
pk_shops = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    pk_total += 1
    pk_shops[str(r[0]).strip()] += 1
wb.close()

print(f'\n平卡总条数: {pk_total}')
print(f'平卡店铺分布:')
for s, c in pk_shops.most_common():
    print(f'  [{c:>6}] {s}')

# 所有输出文件汇总
print(f'\n=== 输出文件汇总 ===')

# 换绑文件
hb_total = 0
for f in sorted(os.listdir(out)):
    if '换绑' in f and f.endswith('.xlsx'):
        wb = oxl.load_workbook(os.path.join(out, f))
        for sn in wb.sheetnames:
            cnt = 0
            for r in wb[sn].iter_rows(min_row=3, values_only=True):
                if r and r[0]: cnt += 1
            hb_total += cnt
            print(f'  {f}[{sn}]: {cnt}条')
        wb.close()
print(f'换绑文件总计: {hb_total}条')

# 定制类
dz_total = 0
for f in sorted(os.listdir(out)):
    if '定制' in f and f.endswith('.xlsx'):
        wb = oxl.load_workbook(os.path.join(out, f))
        for sn in wb.sheetnames:
            cnt = 0
            for r in wb[sn].iter_rows(min_row=3, values_only=True):
                if r and r[0]: cnt += 1
            dz_total += cnt
            print(f'  {f}[{sn}]: {cnt}条')
        wb.close()
print(f'定制文件总计: {dz_total}条')

# 无匹配
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
wm_total = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]: wm_total += 1
wb.close()
print(f'无匹配_待处理: {wm_total}条')

print(f'\n=== 总数验证 ===')
print(f'换绑+定制+平卡+无匹配 = {hb_total}+{dz_total}+{pk_total}+{wm_total} = {hb_total+dz_total+pk_total+wm_total}')
print(f'原始数据: {raw_total}')

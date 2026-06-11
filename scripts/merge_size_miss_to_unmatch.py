# -*- coding: utf-8 -*-
"""将尺寸不足剩余75条合并到无匹配文件"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl
from openpyxl import load_workbook

out = r'd:\Desktop\换绑输出'

# 读取尺寸不足剩余75条
wb = load_workbook(os.path.join(out, '尺寸不足_待处理.xlsx'), data_only=True)
ws = wb['尺寸不足']
miss_rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"尺寸不足剩余: {len(miss_rows)}条", flush=True)

# 追加到无匹配_待处理.xlsx
f_unmatch = os.path.join(out, '无匹配_待处理.xlsx')
wb = load_workbook(f_unmatch, data_only=True)
ws = wb['无匹配']
for r in miss_rows:
    ws.append(list(r))
wb.save(f_unmatch)
print(f"已追加到无匹配_待处理.xlsx", flush=True)
wb.close()

# 删除尺寸不足_待处理.xlsx
os.remove(os.path.join(out, '尺寸不足_待处理.xlsx'))
print("已删除尺寸不足_待处理.xlsx", flush=True)

# 文件大小
sz = os.path.getsize(f_unmatch)
print(f"无匹配_待处理.xlsx: {sz/1024/1024:.2f}MB", flush=True)
print("完成！", flush=True)

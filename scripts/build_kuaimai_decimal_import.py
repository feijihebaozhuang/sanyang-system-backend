# -*- coding: utf-8 -*-
"""小数点库存 + 刀模未覆盖剩余 → 单份快麦商品导入表（规则同四批 finalize）。"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import openpyxl
import quote_weight as qw
import transform_kuaimai_excel as tk

SRC = Path(r"d:\Desktop\小数点商品_库存_处理完成.xlsx")
SQ_REMAIN = Path(r"d:\Desktop\刀模_正方形_未覆盖剩余.xlsx")
RC_REMAIN = Path(r"d:\Desktop\刀模_长方形_未覆盖剩余.xlsx")
OUT = Path(r"d:\Desktop\小数点商品_快麦导入.xlsx")

COL_MAIN = 1
COL_ALIAS = 29
COL_SUPPLIER = 33
COL_WEIGHT = 39
COL_LEN = 40
COL_W = 41
COL_H = 42
COL_SPEC_STD = 52
COL_EXEC_STD = 115
COL_SHORT = 73
COL_G_WEIGHT = 100
COL_G_LEN = 101
COL_G_W = 102
COL_G_H = 103
COL_REMARK = 105
DATA_START = 4


def _load_sq_remain(path: Path) -> list[tuple[str, str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str]] = []
    for r in range(2, ws.max_row + 1):
        size = tk._norm_size(ws.cell(r, 1).value)
        typ = str(ws.cell(r, 2).value or "").strip()
        code = str(ws.cell(r, 3).value or "").strip()
        if size and typ and code:
            rows.append((size, typ, code))
    wb.close()
    return rows


def _load_rc_remain(path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    grp: dict[str, list[str]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        size = tk._norm_size(ws.cell(r, 1).value)
        code = str(ws.cell(r, 2).value or "").strip()
        if size and code:
            grp[size].append(code)
    wb.close()
    return {k: tk._join_codes(v) for k, v in grp.items()}


def _lookup_dimold(
    main,
    sq_entries: list[tuple[str, str, str]],
    rc_map: dict[str, str],
    sq_sizes: set[str],
) -> str | None:
    size, typ = tk._parse_main_dimold(main)
    if not size:
        return None

    if typ == "带扣":
        codes = [c for s, t, c in sq_entries if s == size and t == "带扣"]
        if codes:
            return tk._join_codes(codes)

    if typ == "内":
        codes = [c for s, t, c in sq_entries if s == size and t == "内"]
        if codes:
            return tk._join_codes(codes)

    if typ == "外":
        outer = [
            (t, c)
            for s, t, c in sq_entries
            if s == size and t not in ("内", "带扣") and "外" in t
        ]
        if outer:
            parts: list[str] = []
            for t, c in outer:
                for p in tk._format_outer_entry(c, t).split(","):
                    if p and p not in parts:
                        parts.append(p)
            return ",".join(parts)
        plain = [c for s, t, c in sq_entries if s == size and t == "外"]
        if plain:
            return tk._join_codes(plain)

    if size in rc_map:
        return rc_map[size]
    return None


def _simplify_main_code(main) -> str | None:
    main = str(main or "").strip()
    m = re.match(r"^([\d.*×xX]+)", main)
    if not m:
        return None
    parts = re.split(r"[*×xX]", m.group(1))
    chunks: list[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        d = re.sub(r"\D", "", p)
        if d:
            chunks.append(d)
    return "".join(chunks) if chunks else None


def _parse_dims(main) -> tuple[float, float, float]:
    main = str(main or "").strip()
    m = re.match(r"^([\d.*×xX]+)", main)
    if not m:
        return 0.0, 0.0, 0.0
    parts = re.split(r"[*×xX]", m.group(1))
    vals: list[float] = []
    for p in parts[:3]:
        p = p.strip()
        if not p:
            continue
        try:
            vals.append(float(p))
        except ValueError:
            nums = re.findall(r"\d+(?:\.\d+)?", p)
            vals.append(float(nums[0]) if nums else 0.0)
    while len(vals) < 3:
        vals.append(0.0)
    return vals[0], vals[1], vals[2]


def _material_from_main(main: str) -> str:
    parts = (main or "").split("-")
    if len(parts) >= 2:
        tail = parts[-1].strip()
        if tail and tail not in ("内径", "外径"):
            return tail
    return ""


def _move_spec_to_goods(ws, row: int) -> None:
    for src, dst in (
        (COL_WEIGHT, COL_G_WEIGHT),
        (COL_LEN, COL_G_LEN),
        (COL_W, COL_G_W),
        (COL_H, COL_G_H),
    ):
        val = ws.cell(row, src).value
        if val is not None:
            ws.cell(row, dst).value = val
            ws.cell(row, src).value = None


def build() -> dict:
    if not SRC.is_file():
        raise FileNotFoundError(SRC)
    if not SQ_REMAIN.is_file():
        raise FileNotFoundError(SQ_REMAIN)
    if not RC_REMAIN.is_file():
        raise FileNotFoundError(RC_REMAIN)

    sq_entries = _load_sq_remain(SQ_REMAIN)
    sq_entries.extend(tk.EXTRA_SQ_ENTRIES)
    sq_sizes = {s for s, _, _ in sq_entries}
    rc_map = _load_rc_remain(RC_REMAIN)

    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    stats = {
        "rows": 0,
        "dimold": 0,
        "weight": 0,
        "loc": 0,
    }

    for row in range(DATA_START, ws.max_row + 1):
        main = ws.cell(row, COL_MAIN).value
        if main is None or str(main).strip() == "":
            continue
        stats["rows"] += 1
        main_s = str(main).strip()

        knife = _lookup_dimold(main_s, sq_entries, rc_map, sq_sizes)
        if knife:
            ws.cell(row, COL_SUPPLIER).value = knife
            stats["dimold"] += 1

        if ws.cell(row, COL_REMARK).value:
            stats["loc"] += 1

        ln, wd, ht = _parse_dims(main_s)
        if ln and wd and ht:
            ws.cell(row, COL_LEN).value = ln
            ws.cell(row, COL_W).value = wd
            ws.cell(row, COL_H).value = ht

        mat = _material_from_main(main_s)
        text = " ".join(
            p for p in [main_s, str(ws.cell(row, 2).value or ""), mat] if p
        )
        est = qw.estimate_unit_weight(
            ln, wd, ht, text=text, product_type="", material_key=mat
        )
        if est.get("ok"):
            ws.cell(row, COL_WEIGHT).value = round(float(est["weight_per_unit_kg"]), 4)
            stats["weight"] += 1

        simp = _simplify_main_code(main_s)
        if simp:
            ws.cell(row, COL_SHORT).value = simp
        ws.cell(row, COL_ALIAS).value = None
        _move_spec_to_goods(ws, row)

        knife_val = ws.cell(row, COL_SUPPLIER).value
        ws.cell(row, COL_SPEC_STD).value = None
        if knife_val is not None and str(knife_val).strip() != "":
            ws.cell(row, COL_EXEC_STD).value = knife_val

    wb.save(OUT)
    wb.close()
    return stats


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    st = build()
    print(f"完成: {OUT.name}")
    print(
        f"  商品 {st['rows']} 行 | 刀模 {st['dimold']} | 重量 {st['weight']} | 库位 {st['loc']}"
    )
    print("规则: 主商家编码/商品名称不动 | 简称=编码简化 | 规格重/长宽高→商品列")
    print("      刀模→供应商名称+执行标准(115) | 库位→商品备注 | 规格别名清空")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

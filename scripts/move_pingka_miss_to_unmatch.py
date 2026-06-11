# -*- coding: utf-8 -*-
"""平卡中所有'无匹配'的移到未匹配平台商品.xlsx"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# 分离定制/匹配成功的 和 无匹配的
custom_or_matched = []
miss_to_move = []
for r in rows:
    if not r:
        continue
    reason = str(r[4] or '').strip()
    if reason == '无匹配':
        miss_to_move.append(r)
    else:
        custom_or_matched.append(r)

print(f"平卡总数: {len(rows)}", flush=True)
print(f"  定制/匹配成功: {len(custom_or_matched)}", flush=True)
print(f"  无匹配(移到未匹配): {len(miss_to_move)}", flush=True)

# 更新平卡（只留定制/匹配成功的）
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in custom_or_matched:
    if len(r) >= 6:
        ws.append(list(r[:6]))
    else:
        ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: 剩余{len(custom_or_matched)}条", flush=True)
wb.close()

# 追加到未匹配平台商品
f_unmatch = os.path.join(out, '未匹配平台商品.xlsx')
wb = oxl.load_workbook(f_unmatch, data_only=True)
# 找到'未匹配'sheet
ws_names = wb.sheetnames
print(f"未匹配文件sheet: {ws_names}", flush=True)

if '未匹配' in ws_names:
    ws = wb['未匹配']
elif 'Sheet' in ws_names:
    ws = wb.active
else:
    ws = wb.active

for r in miss_to_move:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])

wb.save(f_unmatch)
sz = os.path.getsize(f_unmatch)
print(f"✅ 未匹配平台商品: 已追加{len(miss_to_move)}条, {sz/1024/1024:.2f}MB", flush=True)
wb.close()

print("\n完成！平卡现在只保留定制/匹配成功的记录。")

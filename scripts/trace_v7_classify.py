# -*- coding: utf-8 -*-
"""trace_v7_classify: 模仿v7分类代码，跟踪外径规格的去向"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# 复制v7的extract_all_dims
def extract_all_dims(s):
    results = []
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(3)), float(m.group(1)), float(m.group(2)), 'A1'))
    m = re.search(r'【长度\s*([\d.]+)\s*厘米\s*】.*?【宽度\s*([\d.]+)\s*厘米\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), 0, 'A2'))
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A3'))
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A4'))
    m = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)(?:\s*mm)?', s)
    if m:
        v = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        if any(x > 50 for x in v):
            v = [x/10 for x in v]
        results.append((v[0], v[1], v[2], 'A5'))
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B1'))
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B2'))
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B3'))
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B4'))
    return results

def build_lwh(s):
    all_dims = extract_all_dims(s)
    if not all_dims: return None
    for l, w, h, src in all_dims:
        if h > 0.1 and 0.5 <= l <= 200 and 0.5 <= w <= 200 and 0.5 <= h <= 200:
            return (l, w, h)
    for l, w, h, src in all_dims:
        if 0.5 <= l <= 200 and 0.5 <= w <= 200:
            return (l, w, h)
    l, w, h, src = all_dims[0]
    return (l, w, h)

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量飞机盒'
    elif all_5: return '全.5'
    else: return '整数'

# 扫描外径规格并分类
from openpyxl import load_workbook
source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

cats = {}
total_waijing = 0
sample = 200
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    total_waijing += 1
    if total_waijing > sample:
        break
    
    # 模拟v7分类
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not extract_all_dims(s):
        cat = '定制(关键词)'
    elif '扣底盒' in s or '双插盒' in s:
        cat = '扣底盒'
    elif '纸箱' in s:
        cat = '纸箱'
    else:
        lwh = build_lwh(s)
        if lwh is None:
            cat = '定制(无LWH)'
        else:
            dt = classify_dims(lwh)
            if '内径' in s or '内尺寸' in s or '内寸' in s:
                cat = '内径'
            elif dt == '非全量飞机盒':
                cat = '非全量飞机盒'
            elif '外径' in s:
                cat = '外径全量飞机盒'
            else:
                cat = '其余'
    cats[cat] = cats.get(cat, 0) + 1

wb.close()
print(f'总外径规格: {total_waijing} (采样{min(total_waijing, sample)})')
print('分类分布:')
for k, v in sorted(cats.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')

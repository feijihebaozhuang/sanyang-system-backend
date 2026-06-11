# -*- coding: utf-8 -*-
"""将定制链接的所有格式输出到桌面文件"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

def skeleton(s):
    s = re.sub(r'\d+\.?\d*', 'N', s)
    s = re.sub(r'\s+', '', s)
    return s[:300]

by_shop_skeleton = {}  # key -> (样例, pid, 计数)
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total += 1
    spec = str(row[2] or '').strip() if len(row) > 2 else ''
    shop = str(row[0] or '').strip() if len(row) > 0 else '(无名)'  # 列0=店铺名称
    pid = str(row[3] or '') if len(row) > 3 else ''
    if not spec: continue
    sk = skeleton(spec)
    key = (shop, sk)
    if key in by_shop_skeleton:
        sample, first_pid, count = by_shop_skeleton[key]
        by_shop_skeleton[key] = (sample, first_pid, count + 1)
    else:
        by_shop_skeleton[key] = (spec, pid, 1)
wb.close()

lines = []
lines.append(f'定制链接总数: {total}')
lines.append(f'不同格式（按店铺+结构）: {len(by_shop_skeleton)}')
lines.append('')

by_shop = {}
for (shop, sk), (spec, pid, count) in by_shop_skeleton.items():
    by_shop.setdefault(shop, []).append((sk, spec, pid, count))

for shop in sorted(by_shop):
    items = by_shop[shop]  # list of (sk, spec, pid, count)
    total_in_shop = sum(c for _, _, _, c in items)
    lines.append(f'═══ {shop}（{len(items)} 种格式, 共 {total_in_shop} 条）═══')
    for sk, spec, pid, count in items:
        lines.append(f'  [x{count}] 结构: {sk[:120]}')
        lines.append(f'    样例: {spec[:120]}')
        lines.append(f'    pid={pid}')
        lines.append('')

out = '\n'.join(lines)

# 输出到文件
with open(r'D:\Desktop\定制链接所有格式.txt', 'w', encoding='utf-8') as f:
    f.write(out)

print(f'已输出到 D:\\Desktop\\定制链接所有格式.txt')
print(f'共 {len(lines)} 行')

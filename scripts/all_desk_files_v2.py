# -*- coding: utf-8 -*-
"""
从平台商品.xlsx，全部以cm为准：
- 213mm = 21.3cm（含小数点）
- 1mm = 0.1cm（含小数点）
按cm值判断长宽高是否为整数/含小数/全.5
输出4个文件到桌面
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

f = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(f, header=1, dtype=str)

def normalize_dim(v_str):
    """将数字+单位转成cm值"""
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'):
        unit = 'cm'
        val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'):
        unit = 'mm'
        val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'):
        unit = 'cm'
        val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'):
        unit = 'mm'
        val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        if unit == 'mm':
            v = v / 10
        return v
    except:
        return None

def extract_lwh_cm(s):
    """从规格名称中提取长宽高，全部转为cm"""
    s = str(s)
    results = []
    
    # 模式1: 找所有【数字+单位】x【数字+单位】x【数字+单位】
    # 外尺寸【213x105x0.5】mm 或 外尺寸【21.3x10.5x0.5】cm
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l, w, h = normalize_dim(m.group(1)+m.group(4)) if m.group(4) else float(m.group(1)), \
                   normalize_dim(m.group(2)+m.group(4)) if m.group(4) else float(m.group(2)), \
                   normalize_dim(m.group(3)+m.group(4)) if m.group(4) else float(m.group(3))
        if l and w and h and 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
            return (l, w, h)
    
    # 模式2: 【长Lx宽Wx高H】
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
            return (l, w, h)
    
    # 模式3: 长【Lcm】宽【Wcm】高【Hcm】
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or ''))
        w = normalize_dim(m.group(3)+(m.group(4) or ''))
        h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
            return (l, w, h)
    
    # 模式4: LxWxH+单位（在【】里但没有外尺寸前缀的）
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
            return (l, w, h)
    
    # 模式5: L*W*H 后面有cm/mm 或直接是LxWxH不带【】
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
            return (l, w, h)
    
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

dot_results = []    # 含小数非全.5
dot5_results = []   # 三项全.5
int_results = []    # 整数

total = len(df)
for idx, (_, row) in enumerate(df.iterrows()):
    s = str(row.get('平台规格名称', '')).strip()
    if not s: continue
    
    lwh = extract_lwh_cm(s)
    if not lwh: continue
    
    l, w, h = lwh
    shop = str(row.get('店铺名称', '')).strip()
    pid = str(row.get('平台商品id', '')).strip()
    sid = str(row.get('平台规格id', '')).strip()
    
    all_int = all(abs(n - round(n)) < 0.001 for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    
    if all_int:
        int_results.append((shop, pid, s, int(round(l)), int(round(w)), int(round(h))))
    elif all_dot5:
        dot5_results.append((shop, pid, s, l, w, h))
    else:
        dot_results.append((shop, pid, s, l, w, h))

print(f'总行: {total}')
print(f'长宽高含小数非全.5: {len(dot_results)}')
print(f'长宽高三项全.5: {len(dot5_results)}')
print(f'长宽高为整数: {len(int_results)}')

import openpyxl

def save_file(name, data, title, header):
    path = rf'D:\Desktop\{name}'
    pd.DataFrame(data, columns=header).to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.insert_rows(1)
    ws.cell(1, 2).value = title
    ws.column_dimensions['C'].width = 60
    for c in ['D','E','F']:
        ws.column_dimensions[c].width = 12
    wb.save(path)
    print(f'  ✅ {name}: {len(data)}条')

save_file('长宽高含小数商品.xlsx', dot_results, '长宽高含小数（已排除全.5）', ['店铺名称','平台商品id','规格名称','长','宽','高'])
save_file('长宽高三项全5商品.xlsx', dot5_results, '长宽高三项全.5', ['店铺名称','平台商品id','规格名称','长','宽','高'])
save_file('长宽高为整数商品.xlsx', int_results, '长宽高为整数', ['店铺名称','平台商品id','规格名称','长','宽','高'])

# 整数商品：从原始df排除所有能提取到长宽高的
extracted_ids = set()
for results_list in [dot_results, dot5_results, int_results]:
    for r in results_list:
        extracted_ids.add(r[1])  # 平台商品id

mask = ~df['平台商品id'].astype(str).isin(extracted_ids)
remaining = df[mask].copy()
remaining.to_excel(r'D:\Desktop\整数商品.xlsx', index=False, columns=['店铺名称', '平台商品id', '平台规格名称', '平台规格id'])
wb = openpyxl.load_workbook(r'D:\Desktop\整数商品.xlsx')
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '整数商品（已排除所有含小数的商品）'
ws.column_dimensions['C'].width = 60
wb.save(r'D:\Desktop\整数商品.xlsx')
print(f'  ✅ 整数商品.xlsx: {len(remaining)}条')

print('\n=== 桌面文件 ===')
print(f'1. 长宽高含小数商品.xlsx → {len(dot_results)}条')
print(f'2. 长宽高三项全5商品.xlsx → {len(dot5_results)}条')
print(f'3. 长宽高为整数商品.xlsx → {len(int_results)}条')
print(f'4. 整数商品.xlsx → {len(remaining)}条')
print(f'\n合计: {len(dot_results)+len(dot5_results)+len(int_results)+len(remaining)} = {total}')

# -*- coding: utf-8 -*-
"""
从平台商品.xlsx中：
1. 长宽高含小数非全.5 → 长宽高含小数商品.xlsx (已有17276条，跳过)
2. 长宽高三项全.5 → 长宽高三项全5商品.xlsx
3. 长宽高为整数 → 长宽高为整数商品.xlsx
4. 剩余全部(排除小数17276条) → 整数商品.xlsx (已有471573条，跳过)
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

f = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(f, header=1, dtype=str)

def extract_lwh(s):
    s = str(s)
    patterns = [
        r'外[尺寸]*寸*[大小]*[：:]?\s*【\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*cm*\s*】',
        r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】',
        r'长[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*cm*\s*】',
        r'(?:外尺寸|外径|外寸)[^；;]*?【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】',
        r'(?:长|长度)[：:]?\s*([\d.]+)\s*cm*\s*(?:宽|宽度)[：:]?\s*([\d.]+)\s*cm*\s*(?:高|高度)[：:]?\s*([\d.]+)\s*cm*',
        r'^\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*(?:cm|mm|C)?',
        r'(?:^|[；;，,\s])+\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(?:cm|mm|C)?(?:\s|$|【|；|;|,)',
    ]
    for pat in patterns:
        m = re.search(pat, s, re.I | re.S)
        if m:
            try:
                l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
                if 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
                    return (l, w, h)
            except: pass
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)(?:\s*cm|\s*mm|\s*C)?(?![\d.])', s)
    if m:
        try:
            l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
            if 0 < l <= 200 and 0 < w <= 200 and 0 < h <= 200:
                return (l, w, h)
        except: pass
    return None

def is_dot5(v):
    return v != int(v) and v * 2 == int(v * 2)

int_results = []
dot5_results = []

total = len(df)
for idx, (_, row) in enumerate(df.iterrows()):
    s = str(row.get('平台规格名称', '')).strip()
    if not s: continue
    
    lwh = extract_lwh(s)
    if not lwh: continue
    
    l, w, h = lwh
    shop = str(row.get('店铺名称', '')).strip()
    pid = str(row.get('平台商品id', '')).strip()
    sid = str(row.get('平台规格id', '')).strip()
    
    all_int = all(n == int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    
    if all_int:
        int_results.append((shop, pid, s, int(l), int(w), int(h)))
    elif all_dot5:
        dot5_results.append((shop, pid, s, l, w, h))

print(f'总行: {total}')
print(f'长宽高为整数: {len(int_results)}')
print(f'长宽高三项全.5: {len(dot5_results)}')
print(f'含小数非全.5(已有): 17276')

import openpyxl

# === 2. 三项全.5 ===
out2 = r'D:\Desktop\长宽高三项全5商品.xlsx'
pd.DataFrame(dot5_results, columns=['店铺名称', '平台商品id', '规格名称', '长', '宽', '高']).to_excel(out2, index=False)
wb2 = openpyxl.load_workbook(out2)
ws2 = wb2.active
ws2.insert_rows(1)
ws2.cell(1, 2).value = '长宽高三项全为.5的商品'
ws2.column_dimensions['C'].width = 60
for c, v in [('D','长'),('E','宽'),('F','高')]:
    ws2.column_dimensions[c].width = 12
wb2.save(out2)
print(f'\n✅ 长宽高三项全5商品.xlsx: {len(dot5_results)}条')

# === 3. 长宽高为整数 ===
out3 = r'D:\Desktop\长宽高为整数商品.xlsx'
pd.DataFrame(int_results, columns=['店铺名称', '平台商品id', '规格名称', '长', '宽', '高']).to_excel(out3, index=False)
wb3 = openpyxl.load_workbook(out3)
ws3 = wb3.active
ws3.insert_rows(1)
ws3.cell(1, 2).value = '长宽高为整数的商品'
ws3.column_dimensions['C'].width = 60
for c, v in [('D','长'),('E','宽'),('F','高')]:
    ws3.column_dimensions[c].width = 12
wb3.save(out3)
print(f'\n✅ 长宽高为整数商品.xlsx: {len(int_results)}条')

print('\n=== 桌面文件清单 ===')
print('1. 长宽高含小数商品.xlsx (17276条) ✅')
print('2. 长宽高三项全5商品.xlsx ({}条) ✅'.format(len(dot5_results)))
print('3. 长宽高为整数商品.xlsx ({}条) ✅'.format(len(int_results)))
print('4. 整数商品.xlsx (471573条) ✅')

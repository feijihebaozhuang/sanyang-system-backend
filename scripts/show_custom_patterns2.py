# -*- coding: utf-8 -*-
"""定制链接：按店铺+规格模式分类，输出所有格式（每个模式一个样例）"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

# 先看第一行实际数据确定列
header = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:  # 原始数据的第一行
        header = list(row)
        break

print(f'列数: {len(header)}')
print(f'列: {[str(h)[:20] for h in header]}')
print()

# 确定店铺简称列和规格名列
# 常见: index2=规格名称, index3=平台商品id, index4=店铺简称
shop_col = None
spec_col = None
pid_col = None
for j, h in enumerate(header):
    if h and '简称' in str(h):
        shop_col = j
    if h and ('规格' in str(h) or '规格名称' in str(h)):
        spec_col = j
    if h and '商品id' in str(h) or h and 'pid' in str(h).lower():
        pid_col = j

print(f'店铺简称列: {shop_col} ({header[shop_col] if shop_col is not None else "?"})')
print(f'规格名称列: {spec_col} ({header[spec_col] if spec_col is not None else "?"})')
print(f'商品id列: {pid_col}')
wb.close()

# 重新遍历
wb2 = load_workbook(f, read_only=True)
ws2 = wb2.active

patterns = {}
total_rows = 0
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total_rows += 1
    
    if spec_col is not None and len(row) > spec_col:
        spec = str(row[spec_col] or '').strip()
    elif len(row) > 2:
        spec = str(row[2] or '').strip()
    else:
        continue
    
    if shop_col is not None and len(row) > shop_col:
        shop = str(row[shop_col] or '').strip()
    elif len(row) > 4:
        shop = str(row[4] or '').strip()
    else:
        shop = '(无名)'
    
    pid = str(row[pid_col] or '') if pid_col is not None and len(row) > pid_col else ''
    if not pid and len(row) > 3:
        pid = str(row[3] or '')
    
    if not spec: continue
    
    # 生成模式签名
    sig = spec
    sig = re.sub(r'\d+\.?\d*', 'N', sig)
    sig = re.sub(r'\s+', ' ', sig)
    sig = sig[:200]
    
    key = (shop, sig)
    if key not in patterns:
        patterns[key] = (spec, pid)

wb2.close()

print(f'定制链接总数: {total_rows}')
print(f'不同格式数: {len(patterns)}')
print()

# 按店铺分组输出
by_shop = {}
for (shop, sig), (spec, pid) in patterns.items():
    by_shop.setdefault(shop, []).append((sig, spec, pid))

for shop in sorted(by_shop):
    items = by_shop[shop]
    print(f'═══ {shop}（{len(items)} 种格式）═══')
    for sig, spec, pid in items:
        print(f'  例: {spec[:150]}')
    print()

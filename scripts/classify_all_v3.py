# -*- coding: utf-8 -*-
"""
从平台商品.xlsx 49万多条中：
1. 先分出【定制链接】— 没有准确长宽高数值的（定制、订制、定做、珍珠棉、咨询客服等）
2. 剩下的再按长宽高是否含小数/整数分类
输出到桌面文件
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re
from collections import Counter

f = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(f, header=1, dtype=str)

specs = df['平台规格名称'].dropna().astype(str).str.strip()

# ===== 1. 识别定制链接 =====
def is_custom(s):
    """判断是否为定制链接（没有准确尺寸）"""
    # 关键词
    keywords = [
        '定制', '订制', '定做', '加工定制', '不接受退货',
        '咨询客服', '拍下联系客服', '定制产品', '定制拍单',
        '定制尺寸', '万款现货', '联系客服备注', '详情咨询',
    ]
    for kw in keywords:
        if kw in s: return True
    
    # 珍珠棉
    if '珍珠棉' in s: return True
    
    # 没有具体数字的纯描述
    # 比如只有层数和材料名，没有长宽高数字
    # 检查是否有LxWxH格式或【数字】格式
    has_dim = bool(re.search(r'[\d.]+\s*[xX*×]\s*[\d.]+', s)) or \
              bool(re.search(r'【\s*[\d.]+', s)) or \
              bool(re.search(r'(?:长|宽|高)[度]*\s*【\s*[\d.]+', s))
    
    # 如果完全没有尺寸数字且很短的描述，也算定制
    nums = re.findall(r'[\d.]+', s)
    if len(nums) < 3 and not has_dim:
        # 排除有明确品名+尺寸的
        if '飞机盒' in s or '纸箱' in s or '扣底盒' in s or '信封' in s:
            return False
        # 没有数字的简短描述=定制
        if len(re.sub(r'[\s【】；;，,、\-/（）()（）【】]', '', s)) < 20:
            return False
        return True
    
    return False

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    # 外尺寸【LxWxH】+单位
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    # 【长Lx宽Wx高H】
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    # 长【L】宽【W】高【H】
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or ''));
        w = normalize_dim(m.group(3)+(m.group(4) or ''));
        h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    # 【LxWxH】
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    # LxWxH+单位
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

# ===== 分类 =====
custom_data = []      # 定制链接
remain_data = []      # 剩下的（有明确尺寸的）

total = len(df)
found_custom = 0
found_remain = 0

for idx, (_, row) in enumerate(df.iterrows()):
    s = str(row.get('平台规格名称', '')).strip()
    if not s: continue
    
    shop = str(row.get('店铺名称', '')).strip()
    pid = str(row.get('平台商品id', '')).strip()
    sid = str(row.get('平台规格id', '')).strip()
    
    if is_custom(s):
        custom_data.append((shop, pid, s, sid))
        found_custom += 1
    else:
        remain_data.append((shop, pid, s, sid))
        found_remain += 1

print(f'总行: {total}')
print(f'定制链接: {len(custom_data)}')
print(f'剩余: {len(remain_data)}')

# 再从剩余中分出含小数/整数
dot_data = []
dot5_data = []
dot5_nodim = []
int_data = []

for row in remain_data:
    shop, pid, s, sid = row
    lwh = extract_lwh(s)
    if not lwh: 
        dot5_nodim.append((shop, pid, s, sid))
        continue
    l, w, h = lwh
    all_int = all(abs(n - round(n)) < 0.001 for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    if all_int:
        int_data.append((shop, pid, s, sid, int(round(l)), int(round(w)), int(round(h))))
    elif all_dot5:
        dot5_data.append((shop, pid, s, sid, l, w, h))
    else:
        dot_data.append((shop, pid, s, sid, l, w, h))

print(f'  含小数非全.5: {len(dot_data)}')
print(f'  三项全.5: {len(dot5_data)}')
print(f'  未提取尺寸(也放全.5文件): {len(dot5_nodim)}')
print(f'  整数: {len(int_data)}')

import openpyxl
def save_file(name, data, title, header):
    path = rf'D:\Desktop\{name}'
    pd.DataFrame(data, columns=header).to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.insert_rows(1)
    ws.cell(1, 2).value = title
    if '规格名称' in header:
        try: ws.column_dimensions['C'].width = 60
        except: pass
    for c in ['D','E','F']:
        try: ws.column_dimensions[c].width = 12
        except: pass
    wb.save(path)
    print(f'  ✅ {name}: {len(data)}条')

save_file('定制链接商品.xlsx', custom_data, '定制链接（无准确尺寸）', ['店铺名称','平台商品id','规格名称','平台规格id'])
save_file('长宽高含小数商品.xlsx', dot_data, '长宽高含小数（已排除全.5）', ['店铺名称','平台商品id','规格名称','平台规格id','长','宽','高'])
save_file('长宽高三项全5商品.xlsx', dot5_data, '长宽高三项全.5', ['店铺名称','平台商品id','规格名称','平台规格id','长','宽','高'])
if dot5_nodim:
    df_tmp = pd.DataFrame(dot5_nodim, columns=['店铺名称','平台商品id','规格名称','平台规格id'])
    # 追加到全.5文件
    import openpyxl
    wb = openpyxl.load_workbook(r'D:\Desktop\长宽高三项全5商品.xlsx')
    ws = wb.active
    start = ws.max_row + 1
    for r in dot5_nodim:
        ws.cell(start, 1).value = r[0]
        ws.cell(start, 2).value = r[1]
        ws.cell(start, 3).value = r[2]
        start += 1
    wb.save(r'D:\Desktop\长宽高三项全5商品.xlsx')
    print(f'  ✅ 追加未提取尺寸 {len(dot5_nodim)}条到三项全5文件')
save_file('长宽高为整数商品.xlsx', int_data, '长宽高为整数', ['店铺名称','平台商品id','规格名称','平台规格id','长','宽','高'])

print('\n=== 桌面文件清单 ===')
print(f'1. 定制链接商品.xlsx → {len(custom_data)}条（没有准确尺寸的）')
print(f'2. 长宽高含小数商品.xlsx → {len(dot_data)}条')
print(f'3. 长宽高三项全5商品.xlsx → {len(dot5_data)}条（含未提取到尺寸的）')
print(f'4. 长宽高为整数商品.xlsx → {len(int_data)}条')
print(f'\n合计: {len(custom_data)+len(dot_data)+len(dot5_data)+len(int_data)} = {total}')

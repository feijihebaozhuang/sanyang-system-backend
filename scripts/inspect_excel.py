# -*- coding: utf-8 -*-
"""快麦商品文件: 检查所有列"""
import openpyxl, sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

fp = r"d:\Desktop\快麦商品.xlsx"
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
ws = wb['报表1']

# 读前20行，看所有列
max_c = 0
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False)):
    vals = [c.value for c in row]
    non_null = [(j, v) for j, v in enumerate(vals) if v is not None]
    if non_null:
        max_c = max(max_c, max(j for j, _ in non_null))
        print(f"  第{i+1}行: {non_null}")

print(f"\n最高列号: {max_c}（共{max_c+1}列）")

# 再读第4行全部列看看有什么
print("\n=== 第4行全部列（前50列）===")
for i, row in enumerate(ws.iter_rows(min_row=4, max_row=4, values_only=False)):
    vals = [c.value for c in row]
    for j, v in enumerate(vals[:50]):
        if v is not None:
            print(f"  列{j}: {repr(v)}")

wb.close()

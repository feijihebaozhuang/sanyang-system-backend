# -*- coding: utf-8 -*-
"""
从原始平台商品.xlsx直接做剩余操作：
1. 找出 777564404927 的"特硬;双插盒""特硬;扣底盒" → 确认在哪个分类
2. 其余商品中分出带小数点的(非全.5) → 非全量飞机盒.xlsx
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

src = r'D:\Desktop\平台商品.xlsx'
df = pd.read_excel(src, header=1, dtype=str)

# 找 777564404927
pid = '777564404927'
rows = df[df['平台商品id'].astype(str).str.strip() == pid]
print(f'777564404927 共 {len(rows)} 条规格:')
for _, r in rows.iterrows():
    print(f'  {r["平台规格名称"]}')

# 判断这2条是不是在"其余商品"里（不是纸箱/扣底盒/定制）
print('\n判断分类:')
for _, r in rows.iterrows():
    s = str(r['平台规格名称'])
    if '双插盒' in s or '扣底盒' in s:
        print(f'  "{s}" → 扣底盒双插盒文件')
    elif '纸箱' in s:
        print(f'  "{s}" → 纸箱文件')
    else:
        print(f'  "{s}" → 其余商品')

# ===== 直接读其余商品文件 =====
print('\n===== 直接处理其余商品 =====')
other_file = r'D:\Desktop\其余商品.xlsx'
if not os.path.exists(other_file):
    print('其余商品.xlsx 不存在，从原始分离...')
    # 从原始df过滤
    def is_custom(s):
        if any(kw in s for kw in ['定制','订制','定做','加工定制','不接受退货','咨询客服','拍下联系客服','定制产品','定制拍单','定制尺寸','万款现货','联系客服备注','详情咨询','更多尺寸','下拉查看','下拉-','1000款现模','更多尺寸看详情','详情-现模','1000个尺寸']): return True
        if '珍珠棉' in s: return True
        has_dim = bool(re.search(r'[\d.]+\s*[xX*×]\s*[\d.]+\s*[xX*×]\s*[\d.]+', s)) or bool(re.search(r'【\s*[\d.]+', s))
        if has_dim: return False
        nums = re.findall(r'[\d.]+', s)
        if len(nums) == 0: return True
        return False
    
    custom_rows = []
    kd_rows = []
    zx_rows = []
    other_rows = []
    
    for _, row in df.iterrows():
        s = str(row.get('平台规格名称', '')).strip()
        if not s: continue
        if is_custom(s):
            custom_rows.append(row)
        elif '扣底盒' in s or '扣抵盒' in s or '双插盒' in s:
            kd_rows.append(row)
        elif '纸箱' in s:
            zx_rows.append(row)
        else:
            other_rows.append(row)
    
    df_other = pd.DataFrame(other_rows, columns=df.columns)
    print(f'其余商品: {len(df_other)} 行')
else:
    # 用openpyxl直接读
    from openpyxl import load_workbook
    wb = load_workbook(other_file, read_only=True)
    ws = wb.active
    data = list(ws.values)
    if len(data) > 1:
        headers = data[1]  # 跳过标题行
        rows_data = [dict(zip(headers, row)) for row in data[2:]] if len(data) > 2 else []
        df_other = pd.DataFrame(rows_data)
        print(f'其余商品: {len(df_other)} 行, col: {list(df_other.columns[:3])}')
    wb.close()

# 分拆带小数点的
def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or '')); w = normalize_dim(m.group(3)+(m.group(4) or '')); h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

dot_rows = []
remain_rows = []

for idx, (_, row) in enumerate(df_other.iterrows()):
    s = str(row.get('规格名称', row.get('平台规格名称', ''))).strip()
    if not s:
        remain_rows.append(row)
        continue
    lwh = extract_lwh(s)
    if not lwh:
        remain_rows.append(row)
        continue
    l, w, h = lwh
    has_dot = any(n != int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    if has_dot and not all_dot5:
        dot_rows.append(row)
    else:
        remain_rows.append(row)

print(f'\n其余商品共{len(df_other)}条')
print(f'  带小数点(非全.5): {len(dot_rows)} 条 → 非全量飞机盒.xlsx')
print(f'  剩余: {len(remain_rows)} 条')

import openpyxl

# 输出非全量飞机盒
feiji_file = r'D:\Desktop\非全量飞机盒.xlsx'
COLS = list(df_other.columns)
pd.DataFrame(dot_rows, columns=COLS).to_excel(feiji_file, index=False)
wb = openpyxl.load_workbook(feiji_file)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '其余商品中带小数点的（已排除全.5）- 非全量飞机盒'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(feiji_file)
print(f'✅ 非全量飞机盒.xlsx: {len(dot_rows)}条')

# 更新其余商品
pd.DataFrame(remain_rows, columns=COLS).to_excel(other_file, index=False)
wb = openpyxl.load_workbook(other_file)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '其余商品（已移除小数项）'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(other_file)
print(f'✅ 其余商品已更新: {len(remain_rows)}条')

# 定制链接加那2条
# 先找那2条在扣底盒文件中的记录
kd_file = r'D:\Desktop\扣底盒双插盒商品.xlsx'
wb_kd = openpyxl.load_workbook(kd_file)
ws_kd = wb_kd.active
kd_data = list(ws_kd.values)
kd_headers = kd_data[1]  # 第二行是表头
found = []
new_kd = [kd_data[0], kd_data[1]]  # 保留标题行+表头行
for row in kd_data[2:]:
    rid = str(row[1] or '').strip() if len(row) > 1 else ''
    spec = str(row[2] or '').strip() if len(row) > 2 else ''
    if rid == pid and spec in ['特硬;双插盒', '特硬;扣底盒']:
        found.append(row)
    else:
        new_kd.append(row)

# 重写扣底盒文件
ws_kd.delete_rows(1, ws_kd.max_row)
for i, row in enumerate(new_kd, 1):
    for j, v in enumerate(row):
        ws_kd.cell(i, j+1, v)
wb_kd.save(kd_file)

# 加到定制链接
custom_file = r'D:\Desktop\定制链接商品.xlsx'
wb_ct = openpyxl.load_workbook(custom_file)
ws_ct = wb_ct.active
ct_data = list(ws_ct.values)
ct_headers = ct_data[1]  # 第二行是表头

for row in found:
    add_row = [row[0], row[1], row[2], row[3]]  # 店铺,id,规格,规格id
    ws_ct.append(add_row)

wb_ct.save(custom_file)

print(f'\n✅ 扣底盒: 移除了{len(found)}条')
print(f'✅ 定制链接: 新增了{len(found)}条')
print(f'✅ 非全量飞机盒.xlsx: {len(dot_rows)}条')

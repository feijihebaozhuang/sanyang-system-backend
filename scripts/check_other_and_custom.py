# -*- coding: utf-8 -*-
"""快速看其余商品和定制里各是什么"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

# 1. 其余商品前20条
f = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active
print('=== 其余商品前20条 ===')
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    count += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    print(f'  [{count}] {s[:130]}')
    if count >= 20: break
wb.close()
print()

# 2. 定制链接前20条
f2 = r'D:\Desktop\定制链接商品.xlsx'
wb2 = load_workbook(f2, read_only=True)
ws2 = wb2.active
print('=== 定制链接前20条 ===')
count = 0
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    count += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    has_dim_tag = '✅有LxWxH' if re.search(r'\d[\d.]*\s*[xX*]\s*\d[\d.]*\s*[xX*]\s*\d[\d.]*', s) else ''
    print(f'  [{count}] {has_dim_tag} {s[:130]}')
    if count >= 20: break
wb2.close()

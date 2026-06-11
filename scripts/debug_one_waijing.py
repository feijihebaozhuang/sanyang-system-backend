# -*- coding: utf-8 -*-
"""手动调试外径规格"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import re

# 取平台商品.xlsx中第一个外径规格
from openpyxl import load_workbook
from split_neijing_v6 import extract_all_dims, best_lwh, classify_dims, has_any_dim

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

waijing_s = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' in s:
        waijing_s = s
        break
wb.close()

print(f'规格: {waijing_s}')
print()

all_dims = extract_all_dims(waijing_s)
print(f'extract_all_dims: {all_dims}')
print(f'has_any_dim: {has_any_dim(waijing_s)}')
blwh = best_lwh(waijing_s)
print(f'best_lwh: {blwh}')
if blwh:
    print(f'classify_dims: {classify_dims(blwh)}')

print()

# 跟随v6代码的分支逻辑
s = waijing_s
if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
    print('→ 定制 (关键词+无尺寸)')
elif '扣底盒' in s or '双插盒' in s:
    print('→ 扣底盒/双插盒')
elif '纸箱' in s:
    print('→ 纸箱')
else:
    lwh = best_lwh(s)
    if lwh is None:
        print('→ 定制 (无LWH)')
    else:
        dt = classify_dims(lwh)
        print(f'维度类型: {dt}')
        if '内径' in s or '内尺寸' in s or '内寸' in s:
            print('→ 内径全量飞机盒')
        elif dt == '非全量飞机盒':
            print('→ 非全量飞机盒')
        elif '外径' in s:
            print('→ 外径全量飞机盒')
        else:
            print('→ 其余')

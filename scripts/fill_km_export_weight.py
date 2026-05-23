#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快麦「普通商品明细导出表」批量填 规格重量(千克) — 保持原模板 1~3 行表头，可直接回导快麦。

用法:
  python scripts/fill_km_export_weight.py "d:\\Desktop\\快麦商品\\ok\\已填充_规格别名_1_第1批.xlsx"
  python scripts/fill_km_export_weight.py "d:\\Desktop\\快麦商品\\ok\\*.xlsx" --out-dir "d:\\Desktop\\快麦商品\\ok\\已填重量"
"""
from __future__ import annotations

import argparse
import glob
import re
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import quote_weight as qw

HEADER_ROW = 3
DATA_START = 4
COL_MAIN_CODE = 0
COL_NAME = 1
COL_ALIAS = 28
COL_WEIGHT_KG = 38
COL_L = 39
COL_W = 40
COL_H = 41
COL_CAT1 = 73


def _float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        m = re.search(r"(\d+(?:\.\d+)?)", str(v))
        return float(m.group(1)) if m else 0.0


def _weight_empty(v) -> bool:
    if v is None or v == "":
        return True
    try:
        return float(v) <= 0
    except (TypeError, ValueError):
        return True


def _material_from_main_code(code: str) -> str:
    """10*10*10-内径-P6D → P6D；10*10*10-内径-特硬 → 特硬"""
    s = (code or "").strip()
    if not s:
        return ""
    parts = s.split("-")
    if len(parts) >= 2:
        tail = parts[-1].strip()
        if tail and tail not in ("内径", "外径"):
            return tail
    return ""


def _product_type_from_cat(cat: str) -> str:
    c = (cat or "").strip()
    if "纸箱" in c:
        return "qita"
    if "扣底" in c:
        return "koudi"
    if "双插" in c:
        return "shuangcha"
    if "珍珠棉" in c or "PE" in c.upper():
        return "pe"
    return ""


def _build_text(row: list) -> str:
    parts = [
        str(row[COL_MAIN_CODE] or ""),
        str(row[COL_NAME] or ""),
        str(row[COL_ALIAS] or ""),
        _material_from_main_code(str(row[COL_MAIN_CODE] or "")),
    ]
    if len(row) > COL_CAT1:
        parts.append(str(row[COL_CAT1] or ""))
    return " ".join(p for p in parts if p)


def _dims_from_row(row: list) -> tuple[float, float, float]:
    l = _float(row[COL_L] if len(row) > COL_L else 0)
    w = _float(row[COL_W] if len(row) > COL_W else 0)
    h = _float(row[COL_H] if len(row) > COL_H else 0)
    if l and w and h:
        return l, w, h
    text = _build_text(row)
    if text:
        import km_sku_map_store as kms

        pl, pw, ph, _ = kms.parse_spec_alias_dims(text)
        if pl and pw:
            return pl, pw, ph or h
    return l, w, h


def process_file(
    inp: Path,
    out: Path,
    *,
    overwrite: bool = False,
    limit: int = 0,
) -> dict[str, int]:
    from openpyxl import Workbook, load_workbook

    t0 = time.time()
    wb_in = load_workbook(inp, read_only=True, data_only=True)
    ws_in = wb_in.active

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet("Sheet1")
    ncol = 143
    stats = {"total": 0, "filled": 0, "skipped": 0, "failed": 0}

    for i, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
        cells = _pad_row(list(row or []), ncol)
        if i <= HEADER_ROW:
            ws_out.append(cells)
            if i == HEADER_ROW:
                ncol = max(ncol, len(cells))
            continue
        if limit and stats["total"] >= limit:
            break
        stats["total"] += 1
        if not overwrite and not _weight_empty(cells[COL_WEIGHT_KG]):
            stats["skipped"] += 1
            ws_out.append(cells)
            continue
        l, w, h = _dims_from_row(cells)
        text = _build_text(cells)
        ptype = _product_type_from_cat(str(cells[COL_CAT1] or "")) if len(cells) > COL_CAT1 else ""
        est = qw.estimate_unit_weight(l, w, h, text=text, product_type=ptype)
        if not est.get("ok"):
            stats["failed"] += 1
            ws_out.append(cells)
            continue
        cells[COL_WEIGHT_KG] = round(float(est["weight_per_unit_kg"]), 4)
        stats["filled"] += 1
        ws_out.append(cells)
        if stats["total"] % 10000 == 0:
            print(f"  ... {stats['total']} 行", flush=True)

    wb_in.close()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out)
    stats["seconds"] = int(time.time() - t0)
    return stats


def _pad_row(row: list, ncol: int) -> list:
    if len(row) < ncol:
        row = row + [None] * (ncol - len(row))
    return row[:ncol]


def main() -> None:
    ap = argparse.ArgumentParser(description="快麦商品导出表填规格重量(千克)")
    ap.add_argument("inputs", nargs="+", help="xlsx 路径，支持通配符")
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=None,
        help="输出目录（默认同目录，文件名加 _已填重量）",
    )
    ap.add_argument("--overwrite", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    paths: list[Path] = []
    for pat in args.inputs:
        g = glob.glob(pat)
        if g:
            paths.extend(Path(p) for p in g)
        else:
            paths.append(Path(pat))

    for inp in paths:
        if not inp.is_file():
            print(f"跳过（不存在）: {inp}")
            continue
        if args.out_dir:
            out = args.out_dir / (inp.stem + "_已填重量.xlsx")
        else:
            out = inp.with_name(inp.stem + "_已填重量.xlsx")
        print(f"处理: {inp.name}")
        if args.dry_run:
            from openpyxl import load_workbook

            wb = load_workbook(inp, read_only=True, data_only=True)
            n = sum(1 for _ in wb.active.iter_rows(min_row=DATA_START)) 
            wb.close()
            print(f"  (dry-run) 数据行约 {n}，将写出 {out}")
            continue
        st = process_file(inp, out, overwrite=args.overwrite, limit=args.limit)
        print(
            f"  完成: 填 {st['filled']} / {st['total']} 行, "
            f"跳过 {st['skipped']}, 无法算 {st['failed']}, "
            f"耗时 {st['seconds']}s → {out}"
        )


if __name__ == "__main__":
    main()

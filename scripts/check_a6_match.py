# -*- coding: utf-8 -*-
"""检查A6外径2+1格式匹配情况"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

total_waijing = 0
matched_a6 = 0
matched_any = 0
examples_no = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if '外径' not in s:
        continue
    total_waijing += 1
    
    # A6 测试
    m_a6 = re.search(r'外径[】\]]*\s*高度【\s*([\d.]+)\s*mm?\s*】.*?长\s*\*\s*宽【\s*([\d.]+)\s*\*\s*([\d.]+)\s*mm?\s*】', s)
    if m_a6:
        matched_a6 += 1
    else:
        # 检查其他格式
        m_other = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s) or \
                  re.search(r'宽【.*?高【.*?长【', s) or \
                  re.search(r'【长度.*?【宽度', s)
        if not m_other and len(examples_no) < 10:
            examples_no.append(s)

wb.close()
print(f'总外径规格: {total_waijing}')
print(f'A6匹配: {matched_a6}')
print(f'A6未匹配示例:')
for i, s in enumerate(examples_no):
    print(f'  [{i+1}] {s[:130]}')

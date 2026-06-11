# -*- coding: utf-8 -*-
"""
从其余商品.xlsx中分出带小数点的（排除全.5）
重点：商品规格是 "优质进口纸【100个】;51.5*25.5*4.5" 这种格式
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''; val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    
    # 外尺寸【LxWxH】+单位
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m and all(c for c in m.groups()[:3]):
        l = normalize_dim(m.group(1)+(m.group(4) or ''))
        w = normalize_dim(m.group(2)+(m.group(4) or ''))
        h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    
    # 【长Lx宽Wx高H】
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    
    # 长【L】宽【W】高【H】
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or '')); w = normalize_dim(m.group(3)+(m.group(4) or '')); h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    
    # 【LxWxH】
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m and all(c for c in m.groups()[:3]):
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    
    # 无括号 LxWxH（这是其余商品的主要格式！）
    # 规格如: "优质进口纸-黄色【100个】;51.5*25.5*4.5"
    # 用分号或逗号或空格后的 LxWxH
    m = re.search(r'(?:[；;，,]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m:
        try:
            l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
            if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
        except: pass
    
    # 也试试任意位置的 LxWxH（但前后不是数字）
    m = re.search(r'(?<!\d)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)(?!\d)', s)
    if m:
        try:
            l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
            if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
        except: pass
    
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

# 读其余商品
other_file = r'D:\Desktop\其余商品.xlsx'
wb = load_workbook(other_file, read_only=True)
ws = wb.active

header = None
feiji = []
remaining = []
total = 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if not s:
        remaining.append(row)
        continue
    
    lwh = extract_lwh(s)
    if not lwh:
        remaining.append(row)
        continue
    
    l, w, h = lwh
    has_dot = any(n != int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    
    if has_dot and not all_dot5:
        feiji.append(row)
    else:
        remaining.append(row)

wb.close()

print(f'其余商品总: {total}')
print(f'非全量飞机盒(带小数非全.5): {len(feiji)} 条')
print(f'剩余: {len(remaining)} 条')

feiji_file = r'D:\Desktop\非全量飞机盒.xlsx'
pd.DataFrame(feiji, columns=header).to_excel(feiji_file, index=False)
wb2 = opx.load_workbook(feiji_file)
ws2 = wb2.active
ws2.insert_rows(1)
ws2.cell(1, 2).value = '其余商品中带小数点的（已排除全.5）- 非全量飞机盒'
ws2.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws2.column_dimensions[c].width = 10
    except: pass
wb2.save(feiji_file)
print(f'✅ 非全量飞机盒.xlsx')

# 更新其余商品
pd.DataFrame(remaining, columns=header).to_excel(other_file, index=False)
wb3 = opx.load_workbook(other_file)
ws3 = wb3.active
ws3.insert_rows(1)
ws3.cell(1, 2).value = '其余商品（已移除小数项）'
ws3.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws3.column_dimensions[c].width = 10
    except: pass
wb3.save(other_file)
print(f'✅ 其余商品已更新')

print(f'\n验证: 256+7266+33214+{len(feiji)}+{len(remaining)} = {256+7266+33214+len(feiji)+len(remaining)}')

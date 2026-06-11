# -*- coding: utf-8 -*-
"""
v10: 加入天猫彩色的所有格式支持
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def to_cm(v_str):
    """提取数值并统一转成cm"""
    v_str = str(v_str).strip()
    v = float(re.sub(r'[^\d.]', '', v_str.replace(',', '.')))
    v_str_lower = v_str.lower()
    if 'mm' in v_str_lower:
        v = v / 10.0
    elif v_str_lower.endswith('m') and not v_str_lower.endswith('cm') and not v_str_lower.endswith('mm'):
        v = v * 100.0
    return v

def extract_all_dims(s):
    """全覆盖提取LWH"""
    results = []
    
    def clean_dot(v):
        parts = v.split('.')
        if len(parts) > 2:
            return parts[0] + '.' + ''.join(parts[1:])
        return v
    
    # ===== 1. 宽【W】高【H】;长【L】（通用）=====
    # 宽【39m】高【7cm】;【100个】长【39cm】
    # 宽【10cm】高【8cm】内径;【100个】长【38cm】
    # 宽【12cm】高【12cm】;【100个】长度【12cm】
    m = re.search(r'宽[\[【]\s*([\d.]+)\s*(\w*)\s*[\]】]'
                  r'.*?'
                  r'高[\[【]\s*([\d.]+)\s*(\w*).*?[\]】]'
                  r'.*?'
                  r'(?:长度?)[\[【]\s*([\d.]+)\s*(\w*).*?[\]】]', s)
    if m:
        results.append((to_cm(m.group(5)+m.group(6)), to_cm(m.group(1)+m.group(2)), to_cm(m.group(3)+m.group(4)), 'A1'))
    
    # ===== 2. 【宽度N厘米】【高N厘米】【长度N厘米】（新格式）=====
    # 【宽度12厘米】【高12厘米】;【数量100个】【长度12厘米】
    m = re.search(r'【宽度\s*([\d.]+)\s*(\w*)\s*】.*?【高\s*([\d.]+)\s*(\w*)\s*】.*?【长度\s*([\d.]+)\s*(\w*)\s*】', s)
    if m:
        results.append((to_cm(m.group(5)+m.group(6)), to_cm(m.group(1)+m.group(2)), to_cm(m.group(3)+m.group(4)), 'A2'))
    
    # ===== 3. 宽度NCM...长N*宽?*高N【内径/外径】（新格式）=====
    # 宽度 10 CM 特硬-牛皮色【100个】;长 11 *宽 ? * 高 6 【所量即所装=内径】
    # 白色【100个】宽度 10 CM;长 11 *宽 ? * 高 6 【所量即所装=内径】
    # 宽 30  CM 特硬-牛皮色【100个】;长 31 *宽  ?* 高 8 【所量即所装=内径】
    m = re.search(r'(?:宽度?)\s*([\d.]+)\s*(\w*).*?长\s*([\d.]+)\s*\*.*?(?:宽\s*[?？]\s*)\*.*?高\s*([\d.]+)', s)
    if m:
        results.append((to_cm(m.group(3)), to_cm(m.group(1)+m.group(2)), to_cm(m.group(4)), 'A3'))
    
    # ===== 4. 宽度：Ncm---高度：Ncm;长度：Ncm（新格式）=====
    # 宽度：10 cm---高度：10 cm;白色 100个 长度：10 cm
    m = re.search(r'宽度[：:]\s*([\d.]+)\s*(\w*)\s*---高度[：:]\s*([\d.]+)\s*(\w*).*?长度[：:]\s*([\d.]+)\s*(\w*)', s)
    if m:
        results.append((to_cm(m.group(5)+m.group(6)), to_cm(m.group(1)+m.group(2)), to_cm(m.group(3)+m.group(4)), 'A4'))
    
    # ===== 5. 长宽【L*W】cm;高度Hcm【扣底纸盒】（新格式）=====
    # 长宽【15*15】cm;高度10cm【扣底纸盒】100个组
    m = re.search(r'长宽[\[【]\s*([\d.]+)\s*\*\s*([\d.]+)\s*\]?\s*cm.*?高度\s*([\d.]+)\s*cm\s*【', s)
    if m:
        results.append((to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3)), 'A5扣底'))
    
    # ===== 6. 【内径/内尺寸/外尺寸】;【正方形飞机盒高度】NxNxN（新格式）=====
    # 白色【内径】100个;【正方形飞机盒 高度 10 厘米】10x10x10 cm
    # 黑色【内尺寸】【100个】;【正方形飞机盒 高度 10 厘米】10x10x10 cm
    # 黑色【外尺寸】【100个】;【正方形飞机盒 高度 10 厘米】10x10x10 cm
    m = re.search(r'【正方形飞机盒\s*高度\s*([\d.]+)\s*(\w*)\s*】\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*', s)
    if m:
        # 高度在前，后跟LxWxH
        h = to_cm(m.group(1)+m.group(2))
        l = to_cm(m.group(3))
        w = to_cm(m.group(4))
        results.append((l, w, h, 'A6'))
    
    # ===== 7. 外尺寸/内尺寸 + 【厘米/毫米】NxNxN（新格式）=====
    # 外尺寸【双面白色】100个;【厘米】10x10x5
    # 外尺寸【双面白色】100个;【毫米】100x100x30
    # 进口牛皮色【内尺寸】;【厘米】10x10x5
    m = re.search(r'【([厘毫]米)】\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)', s)
    if m:
        unit = m.group(1)
        l = float(m.group(2))
        w = float(m.group(3))
        h = float(m.group(4))
        if unit == '毫米':
            l, w, h = l/10, w/10, h/10
        results.append((l, w, h, 'A7'))
    
    # ===== 8. 外尺寸/内尺寸 + 长N宽N高N(cm)（新格式）=====
    # 外尺寸【双面白色】100个;长10宽10高10(cm)
    # 进口牛皮色【内尺寸】;长10宽10高10(cm)
    m = re.search(r'长\s*([\d.]+)\s*(?:cm|厘米|毫米|mm)?\s*宽\s*([\d.]+)\s*(?:cm|厘米|毫米|mm)?\s*高\s*([\d.]+)\s*', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A8'))
    
    # ===== 9. 宽度---Nmm【高N厘米】;长度---Nmm / 【长度N厘米】（新格式）=====
    # 宽度---100mm【高5厘米】;长度---100mm【数量100个】
    # 宽度---100mm【高4厘米】;【数量100个】【长度10厘米】
    # 【宽度10厘米】【高6厘米】;长度---100mm【数量100个】
    m = re.search(r'(?:宽度[--]?\s*)?([\d.]+)\s*mm\s*(?:\[?【?[^【】]*?)?高\s*([\d.]+)\s*厘米', s)
    if m:
        w_mm = float(m.group(1))
        h_cm = float(m.group(2))
        # 找长度
        m_l = re.search(r'长度[--]?\s*([\d.]+)\s*(mm|厘米)?', s)
        if m_l:
            lv = float(m_l.group(1))
            if m_l.group(2) == 'mm':
                lv = lv / 10.0
            results.append((lv, w_mm/10.0, h_cm, 'A9'))
        # 【宽度N厘米】【高N厘米】;长度---Nmm【数量N个】的变体
        m_w = re.search(r'【宽度\s*([\d.]+)\s*厘米\s*】.*?【高\s*([\d.]+)\s*厘米\s*】.*?长度[--]?\s*([\d.]+)\s*mm', s)
        if m_w:
            results.append((float(m_w.group(3))/10, float(m_w.group(1)), float(m_w.group(2)), 'A9b'))
    
    # ===== 10. 【长度 L 厘米】【宽度 W 厘米】（中文"厘米"）=====
    m = re.search(r'【长度\s*([\d.]+)\s*厘米\s*】.*?【宽度\s*([\d.]+)\s*厘米\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), 0, 'A10'))
    
    # ===== 11. 长【L】宽【W】高【H】=====
    m = re.search(r'长[\[【]\s*([\d.]+)\s*(\w*)\s*[\]】].*?宽[\[【]\s*([\d.]+)\s*(\w*)\s*[\]】].*?高[\[【]\s*([\d.]+)\s*(\w*)\s*[\]】]', s)
    if m: results.append((to_cm(m.group(1)+m.group(2)), to_cm(m.group(3)+m.group(4)), to_cm(m.group(5)+m.group(6)), 'A11'))
    
    # ===== 12. 飞机盒【长度L】...【宽度W】X【高度H】=====
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A12'))
    
    # ===== 13. 外径 130x100x20 mm（标准3值）=====
    m = re.search(r'外径[：:]*\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)(?:\s*mm)?', s)
    if m:
        v = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        if any(x > 50 for x in v):
            v = [x/10 for x in v]
        results.append((v[0], v[1], v[2], 'A13'))
    
    # ===== 14. 外径+高度+长*宽（2+1格式）=====
    m = re.search(r'(?:外径[^；;]*高度[\[【]\s*([\d.]+)\s*mm?\s*[\]】]|高度[\[【]\s*([\d.]+)\s*mm?\s*[\]】][^；;]*外径).*?长\s*\*\s*宽[\[【]?\s*([\d.]+)\s*\*\s*([\d.]+)\s*mm?\s*[\]】]?', s)
    if m:
        h_val = m.group(1) if m.group(1) else m.group(2)
        h = float(h_val) / 10.0
        l = float(m.group(3)) / 10.0
        w = float(m.group(4)) / 10.0
        results.append((l, w, h, 'A14'))
    
    # ===== 15. 高度 + 内/外径 + 长*宽【L*W】mm（顺序不同）=====
    # 高度【20mm】【内径】黄色;【100个】长*宽【100*60】mm
    # 高度【20mm】白色【内径】;【100个】长*宽【100*60】mm
    m = re.search(r'高度[\[【]\s*(\d+\.?\d*)\s*mm?\s*[\]】].*?[内外]径.*?长\s*\*\s*宽[\[【]\s*(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)\s*[\]】]?\s*mm', s)
    if m:
        h = float(m.group(1)) / 10.0
        l = float(m.group(2)) / 10.0
        w = float(m.group(3)) / 10.0
        results.append((l, w, h, 'A15'))
    
    # ===== 16. 宽--宽格式（如"37宽--38宽"）=====
    # 白色内径】高度【100mm】;37宽----------------------------------------------38宽
    # 取较大的宽，高度来自【高度】
    m_h = re.search(r'高度[\[【]\s*(\d+\.?\d*)\s*mm', s)
    m_wr = re.search(r'(\d+\.?\d*)\s*宽[^；;]*[-\u2500-\u27ff]{5,}\s*(\d+\.?\d*)\s*宽', s)
    if m_h and m_wr:
        h = float(m_h.group(1)) / 10.0
        w2 = float(m_wr.group(2))
        l2 = w2  # 无长度则默认正方形
        results.append((l2, w2, h, 'A16宽宽'))
        # 如果也有明确的长度，后续build_lwh会选更好的
    
    # ===== 17. 【正方形飞机盒】N*N*Ncm（不带"高度"关键词）=====
    # 特硬【内尺寸】单个;【正方形飞机盒】10*10*2 cm
    m = re.search(r'【正方形飞机盒】\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*', s)
    if m:
        results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A17'))
    
    # ===== 15. 【LxWxH】=====
    m = re.search(r'【\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B1'))
    
    # ===== 16. ;LxWxH在末尾（排除多点号的异常值，如58..7）=====
    m = re.search(r'(?:[；;]\s*|^)([\d.]{1,10})\s*[xX*]\s*([\d.]{1,10})\s*[xX*]\s*([\d.]{1,10})\s*$', s)
    if m:
        def clean_dot(v):
            parts = v.split('.')
            if len(parts) > 2:
                return parts[0] + '.' + ''.join(parts[1:])
            return v
        results.append((float(clean_dot(m.group(1))), float(clean_dot(m.group(2))), float(clean_dot(m.group(3))), 'B3'))
    
    # ===== 17. 裸LxWxH =====
    m = re.search(r'(?:^|[\s；;])([\d.]{1,10})\s*[xX*]\s*([\d.]{1,10})\s*[xX*]\s*([\d.]{1,10})', s)
    if m: results.append((float(clean_dot(m.group(1))), float(clean_dot(m.group(2))), float(clean_dot(m.group(3))), 'B4'))
    
    # ===== 18. 含双点号的特殊格式: 58..7*27.5*6.5 =====
    m = re.search(r'(?:[；;]\s*|^)([\d.]{1,10})(?:\s*cm)?\s*[xX*]\s*([\d.]{1,10})(?:\s*cm)?\s*[xX*]\s*([\d.]{1,10})', s)
    if m: results.append((float(clean_dot(m.group(1))), float(clean_dot(m.group(2))), float(clean_dot(m.group(3))), 'B5'))
    
    # ===== 19. 两个星号: 37*14**3.8 或 22.5**14.4*7.9 =====
    m = re.search(r'(?:[；;]\s*|^)([\d.]{1,10})\s*\*+\s*([\d.]{1,10})\s*\*+\s*([\d.]{1,10})', s)
    if m: results.append((float(clean_dot(m.group(1))), float(clean_dot(m.group(2))), float(clean_dot(m.group(3))), 'B6'))
    
    return results

def build_lwh(s):
    all_dims = extract_all_dims(s)
    if not all_dims: return None
    for l, w, h, src in all_dims:
        if h > 0.1 and 0.5 <= l <= 200 and 0.5 <= w <= 200 and 0.5 <= h <= 200:
            return (l, w, h)
    for l, w, h, src in all_dims:
        if 0.5 <= l <= 200 and 0.5 <= w <= 200:
            return (l, w, h)
    if all_dims:
        l, w, h, src = all_dims[0]
        return (l, w, h)
    return None

def has_any_dim(s):
    return len(extract_all_dims(s)) > 0

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量飞机盒'
    elif all_5: return '全.5'
    else: return '整数'

# === 判断含黑&红 ===
def has_black_red(s):
    """黑和红同时出现（不同颜色组合在同一规格中）"""
    return '黑' in s and '红' in s

# === 全量分类 ===
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
waijing_rows = []
feiji_rows = []
other_rows = []

for idx, row in enumerate(rows_list):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 0. 黑&红同时出现 -> 定制
    if has_black_red(s):
        custom_rows.append(row)
        continue
    
    # 1. 定制关键词 + 无尺寸
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
        custom_rows.append(row)
        continue
    
    # 2. 扣底盒/双插盒
    if '扣底盒' in s or '双插盒' in s:
        dikoudi_rows.append(row)
        continue
    
    # 3. 扣底纸盒 → 扣底盒
    if '扣底纸盒' in s:
        dikoudi_rows.append(row)
        continue
    
    # 4. 纸箱
    if '纸箱' in s:
        zhixiang_rows.append(row)
        continue
    
    # 5. 提取尺寸
    lwh = build_lwh(s)
    if lwh is None:
        custom_rows.append(row)
        continue
    
    dt = classify_dims(lwh)
    
    # 6. 内径
    if '内径' in s or '内尺寸' in s or '内寸' in s or '所量即所装=内径' in s:
        neijing_rows.append(row)
        continue
    
    # 7. 非全量飞机盒
    if dt == '非全量飞机盒':
        feiji_rows.append(row)
        continue
    
    # 8. 外径全量（明确外径 + 没写内/外的默认为外径）
    if '外径' in s or '外尺寸' in s or ('内径' not in s and '内尺寸' not in s and '内寸' not in s and '所量即所装=内径' not in s):
        waijing_rows.append(row)
        continue
    
    # 9. 其余（不应有）
    other_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
print(f'其余: {len(other_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)+len(other_rows)
print(f'和: {s} (应={total})')
if s != total:
    print(f'!!! 差值: {total - s}')

def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(waijing_rows, r'D:\Desktop\外径全量飞机盒.xlsx', '外径全量飞机盒')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
# write_xlsx(其他, ...)

print('\n✅ 全部完成！')

# -*- coding: utf-8 -*-
"""重建平卡v3 - 读取OK文件夹所有已处理文件"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')
raw_file = r'd:\Desktop\平台商品.xlsx'

# 1. 收集所有已处理spec_id - 包括OK文件夹所有文件
processed_specs = set()
file_stats = {}

# OK文件夹所有文件
for f in sorted(os.listdir(ok_dir)):
    if not f.endswith('.xlsx'):
        continue
    fp = os.path.join(ok_dir, f)
    try:
        wb = oxl.load_workbook(fp)
        ws = wb[wb.sheetnames[0]]
        cnt = 0
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r and len(r) >= 3 and r[2]:
                processed_specs.add(str(r[2]).strip())
                cnt += 1
        wb.close()
        file_stats[f] = cnt
    except: pass

# 无匹配_待处理
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) >= 3 and r[2]:
        processed_specs.add(str(r[2]).strip())
wb.close()

print('OK文件夹处理文件:')
for f, c in sorted(file_stats.items()):
    print(f'  {f}: {c}条')
print(f'无匹配_待处理: spec_id数待统计')
print(f'\n去重后已处理spec_id: {len(processed_specs)}')

# 2. 从原始数据找未处理的
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
new_rows = []
shop_new = Counter()

for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    shop = str(r[0]).strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if spec_id not in processed_specs:
        shop_new[shop] += 1
        new_rows.append([shop, pid, spec_name, spec_id, '平卡/待处理', ''])

wb.close()

print(f'\n=== 平卡新分布 ===')
for shop in sorted(shop_new.keys()):
    print(f'  {shop}: {shop_new[shop]}条')

# 3. 写入平卡
wb = oxl.Workbook()
ws = wb.active
ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for row in new_rows:
    ws.append(row)
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()

print(f'\n✅ 平卡_待处理.xlsx 已重建: {len(new_rows)}条')
print(f'\n6个阿里店铺:')
ali = ['阿里友尚', '阿里亚润', '阿里三羊', '阿里大鱼', '阿里正方形', '阿里新鑫星']
for s in ali:
    print(f'  {s}: {shop_new.get(s, 0)}条')

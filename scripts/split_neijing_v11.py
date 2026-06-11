# -*- coding: utf-8 -*-
"""
v11: 简单粗暴法
1. 直接用正则提取所有可能的 LxWxH（不管什么括号、什么分隔符）
2. 按关键词归类：外径、内径、扣底盒、纸箱
3. 没写内外的默认为外径
4. 定制 = 没有三值 + 定制关键词
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def clean_num(s):
    """清理数值字符串：去掉多余的点号"""
    s = s.strip()
    # 58..7 → 58.7
    parts = s.split('.')
    if len(parts) > 2:
        return parts[0] + '.' + ''.join(parts[1:])
    return s

def to_cm(v_str):
    """数值转cm"""
    if not v_str: return None
    v_str = clean_num(v_str)
    try:
        v = float(v_str)
    except:
        return None
    return v

def extract_all_lwh(s):
    """从规格中提取所有可能的(L,W,H)三值组合，返回列表"""
    results = []
    
    # ===== A. 独立维度格式（宽+高+长）=====
    # 宽【10cm】高【8cm】;【100个】长【38cm】
    # 宽度---100mm【高5厘米】;长度---100mm
    # 【宽度12厘米】【高12厘米】;【数量100个】【长度12厘米】
    # 宽度：10 cm---高度：10 cm;白色 100个 长度：10 cm
    patterns = [
        # 宽+高+长（各种写法）
        r'(?:宽[度]?\s*[：:【\[]?\s*)([\d.]+)\s*(?:cm|mm|厘米|m)?\s*[】\]）]?\s*(?:高[度]?\s*[：:【\[]?\s*)([\d.]+)\s*(?:cm|mm|厘米|m)?\s*[】\]）]?.*?(?:长[度]?\s*[：:【\[]?\s*)([\d.]+)\s*(?:cm|mm|厘米|m)?',
        # 宽在前高在前，长在【】中的各种变体
        r'(?:宽[度]?\s*[：:【\[]?\s*)([\d.]+).*?(?:高[度]?\s*[：:【\[]?\s*)([\d.]+).*?(?:长[度]?\s*[：:【\[]?\s*)([\d.]+)',
    ]
    
    for pat in patterns:
        m = re.search(pat, s, re.IGNORECASE | re.DOTALL)
        if m:
            try:
                l, w, h = to_cm(m.group(3)), to_cm(m.group(1)), to_cm(m.group(2))
                if l and w and h:
                    results.append((l, w, h, 'A独立'))
            except:
                pass
    
    # ===== B. 高度在前，长宽在后的组合（内径/外径2+1格式）=====
    # 白色内径】高度【100mm】;长*宽【320*310mm】
    # 高度【100mm】黄色【内径】;【100个】长*宽【320*310mm】
    #  高度是单独值，长宽是乘积
    m = re.search(r'(?:高[度]?\s*[：:【\[]?\s*)([\d.]+)\s*(?:mm)', s)
    if m:
        h_val = float(clean_num(m.group(1))) / 10.0
        # 找长*宽
        m_lw = re.search(r'(?:长\s*\*\s*宽)?[\[【]?\s*([\d.]+)\s*\*\s*([\d.]+)\s*(?:mm)?', s)
        if m_lw:
            try:
                l = float(clean_num(m_lw.group(1))) / 10.0
                w = float(clean_num(m_lw.group(2))) / 10.0
                results.append((l, w, h_val, 'B2+1mm'))
            except:
                pass
    
    # ===== C. 宽--宽格式（宽度范围）=====
    m_h = re.search(r'(?:高[度]?\s*[：:【\[]?\s*)([\d.]+)\s*(?:mm)', s)
    m_w = re.search(r'([\d.]+)\s*宽[^；;]{2,}([\d.]+)\s*宽', s)
    if m_h and m_w:
        try:
            h = float(clean_num(m_h.group(1))) / 10.0
            w = float(clean_num(m_w.group(2)))
            results.append((w, w, h, 'C宽宽'))
        except:
            pass
    
    # ===== D. 【LxWxH】或 L*W*H 各种三值乘积格式 =====
    # 10x10x2, 10*10*2, 100×100×100, 58..7*27.5*6.5, 22.5**14.4*7.9
    # 37*14**3.8, 160×100×100, 39*.3*8.2*6.3
    
    # D1: 括号内的 LxWxH
    m = re.search(r'[\[【]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[\]】]', s)
    if m:
        try:
            l, w, h = to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3))
            if l and w and h: results.append((l, w, h, 'D1【】'))
        except:
            pass
    
    # D2: 分号后或末尾的三值乘积（58..7*27.5*6.5）
    m = re.search(r'(?:[；;]\s*|^)([\d.]+)\s*[*×xX]\s*([\d.]+)\s*[*×xX]\s*([\d.]+)', s)
    if m:
        try:
            l, w, h = to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3))
            if l and w and h: results.append((l, w, h, 'D2裸乘'))
        except:
            pass
    
    # D3: 双星号 37*14**3.8 → 37 14 3.8
    m = re.search(r'([\d.]+)\s*\*+\s*([\d.]+)\s*\*+\s*([\d.]+)', s)
    if m:
        try:
            l, w, h = to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3))
            if l and w and h: results.append((l, w, h, 'D3双星'))
        except:
            pass
    
    # D4: cm在末尾的 LxWxH cm
    m = re.search(r'(?:[；;]\s*|^)([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*cm', s)
    if m:
        try:
            results.append((float(clean_num(m.group(1))), float(clean_num(m.group(2))), float(clean_num(m.group(3))), 'D4cm'))
        except:
            pass
    
    # D5: 外径:130x100x20 mm
    m = re.search(r'外径[：:]*\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)', s)
    if m:
        try:
            v = [float(clean_num(m.group(1))), float(clean_num(m.group(2))), float(clean_num(m.group(3)))]
            if any(x > 50 for x in v):
                v = [x/10 for x in v]
            results.append((v[0], v[1], v[2], 'D5外径'))
        except:
            pass
    
    # D6: 厘米/毫米格式 【厘米】10x10x5 / 【毫米】100x100x30
    m = re.search(r'【[厘毫]米】\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)', s)
    if m:
        try:
            v = [float(clean_num(m.group(1))), float(clean_num(m.group(2))), float(clean_num(m.group(3)))]
            if '毫' in m.group(0):
                v = [x/10 for x in v]
            results.append((v[0], v[1], v[2], 'D6单位'))
        except:
            pass
    
    # D7: 长宽高10*10*2cm / 长10宽10高10(cm)
    m = re.search(r'(?:长[度]?\s*)?([\d.]+)\s*(?:cm|厘米)?\s*(?:宽[度]?\s*)?([\d.]+)\s*(?:cm|厘米)?\s*(?:高[度]?\s*)?([\d.]+)\s*(?:cm|厘米|mm)?', s)
    if m and '长' in s and '宽' in s and '高' in s:
        try:
            results.append((float(clean_num(m.group(1))), float(clean_num(m.group(2))), float(clean_num(m.group(3))), 'D7长宽高'))
        except:
            pass
    
    return results

def best_lwh(s):
    """取最佳的LWH"""
    all_dims = extract_all_lwh(s)
    if not all_dims:
        return None
    
    # 优先完整3值且合理
    for l, w, h, src in all_dims:
        if h > 0.1 and 0.5 <= l <= 150 and 0.5 <= w <= 150 and 0.5 <= h <= 150:
            return (l, w, h)
    
    # 降级
    for l, w, h, src in all_dims:
        if 0.5 <= l <= 150 and 0.5 <= w <= 150:
            return (l, w, h)
    
    if all_dims:
        l, w, h, src = all_dims[0]
        return (l, w, h)
    return None

def has_black_red(s):
    return '黑' in s and '红' in s

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量'
    elif all_5: return '全.5'
    else: return '整数'

# === 检查定制关键词 ===
CUSTOM_KEYWORDS = ['定制', '珍珠棉', '咨询客服', '订做专拍', '不接受退货', '不接受退换', '接拍下联系客服']

# === 全量分类 ===
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
waijing_rows = []
feiji_rows = []

for idx, row in enumerate(rows_list):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 黑&红同时出现 → 定制
    if has_black_red(s):
        custom_rows.append(row)
        continue
    
    # 定制关键词检查
    is_custom_keyword = any(kw in s for kw in CUSTOM_KEYWORDS)
    lwh = best_lwh(s)
    
    if is_custom_keyword and lwh is None:
        custom_rows.append(row)
        continue
    
    # 扣底盒/双插盒/扣底纸盒
    if '扣底盒' in s or '双插盒' in s or '扣底纸盒' in s:
        if lwh: dikoudi_rows.append(row)
        else: custom_rows.append(row)
        continue
    
    # 纸箱（含五层）
    if '纸箱' in s:
        if lwh: zhixiang_rows.append(row)
        else: custom_rows.append(row)
        continue
    
    # 无LWH → 定制
    if lwh is None:
        custom_rows.append(row)
        continue
    
    dt = classify_dims(lwh)
    
    # 内径
    if '内径' in s or '内尺寸' in s or '内寸' in s or '所量即所装=内径' in s:
        neijing_rows.append(row)
        continue
    
    # 非全量
    if dt == '非全量':
        feiji_rows.append(row)
        continue
    
    # 外径（包含明确外径 + 无内外关键词默认为外径）
    if '外径' in s or '外尺寸' in s:
        waijing_rows.append(row)
        continue
    
    # 无内外关键词，默认为外径
    waijing_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)
print(f'和: {s} (应={total})')
if s != total:
    print(f'!!! 差值: {total - s}')

def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(waijing_rows, r'D:\Desktop\外径全量飞机盒.xlsx', '外径全量飞机盒')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')

print('\n✅ 全部完成！')

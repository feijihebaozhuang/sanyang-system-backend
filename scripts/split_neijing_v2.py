# -*- coding: utf-8 -*-
"""
放宽条件: 包含'内径'且包含LxWxH的规格都算
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd
import re

other_file = r'D:\Desktop\其余商品.xlsx'

# 先读原文件（原其余商品在上一步已被修改，需要从上一版来？去桌面找备份）
# 先找原版其余商品——用 平台商品.xlsx 回溯是最准的
source = r'D:\Desktop\平台商品.xlsx'

print('从平台商品.xlsx 重新分……')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
custom_idxs = set()
dikoudi_idxs = set()
zhixiang_idxs = set()
feiji16835_idxs = set()
rows_all = []
total = 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    pid = row[4] if len(row) > 4 else None
    idx = i - 2
    rows_all.append({'idx': idx, 'row': row, 's': s})

# === 分类（与之前一致的逻辑）===
import re as re2

def has_dims(s):
    """判断规格是否包含可提取的长宽高"""
    # 标准格式
    if re2.search(r'(?:【|长\s*[：:]?\s*)(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s):
        return True
    # 尾部分号后格式
    if re2.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s):
        return True
    return False

for d in rows_all:
    s = d['s']
    idx = d['idx']
    row = d['row']
    if '定制' in s or '珍珠棉' in s or '咨询客服' in s or ('特硬;双插盒' in s and not has_dims(s)):
        custom_idxs.add(idx)
        continue
    if not has_dims(s):
        custom_idxs.add(idx)
        continue
    # 扣底盒 / 双插盒
    if '扣底盒' in s or '双插盒' in s:
        dikoudi_idxs.add(idx)
        continue
    # 纸箱
    if '纸箱' in s or '纸箱' in s:
        zhixiang_idxs.add(idx)
        continue
    # 非全量飞机盒（含小数非全.5）- 复用原判断
    lwh = None
    m2 = re2.search(r'(?:【|长\s*[：:]?\s*)(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s, re2.I)
    m3 = re2.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m2:
        lwh = [float(m2.group(1)), float(m2.group(2)), float(m2.group(3))]
    elif m3:
        lwh = [float(m3.group(1)), float(m3.group(2)), float(m3.group(3))]
    if lwh:
        if any(v != int(v) or v != round(v, 0) for v in lwh):
            if not all(v % 1 == 0.5 for v in lwh):
                feiji16835_idxs.add(idx)
                continue

# 现在分内径: 从上一步剩下的"其余"中挑
other_remaining_idxs = set(d['idx'] for d in rows_all) - custom_idxs - dikoudi_idxs - zhixiang_idxs - feiji16835_idxs

# 判断是否内径飞机盒
def is_neijing(s):
    """宽【Wcm】高【Hcm】内径;【N个】长【Lcm】 或 含内径且有LxWxH"""
    s_lower = s.lower()
    if '内径' not in s:
        return False
    # 看是否包含WxH内径 + L
    # 格式1: 宽【Wcm】高【Hcm】内径;【N个】长【Lcm】
    if '宽【' in s and '高【' in s and '内径' in s and '长【' in s:
        return True
    # 格式2: 含内径且有可提取的LxWxH
    if has_dims(s):
        return True
    # 格式3: 内尺寸 / 内寸
    if '内尺寸' in s or '内寸' in s:
        return True
    return False

neijing_idxs = set()
other_final_idxs = set()

for idx in other_remaining_idxs:
    s = rows_all[idx]['s']
    if is_neijing(s):
        neijing_idxs.add(idx)
    else:
        other_final_idxs.add(idx)

print(f'总数: {total}')
print(f'定制: {len(custom_idxs)}')
print(f'扣底盒/双插盒: {len(dikoudi_idxs)}')
print(f'纸箱: {len(zhixiang_idxs)}')
print(f'非全量飞机盒: {len(feiji16835_idxs)}')
print(f'内径全量飞机盒: {len(neijing_idxs)}')
print(f'其余: {len(other_final_idxs)}')
s = len(custom_idxs)+len(dikoudi_idxs)+len(zhixiang_idxs)+len(feiji16835_idxs)+len(neijing_idxs)+len(other_final_idxs)
print(f'和: {s}')

# 输出
def write_xlsx(idxs, fpath, label, col_c=60):
    data = [rows_all[idx]['row'] for idx in sorted(idxs)]
    if not data:
        print(f'{label}: 空，不写')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = col_c
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'✅ {fpath}')

import openpyxl as opx

write_xlsx(custom_idxs, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_idxs, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_idxs, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(feiji16835_idxs, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
write_xlsx(neijing_idxs, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(other_final_idxs, r'D:\Desktop\其余商品.xlsx', '其余商品')

print('\n完成！')

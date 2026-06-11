# -*- coding: utf-8 -*-
"""
从平台商品.xlsx 读取全部数据。
排除已分类的结构（1/2/3/4/4a 的商品pid），
剩下的商品提取长宽高数值，按尺寸分3类：
- 外径飞机盒：三项同时为整数（含外径/外尺寸关键字，或沉默=外径）
- 内径飞机盒：三项同时为.5（含内径/内尺寸关键字，或沉默=外径）
- 非全量飞机盒：除去以上的小数

mm一律转为cm。
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import pandas as pd
import openpyxl as opx

# ===== 1. 收集已分类结构的pid =====
EXCLUDED_PIDS = set()
files_to_exclude = [
    r'D:\Desktop\1-定制商品结构.txt',
    r'D:\Desktop\2-扣底盒商品结构.txt',
    r'D:\Desktop\3-双插盒商品结构.txt',
    r'D:\Desktop\4-三层纸箱商品结构.txt',
    r'D:\Desktop\4a-五层纸箱商品结构.txt',
]
for fpath in files_to_exclude:
    with open(fpath, 'r', encoding='utf-8') as f:
        for line in f:
            m = re.match(r'\s*pid=(.+)', line)
            if m:
                EXCLUDED_PIDS.add(m.group(1).strip())

# 还要排除6-不属于定制的60个结构.txt里的pid
with open(r'D:\Desktop\6-不属于定制的60个结构.txt', 'r', encoding='utf-8') as f:
    for line in f:
        m = re.match(r'\s*pid=(.+)', line)
        if m:
            EXCLUDED_PIDS.add(m.group(1).strip())

print(f'已分类的pid数: {len(EXCLUDED_PIDS)}')

# ===== 2. 读取平台商品.xlsx =====
source = r'D:\Desktop\平台和快麦原始商品\原.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
all_rows = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    all_rows.append(row)
wb.close()
print(f'总条数: {total}')

# ===== 3. 过滤出未分类的商品 =====
remain_rows = []
remain_count = 0
for row in all_rows:
    pid = str(row[3] or '') if len(row) > 3 else ''
    if pid not in EXCLUDED_PIDS:
        remain_rows.append(row)
        remain_count += 1
    else:
        pass

print(f'未分类商品: {remain_count} 条')
print(f'已分类商品: {total - remain_count} 条')

# ===== 4. 提取长宽高 =====
def clean_num(v):
    """处理 58..7 这类问题"""
    v = str(v).strip()
    parts = v.split('.')
    if len(parts) > 2:
        return parts[0] + '.' + ''.join(parts[1:])
    return v

def to_cm(v_str):
    """把任何数值转成cm"""
    v_str = str(v_str).strip().lower()
    unit = ''
    # 提取单位和数值
    m = re.match(r'([\d.]+)\s*(mm|cm|m|厘米|毫米|米)?', v_str)
    if not m:
        return None
    v = m.group(1)
    unit = m.group(2) or ''
    
    try:
        v = float(clean_num(v))
    except:
        return None
    
    if unit in ('mm', '毫米'):
        v = v / 10.0
    elif unit in ('m', '米'):
        v = v * 100
    # cm/厘米 或没有单位就默认cm
    
    return round(v, 2)

def extract_lwh(s):
    """从规格名称提取长宽高，返回 (l, w, h) 或 (None, None, None)"""
    l = w = h = None
    
    # === 长宽高三字分别提取 ===
    # 长
    m = re.search(r'长[度]?\s*[：:＝=]?\s*([\d.]+(?:\s*(?:cm|mm|m|厘米|毫米|米))?)', s)
    if m: l = to_cm(m.group(1))
    
    # 宽
    m = re.search(r'宽[度]?\s*[：:＝=]?\s*([\d.]+(?:\s*(?:cm|mm|m|厘米|毫米|米))?)', s)
    if m: w = to_cm(m.group(1))
    
    # 高
    m = re.search(r'高[度]?\s*[：:＝=]?\s*([\d.]+(?:\s*(?:cm|mm|m|厘米|毫米|米))?)', s)
    if m: h = to_cm(m.group(1))
    
    if l and w and h:
        return (l, w, h)
    
    # === 长宽在一起 + 高单独 ===
    # 长宽N*N  或  长宽NxN
    m = re.search(r'长宽\s*[【\[]?\s*([\d.]+)\s*[×*xX]\s*([\d.]+)', s)
    if m and not (l and w):
        l = to_cm(m.group(1))
        w = to_cm(m.group(2))
        if not h:
            m2 = re.search(r'高[度]?\s*[：:＝=]?\s*([\d.]+(?:\s*(?:cm|mm|m)?))', s)
            if m2: h = to_cm(m2.group(1))
    
    # 高 + 长宽 
    if not (l and w):
        m = re.search(r'高[度]?\s*[【\[]?\s*([\d.]+)', s)
        if m and not h: h = to_cm(m.group(1))
        m = re.search(r'[（(]?\s*长宽\s*[【\[]?\s*([\d.]+)\s*[×*xX]\s*([\d.]+)', s)
        if m and not (l and w):
            l = to_cm(m.group(1))
            w = to_cm(m.group(2))
    
    # === 拆解格式 N*N*N / NxNxN ===
    if not (l and w and h):
        # 找3个数值用乘号连接的
        m = re.search(r'([\d.]+)\s*[×*xX]\s*([\d.]+)\s*[×*xX]\s*([\d.]+)', s)
        if m:
            l, w, h = to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3))
    
    # === 长宽高连在一起（如 长宽高N*N*N）===
    if not (l and w and h):
        m = re.search(r'长宽高\s*[【\[]?\s*([\d.]+)\s*[×*xX]\s*([\d.]+)\s*[×*xX]\s*([\d.]+)', s)
        if m:
            l, w, h = to_cm(m.group(1)), to_cm(m.group(2)), to_cm(m.group(3))
    
    # === NxN cm【长宽】;高度Ncm ===
    if not (l and w and h):
        m = re.search(r'([\d.]+)\s*[×*xX]\s*([\d.]+)\s*cm?\s*【长宽】', s)
        if m:
            l = to_cm(m.group(1))
            w = to_cm(m.group(2))
            m2 = re.search(r'高度?\s*([\d.]+)\s*cm', s)
            if m2: h = to_cm(m2.group(1))
    
    # === 宽度 长度 高度 ===
    if not (l and w and h):
        m = re.search(r'宽度?\s*[：:＝=]?\s*([\d.]+)\s*cm?', s)
        if m and not w: w = to_cm(m.group(1))
        m = re.search(r'长度?\s*[：:＝=]?\s*([\d.]+)\s*cm?', s)
        if m and not l: l = to_cm(m.group(1))
        m = re.search(r'高度?\s*[：:＝=]?\s*([\d.]+)\s*cm?', s)
        if m and not h: h = to_cm(m.group(1))
    
    # === 过滤不合理值 ===
    if l and w and h:
        # 过滤太大或太小
        if l > 200 or w > 200 or h > 200:
            return (None, None, None)
        if l < 0.1 or w < 0.1 or h < 0.1:
            return (None, None, None)
        return (l, w, h)
    
    return (None, None, None)

# ===== 5. 判断内外径 =====
def is_waijing(s):
    """默认为外径，除非明确写了内径/内尺寸"""
    s_lower = s.lower()
    if '内径' in s or '内尺寸' in s or '内' in s.split('尺')[0] if '尺' in s else False:
        # 但也要看有没有明确的外径
        if '外径' in s or '外尺寸' in s:
            return True  # 外径优先
        return False
    return True  # 沉默=外径

# ===== 6. 分类 =====
waijing_rows = []  # 外径全量（整数）
neijing_rows = []  # 内径全量（.5）
feiquan_rows = []  # 非全量（其他小数）
no_dim_rows = []   # 未提取到尺寸

for row in remain_rows:
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if not s:
        no_dim_rows.append(row)
        continue
    
    l, w, h = extract_lwh(s)
    if l is None:
        no_dim_rows.append(row)
        continue
    
    # 判断内外
    is_wai = is_waijing(s)
    
    # 判断数值类型
    is_int = all(v == int(v) for v in [l, w, h])
    is_all5 = all(v % 1 == 0.5 for v in [l, w, h])
    
    if is_int:
        if is_wai:
            waijing_rows.append((row, l, w, h, '外径'))
        else:
            neijing_rows.append((row, l, w, h, '内径'))
    elif is_all5:
        if is_wai:
            neijing_rows.append((row, l, w, h, '外径'))
        else:
            neijing_rows.append((row, l, w, h, '内径'))
    else:
        feiquan_rows.append((row, l, w, h, '外径' if is_wai else '内径'))

print(f'\n分类结果:')
print(f'外径全量飞机盒: {len(waijing_rows)} 条')
print(f'内径全量飞机盒: {len(neijing_rows)} 条')
print(f'非全量飞机盒: {len(feiquan_rows)} 条')
print(f'未提取到尺寸: {len(no_dim_rows)} 条')
total_check = len(waijing_rows) + len(neijing_rows) + len(feiquan_rows) + len(no_dim_rows)
print(f'合计: {total_check} 条 (应={remain_count})')
if total_check == remain_count:
    print('✅ 数量正确！')
else:
    print(f'❌ 差 {abs(total_check - remain_count)}')

# ===== 7. 生成结构文件 =====
def skeleton(s):
    s = re.sub(r'\d+\.?\d*', 'N', s)
    s = re.sub(r'\s+', '', s)
    return s[:300]

def write_struct_file(data_list, fpath, title, label):
    """按店铺+结构分组输出"""
    # 按店铺+结构分组
    groups = {}
    for item in data_list:
        row = item[0]
        shop = str(row[0] or '').strip() if len(row) > 0 else '(无名)'
        spec = str(row[2] or '').strip() if len(row) > 2 else ''
        pid = str(row[3] or '') if len(row) > 3 else ''
        sk = skeleton(spec)
        key = (shop, sk)
        if key in groups:
            sample, first_pid, cnt = groups[key]
            groups[key] = (sample, first_pid, cnt + 1)
        else:
            groups[key] = (spec, pid, 1)
    
    by_shop = {}
    for (shop, sk), (spec, pid, cnt) in groups.items():
        by_shop.setdefault(shop, []).append((sk, spec, pid, cnt))
    
    shop_order = sorted(by_shop.keys(), key=lambda sh: sum(v[3] for v in by_shop[sh]), reverse=True)
    total_cnt = len(data_list)
    total_structs = len(groups)
    
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(f'{title}（{label}）\n')
        f.write(f'结构数: {total_structs}, 总商品数: {total_cnt}\n\n')
        gidx = 0
        for shop in shop_order:
            items = by_shop[shop]
            items.sort(key=lambda x: -x[3])
            shop_total = sum(v[3] for v in items)
            f.write(f'═══ {shop}（{len(items)} 种格式, 共 {shop_total} 条）═══\n')
            for sk, spec, pid, cnt in items:
                gidx += 1
                f.write(f'[{gidx}] [x{cnt}] 结构: {sk}\n')
                f.write(f'    样例: {spec}\n')
                f.write(f'    pid={pid}\n')
                f.write('\n')
    
    print(f'  ✅ {fpath}: {len(groups)} 结构, {total_cnt} 条')

outdir = r'D:\Desktop'
write_struct_file(waijing_rows, f'{outdir}\\5a-外径飞机盒.txt', '外径飞机盒', '长宽高同时为整数，沉默外径')
write_struct_file(neijing_rows, f'{outdir}\\5b-内径飞机盒.txt', '内径飞机盒', '长宽高同时为.5，或明确内径')
write_struct_file(feiquan_rows, f'{outdir}\\5c-非全量飞机盒.txt', '非全量飞机盒', '长宽高不是整数也不是全.5')

# 未提取到尺寸的也保存一下方便检查
write_struct_file(no_dim_rows, f'{outdir}\\5d-未提取尺寸.txt', '未提取到尺寸', '未能从规格名称提取长宽高')

print(f'\n总验证:')
total_out = sum(x['count'] for x in ([{'count': len(waijing_rows)}, {'count': len(neijing_rows)}, {'count': len(feiquan_rows)}, {'count': len(no_dim_rows)}]))
cls = len(waijing_rows) + len(neijing_rows) + len(feiquan_rows) + len(no_dim_rows)
print(f'5a+5b+5c+5d = {cls} 条')

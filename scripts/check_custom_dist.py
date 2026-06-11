# -*- coding: utf-8 -*-
"""检查定制链接380518条的分布"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

cats = {}
total_custom = 0
sample = 2000
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 用v8分类逻辑判断是否进定制
    from split_neijing_v8 import has_any_dim, build_lwh
    
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
        total_custom += 1
        cat = '定制关键词'
    elif '扣底盒' in s or '双插盒' in s:
        continue
    elif '纸箱' in s:
        continue
    else:
        lwh = build_lwh(s)
        if lwh is None:
            total_custom += 1
            if '外径' in s:
                # 检查是否有数字乘式
                has_mult = bool(re.search(r'\d[\d.]*\s*[xX*]\s*\d[\d.]*', s))
                has_wh = bool(re.search(r'宽.*?[\d.]+\s*[^；;]*?[\d.]+', s))
                if has_mult:
                    cat = '外径有乘号无LWH'
                elif '高度' in s:
                    cat = '外径有高度无长宽'
                elif 'mm' in s:
                    cat = '外径含mm'
                else:
                    cat = '外径仅有词无尺寸'
            elif '扣底盒' in s or '双插盒' in s:
                cat = '扣底盒无尺寸'
            elif '纸箱' in s:
                cat = '纸箱无尺寸'
            elif '定制' in s or '珍珠棉' in s or '咨询客服' in s:
                cat = '定制关键词'
            else:
                cat = '其他无尺寸'
    
    cats[cat] = cats.get(cat, 0) + 1
    if total_custom >= sample:
        break

wb.close()
print(f'采样{total_custom}条定制:')
for k, v in sorted(cats.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')

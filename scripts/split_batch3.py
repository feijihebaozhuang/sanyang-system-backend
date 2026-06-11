# -*- coding: utf-8 -*-
"""拆分换绑第3批（3.48MB > 3MB）为两个文件"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

fp = os.path.join(out, '换绑文件_第3批.xlsx')
wb = oxl.load_workbook(fp, data_only=True)
ws = wb['Sheet1']
rows = list(ws.iter_rows(min_row=3, values_only=True))
wb.close()

print(f"第3批总条数: {len(rows)}", flush=True)

# 平分两批
mid = math.ceil(len(rows) / 2)
print(f"拆分为: 第3-1批 {mid}条, 第3-2批 {len(rows)-mid}条")

for i, (start, end, suffix) in enumerate([
    (0, mid, '第3-1批'),
    (mid, len(rows), '第3-2批')
]):
    chunk = rows[start:end]
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for r in chunk:
        ws.append(list(r))
    fn = os.path.join(out, f'换绑文件_{suffix}.xlsx')
    wb.save(fn)
    wb.close()
    print(f"  ✅ 换绑文件_{suffix}.xlsx: {os.path.getsize(fn)/1024/1024:.2f}MB, {len(chunk)}条")

print("\n完成！保持原第3批不动，新增3-1和3-2")

# -*- coding: utf-8 -*-
"""阿里大鱼2706条处理"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 找最新的批次数
import glob
exist = [f for f in os.listdir(ok_dir) if f.startswith('换绑文件_第') and f.endswith('.xlsx')]
max_batch = 0
for f in exist:
    m = re.search(r'第(\d+)批', f)
    if m: max_batch = max(max_batch, int(m.group(1)))
new_batch = max_batch + 1

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
km = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in km.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cand = f"{base}-{dk}-{mat}"
    if cand in all_codes: return cand
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if shop != '阿里大鱼':
        remaining.append(r)
        continue
    
    # ===== 内径类: 三种顺序 =====
    # 顺序1: 长宽【41x24cm】；【100个】高【10cm】黄色内径
    m_inner = re.search(r'长宽【(\d+)x(\d+)cm】.*?高【(\d+)cm】.*?内径', spec_name)
    # 顺序2: 高【10cm】黄色内径;长宽【41x29cm】100个
    if not m_inner:
        m_inner = re.search(r'高度?【(\d+)cm】.*?内径;长宽【(\d+)x(\d+)cm】', spec_name)
    # 顺序3: 【100个】高度【10cm】黄色内径;长宽【41x37cm】
    if not m_inner:
        m_inner = re.search(r'【100个】高度?【(\d+)cm】.*?内径;长宽【(\d+)x(\d+)cm】', spec_name)
    # 顺序4: 高度【10cm】黄色内径 100个;长宽【41x12cm】;高度【10cm】黄色内径 100个（重复段）
    if not m_inner:
        m_inner = re.search(r'高度?【(\d+)cm】.*?内径\s*\d*个;长宽【(\d+)x(\d+)cm】;', spec_name)
    if m_inner:
        z = int(m_inner.group(1)); x = int(m_inner.group(2)); y = int(m_inner.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', '特硬')
        if code:
            matched_new.append((shop, pid, spec_id, code))
            stats['内径-特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-特硬'))
            stats['内径-无匹配'] += 1
        continue
    if m_inner:
        x, y, z = int(m_inner.group(1)), int(m_inner.group(2)), int(m_inner.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', '特硬')
        if code:
            matched_new.append((shop, pid, spec_id, code))
            stats['内径-特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-特硬'))
            stats['内径-无匹配'] += 1
        continue
    
    # ===== 外径类 =====
    # 顺序1: 高度【10cm】黄色外径;长宽【41x29cm】100个
    m_outer = re.search(r'高度?【(\d+)cm】.*?外径[^;]*;长宽【(\d+)x(\d+)cm】', spec_name)
    # 顺序2: 高度【2cm】黄色外径【100个】;长宽【41x8cm】;（带重复段）
    if not m_outer:
        m_outer = re.search(r'高度?【(\d+)cm】.*?外径【\d+个】;长宽【(\d+)x(\d+)cm】;', spec_name)
    if not m_outer:
        m_outer = re.search(r'长宽【(\d+)x(\d+)cm】.*?高【(\d+)cm】.*?外径', spec_name)
        if m_outer:
            x = int(m_outer.group(1)); y = int(m_outer.group(2)); z = int(m_outer.group(3))
            l, w = max(x, y), min(x, y)
            code = match_3d(l, w, z, '外径', '特硬')
            if code:
                matched_new.append((shop, pid, spec_id, code))
                stats['外径-长宽在前匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-特硬'))
                stats['外径-长宽在前无匹配'] += 1
            continue
    if m_outer:
        z, x, y = int(m_outer.group(1)), int(m_outer.group(2)), int(m_outer.group(3))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', '特硬')
        if code:
            matched_new.append((shop, pid, spec_id, code))
            stats['外径-特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-特硬'))
            stats['外径-无匹配'] += 1
        continue
    
    # ===== 第三类: 长12cm【100个】；宽-高【13*11cm】外径 =====
    m_type3 = re.search(r'长(\d+)cm【100个】.*?宽-高【(\d+)\*(\d+)cm】.*?外径', spec_name)
    if m_type3:
        l, w, h = int(m_type3.group(1)), int(m_type3.group(2)), int(m_type3.group(3))
        code = match_3d(l, w, h, '外径', '特硬')
        if code:
            matched_new.append((shop, pid, spec_id, code))
            stats['第三类-外径特硬匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{h}-外径-特硬'))
            stats['第三类-外径无匹配'] += 1
        continue
    
    # ===== 纸箱类: 100个   长宽12*12；高14【3层】 =====
    m_box = re.search(r'长宽(\d+)\s*\*\s*(\d+)[^;]*；高(\d+)【(\d+)层】', spec_name)
    if m_box:
        x, y, z, layer = int(m_box.group(1)), int(m_box.group(2)), int(m_box.group(3)), int(m_box.group(4))
        l, w = max(x, y), min(x, y)
        mat = '3B' if layer == 3 else ('EB' if layer == 5 else None)
        if mat:
            code = match_3d(l, w, z, '外径', mat)
            if code:
                matched_new.append((shop, pid, spec_id, code))
                stats['纸箱-%d层匹配' % layer] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
                stats['纸箱-%d层无匹配' % layer] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-未知层'))
            stats['纸箱-未知层'] += 1
        continue
    
    # 未识别
    remaining.append(r)
    stats['未识别'] += 1

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'大鱼未识别: {stats.get("未识别", 0)}')

# 保存换绑
if matched_new:
    f_out = os.path.join(ok_dir, f'换绑文件_第{new_batch}批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f_out)
    wb.close()
    print(f'\n✅ 换绑文件_第{new_batch}批: {len(matched_new)}条')

# 无匹配追加
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 追加{len(nomatch)}条到无匹配')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡已更新: {len(remaining)}条')

print('\n完成！')

# -*- coding: utf-8 -*-
"""
1. 从扣底盒双插盒文件移除 777564404927 的"特硬;双插盒"和"特硬;扣底盒"，加到定制链接
2. 从其余商品.xlsx中分出带小数点的（排除全.5）→ 非全量飞机盒.xlsx
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''
    val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or '')); w = normalize_dim(m.group(3)+(m.group(4) or '')); h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

# ===== 1. 处理扣底盒文件 =====
kd_file = r'D:\Desktop\扣底盒双插盒商品.xlsx'
df_kd = pd.read_excel(kd_file, skiprows=1, dtype=str)

# 找出 777564404927 的"特硬;双插盒"和"特硬;扣底盒"
pid_target = '777564404927'
mask_remove = (df_kd['平台商品id'].astype(str).str.strip() == pid_target) & \
              df_kd['规格名称'].astype(str).str.strip().isin(['特硬;双插盒', '特硬;扣底盒'])
removed = df_kd[mask_remove].copy()
remaining_kd = df_kd[~mask_remove].copy()

print(f'扣底盒文件: 移除{len(removed)}条, 剩余{len(remaining_kd)}条')

# 加到定制链接
custom_file = r'D:\Desktop\定制链接商品.xlsx'
df_custom = pd.read_excel(custom_file, skiprows=1, dtype=str)
if len(removed) > 0:
    extra_cols = [c for c in removed.columns if c not in ['店铺名称','平台商品id','规格名称','平台规格id']]
    add_rows = removed[['店铺名称','平台商品id','规格名称','平台规格id']].copy()
    df_custom = pd.concat([df_custom, add_rows], ignore_index=True)
print(f'定制链接: 现共{len(df_custom)}条')

# 保存
import openpyxl

def save_df(path, df, title, cols=None):
    if cols: df = df[cols]
    df.to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.insert_rows(1)
    ws.cell(1, 2).value = title
    ws.column_dimensions['C'].width = 60
    wb.save(path)

save_df(kd_file, remaining_kd, '扣底盒/双插盒')
save_df(custom_file, df_custom, '定制链接商品（无准确尺寸）')

print('✅ 扣底盒文件和定制链接已更新')

# ===== 2. 从剩余商品中分出带小数点的 =====
other_file = r'D:\Desktop\其余商品.xlsx'
df_other = pd.read_excel(other_file, skiprows=1, dtype=str)

specs = df_other['规格名称'].dropna().astype(str).str.strip()

dot_rows = []
remain_rows = []
dot_count = 0

for idx, (_, row) in enumerate(df_other.iterrows()):
    s = str(row.get('规格名称', '')).strip()
    if not s:
        remain_rows.append(row)
        continue
    
    lwh = extract_lwh(s)
    if not lwh:
        remain_rows.append(row)
        continue
    
    l, w, h = lwh
    has_dot = any(n != int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    
    if has_dot and not all_dot5:
        dot_rows.append(row)
        dot_count += 1
    else:
        remain_rows.append(row)

print(f'\n其余商品共{len(df_other)}条')
print(f'  带小数点(非全.5): {len(dot_rows)} 条 → 非全量飞机盒.xlsx')
print(f'  剩余: {len(remain_rows)} 条')

# 保存两个文件
feiji_file = r'D:\Desktop\非全量飞机盒.xlsx'
pd.DataFrame(dot_rows, columns=df_other.columns).to_excel(feiji_file, index=False)
wb = openpyxl.load_workbook(feiji_file)
ws = wb.active
ws.insert_rows(1)
ws.cell(1, 2).value = '其余商品中带小数点的（已排除全.5）'
ws.column_dimensions['C'].width = 60
for c in ['E','F','G','H']:
    try: ws.column_dimensions[c].width = 10
    except: pass
wb.save(feiji_file)
print(f'✅ 非全量飞机盒.xlsx: {len(dot_rows)}条')

# 更新剩余商品
save_df(other_file, pd.DataFrame(remain_rows, columns=df_other.columns), '其余商品（已移除小数项）')

print('\n=== 完成 ===')
print(f'定制链接商品.xlsx → {len(df_custom)}条')
print(f'扣底盒双插盒商品.xlsx → {len(remaining_kd)}条')
print(f'非全量飞机盒.xlsx → {len(dot_rows)}条')
print(f'其余商品.xlsx → {len(remain_rows)}条')

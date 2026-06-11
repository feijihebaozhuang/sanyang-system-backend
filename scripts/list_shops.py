# -*- coding: utf-8 -*-
"""从平台商品.xlsx提取唯一的店铺全称和简称"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

source = r'D:\Desktop\平台商品.xlsx'
wb = load_workbook(source, read_only=True)
ws = wb.active

shops = set()
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    full = str(row[3] or '').strip() if len(row) > 3 else ''
    short = str(row[4] or '').strip() if len(row) > 4 else ''
    if full:
        shops.add((full, short))

wb.close()

print(f'共有 {len(shops)} 个店铺:\n')
for full, short in sorted(shops):
    print(f'  {full:25s}  →  {short}')

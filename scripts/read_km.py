# -*- coding: utf-8 -*-
"""用 openpyxl 详细读取快麦商品副本，找数据在哪"""
import openpyxl, sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

fp = r"d:\Desktop\快麦商品 - 副本.xlsx"
print(f"读取: {fp}")

wb = openpyxl.load_workbook(fp, read_only=False, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\nSheet: {sn}")
    print(f"行: {ws.max_row}, 列: {ws.max_column}")
    print(f"合并单元格: {list(ws.merged_cells.ranges)}")
    
    # 看第1-5行所有单元格
    for r in range(1, min(6, ws.max_row+1)):
        vals = []
        for c in range(1, min(10, ws.max_column+1)):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append(f"({r},{c})={repr(cell.value)}")
        if vals:
            print(f"  行{r}: {vals}")

    # 统计每列非空数
    print("\n  每列非空数:")
    for c in range(1, min(10, ws.max_column+1)):
        cnt = 0
        for r in range(1, ws.max_row+1):
            if ws.cell(row=r, column=c).value is not None:
                cnt += 1
        if cnt > 0:
            print(f"    列{c}: {cnt}行")

wb.close()

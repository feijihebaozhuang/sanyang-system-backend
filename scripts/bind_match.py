# -*- coding: utf-8 -*-
"""
平台商品 ↔ 快麦商品 换绑匹配脚本

流程:
1. 读取平台商品表（49万条）→ 解析规格名称 → 拼出候选主商家编码
2. 读取快麦商品表（21万条）→ 建立主商家编码索引
3. 匹配 → 输出：
   a. 换绑文件（模板格式：店铺名称、平台商品id、平台规格id、商品编码）
   b. 未匹配的平台商品（让你确认材质映射）
   c. 统计报告

店铺名称 简称→全称 映射表（你提供的）：
"""
import sys, re, os, json
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 0. 店铺名称 简称→全称 映射
# ============================================================
SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂​',
    '淘宝品牌店': '飞机盒品牌店​',
    '阿里友尚': '深圳市友尚包装有限公司​',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

# ============================================================
# 1. 读取快麦商品表 → 建立编码索引
# ============================================================
def load_km_products(km_file):
    """读取快麦商品，返回 {主商家编码: {name, category}}"""
    import openpyxl
    wb = openpyxl.load_workbook(km_file, read_only=True, data_only=True)
    ws = wb['报表1']
    
    km_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        code, name, category = row[0], row[1], row[2]
        if code and str(code).strip():
            code = str(code).strip()
            km_map[code] = {
                'name': str(name or '').strip(),
                'category': str(category or '').strip(),
            }
    wb.close()
    return km_map

# ============================================================
# 2. 解析平台规格名称
# ============================================================
def parse_spec_name(spec_str):
    """
    解析平台规格名称，返回 {长, 宽, 高, 内外径, 材料, 数量, 是否定制}
    
    支持的格式示例:
    - 宽【10cm】高【8cm】内径;【100个】长【38cm】
    - 长19CM；宽19CM；高10CM;【200个】
    - 宽【25cm】高【15cm】;【150个】长【35cm】内径
    - 定制/定做 → 标记为定制
    """
    if not spec_str:
        return None
    
    s = str(spec_str).strip()
    
    # 统一分隔符：中文分号、英文分号、中文逗号 全部替换为英文分号
    s = re.sub(r'[；;]', ';', s)
    
    # 检查是否定制
    if '定制' in s or '定做' in s or '定造' in s:
        return {'custom': True}
    
    # 检查是否珍珠棉
    if '珍珠棉' in s or 'pe' in s.lower():
        return {'custom': True}
    
    result = {}
    
    # ---- 提取尺寸 ----
    # 匹配各种格式：长【38cm】、长19CM、长19 cm、长19
    # 【】格式
    def extract_dim(pattern, text):
        m = re.search(pattern, text)
        if m:
            val = m.group(1)
            if val:
                val = val.strip().rstrip('cmCM）)').strip()
                try:
                    return float(val)
                except:
                    pass
        return None
    
    result['长'] = extract_dim(r'长[度]?[：:，;；]?\s*【?(\d+\.?\d*)\s*cm?[CM]?】?', s)
    if result['长'] is None:
        result['长'] = extract_dim(r'长[度]?[：:，;；]?\s*【?(\d+\.?\d*)\s*cm?[CM]?】?', s.replace(';', '').replace(' ', ''))
    
    result['宽'] = extract_dim(r'宽[度]?[：:，;；]?\s*【?(\d+\.?\d*)\s*cm?[CM]?】?', s)
    
    result['高'] = extract_dim(r'高[度]?[：:，;；]?\s*【?(\d+\.?\d*)\s*cm?[CM]?】?', s)
    if result['高'] is None:
        result['高'] = extract_dim(r'厚[：:，;；]?\s*【?(\d+\.?\d*)\s*cm?[CM]?】?', s)

    # 如果还没解析到，试去掉所有分隔符
    if not (result['长'] and result['宽'] and result['高']):
        s_flat = re.sub(r'[：:，,；;【】、\s]', '', s)
        result2 = {}
        result2['长'] = extract_dim(r'长(?:度)?(\d+\.?\d*)', s_flat)
        result2['宽'] = extract_dim(r'宽(?:度)?(\d+\.?\d*)', s_flat)
        result2['高'] = extract_dim(r'高(?:度)?(\d+\.?\d*)', s_flat)
        if result['长'] is None and result2['长']: result['长'] = result2['长']
        if result['宽'] is None and result2['宽']: result['宽'] = result2['宽']
        if result['高'] is None and result2['高']: result['高'] = result2['高']
    
    # 最后的兜底：按 ; 分割找段落
    if not (result['长'] and result['宽'] and result['高']):
        parts = [p.strip() for p in re.split(r'[;；]', s) if p.strip()]
        for p in parts:
            pl = p.lower()
            if '长' in pl and result['长'] is None:
                m = re.search(r'【?(\d+\.?\d*)\s*cm?[CM]?】?', p)
                if m: result['长'] = float(m.group(1))
            elif '宽' in pl and result['宽'] is None:
                m = re.search(r'【?(\d+\.?\d*)\s*cm?[CM]?】?', p)
                if m: result['宽'] = float(m.group(1))
            elif ('高' in pl or '厚' in pl) and result['高'] is None:
                m = re.search(r'【?(\d+\.?\d*)\s*cm?[CM]?】?', p)
                if m: result['高'] = float(m.group(1))
    
    # 兜底2：直接找 x/X/* 分隔的三维数字
    if not (result['长'] and result['宽'] and result['高']):
        dim_match = re.search(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)', s)
        if dim_match:
            result['长'] = float(dim_match.group(1))
            result['宽'] = float(dim_match.group(2))
            result['高'] = float(dim_match.group(3))
    
    # 从文本中提取内外径和材料
    if '内径' in s or '内尺寸' in s:
        result['内外径'] = '内径'
    elif '外径' in s:
        result['内外径'] = '外径'
    result['材料'] = extract_material(s)
    
    # ---- 提取数量 ----
    qty = extract_dim(r'【?(\d+\.?\d*)\s*个】?', s)
    result['数量'] = qty
    
    # ---- 判断内外径 ----
    if '内径' in s or '内尺寸' in s or '内' in s:
        result['内外径'] = '内径'
    elif '外径' in s or '外尺寸' in s:
        result['内外径'] = '外径'
    else:
        result['内外径'] = '外径'  # 默认
    
    # ---- 提取材料 ----
    # 已知材料关键词
    result['材料'] = extract_material(s)
    
    # 检查是否所有三个尺寸都有了
    if result['长'] and result['宽'] and result['高']:
        return result
    elif result.get('custom'):
        return result
    
    return None


def extract_material(text):
    """从文本中提取材料。返回标准材料名，或 None"""
    t = str(text).strip()
    
    # 直接从规格文本中提取材料
    # 常见格式：材料-颜色【数量】;尺寸 或 材料 颜色 尺寸
    
    # 优先匹配已知材料词
    patterns = [
        (r'特硬', '特硬'),
        (r'超硬|台湾纸|台湾|进口纸', '超硬'),
        (r'白色|双白|双面白', '白色'),
        (r'黑色', '黑色'),
        (r'红色', '红色'),
        (r'蓝色', '蓝色'),
        (r'绿色', '绿色'),
        (r'黄色', '黄色'),
        (r'灰色', '灰色'),
        (r'牛皮', '牛皮'),
        (r'优质[进]?口?纸', '优质'),
        (r'优质(?![进]?口)', '优质'),
        (r'P\d*[A-Z]?', None),  # P6D 等
        (r'珍珠棉|PE|pe|pe', 'pe'),
    ]
    
    for pat, val in patterns:
        m = re.search(pat, t)
        if m:
            if val:
                return val
            return m.group(0)
    
    return None


# ============================================================
# 3. 拼出主商家编码候选
# ============================================================
def build_km_code(parsed):
    """
    从解析结果拼出主商家编码：L*W*H-内外径-材料
    尺寸按大小排序（从小到大）
    """
    if not parsed:
        return None
    if parsed.get('custom'):
        return None  # 定制商品不拼编码
    
    dims = [parsed.get('长'), parsed.get('宽'), parsed.get('高')]
    if not all(dims):
        return None
    
    # 尺寸按长宽高顺序
    dim_str = '*'.join(
        str(int(d) if d == int(d) else d) 
        for d in [parsed.get('长'), parsed.get('宽'), parsed.get('高')]
    )
    
    dim_kind = parsed.get('内外径', '外径')
    material = parsed.get('材料') or ''
    
    # 先试试最简格式
    candidates = []
    
    if material:
        candidates.append(f"{dim_str}-{dim_kind}-{material}")
    else:
        candidates.append(f"{dim_str}-{dim_kind}")
    
    return candidates


# ============================================================
# 4. 主匹配流程
# ============================================================
def match_platform_to_km(platform_file, km_file, output_dir):
    """主流程"""
    print("="*60)
    print("步骤1: 加载快麦商品表...")
    print(f"  文件: {km_file}")
    km_map = load_km_products(km_file)
    print(f"  加载快麦商品: {len(km_map)} 条")
    
    # 统计快麦商品的编码格式
    code_formats = Counter()
    for code in km_map:
        parts = code.split('-')
        code_formats[len(parts)] += 1
    print(f"  编码格式统计: {dict(code_formats)}")
    
    # 建立快麦编码的快速查找（精确匹配）
    km_codes = set(km_map.keys())
    
    # 也建立归一化匹配（去掉材料后缀、盒子类型后缀等）
    # 比如 10*10*10-外径-特硬 和 10*10*10-外径-特硬-双插盒
    km_normalized = {}  # L*W*H-内外径-材料 → [完整编码列表]
    for code in km_codes:
        # 尝试提取核心部分
        parts = code.split('-')
        if len(parts) >= 3:
            dims = parts[0]
            dim_kind = parts[1]
            material = parts[2]
            key = f"{dims}-{dim_kind}-{material}"
            km_normalized.setdefault(key, []).append(code)
    
    print(f"  归一化编码: {len(km_normalized)} 个")
    
    # ---- 读取平台商品 ----
    print("\n步骤2: 读取平台商品表并匹配...")
    import openpyxl
    
    wb = openpyxl.load_workbook(platform_file, read_only=True, data_only=True)
    ws = wb['报表1']
    
    matched = []      # 匹配上的 → 输出换绑文件
    unmatched = []    # 没匹配上的 → 输出给你确认
    stats = Counter()
    
    total = 0
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        shop_short, pid, spec_name, spec_id = row
        if not spec_name:
            continue
        total += 1
        
        # 解析规格
        parsed = parse_spec_name(str(spec_name))
        
        if parsed is None:
            unmatched.append((shop_short, pid, spec_id, spec_name, '解析失败'))
            stats['解析失败'] += 1
            continue
        
        if parsed.get('custom'):
            unmatched.append((shop_short, pid, spec_id, spec_name, '定制/珍珠棉'))
            stats['定制'] += 1
            continue
        
        # 拼编码
        candidates = build_km_code(parsed)
        if not candidates:
            unmatched.append((shop_short, pid, spec_id, spec_name, '缺尺寸'))
            stats['缺尺寸'] += 1
            continue
        
        # 尝试匹配
        matched_code = None
        for cand in candidates:
            if cand in km_codes:
                matched_code = cand
                break
        
        # 尝试归一化匹配
        if not matched_code:
            for cand in candidates:
                if cand in km_normalized:
                    # 取第一个匹配的
                    matched_code = km_normalized[cand][0]
                    break
        
        if matched_code:
            # 店铺名称 简称→全称
            full_name = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched.append((full_name, str(pid).strip(), str(spec_id).strip(), matched_code))
            stats['匹配成功'] += 1
        else:
            # 没匹配上 → 记录解析结果
            dims = [parsed.get('长'), parsed.get('宽'), parsed.get('高')]
            dim_str = '*'.join(str(d) for d in sorted(dims))
            expected_code = f"{dim_str}-{parsed.get('内外径','?')}-{parsed.get('材料','?')}"
            unmatched.append((shop_short, pid, spec_id, spec_name, expected_code))
            stats['无匹配编码'] += 1
        
        if total % 50000 == 0:
            print(f"  已处理 {total} 行...")
    
    wb.close()
    
    print(f"\n处理完成: 共 {total} 行")
    print(f"  匹配成功: {stats['匹配成功']}")
    print(f"  定制/珍珠棉: {stats['定制']}")
    print(f"  解析失败: {stats['解析失败']}")
    print(f"  缺尺寸: {stats['缺尺寸']}")
    print(f"  无匹配编码: {stats['无匹配编码']}")
    
    # ---- 输出文件 ----
    os.makedirs(output_dir, exist_ok=True)
    
    # 换绑文件
    bind_file = os.path.join(output_dir, '换绑文件.xlsx')
    import openpyxl as oxl
    wb_out = oxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    ws_out.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for row in matched:
        ws_out.append(list(row))
    wb_out.save(bind_file)
    print(f"\n✅ 换绑文件: {bind_file} ({len(matched)}条)")
    
    # 未匹配文件
    unmatch_file = os.path.join(output_dir, '未匹配平台商品.xlsx')
    wb_u = oxl.Workbook()
    ws_u = wb_u.active
    ws_u.title = 'Sheet1'
    ws_u.append(['店铺名称', '平台商品id', '平台规格id', '规格名称', '期望编码/原因'])
    for row in unmatched:
        ws_u.append(list(row))
    wb_u.save(unmatch_file)
    print(f"✅ 未匹配文件: {unmatch_file} ({len(unmatched)}条)")
    
    # 统计报告
    report = {
        '总计': total,
        '匹配成功': stats['匹配成功'],
        '未匹配': {
            '定制珍珠棉': stats['定制'],
            '解析失败': stats['解析失败'],
            '缺尺寸': stats['缺尺寸'],
            '无匹配编码': stats['无匹配编码'],
        }
    }
    report_file = os.path.join(output_dir, '匹配报告.json')
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"✅ 报告: {report_file}")
    
    return matched, unmatched


# ============================================================
# 5. 执行
# ============================================================
if __name__ == '__main__':
    import time
    t0 = time.time()
    
    PLATFORM_FILE = r"d:\Desktop\平台商品.xlsx"
    KM_FILE = r"d:\Desktop\快麦商品 - 副本.xlsx"
    OUTPUT_DIR = r"d:\Desktop\换绑输出"
    
    matched, unmatched = match_platform_to_km(PLATFORM_FILE, KM_FILE, OUTPUT_DIR)
    
    elapsed = time.time() - t0
    print(f"\n总耗时: {elapsed:.1f}秒")

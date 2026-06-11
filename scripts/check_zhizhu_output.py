# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

wb = oxl.load_workbook(r'D:\Desktop\换绑_飞机盒止合专卖店.xlsx')
ws = wb.active
# 看商品编码列不同值
codes = set()
total = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    code = str(row[3] or '')
    codes.add(code)
    total += 1
wb.close()

print(f'总条数: {total}')
print(f'不同编码数: {len(codes)}')
for c in sorted(codes):
    cnt = sum(1 for _ in open(r'D:\Desktop\换绑_飞机盒止合专卖店.xlsx', 'rb') if c.encode('utf-8') in _)
    # 直接用openpyxl重读
print()

# 采样看前20行的编码
wb = oxl.load_workbook(r'D:\Desktop\换绑_飞机盒止合专卖店.xlsx')
ws = wb.active
counts = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    code = str(row[3] or '').strip()
    counts[code] = counts.get(code, 0) + 1
wb.close()

for c, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f'  {c}: {n} 条')

# -*- coding: utf-8 -*-
"""
平台商品 → 快麦商品 换绑匹配 v3
性能优化版：提前编译正则
"""
import sys, re, os, time
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
    '阿里友尚': '深圳市友尚包装有限公司',
    '阿里亚润': '深圳市亚润包装材料有限公司',
    '阿里三羊': '深圳市三羊包装材料有限公司',
    '阿里正方形': '深圳市正方形纸制品有限公司',
    '阿里大鱼': '深圳市大鱼包装材料有限公司',
    '阿里新鑫星': '东莞市新鑫星包装材料有限公司',
}

# ---- 预编译正则 ----
RE_CUSTOM = re.compile(r'定制|定做|定造|订做|订制')
RE_PEARL = re.compile(r'珍珠棉')
RE_BLUE_GREEN = re.compile(r'蓝色|蓝|绿色|绿')
RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_WHITE = re.compile(r'双面白|双白|白色')
RE_RED = re.compile(r'红色|红')
RE_BLACK = re.compile(r'黑色|黑')
RE_SPECIAL_PRICE = re.compile(r'特价')
RE_NEW_MATERIAL = re.compile(r'新材质|新材')
RE_FIVELAYER = re.compile(r'五层|5层')
RE_THREELAYER = re.compile(r'三层|3层')
RE_DOUBLE_BOX = re.compile(r'双插盒')
RE_BUCKLE_BOX = re.compile(r'扣底盒')

# 内外径
RE_INNER = re.compile(r'内径|内尺寸')
RE_OUTER = re.compile(r'外径|外尺寸')

# 尺寸标签
LABEL_PATTERNS = [
    ('长', [
        re.compile(r'长[度]?\s*[:：]?\s*【?(\d+\.?\d*)\s*(?:cm|mm|厘米|毫米)?】?'),
        re.compile(r'长[度]?\s*[:：]?\s*【?(\d+\.?\d*)】?'),
    ]),
    ('宽', [
        re.compile(r'宽[度]?\s*[:：]?\s*【?(\d+\.?\d*)\s*(?:cm|mm|厘米|毫米)?】?'),
        re.compile(r'宽[度]?\s*[:：]?\s*【?(\d+\.?\d*)】?'),
    ]),
    ('高', [
        re.compile(r'高[度]?\s*[:：]?\s*【?(\d+\.?\d*)\s*(?:cm|mm|厘米|毫米)?】?'),
        re.compile(r'高[度]?\s*[:：]?\s*【?(\d+\.?\d*)】?'),
        re.compile(r'厚[度]?\s*[:：]?\s*【?(\d+\.?\d*)\s*(?:cm|mm|厘米|毫米)?】?'),
        re.compile(r'厚[度]?\s*[:：]?\s*【?(\d+\.?\d*)】?'),
    ]),
]
RE_DIMS_3D = re.compile(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)')
RE_DIMS_2D = re.compile(r'(\d+\.?\d*)\s*[xX*×]\s*(\d+\.?\d*)')
RE_DIGIT = re.compile(r'(\d+\.?\d*)\s*(cm|mm|厘米|毫米)?', re.IGNORECASE)

# ============================================================
# 1. 加载快麦商品
# ============================================================
def load_km_products(km_file):
    import pandas as pd
    t0 = time.time()
    print("  读取快麦文件...", flush=True)
    df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
    df.columns = ['code', 'name', 'category']
    km_map = {}
    for _, row in df.iterrows():
        c = str(row['code']).strip()
        if c and c != 'nan':
            km_map[c] = {
                'name': str(row['name'] or '').strip(),
                'category': str(row['category'] or '').strip(),
            }
    print(f"  → 加载 {len(km_map)} 条, 耗时 {time.time()-t0:.1f}s", flush=True)
    return km_map


# ============================================================
# 2. 构建快麦索引
# ============================================================
def build_km_index(km_map):
    t0 = time.time()
    exact_idx = {}
    fuzzy_idx = {}
    dim_only_idx = {}
    all_codes = set(km_map.keys())
    
    for code in km_map:
        parts = code.split('-')
        if len(parts) >= 3:
            try:
                dims = parts[0].split('*')
                if len(dims) == 3:
                    l, w, h = float(dims[0]), float(dims[1]), float(dims[2])
                    dk = parts[1]
                    mat = parts[2]
                    key = (l, w, h, dk, mat)
                    exact_idx.setdefault(key, []).append(code)
                    vals = tuple(sorted([l, w, h]))
                    fkey = (vals[0], vals[1], vals[2], dk)
                    fuzzy_idx.setdefault(fkey, []).append(code)
                    # 按(l,w,h,dk)建索引用于无材料匹配
                    dimkey = (l, w, h, dk)
                    dim_only_idx.setdefault(dimkey, []).append(code)
            except:
                pass
    
    print(f"  → 索引: {len(exact_idx)}精确 {len(fuzzy_idx)}模糊 {len(dim_only_idx)}尺寸 耗时 {time.time()-t0:.1f}s", flush=True)
    return exact_idx, fuzzy_idx, dim_only_idx, all_codes


# ============================================================
# 3. 规格解析（性能版）
# ============================================================
def guess_dim_kind(s):
    if RE_INNER.search(s):
        return '内径'
    if RE_OUTER.search(s):
        return '外径'
    return '外径'


def guess_material(s):
    if RE_CHAOYING.search(s):
        return '超硬'
    if RE_WHITE.search(s):
        return '白色'
    if RE_RED.search(s):
        return '红色'
    if RE_BLACK.search(s):
        return '黑色'
    # 特价/新材质 → 优质
    if RE_SPECIAL_PRICE.search(s) or RE_NEW_MATERIAL.search(s):
        return '优质'
    if RE_FIVELAYER.search(s):
        return 'EB'
    if RE_THREELAYER.search(s):
        return '3B'
    return '特硬'


def extract_labeled_dims(s):
    dims = {}
    for label, pats in LABEL_PATTERNS:
        for pat in pats:
            m = pat.search(s)
            if m and m.group(1):
                val = float(m.group(1))
                ctx = s[max(0, m.start()-5):m.end()+5]
                if 'mm' in ctx:
                    val = val / 10.0
                dims[label] = val
                break
    return dims


def extract_nums(s):
    nums = []
    for m in RE_DIGIT.finditer(s):
        val = float(m.group(1))
        unit = (m.group(2) or '').lower()
        if unit in ('mm', '毫米'):
            val = val / 10.0
        if 0.5 < val <= 500:
            nums.append(val)
    return sorted(set(nums), reverse=True)


def parse_spec_v3(text, shop):
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None
    
    if RE_PEARL.search(s):
        return {'custom': True, 'reason': '珍珠棉'}
    
    dims = extract_labeled_dims(s)
    
    # a*b*c
    m = RE_DIMS_3D.search(s)
    if m:
        vals = [float(m.group(i)) for i in range(1, 4)]
        ctx = s[max(0, m.start()-10):m.end()+10]
        if 'mm' in ctx:
            vals = [v/10 for v in vals]
        for lbl, val in zip(['长', '宽', '高'], vals):
            dims.setdefault(lbl, val)
    
    # a*b
    if not (dims.get('长') and dims.get('宽')):
        m = RE_DIMS_2D.search(s)
        if m:
            v1, v2 = float(m.group(1)), float(m.group(2))
            ctx = s[max(0, m.start()-10):m.end()+10]
            if 'mm' in ctx:
                v1, v2 = v1/10, v2/10
            dims.setdefault('长', v1)
            dims.setdefault('宽', v2)
    
    # 兜底数字
    if not (dims.get('长') and dims.get('宽') and dims.get('高')):
        nums = extract_nums(s)
        if '长' not in dims and len(nums) >= 3:
            dims['长'], dims['宽'], dims['高'] = nums[0], nums[1], nums[2]
        elif '长' not in dims and len(nums) == 2:
            dims['长'], dims['宽'] = nums[0], nums[1]
    
    have_3d = dims.get('长') and dims.get('宽') and dims.get('高')
    have_2d = dims.get('长') and dims.get('宽') and not dims.get('高')
    
    # 定制判断：有关键词但无三围
    if RE_CUSTOM.search(s) and not have_3d:
        return {'custom': True, 'reason': '定制关键词'}
    
    # 蓝/绿 → 定制
    if RE_BLUE_GREEN.search(s):
        return {'custom': True, 'reason': '蓝绿颜色'}
    
    # 平卡
    if have_2d:
        return {'custom': True, 'reason': '只有长宽'}
    
    if not have_3d:
        return {'custom': True, 'reason': '尺寸不足'}
    
    return {
        '长': dims['长'],
        '宽': dims['宽'],
        '高': dims['高'],
        '内外径': guess_dim_kind(s),
        '材料': guess_material(s),
        '类型': '双插盒' if RE_DOUBLE_BOX.search(s) else ('扣底盒' if RE_BUCKLE_BOX.search(s) else ''),
    }


def build_candidates(parsed):
    if not parsed or parsed.get('custom'):
        return []
    dim_str = '*'.join(
        str(int(d) if d == int(d) else d)
        for d in [parsed['长'], parsed['宽'], parsed['高']]
    )
    base = f"{dim_str}-{parsed['内外径']}-{parsed['材料']}"
    xtype = parsed.get('类型', '')
    if xtype:
        return [f"{base}-{xtype}", base]
    return [base]


# ============================================================
# 主逻辑（批量处理，减少函数调用开销）
# ============================================================
def match(platform_file, km_file, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    km_map = load_km_products(km_file)
    exact_idx, fuzzy_idx, dim_only_idx, all_codes = build_km_index(km_map)
    
    print("读取平台商品...", flush=True)
    import pandas as pd
    t0 = time.time()
    df = pd.read_excel(platform_file, sheet_name='报表1', header=2, dtype=str)
    print(f"  → 读取 {len(df)} 行, 耗时 {time.time()-t0:.1f}s", flush=True)
    
    # 转成numpy数组加速迭代
    arr = df.values
    total_rows = len(arr)
    
    matched = []
    unmatched = []
    stats = Counter()
    total = 0
    failed_samples = []
    custom_samples = []
    
    for row in arr:
        total += 1
        shop_short = str(row[0] or '').strip() if len(row) >= 1 else ''
        pid = str(row[1] or '').strip() if len(row) >= 2 else ''
        spec_name = str(row[2] or '').strip() if len(row) >= 3 else ''
        spec_id = str(row[3] or '').strip() if len(row) >= 4 else ''
        
        if not spec_name:
            continue
        
        parsed = parse_spec_v3(spec_name, shop_short)
        
        if parsed is None or parsed.get('custom'):
            reason = parsed.get('reason', '') if parsed else '解析失败'
            if not reason:
                reason = '解析失败'
            unmatched.append((shop_short, str(pid)[:30], str(spec_id)[:30], str(spec_name)[:120], reason, ''))
            stats[reason] += 1
            if parsed and parsed.get('custom') and len(custom_samples) < 50:
                custom_samples.append((shop_short, pid, spec_name))
            elif not parsed and len(failed_samples) < 50:
                failed_samples.append((shop_short, pid, spec_name))
            continue
        
        candidates = build_candidates(parsed)
        matched_code = None
        
        for cand in candidates:
            if cand in all_codes:
                matched_code = cand
                break
        
        if not matched_code:
            dims_arr = [parsed['长'], parsed['宽'], parsed['高']]
            dk = parsed['内外径']
            vals = tuple(sorted(dims_arr))
            fkey = (vals[0], vals[1], vals[2], dk)
            fuzz_cands = fuzzy_idx.get(fkey, [])
            if fuzz_cands:
                mat = parsed['材料']
                for c in fuzz_cands:
                    if mat in c:
                        matched_code = c
                        break
                if not matched_code:
                    matched_code = fuzz_cands[0]
        
        if not matched_code:
            dimkey = (parsed['长'], parsed['宽'], parsed['高'], parsed['内外径'])
            dim_cands = dim_only_idx.get(dimkey, [])
            if dim_cands:
                matched_code = dim_cands[0]
        
        if matched_code:
            full_name = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched.append((full_name, str(pid).strip(), str(spec_id).strip(), matched_code))
            stats['匹配成功'] += 1
        else:
            dim_str = '*'.join(str(int(d) if d == int(d) else d) for d in [parsed['长'], parsed['宽'], parsed['高']])
            expected = f"{dim_str}-{parsed['内外径']}-{parsed['材料']}"
            unmatched.append((shop_short, str(pid)[:30], '', str(spec_name)[:120], '无匹配', expected))
            stats['无匹配'] += 1
        
        if total % 100000 == 0:
            print(f"  → {total}/{total_rows} 已完成...", flush=True)
    
    print(f"\n{'='*60}", flush=True)
    print(f"处理完成: 共 {total} 行", flush=True)
    print(f"  匹配成功: {stats.get('匹配成功', 0)}", flush=True)
    for k, v in sorted(stats.items()):
        if k != '匹配成功':
            print(f"  {k}: {v}", flush=True)
    
    import openpyxl as oxl
    
    bind_file = os.path.join(output_dir, '换绑文件.xlsx')
    wb_out = oxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    ws_out.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for row in matched:
        ws_out.append(list(row))
    wb_out.save(bind_file)
    print(f"\n✅ 换绑文件: {bind_file} ({len(matched)}条)", flush=True)
    
    unmatch_file = os.path.join(output_dir, '未匹配平台商品.xlsx')
    wb_u = oxl.Workbook()
    ws_u = wb_u.active
    ws_u.title = '未匹配'
    ws_u.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for row in unmatched:
        ws_u.append(list(row))
    
    ws_c = wb_u.create_sheet('定制样本')
    ws_c.append(['店铺', '商品id', '规格名称'])
    for s, pid, name in custom_samples:
        ws_c.append([s, str(pid)[:20], str(name)[:100]])
    
    ws_f = wb_u.create_sheet('解析失败样本')
    ws_f.append(['店铺', '商品id', '规格名称'])
    for s, pid, name in failed_samples:
        ws_f.append([s, str(pid)[:20], str(name)[:100]])
    
    wb_u.save(unmatch_file)
    print(f"✅ 未匹配文件: {unmatch_file} ({len(unmatched)}条)", flush=True)


if __name__ == '__main__':
    t0 = time.time()
    match(
        r"d:\Desktop\平台商品.xlsx",
        r"d:\Desktop\快麦商品 - 副本.xlsx",
        r"d:\Desktop\换绑输出"
    )
    print(f"\n总耗时: {time.time()-t0:.1f}秒", flush=True)

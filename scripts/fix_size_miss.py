# -*- coding: utf-8 -*-
"""
处理尺寸不足中 "宽度：X cm---高度：Y cm;白色 100个 长度：Z cm" 格式
全部有完整尺寸，重新解析并匹配快麦
"""
import sys, os, re, time, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ============================================================
# 读取快麦索引（复现bind_match_v4的逻辑）
# ============================================================
def load_km(km_file):
    import pandas as pd
    df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
    km_map = {}
    for row in df.values:
        code = str(row[0] or '').strip()
        name = str(row[1] or '').strip()
        cat = str(row[2] or '').strip()
        if code:
            km_map[code] = (name, cat)
    return km_map

def build_idx(km_map):
    exact = {}
    fuzzy = {}
    dim_dk = {}
    all_codes = set()
    for code in km_map:
        all_codes.add(code)
        parts = code.split('-')
        if len(parts) >= 3:
            dim_part = parts[0]
            dk = parts[1]
            mat = '-'.join(parts[2:])
            dims = dim_part.split('*')
            if len(dims) == 3:
                try:
                    vals = tuple(sorted([float(d) for d in dims]))
                    ek = (vals[0], vals[1], vals[2], dk, mat)
                    exact[ek] = code
                    fk = (vals[0], vals[1], vals[2], dk)
                    fuzzy.setdefault(fk, []).append(code)
                    k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                    dim_dk.setdefault(k3, []).append(code)
                except:
                    pass
    return exact, fuzzy, dim_dk, all_codes

km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
print("加载快麦索引...", flush=True)
km_map = load_km(km_file)
exact_idx, fuzzy_idx, dim_dk_idx, all_codes = build_idx(km_map)
print(f"快麦商品: {len(km_map)}", flush=True)

# ============================================================
# 读取尺寸不足文件
# ============================================================
wb = oxl.load_workbook(os.path.join(out, '尺寸不足_待处理.xlsx'), data_only=True)
ws = wb['尺寸不足']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

print(f"尺寸不足条数: {len(rows)}", flush=True)

# ============================================================
# 解析 "宽度：X cm---高度：Y cm;白色 100个 长度：Z cm"
# ============================================================
# 正则: 宽度：X cm---高度：Y cm;白色 100个 长度：Z cm
RE_WIDTH_HEIGHT_LEN = re.compile(
    r'宽度[：:]\s*(\d+\.?\d*)\s*cm?\s*-+\s*高度[：:]\s*(\d+\.?\d*)\s*cm?\s*;'
    r'\s*白色\s+\d+\s*个\s+长度[：:]\s*(\d+\.?\d*)\s*cm?'
)

# 颜色/材料提取
RE_WHITE = re.compile(r'白色')
RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')

matched = []   # (店铺全称, 平台商品id, 平台规格id, 商品编码)
still_miss = [] # 还是没匹配上的

for r in rows:
    shop_short = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    m = RE_WIDTH_HEIGHT_LEN.search(spec_name)
    if m:
        w_val = float(m.group(1))
        h_val = float(m.group(2))
        l_val = float(m.group(3))
        
        # 排序: 长 ≥ 宽 ≥ 高
        dims_orig = [l_val, w_val, h_val]
        dims_orig.sort(reverse=True)
        
        # 判断材料
        if RE_WHITE.search(spec_name):
            mat = '白色'
        elif RE_CHAOYING.search(spec_name):
            mat = '超硬'
        else:
            mat = '特硬'
        
        dk = '外径'  # 无内外径关键词，沉默默认外径
        
        def dim_fmt(v):
            if v == int(v):
                return str(int(v))
            return f"{v:.1f}"
        
        base = f"{dim_fmt(dims_orig[0])}*{dim_fmt(dims_orig[1])}*{dim_fmt(dims_orig[2])}"
        
        # 构建候选编码
        candidates = [
            f"{base}-{dk}-{mat}",
            f"{base}-{dk}-特硬",   # 白色也可以匹配特硬
        ]
        if mat == '白色':
            candidates.append(f"{base}-{dk}-白色")
        
        hit = None
        for cand in candidates:
            if cand in all_codes:
                hit = cand
                break
        
        # 模糊匹配
        if not hit:
            vals = tuple(sorted(dims_orig))
            fkey = (vals[0], vals[1], vals[2], dk)
            fuzz = fuzzy_idx.get(fkey, [])
            if fuzz:
                for c in fuzz:
                    if mat in c:
                        hit = c
                        break
                if not hit:
                    # 如果材料是白色，尝试特硬
                    if mat == '白色':
                        for c in fuzz:
                            if '特硬' in c:
                                hit = c
                                break
                if not hit:
                    hit = fuzz[0]
        
        # 内外径转换
        if not hit and dk == '外径':
            il = dims_orig[0] - 1.5
            iw = dims_orig[1] - 0.5
            ih = dims_orig[2] - 0.5
            if il > 0 and iw > 0 and ih > 0:
                ivals = tuple(sorted([il, iw, ih]))
                fkey2 = (ivals[0], ivals[1], ivals[2], '内径')
                fuzz2 = fuzzy_idx.get(fkey2, [])
                if fuzz2:
                    for c in fuzz2:
                        if mat in c:
                            hit = c
                            break
                    if not hit and mat == '白色':
                        for c in fuzz2:
                            if '特硬' in c:
                                hit = c
                                break
                    if not hit:
                        hit = fuzz2[0]
        
        if hit:
            full_name = '飞机盒小批量专卖店'  # 从规格看是小批量的
            # 用shop_short查全称
            SHOP_NAME_MAP = {
                '天猫小批量': '飞机盒小批量专卖店',
            }
            full_name = SHOP_NAME_MAP.get(shop_short, shop_short)
            matched.append((full_name, pid, spec_id, hit))
        else:
            expected = f"{base}-{dk}-{mat}"
            still_miss.append((shop_short, pid, spec_id, spec_name, '无匹配', expected))
    else:
        # 不是这个格式的，保留到尺寸不足
        still_miss.append(r)

print(f"\n匹配结果:", flush=True)
print(f"  新匹配成功: {len(matched)}", flush=True)
print(f"  仍无匹配: {len(still_miss)}", flush=True)

# ============================================================
# 保存新匹配的 → 换绑文件_第5批.xlsx
# ============================================================
f5 = os.path.join(out, '换绑文件_第5批.xlsx')
wb = oxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws.append([None, '商品对应表', None, None])
ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
for m in matched:
    ws.append(list(m))
wb.save(f5)
sz = os.path.getsize(f5)
print(f"\n✅ 换绑_第5批: {f5} ({sz/1024:.1f}KB, {len(matched)}条)", flush=True)
wb.close()

# ============================================================
# 更新尺寸不足_待处理.xlsx（只保留仍无匹配的）
# ============================================================
f2 = os.path.join(out, '尺寸不足_待处理.xlsx')
wb = oxl.Workbook()
ws = wb.active
ws.title = '尺寸不足'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in still_miss:
    ws.append(list(r))
wb.save(f2)
print(f"✅ 尺寸不足_待处理: {f2} ({len(still_miss)}条)", flush=True)
wb.close()

print(f"\n{'='*50}")
print("处理完成！")
print(f"  换绑文件_第5批: {len(matched)}条（可直接上传）")
if still_miss:
    print(f"  尺寸不足剩余: {len(still_miss)}条（需进一步处理）")
print(f"{'='*50}")

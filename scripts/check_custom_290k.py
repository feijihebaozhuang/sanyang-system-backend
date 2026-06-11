# -*- coding: utf-8 -*-
"""检查定制链接290003条的规格分布"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import re

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    rows.append(s)
wb.close()

print(f'定制链接总数: {len(rows)}')
print()

# 分类
stats = {}
has_neijing_count = 0
has_waijing_count = 0
has_feiji = 0
has_dikoudi = 0  # 实际完整的扣底盒
has_zhixiang = 0

# 看看有多少是宽高+长格式但没被extract_lwh识别的
whl_format = 0  # 宽【W】高【H】; 长【L】
other_3value = 0  # 其他3值格式

for s in rows:
    has_neijing = '内径' in s or '内尺寸' in s or '内寸' in s
    has_waijing = '外径' in s
    has_feiji_ = '飞机盒' in s
    has_dikoudi_ = '扣底盒' in s or '双插盒' in s
    has_zhixiang_ = '纸箱' in s
    
    if has_neijing: has_neijing_count += 1
    if has_waijing: has_waijing_count += 1
    if has_feiji_: has_feiji += 1
    if has_dikoudi_: has_dikoudi += 1
    if has_zhixiang_: has_zhixiang += 1
    
    # 宽【W】高【H】; 长【L】（没有写"内径"但也是3值）
    m1 = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    m2 = re.search(r'长【\s*([\d.]+)\s*cm?\s*】', s)
    if m1 and m2:
        whl_format += 1
    
    # 定制/珍珠棉/咨询客服
    is_custom_keyword = ('定制' in s or '珍珠棉' in s or '咨询客服' in s)
    if is_custom_keyword:
        key = '定制/珍珠棉/咨询客服'
    elif has_neijing:
        key = '内径(宽高+长)'
    elif has_waijing:
        key = '外径'
    elif m1 and m2:
        key = '宽高+长(无内径)'
    elif has_zhixiang_:
        key = '纸箱(残)'
    elif has_dikoudi_:
        key = '扣底盒双插盒(残)'
    else:
        # 取前60字符作为示例
        key = s[:60]
    stats[key] = stats.get(key, 0) + 1

print(f'  内径: {has_neijing_count} 条')
print(f'  外径: {has_waijing_count} 条')
print(f'  飞机盒: {has_feiji} 条')
print(f'  扣底盒/双插盒: {has_dikoudi} 条')
print(f'  纸箱: {has_zhixiang} 条')
print(f'  宽高+长格式(但无内径关键词): {whl_format} 条')
print()

# 显示前20个分类
print('=== 定制链接规格分布 Top 30 ===')
for i, (k, v) in enumerate(sorted(stats.items(), key=lambda x: -x[1])[:30]):
    print(f'  {i+1}. [{k[:70]}]: {v}')

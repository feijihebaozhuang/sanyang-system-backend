# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re

def normalize_dim(v_str):
    v_str = v_str.strip()
    unit = ''; val_str = v_str
    if v_str.lower().endswith('cm'): unit = 'cm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('mm'): unit = 'mm'; val_str = v_str[:-2].strip()
    elif v_str.lower().endswith('c'): unit = 'cm'; val_str = v_str[:-1].strip()
    elif v_str.lower().endswith('m'): unit = 'mm'; val_str = v_str[:-1].strip()
    try:
        v = float(val_str)
        return v / 10 if unit == 'mm' else v
    except: return None

def extract_lwh(s):
    s = str(s)
    m = re.search(r'(?:外[尺寸]*寸*[大小]*|内[尺寸]*寸*[大小]*)[：:]?\s*【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*长\s*([\d.]+)\s*[xX*×]\s*宽\s*([\d.]+)\s*[xX*×]\s*高\s*([\d.]+)\s*】', s)
    if m:
        l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
        if 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'长[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?宽[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】.*?高[度]*[：:]?\s*【\s*([\d.]+)\s*(cm|mm|c|m)?\s*】', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(2) or '')); w = normalize_dim(m.group(3)+(m.group(4) or '')); h = normalize_dim(m.group(5)+(m.group(6) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'【\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*[xX*×]\s*([\d.]+)\s*】\s*(cm|mm|c|m)?', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    m = re.search(r'(?<![\d.])([\d.]+)\s*[xX*]\s*([\d.]+)\s*[xX*]\s*([\d.]+)\s*(cm|mm|c|m)?(?![\d.])', s)
    if m:
        l = normalize_dim(m.group(1)+(m.group(4) or '')); w = normalize_dim(m.group(2)+(m.group(4) or '')); h = normalize_dim(m.group(3)+(m.group(4) or ''))
        if l and w and h and 0 < l <= 300 and 0 < w <= 300 and 0 < h <= 300: return (l, w, h)
    return None

def is_dot5(v):
    return v != int(v) and abs(v * 2 - round(v * 2)) < 0.01

other_file = r'D:\Desktop\其余商品.xlsx'
from openpyxl import load_workbook

# 流式读（不占内存）
feiji_rows = []
other_rows = []

wb = load_workbook(other_file, read_only=True)
ws = wb.active
header = None
HEADER_ROW = 1  # 第2行是表头（索引1），数据从索引2开始
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # 标题行
    if i == 1:
        header = list(row)
        continue
    # 数据行
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    if not s:
        other_rows.append(row)
        continue
    lwh = extract_lwh(s)
    if not lwh:
        other_rows.append(row)
        continue
    l, w, h = lwh
    has_dot = any(n != int(n) for n in (l, w, h))
    all_dot5 = all(is_dot5(n) for n in (l, w, h))
    if has_dot and not all_dot5:
        feiji_rows.append(row)
    else:
        other_rows.append(row)

wb.close()

print(f'其余商品总行: {len(feiji_rows) + len(other_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)} 条')
print(f'剩余: {len(other_rows)} 条')

# 输出
import openpyxl as opx
feiji_file = r'D:\Desktop\非全量飞机盒.xlsx'
if header:
    pd.DataFrame(feiji_rows, columns=header).to_excel(feiji_file, index=False)
    wb2 = opx.load_workbook(feiji_file)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = '其余商品中带小数点的（已排除全.5）- 非全量飞机盒'
    ws2.column_dimensions['C'].width = 60
    for c in ['E','F','G','H']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(feiji_file)
    print(f'✅ 非全量飞机盒.xlsx: {len(feiji_rows)}条')

    # 更新其余商品
    pd.DataFrame(other_rows, columns=header).to_excel(other_file, index=False)
    wb3 = opx.load_workbook(other_file)
    ws3 = wb3.active
    ws3.insert_rows(1)
    ws3.cell(1, 2).value = '其余商品（已移除小数项）'
    ws3.column_dimensions['C'].width = 60
    for c in ['E','F','G','H']:
        try: ws3.column_dimensions[c].width = 10
        except: pass
    wb3.save(other_file)
    print(f'✅ 其余商品已更新: {len(other_rows)}条')

print('\n=== 桌面文件清单 ===')
print(f'1. 定制链接商品.xlsx ✅ (16KB)')
print(f'2. 扣底盒双插盒商品.xlsx ✅ (202KB)')
print(f'3. 纸箱商品.xlsx ✅ (1.2MB)')
print(f'4. 非全量飞机盒.xlsx ✅ ({len(feiji_rows)}条)')
print(f'5. 其余商品.xlsx ✅ ({len(other_rows)}条)')

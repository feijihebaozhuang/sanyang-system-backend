#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导入快麦商家编码映射（Excel/CSV/TSV）。不全量镜像商品，只写生产用字段。"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import km_sku_map_store as kms

# 表头别名 → 字段名
_HEADER_ALIASES: dict[str, str] = {
    "outer_id": "outer_id",
    "商家编码": "outer_id",
    "编码": "outer_id",
    "sysouterid": "outer_id",
    "outerid": "outer_id",
    "spec_alias": "spec_alias",
    "规格别名": "spec_alias",
    "规格": "spec_alias",
    "product_type": "product_type",
    "产品类型": "product_type",
    "类型": "product_type",
    "length": "length",
    "长": "length",
    "长cm": "length",
    "width": "width",
    "宽": "width",
    "宽cm": "width",
    "height": "height",
    "高": "height",
    "高cm": "height",
    "dim_kind": "dim_kind",
    "内外径": "dim_kind",
    "内/外径": "dim_kind",
    "material": "material",
    "材料": "material",
    "材质": "material",
    "km_title": "km_title",
    "商品名": "km_title",
    "快麦商品名": "km_title",
    "标题": "km_title",
}


def _norm_header(h: str) -> str:
    s = (h or "").strip().lower().replace(" ", "").replace("_", "")
    for k, v in _HEADER_ALIASES.items():
        if s == k.lower().replace(" ", "").replace("_", ""):
            return v
    return ""


def _detect_col_map(headers: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        field = _norm_header(str(h or ""))
        if field and field not in col:
            col[field] = i
    return col


def _row_from_cells(cells: list, col_map: dict[str, int]) -> dict | None:
    def cell(name: str) -> str:
        idx = col_map.get(name)
        if idx is None or idx >= len(cells):
            return ""
        v = cells[idx]
        return "" if v is None else str(v).strip()

    outer_id = cell("outer_id")
    spec_alias = cell("spec_alias")
    if not outer_id and not spec_alias:
        return None

    l = w = h = 0.0
    material = cell("material")
    try:
        l = float(cell("length") or 0)
        w = float(cell("width") or 0)
        h = float(cell("height") or 0)
    except ValueError:
        pass
    if not (l and w) and spec_alias:
        pl, pw, ph, mat_tail = kms.parse_spec_alias_dims(spec_alias)
        if pl and pw:
            l, w, h = pl, pw, ph
        if not material and mat_tail:
            material = mat_tail

    pt_raw = cell("product_type")
    product_type = kms.normalize_product_type(pt_raw, l=l, w=w)
    dim_kind = kms.normalize_dim_kind(cell("dim_kind"))

    if not outer_id and spec_alias:
        # 无编码时用规格别名作占位（便于后续补编码）
        outer_id = f"ALIAS:{spec_alias[:120]}"

    return {
        "outer_id": outer_id,
        "spec_alias": spec_alias,
        "product_type": product_type,
        "length": l,
        "width": w,
        "height": h,
        "dim_kind": dim_kind,
        "material": material,
        "km_title": cell("km_title"),
    }


def _read_excel(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("需要 openpyxl: pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(x or "") for x in next(rows_iter, [])]
    col_map = _detect_col_map(headers)
    if "outer_id" not in col_map and "spec_alias" not in col_map:
        raise SystemExit(f"未识别表头: {headers[:12]}")
    out: list[dict] = []
    for row in rows_iter:
        cells = list(row or [])
        r = _row_from_cells(cells, col_map)
        if r:
            out.append(r)
    wb.close()
    return out


def _read_csv(path: Path, delimiter: str | None):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        if delimiter is None:
            delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.reader(f, delimiter=delimiter)
        headers = next(reader, [])
        col_map = _detect_col_map([str(h) for h in headers])
        if "outer_id" not in col_map and "spec_alias" not in col_map:
            raise SystemExit(f"未识别表头: {headers[:12]}")
        out: list[dict] = []
        for row in reader:
            r = _row_from_cells(list(row), col_map)
            if r:
                out.append(r)
        return out


def _read_txt_lines(path: Path) -> list[dict]:
    """解析 export_unmatched_dimoldb.txt 或类似行：类型\\t规格\\t库存"""
    out: list[dict] = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or s.startswith("-"):
            continue
        if "成品" in s and "总数" in s:
            continue
        parts = [p.strip() for p in s.replace("|", "\t").split("\t") if p.strip()]
        if len(parts) < 2:
            parts = s.split(None, 2)
        if len(parts) < 2:
            continue
        product_type = parts[0]
        spec_alias = parts[1]
        stock = 0
        if len(parts) >= 3:
            tail = parts[2]
            m = __import__("re").search(r"(\d+)", tail)
            if m:
                try:
                    stock = int(m.group(1))
                except ValueError:
                    pass
        l, w, h, material = kms.parse_spec_alias_dims(spec_alias)
        out.append(
            {
                "outer_id": f"SPEC:{spec_alias[:120]}",
                "spec_alias": spec_alias,
                "product_type": kms.normalize_product_type(product_type, l=l, w=w),
                "length": l,
                "width": w,
                "height": h,
                "dim_kind": "",
                "material": material,
                "km_title": "",
                "_stock_hint": stock,
            }
        )
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="导入 km_sku_map 映射表")
    ap.add_argument("path", type=Path, help="Excel(.xlsx)/CSV/TSV/txt")
    ap.add_argument("--delimiter", default=None, help="CSV 分隔符，默认自动")
    ap.add_argument("--txt-unmatched", action="store_true", help="按未匹配刀模导出格式解析")
    ap.add_argument("--dry-run", action="store_true", help="只统计不写库")
    args = ap.parse_args()
    path = args.path
    if not path.is_file():
        raise SystemExit(f"文件不存在: {path}")

    suffix = path.suffix.lower()
    if args.txt_unmatched or path.name.startswith("export_unmatched"):
        rows = _read_txt_lines(path)
    elif suffix in (".xlsx", ".xlsm", ".xltx"):
        rows = _read_excel(path)
    else:
        rows = _read_csv(path, args.delimiter)

    print(f"解析 {len(rows)} 行")
    if args.dry_run:
        for r in rows[:5]:
            print(r)
        print("...(dry-run，未写入)")
        return

    n = kms.upsert_rows(rows)
    total = kms.row_count()
    print(f"写入/更新 {n} 行，表内共 {total} 行")


if __name__ == "__main__":
    main()

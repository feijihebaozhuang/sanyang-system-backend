# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd

source = r'D:\Desktop\未识别飞机盒_待分析.xlsx'
shop_name = '东莞市新鑫星包装材料有限公司'

print('读取中...')
df = pd.read_excel(source, dtype=str)

# 第0行是标题"店铺名称"等，第1行才是真正的数据开始
data = df.iloc[1:].copy()
data.columns = ['店铺名称', '平台商品id', '平台规格名称', '平台规格id']

# 过滤店铺
shop_data = data[data['店铺名称'].str.strip() == shop_name].copy()
print(f'{shop_name}: {len(shop_data)} 条')

# 构建换绑文件
out = pd.DataFrame({
    '店铺名称': shop_data['店铺名称'],
    '平台商品id': shop_data['平台商品id'],
    '平台规格id': shop_data['平台规格id'],
    '商品编码': ''  # 留空，等你填
})

outpath = r'D:\Desktop\换绑_东莞市新鑫星包装材料有限公司.xlsx'
out.to_excel(outpath, index=False, header=False, startrow=1)
# 第1行写入标题
import openpyxl as oxl
wb = oxl.load_workbook(outpath)
ws = wb.active
ws.insert_rows(1)
ws['A1'] = None
ws['B1'] = '商品对应表'
ws['C1'] = None
ws['D1'] = None
# 第2行列头
ws['A2'] = '店铺名称'
ws['B2'] = '平台商品id'
ws['C2'] = '平台规格id'
ws['D2'] = '商品编码'

# 调整列宽
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15

wb.save(outpath)
wb.close()

print(f'✅ 已生成: {outpath}')
print(f'   共 {len(shop_data)} 条')

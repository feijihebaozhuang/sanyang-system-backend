# -*- coding: utf-8 -*-
"""处理品牌店剩余576条：宽度【X*Y】cm;外径【100个】长度 Z cm → 外径 特硬"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd

out = r'd:\Desktop\换绑输出'

# ========== 快麦索引 ==========
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

RE_WHL = re.compile(r'宽度[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*[】]*\s*cm?.*?长度\s*(\d+\.?\d*)\s*cm')

# 读待确认文件
wb = oxl.load_workbook(os.path.join(out, '品牌店_待确认.xlsx'), data_only=True)
ws = wb['待确认']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    m = RE_WHL.search(spec_name)
    if m:
        w_cm = float(m.group(1))  # 宽度=宽
        h_cm = float(m.group(2))  # 高度=高
        l_cm = float(m.group(3))  # 长度=长
        dk = '外径'
        mat = '特硬'
        dims = sorted([l_cm, w_cm, h_cm], reverse=True)
        
        code = match_3d(dims[0], dims[1], dims[2], dk, mat)
        if code:
            matched_new.append((fn(shop), pid, spec_id, code))
        else:
            # 外转内
            il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
            if il > 0 and iw > 0 and ih > 0:
                code_i = match_3d(il, iw, ih, '内径', mat)
                if code_i:
                    matched_new.append((fn(shop), pid, spec_id, code_i))
                else:
                    nomatch.append((shop, pid, spec_id, spec_name, '无匹配',
                        f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
            else:
                nomatch.append((shop, pid, spec_id, spec_name, '无匹配',
                    f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
    else:
        nomatch.append(r)

print(f'待处理: {len(rows)}')
print(f'匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')

if nomatch:
    print(f'\n无匹配详情:')
    for d in nomatch[:10]:
        print(f'  {str(d[3])[:80]} -> {str(d[5])[:60] if len(d)>5 else "?"}')

# 追加到第11批
if matched_new:
    f11 = os.path.join(out, '换绑文件_第11批.xlsx')
    wb = oxl.load_workbook(f11)
    ws = wb['Sheet1']
    for m in matched_new:
        ws.append(list(m))
    wb.save(f11)
    wb.close()
    print(f'\n✅ 换绑文件_第11批追加完成: 共{len(matched_new)}条新增')

# 更新待确认（剩余无匹配）
if nomatch:
    wb = oxl.Workbook()
    ws = wb.active; ws.title = '待确认'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '品牌店_待确认.xlsx'))
    wb.close()
    print(f'✅ 品牌店_待确认: {len(nomatch)}条')
elif os.path.exists(os.path.join(out, '品牌店_待确认.xlsx')):
    os.remove(os.path.join(out, '品牌店_待确认.xlsx'))
    print(f'✅ 品牌店_待确认已清空')

# 更新平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
pk_rows = list(ws.iter_rows(min_row=2, values_only=True))
# 品牌店已全部处理完，只需保留非品牌店
non_pp = [r for r in pk_rows if r and '品牌店' not in str(r[0] or '')]
wb.close()

# 写回平卡（去掉所有品牌店）
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in non_pp:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡_待处理已更新: {len(non_pp)}条')

print('\n品牌店全部处理完毕！')

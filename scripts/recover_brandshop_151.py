# -*- coding: utf-8 -*-
"""找回品牌店151条无匹配数据-从平台商品xlsx直接读取"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'
raw_file = r'd:\Desktop\平台商品.xlsx'

# 1. 收集所有已处理过的spec_id（换绑+定制+无匹配+平卡）
processed_specs = set()

# 换绑文件
for f in sorted(os.listdir(out)):
    if '换绑' in f and f.endswith('.xlsx'):
        try:
            wb = oxl.load_workbook(os.path.join(out, f))
            for sn in wb.sheetnames:
                for r in wb[sn].iter_rows(min_row=3, values_only=True):
                    if r and len(r) >= 3:
                        processed_specs.add(str(r[2] or '').strip())
            wb.close()
        except: pass

# 定制文件
for f in sorted(os.listdir(out)):
    if '定制' in f and f.endswith('.xlsx'):
        try:
            wb = oxl.load_workbook(os.path.join(out, f))
            for sn in wb.sheetnames:
                for r in wb[sn].iter_rows(min_row=3, values_only=True):
                    if r and len(r) >= 3:
                        processed_specs.add(str(r[2] or '').strip())
            wb.close()
        except: pass

# 无匹配
wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
ws = wb['无匹配']
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) >= 3:
        processed_specs.add(str(r[2] or '').strip())
wb.close()

# 平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) >= 3:
        processed_specs.add(str(r[2] or '').strip())
wb.close()

print(f'已处理spec_id总数: {len(processed_specs)}')

# 2. 从原始数据找品牌店中 宽度【X*Y】长度 Z cm 格式且未被处理的数据
wb = oxl.load_workbook(raw_file, data_only=True)
ws = wb['报表1']
found = []
total_pp = 0
for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not r[0]: continue
    shop = str(r[0]).strip()
    if '品牌店' not in shop: continue
    total_pp += 1
    
    spec_name = str(r[2] or '').strip() if len(r) >= 3 else ''
    spec_id = str(r[3] or '').strip() if len(r) >= 4 else ''
    pid = str(r[1] or '').strip() if len(r) >= 2 else ''
    
    if spec_id in processed_specs:
        continue
    
    # 匹配宽度【X*Y】cm;外径【100个】长度 Z cm
    m = re.search(r'宽度[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*[】]*\s*cm?.*?长度\s*(\d+\.?\d*)\s*cm', spec_name)
    if m:
        w = float(m.group(1))
        h = float(m.group(2))
        l = float(m.group(3))
        found.append((shop, pid, spec_id, spec_name, l, w, h))

wb.close()

print(f'品牌店原始总数: {total_pp}')
print(f'未处理且匹配宽度*长度格式: {len(found)}')

if found:
    # 追加到无匹配
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for shop, pid, spec_id, spec_name, l, w, h in found:
        ws.append(['飞机盒品牌店', pid, spec_id, spec_name, '无匹配',
            f'{int(l)}*{int(w)}*{int(h)}-外径-特硬'])
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 已追加{len(found)}条到无匹配_待处理')
    
    # 也列出前几条确认
    for f2 in found[:5]:
        print(f'  {f2[3][:80]} -> {int(f2[4])}*{int(f2[5])}*{int(f2[6])}-外径-特硬')
else:
    print('没有需要找回的数据')

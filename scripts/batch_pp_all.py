# -*- coding: utf-8 -*-
"""一次性处理品牌店全部911条 - 3种模式
模式A: （宽）X mm 外径;【100个】长度 X mm;Y mm [双面白/无]
模式B: 进口优质特硬E瓦-内径/外经;长x宽【100x100】mm;100mm【高】
模式C: 【双面白色/台湾纸/特硬原色】外径;【XxY】;Z cm高度

匹配上→第11批换绑，匹配不上→列出来让用户看
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ========== 快麦索引 ==========
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}", f"{base}-{dk}-特硬"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ========== 正则 ==========
# 模式A: （宽）135 mm 外径;【100个】 长度 135 mm;105 mm 双面白
RE_A = re.compile(
    r'[（(]宽[）)]\s*(\d+\.?\d*)\s*mm?\s*外径\s*;\s*[^长]*?长度\s*'
    r'(\d+\.?\d*)\s*mm?\s*;\s*(\d+\.?\d*)\s*mm?\s*(.*)'
)
# 模式B: 进口优质特硬E瓦-内径;长x宽【100x100】mm;100mm【高】
RE_B = re.compile(
    r'进口优质.*?(内径|外经|外径)\s*;\s*长x宽[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*mm?[】]*\s*;\s*(\d+\.?\d*)\s*mm?【高】'
)
# 模式C: 【双面白色】外径;【36x36 】;10 cm高度
RE_C = re.compile(
    r'【([^】]*)】\s*(内径|外径)\s*;\s*[【】\s]*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)[】]*\s*;\s*(\d+\.?\d*)\s*cm?\s*高度'
)

RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_BAISE = re.compile(r'双面白色|双面白|白色|双白')

def guess_mat(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_BAISE.search(s): return '白色'
    return '特硬'

# ========== 读平卡 ==========
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# ========== 处理 ==========
matched_new = []
nomatch_details = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    if '品牌店' not in shop:
        remaining.append(r)
        continue
    
    handled = False
    
    # ---- 模式B: 进口优质特硬 ----
    m = RE_B.search(spec_name)
    if m:
        dk_raw = m.group(1).strip()
        dk = '内径' if '内' in dk_raw else '外径'
        x_mm = float(m.group(2))
        y_mm = float(m.group(3))
        h_mm = float(m.group(4))
        l, w, h = x_mm/10, y_mm/10, h_mm/10
        mat = '特硬'
        code = match_3d(l, w, h, dk, mat)
        if code:
            matched_new.append((fn(shop), pid, spec_id, code))
            stats['模式B-匹配'] += 1
        else:
            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                f"{dfmt(l)}*{dfmt(w)}*{dfmt(h)}-{dk}-{mat}"))
            stats['模式B-无匹配'] += 1
        handled = True
    
    # ---- 模式A: （宽）Xmm 外径... ----
    if not handled and '进口' not in spec_name:
        m = RE_A.search(spec_name)
        if m:
            w_mm = float(m.group(1))
            l_mm = float(m.group(2))
            h_mm = float(m.group(3))
            suffix = (m.group(4) or '').strip()
            
            l, w, h = l_mm/10, w_mm/10, h_mm/10
            dk = '外径'
            mat = guess_mat(spec_name)  # 双面白→白色, otherwise 特硬
            
            # 先试外径匹配
            code = match_3d(l, w, h, dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['模式A-外径匹配'] += 1
            else:
                # 外径对不上，转内径：长-1.5, 宽-0.5, 高-0.5
                il, iw, ih = l-1.5, w-0.5, h-0.5
                if il > 0 and iw > 0 and ih > 0:
                    code_i = match_3d(il, iw, ih, '内径', mat)
                    if code_i:
                        matched_new.append((fn(shop), pid, spec_id, code_i))
                        stats['模式A-内径转换匹配'] += 1
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                        stats['模式A-内径无匹配'] += 1
                else:
                    nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                        f"{dfmt(l)}*{dfmt(w)}*{dfmt(h)}-{dk}-{mat}"))
                    stats['模式A-外径无匹配'] += 1
            handled = True
    
    # ---- 模式C: 【材料】外径;【XxY】;Zcm高度 ----
    if not handled:
        m = RE_C.search(spec_name)
        if m:
            mat_ctx = m.group(1).strip()
            dk = m.group(2).strip()
            x = float(m.group(3))
            y = float(m.group(4))
            h = float(m.group(5))
            
            l, w = max(x, y), min(x, y)
            dims = sorted([l, w, h], reverse=True)
            mat = guess_mat(spec_name)
            
            code = match_3d(dims[0], dims[1], dims[2], dk, mat)
            if code:
                matched_new.append((fn(shop), pid, spec_id, code))
                stats['模式C-匹配'] += 1
            else:
                if dk == '外径':
                    il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
                    if il > 0 and iw > 0 and ih > 0:
                        code_i = match_3d(il, iw, ih, '内径', mat)
                        if code_i:
                            matched_new.append((fn(shop), pid, spec_id, code_i))
                            stats['模式C-内径转换匹配'] += 1
                        else:
                            nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                                f"{dfmt(il)}*{dfmt(iw)}*{dfmt(ih)}-内径-{mat}"))
                            stats['模式C-无匹配'] += 1
                    else:
                        nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                            f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                        stats['模式C-无匹配'] += 1
                else:
                    nomatch_details.append((shop, pid, spec_id, spec_name, '无匹配',
                        f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
                    stats['模式C-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f"\n====== 品牌店处理结果 ======")
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}")
print(f"\n匹配成功: {len(matched_new)}")
print(f"无匹配(需你确认): {len(nomatch_details)}")
print(f"平卡剩余(非品牌店): {len(remaining)}")

# ========== 保存匹配结果到第11批 ==========
if matched_new:
    f11 = os.path.join(out, '换绑文件_第11批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f11)
    sz = os.path.getsize(f11)
    print(f"\n✅ 换绑文件_第11批: {f11} ({sz/1024:.1f}KB, {len(matched_new)}条)")
    wb.close()

# ========== 保存无匹配待确认 ==========
if nomatch_details:
    wb = oxl.Workbook()
    ws = wb.active; ws.title = '待确认'
    ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
    for d in nomatch_details:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '品牌店_待确认.xlsx'))
    print(f"✅ 品牌店_待确认: {len(nomatch_details)}条（你告诉我对应的编码）")
    wb.close()

# ========== 更新平卡 ==========
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理已更新: {len(remaining)}条")
wb.close()

print("\n完成！请检查 品牌店_待确认.xlsx，告诉我对应的编码")

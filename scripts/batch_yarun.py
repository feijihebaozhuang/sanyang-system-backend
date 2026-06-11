# -*- coding: utf-8 -*-
"""阿里亚润4433条处理"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'
ok_dir = os.path.join(out, 'OK文件')

# 找最新批次数
exist = [f for f in os.listdir(ok_dir) if f.startswith('换绑文件_第') and f.endswith('.xlsx')]
max_batch = 0
for f in exist:
    m = re.search(r'第(\d+)批', f)
    if m: max_batch = max(max_batch, int(m.group(1)))
new_batch = max_batch + 1

SHOP_FULL = '深圳市亚润包装材料有限公司'

def mat_map(text):
    """材料映射"""
    if '白色' in text or '双面白色' in text: return '白色'
    if '黑色' in text: return '黑色'
    if '红色' in text: return '红色'
    if '超硬' in text or '超硬200克' in text or ('台湾' in text): return '超硬'
    if '双插盒' in text: return 'P6D'
    if '扣底盒' in text: return 'P6D'
    if '优质' in text or '特价' in text: return '优质'
    if '特特硬' in text: return '特硬'
    return '特硬'  # 黄色、特硬黄、特硬牛皮色、特硬180克等全部

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
km = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in km.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cand = f"{base}-{dk}-{mat}"
    if cand in all_codes: return cand
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'))
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []
nomatch = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_name = str(r[2] or '').strip()
    spec_id = str(r[3] or '').strip()
    
    if '亚润' not in shop:
        remaining.append(r)
        continue
    
    # ===== ① 长度X cm;宽*高【Y*Z】材料 =====
    m1 = re.search(r'长度\s*(\d+)\s*cm;宽\*高【(\d+)\*(\d+)】([^;]*?)(?:;|$)', spec_name)
    if m1:
        l = int(m1.group(1)); w = int(m1.group(2)); h = int(m1.group(3))
        mat = mat_map(m1.group(4))
        dims = sorted([l, w, h], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['①长度宽高-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['①长度宽高-无匹配'] += 1
        continue
    
    # ===== ② 内径/外径类: 长*宽【X*Y】；内径 Z cm高度-材料 或 外径 Z cm材料高 =====
    m2 = re.search(r'长\*宽【(\d+)\*(\d+)】[^;]*[；;]\s*内径\s*(\d+)\s*cm高度-([^;]*?)(?:;|$)', spec_name)
    if not m2:
        m2 = re.search(r'长\*宽【(\d+)\*(\d+)】[^;]*[；;]\s*外径\s*(\d+)\s*cm([^；;]*?)高(?:;|$)', spec_name)
        if m2:
            x, y, z = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
            mat = mat_map(m2.group(4))
            l, w = max(x, y), min(x, y)
            code = match_3d(l, w, z, '外径', mat)
            if code:
                matched_new.append((SHOP_FULL, pid, spec_id, code))
                stats['②外径白色-匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
                stats['②外径白色-无匹配'] += 1
            continue
    if m2:
        x, y, z = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        mat = mat_map(m2.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['②内径-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-{mat}'))
            stats['②内径-无匹配'] += 1
        continue
    
    # ===== ③ 外径类: 【外径】 Zmm/cm高度-材料;长x宽【XxY】mm =====
    m3 = re.search(r'[（(【]外径[】)）]\s*(\d+)\s*(mm|cm)\s*高度-([^;]*?);长x宽【(\d+)x(\d+)】\s*(mm|cm)', spec_name)
    if m3:
        z_val = int(m3.group(1)); unit = m3.group(2); mat = mat_map(m3.group(3))
        x_val = int(m3.group(4)); y_val = int(m3.group(5)); xy_unit = m3.group(6)
        z = z_val if unit == 'cm' else round(z_val / 10, 1)
        x = x_val if xy_unit == 'cm' else round(x_val / 10, 1)
        y = y_val if xy_unit == 'cm' else round(y_val / 10, 1)
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['③外径mm-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
            stats['③外径mm-无匹配'] += 1
        continue
    
    # ===== ③b 内径类mm: （内径) Zmm高度-材料;长x宽【XxY】mm =====
    m3b = re.search(r'[（(【]内径[】)）]\s*(\d+)\s*(mm|cm)\s*高度-([^;]*?);长x宽【(\d+)x(\d+)】\s*(mm|cm)', spec_name)
    if m3b:
        z_val = int(m3b.group(1)); unit = m3b.group(2); mat = mat_map(m3b.group(3))
        x_val = int(m3b.group(4)); y_val = int(m3b.group(5)); xy_unit = m3b.group(6)
        z = z_val if unit == 'cm' else round(z_val / 10, 1)
        x = x_val if xy_unit == 'cm' else round(x_val / 10, 1)
        y = y_val if xy_unit == 'cm' else round(y_val / 10, 1)
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['③b内径mm-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-{mat}'))
            stats['③b内径mm-无匹配'] += 1
        continue
    
    # ===== ④ 其他mm: 133*113*113 mm；材料 =====
    m4 = re.search(r'^(\d+)\s*\*\s*(\d+)\s*\*\s*(\d+)\s*mm[^;]*[；;]\s*([^;]*?)(?:;|$)', spec_name)
    if m4:
        x = int(m4.group(1)); y = int(m4.group(2)); z = int(m4.group(3))
        mat = mat_map(m4.group(4))
        dims = sorted([x, y, z], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['④mm尺寸-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['④mm尺寸-无匹配'] += 1
        continue
    
    # ===== ⑤ 纸箱类: 3层纸箱13*12*11 或 3层纸箱22.5*9.5*14 =====
    m5 = re.search(r'(?:(\d+)层|五(\d*)层)纸箱([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)', spec_name)
    if m5:
        layer = int(m5.group(1) or m5.group(2) or 3)
        x, y, z = float(m5.group(3)), float(m5.group(4)), float(m5.group(5))
        mat = '3B' if layer == 3 else 'EB'
        dims = sorted([x, y, z], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑤纸箱-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['⑤纸箱-无匹配'] += 1
        continue
    
    # ===== ⑥ 外尺寸类: 长宽【41*7】外尺寸；【高2cm】黄色 =====
    m6 = re.search(r'长宽【(\d+)\*(\d+)】外尺寸[^;]*[；;]【高(\d+)cm】([^;]*?)(?:;|$)', spec_name)
    if not m6:
        m6 = re.search(r'(\d+\.?\d*)[×x](\d+\.?\d*)[×x](\d+\.?\d*)\s*cm[^;]*外尺寸', spec_name)
        if m6:
            x = float(m6.group(1)); y = float(m6.group(2)); z = float(m6.group(3))
            dims = sorted([x, y, z], reverse=True)
            # 找材料
            mat = '白色' if '白色' in spec_name else '特硬'
            code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
            if code:
                matched_new.append((SHOP_FULL, pid, spec_id, code))
                stats['⑥外尺寸-匹配'] += 1
            else:
                nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
                stats['⑥外尺寸-无匹配'] += 1
            continue
    if m6:
        z = int(m6.group(3)); x = int(m6.group(1)); y = int(m6.group(2))
        mat = mat_map(m6.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑥外尺寸-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{mat}'))
            stats['⑥外尺寸-无匹配'] += 1
        continue
    
    # ===== ⑦ mm类颜色: 【红色】或（黑色) 100mm高度-特硬【100个】;长x宽【100x100】mm =====
    m7 = re.search(r'[（(【]([^】)）]+)[】)）]\s*(\d+)\s*(mm|cm)\s*高度-([^;【]*?)(?:【\d+个】)?;长x宽【(\d+)x(\d+)】mm', spec_name)
    if m7:
        color = m7.group(1); z_val = int(m7.group(2)); unit = m7.group(3); mat = mat_map(m7.group(4)); x_mm = int(m7.group(5)); y_mm = int(m7.group(6))
        z = z_val if unit == 'cm' else round(z_val / 10, 1)
        x = round(x_mm / 10, 1); y = round(y_mm / 10, 1)
        l, w = max(x, y), min(x, y)
        # 如果mat不是特硬/白色等，用颜色名
        final_mat = color if color in ('红色', '白色', '黑色') else mat
        code = match_3d(l, w, z, '外径', final_mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑦mm红色-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-外径-{final_mat}'))
            stats['⑦mm红色-无匹配'] += 1
        continue
    
    # ===== ⑧ 外径直接尺寸类: 20*15.3*5.5；特硬牛皮色；E瓦楞 外径 =====
    m8 = re.search(r'^([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)[^；;]*[；;]([^；;]*?)[；;].*?外径', spec_name)
    if not m8:
        m8 = re.search(r'^([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)外径(双插盒|扣底盒)', spec_name)
    if m8:
        x = float(m8.group(1)); y = float(m8.group(2)); z = float(m8.group(3))
        dims = sorted([x, y, z], reverse=True)
        mat = mat_map(spec_name)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑧外径直尺-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['⑧外径直尺-无匹配'] += 1
        continue
    
    # ===== ⑨ 内径类: 长宽【10*10】超硬黄色；内径 10 cm高 =====
    m9 = re.search(r'长宽【(\d+)\*(\d+)】([^；;]*?)[；;]\s*内径\s*(\d+)\s*cm\s*高', spec_name)
    if m9:
        x, y = int(m9.group(1)), int(m9.group(2))
        mat = mat_map(m9.group(3))
        z = int(m9.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑨内径超硬-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-{mat}'))
            stats['⑨内径超硬-无匹配'] += 1
        continue
    
    # ===== ⑩ 其他尺寸类: 31.5*30.5*15.4；特硬180克；牛皮色 或 25.1x15x7.7；白色 =====
    m10 = re.search(r'^([\d.]+)\s*[×x*]\s*([\d.]+)\s*[×x*]\s*([\d.]+)[^；;]*[；;]([^；;]*)', spec_name)
    if m10:
        x = float(m10.group(1)); y = float(m10.group(2)); z = float(m10.group(3))
        dims = sorted([x, y, z], reverse=True)
        mat = mat_map(m10.group(4))
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑩其他尺寸-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['⑩其他尺寸-无匹配'] += 1
        continue
    
    # ===== ⑪ 内径尺寸类: 长宽【41*6】；【高度 3 cm】-牛皮色【内径尺寸】 =====
    m11 = re.search(r'长宽【(\d+)\*(\d+)】[^；;]*[；;]【高度\s*(\d+)\s*cm-?】([^【]*?)【内径尺寸】', spec_name)
    if m11:
        x, y, z = int(m11.group(1)), int(m11.group(2)), int(m11.group(3))
        mat = mat_map(m11.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑪内径牛皮色-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-{mat}'))
            stats['⑪内径牛皮色-无匹配'] += 1
        continue
    
    # ===== ⑫ 【高度X cm】黑色内径/外径;长宽【Y*Z】 =====
    m12 = re.search(r'【高度\s*(\d+)\s*cm】([^;]*?)(内径|外径);长宽【(\d+)\*(\d+)】', spec_name)
    if m12:
        z = int(m12.group(1)); mat = mat_map(m12.group(2)); dk = m12.group(3); x = int(m12.group(4)); y = int(m12.group(5))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, dk, mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑫【】内外径-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-{dk}-{mat}'))
            stats['⑫【】内外径-无匹配'] += 1
        continue
    
    # ===== ⑬ 飞机盒型号类: A10【18X8X5.5CM】；三层 特价 A+；牛皮色 =====
    m13 = re.search(r'[\w]+【([\d.]+)X([\d.]+)X([\d.]+)CM】', spec_name)
    if m13:
        x, y, z = float(m13.group(1)), float(m13.group(2)), float(m13.group(3))
        dims = sorted([x, y, z], reverse=True)
        mat = mat_map(spec_name)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑬飞机盒-%s匹配' % mat] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['⑬飞机盒-无匹配'] += 1
        continue
    
    # ===== ⑭ 【12*10】黄色；长12 cm =====
    m14 = re.search(r'【(\d+)\*(\d+)】([^；;]*?)[；;]长\s*(\d+)\s*cm', spec_name)
    if m14:
        x, y = int(m14.group(1)), int(m14.group(2))
        mat = mat_map(m14.group(3))
        l_val = int(m14.group(4))
        dims = sorted([l_val, x, y], reverse=True)
        code = match_3d(dims[0], dims[1], dims[2], '外径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑭【】长-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-外径-{mat}'))
            stats['⑭【】长-无匹配'] += 1
        continue
    
    # ===== ⑮ 长宽【X*Y】cm；【高Zcm】黄色内径/外径 =====
    m15 = re.search(r'长宽【(\d+)\*(\d+)】\s*cm[^；;]*[；;]【高(\d+)cm】([^；;]*?)(内径|外径)', spec_name)
    if m15:
        x, y, z = int(m15.group(1)), int(m15.group(2)), int(m15.group(3))
        mat = mat_map(m15.group(4))
        dk = m15.group(5)
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, dk, mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑮长宽高-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-{dk}-{mat}'))
            stats['⑮长宽高-无匹配'] += 1
        continue
    
    # ===== ⑯ 长宽【X*Y】；【高度 Z cm-】牛皮色（无【内径尺寸】标记） =====
    m16 = re.search(r'长宽【(\d+)\*(\d+)】[^；;]*[；;]【高度\s*(\d+)\s*cm-?】([^【]*?)(?:;|$)', spec_name)
    if m16 and '内径' not in spec_name and '外径' not in spec_name:
        x, y, z = int(m16.group(1)), int(m16.group(2)), int(m16.group(3))
        mat = mat_map(m16.group(4))
        l, w = max(x, y), min(x, y)
        code = match_3d(l, w, z, '内径', mat)
        if code:
            matched_new.append((SHOP_FULL, pid, spec_id, code))
            stats['⑯牛皮色内径-匹配'] += 1
        else:
            nomatch.append((shop, pid, spec_name, spec_id, '无匹配', f'{l}*{w}*{z}-内径-{mat}'))
            stats['⑯牛皮色内径-无匹配'] += 1
        continue
    
    remaining.append(r)
    stats['未识别'] += 1

print(f'\n=== 处理结果 ===')
for k, v in sorted(stats.items()):
    print(f'  {k}: {v}')
print(f'\n匹配成功: {len(matched_new)}')
print(f'无匹配: {len(nomatch)}')
print(f'未识别: {stats.get("未识别", 0)}')

# 保存换绑
if matched_new:
    f_out = os.path.join(ok_dir, f'换绑文件_第{new_batch}批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f_out)
    wb.close()
    print(f'\n✅ 换绑文件_第{new_batch}批: {len(matched_new)}条')

# 无匹配追加
if nomatch:
    wb = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'))
    ws = wb['无匹配']
    for d in nomatch:
        ws.append(list(d[:6]))
    wb.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb.close()
    print(f'✅ 追加{len(nomatch)}条到无匹配')

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '规格名称', '平台规格id', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r[:6]) if len(r) >= 6 else list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print(f'✅ 平卡已更新: {len(remaining)}条')

print('\n完成！')

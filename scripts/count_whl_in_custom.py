# -*- coding: utf-8 -*-
"""检查定制链接中宽高+长格式的规格"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

count_whl = 0
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1: continue
    total += 1
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    # 宽+高+长
    if re.search(r'宽[【\[]\s*[\d.]+.*?高[【\[]\s*[\d.]+.*?(?:长度?)[【\[]', s):
        count_whl += 1
    if total >= 5000:
        break

wb.close()
print(f'采样{total}条定制:')
print(f'  含宽高+长格式: {count_whl}')

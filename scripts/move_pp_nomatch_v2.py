# -*- coding: utf-8 -*-
"""处理品牌店151条无匹配"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 列出换绑输出目录下所有文件
files = os.listdir(out)
for f in files:
    print(repr(f))

# 找品牌店_待确认文件
wm_file = None
for f in files:
    if '品牌' in f and '待确认' in f:
        wm_file = f
        break

if wm_file is None:
    print('\n找不到品牌店_待确认文件')
else:
    full_path = os.path.join(out, wm_file)
    print(f'\n找到: {full_path}')
    
    wb = oxl.load_workbook(full_path, data_only=True)
    ws = wb['待确认']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    
    print(f'条数: {len(rows)}')
    
    # 追加到无匹配
    wb2 = oxl.load_workbook(os.path.join(out, '无匹配_待处理.xlsx'), data_only=True)
    ws2 = wb2['无匹配']
    for r in rows:
        if len(r) >= 6:
            ws2.append(list(r[:6]))
        else:
            ws2.append(list(r) + ['', ''])
    wb2.save(os.path.join(out, '无匹配_待处理.xlsx'))
    wb2.close()
    
    # 删除待确认
    os.remove(full_path)
    print(f'✅ 已追加 {len(rows)}条 到 无匹配_待处理.xlsx')
    print(f'✅ 已删除 {wm_file}')

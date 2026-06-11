# -*- coding: utf-8 -*-
"""检查定制链接文件前几行的店铺简称"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

f = r'D:\Desktop\定制链接商品.xlsx'
wb = load_workbook(f, read_only=True)
ws = wb.active

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i > 5: break
    for j, v in enumerate(row):
        if v:
            print(f'  行{i} 列{j}: {str(v)[:50]}')
    print('---')
wb.close()

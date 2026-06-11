# -*- coding: utf-8 -*-
"""批量处理平卡中所有天猫店铺（淘宝品牌店、俊鑫、当下家、扣底盒、彩色）"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# ============================================================
# 快麦索引
# ============================================================
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
exact_idx = {}
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                exact_idx[(vals[0], vals[1], vals[2], dk, mat)] = code
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dim_fmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}", f"{base}-{dk}-特硬"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            fuzz2 = fuzzy_idx.get((tuple(sorted([il, iw, ih]))[0], tuple(sorted([il, iw, ih]))[1], tuple(sorted([il, iw, ih]))[2], '内径'), [])
            if not fuzz2:
                k3 = (dims[0], dims[1], dims[2], dk)
                dc = dim_dk_idx.get(k3, [])
                if dc: return dc[0]
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
    '天猫扣底盒': '飞机盒扣底盒专卖店',
    '天猫止合': '飞机盒止合专卖店',
    '淘宝当下家': '当下家包装',
    '淘宝俊鑫': '俊鑫纸品厂',
    '淘宝品牌店': '飞机盒品牌店',
}

def full_name(short):
    return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# ============================================================
# 读取平卡
# ============================================================
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"平卡总条数: {len(rows)}", flush=True)

# ============================================================
# 正则预编译
# ============================================================
# 淘宝品牌店: 牛皮色;长度23cm;宽17cm;高度：10cm
RE_PP_LEN = re.compile(r'长度\s*(\d+\.?\d*)\s*cm')
RE_PP_WID = re.compile(r'宽[度]?\s*(\d+\.?\d*)\s*cm')
RE_PP_HGT = re.compile(r'高度[：:]?\s*(\d+\.?\d*)\s*cm')
RE_PP_HGT2 = re.compile(r'高度\s*[：:]\s*(\d+\.?\d*)\s*cm')

# 俊鑫: 内尺寸【400*400】mm 长*宽;【100 mm 】
RE_JX_DIMS = re.compile(r'[【】\s]*(\d+)\s*\*+\s*(\d+)[】\s]*mm?\s*长\*宽')
RE_JX_HEIGHT = re.compile(r'[【】\s]*(\d+\.?\d*)\s*mm?\s*')

# 当下家: 10cm内高【白色】100个;10*10【长宽内径】
RE_DXJ = re.compile(r'(\d+\.?\d*)\s*cm\s*内高.*?(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)')

# 扣底盒: 31*31 cm;12cm 高;五层
RE_KOUDI = re.compile(r'(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*cm?\s*;\s*(\d+\.?\d*)\s*cm?\s*高')

# 彩色: 宽度：20 cm---高度：5 cmm;白色 100个 长度：20 cm
RE_CAISE = re.compile(
    r'宽度[：:]\s*(\d+\.?\d*)\s*cm?\s*-+\s*高度[：:]\s*(\d+\.?\d*)\s*cm?\s*;'
    r'\s*白色\s+\d+\s*个\s*长度[：:]\s*(\d+\.?\d*)\s*cm?'
)
# 有cmm写错的
RE_CAISE_CMM = re.compile(
    r'宽度[：:]\s*(\d+\.?\d*)\s*cm?\s*-+\s*高度[：:]\s*(\d+\.?\d*)\s*cmm\s*;'
    r'\s*白色\s+\d+\s*个\s*长度[：:]\s*(\d+\.?\d*)\s*cm?'
)

# 材料
RE_WHITE = re.compile(r'白色|双白色|双面白')
RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_BLACK = re.compile(r'黑色|黑')
RE_RED = re.compile(r'红色|红')
RE_FIVELAYER = re.compile(r'五层|5层')
RE_THREELAYER = re.compile(r'三层|3层')
RE_BLACK_RED = re.compile(r'黑.*红|红.*黑')

def guess_mat(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_WHITE.search(s): return '白色'
    if RE_RED.search(s): return '红色'
    if RE_BLACK.search(s): return '黑色'
    if RE_FIVELAYER.search(s): return 'EB'
    if RE_THREELAYER.search(s): return '3B'
    return '特硬'

matched_new = []
custom_new = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop_short = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    handled = False
    
    # ============================================================
    # 1. 扣底盒 X*X cm;Xcm 高;五层/三层 → EB/3B
    # ============================================================
    if not handled and '扣底盒' in shop_short:
        m = RE_KOUDI.search(spec_name)
        if m:
            x = float(m.group(1))
            y = float(m.group(2))
            h = float(m.group(3))
            mat = guess_mat(spec_name)  # 五层→EB
            code = match_3d(x, y, h, '外径', mat)
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['扣底盒-匹配'] += 1
            else:
                # 尝试特硬
                code = match_3d(x, y, h, '外径', '特硬')
                if code:
                    matched_new.append((full_name(shop_short), pid, spec_id, code))
                    stats['扣底盒-特硬兜底'] += 1
                else:
                    remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', f"{dim_fmt(max(x,y,h))}*{dim_fmt(sorted([x,y,h])[1])}*{dim_fmt(min(x,y,h))}-外径-{mat}"))
                    stats['扣底盒-无匹配'] += 1
            handled = True
    
    # ============================================================
    # 2. 彩色 宽度：x---高度：y;白色 长度：z
    # ============================================================
    if not handled and '彩色' in shop_short:
        m = RE_CAISE.search(spec_name) or RE_CAISE_CMM.search(spec_name)
        if m:
            w = float(m.group(1))
            h = float(m.group(2))
            l = float(m.group(3))
            if h < 0.5: h = h * 10  # cmm可能写错
            code = match_3d(l, w, h, '外径', '白色')
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['彩色-匹配'] += 1
            else:
                dims = sorted([l, w, h], reverse=True)
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}-外径-白色"))
                stats['彩色-无匹配'] += 1
            handled = True
    
    # ============================================================
    # 3. 淘宝品牌店: 牛皮色;长度23cm;宽17cm;高度：10cm
    # ============================================================
    if not handled and '品牌店' in shop_short:
        m_l = RE_PP_LEN.search(spec_name)
        m_w = RE_PP_WID.search(spec_name)
        m_h = RE_PP_HGT.search(spec_name) or RE_PP_HGT2.search(spec_name)
        if m_l and m_w and m_h:
            l = float(m_l.group(1))
            w = float(m_w.group(1))
            h = float(m_h.group(1))
            mat = guess_mat(spec_name)
            code = match_3d(l, w, h, '外径', mat)
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['品牌店-匹配'] += 1
            else:
                dims = sorted([l, w, h], reverse=True)
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}-外径-{mat}"))
                stats['品牌店-无匹配'] += 1
            handled = True
    
    # ============================================================
    # 4. 俊鑫: 内尺寸【400*400】mm 长*宽;【100 mm】
    # ============================================================
    if not handled and '俊鑫' in shop_short:
        m = RE_JX_DIMS.search(spec_name)
        if m:
            x = float(m.group(1)) / 10.0  # mm→cm
            y = float(m.group(2)) / 10.0
            # 找高度
            h_match = RE_JX_HEIGHT.search(spec_name)
            h = None
            if h_match:
                h = float(h_match.group(1)) / 10.0
            
            if h and h > 0.5:
                # 内尺寸→内径
                dims = sorted([x, y, h], reverse=True)
                mat = guess_mat(spec_name)
                code = match_3d(dims[0], dims[1], dims[2], '内径', mat)
                if code:
                    matched_new.append((full_name(shop_short), pid, spec_id, code))
                    stats['俊鑫-匹配'] += 1
                else:
                    remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}-内径-{mat}"))
                    stats['俊鑫-无匹配'] += 1
                handled = True
    
    # ============================================================
    # 5. 当下家: 10cm内高【白色】100个;10*10【长宽内径】
    # ============================================================
    if not handled and '当下家' in shop_short:
        m = RE_DXJ.search(spec_name)
        if m:
            h_cm = float(m.group(1))
            x = float(m.group(2))
            y = float(m.group(3))
            dims = sorted([x, y, h_cm], reverse=True)
            mat = guess_mat(spec_name)
            code = match_3d(dims[0], dims[1], dims[2], '内径', mat)
            if code:
                matched_new.append((full_name(shop_short), pid, spec_id, code))
                stats['当下家-匹配'] += 1
            else:
                remaining.append((shop_short, pid, spec_id, spec_name, '无匹配', f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}-内径-{mat}"))
                stats['当下家-无匹配'] += 1
            handled = True
    
    if not handled:
        remaining.append(r)

print(f"\n统计:", flush=True)
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}", flush=True)
print(f"\n匹配成功: {len(matched_new)}", flush=True)
print(f"定制链接: {len(custom_new)}", flush=True)
print(f"仍留平卡: {len(remaining)}", flush=True)

# ============================================================
# 换绑第8批
# ============================================================
if matched_new:
    f8 = os.path.join(out, '换绑文件_第8批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f8)
    print(f"✅ 换绑_第8批: {f8} ({os.path.getsize(f8)/1024:.1f}KB, {len(matched_new)}条)", flush=True)
    wb.close()

# ============================================================
# 定制类第5批
# ============================================================
if custom_new:
    f5 = os.path.join(out, '定制类_换绑文件_第5批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for c in custom_new:
        ws.append(list(c))
    wb.save(f5)
    print(f"✅ 定制类_第5批: {f5} ({os.path.getsize(f5)/1024:.1f}KB, {len(custom_new)}条)", flush=True)
    wb.close()

# ============================================================
# 更新平卡
# ============================================================
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: 剩余{len(remaining)}条", flush=True)
wb.close()

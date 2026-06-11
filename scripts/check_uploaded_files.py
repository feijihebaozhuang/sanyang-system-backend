# -*- coding: utf-8 -*-
"""检查已上传的换绑文件是否重复"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 已上传的文件（批量1-4,5,6,7,8,9,10,11 + 定制 + 主文件）
# 注意：换绑文件.xlsx(主文件)和第1批是重复的
files_to_check = [
    '换绑文件_第1批.xlsx',
    '换绑文件_第2批.xlsx',
    '换绑文件_第3-1批.xlsx',
    '换绑文件_第3-2批.xlsx',
    '换绑文件_第5批.xlsx',
    '换绑文件_第6批.xlsx',
    '换绑文件_第7批.xlsx',
    '换绑文件_第8批.xlsx',
    '换绑文件_第9批.xlsx',
    '换绑文件_第10批.xlsx',
    '换绑文件_第11批.xlsx',
    '定制类_换绑文件_第6批.xlsx',
]

# 收集每个文件的spec_id
file_specs = {}
all_specs = Counter()  # spec_id -> 出现次数

for f in files_to_check:
    fp = os.path.join(out, f)
    if not os.path.exists(fp):
        print(f'❌ {f}: 文件不存在')
        continue
    specs = []
    try:
        wb = oxl.load_workbook(fp)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r and len(r) >= 3 and r[2]:
                sid = str(r[2]).strip()
                specs.append(sid)
                all_specs[sid] += 1
        wb.close()
        file_specs[f] = specs
        print(f'  {f}: {len(specs)}条')
    except Exception as e:
        print(f'❌ {f}: 读取失败 {e}')

# 检查文件间重复
print(f'\n=== 文件间重复检查 ===')
total_unique = set()
total_all = 0
for f, specs in file_specs.items():
    before = len(total_unique)
    total_unique.update(specs)
    new = len(total_unique) - before
    dup = len(specs) - new
    total_all += len(specs)
    print(f'  {f}: {len(specs)}条, 新增唯一={new}, 与之前重复={dup}')

print(f'\n总条数(含重复): {total_all}')
print(f'去重后唯一spec_id: {len(total_unique)}')

# 看哪些spec_id重复了
multi = {k: v for k, v in all_specs.items() if v > 1}
print(f'\n跨文件重复spec_id: {len(multi)}个')
if multi:
    print(f'重复最多的前10个:')
    for sid, cnt in Counter(multi).most_common(10):
        # 找出现在哪些文件有
        in_files = [f for f, specs in file_specs.items() if sid in specs]
        print(f'  {sid[:30]}... 出现{cnt}次, 在: {in_files}')

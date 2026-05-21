#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
7 列 Excel 导入刀模库（覆盖/追加）。表头与改表头前一致，不读「生产规格」「快麦商品映射」。

用法:
  python3 scripts/import_dimoldb_7col.py 刀模.xlsx --overwrite
  python3 scripts/import_dimoldb_7col.py 刀模.xlsx --append
"""
from __future__ import annotations

import argparse
import datetime
import os
import sys
import time

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import dimoldb_store as ds
from settings import DB_CONFIG  # noqa: E402

import pymysql


def get_db():
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)


def load_workbook(path: str):
    import openpyxl

    return openpyxl.load_workbook(path, data_only=True)


def find_header_row(ws) -> tuple[int, list[str]]:
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [str(v or "").strip() for v in row]
        if "名称" in vals or "刀模名称" in vals:
            return r_idx, vals
    raise SystemExit("未找到表头行（需包含「名称」列）")


def import_file(path: str, *, overwrite: bool) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_row, headers = find_header_row(ws)
    col_map = ds.map_dimoldb_import_headers(headers)
    if "name" not in col_map:
        raise SystemExit(f"未找到名称列，表头: {headers}")

    db = [] if overwrite else ds.load_dimoldb(get_db, force=True)
    by_code = {str(d.get("code") or "").strip(): i for i, d in enumerate(db) if d.get("code")}
    by_name = {str(d.get("name") or "").strip(): i for i, d in enumerate(db)}
    added = updated = 0
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = (
            str(row[col_map["name"]]).strip()
            if col_map["name"] < len(row) and row[col_map["name"]]
            else ""
        )
        if not name or name.startswith("=") or name == "None":
            continue
        item = {
            "id": f"dm_{int(time.time())}_{added}_{len(db)}",
            "name": name,
            "product_type": ds.cell_str(row, col_map.get("product_type")) or "zhengsquare",
            "code": ds.cell_str(row, col_map.get("code")),
            "production_spec": "",
            "km_mapping_code": "",
            "remark": ds.cell_str(row, col_map.get("remark")),
            "length": ds.cell_float(row, col_map.get("length")),
            "width": ds.cell_float(row, col_map.get("width")),
            "height": ds.cell_float(row, col_map.get("height")),
            "stock": 0,
            "created_at": now,
        }
        key_code = (item.get("code") or "").strip()
        key_name = item.get("name") or ""
        idx = by_code.get(key_code) if key_code else None
        if idx is None and key_name:
            idx = by_name.get(key_name)
        if idx is not None:
            old = db[idx]
            item["id"] = old.get("id") or item["id"]
            item["created_at"] = old.get("created_at") or now
            db[idx] = {**old, **item}
            updated += 1
        else:
            db.append(item)
            if key_code:
                by_code[key_code] = len(db) - 1
            by_name[key_name] = len(db) - 1
            added += 1

    if not ds.save_dimoldb(get_db, db):
        raise SystemExit("save_dimoldb 失败")
    ds.invalidate_dimoldb_cache()
    print(f"完成: 新增 {added} 更新 {updated} 合计 {len(db)} 条 (overwrite={overwrite})")


def main() -> None:
    parser = argparse.ArgumentParser(description="7 列刀模 Excel 导入")
    parser.add_argument("xlsx", help="Excel 路径")
    parser.add_argument("--overwrite", action="store_true", help="清空后全量导入")
    parser.add_argument("--append", action="store_true", help="追加（默认）")
    args = parser.parse_args()
    if not os.path.isfile(args.xlsx):
        raise SystemExit(f"文件不存在: {args.xlsx}")
    import_file(args.xlsx, overwrite=args.overwrite and not args.append)


if __name__ == "__main__":
    main()

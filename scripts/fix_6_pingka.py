# -*- coding: utf-8 -*-
"""
处理平卡中的6条记录
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 加载快麦索引
import pandas as pd
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
km_map = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if code:
        km_map[code] = True

# 构建索引
exact_idx = {}
fuzzy_idx = {}
all_codes = set(km_map.keys())
for code in all_codes:
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                ek = (vals[0], vals[1], vals[2], dk, mat)
                exact_idx[ek] = code
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
            except:
                pass

def dim_fmt(v):
    if v == int(v):
        return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dim_fmt(dims[0])}*{dim_fmt(dims[1])}*{dim_fmt(dims[2])}"
    
    cands = [f"{base}-{dk}-{mat}"]
    if mat == '白色':
        cands.append(f"{base}-{dk}-特硬")
    cands.append(f"{base}-{dk}-特硬")
    
    for c in cands:
        if c in all_codes:
            return c
    
    vals = tuple(sorted(dims))
    fk = (vals[0], vals[1], vals[2], dk)
    fuzz = fuzzy_idx.get(fk, [])
    if fuzz:
        for c in fuzz:
            if mat in c:
                return c
        return fuzz[0]
    
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fk2 = (ivals[0], ivals[1], ivals[2], '内径')
            fuzz2 = fuzzy_idx.get(fk2, [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c:
                        return c
                return fuzz2[0]
    
    return None

SHOP_NAME_MAP = {
    '天猫小批量': '飞机盒小批量专卖店',
    '天猫正方形': '飞机盒正方形专卖店',
    '天猫彩色': '飞机盒彩色专卖店',
}
def get_full_name(short):
    return SHOP_NAME_MAP.get(short, short)

# 读取平卡文件
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

matched_new = []  # (店铺全称, pid, 规格id, 编码)
custom_new = []   # 定制链接

# 37*14**3.8 → 37, 14, 3.8 （**替换为*）
RE_37 = re.compile(r'37\s*\*+\s*14\s*\*+\s*3\.8')
RE_40 = re.compile(r'40\.12\s*\*+\s*7')

remaining = []
for r in rows:
    if not r: continue
    shop_short = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    is_custom = False
    is_matched = None
    
    # 双面纯色黑&红 → 定制
    if '双面纯色' in spec_name and '黑&红' in spec_name:
        custom_new.append((get_full_name(shop_short), pid, spec_id, '定制链接'))
        is_custom = True
    
    # 37*14**3.8 系列
    elif '37*14' in spec_name and '3.8' in spec_name:
        dk = '外径'
        if '双白色' in spec_name:
            mat = '白色'
        elif '优质进口纸' in spec_name:
            mat = '特硬'
        else:
            mat = '特硬'
        is_matched = match_3d(37, 14, 3.8, dk, mat)
        if is_matched:
            matched_new.append((get_full_name(shop_short), pid, spec_id, is_matched))
        else:
            remaining.append(r)
            continue
    
    # 40.12*7 系列 → 40 12 7
    elif '40.12' in spec_name:
        dk = '外径'
        if '双白色' in spec_name:
            mat = '白色'
        elif '优质进口纸' in spec_name:
            mat = '特硬'
        else:
            mat = '特硬'
        is_matched = match_3d(40, 12, 7, dk, mat)
        if is_matched:
            matched_new.append((get_full_name(shop_short), pid, spec_id, is_matched))
        else:
            remaining.append(r)
            continue
    
    else:
        remaining.append(r)

print(f"处理结果:", flush=True)
print(f"  匹配成功: {len(matched_new)}")
print(f"  定制链接: {len(custom_new)}")
print(f"  仍留平卡: {len(remaining)}")

# 写入换绑文件_第6批.xlsx
if matched_new:
    f6 = os.path.join(out, '换绑文件_第6批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for m in matched_new:
        ws.append(list(m))
    wb.save(f6)
    print(f"✅ 换绑_第6批: {f6} ({os.path.getsize(f6)/1024:.1f}KB, {len(matched_new)}条)", flush=True)
    wb.close()

# 定制类第3批
if custom_new:
    f3 = os.path.join(out, '定制类_换绑文件_第3批.xlsx')
    wb = oxl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.append([None, '商品对应表', None, None])
    ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    for r in custom_new:
        ws.append(list(r))
    wb.save(f3)
    print(f"✅ 定制类_第3批: {f3} ({os.path.getsize(f3)/1024:.1f}KB, {len(custom_new)}条)", flush=True)
    wb.close()

# 更新平卡文件
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    ws.append(list(r))
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
print(f"✅ 平卡_待处理: 剩余{len(remaining)}条", flush=True)
wb.close()

print(f"\n{'='*50}")
print("完成！可上传：换绑文件_第6批.xlsx、定制类_换绑文件_第3批.xlsx")
print(f"{'='*50}")

# -*- coding: utf-8 -*-
"""
v8: 终极修复
- 增加 高度+长*宽 格式: 外径】高度【Hmm】;长*宽【LxWmm】
- 全面mm→cm转换
- 优先选择独立维度值
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
import openpyxl as opx
import pandas as pd

source = r'D:\Desktop\平台商品.xlsx'
print('读取平台商品.xlsx ...')
wb = load_workbook(source, read_only=True)
ws = wb.active

header = None
rows_list = []
total = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if i == 1:
        header = list(row)
        continue
    total += 1
    rows_list.append(row)
wb.close()
print(f'总数: {total}')

def norm_val(v_str, unit=None):
    """将带单位的数值转成cm。v_str可以是'320'、'320mm'等"""
    v = float(re.sub(r'[^\d.]', '', v_str))
    if unit == 'mm' or 'mm' in v_str:
        v = v / 10.0
    return v

def extract_all_dims(s):
    """全覆盖提取LWH，返回list of (L,W,H,src)"""
    results = []
    
    # ===== 独立维度格式 =====
    # 1. 宽【Wcm】高【Hcm】... 长【Lcm】
    m = re.search(r'宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】.*?长【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(3)), float(m.group(1)), float(m.group(2)), 'A1'))
    
    # 2. 【长度 L 厘米】【宽度 W 厘米】（中文"厘米"）
    m = re.search(r'【长度\s*([\d.]+)\s*厘米\s*】.*?【宽度\s*([\d.]+)\s*厘米\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), 0, 'A2'))
    
    # 3. 长【L】宽【W】高【H】
    m = re.search(r'长【\s*([\d.]+)\s*cm?\s*】.*?宽【\s*([\d.]+)\s*cm?\s*】.*?高【\s*([\d.]+)\s*cm?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A3'))
    
    # 4. 飞机盒【长度L】...【宽度W】X【高度H】
    m = re.search(r'飞机盒【长度\s*([\d.]+)\s*CM?\s*】.*?【宽度\s*([\d.]+)\s*CM?\s*】\s*X\s*【高度\s*([\d.]+)\s*CM?\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'A4'))
    
    # 5. 外径格式: 外径 130x100x20 mm
    m = re.search(r'外径[：:]*\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)(?:\s*mm)?', s)
    if m:
        v = [float(m.group(1)), float(m.group(2)), float(m.group(3))]
        if any(x > 50 for x in v):
            v = [x/10 for x in v]
        results.append((v[0], v[1], v[2], 'A5'))
    
    # ===== 6. 外径 + 高度 + 长*宽（2值+1值格式）=====
    # 格式A: 白色外径】高度【100mm】;【100个】长*宽【320*310mm】（外径在前，标准括号）
    # 格式B: 高度【100mm】黄色【外径】;【100个】长*宽【320*310mm】（外径在后，标准括号）
    # 格式C: 白色外径】高度【100mm】;【100个】长*宽360*300mm】（缺左括号）
    # → L=320mm=32cm, W=310mm=31cm, H=100mm=10cm
    m = re.search(r'(?:'
                  r'外径[^；;]*高度【\s*(\d+\.?\d*)\s*mm?\s*】'
                  r'|'
                  r'高度【\s*(\d+\.?\d*)\s*mm?\s*】[^；;]*外径'
                  r')'
                  r'.*?(?:长\s*\*\s*宽)?【?\s*(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)\s*mm?\s*】', s)
    if m:
        h_val = m.group(1) if m.group(1) else m.group(2)
        h = float(h_val) / 10.0
        l = float(m.group(3)) / 10.0
        w = float(m.group(4)) / 10.0
        results.append((l, w, h, 'A6外径2+1'))
    
    # ===== 三值合体格式 =====
    # 7. 【LxWxH】
    m = re.search(r'【\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*】', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B1'))
    
    # 8. 长LxWxH
    m = re.search(r'[长L]+\s*[：:]?\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)\s*(?:cm|mm)?\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B2'))
    
    # 9. ;LxWxH在末尾
    m = re.search(r'(?:[；;]\s*|^)(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*$', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B3'))
    
    # 10. 裸LxWxH
    m = re.search(r'(?:^|[\s；;])(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)\s*[xX*]\s*(\d+\.?\d*)', s)
    if m: results.append((float(m.group(1)), float(m.group(2)), float(m.group(3)), 'B4'))
    
    return results

def build_lwh(s):
    """从extract_all_dims结果构建最合理的(L,W,H)"""
    all_dims = extract_all_dims(s)
    if not all_dims:
        return None
    
    # 优先选完整3值且值合理
    for l, w, h, src in all_dims:
        if h > 0.1 and 0.5 <= l <= 200 and 0.5 <= w <= 200 and 0.5 <= h <= 200:
            return (l, w, h)
    
    # 降级：有2个好值就行
    for l, w, h, src in all_dims:
        if 0.5 <= l <= 200 and 0.5 <= w <= 200:
            if h > 0.1:
                return (l, w, h)
            return (l, w, h)
    
    # 最后回退
    if all_dims:
        l, w, h, src = all_dims[0]
        return (l, w, h)
    return None

def has_any_dim(s):
    return len(extract_all_dims(s)) > 0

def classify_dims(lwh):
    l, w, h = lwh
    has_decimal = any(v != int(v) for v in lwh)
    all_5 = all(v % 1 == 0.5 for v in lwh)
    if has_decimal and not all_5: return '非全量飞机盒'
    elif all_5: return '全.5'
    else: return '整数'

# === 全量分类 ===
custom_rows = []
dikoudi_rows = []
zhixiang_rows = []
neijing_rows = []
waijing_rows = []
feiji_rows = []
other_rows = []

for idx, row in enumerate(rows_list):
    s = str(row[2] or '').strip() if len(row) > 2 else ''
    
    # 1. 定制链接关键词 + 无尺寸
    if ('定制' in s or '珍珠棉' in s or '咨询客服' in s) and not has_any_dim(s):
        custom_rows.append(row)
        continue
    
    # 2. 扣底盒/双插盒
    if '扣底盒' in s or '双插盒' in s:
        dikoudi_rows.append(row)
        continue
    
    # 3. 纸箱
    if '纸箱' in s:
        zhixiang_rows.append(row)
        continue
    
    # 4. 提取尺寸
    lwh = build_lwh(s)
    if lwh is None:
        custom_rows.append(row)
        continue
    
    dt = classify_dims(lwh)
    
    # 5. 内径
    if '内径' in s or '内尺寸' in s or '内寸' in s:
        neijing_rows.append(row)
        continue
    
    # 6. 非全量飞机盒
    if dt == '非全量飞机盒':
        feiji_rows.append(row)
        continue
    
    # 7. 外径全量飞机盒
    if '外径' in s:
        waijing_rows.append(row)
        continue
    
    # 8. 其余
    other_rows.append(row)

print(f'定制链接: {len(custom_rows)}')
print(f'扣底盒/双插盒: {len(dikoudi_rows)}')
print(f'纸箱: {len(zhixiang_rows)}')
print(f'内径全量飞机盒: {len(neijing_rows)}')
print(f'外径全量飞机盒: {len(waijing_rows)}')
print(f'非全量飞机盒: {len(feiji_rows)}')
print(f'其余: {len(other_rows)}')
s = len(custom_rows)+len(dikoudi_rows)+len(zhixiang_rows)+len(neijing_rows)+len(waijing_rows)+len(feiji_rows)+len(other_rows)
print(f'和: {s} (应={total})')
if s != total:
    print(f'!!! 差值: {total - s}')

def write_xlsx(data, fpath, label):
    if not data:
        print(f'  {fpath}: 空，跳过')
        return
    pd.DataFrame(data, columns=header).to_excel(fpath, index=False)
    wb2 = opx.load_workbook(fpath)
    ws2 = wb2.active
    ws2.insert_rows(1)
    ws2.cell(1, 2).value = label
    ws2.column_dimensions['D'].width = 60
    for c in ['F','G','H','I']:
        try: ws2.column_dimensions[c].width = 10
        except: pass
    wb2.save(fpath)
    print(f'  ✅ {fpath}')

write_xlsx(custom_rows, r'D:\Desktop\定制链接商品.xlsx', '定制链接商品')
write_xlsx(dikoudi_rows, r'D:\Desktop\扣底盒双插盒商品.xlsx', '扣底盒双插盒商品')
write_xlsx(zhixiang_rows, r'D:\Desktop\纸箱商品.xlsx', '纸箱商品')
write_xlsx(neijing_rows, r'D:\Desktop\内径全量飞机盒.xlsx', '内径全量飞机盒')
write_xlsx(waijing_rows, r'D:\Desktop\外径全量飞机盒.xlsx', '外径全量飞机盒')
write_xlsx(feiji_rows, r'D:\Desktop\非全量飞机盒.xlsx', '非全量飞机盒')
write_xlsx(other_rows, r'D:\Desktop\其余商品.xlsx', '其余商品')

print('\n✅ 全部完成！')

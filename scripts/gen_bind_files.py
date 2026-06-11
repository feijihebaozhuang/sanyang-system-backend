# -*- coding: utf-8 -*-
"""按照模板格式重新生成换绑文件"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl

src = r"d:\Desktop\换绑输出\换绑文件.xlsx"
out_dir = r"d:\Desktop\换绑输出"

print("读取已生成的换绑文件...", flush=True)
wb = oxl.load_workbook(src, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
total = len(rows)
print(f"总数据: {total} 条", flush=True)

# 按模板格式重新生成：第1行"商品对应表"，第2行列头
n = 3
batch = (total + n - 1) // n

for i in range(n):
    start = i * batch
    end = min((i + 1) * batch, total)
    chunk = rows[start:end]
    
    wb2 = oxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'Sheet1'
    # 第1行：商品对应表
    ws2.append([None, '商品对应表', None, None])
    # 第2行：列头
    ws2.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    # 数据
    for r in chunk:
        ws2.append(list(r))
    
    f = os.path.join(out_dir, f'换绑文件_第{i+1}批.xlsx')
    wb2.save(f)
    sz = os.path.getsize(f)
    print(f"第{i+1}批: {sz/1024/1024:.2f}MB, {len(chunk)}条 => {f}", flush=True)
    wb2.close()

wb.close()
print("\n完成！每个文件第1行都有「商品对应表」", flush=True)

# -*- coding: utf-8 -*-
"""处理品牌店剩余2063条：宽度【X*Y】cm【100个起拍】;外径【材料】长度 Z cm"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl
import pandas as pd
from collections import Counter

out = r'd:\Desktop\换绑输出'

# 快麦索引
km_file = r'd:\Desktop\快麦商品 - 副本.xlsx'
df = pd.read_excel(km_file, sheet_name='报表1', header=3, dtype=str, usecols=[0,1,2])
all_codes = set()
fuzzy_idx = {}
dim_dk_idx = {}
for row in df.values:
    code = str(row[0] or '').strip()
    if not code: continue
    all_codes.add(code)
    parts = code.split('-')
    if len(parts) >= 3:
        dims = parts[0].split('*')
        if len(dims) == 3:
            try:
                vals = tuple(sorted([float(d) for d in dims]))
                dk = parts[1]
                mat = '-'.join(parts[2:])
                fk = (vals[0], vals[1], vals[2], dk)
                fuzzy_idx.setdefault(fk, []).append(code)
                k3 = (float(dims[0]), float(dims[1]), float(dims[2]), dk)
                dim_dk_idx.setdefault(k3, []).append(code)
            except: pass

def dfmt(v):
    if v == int(v): return str(int(v))
    return f"{v:.1f}"

def match_3d(l, w, h, dk, mat):
    dims = sorted([l, w, h], reverse=True)
    base = f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}"
    cands = [f"{base}-{dk}-{mat}", f"{base}-{dk}-特硬"]
    for c in cands:
        if c in all_codes: return c
    vals = tuple(sorted(dims))
    fuzz = fuzzy_idx.get((vals[0], vals[1], vals[2], dk), [])
    if fuzz:
        for c in fuzz:
            if mat in c: return c
        return fuzz[0]
    if dk == '外径':
        il, iw, ih = dims[0]-1.5, dims[1]-0.5, dims[2]-0.5
        if il > 0 and iw > 0 and ih > 0:
            ivals = tuple(sorted([il, iw, ih]))
            fuzz2 = fuzzy_idx.get((ivals[0], ivals[1], ivals[2], '内径'), [])
            if fuzz2:
                for c in fuzz2:
                    if mat in c: return c
                return fuzz2[0]
    k3 = (dims[0], dims[1], dims[2], dk)
    dc = dim_dk_idx.get(k3, [])
    if dc: return dc[0]
    return None

SHOP_NAME_MAP = {'淘宝品牌店': '飞机盒品牌店'}
def fn(short): return SHOP_NAME_MAP.get(str(short).strip(), str(short).strip())

# 正则：宽度【11*10】cm【100个起拍】;外径【台湾纸】长度 35 cm
# 宽*高 在【】里, 材料在外径【】里, 长度在末尾
RE_PP_WH = re.compile(
    r'宽度[【】\s]*(\d+\.?\d*)\s*\*+\s*(\d+\.?\d*)\s*[】]*\s*cm?\s*[【】][^】]*[】]\s*;\s*'
    r'外径[【】\s]*([^】]*)[】]*\s*长度\s*(\d+\.?\d*)\s*cm?'
)

RE_CHAOYING = re.compile(r'台湾纸|台湾|超硬')
RE_BAISE = re.compile(r'双面白色|双面白|白色|双白')

def guess_mat(s):
    if RE_CHAOYING.search(s): return '超硬'
    if RE_BAISE.search(s): return '白色'
    return '特硬'

# 读平卡
wb = oxl.load_workbook(os.path.join(out, '平卡_待处理.xlsx'), data_only=True)
ws = wb['平卡']
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()
print(f"平卡总数: {len(rows)}", flush=True)

matched_new = []
remaining = []
stats = Counter()

for r in rows:
    if not r: continue
    shop = str(r[0] or '').strip()
    pid = str(r[1] or '').strip()
    spec_id = str(r[2] or '').strip()
    spec_name = str(r[3] or '').strip()
    
    if '品牌店' not in shop:
        remaining.append(r)
        continue
    
    m = RE_PP_WH.search(spec_name)
    if m:
        w_cm = float(m.group(1))
        h_cm = float(m.group(2))
        mat_ctx = m.group(3) or ''
        l_cm = float(m.group(4))
        
        l, w, h = l_cm, w_cm, h_cm
        dk = '外径'
        mat = guess_mat(spec_name)
        dims = sorted([l, w, h], reverse=True)
        
        code = match_3d(dims[0], dims[1], dims[2], dk, mat)
        if code:
            matched_new.append((fn(shop), pid, spec_id, code))
            stats['匹配'] += 1
        else:
            remaining.append((shop, pid, spec_id, spec_name, '无匹配', 
                f"{dfmt(dims[0])}*{dfmt(dims[1])}*{dfmt(dims[2])}-{dk}-{mat}"))
            stats['无匹配'] += 1
    else:
        remaining.append(r)

print(f"\n品牌店: 匹配{stats['匹配']}, 无匹配{stats['无匹配']}")
print(f"平卡剩余: {len(remaining)}")

if matched_new:
    # 看是否<3MB，不够就追加到第10批
    f10 = os.path.join(out, '换绑文件_第10批.xlsx')
    f11 = os.path.join(out, '换绑文件_第11批.xlsx')
    
    wb10 = oxl.load_workbook(f10, data_only=True)
    existing = list(wb10['Sheet1'].iter_rows(min_row=3, values_only=True))
    wb10.close()
    
    total = len(existing) + len(matched_new)
    
    if total > 100000:
        # 拆到第11批
        wb = oxl.Workbook()
        ws = wb.active; ws.title = 'Sheet1'
        ws.append([None, '商品对应表', None, None])
        ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
        for m in matched_new:
            ws.append(list(m))
        wb.save(f11)
        print(f"✅ 换绑_第11批: {f11} ({os.path.getsize(f11)/1024:.1f}KB, {len(matched_new)}条)")
        wb.close()
    else:
        # 追加到第10批
        wb = oxl.Workbook()
        ws = wb.active; ws.title = 'Sheet1'
        ws.append([None, '商品对应表', None, None])
        ws.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
        for m in existing:
            ws.append(list(m))
        for m in matched_new:
            ws.append(list(m))
        wb.save(f10)
        sz = os.path.getsize(f10)
        print(f"✅ 换绑_第10批(追加后): {sz/1024:.1f}KB, {len(existing)+len(matched_new)}条")
        wb.close()

# 更新平卡
wb = oxl.Workbook()
ws = wb.active; ws.title = '平卡'
ws.append(['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码'])
for r in remaining:
    if len(r) >= 6: ws.append(list(r[:6]))
    else: ws.append(list(r) + ['', ''])
wb.save(os.path.join(out, '平卡_待处理.xlsx'))
wb.close()
print("✅ 平卡已更新")

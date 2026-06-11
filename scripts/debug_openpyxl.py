# -*- coding: utf-8 -*-
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

s = '高12cm【五层】；长宽【29*21】100个;高12cm【五层】;长宽【29*21】100个'

# 直接用openpyxl读取并检查
import openpyxl
wb = openpyxl.load_workbook(r'd:\Desktop\换绑输出\平卡_待处理.xlsx', data_only=True)
ws = wb.active
for row_idx in range(2, 10):
    shop = ws.cell(row_idx, 1).value
    spec = ws.cell(row_idx, 4).value
    if shop and '友尚' in str(shop):
        v = str(spec)
        print(f'row {row_idx}: {repr(v[:50])}')
        m = re.search(r'高(\d+)cm【\d+层】', v)
        print(f'  match: {m.group() if m else "NO"}')
        break

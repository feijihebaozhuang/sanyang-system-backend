# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

okdir = r'D:\Desktop\换绑输出\OK文件'

# 店铺简称 -> 全称映射
FULL_NAMES = {
    '友尚包装': '深圳市友尚包装有限公司',
}

batches = list(range(33, 40))

for b in batches:
    fname = f'\u6362\u7ed1\u6587\u4ef6_\u7b2c{b}\u6279.xlsx'
    fpath = os.path.join(okdir, fname)
    if not os.path.exists(fpath): continue
    
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    changed = 0
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v and v in FULL_NAMES:
            ws.cell(row, 1).value = FULL_NAMES[v]
            changed += 1
    
    wb.save(fpath)
    print(f'\u7b2c{b}\u6279: \u4fee\u6539\u4e86{changed}\u884c')
    
    # 处理拆分部分
    parts = [f for f in os.listdir(okdir) if f.startswith(f'\u6362\u7ed1\u6587\u4ef6_\u7b2c{b}\u6279_') and f.endswith('.xlsx')]
    for p in sorted(parts):
        ppath = os.path.join(okdir, p)
        pwb = openpyxl.load_workbook(ppath)
        pws = pwb.active
        pc = 0
        for row in range(1, pws.max_row + 1):
            v = pws.cell(row, 1).value
            if v and v in FULL_NAMES:
                pws.cell(row, 1).value = FULL_NAMES[v]
                pc += 1
        pwb.save(ppath)
        print(f'  {p}: \u4fee\u6539\u4e86{pc}\u884c')

print('\n\u5b8c\u6210!')

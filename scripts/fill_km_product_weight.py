#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快麦商品导出批量填重量（与三羊报价/重量测试同一公式）。

用法:
  python scripts/fill_km_product_weight.py 快麦商品.xlsx
  python scripts/fill_km_product_weight.py 快麦商品.xlsx -o 快麦商品_已填重量.xlsx
  python scripts/fill_km_product_weight.py 快麦商品.xlsx --dry-run --limit 20

支持 .xlsx / .csv / .tsv；默认只填空白重量列，原文件另存为 -o 路径。
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import quote_weight as qw

_HEADER_ALIASES: dict[str, str] = {
    "x": "length",
    "y": "width",
    "z": "height",
    "长": "length",
    "宽": "width",
    "高": "height",
    "length": "length",
    "width": "width",
    "height": "height",
    "skuouterid": "outer_id",
    "sysouterid": "outer_id",
    "outerid": "outer_id",
    "商家编码": "outer_id",
    "规格商家编码": "outer_id",
    "编码": "outer_id",
    "重量": "weight",
    "weight": "weight",
    "skuweight": "weight",
    "单品重量": "weight",
    "商品重量": "weight",
    "净重": "weight",
    "propertiesalias": "spec_text",
    "propertiesname": "spec_text",
    "规格别名": "spec_text",
    "销售属性": "spec_text",
    "规格": "spec_text",
    "platformspec": "spec_text",
    "shorttitle": "title",
    "title": "title",
    "商品名称": "title",
    "商品标题": "title",
    "名称": "title",
    "producttype": "product_type",
    "产品类型": "product_type",
    "material": "material",
    "材质": "material",
    "材料": "material",
}


def _norm_header(h: str) -> str:
    s = re.sub(r"[\s_（）()]", "", (h or "").strip().lower())
    for k, v in _HEADER_ALIASES.items():
        kk = re.sub(r"[\s_（）()]", "", k.lower())
        if s == kk or s.endswith(kk) or kk in s:
            return v
    return ""


def _detect_col_map(headers: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        field = _norm_header(str(h or ""))
        if field and field not in col:
            col[field] = i
    return col


def _cell(row: list, col_map: dict[str, int], name: str) -> str:
    idx = col_map.get(name)
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _float_cell(row: list, col_map: dict[str, int], name: str) -> float:
    s = _cell(row, col_map, name)
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else 0.0


def _weight_empty(val: str) -> bool:
    s = (val or "").strip()
    if not s:
        return True
    try:
        return float(s) <= 0
    except ValueError:
        return True


def _build_text(row: list, col_map: dict[str, int]) -> str:
    parts = [
        _cell(row, col_map, "title"),
        _cell(row, col_map, "spec_text"),
        _cell(row, col_map, "material"),
        _cell(row, col_map, "outer_id"),
    ]
    return " ".join(p for p in parts if p)


def _process_row(row: list, col_map: dict[str, int]) -> dict:
    l = _float_cell(row, col_map, "length")
    w = _float_cell(row, col_map, "width")
    h = _float_cell(row, col_map, "height")
    text = _build_text(row, col_map)
    if not (l and w and h) and text:
        import km_sku_map_store as kms

        pl, pw, ph, mat = kms.parse_spec_alias_dims(text)
        if pl and pw:
            l, w, h = pl, pw, ph or h
        if mat:
            text = f"{text} {mat}"
    ptype = _cell(row, col_map, "product_type")
    return qw.estimate_unit_weight(l, w, h, text=text, product_type=ptype)


def _read_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("需要 openpyxl: pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(x or "") for x in next(it, [])]
    col_map = _detect_col_map(headers)
    rows = [list(r or []) for r in it]
    wb.close()
    return headers, col_map, rows


def _write_xlsx(out_path: Path, headers: list[str], rows: list[list]):
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(out_path)


def _read_csv(path: Path, delimiter: str | None):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        if delimiter is None:
            delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.reader(f, delimiter=delimiter)
        headers = [str(h or "") for h in next(reader, [])]
        rows = [list(r) for r in reader]
    return headers, _detect_col_map(headers), rows, delimiter


def _write_csv(out_path: Path, headers: list[str], rows: list[list], delimiter: str):
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=delimiter)
        w.writerow(headers)
        w.writerows(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description="快麦商品导出批量填重量")
    ap.add_argument("input", type=Path, help="快麦商品 Excel/CSV")
    ap.add_argument("-o", "--output", type=Path, default=None, help="输出路径")
    ap.add_argument(
        "--weight-unit",
        choices=("g", "kg"),
        default="g",
        help="写入重量列的单位（快麦常见为克 g）",
    )
    ap.add_argument(
        "--overwrite",
        action="store_true",
        help="覆盖已有重量（默认只填空白）",
    )
    ap.add_argument("--dry-run", action="store_true", help="只统计不写文件")
    ap.add_argument("--limit", type=int, default=0, help="最多处理行数（调试）")
    args = ap.parse_args()

    inp = args.input
    if not inp.is_file():
        raise SystemExit(f"文件不存在: {inp}")

    delimiter = None
    suffix = inp.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xltx"):
        headers, col_map, rows = _read_xlsx(inp)
        is_xlsx = True
    else:
        headers, col_map, rows, delimiter = _read_csv(inp, None)
        is_xlsx = False

    if "weight" not in col_map:
        headers = list(headers) + ["重量"]
        col_map = dict(col_map)
        col_map["weight"] = len(headers) - 1
        rows = [list(r) + [""] for r in rows]

    if not any(k in col_map for k in ("length", "width", "height")):
        raise SystemExit(f"未找到 x/y/z 或 长/宽/高 列，当前表头: {headers[:20]}")

    w_idx = col_map["weight"]
    filled = failed = already = 0
    samples: list[str] = []

    for i, row in enumerate(rows):
        if args.limit and i >= args.limit:
            break
        while len(row) <= w_idx:
            row.append("")
        cur_w = _cell(row, col_map, "weight")
        if not args.overwrite and not _weight_empty(cur_w):
            already += 1
            continue
        est = _process_row(row, col_map)
        if not est.get("ok"):
            failed += 1
            continue
        val = (
            est["weight_per_unit_g"]
            if args.weight_unit == "g"
            else est["weight_per_unit_kg"]
        )
        row[w_idx] = round(float(val), 1 if args.weight_unit == "g" else 4)
        filled += 1
        if len(samples) < 5:
            oid = _cell(row, col_map, "outer_id") or f"row{i+2}"
            samples.append(
                f"  {oid}: {est.get('weight_per_unit_g')}g "
                f"({est.get('product_type')}/{est.get('material_key')})"
            )

    print(f"总行数: {len(rows)}")
    print(f"已填重量: {filled}  已有跳过: {already}  无法算: {failed}")
    if samples:
        print("样例:")
        print("\n".join(samples))

    if args.dry_run:
        print("(dry-run，未写文件)")
        return

    out = args.output or inp.with_name(inp.stem + "_已填重量" + inp.suffix)
    if is_xlsx:
        _write_xlsx(out, headers, rows)
    else:
        _write_csv(out, headers, rows, delimiter or ",")
    print(f"已写出: {out}")


if __name__ == "__main__":
    main()

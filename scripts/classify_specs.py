# -*- coding: utf-8 -*-
"""
一次性分析平台商品所有规格名称格式，归类输出给你看
"""
import openpyxl, sys, re
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

fp = r"d:\Desktop\平台商品.xlsx"
print("读取平台商品表...")
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
ws = wb['报表1']

# 收集所有规格名称样本
spec_samples = []
seen = set()
total = 0

for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
    shop, pid, spec_name, spec_id = row
    if not spec_name:
        continue
    total += 1
    s = str(spec_name).strip()
    if s not in seen:
        seen.add(s)
        spec_samples.append((s, shop, pid, spec_id))

wb.close()
print(f"总计 {total} 行, 去重后 {len(spec_samples)} 个")

# 分析模式
patterns = Counter()

for s, shop, pid, spec_id in spec_samples:
    if re.match(r'^[\d.*xX×\-]+$', s) or ('【' not in s and not re.search(r'[长宽高厚内径外径]', s)):
        patterns['纯数字/符号'] += 1
    elif '【' in s and ('长' in s or '宽' in s or '高' in s or '厚度' in s or '长度' in s):
        patterns['【】格式+有长宽高'] += 1
    elif '【' in s and not ('长' in s or '宽' in s or '高' in s):
        patterns['【】格式+无长宽高'] += 1
    elif ('长' in s or '宽' in s or '高' in s) and '【' not in s:
        patterns['有长宽高+无【】'] += 1
    elif re.search(r'\d+\.?\d*\s*[xX*×]\s*\d+', s):
        patterns['有x分隔尺寸'] += 1
    elif '定制' in s or '定做' in s or '定造' in s:
        patterns['定制'] += 1
    elif '珍珠棉' in s or 'pe' in s.lower():
        patterns['珍珠棉'] += 1
    else:
        patterns['其他'] += 1

print(f"\n=== 格式分类统计 ===")
for name, cnt in patterns.most_common():
    print(f"{name}: {cnt}")

# 输出全部分类样本
print("\n正在输出详细分类文件...")
import openpyxl as oxl

wb_out = oxl.Workbook()
ws_out = wb_out.active
ws_out.title = '格式分类'

# 表头
ws_out.append(['分类', '序号', '规格名称', '店铺简称', '平台商品id', '平台规格id'])

for cat_name, _ in patterns.most_common():
    idx = 0
    for s, shop, pid, spec_id in spec_samples:
        in_cat = False
        if cat_name == '纯数字/符号':
            in_cat = re.match(r'^[\d.*xX×\-]+$', s) or ('【' not in s and not re.search(r'[长宽高厚内径外径]', s))
        elif cat_name == '【】格式+有长宽高':
            in_cat = '【' in s and ('长' in s or '宽' in s or '高' in s or '厚度' in s or '长度' in s)
        elif cat_name == '【】格式+无长宽高':
            in_cat = '【' in s and not ('长' in s or '宽' in s or '高' in s)
        elif cat_name == '有长宽高+无【】':
            in_cat = ('长' in s or '宽' in s or '高' in s) and '【' not in s
        elif cat_name == '有x分隔尺寸':
            in_cat = bool(re.search(r'\d+\.?\d*\s*[xX*×]\s*\d+', s))
            # 排除已被其他分类收录的
            if '【' in s or '长' in s or '宽' in s or '高' in s:
                in_cat = False
        elif cat_name == '定制':
            in_cat = '定制' in s or '定做' in s or '定造' in s
        elif cat_name == '珍珠棉':
            in_cat = '珍珠棉' in s or 'pe' in s.lower()
        
        if in_cat and idx < 100:  # 每类最多100个样本
            idx += 1
            ws_out.append([cat_name, idx, s, str(shop)[:20], str(pid)[:20], str(spec_id)[:20]])

out_file = r"d:\Desktop\规格格式分类.xlsx"
wb_out.save(out_file)
print(f"✅ 文件已保存: {out_file}")

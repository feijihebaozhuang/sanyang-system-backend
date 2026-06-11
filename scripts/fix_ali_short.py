# -*- coding: utf-8 -*-
"""定制类换绑文件：阿里店铺用简称"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl as oxl

out = r'd:\Desktop\换绑输出'

# 简称映射（反向：全称→简称）
FULL_TO_SHORT = {
    '深圳市友尚包装有限公司': '阿里友尚',
    '深圳市亚润包装材料有限公司': '阿里亚润',
    '深圳市三羊包装材料有限公司': '阿里三羊',
    '深圳市正方形纸制品有限公司': '阿里正方形',
    '深圳市大鱼包装材料有限公司': '阿里大鱼',
    '东莞市新鑫星包装材料有限公司': '阿里新鑫星',
}

def fix_file(fname):
    fp = os.path.join(out, fname)
    if not os.path.exists(fp):
        print(f"{fname}: 不存在")
        return
    
    wb = oxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    
    # 重建
    wb2 = oxl.Workbook()
    ws2 = wb2.active; ws2.title = 'Sheet1'
    ws2.append([None, '商品对应表', None, None])
    ws2.append(['店铺名称', '平台商品id', '平台规格id', '商品编码'])
    
    changes = 0
    for r in rows:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        new_name = FULL_TO_SHORT.get(name, name)
        if new_name != name:
            changes += 1
        ws2.append([new_name, str(r[1] or '').strip(), str(r[2] or '').strip(), str(r[3] or '').strip()])
    
    wb2.save(fp)
    wb2.close()
    wb.close()
    sz = os.path.getsize(fp)
    print(f"✅ {fname}: 改了{changes}条, {sz/1024:.1f}KB")

fix_file('定制类_换绑文件.xlsx')
fix_file('定制类_换绑文件_第2批.xlsx')

print("\n可以先上传 定制类_换绑文件.xlsx 试试了！")

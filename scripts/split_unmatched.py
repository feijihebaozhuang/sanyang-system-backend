# -*- coding: utf-8 -*-
"""从未匹配平台商品拆分出 定制类 和 待修改 两个文件"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl as oxl

src = r"d:\Desktop\换绑输出\未匹配平台商品.xlsx"
out_dir = r"d:\Desktop\换绑输出"

print("读取未匹配文件...", flush=True)
wb = oxl.load_workbook(src, data_only=True)
ws = wb['未匹配']

# 表头: 店铺简称, 平台商品id, 平台规格id, 规格名称, 原因, 期望编码
rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"总未匹配: {len(rows)} 条", flush=True)

custom_rows = []    # 定制类：你自己搞定
fix_rows = []       # 待修改

custom_reasons = {'珍珠棉', '定制关键词', '尺寸不足', '蓝绿颜色'}
# 注意："只有长宽"也放到待修改，因为你可能想改规则

for r in rows:
    if not r:
        continue
    reason = str(r[4] or '').strip()
    if reason in custom_reasons:
        custom_rows.append(r)
    else:
        fix_rows.append(r)

print(f"定制类（你自己搞定）: {len(custom_rows)} 条", flush=True)
print(f"  其中珍珠棉: {sum(1 for r in custom_rows if r[4]=='珍珠棉')}", flush=True)
print(f"  其中定制关键词: {sum(1 for r in custom_rows if r[4]=='定制关键词')}", flush=True)
print(f"  其中尺寸不足: {sum(1 for r in custom_rows if r[4]=='尺寸不足')}", flush=True)
print(f"  其中蓝绿颜色: {sum(1 for r in custom_rows if r[4]=='蓝绿颜色')}", flush=True)
print(f"待修改: {len(fix_rows)} 条", flush=True)
print(f"  其中无匹配: {sum(1 for r in fix_rows if r[4]=='无匹配')}", flush=True)
print(f"  其中只有长宽: {sum(1 for r in fix_rows if r[4]=='只有长宽')}", flush=True)

header = ['店铺简称', '平台商品id', '平台规格id', '规格名称', '原因', '期望编码']

# 定制类
wb1 = oxl.Workbook()
ws1 = wb1.active
ws1.title = '定制类'
ws1.append(header)
for r in custom_rows:
    ws1.append(list(r))

# 加统计sheet
ws1s = wb1.create_sheet('统计')
ws1s.append(['原因', '数量'])
from collections import Counter
for reason, cnt in Counter(r[4] for r in custom_rows).most_common():
    ws1s.append([reason, cnt])

f1 = os.path.join(out_dir, '剩余平台商品_定制类.xlsx')
wb1.save(f1)
fs1 = os.path.getsize(f1)
print(f"\n✅ 定制类: {f1} ({fs1/1024:.0f}KB, {len(custom_rows)}条)", flush=True)
wb1.close()

# 待修改
wb2 = oxl.Workbook()
ws2 = wb2.active
ws2.title = '待修改'
ws2.append(header)
for r in fix_rows:
    ws2.append(list(r))

ws2s = wb2.create_sheet('统计')
ws2s.append(['原因', '数量'])
for reason, cnt in Counter(r[4] for r in fix_rows).most_common():
    ws2s.append([reason, cnt])

f2 = os.path.join(out_dir, '剩余平台商品_待修改.xlsx')
wb2.save(f2)
fs2 = os.path.getsize(f2)
print(f"✅ 待修改: {f2} ({fs2/1024:.0f}KB, {len(fix_rows)}条)", flush=True)
wb2.close()

wb.close()
print("\n完成！", flush=True)

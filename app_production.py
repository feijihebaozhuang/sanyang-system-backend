#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞机盒智能生产管理系统 - 后端API
小马哥专属
"""

from flask import Flask, jsonify, send_from_directory, request, make_response, Response, session, redirect
from flask_cors import CORS
from flask_session import Session as FlaskSession
import json, datetime, csv, io, os, hashlib, copy, time, re, hmac, urllib.parse, urllib.request
from datetime import timedelta
from pypinyin import lazy_pinyin
import pymysql

from werkzeug.middleware.proxy_fix import ProxyFix

from settings import DB_CONFIG, FLASK_SECRET_KEY
import production_helpers as ph
from quote_config_merge import merge_quote_config


def _auth_token_for(username: str) -> str:
    return hashlib.sha256(f"{FLASK_SECRET_KEY}:{username}".encode()).hexdigest()[:32]


def resolve_login_user() -> str | None:
    un = session.get("username")
    if un and un in USERS:
        return un
    hdr_user = (request.headers.get("X-Sanyang-User") or "").strip()
    hdr_tok = (request.headers.get("X-Sanyang-Token") or "").strip()
    if hdr_user in USERS and hdr_tok and hdr_tok == _auth_token_for(hdr_user):
        _bind_session_for_user(hdr_user)
        return hdr_user
    return None


def _active_user(username: str | None = None) -> dict | None:
    """当前登录用户（规范化 role，并回写 Session，避免旧 Session 仍为「员工」）。"""
    import permission_resolve as _pr

    un = (username or session.get("username") or "").strip()
    if not un:
        hdr_user = (request.headers.get("X-Sanyang-User") or "").strip()
        hdr_tok = (request.headers.get("X-Sanyang-Token") or "").strip()
        if hdr_user in USERS and hdr_tok and hdr_tok == _auth_token_for(hdr_user):
            un = hdr_user
    if not un:
        return None
    user = USERS.get(un)
    if not user:
        return None
    user = _pr.normalize_user_record(un, user)
    USERS[un] = user
    if session.get("username") == un:
        if session.get("role") != user.get("role"):
            session["role"] = user["role"]
        if session.get("employee_name") != user.get("employee_name"):
            session["employee_name"] = user.get("employee_name", "")
        session.modified = True
    return user


def _bind_session_for_user(username: str) -> None:
    """Header 令牌鉴权通过后写入 Session，避免刷新后仅 Cookie 失效导致接口 401。"""
    user = _active_user(username)
    if not user:
        return
    session.permanent = True
    session["username"] = username
    session["user_name"] = user.get("name") or username
    session["role"] = user.get("role") or "员工"
    session["employee_name"] = user.get("employee_name") or ""
    session.modified = True


def get_db():
    """获取数据库连接，每次调用新建（线程安全）"""
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)

app = Flask(__name__, static_folder='.')
app.secret_key = FLASK_SECRET_KEY
CORS(app, supports_credentials=True)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=31)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_COOKIE_SECURE'] = os.getenv('SESSION_COOKIE_SECURE', '').lower() in (
    '1', 'true', 'yes'
)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = '/tmp/flask_session_prod'
app.config['SESSION_FILE_THRESHOLD'] = 100
FlaskSession(app)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

import embed_frame_policy as _embed_frame_policy

_embed_frame_policy.register_embed_parents(app)

# 小程序 API 同时挂到 3002（feijihe.top 证书可用；guanli 子域 SSL 未配时走此入口）
try:
    from mp_api import mp_bp as _mp_bp

    app.register_blueprint(_mp_bp)
except ImportError:  # pragma: no cover
    pass

try:
    import co_admin_proxy as _co_proxy

    _co_proxy.register_co_admin_proxy(app)
except ImportError:  # pragma: no cover
    pass

try:
    import guanli_embed_proxy as _embed_proxy

    _embed_proxy.register_guanli_embed_proxy(app)
except ImportError:  # pragma: no cover
    pass


@app.before_request
def _sync_login_session_from_token():
    if not request.path.startswith("/api/"):
        return
    un = resolve_login_user()
    if un:
        _bind_session_for_user(un)


import webhook_routes as _webhook_routes

_webhook_routes.register_webhook_routes(app)

import order_cache_store as _order_cache_store

_order_cache_store.schedule_startup_migration()

# ==================== 用户系统 ====================
USERS = {
    "admin": {
        "password": hashlib.sha256("admin888".encode()).hexdigest(),
        "name": "戴雅利",
        "role": "超级管理员",
        "employee_name": "戴雅利"
    },
    "manager": {
        "password": hashlib.sha256("manager666".encode()).hexdigest(),
        "name": "邓涛",
        "role": "管理",
        "employee_name": "邓涛"
    },
    "worker": {
        "password": hashlib.sha256("worker123".encode()).hexdigest(),
        "name": "李周海",
        "role": "员工",
        "employee_name": "李周海"
    },
    "sushiting": {
        "password": hashlib.sha256("sushiting123".encode()).hexdigest(),
        "name": "苏世婷",
        "role": "客服",
        "employee_name": "苏世婷"
    },
    "liaosimei": {
        "password": hashlib.sha256("liaosimei123".encode()).hexdigest(),
        "name": "廖思美",
        "role": "客服",
        "employee_name": "廖思美"
    }
}

def require_login():
    """检查是否已登录（Session 或 X-Sanyang 令牌）"""
    if not resolve_login_user():
        return jsonify({"error": "未登录", "code": 401}), 401
    return None

def require_admin():
    """3001/3002 不再做权限分级，登录即可（细粒度权限仅在 3003）。"""
    return require_login()


def _can_edit_employee_info(user: dict | None) -> bool:
    """3001/3002：登录即可查看/编辑员工基础信息（权限矩阵在 3003）。"""
    return bool(user)

# ==================== 数据持久化 ====================
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')

def load_data():
    """从JSON文件加载持久化数据"""
    # 生成默认用户密码哈希
    default_users = {
        "admin": {
            "password": hashlib.sha256("admin888".encode()).hexdigest(),
            "name": "admin",
            "role": "超级管理员",
            "employee_name": "",
            "is_system": True
        },
        "daiyali": {
            "password": hashlib.sha256("admin888".encode()).hexdigest(),
            "name": "戴雅利",
            "role": "超级管理员",
            "employee_name": "戴雅利"
        },
        "manager": {
            "password": hashlib.sha256("manager666".encode()).hexdigest(),
            "name": "邓涛",
            "role": "管理",
            "employee_name": "邓涛"
        },
        "worker": {
            "password": hashlib.sha256("worker123".encode()).hexdigest(),
            "name": "李周海",
            "role": "员工",
            "employee_name": "李周海"
        },
        "sushiting": {
            "password": hashlib.sha256("sushiting123".encode()).hexdigest(),
            "name": "苏世婷",
            "role": "客服",
            "employee_name": "苏世婷"
        },
        "liaosimei": {
            "password": hashlib.sha256("liaosimei123".encode()).hexdigest(),
            "name": "廖思美",
            "role": "客服",
            "employee_name": "廖思美"
        }
    }
    default = {
        "users": default_users,  # 持久化的用户密码
        "employee_status": {},  # 考勤状态
        "employees_master": [],  # 员工主数据（admin 修改后持久化，部署不覆盖）
        "permission_data": {    # 权限数据
            "processes": [  # 树形结构：美丽湾工厂部/纸箱部 两部门，每部门下工序步骤
                {
                    "dept": "美丽湾工厂部",
                    "steps": [
                        {"id": 1, "name": "客服接单"},
                        {"id": 2, "name": "黄厂打印"},
                        {"id": 3, "name": "审单分单"},
                        {"id": 4, "name": "算料"},
                        {"id": 5, "name": "分纸"},
                        {"id": 6, "name": "啤机(自动平压平)"},
                        {"id": 7, "name": "啤机(机械手)"},
                        {"id": 8, "name": "手啤"},
                        {"id": 9, "name": "印刷"},
                        {"id": 10, "name": "清废"},
                        {"id": 11, "name": "打包发货"}
                    ]
                },
                {
                    "dept": "纸箱部",
                    "steps": [
                        {"id": 1, "name": "客服接单"},
                        {"id": 2, "name": "黄厂打印"},
                        {"id": 3, "name": "审单分单"},
                        {"id": 4, "name": "算料"},
                        {"id": 5, "name": "分纸"},
                        {"id": 6, "name": "印刷"},
                        {"id": 7, "name": "开槽/打角"},
                        {"id": 8, "name": "粘胶/打钉"},
                        {"id": 9, "name": "打包发货"}
                    ]
                }
            ],
            "positions": ["超级管理员", "主管", "客服", "员工", "财务", "业务员"],
            "employee_enabled": {},  # {name: true/false} 员工启用/禁用状态
            "employees": [
                {"name": "戴雅利", "position": "超级管理员"},
                {"name": "邓涛", "position": "主管"},
                {"name": "黄兴", "position": "主管"},
                {"name": "覃海霞", "position": "主管"},
                {"name": "蒋义红", "position": "主管"},
                {"name": "沈齐豪", "position": "主管"},
                {"name": "苏世婷", "position": "客服"},
                {"name": "廖思美", "position": "客服"},
                {"name": "张慧平", "position": "客服"},
                {"name": "陈贝贝", "position": "客服"},
                {"name": "罗怡", "position": "客服"},
                {"name": "周井梅", "position": "客服"},
                {"name": "戴志美", "position": "客服"},
                {"name": "石梅清", "position": "客服"},
                {"name": "张文杰", "position": "客服"},
                {"name": "何水单", "position": "业务员"},
                {"name": "李四军", "position": "员工"},
                {"name": "陈贤聪", "position": "业务员"},
                {"name": "姚斌", "position": "员工"},
                {"name": "隆浪", "position": "财务"},
            ],
            "permissions": {
                "戴雅利": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "数据看板": True, "刀模": True, "库存": True, "快麦ERP": True, "员工": True, "权限管理": True, "报价": True},
                "邓涛": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "数据看板": True, "刀模": True, "库存": True, "快麦ERP": True, "员工": True, "权限管理": False, "报价": False},
                "李周海": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "员工": True, "刀模": False, "库存": False, "快麦ERP": False, "数据看板": False, "权限管理": False, "报价": False},
            }
        }
    }
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 确保所有字段存在
                for key in default:
                    if key not in data:
                        data[key] = default[key]
                # 合并权限数据中的嵌套字段
                for key in default["permission_data"]:
                    if key not in data["permission_data"]:
                        data["permission_data"][key] = default["permission_data"][key]
                # 自动转换旧格式流程（字符串列表→树形）
                processes = data["permission_data"].get("processes", [])
                if processes and isinstance(processes[0], str):
                    data["permission_data"]["processes"] = [
                        {"dept": "美丽湾工厂部", "steps": [{"id": i+1, "name": s} for i, s in enumerate(processes)]}
                    ]
                return data
        except:
            pass
    return default

def save_data(data):
    """保存数据到JSON文件"""
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

# 初始化数据
persistent_data = load_data()

import hermes_boot_fix as _hermes_fix

_hermes_fix.auto_fix_hermes_local_backend()

import agent_self_repair as _agent_repair

_agent_repair.run_boot_repair()

# 从data.json加载所有用户（覆盖并扩展USERS字典）
import permission_resolve as _perm_resolve_boot

_loaded_users = persistent_data.get("users", {})
_perm_resolve_boot.install_users_from_persistent(USERS, _loaded_users)
_employee_today_status = persistent_data.get("employee_status", {})
_employee_leave_counts = persistent_data.get("employee_leave_counts", {})
_order_extra = persistent_data.get("order_extra", {})
import order_extra_store as _order_extra_store

_order_extra_store.merge_urgent_into(_order_extra, get_db)
_permission_data = persistent_data.get("permission_data", {})
_resigned_employees = persistent_data.get("resigned_employees", [])  # 离职员工列表[{name, position, group, dept, phone, resigned_time}]

import config_json as _cfg_json

_cfg_json.apply_permission_overlay(_permission_data)

# 将员工状态变更为按日期存储格式（兼容旧格式）
if isinstance(_employee_today_status, dict):
    # 检查是 "date -> name -> status" 还是老格式
    is_old_format = False
    for k, v in _employee_today_status.items():
        if isinstance(v, str):
            is_old_format = True
            break
    if is_old_format:
        today = datetime.date.today().isoformat()
        _employee_today_status = {today: _employee_today_status}

_perm_save_detail: dict = {}


def _sync_order_extra_from_db() -> None:
    """从 MySQL order_extras 刷新加急（客服端改完后生产端立即可见）。"""
    _order_extra_store.merge_urgent_into(_order_extra, get_db)


def persist():
    """持久化当前数据到文件（data.json，不覆盖 admin 未改动的其它键）。"""
    global _employee_today_status, _employee_leave_counts, _permission_data, USERS, _resigned_employees, _employees_master_list, _perm_save_detail
    # 把USERS的密码也存到文件（用户改密码后重启不丢）
    users_data = {}
    for uid, u in USERS.items():
        entry = {
            "password": u["password"],
            "name": u["name"],
            "role": u["role"],
            "employee_name": u["employee_name"]
        }
        if u.get("is_system"):
            entry["is_system"] = True
        users_data[uid] = entry
    data = {
        "users": users_data,
        "employee_status": _employee_today_status,
        "employee_leave_counts": _employee_leave_counts,
        "order_extra": _order_extra,
        "permission_data": _permission_data,
        "resigned_employees": _resigned_employees,
        "employees_master": _employees_master_list,
    }
    if not save_data(data):
        _perm_save_detail = {"ok": False, "local_ok": False, "vault_error": "save_data 失败"}
        return False
    _perm_save_detail = _cfg_json.write_permission_overlay_detail(_permission_data)
    return bool(_perm_save_detail.get("local_ok"))

# ==================== 登录API ====================
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({"success": False, "message": "请输入账号和密码"})
    
    user = USERS.get(username)
    if not user:
        return jsonify({"success": False, "message": "账号或密码错误"})
    import permission_resolve as _pr_login
    user = _pr_login.normalize_user_record(username, user)
    USERS[username] = user

    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    if user['password'] != pwd_hash:
        return jsonify({"success": False, "message": "账号或密码错误"})
    
    # 检查员工是否被禁用
    employee_name = user.get('employee_name', '')
    enabled_map = _permission_data.get("employee_enabled", {})
    # 默认启用，只有明确设为 False 才禁用
    if enabled_map.get(employee_name, True) is False:
        return jsonify({"success": False, "message": "该账号已被禁用，请联系管理员"})
    
    session.permanent = True
    session['username'] = username
    session['user_name'] = user.get('name') or username
    session['role'] = user['role']
    session['employee_name'] = user.get('employee_name', '')
    session.modified = True
    if user.get("is_system"):
        persist()

    # 查找用户所属部门
    emp_dept = ''
    for emp in _employees_master_list:
        if emp['name'] == user.get('employee_name'):
            emp_dept = emp.get('dept', '')
            break

    pub = _pr_login.user_public_payload(username, user)
    pub["dept"] = emp_dept
    return jsonify({
        "success": True,
        "message": "登录成功",
        "auth_token": _auth_token_for(username),
        "user": pub
    })

@app.route('/api/logout')
def logout():
    session.clear()
    return jsonify({"success": True, "message": "已退出登录"})

@app.route('/api/change_password', methods=['POST'])
def change_password():
    """仅允许修改当前登录账号自己的密码。"""
    username = resolve_login_user()
    if not username:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    target = (data.get('username') or data.get('target_user') or '').strip()
    if target and target != username:
        return jsonify({"success": False, "message": "只能修改自己的密码"}), 403

    old_pwd = data.get('old_password', '').strip()
    new_pwd = data.get('new_password', '').strip()

    if not old_pwd or not new_pwd:
        return jsonify({"success": False, "message": "请填写旧密码和新密码"})

    if len(new_pwd) < 4:
        return jsonify({"success": False, "message": "新密码至少4位"})

    user = USERS.get(username)
    if not user:
        return jsonify({"success": False, "message": "用户不存在"})

    old_hash = hashlib.sha256(old_pwd.encode()).hexdigest()
    if user['password'] != old_hash:
        return jsonify({"success": False, "message": "旧密码错误"})

    user['password'] = hashlib.sha256(new_pwd.encode()).hexdigest()
    persist()
    return jsonify({"success": True, "message": "密码修改成功"})


@app.route('/api/admin/reset_password', methods=['POST'])
def admin_reset_password():
    """超级管理员重置他人密码。"""
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "message": "未登录"}), 401
    actor = USERS.get(un)
    if not actor or actor.get('role') != '超级管理员':
        return jsonify({"success": False, "message": "无权限"}), 403

    data = request.get_json() or {}
    target = (data.get('username') or '').strip()
    new_pwd = (data.get('new_password') or '').strip()
    if not target or not new_pwd:
        return jsonify({"success": False, "message": "请指定账号和新密码"})
    if len(new_pwd) < 4:
        return jsonify({"success": False, "message": "新密码至少4位"})
    if target == un:
        return jsonify({"success": False, "message": "请使用「修改密码」修改自己的密码"})

    user = USERS.get(target)
    if not user:
        return jsonify({"success": False, "message": "目标账号不存在"})
    user['password'] = hashlib.sha256(new_pwd.encode()).hexdigest()
    persist()
    return jsonify({"success": True, "message": f"已重置账号 {target} 的密码"})

@app.route('/api/me')
def get_current_user():
    un = resolve_login_user()
    if un:
        user = _active_user(un)
        if user:
            import permission_resolve as _pr_me
            pub = _pr_me.user_public_payload(un, user)
            return jsonify({
                "logged_in": True,
                "auth_token": _auth_token_for(un),
                "user": pub
            })
    return jsonify({"logged_in": False})

# ==================== 首页 - 生产概览 ====================
@app.route('/api/dashboard')
def dashboard():
    _sync_order_extra_from_db()
    date = request.args.get('date', datetime.date.today().strftime('%Y-%m-%d'))
    shop_config = load_shop_config()

    # 订单缓存：MySQL + 访问即同步
    real_orders = []
    try:
        import order_visit_sync as _ovs

        _ovs.schedule_incremental_sync(memo_getter=get_order_memo, force=False)
        real_orders = ph.load_cache_orders()
    except Exception as e:
        print(f'[Dashboard] 读缓存失败: {e}')
        real_orders = []

    # 构建店铺->客服映射（从shop_config读取）
    store_workers = {}
    for sc in shop_config:
        name = sc.get('shop_name', '')
        cs = sc.get('customer_service', '')
        if name and cs:
            store_workers[name] = cs

    try:
        import km_api as _km
        for o in real_orders:
            _km.finalize_cache_order(o)
    except ImportError:
        pass

    all_orders = []
    for o in real_orders:
        items = o.get('items', [])
        product_parts = []
        total_qty = 0
        for item in items:
            spec = item.get('spec', '') or ''
            qty = item.get('qty', 0)
            total_qty += qty
            if spec:
                product_parts.append(f"{spec}×{qty}")
        product_str = '; '.join(product_parts) if product_parts else '待确认'

        oid = ph.internal_order_id(o)
        saved = _order_extra.get(oid, {})
        shop_name = o.get('shop_name', '亚润')
        all_orders.append({
            "id": oid,
            "store": shop_name,
            "worker": store_workers.get(shop_name, ""),
            "product": product_str,
            "qty": total_qty,
            "items": items,
            "process": "待发货",
            "status": "待发货",
            "urgent": saved.get("urgent", False),
            "remark": saved.get("remark", "") or o.get('seller_memo', ''),
            "pay_time": o.get('pay_time', '') or o.get('created', ''),
        })

    total_all = len(all_orders)
    total_waiting = len([o for o in all_orders if o['status'] == '待发货'])
    urgent_orders = [o for o in all_orders if o.get('urgent')]
    urgent_count = len(urgent_orders)
    completed_count = len([
        o for o in all_orders
        if o.get('status') == '已完成'
    ])

    # 筛选今日必发：pay_time/created 日期 = 今天的订单
    today_str = datetime.date.today().isoformat()
    today_orders = []
    for o in all_orders:
        t = o.get("pay_time", "") or o.get("created", "")
        if t[:10] == today_str:
            today_orders.append(o)

    return jsonify({
        "date": date,
        "summary": {
            "total_orders": total_all,
            "in_production": total_waiting,
            "waiting": total_waiting,
            "urgent_orders": urgent_count,
            "completed": completed_count,
        },
        "today_orders": today_orders,
        "urgent_orders": urgent_orders,
    })

# ==================== 订单生产进度 ====================
@app.route('/api/production_orders')
def production_orders():
    """待发货订单 + 工序树进度（扫码报工/进度查询）"""
    _sync_order_extra_from_db()
    process_tree = _permission_data.get("processes", [])
    orders_data = ph.build_production_orders(
        _orders_cache_path(),
        DB_CONFIG,
        process_tree,
        _order_extra,
    )
    return jsonify({"orders": orders_data})


@app.route('/api/scan_order/<order_id>')
def scan_order_detail(order_id):
    """单条订单工序（车间扫码报工小程序，避免拉全量 production_orders）。"""
    query = (order_id or "").strip()
    if not query:
        return jsonify({"success": False, "error": "缺少单号"})
    process_tree = _permission_data.get("processes", [])
    order = ph.build_production_order_one(
        _orders_cache_path(),
        DB_CONFIG,
        process_tree,
        _order_extra,
        query,
    )
    if not order:
        return jsonify({"success": False, "error": f"未找到订单 {query}"})
    return jsonify({"success": True, "order": order})


# ==================== 扫码报工 ====================
@app.route('/api/scan_report', methods=['POST'])
def scan_report():
    data = request.get_json() or {}
    order_id = (data.get('order_id') or '').strip()
    step = (data.get('step') or '').strip()
    worker = (data.get('worker') or '').strip()
    if not order_id or not step:
        return jsonify({"success": False, "message": "缺少单号或工序"})
    if not worker:
        user = USERS.get(session.get('username', ''), {})
        worker = user.get('employee_name') or user.get('name') or ''
    result = ph.apply_scan_report(
        DB_CONFIG,
        _permission_data.get("processes", []),
        order_id,
        step,
        worker,
        _orders_cache_path(),
    )
    if not result.get("success"):
        return jsonify({"success": False, "message": result.get("msg", "报工失败")})
    return jsonify({
        "success": True,
        "message": f"订单 {result['order_id']} 工序「{step}」已报工",
        "all_done": result.get("all_done"),
        "record": {
            "time": result.get("time"),
            "order_id": result.get("order_id"),
            "step": step,
            "worker": worker,
            "status": "已完成",
        },
    })


@app.route('/api/scan_logs')
def scan_logs():
    return jsonify({"logs": ph.fetch_scan_logs(DB_CONFIG, 80)})

# ==================== 日报表 ====================
@app.route('/api/daily_report')
def daily_report():
    report_date = request.args.get('date', datetime.date.today().strftime('%Y-%m-%d'))
    today = datetime.date.today().isoformat()
    emp_statuses = _employee_today_status.get(today, {})
    cache_path = _orders_cache_path()
    payload = ph.build_daily_report(
        cache_path,
        DB_CONFIG,
        report_date,
        _order_extra,
        emp_statuses,
    )
    return jsonify(payload)


@app.route('/api/databoard')
def databoard():
    range_type = request.args.get('range', 'week')
    cache_path = _orders_cache_path()
    return jsonify(
        ph.build_databoard(cache_path, range_type, _order_extra)
    )

def _orders_cache_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "orders_cache.json")


def _find_cached_order(query):
    import order_cache_store as ocs

    return ocs.find_order(query)


def _order_detail_from_cache(o):
    items = o.get("items") or []
    first = items[0] if items else {}
    product = first.get("name", "") or first.get("spec", "") or "?"
    if len(items) > 1:
        product += f" 等{len(items)}种"
    qty = sum(int(i.get("qty", 0) or 0) for i in items)
    amount = o.get("payment") or o.get("pay_amount") or o.get("total_fee") or 0
    return {
        "order_id": o.get("so_id", ""),
        "store": o.get("shop_name", ""),
        "customer": o.get("receiver_name", "") or o.get("buyer_nick", "") or "",
        "product": product,
        "qty": qty,
        "amount": amount,
        "order_date": (o.get("created") or o.get("pay_time") or "")[:10],
        "delivery_date": (o.get("plan_delivery_date") or o.get("consign_time") or "")[:10],
        "address": o.get("receiver_address", "") or "",
        "logistics": o.get("logistics_company", "") or o.get("express", "") or "",
    }


def _production_status_from_flow(flow):
    flow = flow or []
    done = [s for s in flow if s.get("done")]
    total = len(flow) or 1
    progress = int(len(done) * 100 / total)
    current = "-"
    for s in flow:
        if not s.get("done"):
            current = s.get("step") or s.get("name", "-")
            break
    else:
        if flow:
            current = flow[-1].get("step") or flow[-1].get("name", "已完成")
    status = "已完成" if progress >= 100 and flow else ("生产中" if done else "待处理")
    steps = [
        {"name": s.get("step") or s.get("name", ""), "done": bool(s.get("done")), "time": s.get("time", "-")}
        for s in flow
    ]
    return {"status": status, "current_process": current, "diemold": "", "progress": progress, "steps": steps}


# ==================== 搜索订单 ====================
@app.route('/api/search_order')
def search_order():
    """扫码/输入单号：从 orders_cache.json 查订单（快麦+1688）。"""
    _sync_order_extra_from_db()
    query = (request.args.get('q') or '').strip()
    if not query:
        return jsonify({"found": False, "message": "请输入单号"})
    o = _find_cached_order(query)
    if not o:
        return jsonify({"found": False, "message": "未找到该订单"})
    so_id = o.get("so_id", "")
    extra = _order_extra.get(so_id, {})
    prod_resp = production_orders().get_json()
    prod = next((x for x in prod_resp.get("orders", []) if x.get("inner_id") == so_id), None)
    flow = (prod or {}).get("flow") or []
    detail = _order_detail_from_cache(o)
    return jsonify({
        "found": True,
        "order_id": so_id,
        "urgent": bool(extra.get("urgent")),
        "production_status": _production_status_from_flow(flow),
        "order_detail": detail,
        "jst_detail": detail,
    })

# ==================== 生产进度 ====================
@app.route('/api/production_timeline')
def production_timeline():
    return jsonify({
        "stages": [
            {"id": 1, "name": "客服接单", "color": "#1677ff", "icon": "📞"},
            {"id": 2, "name": "黄厂打印", "color": "#52c41a", "icon": "🖨️"},
            {"id": 3, "name": "审单分单", "color": "#faad14", "icon": "📋"},
            {"id": 4, "name": "算料", "color": "#722ed1", "icon": "📐"},
            {"id": 5, "name": "生产制作", "color": "#1677ff", "icon": "🏭"},
            {"id": 6, "name": "清废", "color": "#13c2c2", "icon": "🧹"},
            {"id": 7, "name": "打包发货", "color": "#52c41a", "icon": "🚚"},
        ]
    })

# ==================== 员工信息（增强版，带手机号和厂区） ====================
# 全局员工列表（方便增删改API共享）
_employees_master_list = [
    # ===== 洋坑塘运营部（电商运营中心） =====
    {"name": "戴雅利", "position": "总负责人", "group": "管理层", "dept": "洋坑塘运营部", "phone": "13800138001"},
    {"name": "苏世婷", "position": "阿里客服主管", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138002"},
    {"name": "廖思美", "position": "淘天客服主管", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138003"},
    {"name": "何水单", "position": "业务员", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138004"},
    {"name": "李四军", "position": "美工", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138005"},
    {"name": "陈贤聪", "position": "抖音/微信/珍珠棉业务", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138006"},
    {"name": "张慧平", "position": "友尚旗舰店运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138007"},
    {"name": "陈贝贝", "position": "亚润跨境运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138008"},
    {"name": "罗怡", "position": "三羊现货运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138009"},
    {"name": "周井梅", "position": "正方形/大鱼运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138010"},
    {"name": "戴志美", "position": "友尚/新鑫星运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138011"},
    {"name": "石梅清", "position": "止合/扣底盒运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138012"},
    {"name": "张文杰", "position": "小批量/品牌店运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138013"},
    {"name": "姚斌", "position": "采购/跟单/机动", "group": "综合组", "dept": "洋坑塘运营部", "phone": "13800138014"},
    # ===== 美丽湾工厂部-管理层 =====
    {"name": "邓涛", "position": "生产总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138015"},
    {"name": "黄兴", "position": "行政总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138016"},
    {"name": "隆浪", "position": "财务总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138017"},
    {"name": "蒋义红", "position": "飞机盒主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138018"},
    {"name": "沈齐豪", "position": "纸箱主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138019"},
    {"name": "覃海霞", "position": "综合主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138020"},
    # ===== 美丽湾工厂部-手啤机组 =====
    {"name": "李方国", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138021"},
    {"name": "李周海", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138022"},
    {"name": "唐美章", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138023"},
    {"name": "蒋保平", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138024"},
    {"name": "陈章远", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138025"},
    {"name": "黄振辉", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138026"},
    {"name": "汪成桃", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138027"},
    # ===== 美丽湾工厂部-啤机分组 =====
    {"name": "易新明", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138028"},
    {"name": "雷善炎", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138029"},
    {"name": "陈海斌", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138030"},
    {"name": "陈奕升", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138031"},
    {"name": "江旭松", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138032"},
    # ===== 美丽湾工厂部-刀模组 =====
    {"name": "廖金玲", "position": "刀模-找刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138033"},
    {"name": "陈勇", "position": "刀模-组刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138034"},
    {"name": "张吉杰", "position": "刀模-拆刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138035"},
    {"name": "唐孝定", "position": "放刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138036"},
    # ===== 美丽湾工厂部-机械手组 =====
    {"name": "黄恒", "position": "机械手啤工", "group": "机械手组", "dept": "美丽湾工厂部", "phone": "13800138037"},
    {"name": "李荣晖", "position": "机械手啤工", "group": "机械手组", "dept": "美丽湾工厂部", "phone": "13800138038"},
    # ===== 美丽湾工厂部-平压平组 =====
    {"name": "蒋森响", "position": "平压平啤工", "group": "平压平组", "dept": "美丽湾工厂部", "phone": "13800138039"},
    {"name": "周业福", "position": "平压平啤工", "group": "平压平组", "dept": "美丽湾工厂部", "phone": "13800138040"},
    # ===== 美丽湾工厂部-车间打包组 =====
    {"name": "陈桂英", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138041"},
    {"name": "毛良芬", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138042"},
    {"name": "陈辉文", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138043"},
    {"name": "文小梅", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138044"},
    {"name": "帅行朝", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138045"},
    {"name": "黄张华", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138046"},
    # ===== 美丽湾工厂部-仓库组 =====
    {"name": "宋小国", "position": "仓库-找货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138047"},
    {"name": "唐忠群", "position": "仓库-找货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138048"},
    {"name": "蒋仁叶", "position": "仓库-打包", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138049"},
    {"name": "罗照权", "position": "仓库-打包", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138050"},
    {"name": "黄爱小", "position": "仓库-放货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138051"},
    {"name": "龙雪兰", "position": "仓库-打样兼配货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138052"},
    # ===== 美丽湾工厂部-印刷组 =====
    {"name": "李双", "position": "印刷", "group": "印刷组", "dept": "美丽湾工厂部", "phone": "13800138053"},
    # ===== 美丽湾工厂部-扣底盒 =====
    {"name": "蒋军林", "position": "扣底盒", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138054"},
    {"name": "宁哈姝", "position": "扣底盒", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138055"},
    {"name": "黄华", "position": "扣底盒-打包", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138056"},
    # ===== 美丽湾工厂部-纸箱部(分纸) =====
    {"name": "陈伟", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138057"},
    {"name": "陈陶", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138058"},
    {"name": "周石国", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138059"},
    # ===== 美丽湾工厂部-纸箱部(开槽) =====
    {"name": "杨家忠", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138060"},
    {"name": "沈奇严", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138061"},
    {"name": "姚建桥", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138062"},
    # ===== 美丽湾工厂部-纸箱部(粘胶) =====
    {"name": "戴西华", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138063"},
    {"name": "魏小王", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138064"},
    {"name": "黄芳", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138065"},
    {"name": "于天发", "position": "纸箱-粘胶/打角", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138066"},
]

_emp_saved = persistent_data.get("employees_master")
if isinstance(_emp_saved, list) and len(_emp_saved) > 0:
    import copy as _copy
    _employees_master_list = _copy.deepcopy(_emp_saved)
else:
    persist()

@app.route('/api/employees')
def employees():
    # 过滤掉系统账号的员工（admin是系统账号，对应的员工名不出现在列表）
    # 规则：员工的名称如果被某个 is_system 用户引用，且没有被非system用户引用，则过滤
    system_emp_names = set()
    real_emp_names = set()
    for u in USERS.values():
        emp_name = u.get("employee_name", "")
        if u.get("is_system"):
            system_emp_names.add(emp_name)
        else:
            real_emp_names.add(emp_name)
    # 只过滤掉纯system用户的员工名
    exclude_names = system_emp_names - real_emp_names
    filtered = [e for e in _employees_master_list if e["name"] not in exclude_names]
    import employee_dept_filter as _edf
    filtered = _edf.filter_prod_site(filtered)
    emp_users = {}
    for un, u in USERS.items():
        if u.get("is_system"):
            continue
        en = u.get("employee_name") or u.get("name", "")
        if en:
            emp_users[en] = un
    out = []
    for e in filtered:
        row = dict(e)
        row["username"] = emp_users.get(e["name"], "")
        out.append(row)
    return jsonify({"employees": out})


# ==================== 员工编辑API ====================
@app.route('/api/employee/update', methods=['POST'])
def update_employee():
    """编辑员工信息（超级管理员或管理层）"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not _can_edit_employee_info(user):
        return jsonify({"success": False, "message": "仅管理员或管理层可编辑员工"}), 403
    
    data = request.get_json()
    old_name = data.get('old_name', '')
    new_name = data.get('name', '')
    position = data.get('position', '')
    group = data.get('group', '')
    phone = data.get('phone', '')
    dept = data.get('dept', '')
    
    if not old_name or not new_name:
        return jsonify({"success": False, "message": "请提供员工姓名"})
    
    for emp in _employees_master_list:
        if emp['name'] == old_name:
            emp['name'] = new_name
            if position:
                emp['position'] = position
            if group:
                emp['group'] = group
            if phone:
                emp['phone'] = phone
            if dept:
                emp['dept'] = dept
            for u in USERS.values():
                if u.get('employee_name') == old_name:
                    u['employee_name'] = new_name
                    u['name'] = new_name
            perms = _permission_data.setdefault('permissions', {})
            if old_name in perms and old_name != new_name:
                perms[new_name] = perms.pop(old_name)
            enabled = _permission_data.setdefault('employee_enabled', {})
            if old_name in enabled and old_name != new_name:
                enabled[new_name] = enabled.pop(old_name)
            persist()
            return jsonify({"success": True, "message": f"员工 {old_name} 信息已更新"})

    return jsonify({"success": False, "message": f"未找到员工 {old_name}"})

@app.route('/api/employee/add', methods=['POST'])
def add_employee():
    """添加新员工（admin权限）——自动生成登录账号和部门默认权限"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] != '超级管理员':
        return jsonify({"success": False, "message": "仅管理员可添加员工"}), 403
    
    data = request.get_json()
    name = data.get('name', '').strip()
    position = data.get('position', '').strip()
    group = data.get('group', '').strip()
    phone = data.get('phone', '').strip()
    dept = data.get('dept', '美丽湾工厂部').strip()
    
    if not name:
        return jsonify({"success": False, "message": "请填写员工姓名"})
    
    # 检查是否已存在
    for emp in _employees_master_list:
        if emp['name'] == name:
            return jsonify({"success": False, "message": f"员工 {name} 已存在"})
    
    _employees_master_list.append({
        "name": name,
        "position": position or "员工",
        "group": group or "其他",
        "dept": dept,
        "phone": phone or ""
    })
    
    # === 自动生成登录账号 ===
    # 姓名拼音转小写+去空格
    pinyin_parts = lazy_pinyin(name)
    raw_username = ''.join(pinyin_parts).lower().replace(' ', '')
    # 去掉非字母数字字符（某些特殊汉字拼音可能含标点）
    raw_username = re.sub(r'[^a-z0-9]', '', raw_username)
    
    # 如果用户名已存在则加数字后缀
    username = raw_username
    suffix = 1
    while username in USERS:
        suffix += 1
        username = f"{raw_username}{suffix}"
    
    # 默认密码 = 姓名拼音 + "123"
    default_pwd = f"{username}123"
    pwd_hash = hashlib.sha256(default_pwd.encode()).hexdigest()
    
    # 根据部门确定默认角色
    if dept == '洋坑塘运营部':
        default_role = '客服'
    else:
        default_role = '员工'
    
    # 注册到USERS
    USERS[username] = {
        "password": pwd_hash,
        "name": name,
        "role": default_role,
        "employee_name": name
    }
    
    # === 根据部门给默认权限 ===
    perms = _permission_data.setdefault("permissions", {})
    if name not in perms:
        perms[name] = {}
    
    if dept == '洋坑塘运营部':
        # 洋坑塘默认：客服权限——查单/库存，不涉及生产
        default_funcs = {
            '快麦ERP': True, '订单': True, '订单备注': True, '库存': True,
            '刀模': True, '报价': True, '智能报价': True,
            '刀模库': True, '原材料': True, '员工': True,
            '首页-加急单': True
        }
    else:
        # 美丽湾默认：员工权限——报工、考勤、查看
        default_funcs = {
            '员工': True, '首页-加急单': True
        }
    
    for func, val in default_funcs.items():
        perms[name][func] = val
    
    # 保存到data.json
    persist()
    
    return jsonify({
        "success": True,
        "message": f"员工 {name} 已添加，登录账号: {username}，默认密码: {default_pwd}，部门: {dept}",
        "username": username,
        "default_password": default_pwd,
        "role": default_role
    })

@app.route('/api/employee/delete', methods=['POST'])
def delete_employee():
    """删除员工（admin权限）"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] != '超级管理员':
        return jsonify({"success": False, "message": "仅管理员可删除员工"}), 403
    
    data = request.get_json()
    name = data.get('name', '')
    
    for i, emp in enumerate(_employees_master_list):
        if emp['name'] == name:
            _employees_master_list.pop(i)
            for un in list(USERS.keys()):
                u = USERS[un]
                if u.get('employee_name') == name and not u.get('is_system'):
                    del USERS[un]
            _permission_data.get('permissions', {}).pop(name, None)
            _permission_data.get('employee_enabled', {}).pop(name, None)
            persist()
            return jsonify({"success": True, "message": f"员工 {name} 已删除"})

    return jsonify({"success": False, "message": f"未找到员工 {name}"})


# ==================== 员工离职/恢复 ====================
@app.route('/api/employee/deactivate', methods=['POST'])
def deactivate_employee():
    """标记员工为离职 - 核心管理(邓涛/黄兴/隆浪)可操作"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] not in ['超级管理员', '管理']:
        return jsonify({"success": False, "message": "仅管理员可操作"}), 403
    
    global _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    
    for i, emp in enumerate(_employees_master_list):
        if emp['name'] == name:
            # 移到离职列表
            _resigned_employees.append({
                "name": emp['name'],
                "position": emp.get('position', ''),
                "group": emp.get('group', ''),
                "dept": emp.get('dept', ''),
                "phone": emp.get('phone', ''),
                "resigned_time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                "operator": user.get('name', '')
            })
            _employees_master_list.pop(i)
            persist()
            return jsonify({"success": True, "message": f"员工 {name} 已离职"})
    
    return jsonify({"success": False, "message": f"未找到员工 {name}"})

@app.route('/api/employee/restore', methods=['POST'])
def restore_employee():
    """恢复离职员工 - 仅超级管理员"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] != '超级管理员':
        return jsonify({"success": False, "message": "仅超级管理员可恢复"}), 403
    
    global _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    
    for i, emp in enumerate(_resigned_employees):
        if emp['name'] == name:
            # 恢复到员工列表
            _employees_master_list.append({
                "name": emp['name'],
                "position": emp.get('position', ''),
                "group": emp.get('group', ''),
                "dept": emp.get('dept', ''),
                "phone": emp.get('phone', '')
            })
            _resigned_employees.pop(i)
            persist()
            return jsonify({"success": True, "message": f"员工 {name} 已恢复"})
    
    return jsonify({"success": False, "message": f"未找到离职员工 {name}"})

@app.route('/api/employee/resigned')
def list_resigned():
    """获取离职员工列表"""
    return jsonify({"employees": _resigned_employees})

@app.route('/api/employee/delete_resigned', methods=['POST'])
def delete_resigned():
    """彻底删除离职员工记录 - 仅超级管理员"""
    if not resolve_login_user():
        return jsonify({"success": False, "message": "未登录"}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] != '超级管理员':
        return jsonify({"success": False, "message": "仅超级管理员可操作"}), 403
    
    global _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    
    for i, emp in enumerate(_resigned_employees):
        if emp['name'] == name:
            _resigned_employees.pop(i)
            persist()
            return jsonify({"success": True, "message": f"已彻底删除 {name} 的记录"})
    
    return jsonify({"success": False, "message": f"未找到离职员工 {name}"})


# ==================== 员工今日状态 ====================
# 按日期存储；新的一天自动空字典（旧日期数据仍保留在文件中）

@app.route('/api/employee/status', methods=['GET'])
def get_employee_status():
    global _employee_today_status, _employee_leave_counts
    today = datetime.date.today().isoformat()
    if today not in _employee_today_status:
        _employee_today_status[today] = {}
    if today not in _employee_leave_counts:
        _employee_leave_counts[today] = {}
    import employee_dept_filter as _edf
    site_emps = _edf.filter_prod_site(_employees_master_list)
    site_names = {e["name"] for e in site_emps}
    statuses = {
        k: v
        for k, v in _employee_today_status[today].items()
        if k in site_names
    }
    leave_counts = {
        k: v
        for k, v in _employee_leave_counts[today].items()
        if k in site_names
    }
    return jsonify({
        "date": today,
        "statuses": statuses,
        "leave_counts": leave_counts,
    })

@app.route('/api/employee/status', methods=['POST'])
def update_employee_status():
    """更新员工考勤状态 - 员工只能改自己的，管理可改所有人"""
    username = resolve_login_user()
    if not username:
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(username)
    if not user:
        return jsonify({"error": "无权限", "code": 403}), 403
    
    data = request.get_json()
    name = data.get('name', '')
    status = data.get('status', '出勤')
    
    # 员工只能改自己的
    if user['role'] == '员工':
        if name != user['employee_name']:
            return jsonify({"error": "无权限：仅可修改自己的考勤", "code": 403}), 403

    global _employee_today_status, _employee_leave_counts
    today = datetime.date.today().isoformat()
    if today not in _employee_today_status:
        _employee_today_status[today] = {}
    if today not in _employee_leave_counts:
        _employee_leave_counts[today] = {}
    if name:
        old_status = _employee_today_status[today].get(name, '出勤')
        if status == '离岗' and old_status != '离岗':
            _employee_leave_counts[today][name] = _employee_leave_counts[today].get(name, 0) + 1
        _employee_today_status[today][name] = status
        persist()  # 持久化到文件
    return jsonify({
        "date": today,
        "name": name,
        "status": status,
        "leave_count": _employee_leave_counts[today].get(name, 0) if name else 0,
        "leave_counts": _employee_leave_counts[today],
        "success": True
    })

# ==================== 订单备注和加急API ====================
@app.route('/api/order/remark', methods=['POST'])
def order_remark():
    """设置订单备注 - 仅 超级管理员/管理/客服 可操作"""
    if not resolve_login_user():
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] == '员工':
        return jsonify({"error": "无权限", "code": 403}), 403
    global _order_extra
    data = request.get_json()
    oid = data.get('order_id', '')
    remark = data.get('remark', '')
    if oid:
        if oid not in _order_extra:
            _order_extra[oid] = {}
        _order_extra[oid]['remark'] = remark
        persist()
    return jsonify({"success": True, "order_id": oid, "remark": remark})

@app.route('/api/order/urgent', methods=['POST'])
def order_urgent():
    """切换加急状态 - 仅 超级管理员/管理/客服 可操作"""
    if not resolve_login_user():
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] == '员工':
        return jsonify({"error": "无权限", "code": 403}), 403
    global _order_extra
    data = request.get_json()
    oid = data.get('order_id', '')
    urgent = data.get('urgent', False)
    if oid:
        if oid not in _order_extra:
            _order_extra[oid] = {}
        _order_extra[oid]['urgent'] = urgent
        _order_extra_store.upsert_urgent(get_db, oid, urgent)
        persist()
        _order_extra_store.mirror_order_extra_to_data_json(_order_extra)
    return jsonify({"success": True, "order_id": oid, "urgent": urgent})

# ==================== 导出今日考勤CSV ====================
@app.route('/api/employee/export')
def export_attendance():
    global _employee_today_status
    date_str = request.args.get('date', datetime.date.today().isoformat())
    today = date_str
    # 如果请求的日期没有数据，初始化为空
    if today not in _employee_today_status:
        _employee_today_status[today] = {}

    _employees_master_list = [
        # ===== 洋坑塘运营部 =====
        {"name": "戴雅利", "position": "总负责人", "group": "管理层", "dept": "洋坑塘运营部", "phone": "13800138001"},
        {"name": "苏世婷", "position": "阿里客服主管", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138002"},
        {"name": "廖思美", "position": "淘天客服主管", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138003"},
        {"name": "何水单", "position": "业务员", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138004"},
        {"name": "李四军", "position": "美工", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138005"},
        {"name": "陈贤聪", "position": "抖音/微信/珍珠棉业务", "group": "业务组", "dept": "洋坑塘运营部", "phone": "13800138006"},
        {"name": "张慧平", "position": "友尚旗舰店运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138007"},
        {"name": "陈贝贝", "position": "亚润跨境运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138008"},
        {"name": "罗怡", "position": "三羊现货运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138009"},
        {"name": "周井梅", "position": "正方形/大鱼运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138010"},
        {"name": "戴志美", "position": "友尚/新鑫星运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138011"},
        {"name": "石梅清", "position": "止合/扣底盒运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138012"},
        {"name": "张文杰", "position": "小批量/品牌店运营", "group": "客服组", "dept": "洋坑塘运营部", "phone": "13800138013"},
        {"name": "姚斌", "position": "采购/跟单/机动", "group": "综合组", "dept": "洋坑塘运营部", "phone": "13800138014"},
        # ===== 美丽湾工厂部-管理层 =====
        {"name": "邓涛", "position": "生产总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138015"},
        {"name": "黄兴", "position": "行政总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138016"},
        {"name": "隆浪", "position": "财务总监", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138017"},
        {"name": "蒋义红", "position": "飞机盒主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138018"},
        {"name": "沈齐豪", "position": "纸箱主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138019"},
        {"name": "覃海霞", "position": "综合主管", "group": "管理层", "dept": "美丽湾工厂部", "phone": "13800138020"},
        # ===== 美丽湾工厂部-手啤机组 =====
        {"name": "李方国", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138021"},
        {"name": "李周海", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138022"},
        {"name": "唐美章", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138023"},
        {"name": "蒋保平", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138024"},
        {"name": "陈章远", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138025"},
        {"name": "黄振辉", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138026"},
        {"name": "汪成桃", "position": "手啤机-啤工", "group": "手啤机组", "dept": "美丽湾工厂部", "phone": "13800138027"},
        # ===== 美丽湾工厂部-啤机分组 =====
        {"name": "易新明", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138028"},
        {"name": "雷善炎", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138029"},
        {"name": "陈海斌", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138030"},
        {"name": "陈奕升", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138031"},
        {"name": "江旭松", "position": "啤机分纸", "group": "啤机分组", "dept": "美丽湾工厂部", "phone": "13800138032"},
        # ===== 美丽湾工厂部-刀模组 =====
        {"name": "廖金玲", "position": "刀模-找刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138033"},
        {"name": "陈勇", "position": "刀模-组刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138034"},
        {"name": "张吉杰", "position": "刀模-拆刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138035"},
        {"name": "唐孝定", "position": "放刀模", "group": "刀模组", "dept": "美丽湾工厂部", "phone": "13800138036"},
        # ===== 美丽湾工厂部-机械手组 =====
        {"name": "黄恒", "position": "机械手啤工", "group": "机械手组", "dept": "美丽湾工厂部", "phone": "13800138037"},
        {"name": "李荣晖", "position": "机械手啤工", "group": "机械手组", "dept": "美丽湾工厂部", "phone": "13800138038"},
        # ===== 美丽湾工厂部-平压平组 =====
        {"name": "蒋森响", "position": "平压平啤工", "group": "平压平组", "dept": "美丽湾工厂部", "phone": "13800138039"},
        {"name": "周业福", "position": "平压平啤工", "group": "平压平组", "dept": "美丽湾工厂部", "phone": "13800138040"},
        # ===== 美丽湾工厂部-车间打包组 =====
        {"name": "陈桂英", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138041"},
        {"name": "毛良芬", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138042"},
        {"name": "陈辉文", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138043"},
        {"name": "文小梅", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138044"},
        {"name": "帅行朝", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138045"},
        {"name": "黄张华", "position": "车间打包", "group": "车间打包组", "dept": "美丽湾工厂部", "phone": "13800138046"},
        # ===== 美丽湾工厂部-仓库组 =====
        {"name": "宋小国", "position": "仓库-找货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138047"},
        {"name": "唐忠群", "position": "仓库-找货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138048"},
        {"name": "蒋仁叶", "position": "仓库-打包", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138049"},
        {"name": "罗照权", "position": "仓库-打包", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138050"},
        {"name": "黄爱小", "position": "仓库-放货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138051"},
        {"name": "龙雪兰", "position": "仓库-打样兼配货", "group": "仓库组", "dept": "美丽湾工厂部", "phone": "13800138052"},
        # ===== 美丽湾工厂部-印刷组 =====
        {"name": "李双", "position": "印刷", "group": "印刷组", "dept": "美丽湾工厂部", "phone": "13800138053"},
        # ===== 美丽湾工厂部-扣底盒 =====
        {"name": "蒋军林", "position": "扣底盒", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138054"},
        {"name": "宁哈姝", "position": "扣底盒", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138055"},
        {"name": "黄华", "position": "扣底盒-打包", "group": "扣底盒组", "dept": "美丽湾工厂部", "phone": "13800138056"},
        # ===== 美丽湾工厂部-纸箱部(分纸) =====
        {"name": "陈伟", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138057"},
        {"name": "陈陶", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138058"},
        {"name": "周石国", "position": "纸箱-分纸", "group": "纸箱分纸组", "dept": "美丽湾工厂部", "phone": "13800138059"},
        # ===== 美丽湾工厂部-纸箱部(开槽) =====
        {"name": "杨家忠", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138060"},
        {"name": "沈奇严", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138061"},
        {"name": "姚建桥", "position": "纸箱-开槽", "group": "纸箱开槽组", "dept": "美丽湾工厂部", "phone": "13800138062"},
        # ===== 美丽湾工厂部-纸箱部(粘胶) =====
        {"name": "戴西华", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138063"},
        {"name": "魏小王", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138064"},
        {"name": "黄芳", "position": "纸箱-粘胶", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138065"},
        {"name": "于天发", "position": "纸箱-粘胶/打角", "group": "纸箱粘胶组", "dept": "美丽湾工厂部", "phone": "13800138066"},
    ]
    statuses = _employee_today_status.get(today, {})

    lines = []
    lines.append('\ufeff' + '姓名,职务,厂区,手机号,今日状态')
    for emp in _employees_master_list:
        s = statuses.get(emp['name'], '出勤')
        lines.append(f"{emp['name']},{emp['position']},{emp['dept']},{emp['phone']},{s}")

    return Response(
        '\n'.join(lines),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=考勤报表_{today}.csv'}
    )


# ==================== 导出日报表CSV ====================
@app.route('/api/report/export')
def export_report():
    date = request.args.get('date', datetime.date.today().strftime('%Y-%m-%d'))
    # 模拟数据
    summary = {
        "total_orders": 45,
        "completed_orders": 14,
        "output_qty": 12500,
        "defect_rate": "1.2%",
    }
    production_lines = [
        {"line": "飞机盒线-自动平压平", "output": 5800, "status": "正常"},
        {"line": "飞机盒线-机械手", "output": 3200, "status": "正常"},
        {"line": "飞机盒线-手啤", "output": 1500, "status": "正常"},
        {"line": "纸箱线", "output": 2000, "status": "正常"},
    ]
    done_orders = [
        {"id": "JST-20260425-005", "product": "飞机盒20*15*10（现货）", "qty": 800, "last_step": "打包发货", "done_time": "16:00"},
        {"id": "JST-20260425-003", "product": "飞机盒25*15*10", "qty": 300, "last_step": "打包发货", "done_time": "15:30"},
        {"id": "JST-20260426-001", "product": "飞机盒30*20*15", "qty": 200, "last_step": "清废", "done_time": "14:20"},
    ]

    lines = []
    lines.append('\ufeff日报表 - ' + date)
    lines.append('')
    lines.append('指标,数值')
    lines.append(f'总单数,{summary["total_orders"]}')
    lines.append(f'完成数,{summary["completed_orders"]}')
    lines.append(f'总产出,{summary["output_qty"]}')
    lines.append(f'不良率,{summary["defect_rate"]}')
    lines.append('')
    lines.append('生产线,产出数,状态')
    for l in production_lines:
        lines.append(f'{l["line"]},{l["output"]},{l["status"]}')
    lines.append('')
    lines.append('内部单号,产品,数量,完成工序,完成时间')
    for o in done_orders:
        lines.append(f'{o["id"]},{o["product"]},{o["qty"]},{o["last_step"]},{o["done_time"]}')

    return Response(
        '\n'.join(lines),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=日报表_{date}.csv'}
    )


# ==================== 权限管理API ====================

# 功能列表（保持不变）
PERM_FEATURES = ["首页","订单生产进度","扫码报工","日报表","数据看板","刀模","库存","原材料","快麦ERP","员工","权限管理","报价","实时订单"]
_PERM_LEGACY_KEY = "聚水潭"
_PERM_CURRENT_KEY = "快麦ERP"


def _migrate_perm_dict(perms):
    if not isinstance(perms, dict):
        return perms
    if _PERM_LEGACY_KEY in perms and _PERM_CURRENT_KEY not in perms:
        perms[_PERM_CURRENT_KEY] = perms.pop(_PERM_LEGACY_KEY)
    elif _PERM_LEGACY_KEY in perms:
        perms.pop(_PERM_LEGACY_KEY, None)
    return perms


def _normalize_role_name(role: str) -> str:
    """岗位/角色名称统一（管理员→管理）。"""
    r = (role or "员工").strip()
    if r in ("管理员", "主管"):
        return "管理"
    return r


def _users_role_for_perm_role(role: str) -> str:
    """写入 USERS.role 的值。"""
    r = _normalize_role_name(role)
    if r == "超级管理员":
        return "超级管理员"
    if r == "管理":
        return "管理"
    if r in ("客服", "员工", "财务", "业务员"):
        return r
    return "员工"


def _builtin_role_permissions() -> dict[str, dict[str, bool]]:
    all_true = {f: True for f in PERM_FEATURES}
    mgmt = {
        "首页": True,
        "订单生产进度": True,
        "扫码报工": True,
        "日报表": True,
        "数据看板": True,
        "刀模": True,
        "刀模库": True,
        "库存": True,
        "原材料": True,
        "快麦ERP": True,
        "员工": True,
        "报价": True,
        "实时订单": True,
        "权限管理": False,
    }
    cs = {
        "首页": True,
        "订单生产进度": True,
        "快麦ERP": True,
        "刀模": True,
        "刀模库": True,
        "库存": True,
        "原材料": True,
        "员工": True,
        "报价": True,
        "实时订单": True,
    }
    worker = {
        "首页": True,
        "扫码报工": True,
        "日报表": True,
        "员工": True,
    }
    return {
        "超级管理员": dict(all_true),
        "管理": dict(mgmt),
        "管理员": dict(mgmt),
        "主管": dict(mgmt),
        "客服": dict(cs),
        "员工": dict(worker),
        "财务": {"首页": True, "数据看板": True, "快麦ERP": True, "员工": True},
        "业务员": {"首页": True, "快麦ERP": True, "刀模": True, "报价": True, "员工": True},
    }


def _infer_employee_role(employee_name: str) -> str:
    er = _permission_data.get("employee_roles") or {}
    if employee_name in er:
        return _normalize_role_name(er[employee_name])
    for u in USERS.values():
        if u.get("employee_name") == employee_name:
            return _normalize_role_name(u.get("role", "员工"))
    return "员工"


def _role_template_permissions(role: str) -> dict[str, bool]:
    role = _normalize_role_name(role)
    stored = _permission_data.get("role_permissions") or {}
    # 兼容 positions 里写的「管理员」等别名
    for key in (role, "管理员" if role == "管理" else None, "主管" if role == "管理" else None):
        if not key:
            continue
        tpl = stored.get(key)
        if isinstance(tpl, dict) and tpl:
            out = {}
            for f in PERM_FEATURES:
                out[f] = bool(tpl.get(f, False))
            return out
    builtin = _builtin_role_permissions()
    tpl = builtin.get(role) or builtin.get("员工", {})
    out = {f: False for f in PERM_FEATURES}
    for f, v in tpl.items():
        if f in PERM_FEATURES:
            out[f] = bool(v)
    return out


def _get_effective_permissions(employee_name: str, user: dict | None = None) -> dict[str, bool]:
    """角色默认 + 个人覆盖（个人优先）。"""
    if employee_name == "戴雅利":
        return {f: True for f in PERM_FEATURES}
    if user and (user.get("is_system") or user.get("role") == "超级管理员"):
        return {f: True for f in PERM_FEATURES}
    role = _infer_employee_role(employee_name)
    if role == "超级管理员":
        return {f: True for f in PERM_FEATURES}
    effective = _role_template_permissions(role)
    personal = _permission_data.get("permissions", {}).get(employee_name, {})
    if isinstance(personal, dict):
        personal = _migrate_perm_dict(dict(personal))
        for f in PERM_FEATURES:
            if f in personal:
                effective[f] = bool(personal[f])
    return effective


def _apply_employee_roles_to_users(employee_roles: dict) -> None:
    """3001/3002 不再根据 employee_roles 改写 USERS.role（避免 admin 被降为员工）。"""
    for uid, u in USERS.items():
        if _perm_resolve.is_super_admin_account(uid, u):
            u["role"] = "超级管理员"


def _ensure_role_permissions_defaults() -> None:
    rp = _permission_data.setdefault("role_permissions", {})
    positions = _permission_data.get("positions") or list(_builtin_role_permissions().keys())
    builtin = _builtin_role_permissions()
    for pos in positions:
        if pos not in rp or not isinstance(rp.get(pos), dict):
            rp[pos] = dict(builtin.get(pos) or builtin.get(_normalize_role_name(pos)) or builtin["员工"])


def _sync_all_employees_perms():
    """
    同步员工角色映射；permissions 仅存「相对角色的个人覆盖」，不再整表填 false。
    """
    employee_roles = _permission_data.setdefault("employee_roles", {})
    # 戴雅利 / admin 永不被 vault 或推断逻辑降为普通员工
    employee_roles["戴雅利"] = "超级管理员"
    for emp in _employees_master_list:
        name = emp["name"]
        if name not in employee_roles:
            employee_roles[name] = _infer_employee_role(name)
    _ensure_role_permissions_defaults()
    base = _permission_data.setdefault("permissions", {})
    valid_names = {e["name"] for e in _employees_master_list}
    for name in list(base.keys()):
        if name not in valid_names:
            del base[name]
    for emp in _employees_master_list:
        name = emp["name"]
        if name not in base:
            base[name] = {}
        else:
            _migrate_perm_dict(base[name])
            role = _infer_employee_role(name)
            tpl = _role_template_permissions(role)
            for f in list(base[name].keys()):
                if f in tpl and bool(base[name][f]) == bool(tpl.get(f)):
                    del base[name][f]
    _permission_data["permissions"] = base
    _permission_data["employee_roles"] = employee_roles
    _apply_employee_roles_to_users(employee_roles)


import permission_resolve as _perm_resolve


def _merge_employee_permissions_from_db() -> None:
    _perm_resolve.merge_employee_permissions_from_db(_permission_data, get_db)


def _user_has_permission(user: dict | None, username: str, feature: str) -> bool:
    import perm_cs_prod as _pcp
    return _pcp.user_has_feature(user, username, feature)


# 初始化时执行一次（MySQL 仅补缺失权限键，不覆盖 data.json）
_sync_all_employees_perms()
_merge_employee_permissions_from_db()
for _uid, _u in list(USERS.items()):
    USERS[_uid] = _perm_resolve.normalize_user_record(_uid, _u)


@app.before_request
def _refresh_login_user_role():
    """每个 API 请求刷新 Session 中的 role（修复旧 Session 仍为「员工」）。"""
    if request.path.startswith("/api/") and session.get("username"):
        _active_user(session["username"])


DEFAULT_PRODUCTION_MATERIAL_MAPPING = [
    {"keywords": "特硬,外径特硬,内径特硬,D6D,国产,加硬", "label": "特硬"},
    {"keywords": "优质,Q7Q,特价", "label": "优质"},
    {"keywords": "台湾,进口,超硬", "label": "台湾"},
    {"keywords": "白色,W7W,双白", "label": "白色"},
    {"keywords": "黑色,黑卡", "label": "黑色"},
    {"keywords": "红色", "label": "红色"},
    {"keywords": "P6D", "label": "P6D"},
    {"keywords": "B坑,K7K,三层", "label": "B坑"},
    {"keywords": "EB坑,K636K,五层EB", "label": "EB坑"},
    {"keywords": "BC坑,K737K,五层BC", "label": "BC坑"},
    {"keywords": "E瓦,瓦楞", "label": "E瓦"},
]

_permission_data_init = {
    "processes": ["客服接单", "黄厂打印", "审单分单", "算料", "分纸", "啤机(自动平压平)", "啤机(机械手)", "手啤", "印刷", "开槽/打角", "清废", "打包发货"],
    "production_material_mapping": list(DEFAULT_PRODUCTION_MATERIAL_MAPPING),
    "positions": ["超级管理员", "主管", "客服", "员工", "财务", "业务员"],
    "employees": [
        {"name": "戴雅利", "position": "超级管理员"},
        {"name": "邓涛", "position": "主管"},
        {"name": "黄兴", "position": "主管"},
        {"name": "覃海霞", "position": "主管"},
        {"name": "蒋义红", "position": "主管"},
        {"name": "沈齐豪", "position": "主管"},
        {"name": "苏世婷", "position": "客服"},
        {"name": "廖思美", "position": "客服"},
        {"name": "张慧平", "position": "客服"},
        {"name": "陈贝贝", "position": "客服"},
        {"name": "罗怡", "position": "客服"},
        {"name": "周井梅", "position": "客服"},
        {"name": "戴志美", "position": "客服"},
        {"name": "石梅清", "position": "客服"},
        {"name": "张文杰", "position": "客服"},
        {"name": "何水单", "position": "业务员"},
        {"name": "李四军", "position": "员工"},
        {"name": "陈贤聪", "position": "业务员"},
        {"name": "姚斌", "position": "员工"},
        {"name": "隆浪", "position": "财务"},
    ],
    "permissions": {}
}
# 确保默认结构与现有数据合并
for key, default_val in _permission_data_init.items():
    if key not in _permission_data:
        _permission_data[key] = default_val
# 再次同步权限
_sync_all_employees_perms()

@app.route("/api/process_timeout", methods=["GET", "POST"])
def process_timeout_api():
    """工序超时（分钟）：{ timeouts: { 工序名: 分钟 } }"""
    global _permission_data
    if request.method == "GET":
        return jsonify(
            {
                "success": True,
                "timeouts": _permission_data.get("process_timeouts") or {},
            }
        )
    un = resolve_login_user()
    role = USERS.get(un or "", {}).get("role", "")
    if role not in ("超级管理员", "管理", "主管"):
        return jsonify({"success": False, "error": "无权限"}), 403
    body = request.get_json() or {}
    timeouts = body.get("timeouts")
    if not isinstance(timeouts, dict):
        return jsonify({"success": False, "error": "timeouts 格式错误"}), 400
    cleaned = {}
    for k, v in timeouts.items():
        name = str(k).strip()
        if not name:
            continue
        try:
            mins = int(v)
        except (TypeError, ValueError):
            continue
        if mins > 0:
            cleaned[name] = mins
    _permission_data["process_timeouts"] = cleaned
    persist()
    return jsonify({"success": True, "timeouts": cleaned, "message": "已保存"})


@app.route('/api/permissions/data')
def get_permissions_data():
    import perm_cs_prod as _pcp
    return jsonify({
        "permissions_mode": "guanli_only",
        "guanli_admin_url": _pcp.GUANLI_ADMIN_URL,
        "message": "功能权限请在 3003 统一管理后台配置",
        "employees": __import__("employee_dept_filter").filter_prod_site(_employees_master_list),
        "processes": _permission_data.get("processes", []),
    })

@app.route('/api/permissions/save', methods=['POST'])
def save_permissions_data():
    import perm_cs_prod as _pcp
    return jsonify({
        "success": False,
        "error": "3001/3002 已关闭权限保存，请登录 3003 管理后台",
        "guanli_admin_url": _pcp.GUANLI_ADMIN_URL,
    }), 403


def _production_material_mapping() -> list:
    """生产材质映射：读 data.json（admin 页面维护）。"""
    return list(_permission_data.get("production_material_mapping") or [])


@app.route("/api/production/material-mapping", methods=["GET"])
def get_production_material_mapping():
    if not resolve_login_user():
        return jsonify({"error": "未登录"}), 401
    _cfg_json.reload_permission_memory(_permission_data)
    return jsonify({"success": True, "mapping": _production_material_mapping()})


@app.route("/api/production/material-mapping", methods=["POST"])
def save_production_material_mapping():
    if not resolve_login_user():
        return jsonify({"error": "未登录"}), 401
    body = request.get_json() or {}
    mapping = body.get("mapping")
    if not isinstance(mapping, list):
        return jsonify({"success": False, "message": "mapping 须为数组"}), 400
    _permission_data["production_material_mapping"] = mapping
    if not persist():
        return jsonify({"success": False, "message": _perm_save_detail.get("vault_error") or "保存失败"}), 500
    return jsonify({"success": True, "mapping": mapping, "message": "已保存", **_perm_save_detail})

@app.route('/api/processes')
def get_processes():
    _cfg_json.reload_permission_memory(_permission_data)
    return jsonify({"processes": _permission_data.get("processes", [])})

@app.route('/api/process_tree')
def get_process_tree():
    _cfg_json.reload_permission_memory(_permission_data)
    return jsonify({"tree": _permission_data.get("processes", [])})


@app.route('/api/my_permissions')
def get_my_permissions():
    """3001/3002：登录即全部功能可用；权限配置请用 3003。"""
    un = resolve_login_user()
    if not un:
        return jsonify({"error": "未登录", "code": 401}), 401
    user = _active_user(un)
    if not user:
        return jsonify({"error": "用户不存在", "code": 401}), 401
    import perm_cs_prod as _pcp
    import employee_dept_filter as _edf
    return jsonify(
        _pcp.my_permissions_response(
            user, all_employees=_edf.filter_prod_site(_employees_master_list)
        )
    )


@app.route('/api/processes/save', methods=['POST'])
def save_processes():
    global _permission_data
    data = request.get_json()
    flows_resynced = 0
    if data and 'processes' in data:
        _permission_data['processes'] = data['processes']
        if not persist():
            return jsonify({"success": False, "message": _perm_save_detail.get("vault_error") or "保存失败"}), 500
        _cfg_json.reload_permission_memory(_permission_data)
        flows_resynced = ph.resync_active_flows_from_tree(DB_CONFIG, data['processes'])
    return jsonify({
        "success": True,
        "message": "流程已保存",
        "flows_resynced": flows_resynced,
    })

# ==================== 扫码报工-可选操作人（按角色过滤） ====================
@app.route('/api/scan_workers')
def scan_workers():
    """根据当前用户角色返回可选择的操作人列表"""
    if not resolve_login_user():
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user:
        return jsonify({"error": "用户不存在", "code": 401}), 401
    
    role = user['role']
    current_name = user['employee_name']
    import employee_dept_filter as _edf
    all_emps = _edf.filter_prod_site(_employees_master_list)
    
    if role == '超级管理员':
        workers = all_emps
    elif role == '管理':
        workers = all_emps
    elif role == '客服':
        workers = [e for e in all_emps if e['dept'] == '洋坑塘运营部']
    elif role == '员工':
        # 员工只能选自己
        workers = [e for e in all_emps if e['name'] == current_name]
    else:
        workers = []
    
    return jsonify({"workers": workers})

# ==================== 报价系统 ====================
QUOTE_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'quote_data.json')

_quote_data_cache: dict = {"data": None, "ts": 0.0}

def load_quote_data():
    """优先 MySQL quote_config，与客服端一致；失败时回退 quote_data.json。有300s内存缓存。"""
    import time as _t
    now = _t.time()
    if _quote_data_cache["data"] is not None and now - _quote_data_cache["ts"] < 300:
        return _quote_data_cache["data"]
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT config_key, config_value FROM quote_config")
        rows = cur.fetchall()
        cur.close()
        db.close()
        result = {}
        for r in rows:
            key = r["config_key"]
            val = r["config_value"]
            if isinstance(val, str):
                try:
                    val = json.loads(val)
                except Exception:
                    pass
            result[key] = val
        if result:
            from quote_material_defaults import enrich_quote_data
            qd = enrich_quote_data(result)
            _quote_data_cache["data"] = qd
            _quote_data_cache["ts"] = now
            return qd
    except Exception as e:
        print(f"[load_quote_data] MySQL: {e}")
    try:
        with open(QUOTE_DATA_FILE, "r", encoding="utf-8") as f:
            from quote_material_defaults import enrich_quote_data
            qd = enrich_quote_data(json.load(f))
            _quote_data_cache["data"] = qd
            _quote_data_cache["ts"] = now
            return qd
    except Exception:
        return None


def save_quote_data(qd):
    """写入 MySQL quote_config，并同步 quote_data.json 备份。"""
    from quote_material_defaults import enrich_quote_data

    qd = enrich_quote_data(qd or {})
    ok_mysql = False
    try:
        db = get_db()
        cur = db.cursor()
        for key, val in qd.items():
            cur.execute(
                "INSERT INTO quote_config (config_key, config_value) VALUES (%s,%s) "
                "ON DUPLICATE KEY UPDATE config_value=VALUES(config_value)",
                (key, json.dumps(val, ensure_ascii=False)),
            )
        cur.close()
        db.close()
        ok_mysql = True
    except Exception as e:
        print(f"[save_quote_data] MySQL: {e}")
    ok_file = False
    try:
        with open(QUOTE_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(qd, f, ensure_ascii=False, indent=2)
        ok_file = True
    except Exception as e:
        print(f"[save_quote_data] file: {e}")
    return ok_mysql or ok_file


def _quote_material_mapping_for_spec() -> list[dict]:
    """报价参数 material_mapping → 打单/算料匹配（兼容 material_name / label）。"""
    rows = (load_quote_data() or {}).get("material_mapping") or []
    out: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        label = (row.get("material_name") or row.get("label") or "").strip()
        if not label:
            continue
        out.append(
            {
                "label": label,
                "keywords": (row.get("keywords") or "").strip(),
                "material_name": label,
                "material_key": row.get("material_key") or "",
            }
        )
    return out


def ceil_to_half(val):
    """向上取整到0.5：23.32→23.5, 23.62→24"""
    import math
    return math.ceil(val * 2) / 2

def ceil_to_spec(val, spec_list, multiply_limit=3):
    """纸箱度数取整：如果val不在spec_list中，尝试×2/×3取奇数再除"""
    import math
    
    # 先检查val本身是否在规格表中或大于等于最小规格
    min_spec = min(spec_list)
    if val >= min_spec:
        for s in spec_list:
            if val <= s:
                return s
        # 超过最大规格，取最大
        return max(spec_list)
    
    # val小于最小规格，尝试×multiply
    for mult in range(2, multiply_limit + 1):
        multiplied = val * mult
        # 取大于multiplied的奇数（纸箱度数都是奇数）
        ceil_odd = math.ceil(multiplied)
        if ceil_odd % 2 == 0:
            ceil_odd += 1
        # 检查这个奇数是否能在spec_list中找到
        if ceil_odd >= min_spec:
            for s in spec_list:
                if ceil_odd <= s:
                    return s / mult
            return max(spec_list) / mult
    
    # 保底
    return min_spec / multiply_limit

def calc_airbox(length, width, height, qty, material_key, quote_data, is_buckle=False):
    """标准飞机盒/带扣飞机盒报价"""
    import math
    
    # 纸长/纸度公式
    if is_buckle:
        paper_l_cm = 3 * height + 2 * width + 3
    else:
        paper_l_cm = 3 * height + 2 * width
    
    paper_w_cm = length + 4 * height + 4
    
    # 取整到0.5
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价
    mat_config = quote_data.get("materials", {}).get("airbox", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.35)
    mat_name = mat_info.get("name", material_key)
    
    # 材料成本 = 纸长×纸度×单价/1000
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    
    # 零售价 = 材料成本 × 2
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult
    
    import quote_calc_core as qcc

    redline_mult = qcc.redline_multiplier(quote_data)
    batch_mult = qcc.batch_multiplier_for_qty(qty, quote_data)

    # ====== 重量计算 ======
    # 展开面积(平方英寸) = 纸长×纸度
    # 转平方米: 1平方英寸 = 0.00064516 平方米
    gram_weight = mat_info.get("gram_weight", 450)  # 克/平方米
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "带扣飞机盒" if is_buckle else "标准飞机盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": redline_mult,
        "suggested_price": round(material_cost_per_unit * redline_mult, 4),
        "batch_multiplier": batch_mult,
        "batch_suggested_price": (
            round(material_cost_per_unit * batch_mult, 4) if batch_mult else None
        ),
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_koudi(length, width, height, qty, material_key, quote_data):
    """扣底盒报价"""
    import math
    
    # 纸长 = (长×2 + 宽×2 + 3) / 2.54
    paper_l_cm = length * 2 + width * 2 + 3
    # 纸度 = (宽 + 高 + 1/2宽 + 3 + 2) / 2.54
    paper_w_cm = width + height + width / 2 + 3 + 2
    
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价（扣底盒体系）
    mat_config = quote_data.get("materials", {}).get("koudi", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.40)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 450)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "扣底盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": None,
        "suggested_price": None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_shuangcha(length, width, height, qty, material_key, quote_data):
    """双插盒报价"""
    import math
    
    # 纸长 = (长×2 + 宽×2 + 3) / 2.54
    paper_l_cm = length * 2 + width * 2 + 3
    # 纸度 = (宽×2 + 6 + 高) / 2.54
    paper_w_cm = width * 2 + 6 + height
    
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价（扣底盒体系，跟扣底盒用同一套）
    mat_config = quote_data.get("materials", {}).get("koudi", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.40)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 450)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "双插盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": None,
        "suggested_price": None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_carton(length, width, height, qty, material_key, quote_data):
    """纸箱报价"""
    import math
    
    specs = quote_data.get("cardboard_specs", {})
    paper_lengths = specs.get("danbu_paper_lengths", [29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90])
    paper_widths = specs.get("danbu_paper_widths", [29,31,33,35,37,39,41,43,45,47,49])
    max_danbu_cm = specs.get("max_danbu_length_cm", 230)
    
    # 判断单卜还是双卜
    danbu_paper_l_cm = (length + width) * 2 + 3.5
    is_double = danbu_paper_l_cm > max_danbu_cm
    
    if is_double:
        # 双卜
        paper_l_cm = ((length + width) + 3.5) * 2
    else:
        # 单卜
        paper_l_cm = danbu_paper_l_cm
    
    paper_w_cm = width + height + 0.6
    
    # 纸长取整到度数表
    paper_l_raw = paper_l_cm / 2.54
    paper_l_inch = ceil_to_spec(paper_l_raw, paper_lengths)
    
    # 纸度取整到度数表
    paper_w_raw = paper_w_cm / 2.54
    paper_w_inch = ceil_to_spec(paper_w_raw, paper_widths)
    
    # 获取材料单价（纸箱体系）
    mat_config = quote_data.get("materials", {}).get("carton", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.30)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult
    import quote_calc_core as qcc

    redline_mult = qcc.redline_multiplier(quote_data)

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 380)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "双卜纸箱" if is_double else "单卜纸箱",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "is_double_buckle": is_double,
        "suggested_multiplier": redline_mult,
        "suggested_price": round(material_cost_per_unit * redline_mult, 4),
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

@app.route('/api/quote_data')
def get_quote_data():
    """返回报价基础数据（MySQL quote_config 优先）。"""
    qd = load_quote_data()
    if not qd:
        return jsonify({"success": False, "error": "报价数据未加载"})
    return jsonify({"success": True, "data": qd})

@app.route('/api/quote/save_config', methods=['POST'])
def save_quote_config():
    """保存报价配置（仅权限管理角色可操作）"""
    global _permission_data
    try:
        # 检查权限
        if not resolve_login_user():
            return jsonify({"success": False, "error": "未登录"})
        user = USERS.get(session['username'])
        if not user:
            return jsonify({"success": False, "error": "用户不存在"})
        username = session.get("username") or ""
        if not _user_has_permission(user, username, "权限管理"):
            return jsonify({"success": False, "error": "无权限修改报价配置"})
        
        patch = request.get_json()
        if not patch:
            return jsonify({"success": False, "error": "数据为空"})

        existing = load_quote_data() or {}
        merged = merge_quote_config(existing, patch)
        if not save_quote_data(merged):
            return jsonify({"success": False, "error": "保存失败"})
        sync_detail: dict = {}
        if isinstance(patch, dict) and "material_mapping" in patch:
            sync_detail = _cfg_json.sync_production_mapping_from_quote(
                _permission_data, merged.get("material_mapping")
            )
        msg = "报价配置已保存"
        if sync_detail.get("vault_error"):
            msg += "（" + sync_detail["vault_error"] + "）"
        return jsonify({"success": True, "message": msg, "sync": sync_detail})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/quote/calculate', methods=['POST'])
def calculate_quote():
    """计算报价"""
    try:
        data = request.get_json()
        ptype = data.get('type', 'zhengsquare')  # zhengsquare, juxing, daikou, koudi, shuangcha, qita
        length = float(data.get('length', 0))
        width = float(data.get('width', 0))
        height = float(data.get('height', 0))
        qty = int(data.get('qty', 1000))
        material = data.get('material', 'd6d')
        discount = float(data.get('discount', 100)) / 100.0  # 折扣百分比转小数

        if length <= 0 or width <= 0 or height <= 0:
            return jsonify({"error": "请输入正确的尺寸", "success": False})
        
        import math

        qd = load_quote_data()
        if not qd:
            return jsonify({"error": "报价数据未加载", "success": False})

        # 兼容 zhengsquare-outer / zhengsquare-inner
        # 内径转外径：长+1.5cm，宽+0.5cm，高+0.5cm
        if ptype == 'zhengsquare-inner':
            length = length + 1.5
            width = width + 0.5
            height = height + 0.5
        
        base_type = ptype.replace('-outer', '').replace('-inner', '')
        if base_type == 'zhengsquare':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=False)
            # 标注外径/内径
            if ptype == 'zhengsquare-inner':
                detail['dimension_label'] = '内径转外径'
            else:
                detail['dimension_label'] = '外径'
        elif base_type == 'juxing':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=False)
        elif ptype == 'daikou':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=True)
        elif ptype == 'koudi':
            detail = calc_koudi(length, width, height, qty, material, qd)
        elif ptype == 'shuangcha':
            detail = calc_shuangcha(length, width, height, qty, material, qd)
        elif ptype == 'qita':
            detail = calc_carton(length, width, height, qty, material, qd)
        elif ptype == 'pe':
            # 珍珠棉报价 - 输入单位毫米
            length_m = length / 1000  # 毫米转米
            width_m = width / 1000    # 毫米转米
            
            # 厚度向上取整到0.5mm的倍数（0.5阶梯）
            # 0~0.5→0.5, 0.6~1.0→1.0, 1.1~1.5→1.5, 1.6~2.0→2.0
            height_raw = height  # 保留原始值用于显示
            height = math.ceil(height / 0.5) * 0.5
            if height == 0:
                height = 0.5  # 最小0.5
            
            pe_price = qd.get("materials", {}).get("pe", {}).get("price", 0.3)
            material_cost_per_unit = length_m * width_m * height * pe_price
            
            # 客户类型
            customer_type = data.get('customer_type', 'guangdong_wholesale')
            customer_types = qd.get("materials", {}).get("pe", {}).get("customer_types", {})
            ct_info = customer_types.get(customer_type, {})
            multiplier = ct_info.get("multiplier", 2.0)
            ct_name = ct_info.get("name", customer_type)
            
            sell_price_per_unit = material_cost_per_unit * multiplier
            retail_mult = qd.get("profit", {}).get("retail_multiplier", 2.0)
            
            # 总货款判断
            total_price = sell_price_per_unit * qty
            total_price_discounted = total_price * discount
            price_too_low = total_price_discounted < 100  # 总货款低于100元
            unit_price_too_low = sell_price_per_unit < 0.3  # 单价低于0.3元
            
            # 重量预估（泡比5000）和快递费
            height_cm = height / 10  # 毫米转厘米
            length_cm = length / 10  # 毫米转厘米
            width_cm = width / 10    # 毫米转厘米
            weight_per_unit_kg = height_cm * length_cm * width_cm / 5000
            
            # 快递费（首重+续重）
            freight_config = qd.get("materials", {}).get("pe", {}).get("freight", {})
            
            def calc_freight(total_kg, cfg):
                first = cfg.get("first_kg", 10)
                renew = cfg.get("renew_per_kg", 0.9)
                if total_kg <= 1:
                    return round(first, 2)
                else:
                    extra = total_kg - 1
                    return round(first + extra * renew, 2)
            
            freight_per_unit_gd = round(calc_freight(weight_per_unit_kg, freight_config.get("guangdong", {})), 4)
            freight_per_unit_jzh = round(calc_freight(weight_per_unit_kg, freight_config.get("jiangzhehu", {})), 4)
            freight_per_unit_other = round(calc_freight(weight_per_unit_kg, freight_config.get("other", {})), 4)
            
            total_weight_kg = weight_per_unit_kg * qty
            freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
            freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
            freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))
            
            detail = {
                "product": "珍珠棉片材",
                "material_name": "珍珠棉",
                "unit_price": pe_price,
                "length_m": round(length_m, 4),
                "width_m": round(width_m, 4),
                "thickness_mm": height,
                "thickness_input_mm": height_raw,
                "customer_type": ct_name,
                "multiplier": multiplier,
                "material_cost_per_unit": round(material_cost_per_unit, 4),
                "retail_multiplier": retail_mult,
                "sell_price_per_unit": round(sell_price_per_unit, 4),
                "total_cost": round(material_cost_per_unit * qty, 2),
                "total_price": round(total_price, 2),
                "total_price_discounted": round(total_price_discounted, 2),
                "sell_price_per_unit_discounted": round(sell_price_per_unit * discount, 4),
                "qty": qty,
                "weight_per_unit_kg": round(weight_per_unit_kg, 4),
                "weight_total_kg": round(total_weight_kg, 2),
                "freight_per_unit_gd": freight_per_unit_gd,
                "freight_per_unit_jzh": freight_per_unit_jzh,
                "freight_per_unit_other": freight_per_unit_other,
                "freight_total_gd": freight_total_gd,
                "freight_total_jzh": freight_total_jzh,
                "freight_total_other": freight_total_other,
                "discount": round(discount * 100, 1),
                "suggested_multiplier": None,
                "suggested_price": None,
                "pe_warning_price_too_low": price_too_low,
                "pe_warning_unit_price_too_low": unit_price_too_low
            }
        else:
            return jsonify({"error": "未知的产品类型", "success": False})

        # 应用折扣
        detail["discount"] = round(discount * 100, 1)
        detail["sell_price_per_unit_discounted"] = round(detail["sell_price_per_unit"] * discount, 4)
        detail["total_price_discounted"] = round(detail["total_price"] * discount, 2)

        return jsonify({"success": True, "detail": detail})

    except Exception as e:
        return jsonify({"error": str(e), "success": False})

# ==================== 固定刀模库 API ====================
import dimoldb_store as _dimoldb_store
import dimoldb_inventory_api as _dim_inv_api


def load_dimoldb(force: bool = False):
    """从 MySQL 加载刀模库（内存缓存，默认 120s）。"""
    return _dimoldb_store.load_dimoldb_cached(get_db, force=force)


def save_dimoldb(data):
    """保存刀模库到 MySQL。"""
    ok = _dimoldb_store.save_dimoldb(get_db, data)
    if ok:
        _dimoldb_store.invalidate_dimoldb_cache()
        _invalidate_dim_match_index_cache()
    return ok


_inv_mem_cache: dict = {"data": None, "ts": 0.0}
_INV_CACHE_TTL = float(os.getenv("INV_CACHE_TTL_SEC", "300"))

_dim_match_index_cache: dict = {"index": None, "ts": 0.0}


def get_dim_match_index_cached():
    """刀模尺寸索引（库存/刀模列表共用，避免每请求重建）。"""
    import time as _t

    now = _t.time()
    if _dim_match_index_cache["index"] is not None and now - float(
        _dim_match_index_cache["ts"] or 0
    ) < 300.0:
        return _dim_match_index_cache["index"]
    rows = load_dimoldb()
    idx = _dimoldb_store.build_dim_match_index(rows)
    _dim_match_index_cache["index"] = idx
    _dim_match_index_cache["ts"] = now
    return idx


def _invalidate_dim_match_index_cache() -> None:
    _dim_match_index_cache["index"] = None
    _dim_match_index_cache["ts"] = 0.0


def load_inventory_cached():
    """库存列表用，避免同请求内重复全表读。"""
    import time as _t

    now = _t.time()
    if _inv_mem_cache["data"] is not None and now - float(
        _inv_mem_cache["ts"] or 0
    ) < _INV_CACHE_TTL:
        return _inv_mem_cache["data"]
    data = load_inventory()
    _inv_mem_cache["data"] = data
    _inv_mem_cache["ts"] = now
    return data


def _has_dimoldb_edit_perm():
    """检查是否有刀模库编辑权限"""
    import permission_resolve as _pr

    user = _active_user()
    if not user:
        return False
    un = session.get("username") or ""
    if _pr.is_super_admin_account(un, user):
        return True
    if user['role'] in ('超级管理员', '管理'):
        return True
    # 检查功能权限
    _sync_all_employees_perms()
    name = user['employee_name']
    my_perm = _permission_data.get("permissions", {}).get(name, {})
    if my_perm.get('刀模管理', False):
        return True
    return False


def _dimoldb_infer_inner_outer(dm):
    return _dimoldb_store.infer_inner_outer(dm)


@app.route('/api/dimoldb', methods=['GET'])
def get_dimoldb():
    """获取刀模列表（支持分页）"""
    ptype = request.args.get('type', '')
    search = request.args.get('search', '').strip()
    dim_type = request.args.get('dim_type', '').strip()
    data = _dim_inv_api.filter_dimoldb_rows(
        load_dimoldb(),
        ptype=ptype,
        search=search,
        dim_type=dim_type,
        infer_fn=_dimoldb_infer_inner_outer,
    )
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
    except ValueError:
        page, page_size = 1, 50
    page_data, total, page, page_size, total_pages = _dim_inv_api.paginate_rows(
        data, page, page_size
    )
    inv_all = load_inventory_cached()
    inv_items = inv_all.get('finished', inv_all if isinstance(inv_all, list) else [])
    _dim_inv_api.enrich_dimoldb_page(page_data, inv_items)
    return jsonify({
        "success": True,
        "data": page_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "can_edit": _has_dimoldb_edit_perm()
    })

@app.route('/api/dimoldb/single', methods=['GET'])
def get_dimoldb_single():
    """根据ID获取单个刀模"""
    dm_id = request.args.get('id', '').strip()
    if not dm_id:
        return jsonify({"success": False, "error": "缺少id参数"})
    data = load_dimoldb()
    for d in data:
        if d.get('id') == dm_id:
            return jsonify({"success": True, "data": d})
    return jsonify({"success": False, "error": "未找到"})

@app.route('/api/dimoldb', methods=['POST'])
def add_dimoldb():
    """新增刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json(silent=True)
        if not isinstance(item, dict):
            return jsonify({"success": False, "error": "请求体为空或 Content-Type 不是 application/json"})
        required = ['product_type', 'name', 'length', 'width', 'height']
        for k in required:
            if k not in item:
                return jsonify({"success": False, "error": f"缺少字段: {k}"})
        data = load_dimoldb()
        item['id'] = f"dm_{int(time.time())}_{len(data)}"
        item['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        data.append(item)
        save_dimoldb(data)
        return jsonify({"success": True, "message": "刀模已添加"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/dimoldb/<dm_id>', methods=['PUT'])
def update_dimoldb(dm_id):
    """修改刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json(silent=True)
        if not isinstance(item, dict):
            return jsonify({"success": False, "error": "请求体为空或 Content-Type 不是 application/json"})
        data = load_dimoldb()
        for d in data:
            if d.get('id') == dm_id:
                d.update(item)
                d['id'] = dm_id
                save_dimoldb(data)
                return jsonify({"success": True, "message": "刀模已更新"})
        return jsonify({"success": False, "error": "未找到该刀模"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/dimoldb/<dm_id>', methods=['DELETE'])
def delete_dimoldb(dm_id):
    """删除刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    data = load_dimoldb()
    new_data = [d for d in data if d.get('id') != dm_id]
    if len(new_data) == len(data):
        return jsonify({"success": False, "error": "未找到该刀模"})
    save_dimoldb(new_data)
    return jsonify({"success": True, "message": "刀模已删除"})

@app.route('/api/dimoldb/template', methods=['GET'])
def template_dimoldb():
    """刀模导入空模板"""
    return _send_xlsx_template(
        list(_dimoldb_store.LEGACY_IMPORT_HEADERS),
        [
            ["zhengsquare", "飞机盒18×14×5", "FH181405", "外径", 18, 14, 5],
            ["juxing", "长方形20×15×10", "", "", 20, 15, 10],
        ],
        "刀模库导入模板.xlsx",
    )


@app.route('/api/dimoldb/export', methods=['GET'])
def export_dimoldb():
    """导出所有刀模为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        db = load_dimoldb()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '刀模库'
        headers = ['序号', '产品类型', '名称', '编码', '备注', '长(cm)', '宽(cm)', '高(cm)', '创建时间']
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin
        for ri, item in enumerate(db, 2):
            vals = [
                ri - 1,
                item.get('product_type', ''),
                item.get('name', ''),
                item.get('code', '') or '',
                item.get('remark', '') or '',
                item.get('length', ''),
                item.get('width', ''),
                item.get('height', ''),
                item.get('created_at', '')
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 18
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        import flask
        resp = flask.send_file(tmp.name, as_attachment=True, download_name=f'刀模库_{datetime.date.today().isoformat()}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return resp
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()}), 500

@app.route('/api/dimoldb/import', methods=['POST'])
def import_dimoldb():
    """上传Excel导入刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        import openpyxl
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "请上传Excel文件"})
        f = request.files['file']
        if f.filename == '':
            return jsonify({"success": False, "error": "未选择文件"})
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        f.save(tmp.name)
        tmp.close()
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active
        # 探测表头
        header_row = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            vals = [str(v or '').strip() for v in row]
            if '名称' in vals or '刀模名称' in vals or 'name' in vals:
                header_row = r_idx
                break
        if header_row is None:
            return jsonify({"success": False, "error": "未找到表头行（需包含'名称'列）"})
        headers = [str(v or '').strip() for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col_map = _dimoldb_store.map_dimoldb_import_headers(headers)
        if 'name' not in col_map:
            return jsonify({"success": False, "error": f"未找到'名称'列，表头: {headers}"})
        mode = (request.form.get("import_mode") or request.form.get("mode") or "append").strip().lower()
        db = load_dimoldb()
        if mode == "overwrite":
            db = []
        by_code = {str(d.get("code") or "").strip(): i for i, d in enumerate(db) if d.get("code")}
        by_name = {str(d.get("name") or "").strip(): i for i, d in enumerate(db)}
        added = updated = 0
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[col_map['name']]).strip() if col_map['name'] < len(row) and row[col_map['name']] else ''
            if not name or name.startswith('=') or name == 'None':
                continue
            item = {
                'id': f'dm_{int(time.time())}_{added}_{len(db)}',
                'name': name,
                'product_type': _dimoldb_store.cell_str(row, col_map.get('product_type')) or 'zhengsquare',
                'code': _dimoldb_store.cell_str(row, col_map.get('code')),
                'production_spec': _dimoldb_store.cell_str(row, col_map.get('production_spec')),
                'km_mapping_code': _dimoldb_store.cell_str(row, col_map.get('km_mapping_code')),
                'remark': _dimoldb_store.cell_str(row, col_map.get('remark')),
                'length': _dimoldb_store.cell_float(row, col_map.get('length')),
                'width': _dimoldb_store.cell_float(row, col_map.get('width')),
                'height': _dimoldb_store.cell_float(row, col_map.get('height')),
                'created_at': now
            }
            key_code = (item.get("code") or "").strip()
            key_name = item.get("name") or ""
            if mode == "upsert":
                idx = by_code.get(key_code) if key_code else None
                if idx is None and key_name:
                    idx = by_name.get(key_name)
                if idx is not None:
                    old = db[idx]
                    item["id"] = old.get("id") or item["id"]
                    item["created_at"] = old.get("created_at") or now
                    db[idx] = {**old, **item}
                    if key_code:
                        by_code[key_code] = idx
                    by_name[key_name] = idx
                    updated += 1
                    continue
            db.append(item)
            if key_code:
                by_code[key_code] = len(db) - 1
            by_name[key_name] = len(db) - 1
            added += 1
        save_dimoldb(db)
        import os
        try: os.unlink(tmp.name)
        except: pass
        _prod_dash_cache.invalidate_dashboard_cache()
        return jsonify({
            "success": True,
            "message": f"导入完成：新增 {added} 条，更新 {updated} 条（模式 {mode}）",
            "added": added,
            "updated": updated,
            "mode": mode,
        })
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()})

@app.route('/api/inventory/template', methods=['GET'])
def template_inventory():
    """成品库存导入模板（与导出列一致）"""
    tab = request.args.get("tab", "finished")
    label = "成品库存" if tab == "finished" else "退货库存"
    return _send_xlsx_template(
        ["名称", "产品类型", "内/外径", "材质", "库存", "长(cm)", "宽(cm)", "高(cm)", "货位", "备注"],
        [
            ["10.5×7.5×4.5 特硬", "zhengsquare", "外径", "特硬", 100, 10.5, 7.5, 4.5, "A-01", ""],
            ["20×15×10", "juxing", "外径", "", 50, 20, 15, 10, "", ""],
        ],
        f"{label}导入模板.xlsx",
    )


# -------------------- 库存导出 --------------------
@app.route('/api/inventory/export', methods=['GET'])
def export_inventory():
    """导出库存为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        tab = request.args.get('tab', 'finished')
        data = load_inventory()
        items = data.get(tab, [])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '库存'
        headers = ['序号', '名称', '产品类型', '内/外径', '材质', '库存', '长(cm)', '宽(cm)', '高(cm)', '货位', '备注', '更新时间']
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin
        for ri, item in enumerate(items, 2):
            dim_label = {'inner': '内径', 'outer': '外径'}.get(item.get('dim_type', ''), '')
            vals = [
                ri - 1,
                item.get('name', ''),
                item.get('product_type', ''),
                dim_label,
                item.get('material', ''),
                int(item.get('stock', item.get('qty', 0))),
                item.get('length', ''),
                item.get('width', ''),
                item.get('height', ''),
                item.get('location', '') or '',
                item.get('remark', '') or '',
                item.get('updated_at', item.get('created_at', ''))
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
        for col_letter, width in zip('ABCDEFGHIJKL', [6, 30, 12, 10, 12, 8, 10, 10, 10, 12, 20, 18]):
            ws.column_dimensions[col_letter].width = width
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        tab_label = {'finished': '成品', 'returned': '退货'}.get(tab, tab)
        import flask
        resp = flask.send_file(tmp.name, as_attachment=True, download_name=f'{tab_label}库存_{datetime.date.today().isoformat()}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return resp
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()}), 500

@app.route('/api/dimoldb/search', methods=['POST', 'GET'])
def search_dimoldb():
    """查询某尺寸是否有固定刀模（供报价系统 / 小程序 / 首页快查）"""
    try:
        if request.method == 'GET':
            data = {
                'type': (request.args.get('type') or '').strip(),
                'dim_type': (request.args.get('dim_type') or '').strip(),
            }
            for k in ('length', 'width', 'height'):
                v = request.args.get(k)
                if v not in (None, ''):
                    try:
                        data[k] = float(v)
                    except (TypeError, ValueError):
                        pass
        else:
            data = request.get_json(silent=True) or {}
            if not data and request.data:
                try:
                    data = json.loads(request.data)
                except Exception:
                    data = {}
        if (
            not data.get('type')
            and data.get('length') is None
            and data.get('width') is None
            and data.get('height') is None
            and not data.get('dim_type')
        ):
            return jsonify({
                "success": True,
                "matches": [],
                "message": "请填写尺寸或类型后再查询",
            })
        matches = _dim_inv_api.search_dimoldb_matches(
            load_dimoldb(), data, _dimoldb_infer_inner_outer
        )
        return jsonify({"success": True, "matches": matches})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ==================== 库存 API ====================

def load_inventory():
    """从 MySQL 加载库存数据（与 app_cs / app.py 一致）。"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute(
            "SELECT id, name, spec, product_type, material, location, qty, "
            "last_month_qty, length, width, height, dim_type, created_at, updated_at "
            "FROM inventory ORDER BY created_at DESC"
        )
        rows = cur.fetchall()
        finished = []
        for r in rows:
            finished.append({
                'id': r['id'],
                'name': r['name'] or '',
                'spec': r['spec'] or '',
                'product_type': r['product_type'] or '',
                'material': r['material'] or '',
                'location': r['location'] or '',
                'qty': r['qty'] or 0,
                'last_month_qty': r['last_month_qty'] or 0,
                'length': float(r['length']) if r['length'] else 0,
                'width': float(r['width']) if r['width'] else 0,
                'height': float(r['height']) if r['height'] else 0,
                'dim_type': r['dim_type'] or '',
                'created_at': r['created_at'] or '',
                'updated_at': r['updated_at'] or ''
            })
        cur.close()
        db.close()
        return {"finished": finished, "raw": [], "returned": []}
    except Exception as e:
        print(f'[MySQL load_inventory] 错误: {e}')
        return {"finished": [], "raw": [], "returned": []}


def save_inventory(data):
    """保存库存到 MySQL（truncate + 批量 insert）。"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE inventory")
        items = []
        if isinstance(data, dict):
            items = data.get('finished', data.get('raw', data.get('returned', [])))
        elif isinstance(data, list):
            items = data
        for item in items:
            cur.execute(
                "INSERT INTO inventory (id, name, spec, product_type, material, location, "
                "qty, last_month_qty, length, width, height, dim_type, created_at, updated_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    item.get('id', ''),
                    item.get('name', ''),
                    item.get('spec', ''),
                    item.get('product_type', ''),
                    item.get('material', ''),
                    item.get('location', ''),
                    item.get('qty', 0),
                    item.get('last_month_qty', 0),
                    item.get('length', 0),
                    item.get('width', 0),
                    item.get('height', 0),
                    item.get('dim_type', ''),
                    item.get('created_at', ''),
                    item.get('updated_at', '')
                )
            )
        cur.close()
        db.close()
        _inv_mem_cache["data"] = None
        _inv_mem_cache["ts"] = 0.0
        return True
    except Exception as e:
        print(f'[MySQL save_inventory] 错误: {e}')
        return False

def _has_inv_edit_perm():
    """库存修改权限：超级管理员/管理/有'库存'权限"""
    if not resolve_login_user():
        return False
    user = USERS.get(session['username'])
    if not user:
        return False
    if user['role'] in ('超级管理员', '管理'):
        return True
    _sync_all_employees_perms()
    name = user['employee_name']
    my_perm = _permission_data.get("permissions", {}).get(name, {})
    if my_perm.get('库存', False):
        return True
    return False

@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    """获取库存列表（支持分页，同刀模库）"""
    tab = request.args.get('tab', 'finished')
    data = load_inventory_cached()
    items = data.get(tab, [])
    dm_index = get_dim_match_index_cached()
    _dim_inv_api.apply_inventory_material_labels(items)
    ptype = request.args.get('type', '')
    search = request.args.get('search', '').strip()
    length = request.args.get('length', '').strip()
    width = request.args.get('width', '').strip()
    height = request.args.get('height', '').strip()
    items = _dim_inv_api.filter_inventory_rows(
        items,
        ptype=ptype,
        search=search,
        length=length,
        width=width,
        height=height,
        search_field=request.args.get('search_field', 'all'),
    )
    _dim_inv_api.sort_inventory_rows(items)
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
    except ValueError:
        page, page_size = 1, 50
    page_data, total, page, page_size, total_pages = _dim_inv_api.paginate_rows(
        items, page, page_size
    )
    # 刀模快查只要数量/材质/位置，跳过逐行刀模匹配（否则全表 dimoldb 索引很慢）
    lite = request.args.get('lite') == '1'
    dim_only = (
        lite
        or (
            length
            and width
            and height
            and not search
            and not ptype
        )
    )
    if not dim_only:
        _dim_inv_api.enrich_inventory_page(page_data, dm_index, _dimoldb_infer_inner_outer)
    resp = jsonify({
        "success": True,
        "data": page_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "can_edit": _has_inv_edit_perm()
    })
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return resp

@app.route('/api/inventory', methods=['POST'])
def add_inventory():
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        tab = item.get('tab', 'finished')
        if 'name' not in item:
            return jsonify({"success": False, "error": "缺少名称"})
        data = load_inventory()
        item['id'] = f"inv_{int(time.time())}_{len(data.get(tab, []))}"
        item['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        item['updated_at'] = item['created_at']
        if tab not in data:
            data[tab] = []
        data[tab].append(item)
        save_inventory(data)
        return jsonify({"success": True, "message": "已添加"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>', methods=['PUT'])
def update_inventory(item_id):
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        data = load_inventory()
        for tab in ('finished', 'raw'):
            for d in data.get(tab, []):
                if d.get('id') == item_id:
                    d.update(item)
                    d['id'] = item_id
                    d['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    save_inventory(data)
                    return jsonify({"success": True, "message": "已更新"})
        return jsonify({"success": False, "error": "未找到"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>', methods=['DELETE'])
def delete_inventory(item_id):
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    data = load_inventory()
    for tab in ('finished', 'raw'):
        before = len(data.get(tab, []))
        data[tab] = [d for d in data.get(tab, []) if d.get('id') != item_id]
        if len(data[tab]) < before:
            save_inventory(data)
            return jsonify({"success": True, "message": "已删除"})
    return jsonify({"success": False, "error": "未找到"})

@app.route('/api/inventory/stock', methods=['POST'])
def stock_operation():
    """入库/出库操作（带log记录）"""
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        body = request.get_json()
        item_id = body.get('id')
        op = body.get('op', 'in')
        qty = int(body.get('qty', 0))
        remark = body.get('remark', '').strip()
        if not item_id or qty <= 0:
            return jsonify({"success": False, "error": "参数错误"})
        # 获取操作人
        operator = ''
        if 'username' in session:
            operator = USERS.get(session['username'], {}).get('employee_name', session['username'])
        data = load_inventory()
        for tab in ('finished', 'raw'):
            for d in data.get(tab, []):
                if d.get('id') == item_id:
                    current = int(d.get('stock', 0))
                    if op == 'in':
                        d['stock'] = current + qty
                    else:
                        if current < qty:
                            return jsonify({"success": False, "error": f"库存不足！当前{current}，出库{qty}"})
                        d['stock'] = current - qty
                    d['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    # 写log
                    if 'inout_log' not in d:
                        d['inout_log'] = []
                    op_label = '入库' if op == 'in' else '出库'
                    before_stock = current
                    after_stock = d['stock']
                    d['inout_log'].append({
                        "time": d['updated_at'],
                        "op": op,
                        "op_label": op_label,
                        "qty": qty,
                        "before": before_stock,
                        "after": after_stock,
                        "operator": operator,
                        "remark": remark
                    })
                    save_inventory(data)
                    return jsonify({
                        "success": True,
                        "message": f"{op_label}成功（{before_stock} → {after_stock}）",
                        "new_stock": d['stock']
                    })
        return jsonify({"success": False, "error": "未找到该商品"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>/log', methods=['GET'])
def get_inventory_log(item_id):
    """获取某商品的出入库历史记录"""
    data = load_inventory()
    for tab in ('finished', 'raw'):
        for d in data.get(tab, []):
            if d.get('id') == item_id:
                log = d.get('inout_log', [])
                return jsonify({
                    "success": True,
                    "item_name": d.get('name', ''),
                    "current_stock": d.get('stock', 0),
                    "log": log
                })
    return jsonify({"success": False, "error": "未找到该商品"})

@app.route('/api/inventory/single', methods=['GET'])
def get_inventory_single():
    """根据ID获取单个库存商品"""
    item_id = request.args.get('id', '').strip()
    if not item_id:
        return jsonify({"success": False, "error": "缺少ID"})
    data = load_inventory()
    for tab in ('finished', 'raw'):
        for d in data.get(tab, []):
            if d.get('id') == item_id:
                return jsonify({"success": True, "data": d})
    return jsonify({"success": False, "error": "未找到该商品"})


@app.route('/api/inventory/import_excel', methods=['POST'])
def import_inventory_excel():
    """导入excel库存台账"""
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        import openpyxl
        # 检查是否有上传文件
        if 'file' in request.files:
            f = request.files['file']
            if f.filename == '':
                return jsonify({"success": False, "error": "未选择文件"})
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            f.save(tmp.name)
            tmp.close()
            filepath = tmp.name
        else:
            # 兼容旧模式：从缓存目录查找
            data = request.get_json() or {}
            filepath = data.get('filepath', '')
            if not os.path.isabs(filepath):
                cache_dir = '/home/admin/.hermes/cache/documents/'
                if os.path.exists(cache_dir):
                    candidates = [f for f in os.listdir(cache_dir) if filepath in f and f.endswith('.xlsx')]
                    if candidates:
                        filepath = os.path.join(cache_dir, sorted(candidates, key=lambda f: os.path.getmtime(os.path.join(cache_dir, f)), reverse=True)[0])
            if not filepath or not os.path.exists(filepath):
                return jsonify({"success": False, "error": "未找到文件"})
        
        sheet_name = request.form.get('sheet_name', '') if 'file' in request.files else data.get('sheet', '')
        product_type = request.form.get('product_type', 'zhengsquare') if 'file' in request.files else data.get('product_type', 'zhengsquare')
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # 找对应sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return jsonify({"success": False, "error": f"未找到工作表: {sheet_name}，可用的: {wb.sheetnames}"})
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        mode = (request.form.get("import_mode") or request.form.get("mode") or "append").strip().lower()
        inv_data = load_inventory()
        tab = "finished"
        if mode == "overwrite":
            inv_data[tab] = []
        added = updated = 0
        
        # 自动探测表头行（找包含"外尺寸"的行）
        header_row = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            vals = [str(v or '') for v in row[:8]]
            if '外尺寸' in vals or '外尺寸规格' in vals:
                header_row = r_idx
                break
        
        if header_row is None:
            return jsonify({"success": False, "error": "未找到表头行（需包含外尺寸）"})
        
        # 确定列索引
        header = [str(v or '') for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col_map = {}
        for i, h in enumerate(header):
            h_clean = h.strip()
            if '外尺寸' in h_clean:
                col_map['spec'] = i
            elif '材质' in h_clean:
                col_map['material'] = i
            elif '货位' in h_clean or '货 位' in h_clean:
                col_map['location'] = i
            elif '上月' in h_clean or '上月库存' in h_clean:
                col_map['last_month'] = i
            elif '入库' in h_clean and '数量' in h_clean:
                col_map['in_qty'] = i
            elif '出库' in h_clean and '数量' in h_clean:
                col_map['out_qty'] = i
            elif '剩余' in h_clean or '库存' in h_clean:
                col_map['stock'] = i
            elif '序号' in h_clean:
                col_map['seq'] = i
        if 'spec' not in col_map:
            return jsonify({"success": False, "error": f"未找到'外尺寸'列，表头: {header}"})
        
        # 读取数据行：从header_row之后找到第一个有规格的行
        data_start = header_row + 1
        for r_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            if col_map['spec'] < len(row):
                spec_val = row[col_map['spec']]
                spec_str = str(spec_val).strip() if spec_val else ''
                if spec_val and spec_str and not spec_str.startswith('=') and not spec_str.startswith('#'):
                    try:
                        dims = spec_str.replace('×','*').replace('x','*').split('*')
                        if len(dims) >= 3:
                            data_start = r_idx
                            break
                    except:
                        pass
        
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            spec = row[col_map['spec']] if col_map['spec'] < len(row) else None
            if not spec or not str(spec).strip():
                continue
            spec = str(spec).strip()
            # 跳过公式行（=ROW开头）
            if spec.startswith('=') or spec.startswith('#'):
                continue
            
            # 解析规格：如 "10.5×7.5×4.5" 或 "6×6×2"
            dims = spec.replace('×', '*').replace('x', '*').replace('X', '*').split('*')
            length = width = height = None
            try:
                if len(dims) >= 3:
                    length = float(dims[0].strip())
                    width = float(dims[1].strip())
                    height = float(dims[2].strip())
            except:
                pass
            
            # 材质（可能有多个材质行共享同一个规格）
            material = str(row[col_map['material']]).strip() if col_map.get('material') and col_map['material'] < len(row) and row[col_map['material']] else ''
            location = str(row[col_map['location']]).strip() if col_map.get('location') and col_map['location'] < len(row) and row[col_map['location']] else ''
            
            # 数值
            last_month = 0
            stock_qty = 0
            try:
                v = row[col_map['last_month']] if col_map.get('last_month') and col_map['last_month'] < len(row) else 0
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    last_month = int(v)
            except: pass
            try:
                v = row[col_map['stock']] if col_map.get('stock') and col_map['stock'] < len(row) else 0
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    stock_qty = int(v)
            except: pass
            
            # 构造商品名
            name_parts = [spec]
            if material and material not in ('None', ''):
                name_parts.append(material)
            name = ' '.join(name_parts)
            
            item_id = f"inv_{int(time.time())}_{added}_{len(inv_data[tab])}"
            item = {
                "id": item_id,
                "name": name,
                "spec": spec,
                "product_type": product_type,
                "material": material,
                "location": location,
                "qty": stock_qty,
                "stock": stock_qty,
                "last_month_qty": last_month,
                "length": length,
                "width": width,
                "height": height,
                "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                "updated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            idx = None
            for i, d in enumerate(inv_data[tab]):
                if d.get("name") == name and d.get("product_type") == product_type:
                    idx = i
                    break
            if mode == "append" and idx is not None:
                continue
            if idx is not None and mode == "upsert":
                old = inv_data[tab][idx]
                item["id"] = old.get("id") or item_id
                item["created_at"] = old.get("created_at") or item["created_at"]
                inv_data[tab][idx] = {**old, **item}
                updated += 1
            else:
                inv_data[tab].append(item)
                added += 1
        
        save_inventory(inv_data)
        _prod_dash_cache.invalidate_dashboard_cache()
        return jsonify({
            "success": True,
            "message": f"导入完成：新增 {added} 条，更新 {updated} 条（模式 {mode}）",
            "added": added,
            "updated": updated,
            "mode": mode,
        })
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()})

# ==================== 权限列表 ====================
@app.route('/api/permissions')
def permissions():
    return jsonify({
        "roles": [
            {"name": "超级管理员", "desc": "所有权限", "count": 2},
            {"name": "主管", "desc": "生产管理、员工查看", "count": 5},
            {"name": "客服", "desc": "订单查看、生产进度", "count": 10},
            {"name": "员工", "desc": "扫码报工、查看任务", "count": 20},
        ]
    })


# ==================== 静态文件 ====================
@app.route('/login_simple')
def login_simple():
    return send_from_directory('.', 'simple_login.html')

# ==================== 原材料库存 API（MySQL raw_materials） ====================
RAW_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'raw_data.json')


def load_raw_data():
    """从 MySQL 加载纸板规格（与 app.py 客服端一致）。"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute(
            "SELECT id, date, name, supplier, paper_width, paper_length, qty, remark, created_at, updated_at "
            "FROM raw_materials ORDER BY id DESC"
        )
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append(
                {
                    "id": r["id"],
                    "date": r["date"] or "",
                    "name": r["name"] or "",
                    "supplier": r["supplier"] or "",
                    "paper_width": str(r["paper_width"] or ""),
                    "paper_length": str(r["paper_length"] or ""),
                    "qty": r["qty"] or 0,
                    "remark": r["remark"] or "",
                    "created_at": str(r["created_at"] or ""),
                    "updated_at": str(r["updated_at"] or ""),
                }
            )
        cur.close()
        db.close()
        return result
    except Exception as e:
        print(f"[MySQL load_raw_data] 错误: {e}")
        return []


def save_raw_data(data):
    """批量写回 MySQL（管理端导入/维护用）。"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE raw_materials")
        for item in data:
            cur.execute(
                "INSERT INTO raw_materials (id, date, name, supplier, paper_width, paper_length, qty, remark, created_at, updated_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    item.get("id", ""),
                    item.get("date", ""),
                    item.get("name", ""),
                    item.get("supplier", ""),
                    item.get("paper_width", ""),
                    item.get("paper_length", ""),
                    item.get("qty", 0),
                    item.get("remark", ""),
                    item.get("created_at", ""),
                    item.get("updated_at", ""),
                ),
            )
        db.commit()
        cur.close()
        db.close()
        import material_calc as _mc

        _mc._raw_cache = {"ts": 0, "rows": []}
        return True
    except Exception as e:
        print(f"[MySQL save_raw_data] 错误: {e}")
        return False

@app.route('/api/raw/template', methods=['GET'])
def template_raw():
    return _send_xlsx_template(
        ["日期", "材料名称", "供应商", "门幅(cm)", "长度(m)", "数量", "备注"],
        [["2026-05-19", "三层牛卡", "XX纸业", 159, 2000, 5, ""]],
        "原材料导入模板.xlsx",
    )


@app.route('/api/raw/import', methods=['POST'])
def import_raw():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "请上传 Excel 文件"})
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "未选择文件"})
    try:
        import openpyxl
        import tempfile

        mode = (request.form.get("import_mode") or "append").strip().lower()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        f.save(tmp.name)
        tmp.close()
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active
        header_row = 1
        headers = [
            str(v or "").strip()
            for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        col = {}
        for i, h in enumerate(headers):
            if "日期" in h:
                col["date"] = i
            elif "材料" in h or "名称" in h:
                col["name"] = i
            elif "供应商" in h:
                col["supplier"] = i
            elif "门幅" in h or "宽" in h:
                col["paper_width"] = i
            elif "长度" in h or "长" in h:
                col["paper_length"] = i
            elif "数量" in h:
                col["qty"] = i
            elif "备注" in h:
                col["remark"] = i
        if "name" not in col:
            return jsonify({"success": False, "error": "表头需包含「材料名称」列"})
        data = load_raw_data()
        if mode == "overwrite":
            data = []
        by_id_name = {d.get("name", ""): i for i, d in enumerate(data)}
        added = updated = 0
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        def cell(row, key):
            i = col.get(key)
            if i is None or i >= len(row):
                return ""
            return row[i]

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(cell(row, "name") or "").strip()
            if not name:
                continue
            entry = {
                "date": str(cell(row, "date") or "").strip(),
                "name": name,
                "supplier": str(cell(row, "supplier") or "").strip(),
                "paper_width": cell(row, "paper_width"),
                "paper_length": cell(row, "paper_length"),
                "qty": int(cell(row, "qty") or 0),
                "remark": str(cell(row, "remark") or "").strip(),
                "updated_at": now,
            }
            if mode == "upsert" and name in by_id_name:
                i = by_id_name[name]
                old = data[i]
                entry["id"] = old.get("id")
                entry["created_at"] = old.get("created_at") or now
                data[i] = {**old, **entry}
                updated += 1
            elif mode == "append" and name in by_id_name:
                continue
            else:
                entry["id"] = str(int(time.time() * 1000)) + str(len(data))
                entry["created_at"] = now
                data.append(entry)
                by_id_name[name] = len(data) - 1
                added += 1
        save_raw_data(data)
        os.unlink(tmp.name)
        return jsonify({
            "success": True,
            "message": f"原材料导入：新增 {added}，更新 {updated}（{mode}）",
            "added": added,
            "updated": updated,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/api/raw', methods=['GET'])
def get_raw():
    data = load_raw_data()
    search = request.args.get('search', '').strip()
    date = request.args.get('date', '').strip()
    name = request.args.get('name', '').strip()
    supplier = request.args.get('supplier', '').strip()
    paper_width = request.args.get('paper_width', '').strip()
    paper_length = request.args.get('paper_length', '').strip()
    if date:
        data = [d for d in data if date in d.get('date', '')]
    if name:
        data = [d for d in data if name.lower() in d.get('name', '').lower()]
    if supplier:
        data = [d for d in data if supplier.lower() in d.get('supplier', '').lower()]
    if paper_width:
        data = [d for d in data if str(d.get('paper_width', '')) == paper_width]
    if paper_length:
        data = [d for d in data if str(d.get('paper_length', '')) == paper_length]
    if search:
        data = [d for d in data if search.lower() in d.get('name', '').lower() or search.lower() in d.get('remark', '').lower()]
    data.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return jsonify({"success": True, "data": data, "total": len(data)})

@app.route('/api/raw', methods=['POST'])
def add_raw():
    body = request.get_json()
    if not body or not body.get('name', '').strip():
        return jsonify({"success": False, "error": "材料名称不能为空"})
    data = load_raw_data()
    new_id = str(int(time.time() * 1000)) + str(len(data))
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    entry = {
        "id": new_id,
        "date": body.get('date', '').strip(),
        "name": body['name'].strip(),
        "supplier": body.get('supplier', '').strip(),
        "paper_width": body.get('paper_width', ''),
        "paper_length": body.get('paper_length', ''),
        "qty": int(body.get('qty', 0)),
        "remark": body.get('remark', '').strip(),
        "created_at": now,
        "updated_at": now
    }
    data.append(entry)
    save_raw_data(data)
    return jsonify({"success": True, "data": entry, "message": "原材料保存成功"})

@app.route('/api/raw/<raw_id>', methods=['PUT'])
def update_raw(raw_id):
    body = request.get_json()
    data = load_raw_data()
    for item in data:
        if item['id'] == raw_id:
            if body.get('date', '').strip(): item['date'] = body['date'].strip()
            if body.get('name', '').strip(): item['name'] = body['name'].strip()
            if 'supplier' in body: item['supplier'] = body['supplier'].strip()
            if 'paper_width' in body: item['paper_width'] = str(body['paper_width'])
            if 'paper_length' in body: item['paper_length'] = str(body['paper_length'])
            if 'qty' in body: item['qty'] = int(body['qty'])
            if 'remark' in body: item['remark'] = body['remark'].strip()
            item['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            save_raw_data(data)
            return jsonify({"success": True, "data": item, "message": "更新成功"})
    return jsonify({"success": False, "error": "未找到该材料"})

@app.route('/api/raw/stock', methods=['POST'])
def raw_stock():
    body = request.get_json()
    rid = body.get('id', '')
    op = body.get('op', 'in')
    qty = int(body.get('qty', 0))
    if qty <= 0:
        return jsonify({"success": False, "error": "数量必须大于0"})
    data = load_raw_data()
    for item in data:
        if item['id'] == rid:
            if op == 'in':
                item['qty'] = (item.get('qty', 0) or 0) + qty
            else:
                item['qty'] = max(0, (item.get('qty', 0) or 0) - qty)
            item['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            save_raw_data(data)
            return jsonify({"success": True, "data": item, "message": f"{'入库' if op=='in' else '出库'}成功"})
    return jsonify({"success": False, "error": "未找到该材料"})


def _parse_raw_barcode(code: str) -> dict:
    code = (code or "").strip()
    if not code:
        return {}
    if code.startswith("{") and code.endswith("}"):
        try:
            return json.loads(code)
        except json.JSONDecodeError:
            pass
    if "|" in code:
        parts = [p.strip() for p in code.split("|") if p.strip()]
        if parts and parts[0].upper() == "RAW":
            parts = parts[1:]
        keys = ("name", "supplier", "paper_width", "paper_length")
        out = {}
        for i, p in enumerate(parts):
            if i < len(keys):
                out[keys[i]] = p
        return out
    return {"name": code}


@app.route('/api/raw/scan_register', methods=['POST'])
def raw_scan_register():
    """扫码登记原材料入库（每张+1，可指定数量）。"""
    if not resolve_login_user():
        return jsonify({"success": False, "error": "未登录"}), 401
    body = request.get_json() or {}
    code = (body.get("code") or body.get("barcode") or "").strip()
    qty = int(body.get("qty") or 1)
    if not code:
        return jsonify({"success": False, "error": "请扫描或输入条码"})
    if qty <= 0:
        return jsonify({"success": False, "error": "数量须大于0"})

    data = load_raw_data()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    today = datetime.date.today().isoformat()

    for item in data:
        if item.get("id") == code:
            item["qty"] = (item.get("qty", 0) or 0) + qty
            item["updated_at"] = now
            save_raw_data(data)
            return jsonify({
                "success": True,
                "data": item,
                "message": f"「{item.get('name')}」扫码入库 +{qty} 张",
            })

    parsed = _parse_raw_barcode(code)
    name = (parsed.get("name") or code).strip()
    supplier = (parsed.get("supplier") or "").strip()
    pw = str(parsed.get("paper_width") or "")
    pl = str(parsed.get("paper_length") or "")

    for item in data:
        if (
            item.get("name") == name
            and (item.get("supplier") or "") == supplier
            and str(item.get("paper_width") or "") == pw
            and str(item.get("paper_length") or "") == pl
        ):
            item["qty"] = (item.get("qty", 0) or 0) + qty
            if not item.get("date"):
                item["date"] = today
            item["updated_at"] = now
            save_raw_data(data)
            return jsonify({
                "success": True,
                "data": item,
                "message": f"「{name}」扫码入库 +{qty} 张（累计 {item['qty']}）",
            })

    new_id = str(int(time.time() * 1000)) + str(len(data))
    entry = {
        "id": new_id,
        "date": today,
        "name": name,
        "supplier": supplier,
        "paper_width": pw,
        "paper_length": pl,
        "qty": qty,
        "remark": "扫码登记",
        "created_at": now,
        "updated_at": now,
    }
    data.append(entry)
    save_raw_data(data)
    return jsonify({
        "success": True,
        "data": entry,
        "message": f"已登记新材料「{name}」+{qty} 张",
    })


from io import BytesIO
import qrcode
import barcode
from barcode.writer import ImageWriter
import base64

# ===== 二维码 & 条码生成 =====
@app.route('/api/barcode/<order_id>')
def generate_barcode(order_id):
    """生成订单条码（Code128，base64图片）"""
    try:
        code = barcode.get('code128', order_id, writer=ImageWriter())
        buf = BytesIO()
        code.write(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()
        return jsonify({'success': True, 'barcode_base64': b64})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/qrcode/<order_id>')
def generate_qrcode(order_id):
    """生成订单二维码（base64图片）"""
    # 二维码内容：扫码后打开报工页面
    qr_data = f"{request.host_url}scan?order={order_id}"
    img = qrcode.make(qr_data, box_size=8)
    buf = BytesIO()
    img.save(buf, format='PNG')
    b64 = base64.b64encode(buf.getvalue()).decode()
    return jsonify({'success': True, 'qr_base64': b64, 'url': qr_data})

# ===== 扫码报工手机专用页 =====
@app.route('/scan')
def scan_page():
    """手机扫码后打开的报工页面"""
    order_id = request.args.get('order', '')
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>三羊包装 - 扫码报工</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,sans-serif; background:#f5f6fa; min-height:100vh; display:flex; flex-direction:column; align-items:center; padding:20px; }}
.card {{ background:#fff; border-radius:16px; padding:24px; width:100%; max-width:420px; box-shadow:0 2px 12px rgba(0,0,0,0.08); text-align:center; }}
.order-id {{ font-size:24px; font-weight:800; color:#1a1a2e; margin:12px 0; word-break:break-all; }}
.step-btn {{ display:block; width:100%; padding:14px; margin:8px 0; border:none; border-radius:10px; font-size:15px; font-weight:600; cursor:pointer; transition:all 0.2s; }}
.step-btn:active {{ transform:scale(0.97); }}
.btn-primary {{ background:#1677ff; color:#fff; }}
.btn-success {{ background:#52c41a; color:#fff; }}
.btn-gray {{ background:#f0f0f0; color:#666; }}
.info {{ font-size:13px; color:#999; margin:8px 0; }}
.result {{ margin-top:12px; padding:12px; border-radius:8px; font-size:14px; display:none; }}
.result.ok {{ display:block; background:#f6ffed; border:1px solid #b7eb8f; color:#52c41a; }}
.result.err {{ display:block; background:#fff2f0; border:1px solid #ffccc7; color:#e94560; }}
.back-btn {{ display:inline-block; margin-top:16px; padding:8px 24px; background:#f0f0f0; border:none; border-radius:8px; font-size:13px; color:#666; cursor:pointer; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
    <div style="font-size:40px;margin-bottom:8px;">📦</div>
    <h3 style="font-size:16px;color:#666;font-weight:400;">扫码报工</h3>
    <div class="order-id">{order_id}</div>
    <div id="stepList"></div>
    <div id="result" class="result"></div>
    <div style="font-size:11px;color:#bbb;margin-top:12px;">请选择要报工的工序</div>
</div>
<a class="back-btn" href="javascript:history.back()">← 返回</a>
<script>
const ORDER_ID = "{order_id}";
const WORKER = prompt("请输入你的姓名：", "");
if (!WORKER) {{ document.body.innerHTML = '<div style="text-align:center;padding:40px;color:#e94560;">❌ 需要姓名才能报工</div>'; }}

// 获取订单工序
fetch('/api/production_orders').then(r=>r.json()).then(data => {{
    const orders = data.orders || [];
    const order = orders.find(o => o.inner_id === ORDER_ID);
    const steps = order ? (order.flow || []) : [];
    const container = document.getElementById('stepList');
    if (steps.length === 0) {{
        container.innerHTML = '<div class="info">该订单暂无工序数据</div>';
        return;
    }}
    steps.forEach((s, i) => {{
        const done = s.done;
        const btn = document.createElement('button');
        btn.className = 'step-btn ' + (done ? 'btn-gray' : 'btn-primary');
        btn.textContent = (done ? '✅ ' : '') + s.step + (done ? ' (已完成)' : '');
        btn.disabled = done;
        if (!done) {{
            btn.onclick = () => submitReport(i, s.step);
        }}
        container.appendChild(btn);
    }});
}}).catch(e => {{
    document.getElementById('stepList').innerHTML = '<div class="info">加载失败: ' + e.message + '</div>';
}});

async function submitReport(stepIndex, stepName) {{
    const resultEl = document.getElementById('result');
    resultEl.className = 'result';
    resultEl.textContent = '提交中...';
    resultEl.style.display = 'block';
    try {{
        const res = await fetch('/api/scan_report', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{order_id: ORDER_ID, step: stepName, worker: WORKER}})
        }});
        const data = await res.json();
        if (data.success) {{
            resultEl.className = 'result ok';
            resultEl.textContent = '✅ ' + (data.message || '报工成功');
            // 刷新页面
            setTimeout(() => location.reload(), 1500);
        }} else {{
            resultEl.className = 'result err';
            resultEl.textContent = '❌ ' + (data.error || '报工失败');
        }}
    }} catch(e) {{
        resultEl.className = 'result err';
        resultEl.textContent = '❌ 网络错误: ' + e.message;
    }}
}}
</script>
</body>
</html>'''
    return html


# ==================== 1688开放平台集成 ====================

ALIBABA_CONFIG = {
    'app_key': 1452593,
    'app_secret': 'f4UOJ5NO1X',
    'access_token': 'bdb56a38-b85b-410d-acb2-74a04fc20496',
    'server': 'gw.open.1688.com'
}

def _1688_sign(url_path, params, secret):
    """1688签名：HMAC-SHA1"""
    param_list = sorted([str(k) + str(v) for k, v in params.items()])
    msg = url_path.encode('utf-8')
    for p in param_list:
        msg += p.encode('utf-8')
    sha = hmac.new(secret.encode('utf-8'), msg, hashlib.sha1)
    return sha.hexdigest().upper()

def _1688_api(api_uri, biz_params=None):
    """调用1688开放平台API"""
    cfg = ALIBABA_CONFIG
    url_path = f'param2/{api_uri}/{cfg["app_key"]}'
    params = {'access_token': cfg['access_token']}
    if biz_params:
        params.update(biz_params)
    params['_aop_signature'] = _1688_sign(url_path, params, cfg['app_secret'])
    url = f'https://{cfg["server"]}/openapi/{url_path}'
    form_data = urllib.parse.urlencode(params)
    try:
        req = urllib.request.Request(url, data=form_data.encode('utf-8'), method='POST')
        resp = urllib.request.urlopen(req, timeout=30)
        return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print(f'[1688API] {api_uri} 调用失败: {e}')
        return {}


def _1688_format_order(o):
    """格式化1688订单为统一格式"""
    info = o.get('baseInfo', {})
    items = o.get('productItems', [])
    product_list = []
    for item in items:
        product_list.append({
            'name': item.get('name', ''),
            'sku': item.get('productCargoNumber', ''),
            'qty': item.get('quantity', 0),
            'price': item.get('price', 0),
            'spec': item.get('skuInfos', [{'value': ''}])[0].get('value', ''),
            'skuId': item.get('skuID', ''),
            'productId': item.get('productID', '')
        })
    so_id = str(info.get('idOfStr', info.get('id', '')))
    return {
        'so_id': so_id,
        'platform': '1688',
        'order_status': info.get('status', ''),
        'status_label': '待发货' if info.get('status') == 'waitsellersend' else info.get('status', ''),
        'created': info.get('createTime', '')[:8] if info.get('createTime') else '',
        'pay_time': info.get('payTime', '')[:8] if info.get('payTime') else '',
        'total_amount': info.get('totalAmount', 0),
        'shipping_fee': info.get('shippingFee', 0) / 100,
        'discount': info.get('discount', 0) / 100,
        'receiver_name': info.get('receiverName', info.get('buyerLoginId', '')),
        'receiver_mobile': info.get('receiverMobile', ''),
        'receiver_address': info.get('receiverAddress', ''),
        'receiver_phone': info.get('receiverPhone', ''),
        'shop_name': info.get('buyerLoginId', '1688'),
        'items': product_list,
        'buyer_login_id': info.get('buyerLoginId', ''),
        'alipay_trade_id': info.get('alipayTradeId', ''),
        'trade_type': info.get('tradeType', ''),
        'buyer_memo': '',
        'seller_memo': get_order_memo(so_id),
    }

# ==================== 订单备注本地存储 ====================

_ORDER_MEMO = {}

def get_order_memo(so_id):
    return _ORDER_MEMO.get(so_id, '')

def set_order_memo(so_id, memo):
    _ORDER_MEMO[so_id] = memo

@app.route('/api/order/1688-memo', methods=['POST'])
def api_order_1688_memo():
    """修改卖家备注并同步回1688"""
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    data = request.get_json() or {}
    so_id = data.get('so_id', '')
    memo = data.get('memo', '')
    if not so_id:
        return jsonify({'success': False, 'error': '缺少订单ID'})
    # 先本地存储
    set_order_memo(so_id, memo)
    # 同步到1688
    try:
        # 调用1688 memoAdd API
        result = _1688_api('1/com.alibaba.trade/alibaba.order.memoAdd', {
            'orderId': so_id,
            'memo': memo,
            'remarkIcon': '0'
        })
        if result.get('errorCode') or result.get('errorMessage'):
            return jsonify({'success': False, 'error': f'1688同步失败: {result.get("errorMessage", result.get("errorCode", "未知错误"))}', 'local_saved': True})
    except Exception as e:
        return jsonify({'success': True, 'error': f'1688同步失败: {str(e)}', 'local_saved': True})
    return jsonify({'success': True, 'message': '备注已保存并同步到1688'})

@app.route('/api/realtime/orders', methods=['GET'])
def api_realtime_orders():
    """实时订单：MySQL 秒回 + 后台增量同步。"""
    import order_sync as _osync

    payload = _osync.read_realtime_cache_payload(
        shop_config=load_shop_config(),
        memo_getter=get_order_memo,
        trigger_background_sync=True,
    )
    n = payload.get("total", 0)
    print(f"[实时订单] MySQL {payload.get('cache_status')}: {n} 条")
    return jsonify(payload)


# ==================== 店铺客服配置 API ====================

SHOP_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'shop_config.json')

DEFAULT_SHOP_CONFIG = [
    {"id": "shop_1688_1", "platform": "1688", "shop_name": "友尚", "customer_service": "张慧平,戴志美", "sort_order": 1},
    {"id": "shop_1688_2", "platform": "1688", "shop_name": "亚润", "customer_service": "陈贝贝", "sort_order": 2},
    {"id": "shop_1688_3", "platform": "1688", "shop_name": "三羊", "customer_service": "罗怡", "sort_order": 3},
    {"id": "shop_1688_4", "platform": "1688", "shop_name": "正方形", "customer_service": "周井梅", "sort_order": 4},
    {"id": "shop_1688_5", "platform": "1688", "shop_name": "大鱼", "customer_service": "周井梅", "sort_order": 5},
    {"id": "shop_1688_6", "platform": "1688", "shop_name": "阿里新鑫星", "customer_service": "戴志美", "sort_order": 6},
    {"id": "shop_tmall_1", "platform": "tmall", "shop_name": "飞机盒彩色专卖店", "customer_service": "廖思美", "sort_order": 7},
    {"id": "shop_tmall_2", "platform": "tmall", "shop_name": "飞机盒正方形专卖店", "customer_service": "廖思美", "sort_order": 8},
    {"id": "shop_tmall_3", "platform": "tmall", "shop_name": "飞机盒止合专卖店", "customer_service": "石梅清", "sort_order": 9},
    {"id": "shop_tmall_4", "platform": "tmall", "shop_name": "飞机盒扣底盒专卖店", "customer_service": "石梅清", "sort_order": 10},
    {"id": "shop_tmall_5", "platform": "tmall", "shop_name": "飞机盒小批量专卖店", "customer_service": "张文杰", "sort_order": 11},
    {"id": "shop_taobao_1", "platform": "taobao", "shop_name": "飞机盒品牌店", "customer_service": "张文杰", "sort_order": 12},
    {"id": "shop_taobao_2", "platform": "taobao", "shop_name": "俊鑫纸品厂", "customer_service": "石梅清", "sort_order": 13},
    {"id": "shop_taobao_3", "platform": "taobao", "shop_name": "当下家包装", "customer_service": "张文杰", "sort_order": 14}
]

def load_shop_config():
    """MySQL + shop_config.json（与 3003 一致）；无数据时回退本地 JSON/默认。"""
    try:
        import admin_shared_store as asc

        rows = asc.load_shop_config()
        if rows:
            return rows
    except Exception as e:
        print(f"[load_shop_config] admin_shared_store: {e}")
    if not os.path.exists(SHOP_CONFIG_FILE):
        return list(DEFAULT_SHOP_CONFIG)
    try:
        with open(SHOP_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return list(DEFAULT_SHOP_CONFIG)

def save_shop_config(config):
    """Admin 保存店铺配置 → shop_config.json。"""
    with open(SHOP_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

@app.route('/api/shop-config', methods=['GET'])
def api_get_shop_config():
    config = load_shop_config()
    return jsonify({'success': True, 'config': config})

@app.route('/api/shop-config', methods=['POST'])
def api_save_shop_config():
    try:
        data = request.get_json(force=True) or {}
        config = data.get('config', [])
        save_shop_config(config)
        return jsonify({'success': True, 'message': '保存成功'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop-config/reset', methods=['POST'])
def api_reset_shop_config():
    save_shop_config(DEFAULT_SHOP_CONFIG)
    return jsonify({'success': True, 'config': list(DEFAULT_SHOP_CONFIG)})


# ========== 后台定时同步订单（快麦 ERP）==========
ORDERS_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'orders_cache.json')


def _km_sync_orders_to_cache(days_back=14):
    import order_sync as _osync
    r = _osync.sync_orders_to_cache(
        ORDERS_CACHE_FILE,
        days_back=days_back,
        memo_getter=get_order_memo,
        include_1688_direct=False,
    )
    _prod_dash_cache.invalidate_dashboard_cache()
    return r


def _run_auto_material_calc():
    import material_calc as _mc

    try:
        orders = ph.load_cache_orders(_orders_cache_path())
        rep = _mc.auto_calc_all_orders(
            orders,
            quote_data=load_quote_data() or {},
            load_raw_fn=load_raw_data,
            load_dimoldb_fn=load_dimoldb,
            material_mapping=_quote_material_mapping_for_spec(),
            mark_flow=True,
            db_config=DB_CONFIG,
            processes=_permission_data.get("processes", []),
            internal_order_id_fn=ph.internal_order_id,
            infer_order_type_fn=ph.infer_order_type,
            get_or_create_flow_steps_fn=ph.get_or_create_flow_steps,
            save_flow_row_fn=ph.save_flow_row,
        )
        print(
            f"[自动算料] 完成 {rep.get('lines_done', 0)} 行, "
            f"失败 {rep.get('lines_failed', 0)} 行"
        )
    except Exception as e:
        import traceback

        print(f"[自动算料] 失败: {e}")
        traceback.print_exc()
    _prod_dash_cache.invalidate_dashboard_cache()
    try:
        _prod_dash_cache.get_dashboard_cache(_rebuild_prod_dashboard_cache, force=True)
    except Exception as e:
        print(f"[打单缓存] 同步后重建失败: {e}")


def _on_background_sync_done(_report):
    _th.Thread(target=_run_auto_material_calc, daemon=True).start()

@app.route('/api/sync/force', methods=['POST'])
def api_sync_force():
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    import order_sync as _osync
    ok, msg = _osync.start_force_sync_async(
        ORDERS_CACHE_FILE,
        days_back=30,
        memo_getter=get_order_memo,
        include_1688_direct=False,
    )
    if not ok:
        return jsonify({'success': False, 'error': msg}), 409
    return jsonify({'success': True, 'async': True, 'message': msg})


@app.route('/api/sync/status', methods=['GET'])
def api_sync_status():
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    import order_sync as _osync
    import order_visit_sync as _ovs

    st = _osync.force_sync_status()
    last = st.get('last') or {}
    visit = _ovs.visit_sync_status()
    return jsonify({
        'success': True,
        'running': st.get('running') or visit.get('running'),
        'error': st.get('error') or visit.get('last_error'),
        'phase': st.get('phase'),
        'detail': st.get('detail'),
        'last': last,
        'count': last.get('pending_count') if isinstance(last, dict) else None,
        'cache_updated_ago_sec': visit.get('updated_ago_sec'),
        'sync_interval_sec': int(os.getenv('ORDER_SYNC_INTERVAL_SEC', '45')),
        'stale_force_sec': visit.get('stale_force_sec'),
    })


@app.route('/api/km/probe', methods=['GET'])
def api_km_probe():
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    import km_api as _km
    return jsonify({'success': True, 'data': _km.km_probe()})


import threading as _th
import production_dashboard_cache as _prod_dash_cache


def _rebuild_prod_dashboard_cache():
    try:
        import order_sync as _osync

        norm_shop = _osync.normalize_shop_display
    except ImportError:

        def norm_shop(s):
            return (s or "").strip() or "未知店铺"

    return _prod_dash_cache.rebuild_dashboard_cache(
        orders_cache_file=ORDERS_CACHE_FILE,
        permission_data=_permission_data,
        material_mapping=_quote_material_mapping_for_spec(),
        load_inventory_fn=load_inventory,
        load_dimoldb_fn=load_dimoldb,
        load_cache_orders_fn=lambda: ph.load_cache_orders(_orders_cache_path()),
        get_db_fn=get_db,
        infer_order_type_fn=ph.infer_order_type,
        internal_order_id_fn=ph.internal_order_id,
        parse_flow_steps_fn=ph.parse_flow_steps,
        template_steps_fn=ph.template_steps_for_order,
        normalize_shop_fn=norm_shop,
    )


def _warm_prod_dashboard_cache():
    first = True
    while True:
        try:
            _prod_dash_cache.get_dashboard_cache(
                _rebuild_prod_dashboard_cache, force=first
            )
            if first:
                first = False
                print("[打单缓存] 首次构建完成，启动后台算料…")
                _th.Thread(target=_run_auto_material_calc, daemon=True).start()
        except Exception as e:
            print(f"[打单缓存] 刷新失败: {e}")
        time.sleep(int(os.getenv("PROD_DASH_WARM_INTERVAL_SEC", "120")))


_dash_warm_thread = _th.Thread(target=_warm_prod_dashboard_cache, daemon=True)
_dash_warm_thread.start()


def _warm_dimoldb_cache():
    while True:
        try:
            load_dimoldb(force=True)
            print("[刀模缓存] 内存索引已刷新")
        except Exception as e:
            print(f"[刀模缓存] 刷新失败: {e}")
        time.sleep(int(os.getenv("DIMOLDB_CACHE_TTL_SEC", "120")))


_th.Thread(target=_warm_dimoldb_cache, daemon=True, name="dimoldb-cache-warm").start()

import order_sync_scheduler as _order_sched

_order_sched.start_background_order_sync(
    ORDERS_CACHE_FILE,
    memo_getter=get_order_memo,
    include_1688_direct=False,
    full_days_back=30,
    incremental_days_back=int(os.getenv("ORDER_SYNC_INCREMENTAL_DAYS", "14")),
    interval_sec=int(os.getenv("ORDER_SYNC_INTERVAL_SEC", "45")),
    on_after_sync=_on_background_sync_done,
)


def _send_xlsx_template(headers: list, sample_rows: list, filename: str):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据"
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
    for ri, row in enumerate(sample_rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
    import tempfile
    import flask

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return flask.send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ==================== 生产公告 ====================
try:
    import production_announcements as prod_ann

    prod_ann.ensure_tables(get_db)
except Exception as _ann_ex:
    print(f"[公告] 表初始化: {_ann_ex}")


def _require_manager_json():
    un = resolve_login_user()
    if not un or un not in USERS:
        return jsonify({"success": False, "error": "未登录"}), 401
    if USERS[un].get("role") not in ("超级管理员", "管理"):
        return jsonify({"success": False, "error": "仅管理员可编辑公告"}), 403
    return None


@app.route("/api/production/announcements")
def production_announcements_list():
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "error": "未登录"}), 401
    items = prod_ann.list_active_for_user(get_db, un)
    unread = prod_ann.unread_count(get_db, un)
    role = USERS.get(un, {}).get("role", "")
    return jsonify(
        {
            "success": True,
            "announcements": items,
            "unread_count": unread,
            "can_edit": role in ("超级管理员", "管理"),
        }
    )


@app.route("/api/production/announcements", methods=["POST"])
def production_announcements_save():
    err = _require_manager_json()
    if err:
        return err
    body = request.get_json() or {}
    title = (body.get("title") or "").strip()
    content = (body.get("content") or "").strip()
    ann_id = body.get("id")
    if not title or not content:
        return jsonify({"success": False, "error": "标题和内容不能为空"})
    un = resolve_login_user()
    row = prod_ann.save_announcement(
        get_db, title=title, content=content, created_by=un or "", ann_id=ann_id
    )
    return jsonify({"success": True, "announcement": row})


@app.route("/api/production/announcements/<int:ann_id>/read", methods=["POST"])
def production_announcements_read(ann_id: int):
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "error": "未登录"}), 401
    prod_ann.mark_read(get_db, un, ann_id)
    return jsonify({"success": True, "unread_count": prod_ann.unread_count(get_db, un)})


@app.route("/api/production/announcements/read-all", methods=["POST"])
def production_announcements_read_all():
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "error": "未登录"}), 401
    prod_ann.mark_all_read(get_db, un)
    return jsonify({"success": True, "unread_count": 0})


# ==================== 算料 / 纸板匹配 ====================
import material_calc as mcalc


@app.route("/api/production/match-paper", methods=["POST"])
def production_match_paper():
    """纸板规格匹配（算料基础）。"""
    if not resolve_login_user():
        return jsonify({"success": False, "error": "未登录"}), 401
    body = request.get_json() or {}
    paper_l_inch = float(
        body.get("paper_l_inch") or body.get("spread_l_inch") or body.get("spread_l") or 0
    )
    paper_w_inch = float(
        body.get("paper_w_inch") or body.get("spread_w_inch") or body.get("spread_w") or 0
    )
    if not paper_l_inch and body.get("spread_l_cm"):
        paper_l_inch = float(body["spread_l_cm"]) / 2.54
    if not paper_w_inch and body.get("spread_w_cm"):
        paper_w_inch = float(body["spread_w_cm"]) / 2.54
    material = (body.get("material") or body.get("material_text") or "").strip()
    product_type = (body.get("product_type") or "").strip()
    attrs_hint = (body.get("attrs") or "").strip()
    rows = mcalc.load_raw_rows(load_raw_data)
    result = mcalc.match_paper(
        paper_l_inch,
        paper_w_inch,
        material,
        rows,
        product_type=product_type,
        attrs=attrs_hint,
    )
    return jsonify(result)


def _calc_material_for_request(so_id: str, line_index: int, *, mark_flow: bool = True) -> dict:
    import production_spec as pspec

    orders = ph.load_cache_orders(_orders_cache_path())
    order = ph.find_order_in_cache(orders, so_id)
    if not order:
        return {"success": False, "error": "未找到订单"}

    result = mcalc.calc_order_line(
        order,
        line_index,
        quote_data=load_quote_data() or {},
        raw_rows=mcalc.load_raw_rows(load_raw_data),
        dimoldb=load_dimoldb(),
        material_mapping=_quote_material_mapping_for_spec(),
        mark_flow=mark_flow,
        db_config=DB_CONFIG,
        processes=_permission_data.get("processes", []),
        internal_order_id_fn=ph.internal_order_id,
        infer_order_type_fn=ph.infer_order_type,
        get_or_create_flow_steps_fn=ph.get_or_create_flow_steps,
        save_flow_row_fn=ph.save_flow_row,
        fetch_km_product=True,
    )
    _prod_dash_cache.patch_line_material(so_id, line_index, result)
    return {"success": True, "result": result, "so_id": so_id, "line_index": line_index}


@app.route("/api/production/calc-material", methods=["POST"])
def production_calc_material():
    """订单行算料（展开+纸板+刀模）。"""
    if not resolve_login_user():
        return jsonify({"success": False, "error": "未登录"}), 401
    body = request.get_json() or {}
    so_id = (body.get("so_id") or body.get("order_id") or "").strip()
    line_index = int(body.get("line_index", body.get("line_idx", 0)))
    return jsonify(_calc_material_for_request(so_id, line_index, mark_flow=True))


@app.route("/api/production/calc-material/batch", methods=["POST"])
def production_calc_material_batch():
    """批量算料：body.so_ids 为订单号列表。"""
    if not resolve_login_user():
        return jsonify({"success": False, "error": "未登录"}), 401
    body = request.get_json() or {}
    so_ids = body.get("so_ids") or body.get("order_ids") or []
    if not isinstance(so_ids, list) or not so_ids:
        return jsonify({"success": False, "error": "请提供 so_ids 列表"}), 400

    orders = ph.load_cache_orders(_orders_cache_path())
    by_id = {str(o.get("so_id") or ""): o for o in orders}
    done = failed = 0
    errors: list[str] = []
    total_lines = 0

    for sid in so_ids:
        sid = str(sid).strip()
        order = by_id.get(sid)
        if not order:
            failed += 1
            errors.append(f"{sid}: 未找到订单")
            continue
        for idx in range(len(order.get("items") or [])):
            total_lines += 1
            try:
                r = _calc_material_for_request(sid, idx, mark_flow=True)
                st = (r.get("result") or {}).get("status")
                if st in ("done", "shortage"):
                    done += 1
                else:
                    failed += 1
                    errors.append(
                        f"{sid}#{idx}: {(r.get('result') or {}).get('error', '')}"
                    )
            except Exception as ex:
                failed += 1
                errors.append(f"{sid}#{idx}: {ex}")

    return jsonify(
        {
            "success": True,
            "total_orders": len(so_ids),
            "total_lines": total_lines,
            "lines_done": done,
            "lines_failed": failed,
            "errors": errors[:100],
        }
    )


# ==================== 打单管理 API ====================

@app.route('/api/production/dashboard')
def production_dashboard():
    """打单管理：缓存 + 分页（page/page_size + 筛选）"""
    try:
        import order_visit_sync as _ovs

        _ovs.schedule_incremental_sync(memo_getter=get_order_memo, force=False)
    except Exception:
        pass
    force = request.args.get("refresh") == "1"
    cache = _prod_dash_cache.get_dashboard_cache(
        _rebuild_prod_dashboard_cache, force=force
    )
    orders = list(cache.get("orders") or [])
    _sync_order_extra_from_db()
    for o in orders:
        sid = o.get("so_id") or ""
        o["urgent"] = bool(_order_extra.get(sid, {}).get("urgent"))

    shop = (request.args.get("shop") or "").strip()
    ptype = (request.args.get("type") or "").strip()
    print_status = (request.args.get("print") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    search = (request.args.get("search") or "").strip().lower()

    if shop:
        orders = [o for o in orders if o.get("shop") == shop]
    if ptype == "现货":
        # 现货 = 子单尺寸匹配成品库存且数量足够（非属性里含「现货」字样）
        orders = [o for o in orders if o.get("has_stock")]
    elif ptype:
        orders = [o for o in orders if o.get("product_type") == ptype]
    if print_status == "printed":
        orders = [o for o in orders if o.get("printed")]
    elif print_status == "unprinted":
        orders = [o for o in orders if not o.get("printed")]
    if date_from:
        orders = [o for o in orders if (o.get("created") or "") >= date_from]
    if date_to:
        orders = [o for o in orders if (o.get("created") or "") <= date_to]
    if search:
        orders = [
            o
            for o in orders
            if search in (o.get("so_id") or "").lower()
            or search in (o.get("product_name") or "").lower()
            or search in (o.get("shop") or "").lower()
        ]

    total = len(orders)
    printed_n = sum(1 for o in orders if o.get("printed"))
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    try:
        page_size = int(request.args.get("page_size", 15))
    except ValueError:
        page_size = 15
    page_size = max(5, min(50, page_size))
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    page_orders = orders[start : start + page_size]

    prod_mat_map = _production_material_mapping()
    return jsonify(
        {
            "success": True,
            "orders": page_orders,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "stats": {
                "total": total,
                "printed": printed_n,
                "unprinted": total - printed_n,
            },
            "cached_at": cache.get("ts"),
            "filters": {
                "shops": cache.get("shops") or [],
                "types": cache.get("types") or [],
            },
            "material_mapping": load_quote_data().get("material_mapping", [])
            if load_quote_data()
            else [],
            "production_material_mapping": prod_mat_map,
        }
    )


@app.route('/api/production/dashboard/order/<order_id>')
def production_dashboard_order(order_id):
    """单条订单详情（从缓存取，供分页列表点详情）"""
    o = _prod_dash_cache.find_order_in_cache(order_id)
    if not o:
        _prod_dash_cache.get_dashboard_cache(
            _rebuild_prod_dashboard_cache, force=False
        )
        o = _prod_dash_cache.find_order_in_cache(order_id)
    if not o:
        return jsonify({"success": False, "error": "未找到订单"})
    return jsonify({"success": True, "order": o})

@app.route('/api/production/print_log', methods=['POST'])
def production_print_log():
    """记录打印日志"""
    data = request.get_json() or {}
    order_id = data.get('order_id', '')
    shop_name = data.get('shop_name', '')
    product_type = data.get('product_type', '')
    printed_by = data.get('printed_by', '')
    if not order_id:
        return jsonify({'success': False, 'error': '缺少order_id'})
    try:
        db = get_db()
        cur = db.cursor()
        # 先删除旧记录
        cur.execute("DELETE FROM print_logs WHERE order_id=%s AND status='printed'", (order_id,))
        cur.execute("INSERT INTO print_logs (order_id, shop_name, product_type, printed_by, status) VALUES (%s,%s,%s,%s,'printed')",
                    (order_id, shop_name, product_type, printed_by))
        db.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/production/print_log/<order_id>', methods=['DELETE'])
def production_unmark_print(order_id):
    """取消打印标记"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("DELETE FROM print_logs WHERE order_id=%s AND status='printed'", (order_id,))
        db.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/production/print')
def production_print():
    """打印工单（跳转原生打印）"""
    ids = request.args.get('ids', '')
    sku_id = request.args.get('sku_id', '')
    
    html = '<html><head><meta charset="utf-8"><style>body{font-family:monospace;font-size:12px;padding:20px;}h2{text-align:center;}table{width:100%;border-collapse:collapse;margin-top:12px;}th,td{border:1px solid #ccc;padding:6px 8px;text-align:left;font-size:11px;}th{background:#f0f0f0;}.label{font-weight:600;color:#333;}.value{color:#555;}</style></head><body>'
    
    if sku_id:
        # 按单规格打印
        order_id = ids
        try:
            cache_file = ORDERS_CACHE_FILE
            if not os.path.isabs(cache_file):
                cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), cache_file)
            if os.path.exists(cache_file):
                with open(cache_file, 'r', encoding='utf-8') as f:
                    cache = json.load(f)
                for o in cache.get('orders', []):
                    if o.get('so_id') == order_id:
                        for item in (o.get('items') or []):
                            if item.get('skuId', '') == sku_id:
                                name = item.get('name', '')
                                spec = item.get('spec', '')
                                qty = item.get('qty', 0)
                                sku = item.get('sku', '')
                                html += f'<h2>📋 规格工单</h2>'
                                html += f'<p><span class="label">订单号：</span><span class="value">{order_id}</span></p>'
                                html += f'<p><span class="label">商品：</span><span class="value">{name}</span></p>'
                                html += f'<p><span class="label">规格：</span><span class="value">{spec}</span></p>'
                                html += f'<p><span class="label">数量：</span><span class="value">{qty}</span></p>'
                                if sku:
                                    html += f'<p><span class="label">SKU：</span><span class="value">{sku}</span></p>'
                                break
                        break
        except:
            pass
        if html == '<html><head><meta charset="utf-8"><style>body{font-family:monospace;font-size:12px;padding:20px;}h2{text-align:center;}table{width:100%;border-collapse:collapse;margin-top:12px;}th,td{border:1px solid #ccc;padding:6px 8px;text-align:left;font-size:11px;}th{background:#f0f0f0;}.label{font-weight:600;color:#333;}.value{color:#555;}</style></head><body>':
            html += '<p style="color:#e94560;">未找到该规格</p>'
    else:
        # 整单打印（原有逻辑）
        html += '<h2>📋 生产工单</h2>'
        html += f'<p>订单号: {ids}</p>'
        html += '<p>请使用浏览器打印功能 (Ctrl+P)</p>'
    
    html += '<script>window.print();</script></body></html>'
    return html

@app.route('/api/flow/create', methods=['POST'])
def production_flow_create():
    """创建生产工单（工序来自权限树：岗位→接单→工序）"""
    data = request.get_json() or {}
    order_id = (data.get('order_id') or '').strip()
    if not order_id:
        return jsonify({'success': False, 'error': '缺少order_id'})
    try:
        orders = ph.load_cache_orders(_orders_cache_path())
        o = ph.find_order_in_cache(orders, order_id)
        oid = ph.internal_order_id(o) if o else order_id
        order_type = ph.infer_order_type(o) if o else '飞机盒'
        dept = ph.get_process_dept(_permission_data.get("processes", []), order_type)
        steps = ph.steps_from_dept(dept)
        existed = ph.get_flow_row(DB_CONFIG, oid) is not None
        ph.save_flow_row(DB_CONFIG, oid, order_type, steps)
        msg = '工单已重置' if existed else '工单已创建'
        return jsonify({
            'success': True,
            'message': msg,
            'flow': {
                'order_id': oid,
                'product_type': order_type,
                'total_steps': len(steps),
                'dept': (dept or {}).get('dept', ''),
            },
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/flow/step', methods=['POST'])
def production_flow_step():
    """推进/回退工单工序"""
    data = request.get_json() or {}
    order_id = data.get('order_id', '')
    action = data.get('action', 'next')  # next/prev
    if not order_id:
        return jsonify({'success': False, 'error': '缺少order_id'})
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT * FROM production_flows WHERE order_id=%s", (order_id,))
        flow = cur.fetchone()
        if not flow:
            return jsonify({'success': False, 'error': '工单不存在'})
        idx = flow['current_step_index'] or 0
        total = flow['total_steps'] or len(json.loads(flow['steps_json']) if isinstance(flow['steps_json'], str) else flow['steps_json'])
        if action == 'next' and idx < total - 1:
            idx += 1
        elif action == 'prev' and idx > 0:
            idx -= 1
        status = 'active' if idx < total - 1 else 'done'
        cur.execute("UPDATE production_flows SET current_step_index=%s, status=%s, updated_at=NOW() WHERE order_id=%s", (idx, status, order_id))
        db.close()
        return jsonify({'success': True, 'current_step_index': idx, 'status': status, 'total_steps': total})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ==================== 静态文件（必须放在所有 /api 路由之后）====================
@app.route('/guanli/login/submit', methods=['POST'])
def guanli_login_submit():
    import guanli_server_login as _gsl
    return _gsl.handle_guanli_form_login()


@app.route('/guanli/login')
@app.route('/login_guanli.html')
def guanli_login_entry():
    resp = make_response(send_from_directory('.', 'login_guanli.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    return resp


@app.route('/guanli/logout')
def guanli_logout():
    resp = redirect('/guanli/login')
    for key in ('sanyang_auth_user', 'sanyang_auth_token'):
        resp.delete_cookie(key, path='/')
    return resp


@app.route('/guanli')
@app.route('/guanli/')
def guanli_admin_entry():
    """3003 统一管理后台：须先 /guanli/login 表单登录，服务端校验 Cookie。"""
    import guanli_server_login as _gsl

    un, tok = _gsl.read_guanli_auth_from_request()
    if not un:
        return redirect('/guanli/login')
    user = _gsl.fetch_co_user(un, tok)
    if not user:
        resp = redirect('/guanli/login?error=auth')
        resp.delete_cookie('sanyang_auth_user', path='/')
        resp.delete_cookie('sanyang_auth_token', path='/')
        return resp
    base_dir = os.path.dirname(os.path.abspath(__file__))
    html_path = os.path.join(base_dir, 'index_customer_order.html')
    with open(html_path, encoding='utf-8') as f:
        html = f.read()
    body = _gsl.inject_preauth_html(html, user)
    resp = make_response(body)
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    return resp


@app.route('/')
def index():
    resp = make_response(send_from_directory('.', 'index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/static/<path:filename>')
def serve_static_assets(filename):
    """显式提供 static/（避免反代或通配路由导致 auth_session.js 404）"""
    return send_from_directory('static', filename)


@app.route('/<path:path>')
def static_files(path):
    if path.startswith('api/') or path == 'api':
        from werkzeug.exceptions import NotFound
        raise NotFound()
    return send_from_directory('.', path)


if __name__ == '__main__':
    print("🏭 飞机盒智能生产管理系统启动中...")
    print("📡 http://0.0.0.0:3002")
    app.run(host='0.0.0.0', port=3002, threaded=True)

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞机盒智能生产管理系统 - 后端API
小马哥专属
"""

from flask import Flask, jsonify, send_from_directory, request, make_response, Response, session
from flask_cors import CORS
import json, datetime, csv, io, os, hashlib, copy, time, re, hmac, urllib.parse, urllib.request
from datetime import timedelta
from pypinyin import lazy_pinyin
import pymysql

from settings import (
    ALIBABA_CONFIG,
    ALIBABA_SHOPS,
    DB_CONFIG,
    FLASK_SECRET_KEY,
    get_wechat_token,
)

def get_db():
    """获取数据库连接，每次调用新建（线程安全）"""
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)

app = Flask(__name__, static_folder='.')
app.secret_key = FLASK_SECRET_KEY
CORS(app, supports_credentials=True)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=31)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_COOKIE_SECURE'] = os.getenv('SESSION_COOKIE_SECURE', '').lower() in (
    '1', 'true', 'yes'
)

# 报价数据文件路径
QUOTE_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'quote_data.json')

WECHAT_TOKEN = get_wechat_token()

# ==================== 用户系统 ====================
USERS = {
    "admin": {
        "password": hashlib.sha256("admin888".encode()).hexdigest(),
        "name": "戴雅利",
        "role": "超级管理员",
        "employee_name": "戴雅利"
    },
    "manager": {
        "password": hashlib.sha256("manager666".encode()).hexdigest(),
        "name": "邓涛",
        "role": "管理",
        "employee_name": "邓涛"
    },
    "worker": {
        "password": hashlib.sha256("worker123".encode()).hexdigest(),
        "name": "李周海",
        "role": "员工",
        "employee_name": "李周海"
    },
    "sushiting": {
        "password": hashlib.sha256("sushiting123".encode()).hexdigest(),
        "name": "苏世婷",
        "role": "客服",
        "employee_name": "苏世婷"
    },
    "liaosimei": {
        "password": hashlib.sha256("liaosimei123".encode()).hexdigest(),
        "name": "廖思美",
        "role": "客服",
        "employee_name": "廖思美"
    }
}


def _auth_token_for(username: str) -> str:
    """与前端 sessionStorage 同步，POST 未带 Cookie 时兜底鉴权。"""
    return hashlib.sha256(f"{FLASK_SECRET_KEY}:{username}".encode()).hexdigest()[:32]


def resolve_login_user() -> str | None:
    """Session Cookie 或 X-Sanyang-User + X-Sanyang-Token（登录后 /api/me 下发）。"""
    un = session.get("username")
    if un and un in USERS:
        return un
    hdr_user = (request.headers.get("X-Sanyang-User") or "").strip()
    hdr_tok = (request.headers.get("X-Sanyang-Token") or "").strip()
    if hdr_user in USERS and hdr_tok and hdr_tok == _auth_token_for(hdr_user):
        return hdr_user
    return None


def require_login():
    """检查是否已登录"""
    if not resolve_login_user():
        return jsonify({"success": False, "error": "未登录", "code": 401}), 401
    return None

def require_admin():
    """检查是否为超级管理员"""
    if 'username' not in session:
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] != '超级管理员':
        return jsonify({"error": "无权限", "code": 403}), 403
    return None

# ==================== 数据持久化 ====================

def load_data():
    """从MySQL加载持久化数据"""
    default_users = {
        "admin": {
            "password": hashlib.sha256("admin888".encode()).hexdigest(),
            "name": "admin",
            "role": "超级管理员",
            "employee_name": "",
            "is_system": True
        },
        "daiyali": {
            "password": hashlib.sha256("admin888".encode()).hexdigest(),
            "name": "戴雅利",
            "role": "超级管理员",
            "employee_name": "戴雅利"
        },
        "manager": {
            "password": hashlib.sha256("manager666".encode()).hexdigest(),
            "name": "邓涛",
            "role": "管理",
            "employee_name": "邓涛"
        },
        "worker": {
            "password": hashlib.sha256("worker123".encode()).hexdigest(),
            "name": "李周海",
            "role": "员工",
            "employee_name": "李周海"
        },
        "sushiting": {
            "password": hashlib.sha256("sushiting123".encode()).hexdigest(),
            "name": "苏世婷",
            "role": "客服",
            "employee_name": "苏世婷"
        },
        "liaosimei": {
            "password": hashlib.sha256("liaosimei123".encode()).hexdigest(),
            "name": "廖思美",
            "role": "客服",
            "employee_name": "廖思美"
        }
    }
    default = {
        "users": default_users,
        "employee_status": {},
        "permission_data": {
            "processes": [
                {
                    "dept": "美丽湾工厂部",
                    "steps": [
                        {"id": 1, "name": "客服接单"},
                        {"id": 2, "name": "黄厂打印"},
                        {"id": 3, "name": "审单分单"},
                        {"id": 4, "name": "算料"},
                        {"id": 5, "name": "分纸"},
                        {"id": 6, "name": "啤机(自动平压平)"},
                        {"id": 7, "name": "啤机(机械手)"},
                        {"id": 8, "name": "手啤"},
                        {"id": 9, "name": "印刷"},
                        {"id": 10, "name": "清废"},
                        {"id": 11, "name": "打包发货"}
                    ]
                },
                {
                    "dept": "纸箱部",
                    "steps": [
                        {"id": 1, "name": "客服接单"},
                        {"id": 2, "name": "黄厂打印"},
                        {"id": 3, "name": "审单分单"},
                        {"id": 4, "name": "算料"},
                        {"id": 5, "name": "分纸"},
                        {"id": 6, "name": "印刷"},
                        {"id": 7, "name": "开槽/打角"},
                        {"id": 8, "name": "粘胶/打钉"},
                        {"id": 9, "name": "打包发货"}
                    ]
                }
            ],
            "positions": ["超级管理员", "主管", "客服", "员工", "财务", "业务员"],
            "employee_enabled": {},
            "employees": [
                {"name": "戴雅利", "position": "超级管理员"},
                {"name": "邓涛", "position": "主管"},
                {"name": "黄兴", "position": "主管"},
                {"name": "覃海霞", "position": "主管"},
                {"name": "蒋义红", "position": "主管"},
                {"name": "沈齐豪", "position": "主管"},
                {"name": "苏世婷", "position": "客服"},
                {"name": "廖思美", "position": "客服"},
                {"name": "张慧平", "position": "客服"},
                {"name": "陈贝贝", "position": "客服"},
                {"name": "罗怡", "position": "客服"},
                {"name": "周井梅", "position": "客服"},
                {"name": "戴志美", "position": "客服"},
                {"name": "石梅清", "position": "客服"},
                {"name": "张文杰", "position": "客服"},
                {"name": "何水单", "position": "业务员"},
                {"name": "李四军", "position": "员工"},
                {"name": "陈贤聪", "position": "业务员"},
                {"name": "姚斌", "position": "员工"},
                {"name": "隆浪", "position": "财务"},
            ],
            "permissions": {
                "戴雅利": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "数据看板": True, "刀模": True, "库存": True, "快麦ERP": True, "员工": True, "权限管理": True, "报价": True},
                "邓涛": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "数据看板": True, "刀模": True, "库存": True, "快麦ERP": True, "员工": True, "权限管理": False, "报价": False},
                "李周海": {"首页": True, "订单生产进度": True, "扫码报工": True, "日报表": True, "员工": True, "刀模": False, "库存": False, "快麦ERP": False, "数据看板": False, "权限管理": False, "报价": False},
            }
        }
    }
    try:
        db = get_db()
        cur = db.cursor()
        # 从users表读取
        cur.execute("SELECT username, password, display_name, role, employee_name, enabled FROM users")
        rows = cur.fetchall()
        users = {}
        for r in rows:
            users[r['username']] = {
                'password': r['password'],
                'name': r['display_name'] or r['username'],
                'role': r['role'] or '员工',
                'employee_name': r['employee_name'] or ''
            }
        # 从employees表读取
        cur.execute("SELECT name, position, enabled FROM employees")
        emp_rows = cur.fetchall()
        employees = [{'name': r['name'], 'position': r['position'] or ''} for r in emp_rows]
        employee_enabled = {}
        for r in emp_rows:
            employee_enabled[r['name']] = bool(r['enabled'])

        # 从employee_permissions表读取
        cur.execute("SELECT employee_name, permission_key, enabled FROM employee_permissions")
        perm_rows = cur.fetchall()
        permission_map = {}
        all_perm_keys = set()
        for r in perm_rows:
            en = r['employee_name']
            pk = r['permission_key']
            val = bool(r['enabled'])
            all_perm_keys.add(pk)
            if en not in permission_map:
                permission_map[en] = {}
            permission_map[en][pk] = val

        # 构建permissions字典（聚水潭 → 快麦ERP 兼容）
        permissions = {}
        for emp in employees:
            en = emp['name']
            perms = permission_map.get(en, {})
            row = {pk: perms.get(pk, False) for pk in sorted(all_perm_keys)}
            if '聚水潭' in row and '快麦ERP' not in row:
                row['快麦ERP'] = row.pop('聚水潭')
            elif '聚水潭' in row:
                row.pop('聚水潭', None)
            permissions[en] = row

        # 从order_extras表读取
        cur.execute("SELECT so_id, urgent FROM order_extras")
        extra_rows = cur.fetchall()
        order_extra = {}
        for r in extra_rows:
            order_extra[r['so_id']] = {'urgent': bool(r['urgent'])}

        cur.close()
        db.close()

        result = {
            'users': users,
            'employee_status': {},
            'order_extra': order_extra,
            'permission_data': {
                'processes': default['permission_data']['processes'],
                'positions': default['permission_data']['positions'],
                'employees': employees,
                'permissions': permissions,
                'employee_enabled': employee_enabled
            },
            'resigned_employees': []
        }
        return result
    except Exception as e:
        print(f'[MySQL load_data] 错误: {e}')
        return default

def save_data(data):
    """保存数据到MySQL"""
    try:
        db = get_db()
        cur = db.cursor()
        # 保存users
        users = data.get('users', {})
        for uid, uinfo in users.items():
            if uid == 'admin' and uinfo.get('is_system'):
                continue  # 不覆盖系统admin
            cur.execute(
                "INSERT INTO users (username, password, display_name, role, employee_name, enabled) VALUES (%s,%s,%s,%s,%s,1) ON DUPLICATE KEY UPDATE password=VALUES(password), display_name=VALUES(display_name), role=VALUES(role), employee_name=VALUES(employee_name)",
                (uid, uinfo.get('password',''), uinfo.get('name',uid), uinfo.get('role','员工'), uinfo.get('employee_name',''))
            )

        # 保存employees
        emp_list = data.get('permission_data', {}).get('employees', [])
        for emp in emp_list:
            name = emp.get('name', '')
            pos = emp.get('position', '')
            cur.execute(
                "INSERT INTO employees (name, position, enabled) VALUES (%s,%s,1) ON DUPLICATE KEY UPDATE position=VALUES(position)",
                (name, pos)
            )

        # 保存employee_permissions
        perms = data.get('permission_data', {}).get('permissions', {})
        for emp_name, perm_dict in perms.items():
            for pk, val in perm_dict.items():
                cur.execute(
                    "INSERT INTO employee_permissions (employee_name, permission_key, enabled) VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE enabled=VALUES(enabled)",
                    (emp_name, pk, 1 if val else 0)
                )

        # 保存employee_enabled
        emp_enabled = data.get('permission_data', {}).get('employee_enabled', {})
        for name, enabled in emp_enabled.items():
            cur.execute(
                "UPDATE employees SET enabled=%s WHERE name=%s",
                (1 if enabled else 0, name)
            )

        # 保存order_extras
        order_extra = data.get('order_extra', {})
        for so_id, info in order_extra.items():
            urgent = 1 if info.get('urgent') else 0
            cur.execute(
                "INSERT INTO order_extras (so_id, urgent) VALUES (%s,%s) ON DUPLICATE KEY UPDATE urgent=VALUES(urgent)",
                (so_id, urgent)
            )

        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_data] 错误: {e}')
        return False

# 初始化数据
persistent_data = load_data()

# 从data.json加载所有用户（覆盖并扩展USERS字典）
_loaded_users = persistent_data.get("users", {})
for uid, uinfo in _loaded_users.items():
    USERS[uid] = {
        "password": uinfo.get("password", ""),
        "name": uinfo.get("name", uid),
        "role": uinfo.get("role", "员工"),
        "employee_name": uinfo.get("employee_name", uinfo.get("name", uid))
    }
_employee_today_status = persistent_data.get("employee_status", {})
_order_extra = persistent_data.get("order_extra", {})
_permission_data = persistent_data.get("permission_data", {})
_resigned_employees = persistent_data.get("resigned_employees", [])  # 离职员工列表[{name, position, group, dept, phone, resigned_time}]

# 将员工状态变更为按日期存储格式（兼容旧格式）
if isinstance(_employee_today_status, dict):
    # 检查是 "date -> name -> status" 还是老格式
    is_old_format = False
    for k, v in _employee_today_status.items():
        if isinstance(v, str):
            is_old_format = True
            break
    if is_old_format:
        today = datetime.date.today().isoformat()
        _employee_today_status = {today: _employee_today_status}

def persist():
    """持久化当前数据到文件"""
    global _employee_today_status, _permission_data, USERS, _resigned_employees
    # 把USERS的密码也存到文件（用户改密码后重启不丢）
    users_data = {}
    for uid, u in USERS.items():
        entry = {
            "password": u["password"],
            "name": u["name"],
            "role": u["role"],
            "employee_name": u["employee_name"]
        }
        if u.get("is_system"):
            entry["is_system"] = True
        users_data[uid] = entry
    data = {
        "users": users_data,
        "employee_status": _employee_today_status,
        "order_extra": _order_extra,
        "permission_data": _permission_data,
        "resigned_employees": _resigned_employees
    }
    save_data(data)

# ==================== 登录API ====================
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({"success": False, "message": "请输入账号和密码"})
    
    user = USERS.get(username)
    if not user:
        return jsonify({"success": False, "message": "账号或密码错误"})
    
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    if user['password'] != pwd_hash:
        return jsonify({"success": False, "message": "账号或密码错误"})
    
    # 检查员工是否被禁用
    employee_name = user.get('employee_name', '')
    enabled_map = _permission_data.get("employee_enabled", {})
    # 默认启用，只有明确设为 False 才禁用
    if enabled_map.get(employee_name, True) is False:
        return jsonify({"success": False, "message": "该账号已被禁用，请联系管理员"})
    
    session.permanent = True
    session['username'] = username
    session['user_name'] = user['name']
    session['role'] = user['role']
    session['employee_name'] = user['employee_name']
    session.modified = True

    # 查找用户所属部门
    emp_dept = ''
    for emp in _employees_master_list:
        if emp['name'] == user['employee_name']:
            emp_dept = emp.get('dept', '')
            break

    return jsonify({
        "success": True,
        "message": "登录成功",
        "auth_token": _auth_token_for(username),
        "user": {
            "username": username,
            "name": user['name'],
            "role": user['role'],
            "employee_name": user['employee_name'],
            "dept": emp_dept
        }
    })

@app.route('/api/logout')
def logout():
    session.clear()
    return jsonify({"success": True, "message": "已退出登录"})

@app.route('/api/change_password', methods=['POST'])
def change_password():
    """仅允许修改当前登录账号自己的密码。"""
    username = resolve_login_user()
    if not username:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    target = (data.get('username') or data.get('target_user') or '').strip()
    if target and target != username:
        return jsonify({"success": False, "message": "只能修改自己的密码"}), 403

    old_pwd = data.get('old_password', '').strip()
    new_pwd = data.get('new_password', '').strip()

    if not old_pwd or not new_pwd:
        return jsonify({"success": False, "message": "请填写旧密码和新密码"})

    if len(new_pwd) < 4:
        return jsonify({"success": False, "message": "新密码至少4位"})

    user = USERS.get(username)
    if not user:
        return jsonify({"success": False, "message": "用户不存在"})

    old_hash = hashlib.sha256(old_pwd.encode()).hexdigest()
    if user['password'] != old_hash:
        return jsonify({"success": False, "message": "旧密码错误"})

    user['password'] = hashlib.sha256(new_pwd.encode()).hexdigest()
    persist()
    return jsonify({"success": True, "message": "密码修改成功"})


@app.route('/api/admin/reset_password', methods=['POST'])
def admin_reset_password():
    """超级管理员重置他人密码（普通员工不可用）。"""
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "message": "未登录"}), 401
    actor = USERS.get(un)
    if not actor or actor.get('role') != '超级管理员':
        return jsonify({"success": False, "message": "无权限"}), 403

    data = request.get_json() or {}
    target = (data.get('username') or '').strip()
    new_pwd = (data.get('new_password') or '').strip()
    if not target or not new_pwd:
        return jsonify({"success": False, "message": "请指定账号和新密码"})
    if len(new_pwd) < 4:
        return jsonify({"success": False, "message": "新密码至少4位"})
    if target == un:
        return jsonify({"success": False, "message": "请使用「修改密码」修改自己的密码"})

    user = USERS.get(target)
    if not user:
        return jsonify({"success": False, "message": "目标账号不存在"})
    user['password'] = hashlib.sha256(new_pwd.encode()).hexdigest()
    persist()
    return jsonify({"success": True, "message": f"已重置账号 {target} 的密码"})

@app.route('/api/me')
def get_current_user():
    un = resolve_login_user()
    if un:
        user = USERS.get(un)
        if user:
            return jsonify({
                "logged_in": True,
                "auth_token": _auth_token_for(un),
                "user": {
                    "username": un,
                    "name": user['name'],
                    "role": user['role'],
                    "employee_name": user['employee_name']
                }
            })
    return jsonify({"logged_in": False})

# ==================== 规格解析工具函数 ====================

def _trim_digits(s):
    """把数字转为字符串，去掉不必要的.0"""
    try:
        n = float(s.group())
        if n == int(n):
            return str(int(n))
        return s.group()
    except:
        return s.group()

def _parse_item_display(spec, name, qty=0):
    """从买家下单 spec 解析展示；无规格时不回退商品标题。"""
    del name, qty
    if not spec:
        return ''
    
    # 从报价配置加载材质映射
    try:
        with open(QUOTE_DATA_FILE, 'r', encoding='utf-8') as f:
            qd = json.load(f)
        mat_config = qd.get('materials', {})
    except:
        mat_config = {}
    
    mat_map = {}
    for ptype, cfg in mat_config.items():
        for mk, mv in cfg.get('materials', {}).items():
            display_name = mv.get('name', mk)
            mat_map[display_name] = display_name
    extra_map = {
        '特硬': '特硬 D6D', '超硬': '台湾纸(超硬)', '台湾纸': '台湾纸(超硬)',
        '白色': '白色 W7W', '双白': '白色 W7W', '双面白色': '白色 W7W',
        '黑色': '黑色', '红色': '红色', '优质': '优质 Q7Q', '特价': '优质 Q7Q',
        '牛皮': '牛皮', '三层加硬': '三层B坑 K7K', '黄色': '黄色',
    }
    mat_map.update(extra_map)
    
    mat_keys = ['三层加硬', '双面白色', '台湾纸', '超硬', '特硬', '优质', '特价', '牛皮', '白色', '双白', '黑色', '红色', '黄色']
    mat = ''
    for k in mat_keys:
        if k in spec:
            mat = mat_map.get(k, k)
            break
    
    clean = spec
    hm = re.search(r'(?:高(?:度)?[【\\[]?|【)(\d+\.?\d*)\s*(?:[cC][mM]?|厘米)(?:高|)', clean)
    has_only_height = hm and not re.search(r'[长宽×\*xX]', clean)
    
    # 高【Ncm】或高【N厘米】格式：只有高维度（必须在clean之前提取，避免被替换掉）

    
    clean = re.sub(r'\d+\s*层\s*纸箱', '', clean)
    clean = re.sub(r'\d+\s*级', '', clean)
    clean = re.sub(r'[【\[]?\s*\d+\s*个[）\)]?', '', clean)
    clean = re.sub(r'^\s*\d+\s*个\s*', '', clean)
    clean = re.sub(r'\bmm\b', '', clean)
    clean = re.sub(r'[（(][^）)]*不接受[^）)]*[）)]', '', clean)
    
    # 提取全部数字
    all_nums = [float(x.group()) for x in re.finditer(r'\d+\.?\d*', clean) if float(x.group()) > 0 and float(x.group()) < 2000]
    dim_nums = [n for n in all_nums if abs(n - qty) > 0.01]
    if not dim_nums:
        dim_nums = [n for n in all_nums if n != qty]
    
    fmt = lambda n: str(int(n)) if n == int(n) else str(n)
    
    # 处理只有高度的情况（已提前提取hm和has_only_height）
    if has_only_height:
        hv = float(hm.group(1))
        # 尝试从name提取长宽（如"宽度14CM高度7CM"或"14*7"）
        name_lw = re.search(r'宽度(\d+\.?\d*).*?高度(\d+\.?\d*)|长[度]?(\d+\.?\d*).*?宽[度]?(\d+\.?\d*)|(\d+\.?\d*)\s*[×\*xX]\s*(\d+\.?\d*)', name)
        if name_lw:
            n1 = name_lw.group(1) or ''
            n2 = name_lw.group(2) or ''
            if n1 and n2:
                nm = f'{fmt(float(n1))}×{fmt(float(n2))}×{fmt(hv)}'
                return f'{mat} {nm}' if mat else nm
        return f'{mat} {fmt(hv)}' if mat else fmt(hv)
    
    # 数字×数字×数字 格式
    m3 = re.search(r'(\d+\.?\d*)\s*[*×xX]\s*(\d+\.?\d*)\s*[*×xX]\s*(\d+\.?\d*)', clean)
    if m3:
        d = f'{fmt(float(m3.group(1)))}×{fmt(float(m3.group(2)))}×{fmt(float(m3.group(3)))}'
        hv = float(m3.group(3))
        if abs(hv - qty) < 0.01:
            d = f'{fmt(float(m3.group(1)))}×{fmt(float(m3.group(2)))}'
        return f'{mat} {d}' if mat else d
    
    # 找不到三数字×格式，但可能有"长宽"格式
    lw = re.search(r'[长宽]\s*[*×xX]?\s*(\d+\.?\d*)\s*[*×xX]\s*(\d+\.?\d*)', clean)
    h = re.search(r'[高]\s*(\d+\.?\d*)', clean)
    if lw:
        d = f'{fmt(float(lw.group(1)))}×{fmt(float(lw.group(2)))}'
        if h:
            hv = float(h.group(1))
            if abs(hv - qty) > 0.01:
                d += f'×{fmt(hv)}'
        return f'{mat} {d}' if mat else d
    
    # 数字×数字
    pair = re.search(r'(\d+\.?\d*)\s*[*×xX]\s*(\d+\.?\d*)', clean)
    if pair:
        d = f'{fmt(float(pair.group(1)))}×{fmt(float(pair.group(2)))}'
        if len(dim_nums) >= 3:
            used = {float(pair.group(1)), float(pair.group(2))}
            for n in dim_nums:
                if all(abs(n-u) > 0.01 for u in used):
                    d += f'×{fmt(n)}'
                    break
        return f'{mat} {d}' if mat else d
    
    # 兜底
    if len(dim_nums) >= 3:
        d = f'{fmt(dim_nums[0])}×{fmt(dim_nums[1])}×{fmt(dim_nums[2])}'
        return f'{mat} {d}' if mat else d
    elif len(dim_nums) >= 2:
        d = f'{fmt(dim_nums[0])}×{fmt(dim_nums[1])}'
        return f'{mat} {d}' if mat else d
    elif mat:
        return mat
    
    s = spec.replace(';', ' ').replace('；', ' ').strip()
    return s[:40]

def _order_date_key(o: dict) -> str:
    """从订单字段提取 YYYY-MM-DD（本地日历日）。"""
    for field in ('pay_time', 'created', 'consign_time', 'updTime', 'updated'):
        raw = str(o.get(field) or '').strip()
        if not raw:
            continue
        if len(raw) >= 10 and raw[4] == '-' and raw[7] == '-':
            return raw[:10]
        digits = ''.join(c for c in raw if c.isdigit())
        if len(digits) >= 8:
            return f'{digits[:4]}-{digits[4:6]}-{digits[6:8]}'
    return ''


def _order_on_date(o: dict, date_yyyy_mm_dd: str) -> bool:
    return _order_date_key(o) == date_yyyy_mm_dd


def _filter_orders_by_range(orders: list, range_type: str) -> list:
    """店铺看板时间筛选：live/yesterday/week/month/last30/quarter。"""
    now = datetime.datetime.now()
    today = now.date()
    out = []
    for o in orders:
        dk = _order_date_key(o)
        if not dk:
            continue
        try:
            od = datetime.datetime.strptime(dk, '%Y-%m-%d').date()
        except ValueError:
            continue
        if range_type == 'live':
            if od != today:
                continue
        elif range_type == 'yesterday':
            y = today - datetime.timedelta(days=1)
            if od != y:
                continue
        elif range_type == 'week':
            week_start = today - datetime.timedelta(days=today.weekday())
            if od < week_start:
                continue
        elif range_type == 'month':
            if od.year != today.year or od.month != today.month:
                continue
        elif range_type == 'last30':
            if od < today - datetime.timedelta(days=29):
                continue
        elif range_type == 'quarter':
            q_start_month = ((today.month - 1) // 3) * 3 + 1
            q_start = datetime.date(today.year, q_start_month, 1)
            if od < q_start:
                continue
        else:
            if od < today - datetime.timedelta(days=6):
                continue
        out.append(o)
    return out


def _dashboard_shop_info(shop_name: str) -> dict:
    """店铺名 → 首页展示用 store/worker（结合 shop_config）。"""
    try:
        import order_sync as _osync
        short = _osync.normalize_shop_display(shop_name or '')
    except ImportError:
        short = (shop_name or '').replace('阿里', '').replace('包装', '')[:8]
    worker = ''
    for sc in load_shop_config():
        sn = (sc.get('shop_name') or '').strip()
        if not sn:
            continue
        if sn in (shop_name or '') or (short and sn in short) or (short and short in sn):
            cs = (sc.get('customer_service') or '').strip()
            worker = cs.split(',')[0].strip() if cs else ''
            break
    store = f'深圳{short}' if short else (shop_name or '—')
    return {'store': store, 'worker': worker}


@app.route('/api/dashboard')
def dashboard():
    """今日订单 - 从 orders_cache（快麦+1688）读取"""
    date = request.args.get('date', datetime.date.today().strftime('%Y-%m-%d'))
    
    # 读取订单缓存（优先 MySQL）
    orders = []
    total_sales = 0.0
    try:
        import production_helpers as _ph

        raw_orders = _ph.load_cache_orders(_orders_cache_path())
        if raw_orders:
            try:
                import km_api as _km
                for o in raw_orders:
                    _km.finalize_cache_order(o)
            except ImportError:
                pass
            prod_by_id = {
                x['inner_id']: x
                for x in _build_production_orders_for_cs()
            }
            today_orders_raw = [o for o in raw_orders if _order_on_date(o, date)]
            
            for o in today_orders_raw:
                shop_1688 = o.get('shop_name', '')
                shop_info = _dashboard_shop_info(shop_1688)
                oid = str(o.get('so_id') or o.get('km_sid') or '')
                items = o.get('items', [])
                product_names = '; '.join([f"{i.get('name','?')[:20]} x{i.get('qty',0)}" for i in items[:2]])
                if len(items) > 2:
                    product_names += f' …等{len(items)}种'
                _raw_dt = o.get('pay_time') or o.get('created') or ''
                _digits = ''.join(c for c in str(_raw_dt) if c.isdigit())
                _pay_yyyymmdd = _digits[:8] if len(_digits) >= 8 else ''
                
                # 构建结构化items（含规格/尺寸/数量）
                detail_items = []
                for i in items:
                    name = i.get('name', '') or ''
                    spec = i.get('spec', '') or ''
                    sku = i.get('sku', '') or ''
                    qty = i.get('qty', 0) or 0
                    
                    full_spec = (i.get('display') or spec or sku or '').strip()
                    detail_items.append({
                        'name': name,
                        'spec': full_spec,
                        'qty': qty,
                        'display': full_spec,
                    })
                
                ex = _order_extra.get(oid, {})
                flow = (prod_by_id.get(oid) or {}).get('flow') or []
                ps = _production_status_from_flow(flow)
                try:
                    import km_api as _km
                    amt = _km.km_to_float(o.get('total_amount'))
                except ImportError:
                    try:
                        amt = float(o.get('total_amount') or 0)
                    except (TypeError, ValueError):
                        amt = 0.0
                total_sales += amt
                orders.append({
                    'id': oid,
                    'store': shop_info['store'],
                    'worker': shop_info['worker'],
                    'product': product_names[:50] if product_names else (items[0].get('name','?')[:30] if items else '？'),
                    'qty': sum(i.get('qty', 0) for i in items),
                    'items': detail_items,  # 返回结构化items
                    'process': ps.get('current_process', '待处理'),
                    'status': ps.get('status', '待发货'),
                    'progress': ps.get('progress', 0),
                    'province': (o.get('receiver_province') or '').strip(),
                    'city': (o.get('receiver_city') or '').strip(),
                    'urgent': bool(ex.get('urgent')),
                    'remark': o.get('seller_memo', '') or o.get('buyer_memo', ''),
                    'pay_time': _pay_yyyymmdd,
                    'created': o.get('created', ''),
                    'total_amount': amt,
                })
    except Exception as e:
        print(f'[今日订单] 读取缓存失败: {e}')
    
    return jsonify({
        'date': date,
        'summary': {
            'total_orders': len(orders),
            'in_production': len([o for o in orders if o.get('status') not in ('已完成',)]),
            'urgent_orders': len([o for o in orders if o.get('urgent')]),
            'completed': len([o for o in orders if o.get('status') == '已完成']),
            'total_sales': round(total_sales, 2),
        },
        'today_orders': orders,
    })


@app.route('/api/databoard')
def shop_databoard():
    """店铺数据看板：基于 orders_cache 真实订单统计（客服端）。"""
    from collections import defaultdict

    range_type = request.args.get('range', 'week')
    now = datetime.datetime.now()

    raw_orders: list = []
    try:
        if os.path.exists(ORDERS_CACHE_FILE):
            with open(ORDERS_CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            raw_orders = list(cache.get('orders') or [])
            try:
                import km_api as _km
                for o in raw_orders:
                    _km.finalize_cache_order(o)
            except ImportError:
                pass
    except Exception as e:
        print(f'[店铺看板] 读取缓存失败: {e}')

    in_range = _filter_orders_by_range(raw_orders, range_type)
    if range_type == 'live':
        days_span = 1
    elif range_type == 'yesterday':
        days_span = 1
    elif range_type == 'week':
        days_span = max(1, now.date().weekday() + 1)
    elif range_type == 'month':
        days_span = now.day
    elif range_type == 'last30':
        days_span = 30
    elif range_type == 'quarter':
        q_start_month = ((now.month - 1) // 3) * 3 + 1
        days_span = max(1, (now.date() - datetime.date(now.year, q_start_month, 1)).days + 1)
    else:
        days_span = 7

    try:
        import km_api as _km
        _amt = _km.km_to_float
    except ImportError:
        def _amt(v, default=0.0):
            try:
                return float(v or 0)
            except (TypeError, ValueError):
                return default

    total_amount = sum(_amt(o.get('total_amount')) for o in in_range)
    store_counts: dict[str, int] = defaultdict(int)
    platform_counts: dict[str, int] = defaultdict(int)
    for o in in_range:
        store_counts[(o.get('shop_name') or '未知').strip() or '未知'] += 1
        plat = (o.get('platform_label') or o.get('platform') or '其他').strip()
        platform_counts[plat] += 1

    trend = []
    for i in range(min(14, days_span) - 1, -1, -1):
        d = (now - datetime.timedelta(days=i)).date()
        ds = d.strftime('%Y-%m-%d')
        cnt = sum(1 for o in in_range if _order_on_date(o, ds))
        trend.append({'date': d.strftime('%m-%d'), 'output': cnt, 'orders': cnt})

    total_o = len(in_range) or 1
    store_distribution = [
        {
            'name': name,
            'count': cnt,
            'pct': round(cnt * 100 / total_o),
        }
        for name, cnt in sorted(store_counts.items(), key=lambda x: -x[1])[:10]
    ]
    platform_distribution = [
        {
            'name': name,
            'count': cnt,
            'pct': round(cnt * 100 / total_o),
        }
        for name, cnt in sorted(platform_counts.items(), key=lambda x: -x[1])
    ]

    urgent_n = 0
    for o in in_range:
        oid = str(o.get('so_id') or o.get('km_sid') or '')
        if _order_extra.get(oid, {}).get('urgent'):
            urgent_n += 1

    hourly = [{'hour': f'{h:02}', 'val': 0} for h in range(8, 21)]
    for o in in_range:
        raw = str(o.get('pay_time') or o.get('created') or '')
        digits = ''.join(c for c in raw if c.isdigit())
        if len(digits) >= 10:
            try:
                h = int(digits[8:10])
                if 8 <= h <= 20:
                    hourly[h - 8]['val'] += 1
            except ValueError:
                pass

    recent = sorted(
        in_range,
        key=lambda x: str(x.get('created') or x.get('pay_time') or ''),
        reverse=True,
    )[:8]
    recent_news = []
    for o in recent:
        oid = str(o.get('so_id') or o.get('km_sid') or '')
        shop = o.get('shop_name') or ''
        recent_news.append({
            'icon': '📦',
            'text': f'{shop} #{oid} 待发货',
            'time': (o.get('created') or o.get('pay_time') or '')[-8:],
        })

    return jsonify({
        'stats': {
            'total_output': int(round(total_amount)),
            'total_orders': len(in_range),
            'completed': 0,
            'in_production': len(in_range),
            'avg_daily': round(len(in_range) / days_span, 1),
            'on_time_rate': '-',
            'defect_rate': '-',
            'urgent_count': urgent_n,
        },
        'trend': trend or [{'date': now.strftime('%m-%d'), 'output': 0, 'orders': 0}],
        'store_distribution': store_distribution,
        'platform_distribution': platform_distribution,
        'process_load': [
            {
                'name': d['name'],
                'current': d['count'],
                'capacity': total_o,
                'pct': d['pct'],
            }
            for d in platform_distribution
        ],
        'product_type': store_distribution[:5],
        'hourly_output': hourly,
        'worker_productivity': [],
        'recent_news': recent_news,
    })





# ==================== 订单备注和加急API ====================
@app.route('/api/order/remark', methods=['POST'])
def order_remark():
    """设置订单备注 - 仅 超级管理员/管理/客服 可操作"""
    if 'username' not in session:
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] == '员工':
        return jsonify({"error": "无权限", "code": 403}), 403
    global _order_extra
    data = request.get_json()
    oid = data.get('order_id', '')
    remark = data.get('remark', '')
    if oid:
        if oid not in _order_extra:
            _order_extra[oid] = {}
        _order_extra[oid]['remark'] = remark
        persist()
    return jsonify({"success": True, "order_id": oid, "remark": remark})

@app.route('/api/order/urgent', methods=['POST'])
def order_urgent():
    """切换加急状态 - 仅 超级管理员/管理/客服 可操作"""
    if 'username' not in session:
        return jsonify({"error": "未登录", "code": 401}), 401
    user = USERS.get(session['username'])
    if not user or user['role'] == '员工':
        return jsonify({"error": "无权限", "code": 403}), 403
    global _order_extra
    data = request.get_json()
    oid = data.get('order_id', '')
    urgent = data.get('urgent', False)
    if oid:
        if oid not in _order_extra:
            _order_extra[oid] = {}
        _order_extra[oid]['urgent'] = urgent
        persist()
    return jsonify({"success": True, "order_id": oid, "urgent": urgent})

# 功能列表（保持不变）
PERM_FEATURES = ["首页","订单生产进度","扫码报工","日报表","数据看板","刀模","库存","原材料","快麦ERP","员工","权限管理","报价","实时订单"]
_PERM_LEGACY_KEY = "聚水潭"
_PERM_CURRENT_KEY = "快麦ERP"


def _migrate_perm_dict(perms):
    """权限键 聚水潭 → 快麦ERP（兼容旧数据）。"""
    if not isinstance(perms, dict):
        return perms
    if _PERM_LEGACY_KEY in perms and _PERM_CURRENT_KEY not in perms:
        perms[_PERM_CURRENT_KEY] = perms.pop(_PERM_LEGACY_KEY)
    elif _PERM_LEGACY_KEY in perms:
        perms.pop(_PERM_LEGACY_KEY, None)
    return perms


def _orders_cache_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "orders_cache.json")


def _find_cached_order(query):
    import order_cache_store as ocs

    o = ocs.find_order(query, _orders_cache_path())
    if o:
        try:
            import km_api as _km

            _km.finalize_cache_order(o)
        except ImportError:
            pass
    return o


def _production_status_from_flow(flow):
    flow = flow or []
    done = [s for s in flow if s.get("done")]
    total = len(flow) or 1
    progress = int(len(done) * 100 / total)
    current = "-"
    for s in flow:
        if not s.get("done"):
            current = s.get("step") or s.get("name", "-")
            break
    else:
        if flow:
            current = flow[-1].get("step") or flow[-1].get("name", "已完成")
    status = "已完成" if progress >= 100 and flow else ("生产中" if done else "待处理")
    steps = [
        {
            "name": s.get("step") or s.get("name", ""),
            "done": bool(s.get("done")),
            "time": s.get("time", "-"),
        }
        for s in flow
    ]
    return {
        "status": status,
        "current_process": current,
        "diemold": "",
        "progress": progress,
        "steps": steps,
    }


def _order_detail_from_cache(o):
    items = o.get("items") or []
    first = items[0] if items else {}
    product = first.get("name", "") or first.get("spec", "") or "?"
    if len(items) > 1:
        product += f" 等{len(items)}种"
    qty = sum(int(i.get("qty", 0) or 0) for i in items)
    try:
        import km_api as _km
        amount = _km.km_to_float(o.get("total_amount"))
    except ImportError:
        amount = o.get("payment") or o.get("pay_amount") or o.get("total_fee") or 0
    province = (o.get("receiver_province") or "").strip()
    city = (o.get("receiver_city") or "").strip()
    return {
        "order_id": o.get("so_id") or o.get("km_sid", ""),
        "store": o.get("shop_name", ""),
        "customer": o.get("receiver_name", "") or o.get("buyer_nick", "") or "",
        "product": product,
        "qty": qty,
        "amount": amount,
        "order_date": _order_date_key(o) or (o.get("created") or o.get("pay_time") or "")[:10],
        "delivery_date": (o.get("plan_delivery_date") or o.get("consign_time") or "")[:10],
        "province": province,
        "city": city,
        "address": o.get("receiver_address", "") or "",
        "logistics": o.get("logistics_company", "") or o.get("express", "") or "",
    }


@app.route("/api/search_order")
def search_order():
    """扫码/输入单号：从 orders_cache.json 查订单（快麦+1688）。"""
    query = (request.args.get("q") or "").strip()
    if not query:
        return jsonify({"found": False, "message": "请输入单号"})
    o = _find_cached_order(query)
    if not o:
        return jsonify({"found": False, "message": "未找到该订单"})
    so_id = str(o.get("so_id") or o.get("km_sid") or "")
    extra = _order_extra.get(so_id, {})
    prod_list = _build_production_orders_for_cs()
    prod = next((x for x in prod_list if x.get("inner_id") == so_id), None)
    flow = (prod or {}).get("flow") or []
    detail = _order_detail_from_cache(o)
    payload = {
        "found": True,
        "order_id": so_id,
        "urgent": bool(extra.get("urgent")),
        "production_status": _production_status_from_flow(flow),
        "order_detail": detail,
        "jst_detail": detail,
    }
    return jsonify(payload)


def _sync_all_employees_perms():
    """
    核心：确保_permission_data['permissions'] 包含所有员工且每个功能字段齐全。
    同步规则：
    - 超级管理员: 所有功能 true
    - 已有记录: 保留已有值，补全缺失字段为 false
    - 新员工(无记录): 所有功能 false
    每次保存和API返回前都调用，永不遗漏任何人。
    """
    base = _permission_data.setdefault("permissions", {})
    for emp in _employees_master_list:
        name = emp["name"]
        if name == "戴雅利":
            # 超级管理员全开
            if name not in base:
                base[name] = {}
            for f in PERM_FEATURES:
                base[name][f] = True
        elif name not in base:
            # 新员工默认全false
            base[name] = {}
            for f in PERM_FEATURES:
                base[name][f] = False
        else:
            _migrate_perm_dict(base[name])
            # 已有员工：补全缺失字段
            for f in PERM_FEATURES:
                if f not in base[name]:
                    base[name][f] = False
    # 清理已删除员工的残留权限
    valid_names = {e["name"] for e in _employees_master_list}
    for name in list(base.keys()):
        if name not in valid_names:
            del base[name]
    _permission_data["permissions"] = base

_employees_master_list = list(_permission_data.get("employees", []))

# 初始化时执行一次
_sync_all_employees_perms()

_permission_data_init = {
    "processes": ["客服接单", "黄厂打印", "审单分单", "算料", "分纸", "啤机(自动平压平)", "啤机(机械手)", "手啤", "印刷", "开槽/打角", "清废", "打包发货"],
    "positions": ["超级管理员", "主管", "客服", "员工", "财务", "业务员"],
    "employees": [
        {"name": "戴雅利", "position": "超级管理员"},
        {"name": "邓涛", "position": "主管"},
        {"name": "黄兴", "position": "主管"},
        {"name": "覃海霞", "position": "主管"},
        {"name": "蒋义红", "position": "主管"},
        {"name": "沈齐豪", "position": "主管"},
        {"name": "苏世婷", "position": "客服"},
        {"name": "廖思美", "position": "客服"},
        {"name": "张慧平", "position": "客服"},
        {"name": "陈贝贝", "position": "客服"},
        {"name": "罗怡", "position": "客服"},
        {"name": "周井梅", "position": "客服"},
        {"name": "戴志美", "position": "客服"},
        {"name": "石梅清", "position": "客服"},
        {"name": "张文杰", "position": "客服"},
        {"name": "何水单", "position": "业务员"},
        {"name": "李四军", "position": "员工"},
        {"name": "陈贤聪", "position": "业务员"},
        {"name": "姚斌", "position": "员工"},
        {"name": "隆浪", "position": "财务"},
    ],
    "permissions": {}
}
# 确保默认结构与现有数据合并
for key, default_val in _permission_data_init.items():
    if key not in _permission_data:
        _permission_data[key] = default_val
# 再次同步权限
_sync_all_employees_perms()


# ==================== 报价系统 ====================

def load_quote_data():
    """从MySQL加载报价配置"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT config_key, config_value FROM quote_config")
        rows = cur.fetchall()
        cur.close()
        db.close()
        result = {}
        for r in rows:
            key = r['config_key']
            val = r['config_value']
            if isinstance(val, str):
                try:
                    val = json.loads(val)
                except:
                    pass
            result[key] = val
        if not result:
            return None
        return result
    except Exception as e:
        print(f'[MySQL load_quote_data] 错误: {e}')
        return None

def save_quote_data(qd):
    """保存报价配置到MySQL"""
    try:
        db = get_db()
        cur = db.cursor()
        for key, val in qd.items():
            cur.execute(
                "INSERT INTO quote_config (config_key, config_value) VALUES (%s,%s) ON DUPLICATE KEY UPDATE config_value=VALUES(config_value)",
                (key, json.dumps(val, ensure_ascii=False))
            )
        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_quote_data] 错误: {e}')
        return False

def ceil_to_half(val):
    """向上取整到0.5：23.32→23.5, 23.62→24"""
    import math
    return math.ceil(val * 2) / 2

def ceil_to_spec(val, spec_list, multiply_limit=3):
    """纸箱度数取整：如果val不在spec_list中，尝试×2/×3取奇数再除"""
    import math
    
    # 先检查val本身是否在规格表中或大于等于最小规格
    min_spec = min(spec_list)
    if val >= min_spec:
        for s in spec_list:
            if val <= s:
                return s
        # 超过最大规格，取最大
        return max(spec_list)
    
    # val小于最小规格，尝试×multiply
    for mult in range(2, multiply_limit + 1):
        multiplied = val * mult
        # 取大于multiplied的奇数（纸箱度数都是奇数）
        ceil_odd = math.ceil(multiplied)
        if ceil_odd % 2 == 0:
            ceil_odd += 1
        # 检查这个奇数是否能在spec_list中找到
        if ceil_odd >= min_spec:
            for s in spec_list:
                if ceil_odd <= s:
                    return s / mult
            return max(spec_list) / mult
    
    # 保底
    return min_spec / multiply_limit

def calc_airbox(length, width, height, qty, material_key, quote_data, is_buckle=False):
    """标准飞机盒/带扣飞机盒报价"""
    import math
    
    # 纸长/纸度公式
    if is_buckle:
        paper_l_cm = 3 * height + 2 * width + 3
    else:
        paper_l_cm = 3 * height + 2 * width
    
    paper_w_cm = length + 4 * height + 4
    
    # 取整到0.5
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价
    mat_config = quote_data.get("materials", {}).get("airbox", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.35)
    mat_name = mat_info.get("name", material_key)
    
    # 材料成本 = 纸长×纸度×单价/1000
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    
    # 零售价 = 材料成本 × 2
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult
    
    # 批量建议
    suggestions = quote_data.get("profit", {}).get("suggestions", {}).get("airbox", [])
    suggested_rate = None
    for s in sorted(suggestions, key=lambda x: -x["qty"]):
        if qty >= s["qty"]:
            suggested_rate = s["multiplier"]
            break

    # ====== 重量计算 ======
    # 展开面积(平方英寸) = 纸长×纸度
    # 转平方米: 1平方英寸 = 0.00064516 平方米
    gram_weight = mat_info.get("gram_weight", 450)  # 克/平方米
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "带扣飞机盒" if is_buckle else "标准飞机盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": suggested_rate,
        "suggested_price": round(material_cost_per_unit * suggested_rate, 4) if suggested_rate else None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_koudi(length, width, height, qty, material_key, quote_data):
    """扣底盒报价"""
    import math
    
    # 纸长 = (长×2 + 宽×2 + 3) / 2.54
    paper_l_cm = length * 2 + width * 2 + 3
    # 纸度 = (宽 + 高 + 1/2宽 + 3 + 2) / 2.54
    paper_w_cm = width + height + width / 2 + 3 + 2
    
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价（扣底盒体系）
    mat_config = quote_data.get("materials", {}).get("koudi", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.40)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 450)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "扣底盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": None,
        "suggested_price": None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_shuangcha(length, width, height, qty, material_key, quote_data):
    """双插盒报价"""
    import math
    
    # 纸长 = (长×2 + 宽×2 + 3) / 2.54
    paper_l_cm = length * 2 + width * 2 + 3
    # 纸度 = (宽×2 + 6 + 高) / 2.54
    paper_w_cm = width * 2 + 6 + height
    
    paper_l_inch = ceil_to_half(paper_l_cm / 2.54)
    paper_w_inch = ceil_to_half(paper_w_cm / 2.54)
    
    # 获取材料单价（扣底盒体系，跟扣底盒用同一套）
    mat_config = quote_data.get("materials", {}).get("koudi", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.40)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 450)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "双插盒",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "suggested_multiplier": None,
        "suggested_price": None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

def calc_carton(length, width, height, qty, material_key, quote_data):
    """纸箱报价"""
    import math
    
    specs = quote_data.get("cardboard_specs", {})
    paper_lengths = specs.get("danbu_paper_lengths", [29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90])
    paper_widths = specs.get("danbu_paper_widths", [29,31,33,35,37,39,41,43,45,47,49])
    max_danbu_cm = specs.get("max_danbu_length_cm", 230)
    
    # 判断单卜还是双卜
    danbu_paper_l_cm = (length + width) * 2 + 3.5
    is_double = danbu_paper_l_cm > max_danbu_cm
    
    if is_double:
        # 双卜
        paper_l_cm = ((length + width) + 3.5) * 2
    else:
        # 单卜
        paper_l_cm = danbu_paper_l_cm
    
    paper_w_cm = width + height + 0.6
    
    # 纸长取整到度数表
    paper_l_raw = paper_l_cm / 2.54
    paper_l_inch = ceil_to_spec(paper_l_raw, paper_lengths)
    
    # 纸度取整到度数表
    paper_w_raw = paper_w_cm / 2.54
    paper_w_inch = ceil_to_spec(paper_w_raw, paper_widths)
    
    # 获取材料单价（纸箱体系）
    mat_config = quote_data.get("materials", {}).get("carton", {}).get("materials", {})
    mat_info = mat_config.get(material_key, {})
    unit_price = mat_info.get("price", 1.30)
    mat_name = mat_info.get("name", material_key)
    
    material_cost_per_unit = paper_l_inch * paper_w_inch * unit_price / 1000
    retail_mult = quote_data.get("profit", {}).get("retail_multiplier", 2.0)
    sell_price_per_unit = material_cost_per_unit * retail_mult

    # ====== 重量计算 ======
    gram_weight = mat_info.get("gram_weight", 380)
    area_m2 = paper_l_inch * paper_w_inch * 0.00064516
    weight_per_unit_g = area_m2 * gram_weight
    weight_per_unit_kg = weight_per_unit_g / 1000

    # ====== 快递费（首重+续重） ======
    freight_config = quote_data.get("freight", {})
    total_weight_kg = weight_per_unit_kg * qty

    def calc_freight(total_kg, cfg):
        first = cfg.get("first_kg", 10)
        renew = cfg.get("renew_per_kg", 0.9)
        if total_kg <= 1:
            return round(first, 2)
        else:
            extra = total_kg - 1
            return round(first + extra * renew, 2)

    freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
    freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
    freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))

    freight_per_unit_gd = round(freight_total_gd / qty, 4)
    freight_per_unit_jzh = round(freight_total_jzh / qty, 4)
    freight_per_unit_other = round(freight_total_other / qty, 4)

    return {
        "product": "双卜纸箱" if is_double else "单卜纸箱",
        "material_key": material_key,
        "material_name": mat_name,
        "unit_price": unit_price,
        "gram_weight": gram_weight,
        "paper_l_cm": round(paper_l_cm, 1),
        "paper_w_cm": round(paper_w_cm, 1),
        "paper_l_inch": round(paper_l_inch, 2),
        "paper_w_inch": round(paper_w_inch, 2),
        "material_cost_per_unit": round(material_cost_per_unit, 4),
        "retail_multiplier": retail_mult,
        "sell_price_per_unit": round(sell_price_per_unit, 4),
        "total_cost": round(material_cost_per_unit * qty, 2),
        "total_price": round(sell_price_per_unit * qty, 2),
        "qty": qty,
        "is_double_buckle": is_double,
        "suggested_multiplier": None,
        "suggested_price": None,
        "weight_per_unit_g": round(weight_per_unit_g, 2),
        "weight_per_unit_kg": round(weight_per_unit_kg, 4),
        "weight_total_kg": round(total_weight_kg, 2),
        "freight_per_unit_gd": freight_per_unit_gd,
        "freight_per_unit_jzh": freight_per_unit_jzh,
        "freight_per_unit_other": freight_per_unit_other,
        "freight_total_gd": freight_total_gd,
        "freight_total_jzh": freight_total_jzh,
        "freight_total_other": freight_total_other
    }

@app.route('/api/quote_data')
def get_quote_data():
    """返回报价基础数据（保持JSON key顺序）"""
    try:
        with open(QUOTE_DATA_FILE, 'r', encoding='utf-8') as f:
            raw = f.read()
        return '{"success":true,"data":' + raw + '}', 200, {'Content-Type': 'application/json'}
    except:
        return jsonify({"success": False, "error": "报价数据未加载"})

@app.route('/api/quote/save_config', methods=['POST'])
def save_quote_config():
    """保存报价配置（仅权限管理角色可操作）"""
    try:
        # 检查权限
        if 'username' not in session:
            return jsonify({"success": False, "error": "未登录"})
        user = USERS.get(session['username'])
        if not user:
            return jsonify({"success": False, "error": "用户不存在"})
        name = user['employee_name']
        _sync_all_employees_perms()
        my_perm = _permission_data.get("permissions", {}).get(name, {})
        if not my_perm.get('权限管理', False):
            return jsonify({"success": False, "error": "无权限修改报价配置"})
        
        new_data = request.get_json()
        if not new_data:
            return jsonify({"success": False, "error": "数据为空"})
        
        if save_quote_data(new_data):
            return jsonify({"success": True, "message": "报价配置已保存"})
        else:
            return jsonify({"success": False, "error": "保存失败"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/quote/calculate', methods=['POST'])
def calculate_quote():
    """计算报价"""
    try:
        data = request.get_json()
        ptype = data.get('type', 'zhengsquare')  # zhengsquare, juxing, daikou, koudi, shuangcha, qita
        length = float(data.get('length', 0))
        width = float(data.get('width', 0))
        height = float(data.get('height', 0))
        qty = int(data.get('qty', 1000))
        material = data.get('material', 'd6d')
        discount = float(data.get('discount', 100)) / 100.0  # 折扣百分比转小数

        if length <= 0 or width <= 0 or height <= 0:
            return jsonify({"error": "请输入正确的尺寸", "success": False})
        
        import math

        qd = load_quote_data()
        if not qd:
            return jsonify({"error": "报价数据未加载", "success": False})

        # 兼容 zhengsquare-outer / zhengsquare-inner
        # 内径转外径：长+1.5cm，宽+0.5cm，高+0.5cm
        if ptype == 'zhengsquare-inner':
            length = length + 1.5
            width = width + 0.5
            height = height + 0.5
        
        base_type = ptype.replace('-outer', '').replace('-inner', '')
        if base_type == 'zhengsquare':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=False)
            # 标注外径/内径
            if ptype == 'zhengsquare-inner':
                detail['dimension_label'] = '内径转外径'
            else:
                detail['dimension_label'] = '外径'
        elif base_type == 'juxing':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=False)
        elif ptype == 'daikou':
            detail = calc_airbox(length, width, height, qty, material, qd, is_buckle=True)
        elif ptype == 'koudi':
            detail = calc_koudi(length, width, height, qty, material, qd)
        elif ptype == 'shuangcha':
            detail = calc_shuangcha(length, width, height, qty, material, qd)
        elif ptype == 'qita':
            detail = calc_carton(length, width, height, qty, material, qd)
        elif ptype == 'pe':
            # 珍珠棉报价 - 输入单位毫米
            length_m = length / 1000  # 毫米转米
            width_m = width / 1000    # 毫米转米
            
            # 厚度向上取整到0.5mm的倍数（0.5阶梯）
            # 0~0.5→0.5, 0.6~1.0→1.0, 1.1~1.5→1.5, 1.6~2.0→2.0
            height_raw = height  # 保留原始值用于显示
            height = math.ceil(height / 0.5) * 0.5
            if height == 0:
                height = 0.5  # 最小0.5
            
            pe_price = qd.get("materials", {}).get("pe", {}).get("price", 0.3)
            material_cost_per_unit = length_m * width_m * height * pe_price
            
            # 客户类型
            customer_type = data.get('customer_type', 'guangdong_wholesale')
            customer_types = qd.get("materials", {}).get("pe", {}).get("customer_types", {})
            ct_info = customer_types.get(customer_type, {})
            multiplier = ct_info.get("multiplier", 2.0)
            ct_name = ct_info.get("name", customer_type)
            
            sell_price_per_unit = material_cost_per_unit * multiplier
            retail_mult = qd.get("profit", {}).get("retail_multiplier", 2.0)
            
            # 总货款判断
            total_price = sell_price_per_unit * qty
            total_price_discounted = total_price * discount
            price_too_low = total_price_discounted < 100  # 总货款低于100元
            unit_price_too_low = sell_price_per_unit < 0.3  # 单价低于0.3元
            
            # 重量预估（泡比5000）和快递费
            height_cm = height / 10  # 毫米转厘米
            length_cm = length / 10  # 毫米转厘米
            width_cm = width / 10    # 毫米转厘米
            weight_per_unit_kg = height_cm * length_cm * width_cm / 5000
            
            # 快递费（首重+续重）
            freight_config = qd.get("materials", {}).get("pe", {}).get("freight", {})
            
            def calc_freight(total_kg, cfg):
                first = cfg.get("first_kg", 10)
                renew = cfg.get("renew_per_kg", 0.9)
                if total_kg <= 1:
                    return round(first, 2)
                else:
                    extra = total_kg - 1
                    return round(first + extra * renew, 2)
            
            freight_per_unit_gd = round(calc_freight(weight_per_unit_kg, freight_config.get("guangdong", {})), 4)
            freight_per_unit_jzh = round(calc_freight(weight_per_unit_kg, freight_config.get("jiangzhehu", {})), 4)
            freight_per_unit_other = round(calc_freight(weight_per_unit_kg, freight_config.get("other", {})), 4)
            
            total_weight_kg = weight_per_unit_kg * qty
            freight_total_gd = calc_freight(total_weight_kg, freight_config.get("guangdong", {}))
            freight_total_jzh = calc_freight(total_weight_kg, freight_config.get("jiangzhehu", {}))
            freight_total_other = calc_freight(total_weight_kg, freight_config.get("other", {}))
            
            detail = {
                "product": "珍珠棉片材",
                "material_name": "珍珠棉",
                "unit_price": pe_price,
                "length_m": round(length_m, 4),
                "width_m": round(width_m, 4),
                "thickness_mm": height,
                "thickness_input_mm": height_raw,
                "customer_type": ct_name,
                "multiplier": multiplier,
                "material_cost_per_unit": round(material_cost_per_unit, 4),
                "retail_multiplier": retail_mult,
                "sell_price_per_unit": round(sell_price_per_unit, 4),
                "total_cost": round(material_cost_per_unit * qty, 2),
                "total_price": round(total_price, 2),
                "total_price_discounted": round(total_price_discounted, 2),
                "sell_price_per_unit_discounted": round(sell_price_per_unit * discount, 4),
                "qty": qty,
                "weight_per_unit_kg": round(weight_per_unit_kg, 4),
                "weight_total_kg": round(total_weight_kg, 2),
                "freight_per_unit_gd": freight_per_unit_gd,
                "freight_per_unit_jzh": freight_per_unit_jzh,
                "freight_per_unit_other": freight_per_unit_other,
                "freight_total_gd": freight_total_gd,
                "freight_total_jzh": freight_total_jzh,
                "freight_total_other": freight_total_other,
                "discount": round(discount * 100, 1),
                "suggested_multiplier": None,
                "suggested_price": None,
                "pe_warning_price_too_low": price_too_low,
                "pe_warning_unit_price_too_low": unit_price_too_low
            }
        else:
            return jsonify({"error": "未知的产品类型", "success": False})

        # 应用折扣
        detail["discount"] = round(discount * 100, 1)
        detail["sell_price_per_unit_discounted"] = round(detail["sell_price_per_unit"] * discount, 4)
        detail["total_price_discounted"] = round(detail["total_price"] * discount, 2)

        return jsonify({"success": True, "detail": detail})

    except Exception as e:
        return jsonify({"error": str(e), "success": False})

# ==================== 固定刀模库 API ====================

def load_dimoldb():
    """从MySQL加载刀模库数据"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT id, product_type, name, length, width, height, remark, stock, created_at FROM dimoldb ORDER BY created_at DESC")
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'product_type': r['product_type'] or '',
                'name': r['name'] or '',
                'length': float(r['length']) if r['length'] else 0,
                'width': float(r['width']) if r['width'] else 0,
                'height': float(r['height']) if r['height'] else 0,
                'remark': r['remark'] or '',
                'stock': r['stock'] or 0,
                'created_at': r['created_at'] or ''
            })
        cur.close()
        db.close()
        return result
    except Exception as e:
        print(f'[MySQL load_dimoldb] 错误: {e}')
        return []

def save_dimoldb(data):
    """保存刀模库数据到MySQL（truncate+批量insert）"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE dimoldb")
        if data:
            for item in data:
                cur.execute(
                    "INSERT INTO dimoldb (id, product_type, name, length, width, height, remark, stock, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (
                        item.get('id', ''),
                        item.get('product_type', ''),
                        item.get('name', ''),
                        item.get('length', 0),
                        item.get('width', 0),
                        item.get('height', 0),
                        item.get('remark', ''),
                        item.get('stock', 0),
                        item.get('created_at', '')
                    )
                )
        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_dimoldb] 错误: {e}')
        return False

def _has_dimoldb_edit_perm():
    """检查是否有刀模库编辑权限"""
    if 'username' not in session:
        return False
    user = USERS.get(session['username'])
    if not user:
        return False
    if user['role'] in ('超级管理员', '管理'):
        return True
    # 检查功能权限
    _sync_all_employees_perms()
    name = user['employee_name']
    my_perm = _permission_data.get("permissions", {}).get(name, {})
    if my_perm.get('刀模管理', False):
        return True
    return False


def _dimoldb_infer_inner_outer(dm):
    """内外径：优先 dim_type；为空时用 name 中 (内)/内径/(外)/外径 及 remark 中的 内/外 字推断。返回 inner|outer|''。"""
    dt = (dm.get('dim_type') or '').strip()
    if dt == 'inner':
        return 'inner'
    if dt == 'outer':
        return 'outer'
    name = dm.get('name') or ''
    rk = dm.get('remark') or ''
    if '(内)' in name or '内径' in name or '内' in rk:
        return 'inner'
    if '(外)' in name or '外径' in name or '外' in rk:
        return 'outer'
    return ''


@app.route('/api/dimoldb', methods=['GET'])
def get_dimoldb():
    """获取刀模列表（支持分页）"""
    data = load_dimoldb()
    # 支持筛选
    ptype = request.args.get('type', '')
    search = request.args.get('search', '').strip()
    dim_type = request.args.get('dim_type', '').strip()
    if ptype:
        # 兼容 zhengsquare-outer / zhengsquare-inner（dim_type 为空时 name/remark 推断）
        if ptype == 'zhengsquare-outer':
            data = [d for d in data if d.get('product_type') == 'zhengsquare' and _dimoldb_infer_inner_outer(d) == 'outer']
        elif ptype == 'zhengsquare-inner':
            data = [d for d in data if d.get('product_type') == 'zhengsquare' and _dimoldb_infer_inner_outer(d) == 'inner']
        else:
            data = [d for d in data if d.get('product_type') == ptype]
    if dim_type and not ptype.startswith('zhengsquare-'):
        data = [d for d in data if _dimoldb_infer_inner_outer(d) == dim_type]
    if search:
        # 刀模搜索统一用正则提取数字，兼容所有分隔符（x/*/×/空格...）
        nums = re.findall(r'\d+\.?\d*', search.replace(' ', '*'))
        nums = [n for n in nums if n.strip()]
        if len(nums) >= 3:
            try:
                sl = float(nums[0].strip())
                sw = float(nums[1].strip())
                sh = float(nums[2].strip())
                data = [d for d in data if 
                    d.get('length') is not None and abs(d['length'] - sl) < 0.1 and
                    d.get('width') is not None and abs(d['width'] - sw) < 0.1 and
                    d.get('height') is not None and abs(d['height'] - sh) < 0.1]
            except:
                pass
        else:
            data = [d for d in data if search in d.get('name', '') or f"{d.get('length',0)}x{d.get('width',0)}x{d.get('height',0)}" in search or f"{d.get('length',0)}×{d.get('width',0)}×{d.get('height',0)}" in search]
    # 分页
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
    except ValueError:
        page, page_size = 1, 50
    if page < 1: page = 1
    if page_size < 1: page_size = 50
    if page_size > 200: page_size = 200
    total = len(data)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    end = start + page_size
    page_data = data[start:end] if start < total else []
    # 给每条刀模补充真实库存数（从inventory.json统计）
    inv_all = load_inventory()
    inv_items = inv_all.get('finished', inv_all if isinstance(inv_all, list) else [])
    def calc_stock(dm):
        """根据刀模的尺寸+dim_type统计库存总量"""
        dm_l = dm.get('length')
        dm_w = dm.get('width')
        dm_h = dm.get('height')
        dm_dt = (dm.get('dim_type') or '').strip()
        # 刀模 dim_type 为空：name / remark 推断内外径
        if not dm_dt:
            dm_dt = _dimoldb_infer_inner_outer(dm)
        dm_type = dm.get('product_type', '')
        # 刀模类型→库存类型映射：库存只有 changfang 和 zhengsquare
        _type_map = {
            'zhengsquare': 'zhengsquare',
            'juxing': 'changfang',
            'daikou': 'changfang',
            'koudi': 'changfang',
            'shuangcha': 'changfang',
            'qita': 'changfang',
        }
        inv_type_map = _type_map.get(dm_type, dm_type)
        total = 0
        for iv in inv_items:
            iv_l = iv.get('length')
            iv_w = iv.get('width')
            iv_h = iv.get('height')
            if not (iv_l and iv_w and iv_h and dm_l and dm_w and dm_h):
                continue
            if abs(float(iv_l) - float(dm_l)) >= 0.1: continue
            if abs(float(iv_w) - float(dm_w)) >= 0.1: continue
            if abs(float(iv_h) - float(dm_h)) >= 0.1: continue
            # 匹配dim_type（内外径要分开）
            iv_dt = iv.get('dim_type', '')
            if dm_dt and iv_dt and dm_dt != iv_dt:
                continue
            # 匹配product_type：刀模类型→库存类型映射
            iv_type = iv.get('product_type', '')
            if iv_type and iv_type != inv_type_map:
                continue
            total += int(iv.get('qty') or iv.get('stock') or 0)
        return total
    for d in page_data:
        d['stock'] = calc_stock(d)
    return jsonify({
        "success": True,
        "data": page_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "can_edit": _has_dimoldb_edit_perm()
    })

@app.route('/api/dimoldb/single', methods=['GET'])
def get_dimoldb_single():
    """根据ID获取单个刀模"""
    dm_id = request.args.get('id', '').strip()
    if not dm_id:
        return jsonify({"success": False, "error": "缺少id参数"})
    data = load_dimoldb()
    for d in data:
        if d.get('id') == dm_id:
            return jsonify({"success": True, "data": d})
    return jsonify({"success": False, "error": "未找到"})

@app.route('/api/dimoldb', methods=['POST'])
def add_dimoldb():
    """新增刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        required = ['product_type', 'name', 'length', 'width', 'height']
        for k in required:
            if k not in item:
                return jsonify({"success": False, "error": f"缺少字段: {k}"})
        data = load_dimoldb()
        item['id'] = f"dm_{int(time.time())}_{len(data)}"
        item['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        data.append(item)
        save_dimoldb(data)
        return jsonify({"success": True, "message": "刀模已添加"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/dimoldb/<dm_id>', methods=['PUT'])
def update_dimoldb(dm_id):
    """修改刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        data = load_dimoldb()
        for d in data:
            if d.get('id') == dm_id:
                d.update(item)
                d['id'] = dm_id
                save_dimoldb(data)
                return jsonify({"success": True, "message": "刀模已更新"})
        return jsonify({"success": False, "error": "未找到该刀模"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/dimoldb/<dm_id>', methods=['DELETE'])
def delete_dimoldb(dm_id):
    """删除刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    data = load_dimoldb()
    new_data = [d for d in data if d.get('id') != dm_id]
    if len(new_data) == len(data):
        return jsonify({"success": False, "error": "未找到该刀模"})
    save_dimoldb(new_data)
    return jsonify({"success": True, "message": "刀模已删除"})

@app.route('/api/dimoldb/export', methods=['GET'])
def export_dimoldb():
    """导出所有刀模为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        db = load_dimoldb()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '刀模库'
        headers = ['序号', '产品类型', '名称', '编码', '备注', '长(cm)', '宽(cm)', '高(cm)', '创建时间']
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin
        for ri, item in enumerate(db, 2):
            vals = [
                ri - 1,
                item.get('product_type', ''),
                item.get('name', ''),
                item.get('code', '') or '',
                item.get('remark', '') or '',
                item.get('length', ''),
                item.get('width', ''),
                item.get('height', ''),
                item.get('created_at', '')
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 18
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        import flask
        resp = flask.send_file(tmp.name, as_attachment=True, download_name=f'刀模库_{datetime.date.today().isoformat()}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return resp
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()}), 500

@app.route('/api/dimoldb/import', methods=['POST'])
def import_dimoldb():
    """上传Excel导入刀模"""
    if not _has_dimoldb_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        import openpyxl
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "请上传Excel文件"})
        f = request.files['file']
        if f.filename == '':
            return jsonify({"success": False, "error": "未选择文件"})
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        f.save(tmp.name)
        tmp.close()
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active
        # 探测表头
        header_row = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            vals = [str(v or '').strip() for v in row]
            if '名称' in vals or '刀模名称' in vals or 'name' in vals:
                header_row = r_idx
                break
        if header_row is None:
            return jsonify({"success": False, "error": "未找到表头行（需包含'名称'列）"})
        headers = [str(v or '').strip() for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col_map = {}
        for i, h in enumerate(headers):
            if '名称' in h or h == 'name':
                col_map['name'] = i
            elif '类型' in h or '产品' in h or 'product' in h.lower():
                col_map['product_type'] = i
            elif '编码' in h or 'code' in h.lower():
                col_map['code'] = i
            elif '备注' in h or 'remark' in h.lower():
                col_map['remark'] = i
            elif '长' in h or 'length' in h.lower():
                col_map['length'] = i
            elif '宽' in h or 'width' in h.lower():
                col_map['width'] = i
            elif '高' in h or 'height' in h.lower():
                col_map['height'] = i
        if 'name' not in col_map:
            return jsonify({"success": False, "error": f"未找到'名称'列，表头: {headers}"})
        db = load_dimoldb()
        added = 0
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[col_map['name']]).strip() if col_map['name'] < len(row) and row[col_map['name']] else ''
            if not name or name.startswith('=') or name == 'None':
                continue
            item = {
                'id': f'dm_{int(time.time())}_{added}_{len(db)}',
                'name': name,
                'product_type': str(row[col_map['product_type']]).strip() if col_map.get('product_type') is not None and col_map['product_type'] < len(row) and row[col_map['product_type']] else 'zhengsquare',
                'code': str(row[col_map['code']]).strip() if col_map.get('code') is not None and col_map['code'] < len(row) and row[col_map['code']] else '',
                'remark': str(row[col_map['remark']]).strip() if col_map.get('remark') is not None and col_map['remark'] < len(row) and row[col_map['remark']] else '',
                'length': float(row[col_map['length']]) if col_map.get('length') is not None and col_map['length'] < len(row) and row[col_map['length']] is not None else 0,
                'width': float(row[col_map['width']]) if col_map.get('width') is not None and col_map['width'] < len(row) and row[col_map['width']] is not None else 0,
                'height': float(row[col_map['height']]) if col_map.get('height') is not None and col_map['height'] < len(row) and row[col_map['height']] is not None else 0,
                'created_at': now
            }
            try:
                item['length'] = float(item['length'])
            except: item['length'] = 0
            try:
                item['width'] = float(item['width'])
            except: item['width'] = 0
            try:
                item['height'] = float(item['height'])
            except: item['height'] = 0
            db.append(item)
            added += 1
        save_dimoldb(db)
        import os
        try: os.unlink(tmp.name)
        except: pass
        return jsonify({"success": True, "message": f"导入完成，新增 {added} 条刀模记录", "added": added})
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()})

# -------------------- 库存导出 --------------------
@app.route('/api/inventory/export', methods=['GET'])
def export_inventory():
    """导出库存为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        tab = request.args.get('tab', 'finished')
        data = load_inventory()
        items = data.get(tab, [])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '库存'
        headers = ['序号', '名称', '产品类型', '内/外径', '材质', '库存', '长(cm)', '宽(cm)', '高(cm)', '货位', '备注', '更新时间']
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin
        for ri, item in enumerate(items, 2):
            dim_label = {'inner': '内径', 'outer': '外径'}.get(item.get('dim_type', ''), '')
            vals = [
                ri - 1,
                item.get('name', ''),
                item.get('product_type', ''),
                dim_label,
                item.get('material', ''),
                int(item.get('stock', item.get('qty', 0))),
                item.get('length', ''),
                item.get('width', ''),
                item.get('height', ''),
                item.get('location', '') or '',
                item.get('remark', '') or '',
                item.get('updated_at', item.get('created_at', ''))
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
        for col_letter, width in zip('ABCDEFGHIJKL', [6, 30, 12, 10, 12, 8, 10, 10, 10, 12, 20, 18]):
            ws.column_dimensions[col_letter].width = width
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        tab_label = {'finished': '成品', 'returned': '退货'}.get(tab, tab)
        import flask
        resp = flask.send_file(tmp.name, as_attachment=True, download_name=f'{tab_label}库存_{datetime.date.today().isoformat()}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return resp
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()}), 500

@app.route('/api/dimoldb/search', methods=['POST'])
def search_dimoldb():
    """查询某尺寸是否有固定刀模（供报价系统调用）"""
    try:
        data = request.get_json()
        ptype = data.get('type', '')
        length = data.get('length')
        width = data.get('width')
        height = data.get('height')
        dim_type = data.get('dim_type', '')
        db = load_dimoldb()
        matches = db
        if ptype:
            # 兼容 zhengsquare-outer / zhengsquare-inner
            actual_type = ptype.replace('-outer', '').replace('-inner', '')
            matches = [d for d in matches if d.get('product_type') == actual_type]
            # 如果ptype带内外径但dim_type没传，自动设置
            if 'outer' in ptype and not data.get('dim_type'):
                dim_type = 'outer'
            elif 'inner' in ptype and not data.get('dim_type'):
                dim_type = 'inner'
        if length is not None:
            lv = float(length)
            matches = [d for d in matches if abs(float(d.get('length', 0)) - lv) < 0.1]
        if width is not None:
            wv = float(width)
            matches = [d for d in matches if abs(float(d.get('width', 0)) - wv) < 0.1]
        if height is not None:
            hv = float(height)
            matches = [d for d in matches if abs(float(d.get('height', 0)) - hv) < 0.1]
        # 内外径过滤：刀模数据dim_type字段为空，需要用name里的(内)/(外)来判断
        # 但juxing/koudi/shuangcha等类型没有内外径之分，忽略dim_type过滤
        ptype_for_dim = data.get('type', '')
        if dim_type and ptype_for_dim == 'zhengsquare':
            if dim_type == 'inner':
                matches = [d for d in matches if _dimoldb_infer_inner_outer(d) == 'inner']
            elif dim_type == 'outer':
                matches = [d for d in matches if _dimoldb_infer_inner_outer(d) == 'outer']
        elif dim_type and ptype_for_dim not in ('zhengsquare', '', None):
            # 非正方形类型：若存在内外径信息则过滤
            has_inner_outer = any(_dimoldb_infer_inner_outer(d) for d in matches)
            if has_inner_outer and dim_type:
                if dim_type == 'inner':
                    matches = [d for d in matches if _dimoldb_infer_inner_outer(d) == 'inner']
                elif dim_type == 'outer':
                    matches = [d for d in matches if _dimoldb_infer_inner_outer(d) == 'outer']
        return jsonify({"success": True, "matches": matches})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ==================== 库存 API ====================

def load_inventory():
    """从MySQL加载库存数据"""
    try:
        db = get_db()
        cur = db.cursor()
        # 库存表所有数据作为finished (兼容旧API: 前端主要用finished)
        cur.execute("SELECT id, name, spec, product_type, material, location, qty, last_month_qty, length, width, height, dim_type, created_at, updated_at FROM inventory ORDER BY created_at DESC")
        rows = cur.fetchall()
        finished = []
        for r in rows:
            finished.append({
                'id': r['id'],
                'name': r['name'] or '',
                'spec': r['spec'] or '',
                'product_type': r['product_type'] or '',
                'material': r['material'] or '',
                'location': r['location'] or '',
                'qty': r['qty'] or 0,
                'last_month_qty': r['last_month_qty'] or 0,
                'length': float(r['length']) if r['length'] else 0,
                'width': float(r['width']) if r['width'] else 0,
                'height': float(r['height']) if r['height'] else 0,
                'dim_type': r['dim_type'] or '',
                'created_at': r['created_at'] or '',
                'updated_at': r['updated_at'] or ''
            })
        cur.close()
        db.close()
        return {"finished": finished, "raw": [], "returned": []}
    except Exception as e:
        print(f'[MySQL load_inventory] 错误: {e}')
        return {"finished": [], "raw": [], "returned": []}

def save_inventory(data):
    """保存库存数据到MySQL（truncate+批量insert）"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE inventory")
        # 取 finished、raw、returned 的统一列表
        items = []
        if isinstance(data, dict):
            items = data.get('finished', data.get('raw', data.get('returned', [])))
        elif isinstance(data, list):
            items = data
        for item in items:
            cur.execute(
                "INSERT INTO inventory (id, name, spec, product_type, material, location, qty, last_month_qty, length, width, height, dim_type, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    item.get('id', ''),
                    item.get('name', ''),
                    item.get('spec', ''),
                    item.get('product_type', ''),
                    item.get('material', ''),
                    item.get('location', ''),
                    item.get('qty', 0),
                    item.get('last_month_qty', 0),
                    item.get('length', 0),
                    item.get('width', 0),
                    item.get('height', 0),
                    item.get('dim_type', ''),
                    item.get('created_at', ''),
                    item.get('updated_at', '')
                )
            )
        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_inventory] 错误: {e}')
        return False

def _has_inv_edit_perm():
    """库存修改权限：超级管理员/管理/有'库存'权限"""
    if 'username' not in session:
        return False
    user = USERS.get(session['username'])
    if not user:
        return False
    if user['role'] in ('超级管理员', '管理'):
        return True
    _sync_all_employees_perms()
    name = user['employee_name']
    my_perm = _permission_data.get("permissions", {}).get(name, {})
    if my_perm.get('库存', False):
        return True
    return False

@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    """获取库存列表（支持分页，同刀模库）"""
    tab = request.args.get('tab', 'finished')
    data = load_inventory()
    items = data.get(tab, [])
    
    # 材质映射函数：库存中的材质名 → 显示名
    def map_material(m):
        mapping = {
            '国产纸': '特硬D6D',
            '双白': '白色W7W',
            '台湾纸': '台湾纸',
            '黑色': '黑色',
            '红色': '红色',
            '差材料': '差材料'
        }
        return mapping.get(m.strip(), m.strip()) if m else ''
    
    # 给所有item做材质映射
    for item in items:
        if item.get('material'):
            item['material'] = map_material(item['material'])
        # name 中也替换材质名
        n = item.get('name', '')
        for old, new in [('国产纸','特硬D6D'), ('双白','白色W7W')]:
            if old in n:
                item['name'] = n.replace(old, new)
                n = item['name']
    
    ptype = request.args.get('type', '')
    search = request.args.get('search', '').strip()
    length = request.args.get('length', '').strip()
    width = request.args.get('width', '').strip()
    height = request.args.get('height', '').strip()
    if ptype == 'zhengsquare-outer':
        items = [d for d in items if d.get('product_type') == 'zhengsquare' and d.get('dim_type') == 'outer']
    elif ptype == 'zhengsquare-inner':
        items = [d for d in items if d.get('product_type') == 'zhengsquare' and d.get('dim_type') == 'inner']
    elif ptype:
        # 库存只有 changfang 和 zhengsquare，刀模type需要映射
        _type_to_inv = {'juxing':'changfang', 'daikou':'changfang', 'koudi':'changfang', 'shuangcha':'changfang', 'qita':'changfang', 'changfang':'changfang', 'zhengsquare':'zhengsquare'}
        inv_type = _type_to_inv.get(ptype, ptype)
        items = [d for d in items if d.get('product_type') == inv_type]
    if search:
        # 【通用提取】提取输入中的前三个数字，不管分隔符是什么
        # 支持: 10x10x2, 10*10*2, 10×10×2, 10 10 2, 10,10,2, 10-10-2, 10.10.2 等全部格式
        nums = re.findall(r'\d+\.?\d*', search.replace(' ', '*'))
        nums = [n for n in nums if n.strip()]
        if len(nums) >= 3:
            try:
                sl = float(nums[0].strip())
                sw = float(nums[1].strip())
                sh = float(nums[2].strip())
                items = [d for d in items if 
                    d.get('length') is not None and abs(d['length'] - sl) < 0.1 and
                    d.get('width') is not None and abs(d['width'] - sw) < 0.1 and
                    d.get('height') is not None and abs(d['height'] - sh) < 0.1]
            except:
                pass
        elif search.replace('.','').replace('-','').isdigit() or search.strip().replace('.','').replace('-','').replace(',','').replace('|','').lstrip('-').isdigit():
            # 纯数字搜索
            sv = float(search.strip().lstrip('-').replace(',','.'))
            items = [d for d in items if 
                (d.get('length') is not None and abs(d['length'] - sv) < 0.1) or
                (d.get('width') is not None and abs(d['width'] - sv) < 0.1) or
                (d.get('height') is not None and abs(d['height'] - sv) < 0.1)]
        else:
            search_field = request.args.get('search_field', 'all')
            if search_field == 'name':
                items = [d for d in items if search in d.get('name', '')]
            elif search_field == 'material':
                items = [d for d in items if search in d.get('material', '')]
            else:
                # 通用搜索
                items = [d for d in items if 
                    search in d.get('name', '') or
                    search in d.get('material', '')]
    if length:
        lv = float(length)
        items = [d for d in items if d.get('length') is not None and abs(float(d['length']) - lv) < 0.1]
    if width:
        wv = float(width)
        items = [d for d in items if d.get('width') is not None and abs(float(d['width']) - wv) < 0.1]
    if height:
        hv = float(height)
        items = [d for d in items if d.get('height') is not None and abs(float(d['height']) - hv) < 0.1]
    # 按尺寸从小到大排序
    def sort_key(item):
        try:
            dims = item.get('name', '').replace('×','*').replace('x','*').replace('X','*').split('*')
            l, w, h = float(dims[0]), float(dims[1]), float(dims[2])
            return (l, w, h)
        except:
            return (9999, 9999, 9999)
    items.sort(key=sort_key)
    # 分页
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
    except ValueError:
        page, page_size = 1, 50
    if page < 1: page = 1
    if page_size < 1: page_size = 50
    if page_size > 200: page_size = 200
    total = len(items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    end = start + page_size
    page_data = items[start:end] if start < total else []
    # 兼容前端stock字段 + 在name中标明内外径 + 反缓存
    for item in page_data:
        if 'stock' not in item and 'qty' in item:
            item['stock'] = item['qty']
        if item.get('dim_type') == 'inner' and not item['name'].startswith('内径'):
            item['name'] = '内径' + item['name']
        elif item.get('dim_type') == 'outer' and not item['name'].startswith('外径') and not item['name'].startswith('内'):
            item['name'] = '外径' + item['name']
    # 匹配对应刀模（按尺寸+内外径）——只加载一次刀模库
    dm_match_db = load_dimoldb()
    for item in page_data:
        dm_info = []
        try:
            l, w, h = item.get('length'), item.get('width'), item.get('height')
            idim = item.get('dim_type', '')
            if l and w and h:
                # 按尺寸匹配所有刀模（不限product_type）
                candidates = [d for d in dm_match_db if
                    abs(float(d.get('length', 0)) - float(l)) < 0.1
                    and abs(float(d.get('width', 0)) - float(w)) < 0.1
                    and abs(float(d.get('height', 0)) - float(h)) < 0.1]
                # 内外径过滤（基于刀模name和remark）
                if idim == 'outer':
                    candidates = [d for d in candidates if '(外)' in d.get('name', '') or ('外' in (d.get('remark') or ''))]
                elif idim == 'inner':
                    candidates = [d for d in candidates if '(内)' in d.get('name', '') or ('内' in (d.get('remark') or ''))]
                # 当库存dim_type为空时，按库存name中的"内径"/"外径"前缀来过滤
                if not idim:
                    iname = item.get('name', '')
                    if iname.startswith('内径'):
                        candidates = [d for d in candidates if '(内)' in d.get('name', '') or ('内' in (d.get('remark') or ''))]
                    else:
                        # 没有标记内外径的，默认走外径刀模
                        candidates = [d for d in candidates if '(外)' in d.get('name', '') or ('外' in (d.get('remark') or ''))]
                for dm in candidates:
                    dm_info.append({
                        'id': dm.get('id', ''),
                        'name': dm.get('name', ''),
                        'code': dm.get('code', '') or '',
                        'remark': dm.get('remark', '') or ''
                    })
        except:
            pass
        item['dimoldb_info'] = dm_info
    resp = jsonify({
        "success": True,
        "data": page_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "can_edit": _has_inv_edit_perm()
    })
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return resp

@app.route('/api/inventory', methods=['POST'])
def add_inventory():
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        tab = item.get('tab', 'finished')
        if 'name' not in item:
            return jsonify({"success": False, "error": "缺少名称"})
        data = load_inventory()
        item['id'] = f"inv_{int(time.time())}_{len(data.get(tab, []))}"
        item['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        item['updated_at'] = item['created_at']
        if tab not in data:
            data[tab] = []
        data[tab].append(item)
        save_inventory(data)
        return jsonify({"success": True, "message": "已添加"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>', methods=['PUT'])
def update_inventory(item_id):
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        item = request.get_json()
        data = load_inventory()
        for tab in ('finished', 'raw'):
            for d in data.get(tab, []):
                if d.get('id') == item_id:
                    d.update(item)
                    d['id'] = item_id
                    d['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    save_inventory(data)
                    return jsonify({"success": True, "message": "已更新"})
        return jsonify({"success": False, "error": "未找到"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>', methods=['DELETE'])
def delete_inventory(item_id):
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    data = load_inventory()
    for tab in ('finished', 'raw'):
        before = len(data.get(tab, []))
        data[tab] = [d for d in data.get(tab, []) if d.get('id') != item_id]
        if len(data[tab]) < before:
            save_inventory(data)
            return jsonify({"success": True, "message": "已删除"})
    return jsonify({"success": False, "error": "未找到"})

@app.route('/api/inventory/stock', methods=['POST'])
def stock_operation():
    """入库/出库操作（带log记录）"""
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        body = request.get_json()
        item_id = body.get('id')
        op = body.get('op', 'in')
        qty = int(body.get('qty', 0))
        remark = body.get('remark', '').strip()
        if not item_id or qty <= 0:
            return jsonify({"success": False, "error": "参数错误"})
        # 获取操作人
        operator = ''
        if 'username' in session:
            operator = USERS.get(session['username'], {}).get('employee_name', session['username'])
        data = load_inventory()
        for tab in ('finished', 'raw'):
            for d in data.get(tab, []):
                if d.get('id') == item_id:
                    current = int(d.get('stock', 0))
                    if op == 'in':
                        d['stock'] = current + qty
                    else:
                        if current < qty:
                            return jsonify({"success": False, "error": f"库存不足！当前{current}，出库{qty}"})
                        d['stock'] = current - qty
                    d['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    # 写log
                    if 'inout_log' not in d:
                        d['inout_log'] = []
                    op_label = '入库' if op == 'in' else '出库'
                    before_stock = current
                    after_stock = d['stock']
                    d['inout_log'].append({
                        "time": d['updated_at'],
                        "op": op,
                        "op_label": op_label,
                        "qty": qty,
                        "before": before_stock,
                        "after": after_stock,
                        "operator": operator,
                        "remark": remark
                    })
                    save_inventory(data)
                    return jsonify({
                        "success": True,
                        "message": f"{op_label}成功（{before_stock} → {after_stock}）",
                        "new_stock": d['stock']
                    })
        return jsonify({"success": False, "error": "未找到该商品"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/inventory/<item_id>/log', methods=['GET'])
def get_inventory_log(item_id):
    """获取某商品的出入库历史记录"""
    data = load_inventory()
    for tab in ('finished', 'raw'):
        for d in data.get(tab, []):
            if d.get('id') == item_id:
                log = d.get('inout_log', [])
                return jsonify({
                    "success": True,
                    "item_name": d.get('name', ''),
                    "current_stock": d.get('stock', 0),
                    "log": log
                })
    return jsonify({"success": False, "error": "未找到该商品"})

@app.route('/api/inventory/single', methods=['GET'])
def get_inventory_single():
    """根据ID获取单个库存商品"""
    item_id = request.args.get('id', '').strip()
    if not item_id:
        return jsonify({"success": False, "error": "缺少ID"})
    data = load_inventory()
    for tab in ('finished', 'raw'):
        for d in data.get(tab, []):
            if d.get('id') == item_id:
                return jsonify({"success": True, "data": d})
    return jsonify({"success": False, "error": "未找到该商品"})

@app.route('/api/inventory/import_excel', methods=['POST'])
def import_inventory_excel():
    """导入excel库存台账"""
    if not _has_inv_edit_perm():
        return jsonify({"success": False, "error": "无权限"})
    try:
        import openpyxl
        # 检查是否有上传文件
        if 'file' in request.files:
            f = request.files['file']
            if f.filename == '':
                return jsonify({"success": False, "error": "未选择文件"})
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            f.save(tmp.name)
            tmp.close()
            filepath = tmp.name
        else:
            # 兼容旧模式：从缓存目录查找
            data = request.get_json() or {}
            filepath = data.get('filepath', '')
            if not os.path.isabs(filepath):
                cache_dir = '/home/admin/.hermes/cache/documents/'
                if os.path.exists(cache_dir):
                    candidates = [f for f in os.listdir(cache_dir) if filepath in f and f.endswith('.xlsx')]
                    if candidates:
                        filepath = os.path.join(cache_dir, sorted(candidates, key=lambda f: os.path.getmtime(os.path.join(cache_dir, f)), reverse=True)[0])
            if not filepath or not os.path.exists(filepath):
                return jsonify({"success": False, "error": "未找到文件"})
        
        sheet_name = request.form.get('sheet_name', '') if 'file' in request.files else data.get('sheet', '')
        product_type = request.form.get('product_type', 'zhengsquare') if 'file' in request.files else data.get('product_type', 'zhengsquare')
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # 找对应sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return jsonify({"success": False, "error": f"未找到工作表: {sheet_name}，可用的: {wb.sheetnames}"})
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        inv_data = load_inventory()
        tab = 'finished'
        added = 0
        
        # 自动探测表头行（找包含"外尺寸"的行）
        header_row = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            vals = [str(v or '') for v in row[:8]]
            if '外尺寸' in vals or '外尺寸规格' in vals:
                header_row = r_idx
                break
        
        if header_row is None:
            return jsonify({"success": False, "error": "未找到表头行（需包含外尺寸）"})
        
        # 确定列索引
        header = [str(v or '') for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col_map = {}
        for i, h in enumerate(header):
            h_clean = h.strip()
            if '外尺寸' in h_clean:
                col_map['spec'] = i
            elif '材质' in h_clean:
                col_map['material'] = i
            elif '货位' in h_clean or '货 位' in h_clean:
                col_map['location'] = i
            elif '上月' in h_clean or '上月库存' in h_clean:
                col_map['last_month'] = i
            elif '入库' in h_clean and '数量' in h_clean:
                col_map['in_qty'] = i
            elif '出库' in h_clean and '数量' in h_clean:
                col_map['out_qty'] = i
            elif '剩余' in h_clean or '库存' in h_clean:
                col_map['stock'] = i
            elif '序号' in h_clean:
                col_map['seq'] = i
        if 'spec' not in col_map:
            return jsonify({"success": False, "error": f"未找到'外尺寸'列，表头: {header}"})
        
        # 读取数据行：从header_row之后找到第一个有规格的行
        data_start = header_row + 1
        for r_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            if col_map['spec'] < len(row):
                spec_val = row[col_map['spec']]
                spec_str = str(spec_val).strip() if spec_val else ''
                if spec_val and spec_str and not spec_str.startswith('=') and not spec_str.startswith('#'):
                    try:
                        dims = spec_str.replace('×','*').replace('x','*').split('*')
                        if len(dims) >= 3:
                            data_start = r_idx
                            break
                    except:
                        pass
        
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            spec = row[col_map['spec']] if col_map['spec'] < len(row) else None
            if not spec or not str(spec).strip():
                continue
            spec = str(spec).strip()
            # 跳过公式行（=ROW开头）
            if spec.startswith('=') or spec.startswith('#'):
                continue
            
            # 解析规格：如 "10.5×7.5×4.5" 或 "6×6×2"
            dims = spec.replace('×', '*').replace('x', '*').replace('X', '*').split('*')
            length = width = height = None
            try:
                if len(dims) >= 3:
                    length = float(dims[0].strip())
                    width = float(dims[1].strip())
                    height = float(dims[2].strip())
            except:
                pass
            
            # 材质（可能有多个材质行共享同一个规格）
            material = str(row[col_map['material']]).strip() if col_map.get('material') and col_map['material'] < len(row) and row[col_map['material']] else ''
            location = str(row[col_map['location']]).strip() if col_map.get('location') and col_map['location'] < len(row) and row[col_map['location']] else ''
            
            # 数值
            last_month = 0
            stock_qty = 0
            try:
                v = row[col_map['last_month']] if col_map.get('last_month') and col_map['last_month'] < len(row) else 0
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    last_month = int(v)
            except: pass
            try:
                v = row[col_map['stock']] if col_map.get('stock') and col_map['stock'] < len(row) else 0
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    stock_qty = int(v)
            except: pass
            
            # 构造商品名
            name_parts = [spec]
            if material and material not in ('None', ''):
                name_parts.append(material)
            name = ' '.join(name_parts)
            
            # 去重：同名+同产品类型 不重复导入
            exists = False
            for d in inv_data[tab]:
                if d.get('name') == name and d.get('product_type') == product_type:
                    exists = True
                    break
            
            if not exists:
                item_id = f"inv_{int(time.time())}_{added}_{len(inv_data[tab])}"
                item = {
                    'id': item_id,
                    'name': name,
                    'spec': spec,
                    'product_type': product_type,
                    'material': material,
                    'location': location,
                    'qty': stock_qty,
                    'last_month_qty': last_month,
                    'length': length,
                    'width': width,
                    'height': height,
                    'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                    'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                }
                inv_data[tab].append(item)
                added += 1
        
        save_inventory(inv_data)
        return jsonify({"success": True, "message": f"导入完成，新增 {added} 条记录", "added": added})
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()})

# ==================== 权限列表 ====================
@app.route('/api/permissions')
def permissions():
    return jsonify({
        "roles": [
            {"name": "超级管理员", "desc": "所有权限", "count": 2},
            {"name": "主管", "desc": "生产管理、员工查看", "count": 5},
            {"name": "客服", "desc": "订单查看、生产进度", "count": 10},
            {"name": "员工", "desc": "扫码报工、查看任务", "count": 20},
        ]
    })

@app.route('/api/permissions/data')
def permissions_data():
    _sync_all_employees_perms()
    employees = _employees_master_list
    perms = _permission_data.get("permissions", {})
    return jsonify({
        "employees": employees,
        "permissions": perms,
        "roles": [
            {"name": "超级管理员", "desc": "所有权限", "count": 2},
            {"name": "主管", "desc": "生产管理、员工查看", "count": 5},
            {"name": "客服", "desc": "订单查看、生产进度", "count": 10},
            {"name": "员工", "desc": "扫码报工、查看任务", "count": 20},
        ],
        "menu_modules": [
            {"name": "首页", "perms": ["今日订单", "生产概览", "销售概览"]},
            {"name": "报价", "perms": ["报价计算", "报价参数"]},
            {"name": "刀模", "perms": ["刀模查询", "刀模录入"]},
            {"name": "库存", "perms": ["成品库存", "原材料库存"]},
            {"name": "订单进度", "perms": ["生产进度", "加急管理"]},
            {"name": "数据看板", "perms": ["销售数据", "生产数据"]},
            {"name": "实时订单", "perms": ["订单查看", "客服分配"]},
            {"name": "员工", "perms": ["员工查看", "员工管理"]},
        ]
    })

@app.route('/api/employees')
def employees_list():
    _sync_all_employees_perms()
    emp_users = {}
    for un, u in USERS.items():
        if u.get('is_system'):
            continue
        en = u.get('employee_name') or u.get('name', '')
        if en:
            emp_users[en] = un
    out = []
    for e in _employees_master_list:
        row = dict(e)
        row['username'] = emp_users.get(e.get('name', ''), '')
        row['role'] = USERS.get(row['username'], {}).get('role', '')
        out.append(row)
    return jsonify(out)

# ==================== 静态文件 ====================
@app.route('/login_simple')
def login_simple():
    return send_from_directory('.', 'simple_login.html')

# ==================== 原材料库存 API ====================

def load_raw_data():
    """从MySQL加载原材料数据"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT id, date, name, supplier, paper_width, paper_length, qty, remark, created_at, updated_at FROM raw_materials ORDER BY created_at DESC")
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'date': r['date'] or '',
                'name': r['name'] or '',
                'supplier': r['supplier'] or '',
                'paper_width': str(r['paper_width'] or ''),
                'paper_length': str(r['paper_length'] or ''),
                'qty': r['qty'] or 0,
                'remark': r['remark'] or '',
                'created_at': r['created_at'] or '',
                'updated_at': r['updated_at'] or ''
            })
        cur.close()
        db.close()
        return result
    except Exception as e:
        print(f'[MySQL load_raw_data] 错误: {e}')
        return []

def save_raw_data(data):
    """保存原材料数据到MySQL（truncate+批量insert）"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE raw_materials")
        for item in data:
            cur.execute(
                "INSERT INTO raw_materials (id, date, name, supplier, paper_width, paper_length, qty, remark, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    item.get('id', ''),
                    item.get('date', ''),
                    item.get('name', ''),
                    item.get('supplier', ''),
                    item.get('paper_width', ''),
                    item.get('paper_length', ''),
                    item.get('qty', 0),
                    item.get('remark', ''),
                    item.get('created_at', ''),
                    item.get('updated_at', '')
                )
            )
        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_raw_data] 错误: {e}')
        return False

@app.route('/api/raw', methods=['GET'])
def get_raw():
    data = load_raw_data()
    search = request.args.get('search', '').strip()
    date = request.args.get('date', '').strip()
    name = request.args.get('name', '').strip()
    supplier = request.args.get('supplier', '').strip()
    paper_width = request.args.get('paper_width', '').strip()
    paper_length = request.args.get('paper_length', '').strip()
    if date:
        data = [d for d in data if date in d.get('date', '')]
    if name:
        data = [d for d in data if name.lower() in d.get('name', '').lower()]
    if supplier:
        data = [d for d in data if supplier.lower() in d.get('supplier', '').lower()]
    if paper_width:
        data = [d for d in data if str(d.get('paper_width', '')) == paper_width]
    if paper_length:
        data = [d for d in data if str(d.get('paper_length', '')) == paper_length]
    if search:
        data = [d for d in data if search.lower() in d.get('name', '').lower() or search.lower() in d.get('remark', '').lower()]
    data.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify({"success": True, "data": data})

@app.route('/api/raw', methods=['POST'])
def add_raw():
    body = request.get_json()
    if not body or not body.get('name', '').strip():
        return jsonify({"success": False, "error": "材料名称不能为空"})
    data = load_raw_data()
    new_id = str(int(time.time() * 1000)) + str(len(data))
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    entry = {
        "id": new_id,
        "date": body.get('date', '').strip(),
        "name": body['name'].strip(),
        "supplier": body.get('supplier', '').strip(),
        "paper_width": body.get('paper_width', ''),
        "paper_length": body.get('paper_length', ''),
        "qty": int(body.get('qty', 0)),
        "remark": body.get('remark', '').strip(),
        "created_at": now,
        "updated_at": now
    }
    data.append(entry)
    save_raw_data(data)
    return jsonify({"success": True, "data": entry, "message": "原材料保存成功"})

@app.route('/api/raw/<raw_id>', methods=['PUT'])
def update_raw(raw_id):
    body = request.get_json()
    data = load_raw_data()
    for item in data:
        if item['id'] == raw_id:
            if body.get('date', '').strip(): item['date'] = body['date'].strip()
            if body.get('name', '').strip(): item['name'] = body['name'].strip()
            if 'supplier' in body: item['supplier'] = body['supplier'].strip()
            if 'paper_width' in body: item['paper_width'] = str(body['paper_width'])
            if 'paper_length' in body: item['paper_length'] = str(body['paper_length'])
            if 'qty' in body: item['qty'] = int(body['qty'])
            if 'remark' in body: item['remark'] = body['remark'].strip()
            item['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            save_raw_data(data)
            return jsonify({"success": True, "data": item, "message": "更新成功"})
    return jsonify({"success": False, "error": "未找到该材料"})

@app.route('/api/raw/stock', methods=['POST'])
def raw_stock():
    body = request.get_json()
    rid = body.get('id', '')
    op = body.get('op', 'in')
    qty = int(body.get('qty', 0))
    if qty <= 0:
        return jsonify({"success": False, "error": "数量必须大于0"})
    data = load_raw_data()
    for item in data:
        if item['id'] == rid:
            if op == 'in':
                item['qty'] = (item.get('qty', 0) or 0) + qty
            else:
                item['qty'] = max(0, (item.get('qty', 0) or 0) - qty)
            item['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            save_raw_data(data)
            return jsonify({"success": True, "data": item, "message": f"{'入库' if op=='in' else '出库'}成功"})
    return jsonify({"success": False, "error": "未找到该材料"})

# ==================== 企业微信验证接口 ====================
def wechat_callback():
    import hashlib
    msg_signature = request.args.get('msg_signature', '')
    timestamp = request.args.get('timestamp', '')
    nonce = request.args.get('nonce', '')
    echostr = request.args.get('echostr', '')
    if request.method == 'GET' and echostr:
        tmp_list = sorted([WECHAT_TOKEN, timestamp, nonce])
        tmp_str = ''.join(tmp_list)
        sha1 = hashlib.sha1(tmp_str.encode('utf-8')).hexdigest()
        if sha1 == msg_signature:
            return echostr, 200, {'Content-Type': 'text/plain'}
        return 'signature fail', 403
    if request.method == 'POST':
        return 'success', 200, {'Content-Type': 'text/plain; charset=utf-8'}
    return 'ok', 200

# ==================== 静态文件 ====================
@app.route('/')
def index():
    resp = make_response(send_from_directory('.', 'index_cs.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)


from io import BytesIO
import qrcode
import barcode
from barcode.writer import ImageWriter
import base64


def _1688_fetch_all_shops_orders():
    """遍历所有1688店铺，拉取每个店铺的订单并合并（独立函数，不依赖全局配置）"""
    import time as _t
    api = '1/com.alibaba.trade/alibaba.trade.getSellerOrderList'
    all_shop_orders = []
    max_pages = 3  # 最多翻3页

    for shop in ALIBABA_SHOPS:
        shop_name = shop['shop_name']
        token = shop['access_token']
        app_key = int(shop['app_key'])
        app_secret = shop['app_secret']
        server = shop['server']
        print(f'[1688] 拉取店铺: {shop_name}')
        for page in range(1, max_pages + 1):
            try:
                url_path = f'param2/{api}/{app_key}'
                params = {'access_token': token, 'page': page, 'pageSize': 50}
                param_list = sorted([str(k) + str(v) for k, v in params.items()])
                msg = url_path.encode('utf-8')
                for p in param_list:
                    msg += p.encode('utf-8')
                sign = hmac.new(app_secret.encode('utf-8'), msg, hashlib.sha1).hexdigest().upper()
                params['_aop_signature'] = sign
                url = f'https://{server}/openapi/{url_path}'
                form_data = urllib.parse.urlencode(params)
                req = urllib.request.Request(url, data=form_data.encode('utf-8'), method='POST')
                resp = urllib.request.urlopen(req, timeout=30)
                data = json.loads(resp.read().decode('utf-8'))
                orders = data.get('result', [])
                if not orders:
                    break
                for o in orders:
                    info = o.get('baseInfo', {})
                    info['shop_name'] = shop_name
                all_shop_orders.extend(orders)
                if len(orders) < 50:
                    break
                _t.sleep(0.5)
            except Exception as e:
                print(f'[1688] {shop_name} 第{page}页失败: {e}')
                break
        print(f'[1688] {shop_name}: 拉取完成')
    return all_shop_orders


def _1688_sign(url_path, params, secret):
    """1688签名：HMAC-SHA1"""
    param_list = sorted([str(k) + str(v) for k, v in params.items()])
    msg = url_path.encode('utf-8')
    for p in param_list:
        msg += p.encode('utf-8')
    sha = hmac.new(secret.encode('utf-8'), msg, hashlib.sha1)
    return sha.hexdigest().upper()


def _1688_api(api_uri, biz_params=None):
    """调用1688开放平台API"""
    cfg = ALIBABA_CONFIG
    url_path = f'param2/{api_uri}/{cfg["app_key"]}'
    params = {'access_token': cfg['access_token']}
    if biz_params:
        params.update(biz_params)
    params['_aop_signature'] = _1688_sign(url_path, params, cfg['app_secret'])
    url = f'https://{cfg["server"]}/openapi/{url_path}'
    form_data = urllib.parse.urlencode(params)
    try:
        req = urllib.request.Request(url, data=form_data.encode('utf-8'), method='POST')
        resp = urllib.request.urlopen(req, timeout=30)
        return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print(f'[1688API] {api_uri} 调用失败: {e}')
        return {}


def _km_merge_1688_spec(item: dict) -> str:
    import km_api as _km
    return _km.km_platform_item_attrs(item)


def _1688_format_order(o):
    """格式化1688订单为统一格式"""
    info = o.get('baseInfo', {})
    items = o.get('productItems', [])
    product_list = []
    for item in items:
        spec = _km_merge_1688_spec(item)
        product_list.append({
            'name': item.get('name', ''),
            'sku': item.get('productCargoNumber', ''),
            'qty': item.get('quantity', 0),
            'price': item.get('price', 0),
            'spec': spec,
            'display': spec,
            'skuId': item.get('skuID', ''),
            'productId': item.get('productID', '')
        })
    so_id = str(info.get('idOfStr', info.get('id', '')))
    return {
        'so_id': so_id,
        'platform': '1688',
        'order_status': info.get('status', ''),
        'status_label': '待发货' if info.get('status') == 'waitsellersend' else info.get('status', ''),
        'created': info.get('createTime', '')[:8] if info.get('createTime') else '',
        'pay_time': info.get('payTime', '')[:8] if info.get('payTime') else '',
        'total_amount': info.get('totalAmount', 0),
        'shipping_fee': info.get('shippingFee', 0) / 100,
        'discount': info.get('discount', 0) / 100,
        'receiver_name': info.get('receiverName', info.get('buyerLoginId', '')),
        'receiver_mobile': info.get('receiverMobile', ''),
        'receiver_address': info.get('receiverAddress', ''),
        'receiver_phone': info.get('receiverPhone', ''),
        'shop_name': info.get('shop_name', info.get('buyerLoginId', '1688')),
        'items': product_list,
        'buyer_login_id': info.get('buyerLoginId', ''),
        'alipay_trade_id': info.get('alipayTradeId', ''),
        'trade_type': info.get('tradeType', ''),
        'buyer_memo': '',
        'seller_memo': get_order_memo(so_id),
    }

# ==================== 订单备注本地存储 ====================

_ORDER_MEMO = {}

def get_order_memo(so_id):
    return _ORDER_MEMO.get(so_id, '')

def set_order_memo(so_id, memo):
    _ORDER_MEMO[so_id] = memo

@app.route('/api/order/1688-memo', methods=['POST'])
def api_order_1688_memo():
    """修改卖家备注并同步回1688"""
    if 'username' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    data = request.get_json() or {}
    so_id = data.get('so_id', '')
    memo = data.get('memo', '')
    if not so_id:
        return jsonify({'success': False, 'error': '缺少订单ID'})
    # 先本地存储
    set_order_memo(so_id, memo)
    # 同步到1688
    try:
        # 调用1688 memoAdd API
        result = _1688_api('1/com.alibaba.trade/alibaba.order.memoAdd', {
            'orderId': so_id,
            'memo': memo,
            'remarkIcon': '0'
        })
        if result.get('errorCode') or result.get('errorMessage'):
            return jsonify({'success': False, 'error': f'1688同步失败: {result.get("errorMessage", result.get("errorCode", "未知错误"))}', 'local_saved': True})
    except Exception as e:
        return jsonify({'success': True, 'error': f'1688同步失败: {str(e)}', 'local_saved': True})
    return jsonify({'success': True, 'message': '备注已保存并同步到1688'})

@app.route('/api/realtime/orders', methods=['GET'])
def api_realtime_orders():
    """实时订单：只读 orders_cache.json，毫秒级返回（后台静默同步）。"""
    import order_sync as _osync

    payload = _osync.read_realtime_cache_payload(
        ORDERS_CACHE_FILE, load_shop_config()
    )
    n = payload.get("total", 0)
    print(f"[实时订单] 读缓存 {payload.get('cache_status')}: {n} 条")
    return jsonify(payload)

# ==================== 店铺客服配置 API ====================

DEFAULT_SHOP_CONFIG = [
    {"id": "shop_1688_1", "platform": "1688", "shop_name": "友尚", "customer_service": "张慧平,戴志美", "sort_order": 1},
    {"id": "shop_1688_2", "platform": "1688", "shop_name": "亚润", "customer_service": "陈贝贝", "sort_order": 2},
    {"id": "shop_1688_3", "platform": "1688", "shop_name": "三羊", "customer_service": "罗怡", "sort_order": 3},
    {"id": "shop_1688_4", "platform": "1688", "shop_name": "正方形", "customer_service": "周井梅", "sort_order": 4},
    {"id": "shop_1688_5", "platform": "1688", "shop_name": "大鱼", "customer_service": "周井梅", "sort_order": 5},
    {"id": "shop_1688_6", "platform": "1688", "shop_name": "阿里新鑫星", "customer_service": "戴志美", "sort_order": 6},
    {"id": "shop_tmall_1", "platform": "tmall", "shop_name": "飞机盒彩色专卖店", "customer_service": "廖思美", "sort_order": 7},
    {"id": "shop_tmall_2", "platform": "tmall", "shop_name": "飞机盒正方形专卖店", "customer_service": "廖思美", "sort_order": 8},
    {"id": "shop_tmall_3", "platform": "tmall", "shop_name": "飞机盒止合专卖店", "customer_service": "石梅清", "sort_order": 9},
    {"id": "shop_tmall_4", "platform": "tmall", "shop_name": "飞机盒扣底盒专卖店", "customer_service": "石梅清", "sort_order": 10},
    {"id": "shop_tmall_5", "platform": "tmall", "shop_name": "飞机盒小批量专卖店", "customer_service": "张文杰", "sort_order": 11},
    {"id": "shop_taobao_1", "platform": "taobao", "shop_name": "飞机盒品牌店", "customer_service": "张文杰", "sort_order": 12},
    {"id": "shop_taobao_2", "platform": "taobao", "shop_name": "俊鑫纸品厂", "customer_service": "石梅清", "sort_order": 13},
    {"id": "shop_taobao_3", "platform": "taobao", "shop_name": "当下家包装", "customer_service": "张文杰", "sort_order": 14}
]

def load_shop_config():
    """从MySQL加载店铺配置"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT id, shop_name, platform, customer_service, sort_order FROM shop_config ORDER BY sort_order ASC")
        rows = cur.fetchall()
        cur.close()
        db.close()
        if not rows:
            save_shop_config(list(DEFAULT_SHOP_CONFIG))
            return list(DEFAULT_SHOP_CONFIG)
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'shop_name': r['shop_name'] or '',
                'platform': r['platform'] or '',
                'customer_service': r['customer_service'] or '',
                'sort_order': r['sort_order'] or 0
            })
        return result
    except Exception as e:
        print(f'[MySQL load_shop_config] 错误: {e}')
        return list(DEFAULT_SHOP_CONFIG)

def save_shop_config(config):
    """保存店铺配置到MySQL（truncate+批量insert）"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("TRUNCATE TABLE shop_config")
        for item in config:
            cur.execute(
                "INSERT INTO shop_config (id, shop_name, platform, customer_service, sort_order) VALUES (%s,%s,%s,%s,%s)",
                (
                    item.get('id', ''),
                    item.get('shop_name', ''),
                    item.get('platform', ''),
                    item.get('customer_service', ''),
                    item.get('sort_order', 0)
                )
            )
        cur.close()
        db.close()
        return True
    except Exception as e:
        print(f'[MySQL save_shop_config] 错误: {e}')
        return False

@app.route('/api/shop-config', methods=['GET'])
def api_get_shop_config():
    config = load_shop_config()
    return jsonify({'success': True, 'config': config})

@app.route('/api/shop-config', methods=['POST'])
def api_add_shop_config():
    """添加单个店铺配置"""
    try:
        data = request.get_json(force=True) or {}
        configs = load_shop_config()
        new_id = f"shop_{len(configs)+1}_{int(time.time())}"
        new_item = {
            'id': new_id,
            'platform': data.get('platform', '1688'),
            'shop_name': data.get('shop_name', ''),
            'customer_service': data.get('customer_service', ''),
            'sort_order': data.get('sort_order', len(configs)+1)
        }
        configs.append(new_item)
        save_shop_config(configs)
        return jsonify({'success': True, 'message': '已添加', 'item': new_item})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop-config/<config_id>', methods=['PUT'])
def api_update_shop_config(config_id):
    """更新单个店铺配置"""
    try:
        data = request.get_json(force=True) or {}
        configs = load_shop_config()
        for c in configs:
            if c.get('id') == config_id:
                if 'customer_service' in data:
                    c['customer_service'] = data['customer_service']
                if 'sort_order' in data:
                    c['sort_order'] = data['sort_order']
                if 'shop_name' in data:
                    c['shop_name'] = data['shop_name']
                if 'platform' in data:
                    c['platform'] = data['platform']
                save_shop_config(configs)
                return jsonify({'success': True, 'message': '已更新'})
        return jsonify({'success': False, 'error': '未找到该配置'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop-config/<config_id>', methods=['DELETE'])
def api_delete_shop_config(config_id):
    """删除单个店铺配置"""
    try:
        configs = load_shop_config()
        configs = [c for c in configs if c.get('id') != config_id]
        save_shop_config(configs)
        return jsonify({'success': True, 'message': '已删除'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop-config/reset', methods=['POST'])
def api_reset_shop_config():
    save_shop_config(DEFAULT_SHOP_CONFIG)
    return jsonify({'success': True, 'config': list(DEFAULT_SHOP_CONFIG)})

# ========== 后台定时同步订单（快麦 ERP）==========
ORDERS_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "orders_cache.json"
)

def _km_sync_orders_to_cache(days_back=14):
    """快麦(1688无需奇门) + 1688直连 → orders_cache.json"""
    import order_sync as _osync
    r = _osync.sync_orders_to_cache(
        ORDERS_CACHE_FILE,
        days_back=days_back,
        memo_getter=get_order_memo,
        include_1688_direct=True,
    )
    return r.get('pending_count', 0)


@app.route('/api/sync/force', methods=['POST'])
def api_sync_force():
    """手动触发快麦订单同步（后台执行，立即返回 JSON，避免网关超时）"""
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    import order_sync as _osync
    ok, msg = _osync.start_force_sync_async(
        ORDERS_CACHE_FILE,
        days_back=30,
        memo_getter=get_order_memo,
        include_1688_direct=True,
    )
    if not ok:
        return jsonify({'success': False, 'error': msg}), 409
    return jsonify({'success': True, 'async': True, 'message': msg})


@app.route('/api/sync/status', methods=['GET'])
def api_sync_status():
    """查询后台同步进度（配合 /api/sync/force async）"""
    if not resolve_login_user():
        return jsonify({'success': False, 'error': '未登录'}), 401
    import order_sync as _osync
    st = _osync.force_sync_status()
    last = st.get('last') or {}
    return jsonify({
        'success': True,
        'running': st.get('running'),
        'error': st.get('error'),
        'phase': st.get('phase'),
        'detail': st.get('detail'),
        'last': last,
        'count': last.get('pending_count') if isinstance(last, dict) else None,
    })


@app.route('/api/km/probe', methods=['GET'])
def api_km_probe():
    """快麦连通性探测（店铺数、近3天订单抽样）"""
    if 'username' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    import km_api as _km
    return jsonify({'success': True, 'data': _km.km_probe()})


@app.route('/api/km/refresh_token', methods=['POST'])
def api_km_refresh_token():
    """刷新快麦 session（open.token.refresh）"""
    if 'username' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    import km_api as _km
    return jsonify(_km.km_refresh_token())


import order_sync_scheduler as _order_sched

_order_sched.start_background_order_sync(
    ORDERS_CACHE_FILE,
    memo_getter=get_order_memo,
    include_1688_direct=True,
    full_days_back=30,
    incremental_days_back=7,
    interval_sec=180,
)

# ==================== 缺失API补充（客服端）====================

def _build_production_orders_for_cs():
    """读 orders_cache + production_flows（生产端扫码报工）+ 工序树。"""
    import production_helpers as ph

    return ph.build_production_orders(
        _orders_cache_path(),
        DB_CONFIG,
        _permission_data.get("processes", []),
        _order_extra,
    )


@app.route("/api/production_orders")
def api_production_orders():
    """客服端：订单生产进度（与主站同源数据结构）"""
    return jsonify({"orders": _build_production_orders_for_cs()})

@app.route('/api/scan_logs')
def api_scan_logs():
    """返回扫码日志列表"""
    return jsonify({"logs": []})

@app.route('/api/scan_workers')
def api_scan_workers():
    """返回可选操作人列表"""
    return jsonify({"workers": []})

@app.route('/api/employee/status', methods=['GET', 'POST'])
def api_employee_status():
    """获取/修改员工出勤状态（内存中维护）"""
    global _employee_today_status
    if request.method == 'GET':
        today = datetime.date.today().isoformat()
        day_status = _employee_today_status.get(today, {})
        # 所有在职员工默认"出勤"
        statuses = {}
        for emp in _employees_master_list:
            statuses[emp['name']] = day_status.get(emp['name'], '出勤')
        return jsonify({"statuses": statuses})
    else:
        data = request.get_json()
        name = data.get('name', '')
        status = data.get('status', '出勤')
        today = datetime.date.today().isoformat()
        if today not in _employee_today_status:
            _employee_today_status[today] = {}
        _employee_today_status[today][name] = status
        persist()
        return jsonify({"success": True, "message": "状态已更新"})

@app.route('/api/employee/resigned')
def api_employee_resigned():
    """返回离职员工列表"""
    return jsonify({"employees": _resigned_employees})

@app.route('/api/employee/add', methods=['POST'])
def api_employee_add():
    """新增员工"""
    global _employees_master_list, _permission_data
    data = request.get_json()
    name = data.get('name', '').strip()
    position = data.get('position', '').strip()
    if not name:
        return jsonify({"success": False, "message": "员工姓名不能为空"})
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO employees (name, position, enabled) VALUES (%s, %s, 1)", (name, position))
        db.commit()
        cur.close()
        db.close()
        # 刷新内存列表
        _sync_all_employees_perms()
        return jsonify({"success": True, "message": "添加成功"})
    except Exception as e:
        return jsonify({"success": False, "message": f"添加失败: {str(e)}"})

@app.route('/api/employee/update', methods=['POST'])
def api_employee_update():
    """编辑员工（姓名/职务/手机/部门/角色/密码）"""
    global _employees_master_list
    un = resolve_login_user()
    if not un:
        return jsonify({"success": False, "message": "未登录"}), 401
    actor = USERS.get(un)
    if not actor or actor.get('role') != '超级管理员':
        return jsonify({"success": False, "message": "仅超级管理员可编辑员工资料"}), 403

    data = request.get_json() or {}
    old_name = (data.get('old_name') or '').strip()
    name = (data.get('name') or '').strip()
    position = (data.get('position') or '').strip()
    phone = data.get('phone')
    dept = (data.get('dept') or '').strip()
    group = (data.get('group') or '').strip()
    role = (data.get('role') or '').strip()
    new_password = (data.get('new_password') or '').strip()
    username = (data.get('username') or '').strip()

    if not old_name or not name:
        return jsonify({"success": False, "message": "员工姓名不能为空"})

    target_emp = next((e for e in _employees_master_list if e.get('name') == old_name), None)
    if not target_emp:
        return jsonify({"success": False, "message": f"未找到员工 {old_name}"})
    target_emp['name'] = name
    if position:
        target_emp['position'] = position
    if phone is not None:
        target_emp['phone'] = str(phone).strip()
    if dept:
        target_emp['dept'] = dept
    if group:
        target_emp['group'] = group

    if old_name != name:
        perms = _permission_data.setdefault('permissions', {})
        if old_name in perms:
            perms[name] = perms.pop(old_name)

    user_updated = False
    for ukey, u in USERS.items():
        if u.get('employee_name') == old_name:
            u['employee_name'] = name
            u['name'] = name
            if role:
                u['role'] = role
            if new_password:
                if len(new_password) < 4:
                    return jsonify({"success": False, "message": "密码至少4位"})
                u['password'] = hashlib.sha256(new_password.encode()).hexdigest()
            user_updated = True
            break
    if not user_updated and username and username in USERS:
        u = USERS[username]
        if role:
            u['role'] = role
        if new_password:
            if len(new_password) < 4:
                return jsonify({"success": False, "message": "密码至少4位"})
            u['password'] = hashlib.sha256(new_password.encode()).hexdigest()

    try:
        db = get_db()
        cur = db.cursor()
        cur.execute(
            "UPDATE employees SET name=%s, position=%s WHERE name=%s",
            (name, position or target_emp.get('position', ''), old_name),
        )
        db.commit()
        cur.close()
        db.close()
    except Exception as e:
        return jsonify({"success": False, "message": f"数据库更新失败: {str(e)}"})

    persist()
    _sync_all_employees_perms()
    return jsonify({"success": True, "message": "更新成功"})

@app.route('/api/employee/delete', methods=['POST'])
def api_employee_delete():
    """删除员工（软删除到离职列表）"""
    global _employees_master_list, _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    emp = next((e for e in _employees_master_list if e['name'] == name), None)
    if emp:
        _resigned_employees.append({
            "name": name, "position": emp.get('position', ''),
            "resigned_time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        })
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("UPDATE employees SET enabled=0 WHERE name=%s", (name,))
        db.commit()
        cur.close()
        db.close()
    except:
        pass
    _sync_all_employees_perms()
    return jsonify({"success": True, "message": "已标记为离职"})

@app.route('/api/employee/restore', methods=['POST'])
def api_employee_restore():
    """恢复离职员工"""
    global _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    _resigned_employees = [e for e in _resigned_employees if e.get('name') != name]
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("UPDATE employees SET enabled=1 WHERE name=%s", (name,))
        db.commit()
        cur.close()
        db.close()
    except:
        pass
    _sync_all_employees_perms()
    return jsonify({"success": True, "message": "已恢复"})

@app.route('/api/employee/deactivate', methods=['POST'])
def api_employee_deactivate():
    """标记员工离职（不可逆）"""
    data = request.get_json()
    name = data.get('name', '')
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("UPDATE employees SET enabled=0 WHERE name=%s", (name,))
        db.commit()
        cur.close()
        db.close()
    except:
        pass
    _sync_all_employees_perms()
    return jsonify({"success": True, "message": "已离职"})

@app.route('/api/employee/delete_resigned', methods=['POST'])
def api_employee_delete_resigned():
    """彻底删除离职记录"""
    global _resigned_employees
    data = request.get_json()
    name = data.get('name', '')
    _resigned_employees = [e for e in _resigned_employees if e.get('name') != name]
    return jsonify({"success": True, "message": "已删除"})

@app.route('/api/permissions/save', methods=['POST'])
def api_permissions_save():
    """保存权限配置"""
    global _permission_data
    data = request.get_json()
    if 'permissions' in data:
        _permission_data['permissions'] = data['permissions']
    if 'employee_enabled' in data:
        _permission_data['employee_enabled'] = data['employee_enabled']
    # 保存到数据库
    try:
        db = get_db()
        cur = db.cursor()
        for emp_name, perms in data.get('permissions', {}).items():
            for pk, val in perms.items():
                cur.execute("""
                    INSERT INTO employee_permissions (employee_name, permission_key, enabled)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE enabled=%s
                """, (emp_name, pk, 1 if val else 0, 1 if val else 0))
        for emp_name, ena in data.get('employee_enabled', {}).items():
            cur.execute("UPDATE employees SET enabled=%s WHERE name=%s", (1 if ena else 0, emp_name))
        db.commit()
        cur.close()
        db.close()
    except Exception as e:
        print(f"[权限保存] 错误: {e}")
    persist()
    return jsonify({"success": True, "message": "保存成功"})

@app.route('/api/qrcode/<order_id>')
def api_qrcode(order_id):
    """生成订单二维码（返回占位，无qr库时兜底）"""
    try:
        import qrcode
        from io import BytesIO
        import base64
        img = qrcode.make(str(order_id))
        buf = BytesIO()
        img.save(buf, format='PNG')
        b64 = base64.b64encode(buf.getvalue()).decode()
        return jsonify({"success": True, "qr_base64": b64})
    except ImportError:
        # 没有qrcode库时返回空白
        return jsonify({"success": True, "qr_base64": ""})

@app.route('/api/barcode/<order_id>')
def api_barcode(order_id):
    """生成订单条码（返回占位）"""
    return jsonify({"success": True, "barcode_base64": ""})

@app.route('/api/scan_report', methods=['POST'])
def api_scan_report():
    """扫码报工确认"""
    return jsonify({"success": True})

if __name__ == '__main__':
    print("🏭 飞机盒智能生产管理系统启动中...")
    print("📡 http://0.0.0.0:3001")
    app.run(host='0.0.0.0', port=3001, threaded=True)

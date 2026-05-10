#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
三羊包装 - 库存数据重建脚本
从三个Excel文件读取完整库存数据，重建inventory.json
"""
import json
import openpyxl
from datetime import datetime
import os, re, sys

ROOT = '/www/wwwroot/feijihe'
DOCS = '/home/admin/.hermes/cache/documents'
NOW = datetime.now().strftime('%Y-%m-%d %H:%M')

def parse_spec(spec_str):
    """解析规格，返回 (length, width, height) 或 None"""
    if not spec_str:
        return None
    s = str(spec_str).replace('×', 'x').replace('*', 'x').replace('X','x')
    # 去除材质标记 (白)(台)(黑)(红)(差)(差材料)(卡)等
    # 也去除末尾可能会有的空格
    s = s.strip()
    # 匹配纯尺寸部分：数字x数字x数字，可能有小数
    m = re.match(r'([\d.]+)\s*x\s*([\d.]+)\s*x\s*([\d.]+)', s)
    if m:
        try:
            return (float(m.group(1)), float(m.group(2)), float(m.group(3)))
        except:
            pass
    return None

def detect_material(spec_str, explicit_mat=None):
    """检测材质。如果explicit_mat给了就用它，否则从spec字符串判断"""
    if explicit_mat and str(explicit_mat).strip():
        m = str(explicit_mat).strip()
        # 标准化材质名
        mat_map = {
            '国产纸': '国产纸', '国产纸`': '国产纸',
            '台湾纸': '台湾纸', '台': '台湾纸',
            '双白': '双白', '白色': '双白', '白': '双白',
            '黑色': '黑色', '黑': '黑色',
            '红色': '红色', '红': '红色',
            '差材料': '差材料', '差': '差材料', '差质': '差材料',
        }
        if m in mat_map:
            return mat_map[m]
        # 直接返回
        return m
    
    if not spec_str:
        return '国产纸'
    s = str(spec_str)
    if '(台)' in s or s.endswith('(台）') or s.endswith('(台)'):
        return '台湾纸'
    if '(白)' in s:
        return '双白'
    if '(黑)' in s:
        return '黑色'
    if '(红)' in s:
        return '红色'
    if '(差' in s or '(卡)' in s:
        return '差材料'
    if ' 白' in s or '(白' in s:
        return '双白'
    return '国产纸'

def clean_spec(spec_str):
    """清理规格：去掉材质标记、统一分隔符为x"""
    if not spec_str:
        return ''
    s = str(spec_str).replace('×', 'x').replace('*', 'x').replace('X', 'x')
    # 去掉括号材质标记 (白)(台)(黑)(红)(差)(差材料)等
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'（[^）]*）', '', s)
    s = re.sub(r'\s+白$', '', s.strip())
    s = re.sub(r'\s+台$', '', s.strip())
    s = re.sub(r'\s+黑$', '', s.strip())
    s = re.sub(r'\s+红$', '', s.strip())
    return s.strip()

def make_name(spec, material):
    """生成name字段：规格x材质"""
    return f"{spec} {material}"

def make_zhengsquare(spec_str, material, qty, location, dim_type):
    """生成一条正方形库存记录"""
    spec = clean_spec(spec_str)
    dim = parse_spec(spec_str)
    length = dim[0] if dim else 0
    width = dim[1] if dim else 0
    height = dim[2] if dim else 0
    
    # name: 内径加"内径"前缀
    if dim_type == 'inner':
        name = f"内径{spec} {material}"
    else:
        name = f"{spec} {material}"
    
    return {
        "id": f"inv_{int(datetime.now().timestamp())}_{hash(spec+material+dim_type) % 100000}",
        "name": name,
        "spec": spec,
        "product_type": "zhengsquare",
        "material": material,
        "location": location or "",
        "qty": qty,
        "last_month_qty": 0,
        "length": length,
        "width": width,
        "height": height,
        "created_at": NOW,
        "updated_at": NOW,
        "dim_type": dim_type
    }

def make_changfang(spec_str, material, qty, location):
    """生成一条长方形库存记录"""
    spec = clean_spec(spec_str)
    dim = parse_spec(spec_str)
    length = dim[0] if dim else 0
    width = dim[1] if dim else 0
    height = dim[2] if dim else 0
    
    name = f"{spec} {material}"
    
    return {
        "id": f"inv_{int(datetime.now().timestamp())}_{hash(spec+material) % 100000}",
        "name": name,
        "spec": spec,
        "product_type": "changfang",
        "material": material,
        "location": location or "",
        "qty": qty,
        "last_month_qty": 0,
        "length": length,
        "width": width,
        "height": height,
        "created_at": NOW,
        "updated_at": NOW
    }

def read_zhengsquare(filepath, dim_type):
    """读取正方形Excel文件"""
    print(f"正在读取正方形{dim_type}: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    # 检查列头：第2行
    # B=规格 C=材质 H=剩余库存 D=货位
    records = []
    last_spec = None
    last_location = None
    
    for row in ws.iter_rows(min_row=5, values_only=True):
        spec = row[1]
        mat_raw = row[2]
        location = row[3] if len(row) > 3 else None
        qty = row[7] if len(row) > 7 else 0
        
        # 规格在下行没有时用上一个
        if spec is not None and str(spec).strip():
            last_spec = str(spec).strip()
            if location is not None and str(location).strip():
                last_location = str(location).strip()
            else:
                last_location = ''
        elif location is not None and str(location).strip():
            last_location = str(location).strip()
        
        if mat_raw is None:
            continue
        mat = str(mat_raw).strip()
        if not mat:
            continue
        
        if qty is None:
            qty = 0
        try:
            qty_val = int(float(str(qty)))
        except:
            qty_val = 0
        
        record = make_zhengsquare(last_spec, mat, qty_val, last_location or '', dim_type)
        records.append(record)
    
    wb.close()
    return records

def read_changfang(filepath):
    """读取长方形Excel文件"""
    print(f"正在读取长方形: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    records = []
    
    for row in ws.iter_rows(min_row=3, values_only=True):
        spec = row[0]
        location = row[1]
        qty = row[3]  # D列 = 剩余库存
        
        if spec is None:
            continue
        spec_str = str(spec).strip()
        if not spec_str:
            continue
        
        if location is not None and str(location).strip() and str(location).strip() not in ('None', '空', '-', '无', ''):
            loc = str(location).strip()
        else:
            loc = ''
        
        if qty is None:
            qty = 0
        try:
            qty_val = int(float(str(qty)))
        except:
            qty_val = 0
        
        mat = detect_material(spec_str, None)
        
        record = make_changfang(spec_str, mat, qty_val, loc)
        records.append(record)
    
    wb.close()
    return records


def main():
    print("=" * 60)
    print("三羊包装 - 库存数据重建")
    print(f"时间: {NOW}")
    print("=" * 60)
    
    # 1. 读取正方形外径
    outer_file = os.path.join(DOCS, 'doc_b7beb7641301_正方形外径.xlsx')
    if not os.path.exists(outer_file):
        outer_file = os.path.join(ROOT, '正方形外径.xlsx')
    outer_records = read_zhengsquare(outer_file, 'outer')
    print(f"  外径: {len(outer_records)}条")
    
    # 2. 读取正方形内径
    inner_file = os.path.join(DOCS, 'doc_7cc875a53f60_正方形内径.xlsx')
    if not os.path.exists(inner_file):
        inner_file = os.path.join(ROOT, '正方形内径.xlsx')
    inner_records = read_zhengsquare(inner_file, 'inner')
    print(f"  内径: {len(inner_records)}条")
    
    # 3. 读取长方形
    chang_file = os.path.join(DOCS, 'doc_725ca38570ff_长方形库存.xlsx')
    if not os.path.exists(chang_file):
        chang_file = os.path.join(ROOT, '长方形库存.xlsx')
    chang_records = read_changfang(chang_file)
    print(f"  长方形: {len(chang_records)}条")
    
    # 4. 去重 - 按name去重
    all_records = outer_records + inner_records + chang_records
    total_before = len(all_records)
    
    seen = set()
    unique_records = []
    dup_count = 0
    for r in all_records:
        key = r['name']
        if key in seen:
            dup_count += 1
            continue
        seen.add(key)
        unique_records.append(r)
    
    print(f"\n去重统计:")
    print(f"  去重前: {total_before}条")
    print(f"  重复: {dup_count}条")
    print(f"  去重后: {len(unique_records)}条")
    
    # 5. 按类型统计
    from collections import Counter
    type_count = Counter()
    mat_count = Counter()
    for r in unique_records:
        type_count[r['product_type']] += 1
        mat_count[r['material']] += 1
    
    print(f"\n产品类型分布:")
    for k,v in type_count.most_common():
        print(f"  {k}: {v}")
    
    print(f"\n材质分布:")
    for k,v in mat_count.most_common():
        print(f"  {k}: {v}")
    
    # 6. 写入文件
    output = {
        "finished": unique_records,
        "raw": [],
        "returned": []
    }
    
    output_path = os.path.join(ROOT, 'inventory.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    file_size = os.path.getsize(output_path)
    print(f"\n✅ 已写入: {output_path}")
    print(f"   文件大小: {file_size:,} bytes")
    print(f"   总库存: {len(unique_records)}条")
    print("=" * 60)


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
导入台帐和出入库两个文档到 inventory.json
清空后全部重新导入
"""
import json
import os
import openpyxl

INVENTORY_FILE = '/www/wwwroot/feijihe/data/inventory.json'
TAIZHANG_FILE = '/www/wwwroot/feijihe/台帐-飞机盒.xlsx'
CHURUKU_FILE = '/www/wwwroot/feijihe/出入库正方形.xlsx'

def parse_size(size_str):
    """解析规格字符串为统一格式"""
    if not size_str or size_str == '(空白)':
        return None, None
    s = str(size_str).strip()
    # 去除括号内的材质标记，如 (台)、(白)、(红)
    # 统一乘号
    s = s.replace('*', '×')
    
    # 提取括号标记
    tag = ''
    if '(' in s:
        idx = s.index('(')
        tag = s[idx:]
        s = s[:idx].strip()
    
    # 解析尺寸数字
    parts = s.split('×')
    if len(parts) == 3:
        try:
            w = float(parts[0].strip())
            h = float(parts[1].strip())
            d = float(parts[2].strip())
            # 统一格式化：去掉 .0
            def fmt(x):
                return str(int(x)) if x == int(x) else str(x)
            return f"{fmt(w)}×{fmt(h)}×{fmt(d)}{tag}", (w, h, d)
        except:
            return s, None
    return s, None

def extract_taizhang_square(ws):
    """提取正方形台帐数据"""
    records = []
    # 找到表头行: 序号、外尺寸规格、库存、剩余库存
    header_row = None
    col_map = {}
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        for c, v in enumerate(cells, 1):
            if v and str(v).strip() in ['外尺寸规格', '规格']:
                col_map['size'] = c
            if v and str(v).strip() in ['库存', '剩余库存', '在库']:
                col_map['stock'] = c
            if v and str(v).strip() == '序号':
                header_row = row_idx
    if not col_map.get('size'):
        print(f"  找不到规格列, 使用默认列D(4)")
        col_map['size'] = 4
    if not col_map.get('stock'):
        print(f"  找不到库存列, 使用默认列H(8)")
        col_map['stock'] = 8
    
    data_start = header_row + 1 if header_row else 6
    print(f"  正方形: 规格列={col_map.get('size')}, 库存列={col_map.get('stock')}, 数据起始行={data_start}")
    
    for row_idx in range(data_start, ws.max_row + 1):
        size_cell = ws.cell(row=row_idx, column=col_map['size']).value
        stock_cell = ws.cell(row=row_idx, column=col_map['stock']).value
        if not size_cell or str(size_cell).strip() in ['(空白)', '', None]:
            continue
        size_str = str(size_cell).strip()
        if size_str in ['序号', '外尺寸规格', '规格', '合计']:
            continue
        if size_str.startswith('#'):
            continue
        
        stock_val = 0
        if stock_cell is not None:
            try:
                stock_val = int(float(str(stock_cell).replace(',', '')))
            except:
                stock_val = 0
        
        parsed_size, dims = parse_size(size_str)
        if parsed_size:
            records.append({
                'name': parsed_size,
                'stock': max(0, stock_val),
                'product_type': 'zhengsquare',
                'unit': '个'
            })
    return records

def extract_taizhang_rectangle(ws):
    """提取长方形台帐数据"""
    records = []
    header_row = None
    col_map = {}
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        for c, v in enumerate(cells, 1):
            if v and str(v).strip() in ['外尺寸规格', '规格']:
                col_map['size'] = c
            if v and str(v).strip() in ['库存', '剩余库存', '在库']:
                col_map['stock'] = c
            if v and str(v).strip() == '序号':
                header_row = row_idx
    if not col_map.get('size'):
        col_map['size'] = 4
    if not col_map.get('stock'):
        col_map['stock'] = 8
    
    data_start = header_row + 1 if header_row else 6
    print(f"  长方形: 规格列={col_map.get('size')}, 库存列={col_map.get('stock')}, 数据起始行={data_start}")
    
    for row_idx in range(data_start, ws.max_row + 1):
        size_cell = ws.cell(row=row_idx, column=col_map['size']).value
        stock_cell = ws.cell(row=row_idx, column=col_map['stock']).value
        if not size_cell or str(size_cell).strip() in ['(空白)', '', None]:
            continue
        size_str = str(size_cell).strip()
        if size_str in ['序号', '外尺寸规格', '规格', '合计']:
            continue
        if size_str.startswith('#'):
            continue
        
        stock_val = 0
        if stock_cell is not None:
            try:
                stock_val = int(float(str(stock_cell).replace(',', '')))
            except:
                stock_val = 0
        
        parsed_size, dims = parse_size(size_str)
        if parsed_size:
            records.append({
                'name': parsed_size,
                'stock': max(0, stock_val),
                'product_type': 'juxing',
                'unit': '个'
            })
    return records

def extract_churuku_sheet(ws, size_type):
    """提取出入库统计表数据 - size_type: 'nei'内尺寸 或 'wai'外尺寸"""
    records = []
    # 第2行是表头
    header_row = 2
    col_map = {}
    for c in range(1, min(20, ws.max_column + 1)):
        v = ws.cell(row=header_row, column=c).value
        if v:
            vs = str(v).strip()
            if vs in ['内尺寸规格', '外尺寸规格', '规格', '尺寸规格']:
                col_map['size'] = c
            if vs == '剩余库存':
                col_map['stock'] = c
    if not col_map.get('size'):
        col_map['size'] = 2
    if not col_map.get('stock'):
        col_map['stock'] = 8
    
    # 数据从第5行开始（第4行是空白/子表头，第5行是序号1开始）
    data_start = 5
    print(f"  {size_type}尺寸: 规格列={col_map.get('size')}, 库存列={col_map.get('stock')}, 数据起始行={data_start}")
    
    pt = 'zhengsquare' if size_type == '外' else 'juxing'
    
    for row_idx in range(data_start, ws.max_row + 1):
        seq = ws.cell(row=row_idx, column=1).value
        size_cell = ws.cell(row=row_idx, column=col_map['size']).value
        stock_cell = ws.cell(row=row_idx, column=col_map['stock']).value
        
        if seq is None or str(seq).strip() == '':
            continue
        if not size_cell:
            continue
        size_str = str(size_cell).strip()
        if size_str in ['序号', '规格', '外尺寸规格', '内尺寸规格', '合计']:
            continue
        
        # 获取材质
        cailiao = ''
        mat_cell = ws.cell(row=row_idx, column=3).value
        if mat_cell:
            cailiao = str(mat_cell).strip()
        
        # 获取货位
        huowei = ''
        hw_cell = ws.cell(row=row_idx, column=4).value
        if hw_cell:
            huowei = str(hw_cell).strip()
        
        stock_val = 0
        if stock_cell is not None:
            try:
                stock_val = int(float(str(stock_cell).replace(',', '')))
            except:
                stock_val = 0
        
        parsed_size, dims = parse_size(size_str)
        if parsed_size:
            records.append({
                'name': parsed_size,
                'stock': max(0, stock_val),
                'product_type': pt,
                'unit': '个',
                'material': cailiao,
                'location': huowei
            })
    return records

def merge_records(records):
    """合并重复记录（相同name+product_type去重求和）"""
    merged = {}
    for r in records:
        key = f"{r['name']}|{r['product_type']}"
        if key in merged:
            merged[key]['stock'] = max(merged[key]['stock'], r['stock'])
            # 合并材质/位置信息
            if r.get('material') and not merged[key].get('material'):
                merged[key]['material'] = r['material']
            if r.get('location') and not merged[key].get('location'):
                merged[key]['location'] = r['location']
        else:
            merged[key] = dict(r)
    return list(merged.values())

def main():
    print("=" * 60)
    print("三羊包装 - 成品库存批量导入")
    print("=" * 60)
    
    # 1. 导入台帐文件
    print("\n📄 导入台帐文件 (台帐-飞机盒.xlsx)...")
    wb = openpyxl.load_workbook(TAIZHANG_FILE, data_only=True)
    
    all_records = []
    
    # 正方形
    if '正方形' in wb.sheetnames:
        print("\n--- 正方形台帐 ---")
        rects = extract_taizhang_square(wb['正方形'])
        print(f"  提取到 {len(rects)} 条记录")
        all_records.extend(rects)
    
    # 长方形
    if '长方形' in wb.sheetnames:
        print("\n--- 长方形台帐 ---")
        rects = extract_taizhang_rectangle(wb['长方形'])
        print(f"  提取到 {len(rects)} 条记录")
        all_records.extend(rects)
    
    # 其他Sheet可能有数据
    for sn in wb.sheetnames:
        if sn in ['正方形', '长方形', 'Sheet4', '转换', '出入库记录', 'Sheet2', '台帐记录方法', 'Sheet1', 'Sheet3']:
            continue
        ws = wb[sn]
        if ws.max_row < 5 or ws.max_column < 3:
            continue
        # 看看有没有规格数据
        has_data = False
        for r in range(3, min(10, ws.max_row + 1)):
            for c in range(1, min(5, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if v and '×' in str(v):
                    has_data = True
                    break
        if has_data:
            print(f"\n--- {sn} (探测到数据) ---")
            rects = extract_taizhang_rectangle(ws)
            print(f"  提取到 {len(rects)} 条记录")
            all_records.extend(rects)
    
    print(f"\n📊 台帐合计: {len(all_records)} 条（去重前）")
    
    # 2. 导入出入库文件
    print("\n📄 导入出入库文件 (出入库正方形.xlsx)...")
    wb2 = openpyxl.load_workbook(CHURUKU_FILE, data_only=True)
    
    # 外尺寸 - 正方形
    if '外尺寸' in wb2.sheetnames:
        print("\n--- 外尺寸(正方形) ---")
        wai_records = extract_churuku_sheet(wb2['外尺寸'], '外')
        print(f"  提取到 {len(wai_records)} 条记录")
        all_records.extend(wai_records)
    
    # 内尺寸 - 长方形
    if '内尺寸' in wb2.sheetnames:
        print("\n--- 内尺寸(长方形) ---")
        nei_records = extract_churuku_sheet(wb2['内尺寸'], '内')
        print(f"  提取到 {len(nei_records)} 条记录")
        all_records.extend(nei_records)
    
    print(f"\n📊 合计: {len(all_records)} 条（去重前）")
    
    # 3. 去重合并
    print("\n🔄 去重合并...")
    merged = merge_records(all_records)
    print(f"  去重后: {len(merged)} 条记录")
    
    # 统计
    pt_count = {}
    for r in merged:
        pt = r['product_type']
        pt_count[pt] = pt_count.get(pt, 0) + 1
    print(f"\n  各类型数量:")
    for pt, cnt in sorted(pt_count.items()):
        print(f"    {pt}: {cnt} 条")
    
    # 4. 写入 inventory.json
    print(f"\n💾 写入 {INVENTORY_FILE}...")
    with open(INVENTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    
    total_stock = sum(r['stock'] for r in merged)
    print(f"\n✅ 导入完成！")
    print(f"   总记录数: {len(merged)} 条")
    print(f"   总库存数: {total_stock} 个")
    
    # 打印前10条看看
    print(f"\n📋 前10条示例:")
    for r in merged[:10]:
        print(f"   {r['name']:20s} | 库存: {r['stock']:>6d} | 类型: {r['product_type']:15s} | {r.get('material','')}")

if __name__ == '__main__':
    main()
